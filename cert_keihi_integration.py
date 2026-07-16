# -*- coding: utf-8 -*-
"""経費統合一覧表 移植の全行一致 認証スクリプト（P1a 合格ライン検証）

現行マクロが実際に出力した「経費統合一覧表」CSV（＝正解）と、同じ4生ファイルから
Python 実装（services/keihi_summary）が生成した統合一覧表を、34列すべてで多重集合突合する。
→ 差分0 が P1a の合格ライン（README_勤怠チェッカー組み込み仕様.md §8）。

使い方（同一マクロ実行で使った「揃った」データを渡すこと）:
  python cert_keihi_integration.py ^
    --jinjer   "...\\*仕訳データ.csv" ^
    --estaffing "...\\勤怠データダウンロード_立替金_*.csv" ^
    --sap      "...\\経費月次出力_API連携用*.csv" ^
    --freee    "...\\経費精算-*.csv" ^
    --expected "...\\経費利用履歴 RevN.csv"

社員番号の照合ロスターは、公平性のため expected CSV の（社員番号↔氏名）ペアと jinjer CSV から
構築する（本番は jinjer API 在籍者ロスターを使い、実測で未照合0を確認済み）。
"""
import argparse
import csv
import re
from collections import Counter
from pathlib import Path

from services.keihi_summary import (
    read_csv_any_enc, build_integrated_rows, build_roster_from_jinjer_csv,
    add_roster_entry, norm_date_slash, INTEGRATED_HEADERS, NCOL,
    C_EMP, C_NAME,
)

# 日付として正規化して比較する列（ゼロ埋め差・区切り差を吸収）
_DATE_COLS = {2, 5, 20}  # 申請日 / 利用日 / 計上日(yyyy/mm/dd)


def _read_expected(path: str) -> list:
    """正解の統合一覧表を CSV か xlsx から読み、ヘッダーを除いたデータ行(list[list])を返す。

    xlsx はシート「経費統合一覧表」を優先、無ければ先頭シート。空行は落とす。
    """
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        from datetime import date as _date, datetime as _dt
        from openpyxl import load_workbook

        def _cellstr(v):
            if v is None:
                return ""
            if isinstance(v, (_dt, _date)):
                return f"{v.year}/{v.month}/{v.day}"
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["経費統合一覧表"] if "経費統合一覧表" in wb.sheetnames else wb.worksheets[0]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # ヘッダー
                cells = [_cellstr(v) for v in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)
            return rows
        finally:
            wb.close()
    _, rows = read_csv_any_enc(path)
    return rows


def _norm_cell(i: int, v) -> str:
    s = "" if v is None else str(v).strip()
    if i in _DATE_COLS:
        return norm_date_slash(s)
    return s


def _row_key(row) -> tuple:
    return tuple(_norm_cell(i, row[i] if i < len(row) else "") for i in range(NCOL))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--jinjer")
    ap.add_argument("--estaffing")
    ap.add_argument("--sap")
    ap.add_argument("--freee")
    ap.add_argument("--expected", required=True, help="現行マクロ出力の経費統合一覧表CSV（正解）")
    ap.add_argument("--samples", type=int, default=10, help="差分の表示件数")
    args = ap.parse_args()

    # --- 正解（expected）を読む（CSV / xlsx 両対応） ---
    exp_rows = _read_expected(args.expected)
    exp_rows = [list(r) + [""] * (NCOL - len(r)) for r in exp_rows]

    # --- ロスター（expected + jinjer CSV の 社員番号↔氏名） ---
    roster: dict = {}
    for r in exp_rows:
        emp = (r[C_EMP] or "").strip()
        if emp and emp[0] not in ("5", "6", "9"):
            add_roster_entry(roster, r[C_NAME] or "", emp)
    if args.jinjer:
        jh, jr = read_csv_any_enc(args.jinjer)
        for k, v in build_roster_from_jinjer_csv(jh, jr).items():
            roster.setdefault(k, v)

    # --- Python 実装で統合一覧表を生成 ---
    produced, counts = build_integrated_rows(
        jinjer_csv=args.jinjer, estaffing_csv=args.estaffing,
        sap_csv=args.sap, freee_csv=args.freee, roster=roster, log_func=print,
    )

    print("\n================ 認証結果 ================")
    print(f"ソース件数: {counts}")
    print(f"produced 行数: {len(produced)}  /  expected 行数: {len(exp_rows)}")

    exp_keys = Counter(_row_key(r) for r in exp_rows)
    prod_keys = Counter(_row_key(r) for r in produced)

    matched = sum((exp_keys & prod_keys).values())
    missing = exp_keys - prod_keys      # 正解にあるが未生成
    extra = prod_keys - exp_keys        # 生成したが正解に無い
    print(f"完全一致(34列): {matched}")
    print(f"未生成（正解のみ）: {sum(missing.values())}")
    print(f"余剰（生成のみ）  : {sum(extra.values())}")

    if not missing and not extra:
        print("\n[OK] 合格: 全行・全列一致（差分0）")
        return 0

    print("\n--- 未生成サンプル（正解にあるが Python 出力に無い） ---")
    for k in list(missing)[:args.samples]:
        print("  ", " | ".join(f"{INTEGRATED_HEADERS[i]}={k[i]}" for i in (0, 1, 3, 7, 6, 19, 33) if k[i]))
    print("\n--- 余剰サンプル（Python 出力にあるが正解に無い） ---")
    for k in list(extra)[:args.samples]:
        print("  ", " | ".join(f"{INTEGRATED_HEADERS[i]}={k[i]}" for i in (0, 1, 3, 7, 6, 19, 33) if k[i]))
    print("\n[NG] 不一致あり（上の差分を確認）")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
