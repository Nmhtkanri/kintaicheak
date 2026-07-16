# -*- coding: utf-8 -*-
"""経費分類・集計（P1b）の全セル一致 認証スクリプト

現行マクロ（c設定.Run_経費集計_設定シート版）が出力した「集計」シートと、
同じ統合一覧表を Python 実装（services/keihi_classify）に通した集計を、
社員別 C〜L 列（合計/夜間当番/RINK/手当2/顧客請求分/交通費/非課税精算/その他/テレワーク/請求日）
で突合する。→ 差分0 が P1b の合格ライン。

使い方:
  python cert_keihi_classify.py --book "Y:\\...\\経費利用履歴 Rev5.xlsm"
    → ブック内の「経費統合一覧表」シートを入力、「集計」シートを正解として突合
  python cert_keihi_classify.py --input <統合一覧表.xlsx/csv> --expected <集計.xlsx/csv>
    → 入力と正解を別ファイルで指定

  --intersection : マクロ側集計シートに存在する社員だけ比較し、Python側のみの社員
                   （=マクロが落とした新入社員）を「期待される差分」として別掲する
                   （マクロの自動行追加修正が入る前の正解と比較するときに使う）

実績: live Rev5.xlsm（統合一覧表2401行）で 173名全員一致・差分0、
      シート未登録 = 2026012/2026013（今回の計上漏れ事故の2名）を確認（2026-07-16）。
"""
import argparse
import sys
from datetime import date, datetime, time
from pathlib import Path

from services.keihi_classify import (
    classify_rows, aggregate_by_id,
    B_YAKAN, B_RINK, B_TRANS, B_ETC, B_TW, B_BILL, B_NONTAX,
)
from services.keihi_summary import read_csv_any_enc, in_company_scope

NCOL = 34


def _cellstr(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return f"{v.year}/{v.month}/{v.day}"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def read_integrated(path: str, sheet: str = "経費統合一覧表") -> list[list[str]]:
    """統合一覧表を xlsx/xlsm のシート or CSV から読み、34列の文字列行リストを返す。"""
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                cells = [_cellstr(v) for v in row[:NCOL]]
                cells += [""] * (NCOL - len(cells))
                if any(c.strip() for c in cells):
                    rows.append(cells)
            return rows
        finally:
            wb.close()
    _, rows = read_csv_any_enc(path)
    return [list(r) + [""] * (NCOL - len(r)) for r in rows]


def read_expected_summary(path: str, sheet: str = "集計") -> dict:
    """正解の集計シートを読み {社員番号: (C..L の10値)} を返す。"""
    expected: dict = {}
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                emp = str(row[0]).strip() if row and row[0] is not None else ""
                if emp:
                    expected[emp] = tuple(row[2:12])
        finally:
            wb.close()
        return expected
    _, rows = read_csv_any_enc(path)
    for r in rows:
        emp = (r[0] or "").strip()
        if emp:
            vals = list(r[2:12]) + [""] * (10 - len(r[2:12]))
            expected[emp] = tuple(vals)
    return expected


def canon_num(v):
    """数値セルの正規化。空/None/00:00:00(時刻書式の0)は 0 に。"""
    if v is None or v == "":
        return 0
    if isinstance(v, time):
        return 0 if v == time(0, 0) else v
    if isinstance(v, datetime):
        return 0 if v.year < 1901 else v
    try:
        f = float(str(v).replace(",", ""))
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return v


def canon_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, time):
        return None if v == time(0, 0) else v
    if isinstance(v, str) and v.strip():
        s = v.strip().replace("-", "/").split(" ")[0]
        parts = s.split("/")
        if len(parts) == 3:
            try:
                return date(int(parts[0]), int(parts[1]), int(parts[2]))
            except ValueError:
                return v
    if v == "" or v is None:
        return None
    return v


COLS = ["C合計", "D夜間当番", "E RINK", "F手当2", "G顧客請求分",
        "H交通費", "I非課税精算", "Jその他", "Kテレワーク", "L請求日"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--book", help="経費利用履歴 xlsm（統合一覧表と集計を同じブックから読む）")
    ap.add_argument("--input", help="統合一覧表（xlsx/xlsm/CSV）")
    ap.add_argument("--expected", help="正解の集計シート（xlsx/xlsm/CSV）")
    ap.add_argument("--keywords", help="分類キーワード設定（省略時は内蔵デフォルト）")
    ap.add_argument("--intersection", action="store_true",
                    help="マクロ側に存在する社員だけ比較（Python側のみの新入社員は別掲）")
    ap.add_argument("--samples", type=int, default=15)
    args = ap.parse_args()

    if args.book:
        input_path = expected_path = args.book
    elif args.input and args.expected:
        input_path, expected_path = args.input, args.expected
    else:
        ap.error("--book か、--input と --expected の組を指定してください")

    rows = read_integrated(input_path)
    expected = read_expected_summary(expected_path)
    print(f"統合一覧表: {len(rows)} 行 / 正解の集計: {len(expected)} 名")

    keywords = None
    if args.keywords:
        from services.keihi_classify import load_keywords
        keywords = load_keywords(args.keywords)
        print(f"キーワード設定: {args.keywords}")

    cls = classify_rows(rows, keywords)
    agg = aggregate_by_id(cls)
    print(f"分類: {cls.hits} 件 / 集計ログ {len(cls.log)} 行 / 社員 {len(agg.by_id)} 名")

    n_match = n_diff = 0
    missing_on_sheet: list[tuple] = []
    diffs: list[tuple] = []
    for emp_id, vals in sorted(agg.by_id.items()):
        if not in_company_scope(emp_id):
            continue
        ours = [
            vals[B_YAKAN] + vals[B_RINK] + vals[B_BILL] + vals[B_TRANS]
            + vals[B_NONTAX] + vals[B_ETC] + vals[B_TW],
            vals[B_YAKAN], vals[B_RINK], vals[B_YAKAN] + vals[B_RINK], vals[B_BILL],
            vals[B_TRANS], vals[B_NONTAX], vals[B_ETC], vals[B_TW],
            agg.date_by_id.get(emp_id),
        ]
        exp = expected.get(emp_id)
        if exp is None:
            missing_on_sheet.append((emp_id, int(ours[0])))
            continue
        exp_c = [canon_num(x) for x in exp[:9]] + [canon_date(exp[9])]
        our_c = [canon_num(x) for x in ours[:9]] + [ours[9]]
        if exp_c == our_c:
            n_match += 1
        else:
            n_diff += 1
            detail = [(COLS[i], e, o) for i, (e, o) in enumerate(zip(exp_c, our_c)) if e != o]
            diffs.append((emp_id, detail))

    print("\n================ 認証結果 ================")
    print(f"C〜L 全セル一致: {n_match} 名")
    print(f"差分あり        : {n_diff} 名")
    print(f"マクロ集計シート未登録（Python側のみ）: {len(missing_on_sheet)} 名")
    for emp_id, total in missing_on_sheet:
        print(f"    {emp_id}: 合計 {total:,} 円 ← マクロでは計上漏れになる社員（新入社員等）")

    if diffs:
        print("\n--- 差分サンプル ---")
        for emp_id, detail in diffs[:args.samples]:
            parts = [f"{c}: マクロ={e} Python={o}" for c, e, o in detail]
            print(f"  {emp_id}: " + " | ".join(parts))

    if n_diff == 0 and (args.intersection or not missing_on_sheet):
        print("\n[OK] 合格: 集計 C〜L 全セル一致（差分0）")
        return 0
    if n_diff == 0 and missing_on_sheet:
        print("\n[OK] 共通社員は全セル一致。マクロ側の計上漏れ社員のみ差分"
              "（マクロの自動行追加修正後は完全一致になる想定）")
        return 0
    print("\n[NG] 不一致あり（上の差分を確認）")
    return 1


if __name__ == "__main__":
    sys.exit(main())
