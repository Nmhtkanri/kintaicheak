"""quick_compare.py — 5月本番 MVP 差異一覧生成スクリプト

入力:
  --kintai-dir : 既存 kintai-checker が出力した突合結果xlsx 群のフォルダ
  --jinjer-dir : jinjer 管理画面からダウンロードした「汎用データ（まるめ適用後）」CSV 群のフォルダ
  --output     : 出力する差異一覧xlsx のパス

出力:
  差異一覧_<YYYY-MM>.xlsx
    - 差異一覧（人間判断プルダウン・警告レベル付き）※先頭シート
    - 取込ログ
    - サマリ（最後尾）

設計書: docs/PLAN_5月本番_3営業日MVP.md  /  docs/DESIGN_月次マスター_P0_P3.md
"""

from __future__ import annotations

import argparse
import csv
import sys
import re
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


# ===== jinjer CSV の列マップ（1-indexed、ヘッダー名で照合する想定だが定数として控えておく） =====
JINJER_HEADERS = {
    "name": "名前",
    "emp_id": "*従業員ID",
    "date": "*年月日",
    "punch_in_1": "出勤1",
    "punch_out_1": "退勤1",
    "break_1_start": "休憩1",
    "break_1_end": "復帰1",
    "finalized": "実績確定状況",
    "total_work": "総労働時間",
    "actual_work": "実労働時間",
    "break_total": "休憩時間",
}

# ===== 警告レベル =====
LEVEL_DANGER = "DANGER"
LEVEL_WARN = "WARN"
LEVEL_INFO = "INFO"

# ===== 警告閾値（後で調整可能） =====
PUNCH_DIFF_WARN_MIN = 120  # 出勤/退勤差分が 120分以上で WARN
# ※ 長時間労働(>10h)の注意喚起は廃止（勤怠チェック時には行わない）。
#    差分が無い総労働は他の差分なし行と同じく行を生成しない。
SHORT_BREAK_AT_LONG_WORK = 6  # 勤務 6h 超で休憩 0:00 なら DANGER（労基法的に必要）
OVER_BREAK_HOURS = 2       # 休憩 2h 超で WARN
# 自動修正提案値（採用ラベル）の既定しきい値（分）。差分勤怠チェッカーで選択した範囲。
DEFAULT_RECOMMEND_THRESHOLD_MIN = 10

# ===== 出力xlsx 列定義 =====
# 判断値: 「請求勤怠」=請求勤怠を正としてjinjerへ書き戻す / 「jinjer勤怠」=jinjerを正（書き戻さない） / 「保留」。
# 手入力欄（任意・上級者向け）は右側へ寄せる。
# 横スクロールを減らすため「判断に必要な列」を左へ集約する。
# 識別3列(氏名/対象日付/差異種別)はウィンドウ枠固定。コメント2列は判断材料なので人間判断の左へ。
# 手入力・参考(予定/休日休暇/ID/トレーサビリティ)は右へ寄せる。
# ※ quick_export は列名で読むため、列順を変えても後方互換は保たれる。
# 2026-08-20 谷津さん指定で「自動修正提案値」「打刻修正」を廃止（25列→23列）。
#   自動修正提案値は採用ラベルを出すだけで判断は確認区分の色で足りていた。
#   打刻修正（提案値と違う時刻を人が指定して書き戻す欄）は使われておらず、
#   代わりに jinjer 画面で直接修正し人間判断=保留にする運用へ。
#   ※ quick_export 側は列名で読むので、旧フォーマットの差異一覧は従来どおり処理できる。
DIFF_COLUMNS = [
    # 識別（ウィンドウ枠固定）。従業員IDを先頭(A列)へ。
    "従業員ID", "氏名", "対象日付", "差異種別",
    # すぐ見る（差異の中身）
    "請求勤怠値", "jinjer値", "差分(分)",
    # 確認区分（要確認/自動採用/自動OK/参考のみ）。深刻さは色＋警告理由で表す（警告レベル列は廃止）
    "確認区分",
    # 判断の根拠（人間判断の左にまとめる）。有休も判断材料なので予定/休日休暇もここに置く。
    "打刻時コメント",       # 汎用データ#96「打刻時コメント」より（出勤:/退勤: 両方）
    "打刻修正時コメント",   # 申請データCSVの「理由」より
    "警告理由",
    "スケジュール出勤", "スケジュール退勤", "休憩予定", "復帰予定", "休日休暇名1", "休日休暇名1：種別",
    # 判断（入力）
    "人間判断", "判断メモ",
    # 反映（手入力・普段使わない）。手入力休憩時間は休憩1/復帰1から自動計算する数式列。
    "手入力休憩1", "手入力復帰1", "手入力休憩時間",
    # 参考（右端・普段見ない）。
    "実績確定状況",
]

# 手入力欄（時刻を手で打つ列）。差異一覧の生成時に文字列書式('@')を事前設定する。
# 既定の書式のままだと Excel が「31:00」を経過時間([h]:mm:ss)、「7:00」を時刻型へ
# 自動変換して保存し、openpyxl 経由で timedelta/time となって quick_export の転記が
# 壊れる（2026-07-09 の事故。quick_export 側の正規化と二重の防御）。
# 文字列書式なら入力値がそのまま保存され、入力者にも「31:00」のまま見える。
# ※「手入力休憩時間」は数式列（TEXT()が文字列を返す）ため '@' の対象外。'@' を付けると
#   ユーザーが F2→Enter した瞬間に数式が文字列として表示されてしまう。
MANUAL_INPUT_TEXT_COLUMNS = ["手入力休憩1", "手入力復帰1"]

# 汎用データから転記する予定・有休 列のキャノニカル名 → (候補ヘッダー, 完全一致のみか)
# 有休系は完全一致のみ（部分一致だと「有休」が「AM有休」「PM有休」を誤ヒットするため）。
JINJER_EXTRA_COLUMNS: dict[str, tuple[list[str], bool]] = {
    "出勤予定": (["出勤予定時刻", "出勤予定"], False),
    "退勤予定": (["退勤予定時刻", "退勤予定"], False),
    "休憩予定": (["休憩予定時間", "休憩予定"], False),
    "復帰予定": (["復帰予定時刻1", "復帰予定時刻", "復帰予定"], False),
    "有休": (["有休"], True),
    "AM有休": (["AM有休"], True),
    "PM有休": (["PM有休"], True),
    # 休日休暇名1 は「休日休暇名1：種別」の部分文字列なので完全一致のみ（誤ヒット防止）。
    "休日休暇名1": (["休日休暇名1"], True),
    "休日休暇名1：種別": (["休日休暇名1：種別"], True),
    # 打刻時コメント（汎用データ#96）。"出勤: ○○ , 退勤: ○○" 形式で出勤退勤両方を含む。
    "打刻時コメント": (["打刻時コメント"], True),
    # 休日区分（U列）。実ヘッダーは「休日（0:法定休日1:所定休日2:法休(振替休出)…）」と
    # 長い注記付きのため部分一致で解決する。値: 0=法定休日 / 1=所定休日 / 2〜5=振替・時間外休出。
    "休日区分": (["休日（0:法定休日", "休日(0:法定休日"], False),
    # 勤務状況（CT列）。実ヘッダーは「勤務状況（0:未打刻1:欠勤）」。値: 1=欠勤 / 0=未打刻。
    "勤務状況": (["勤務状況"], False),
}


def resolve_jinjer_extra_columns(columns) -> dict[str, str]:
    """汎用データの実ヘッダーから、予定/有休 のキャノニカル名→実ヘッダー名を解決する。

    完全一致を優先し、exact_only=False の列のみ部分一致でフォールバックする。
    見つからない列は dict に含めない（呼び出し側で空欄転記＋警告ログにする）。
    """
    cols = [str(c).strip() for c in columns]
    resolved: dict[str, str] = {}
    for canonical, (candidates, exact_only) in JINJER_EXTRA_COLUMNS.items():
        found = None
        for cand in candidates:  # 完全一致
            if cand in cols:
                found = cand
                break
        if found is None and not exact_only:  # 部分一致フォールバック
            for cand in candidates:
                for col in cols:
                    if cand in col:
                        found = col
                        break
                if found:
                    break
        if found:
            resolved[canonical] = found
    return resolved

DIFF_KIND_PUNCH_IN = "出勤"
DIFF_KIND_PUNCH_OUT = "退勤"
DIFF_KIND_BREAK = "休憩"
DIFF_KIND_TOTAL = "総労働時間"
# 氏名→従業員ID を解決できず（jinjer 未登録 or 氏名表記の不一致）、請求勤怠側には
# 勤務がある人。従来は黙ってスキップしていたが、未払い見落とし防止のため要確認行で可視化する。
DIFF_KIND_UNMATCHED = "jinjer未登録"
# jinjer 汎用データの「勤務状況」(CT列: 0=未打刻/1=欠勤) の転記（2026-07-10 谷津指定）。
# 新しい列は増やさず、差異種別（D列）にそのまま記載する:
#   ・jinjer打刻なしの出勤/退勤行 → 勤務状況=1なら「欠勤」、0なら「未打刻」に種別を付け替える
#     （書き戻し先が出勤1/退勤1のどちらかは punch_side と警告理由で保持し、quick_export が解決する）
#   ・請求勤怠側にも勤務が無い欠勤日 → 行が無いので日単位の「欠勤」転記行を1行足す
DIFF_KIND_ABSENCE = "欠勤"
DIFF_KIND_NO_PUNCH = "未打刻"
# 実績（請求勤怠の出勤）がスケジュール開始より早い日（2026-08-13 谷津さん指定）。
# jinjer はスケジュール軸で集計する項目があるため、実際より遅い予定が残っていると
# 早出分の扱いがずれる。書き戻し先は打刻ではなく「出勤予定時刻」のみ
# （quick_export が種別で判別する）。遅刻方向（実績が遅い）はスケジュールを動かさない。
DIFF_KIND_SCHED_START = "スケジュール開始"

# 差異行を出さない日の条件（2026-07-10 谷津指定）:
#   ・休日休暇名1 が振休/代休/年次有給の「全日」で、請求勤怠が記載なしまたは0
#   ・休日区分(U列)が法定休日(0)/所定休日(1)で、請求勤怠が記載なしまたは0
#     （SAPの請求勤怠は休日にも0:00の行を出すため、そのままだと休日ぶんの
#       「jinjer出勤なし/請求勤怠側に時刻あり」等のノイズ行が大量に出る）
# 振替休出(2,3)・時間外休出(4,5)は働く日のため対象外。AM/PM半休も対象外（全日のみ）。
SUPPRESS_FULLDAY_LEAVE_KEYWORDS = ("振休", "振替休", "代休", "年次有給")
SUPPRESS_HOLIDAY_CODES = ("0", "1")  # 0:法定休日 / 1:所定休日

# 既存 kintai-checker の services パスを通す（氏名正規化を借りる場合用）
KINTAI_CHECKER_ROOT = Path(__file__).resolve().parent
if str(KINTAI_CHECKER_ROOT) not in sys.path:
    sys.path.insert(0, str(KINTAI_CHECKER_ROOT))

# トリアージ（要確認/自動採用/自動OK の分類）
from services.triage import (  # noqa: E402
    classify as triage_classify,
    is_zero_or_empty,
    TRIAGE_NEEDS_CHECK,
    TRIAGE_AUTO_KINTAI,
    TRIAGE_AUTO_OK,
    TRIAGE_INFO_ONLY,
    TRIAGE_ORDER,
    JUDGE_KINTAI,
    JUDGE_JINJER,
    JUDGE_HOLD,
)


def recommend_judge_label(
    kind: str,
    diff_minutes: str,
    threshold_minutes: int,
    has_comment: bool = False,
) -> str:
    """自動修正提案値（採用ラベル）を決める。

    出退勤の新ルール（2026-06-25 谷津指定）:
      - 打刻時/打刻修正コメントに記載あり → jinjer勤怠
      - コメントなし & 差分が許容しきい値以上、または片側欠落（差分なし）→ 保留
      - コメントなし & 差分が許容しきい値未満 → 請求勤怠
    総労働時間は手順3で書き戻せない計算値のため対象外＝従来通り jinjer勤怠（変更なし）。
    jinjer未登録・欠勤は採用ラベルを付けない（書き戻し対象外・人が確認する）。
    """
    if kind in (DIFF_KIND_UNMATCHED, DIFF_KIND_ABSENCE):
        return ""
    if kind == DIFF_KIND_TOTAL:
        return JUDGE_JINJER
    if kind in (DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT):
        if has_comment:
            return JUDGE_JINJER
        d = to_int_diff(diff_minutes)
        # 片側欠落（差分なし）は安全側に保留。差分が許容しきい値以上も保留。
        if d is None or abs(d) >= threshold_minutes:
            return JUDGE_HOLD
        return JUDGE_KINTAI
    return JUDGE_JINJER


# ----------------------------------------------------------------------
# データクラス
# ----------------------------------------------------------------------

@dataclass
class DiffRow:
    row_id: int
    emp_id: str
    name: str
    target_date: str  # YYYY-MM-DD
    kind: str
    kintai_value: str
    jinjer_value: str
    diff_minutes: str  # 文字列で持つ（空欄表現のため）
    warn_level: str
    warn_reason: str
    auto_fix_value: str
    finalized: str  # 実績確定状況（参考情報。"TRUE"/"FALSE"/空）
    source_file: str
    # 汎用データから転記する予定・有休（参考表示。差異判定には使わない）
    sched_in: str = ""
    sched_out: str = ""
    sched_break: str = ""
    sched_break_end: str = ""  # 復帰予定（汎用データ「復帰予定時刻1」）
    yukyu: str = ""
    am_yukyu: str = ""
    pm_yukyu: str = ""
    # 汎用データの休日休暇名1 / 休日休暇名1：種別（参考表示）
    holiday_name1: str = ""
    holiday_name1_type: str = ""
    # 打刻時コメント（汎用データ#96。出勤:/退勤: 両方を含む。判断材料）
    punch_comment: str = ""
    # 打刻修正時コメント（打刻修正申請の従業員コメント等。同一 emp/date の全差異行に併記）
    jinjer_stamp_comment: str = ""
    # トリアージ区分（要確認/自動採用/自動OK）と既定の人間判断
    triage: str = ""
    judge_default: str = ""
    # 自動修正提案値（採用ラベル）。差分が許容しきい値内→請求勤怠 / 超過→jinjer勤怠 /
    # 総労働・休憩・差分なし→jinjer勤怠。
    recommend_judge: str = ""
    # 内部用（xlsxには出さない）: 片側欠落の打刻行がどちらの打刻か（"in"/"out"）。
    # 差異種別を欠勤/未打刻へ付け替えても、トリアージ・提案・書き戻し先は出勤/退勤として扱う。
    punch_side: str = ""


@dataclass
class LogEntry:
    severity: str  # INFO/WARN/ERROR
    message: str
    source: str = ""


# ----------------------------------------------------------------------
# 時刻ユーティリティ
# ----------------------------------------------------------------------

_TIME_RE = re.compile(r"^(\d{1,3}):(\d{2})(?::\d{2})?$")


def parse_hhmm(value: Any) -> int | None:
    """'H:MM' / 'HH:MM' / 'HH:MM:SS' を分に変換。空欄や不正値は None。
    24時超表記（25:30 等）はそのまま分換算する（時刻 ≠ 時間長 の用途で使い分け）。
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    m = _TIME_RE.match(s)
    if not m:
        return None
    h = int(m.group(1))
    mm = int(m.group(2))
    return h * 60 + mm


def format_minutes_as_hhmm(minutes: int | None) -> str:
    if minutes is None:
        return ""
    h, m = divmod(int(minutes), 60)
    return f"{h}:{m:02d}"


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    return s


def normalize_kintai_result_columns(df: pd.DataFrame) -> pd.DataFrame:
    """新しい「請求勤怠」ヘッダーを内部処理用の旧キーへ寄せる。

    既存ユーザーがすでに出力済みの「勤務表」ヘッダーの差異一覧も読めるよう、
    内部キーは当面「勤務表_*」を維持する。
    """
    aliases = {
        "請求勤怠_出勤": "勤務表_出勤",
        "請求勤怠_出勤時刻": "勤務表_出勤時刻",
        "請求勤怠_退勤": "勤務表_退勤",
        "請求勤怠_退勤時刻": "勤務表_退勤時刻",
        "請求勤怠_総労働": "勤務表_総労働",
        "請求勤怠_総労働時間": "勤務表_総労働時間",
        "請求勤怠_総労働時間(分)": "勤務表_総労働時間(分)",
        "請求勤怠_総労働(分)": "勤務表_総労働(分)",
        "請求勤怠_実働": "勤務表_実働時間",
        "請求勤怠_実働時間": "勤務表_実働時間",
        "請求勤怠_コメント": "勤務表_コメント",
    }
    rename_map = {
        src: dst
        for src, dst in aliases.items()
        if src in df.columns and dst not in df.columns
    }
    if rename_map:
        return df.rename(columns=rename_map)
    return df


def elapsed_minutes_from_values(start: Any, end: Any) -> int | None:
    """出勤・退勤値から経過分を計算する。日跨ぎは翌日退勤として扱う。"""
    start_min = parse_hhmm(start)
    end_min = parse_hhmm(end)
    if start_min is None or end_min is None:
        return None
    if end_min < start_min:
        end_min += 24 * 60
    return end_min - start_min


def to_jinjer_overnight_punch_out(punch_in: Any, punch_out: Any) -> str:
    """日跨ぎ勤務の退勤時刻を jinjer インポート用の 24時超表記へ変換する。

    jinjer は「年月日」を勤務開始日として扱うため、翌朝退勤は 24時を超える表記
    （例: 翌 08:15 → 32:15）でないとインポートできない。請求勤怠の出勤 > 退勤
    （= 退勤が翌日にずれ込んだ夜勤）のときだけ、退勤に 24h を足した 'HH:MM' を返す。

    日跨ぎでない・出退勤のどちらかが空/不正なときは、退勤値をそのまま返す。
    すでに 24時超表記（32:15 等）なら out >= in となり再変換されない（冪等）。
    """
    out_clean = clean_cell(punch_out)
    in_min = parse_hhmm(punch_in)
    out_min = parse_hhmm(punch_out)
    if in_min is None or out_min is None or out_min >= in_min:
        return out_clean
    adjusted = out_min + 24 * 60
    h, m = divmod(adjusted, 60)
    return f"{h:02d}:{m:02d}"


def overnight_display_value(
    value: Any, ref_in: Any, sched_out: Any = None, sched_in: Any = None
) -> str:
    """夜勤など深夜を跨ぐ退勤時刻を 24時超表記(25:00~ / 33:00~)へ揃える（差異一覧の表示用）。

    変換条件（いずれかを満たせば +24h）:
      - 実出勤 ref_in があり、退勤 < 出勤（同日内で逆転＝翌日にずれ込み）
      - 退勤予定 sched_out が24時超（夜勤スケジュール）で、退勤が翌朝側
        （出勤予定 sched_in 未満、無ければ正午未満）
    すでに24時超表記・空・不正値はそのまま返す（冪等）。退勤(out)の表示にのみ使う。
    出退勤の差分は両側を同じだけ +24h するため不変（33:00 と 33:01 の差は 1 分）。
    """
    out_min = parse_hhmm(value)
    if out_min is None or out_min >= 24 * 60:
        return clean_cell(value)
    overnight = False
    in_min = parse_hhmm(ref_in)
    if in_min is not None and out_min < in_min:
        overnight = True
    else:
        sout = parse_hhmm(sched_out)
        if sout is not None and sout >= 24 * 60:
            sin = parse_hhmm(sched_in)
            ref = (sin % (24 * 60)) if sin is not None else 12 * 60
            if out_min < ref:
                overnight = True
    if not overnight:
        return clean_cell(value)
    adjusted = out_min + 24 * 60
    h, m = divmod(adjusted, 60)
    return f"{h:02d}:{m:02d}"


def first_present(row: pd.Series, candidates: list[str]) -> Any:
    for col in candidates:
        if col not in row:
            continue
        value = row.get(col)
        if value is None:
            continue
        s = str(value).strip()
        if s and s.lower() not in ("nan", "none"):
            return value
    return None


def kintai_total_minutes(krow: pd.Series) -> tuple[int | None, str]:
    """勤怠突合結果行から勤務表側の総労働時間を取得する。

    請求勤怠ファイル記載の正味労働時間（勤務表_実働時間）があれば最優先で使う。
    無ければ拘束時間ベースの総労働列、それも無ければ勤務表_出勤/退勤から計算する。
    """
    explicit = first_present(krow, [
        # 請求勤怠ファイル記載の正味労働（休憩控除後）を最優先。
        "勤務表_実働時間",
        "勤務表_実働(分)",
        "勤務表_総労働",
        "勤務表_総労働時間",
        "勤務表_総労働時間(分)",
        "勤務表_総労働(分)",
    ])
    explicit_min = parse_hhmm(explicit)
    if explicit_min is not None:
        return explicit_min, format_minutes_as_hhmm(explicit_min)
    if explicit is not None:
        try:
            numeric = int(float(str(explicit)))
            return numeric, format_minutes_as_hhmm(numeric)
        except (TypeError, ValueError):
            pass

    start = first_present(krow, ["勤務表_出勤", "勤務表_出勤時刻"])
    end = first_present(krow, ["勤務表_退勤", "勤務表_退勤時刻"])
    total = elapsed_minutes_from_values(start, end)
    return total, format_minutes_as_hhmm(total)


def kintai_day_zero_or_empty(krow: pd.Series) -> bool:
    """請求勤怠側の1日ぶんが「記載なし、または0」か。

    出勤・退勤とも空/0:00 で、かつ請求勤怠ファイル記載の実働・総労働も空/0 のとき True。
    全日休暇・法休/所休の行抑制と、欠勤転記の勤務有無判定に使う。
    """
    if not is_zero_or_empty(clean_cell(krow.get("勤務表_出勤"))):
        return False
    if not is_zero_or_empty(clean_cell(krow.get("勤務表_退勤"))):
        return False
    total_min, _ = kintai_total_minutes(krow)
    return total_min in (None, 0)


def normalize_date_iso(value: Any) -> str | None:
    """日付値を 'YYYY-MM-DD' に正規化。jinjer の 'YYYY/M/D' にも対応。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s or s.lower() in ("nan", "nat"):
        return None
    # ISO
    try:
        return datetime.strptime(s, "%Y-%m-%d").date().isoformat()
    except ValueError:
        pass
    # YYYY/M/D
    try:
        return datetime.strptime(s, "%Y/%m/%d").date().isoformat()
    except ValueError:
        pass
    # YYYY/MM/DD
    for fmt in ("%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ----------------------------------------------------------------------
# 入力読み込み
# ----------------------------------------------------------------------

KINTAI_RESULT_SHEET_CANDIDATES = ["突合結果", "突合"]


def _input_files(path: Path, patterns: list[str]) -> list[Path]:
    if path.is_file():
        return [path]

    files: list[Path] = []
    for pattern in patterns:
        files.extend(sorted(path.glob(pattern)))
    return files


def _parse_jp_date(value: Any) -> str | None:
    """'2026年05月07日' / 'YYYY-MM-DD' / 'YYYY/M/D' を 'YYYY-MM-DD' に正規化。"""
    s = str(value or "").strip()
    if not s:
        return None
    m = re.match(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return None
    return normalize_date_iso(value)


def load_stamp_correction_reasons(
    path: Path, logs: list[LogEntry]
) -> dict[tuple[str, str], list[dict]]:
    """jinjer「申請データ（打刻修正申請）」CSV を読み、(従業員ID, 日付ISO)→[{type,method,comment}] を返す。

    打刻修正申請の「理由」列が従業員コメント。理由が空の申請は除外する。
    同一(emp,date)に複数申請があれば全て保持し、フォーマット時に結合する。
    """
    result: dict[tuple[str, str], list[dict]] = {}
    if not path.exists():
        logs.append(LogEntry("WARN", f"申請データCSVが見つかりません: {path}"))
        return result

    total = 0
    for f in _input_files(path, ["*.csv"]):
        if f.name.startswith("~$"):
            continue
        df = None
        for enc in ("cp932", "utf-8-sig", "utf-8"):
            try:
                df = pd.read_csv(f, encoding=enc, dtype=object)
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logs.append(LogEntry("WARN", f"申請データCSV読込失敗: {f.name}: {e}"))
                df = None
                break
        if df is None:
            continue

        cols = {str(c).strip(): c for c in df.columns}
        emp_col, date_col, reason_col = cols.get("従業員ID"), cols.get("日付"), cols.get("理由")
        status_col = cols.get("ステータス")
        if not (emp_col and date_col and reason_col):
            logs.append(LogEntry(
                "WARN", f"申請データCSVに必要列(従業員ID/日付/理由)がありません: {f.name}"))
            continue

        for _, row in df.iterrows():
            reason = str(row.get(reason_col) or "").strip()
            if not reason or reason.lower() == "nan":
                continue
            # 定型語（打刻忘れ／打刻漏れ／打刻修正）は転記対象から除外。残りが無ければスキップ。
            reason = strip_punch_noise_words(reason)
            if not reason:
                continue
            emp_id = str(row.get(emp_col) or "").strip()
            date_iso = _parse_jp_date(row.get(date_col))
            if not emp_id or not date_iso:
                continue
            status = str(row.get(status_col) or "").strip() if status_col else ""
            comment = f"{reason}（{status}）" if status and status.lower() != "nan" else reason
            result.setdefault((emp_id, date_iso), []).append(
                {"type": "", "method": "", "comment": comment}
            )
            total += 1

    logs.append(LogEntry(
        "INFO", f"申請データCSVから打刻修正理由 {total} 件（{len(result)} (emp,date)）を読込"))
    return result


def load_kintai_results(kintai_dir: Path, logs: list[LogEntry]) -> pd.DataFrame:
    """突合結果xlsx を読んで縦結合する。ファイル単体指定にも対応する。"""
    if not kintai_dir.exists():
        logs.append(LogEntry("ERROR", f"突合結果xlsxまたはフォルダが見つかりません: {kintai_dir}"))
        return pd.DataFrame()

    frames = []
    xlsx_files = _input_files(kintai_dir, ["*.xlsx"])
    for xlsx in xlsx_files:
        # 一時ファイル除外
        if xlsx.name.startswith("~$"):
            continue
        try:
            xl = pd.ExcelFile(xlsx)
        except Exception as e:
            logs.append(LogEntry("ERROR", f"xlsx を開けません: {xlsx.name}: {e}"))
            continue
        sheet = None
        for cand in KINTAI_RESULT_SHEET_CANDIDATES:
            if cand in xl.sheet_names:
                sheet = cand
                break
        if sheet is None:
            sheet = xl.sheet_names[0]
            logs.append(LogEntry("WARN", f"突合結果シートが見つからないため先頭シート '{sheet}' を使用: {xlsx.name}"))
        try:
            df = pd.read_excel(xlsx, sheet_name=sheet, dtype=object)
        except Exception as e:
            logs.append(LogEntry("ERROR", f"シート読み込み失敗 {xlsx.name}:{sheet}: {e}"))
            continue
        df = normalize_kintai_result_columns(df)
        df["_source_file"] = xlsx.name
        frames.append(df)
        logs.append(LogEntry("INFO", f"突合結果読込: {xlsx.name} ({len(df)} 行)"))

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def load_jinjer_csvs(jinjer_dir: Path, logs: list[LogEntry]) -> pd.DataFrame:
    """jinjer CSV 群を CP932 で読んで縦結合する。ファイル単体指定にも対応する。"""
    if not jinjer_dir.exists():
        logs.append(LogEntry("ERROR", f"jinjer CSVまたはフォルダが見つかりません: {jinjer_dir}"))
        return pd.DataFrame()

    frames = []
    for path in _input_files(jinjer_dir, ["*.csv", "*.xlsx"]):
        if path.name.startswith("~$"):
            continue
        try:
            if path.suffix.lower() == ".csv":
                df = pd.read_csv(path, encoding="cp932", dtype=object)
            else:
                df = pd.read_excel(path, dtype=object)
        except Exception as e:
            logs.append(LogEntry("ERROR", f"jinjer データ読込失敗 {path.name}: {e}"))
            continue
        df["_jinjer_source"] = path.name
        frames.append(df)
        logs.append(LogEntry("INFO", f"jinjer 読込: {path.name} ({len(df)} 行)"))

    if not frames:
        return pd.DataFrame()
    # 必須列確認
    combined = pd.concat(frames, ignore_index=True)
    required = [JINJER_HEADERS["emp_id"], JINJER_HEADERS["date"]]
    for col in required:
        if col not in combined.columns:
            logs.append(LogEntry("ERROR", f"jinjer CSV に必須列 '{col}' がありません"))
            return pd.DataFrame()
    return combined


# ----------------------------------------------------------------------
# 突合
# ----------------------------------------------------------------------

def build_jinjer_index(jinjer_df: pd.DataFrame) -> dict[tuple[str, str], dict]:
    """(従業員ID, 日付ISO) -> jinjer 行 dict"""
    index: dict[tuple[str, str], dict] = {}
    emp_col = JINJER_HEADERS["emp_id"]
    date_col = JINJER_HEADERS["date"]
    for _, row in jinjer_df.iterrows():
        emp_id = str(row.get(emp_col) or "").strip()
        date_iso = normalize_date_iso(row.get(date_col))
        if not emp_id or not date_iso:
            continue
        index[(emp_id, date_iso)] = row.to_dict()
    return index


def build_name_to_emp_map(jinjer_df: pd.DataFrame) -> dict[str, str]:
    """氏名 → 従業員ID マップ。jinjer CSV の「名前」列から作る。
    姓名のスペース有無バリエーションも登録。同姓同名は最初の1人を採用（後で警告）。
    """
    name_col = JINJER_HEADERS["name"]
    emp_col = JINJER_HEADERS["emp_id"]
    result: dict[str, str] = {}
    if name_col not in jinjer_df.columns:
        return result
    seen_ids = set()
    for _, row in jinjer_df.iterrows():
        emp_id = str(row.get(emp_col) or "").strip()
        name = str(row.get(name_col) or "").strip()
        if not emp_id or not name or emp_id in seen_ids:
            continue
        seen_ids.add(emp_id)
        # 表記バリエーション
        variants = {
            name,
            name.replace(" ", ""),
            name.replace("　", ""),
            name.replace(" ", "").replace("　", ""),
        }
        for v in variants:
            if v and v not in result:
                result[v] = emp_id
    return result


def resolve_emp_id(name: Any, name_map: dict[str, str]) -> str | None:
    if not name:
        return None
    s = str(name).strip()
    if not s:
        return None
    for key in [s, s.replace(" ", ""), s.replace("　", ""), s.replace(" ", "").replace("　", "")]:
        if key in name_map:
            return name_map[key]
    return None


def to_int_diff(value: Any) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


# 差異一覧へ転記する打刻系コメントから除外する定型語。
# これらの語そのものは判断材料にならない（打刻を忘れた／直した、という事実だけ）ため、
# 文中にあってもその語だけ取り除き、残った本文は必ず転記する。
PUNCH_NOISE_WORDS = ("打刻忘れ", "打刻漏れ", "打刻修正")


def strip_punch_noise_words(text: Any) -> str:
    """打刻コメントから定型語（打刻忘れ／打刻漏れ／打刻修正）だけを除去する。

    語を抜いた跡に残る区切り文字・空白を軽く整理し、本文が残ればそれを返す。
    定型語しか無ければ空文字を返す。
    """
    s = str(text or "")
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    for w in PUNCH_NOISE_WORDS:
        s = s.replace(w, "")
    # 連続した空白を1つに
    s = re.sub(r"[ \t　]+", " ", s)
    # 語を抜いた跡に区切りだけが連続して残った箇所を畳む（例: "、 、" → "、"）
    s = re.sub(r"\s*([、，,/／・])\s*(?=[、，,/／・])", "", s)
    # 先頭・末尾に残った区切り・空白を除去
    s = s.strip(" \t　、，,/／・:：-—ー")
    return s


def clean_punch_comment(value: Any) -> str:
    """汎用データ「打刻時コメント」(#96) の値を整形する。

    形式は "出勤: X , 退勤: Y" で、中身が無いと "出勤:  , 退勤:  ," のようなゴミ文字列に
    なるため、中身のあるラベルだけ残す。何も無ければ空文字を返す。
    定型語（打刻忘れ／打刻漏れ／打刻修正）は転記時に各本文から除外する。
    """
    s = clean_cell(value)
    if not s:
        return ""
    parts = []
    found_label = False
    for m in re.finditer(r"(出勤|退勤)\s*[:：]\s*([^,]*)", s):
        found_label = True
        body = strip_punch_noise_words(m.group(2))
        if body:
            parts.append(f"{m.group(1)}: {body}")
    if parts:
        return " / ".join(parts)
    if found_label:
        return ""  # ラベルはあるが中身なし → ゴミなので空
    return strip_punch_noise_words(s)  # 想定外フォーマットも定型語だけは除外


def format_stamp_comments(items: list[dict]) -> str:
    """打刻コメント list（{type,method,comment}）を1セルぶんの文字列に整形する。

    例: "出勤[打刻修正申請] KDX出社 / 退勤[PC] 私用のため早退"
    """
    parts = []
    for it in items or []:
        # 定型語（打刻忘れ／打刻漏れ／打刻修正）は転記対象から除外。残りが無ければスキップ。
        comment = strip_punch_noise_words(it.get("comment"))
        if not comment:
            continue
        stype = str(it.get("type") or "").strip()
        method = str(it.get("method") or "").strip()
        label = "".join([stype, f"[{method}]" if method else ""])
        parts.append(f"{label} {comment}".strip() if label else comment)
    return " / ".join(parts)


def compute_diffs(
    kintai_df: pd.DataFrame,
    jinjer_index: dict[tuple[str, str], dict],
    name_map: dict[str, str],
    logs: list[LogEntry],
    extra_cols: dict[str, str] | None = None,
    stamp_comments: dict[tuple[str, str], list[dict]] | None = None,
    threshold_minutes: int = DEFAULT_RECOMMEND_THRESHOLD_MIN,
) -> list[DiffRow]:
    """突合結果xlsx の各行に対し、4種の差異を生成して返す。

    stamp_comments: ``{(従業員ID, 日付ISO): [{type,method,comment}, ...]}``。
    打刻修正申請の従業員コメント等を、同一 emp/date の各差異行に併記する。
    """
    rows: list[DiffRow] = []
    next_id = 1
    extra_cols = extra_cols or {}
    stamp_comments = stamp_comments or {}

    seen_emp_date: set[tuple[str, str]] = set()
    # 氏名→従業員ID を解決できなかった人を黙って捨てず、請求勤怠側に勤務がある日を
    # 集計して後で「jinjer未登録」要確認行にまとめる（未払いの静かな見落としを防ぐ）。
    unresolved: dict[str, dict] = {}
    # 欠勤転記用: 請求勤怠側の1日ぶんの状態（勤務の有無と表示用の時刻範囲）
    billing_day: dict[tuple[str, str], dict] = {}
    # 差異種別を「欠勤」に付け替えた打刻行を出した日（日単位の欠勤転記行と重複させない）
    absence_punch_days: set[tuple[str, str]] = set()
    # 行を出さなかった日（取込ログ用）: 全日休暇（振休/代休/年次有給）・法休/所休
    suppressed_leave_days: set[tuple[str, str]] = set()
    suppressed_holiday_days: set[tuple[str, str]] = set()

    def _extra_of(jrow_, canonical: str) -> str:
        """汎用データ行から予定・休暇等の転記値を取り出す（列未解決・行なしは空欄）。"""
        if jrow_ is None:
            return ""
        header = extra_cols.get(canonical)
        return clean_cell(jrow_.get(header)) if header else ""

    def _collect_extras(jrow_, emp_id_: str, date_iso_: str) -> dict:
        """同一 emp/date のすべての差異行に併記する参考情報（予定・有休・コメント）。"""
        return {
            "sched_in": _extra_of(jrow_, "出勤予定"),
            "sched_out": _extra_of(jrow_, "退勤予定"),
            "sched_break": _extra_of(jrow_, "休憩予定"),
            "sched_break_end": _extra_of(jrow_, "復帰予定"),
            "yukyu": _extra_of(jrow_, "有休"),
            "am_yukyu": _extra_of(jrow_, "AM有休"),
            "pm_yukyu": _extra_of(jrow_, "PM有休"),
            "holiday_name1": _extra_of(jrow_, "休日休暇名1"),
            "holiday_name1_type": _extra_of(jrow_, "休日休暇名1：種別"),
            "punch_comment": clean_punch_comment(_extra_of(jrow_, "打刻時コメント")),
            "jinjer_stamp_comment": format_stamp_comments(stamp_comments.get((emp_id_, date_iso_))),
        }

    for _, krow in kintai_df.iterrows():
        name = str(krow.get("氏名") or "").strip()
        date_iso = normalize_date_iso(krow.get("日付"))
        if not name or not date_iso:
            continue

        emp_id = resolve_emp_id(name, name_map)
        if not emp_id:
            # jinjer 側に氏名が見つからない（未登録 or 氏名表記の不一致）。
            # 請求勤怠側に実打刻がある日だけ未払い候補として記録する（休/空欄日は無視）。
            k_in0 = clean_cell(krow.get("勤務表_出勤"))
            k_out0 = clean_cell(krow.get("勤務表_退勤"))
            if k_in0 or k_out0:
                u = unresolved.setdefault(
                    name, {"dates": [], "source": str(krow.get("_source_file") or "")}
                )
                u["dates"].append(date_iso)
            logs.append(LogEntry("WARN", f"氏名 → 従業員ID 解決失敗: {name} ({krow.get('_source_file', '')})"))
            continue

        source_file = str(krow.get("_source_file") or "")
        jrow = jinjer_index.get((emp_id, date_iso))
        seen_emp_date.add((emp_id, date_iso))

        # 実績確定状況（参考表示のみ、警告レベル判定には使わない）
        finalized = ""
        if jrow is not None:
            finalized = str(jrow.get(JINJER_HEADERS["finalized"]) or "").strip()

        # 汎用データから予定・有休を転記（同一 emp/date のすべての差異行に併記する参考情報）。
        # 列が解決できない / jinjer 行が無いときは空欄。
        extra = _collect_extras(jrow, emp_id, date_iso)

        # 出退勤の値（行生成と、休暇・休日の行抑制の両方で使う）
        k_in = clean_cell(krow.get("勤務表_出勤"))
        j_in = clean_cell(krow.get("jinjer_出勤"))
        k_out = clean_cell(krow.get("勤務表_退勤"))
        j_out = clean_cell(krow.get("jinjer_退勤"))

        # 欠勤転記用に請求勤怠側の状態を記録（同一日に複数行あれば勤務ありを優先）
        billing_empty = kintai_day_zero_or_empty(krow)
        day_state = billing_day.setdefault((emp_id, date_iso), {"has_work": False, "span": ""})
        if not billing_empty:
            day_state["has_work"] = True
            if not day_state["span"]:
                day_state["span"] = f"{k_in}〜{k_out}" if (k_in and k_out) else (k_in or k_out)

        # ----- 全日休暇・法休/所休の行抑制（2026-07-10 谷津指定） -----
        # 請求勤怠が「記載なし、または0」の日は、jinjer側が
        #   ・法定休日(0)/所定休日(1) → SAPの請求勤怠が休日に出す0:00行等のノイズを出さない
        #     （振替休出・時間外休出(2〜5)は働く日のため通常どおり突合する）
        #   ・振休/代休/年次有給の「全日」 → 休暇として整合しているため差異行を出さない
        # DANGER含め当日の全行を出さない（billing_empty かつ全日休暇/休日の日に限る）。
        if jrow is not None and billing_empty:
            if _extra_of(jrow, "休日区分") in SUPPRESS_HOLIDAY_CODES:
                suppressed_holiday_days.add((emp_id, date_iso))
                continue
            if extra["holiday_name1_type"] == "全日" and any(
                kw in extra["holiday_name1"] for kw in SUPPRESS_FULLDAY_LEAVE_KEYWORDS
            ):
                suppressed_leave_days.add((emp_id, date_iso))
                continue

        # 突合結果の「特記」（月末日跨ぎ・前月末後半・Fieldglass時刻なし等）。
        # 種別ごとに出す行を変え、誤った自動書き戻し提案（打切り24:00での上書きや
        # 前月末勤務の後半を当日行へ書く等）を防ぐ。注記は総労働時間の行として出す
        # （総労働時間は手順3で書き戻し対象外のため、誤操作しても打刻が消えない）。
        tokki = clean_cell(krow.get("特記"))
        skip_in = skip_out = skip_total = False
        month_end_sched = False
        tokki_note = ""
        tokki_note_level = LEVEL_WARN
        if tokki:
            if "月末日跨ぎ" in tokki:
                # 請求側の24:00は暦日打切りで実退勤ではない。出勤のみ通常突合し、
                # 退勤は退勤予定時刻（スケジュール）を暫定の正として別途突合する
                # （月末の日跨ぎ勤務はスケジュール勤怠を暫定で働いたとみなし当月計上する運用ルール）。
                skip_out = skip_total = True
                month_end_sched = True
                tokki_note = tokki
            elif "前月末夜勤の後半" in tokki and "自動書戻し不可" in tokki:
                # 前月分でスケジュール暫定計上済み＝当月の判断は不要。参考の注記のみ出す。
                skip_in = skip_out = skip_total = True
                tokki_note = tokki
                tokki_note_level = LEVEL_INFO
            elif "前月末夜勤の後半" in tokki:
                # 前月末日で突合済み: 退勤（24時超表記で前月末日へ書き戻し）は通常突合。
                # 総労働は後半のみの値のため突合しない。
                skip_total = True
            elif "jinjer側2行分割登録" in tokki:
                # jinjerの実データが2行のまま（開始日行＋翌日行）のため、開始日行の退勤へ
                # 24時超の値を書くと翌日行と二重計上になる。退勤の自動書き戻しは止め、
                # 退勤に差があるときだけ注記行で知らせる（差が無ければ従来どおり行を出さない）。
                skip_out = True
                skip_total = True
                if to_int_diff(krow.get("退勤差分(分)")) not in (None, 0):
                    tokki_note = tokki
            else:
                # Fieldglass時刻なし・未知の特記は安全側: 全て注記行のみ出す
                skip_in = skip_out = skip_total = True
                tokki_note = tokki

        # jinjer打刻なしの行の差異種別（D列）: 勤務状況=1→欠勤 / 0→未打刻 / それ以外→出勤・退勤。
        # 書き戻し先（出勤1/退勤1）は punch_side と警告理由の「jinjer出勤なし/退勤なし」で保持する。
        kinmu_status = _extra_of(jrow, "勤務状況")
        if kinmu_status == "1":
            missing_kind = DIFF_KIND_ABSENCE
            missing_note = "(欠勤・未払いの恐れ)"
        elif kinmu_status == "0":
            missing_kind = DIFF_KIND_NO_PUNCH
            missing_note = "(未打刻)"
        else:
            missing_kind = None
            missing_note = ""

        # ----- 出勤差異 -----
        diff_in = to_int_diff(krow.get("出勤差分(分)"))
        punch_in_row_emitted = False   # 出勤の書き戻し行を出した日はスケジュール開始行を出さない
        if skip_in:
            pass
        elif diff_in is not None and diff_in != 0:
            level, reason = classify_punch_diff(diff_in, "in")
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=DIFF_KIND_PUNCH_IN,
                kintai_value=k_in, jinjer_value=j_in, diff_minutes=str(diff_in),
                warn_level=level, warn_reason=reason,
                auto_fix_value=k_in,
                finalized=finalized,
                source_file=source_file,
                **extra,
            ))
            next_id += 1
            punch_in_row_emitted = True
        elif diff_in is None and k_in and not j_in:
            if missing_kind == DIFF_KIND_ABSENCE:
                absence_punch_days.add((emp_id, date_iso))
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=missing_kind or DIFF_KIND_PUNCH_IN,
                kintai_value=k_in, jinjer_value="", diff_minutes="",
                warn_level=LEVEL_WARN,
                warn_reason=f"jinjer出勤なし{missing_note} / 請求勤怠側に時刻あり",
                auto_fix_value=k_in,
                finalized=finalized,
                source_file=source_file,
                punch_side="in",
                **extra,
            ))
            next_id += 1
            punch_in_row_emitted = True
        elif diff_in is None and j_in and not k_in:
            # 逆向きの片側欠落: jinjer に打刻あり / 請求勤怠側に時刻なし。
            # 請求勤怠が正なら jinjer の打刻は余分なので空で上書きして消す（auto_fix_value=""）。
            # 既定は要確認(WARN)。全日休暇日のjinjer打刻は triage が「自動OK(jinjer)」へ回す。
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=DIFF_KIND_PUNCH_IN,
                kintai_value="", jinjer_value=j_in, diff_minutes="",
                warn_level=LEVEL_WARN,
                warn_reason="請求勤怠なし / jinjer側に時刻あり",
                auto_fix_value="",
                finalized=finalized,
                source_file=source_file,
                **extra,
            ))
            next_id += 1

        # ----- スケジュール開始が実績より遅い（早出の予定合わせ）-----
        # 実績（請求勤怠の出勤）がスケジュール開始より早い日は、出勤予定時刻を請求勤怠に
        # 合わせる行を出す（2026-08-13 谷津さん指定。既定は自動採用＝手順3で書き戻る）。
        # 出勤の書き戻し行が出た日は出さない: 打刻の人間判断に連動して手順3が予定も
        # 同期する（quick_export のおまけ同期）ので、別行を出すと判断が割れたとき矛盾する。
        # 遅刻方向（実績がスケジュールより遅い）は動かさない＝遅刻が隠れるため。
        if not skip_in and not punch_in_row_emitted and k_in:
            sched_in_v = str(extra.get("sched_in") or "").strip()
            sched_out_v = str(extra.get("sched_out") or "").strip()
            k_in_min = parse_hhmm(k_in)
            sched_in_min = parse_hhmm(sched_in_v) if sched_in_v else 0
            # 退勤予定が空の行に出勤予定だけ書くと jinjer が行ごと弾くため、予定が揃う日に限る
            if sched_out_v and k_in_min and sched_in_min and k_in_min < sched_in_min:
                rows.append(DiffRow(
                    row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                    kind=DIFF_KIND_SCHED_START,
                    kintai_value=k_in, jinjer_value=sched_in_v,
                    diff_minutes=str(k_in_min - sched_in_min),
                    warn_level=LEVEL_INFO,
                    warn_reason=(f"実績の開始 {k_in} がスケジュール開始 {sched_in_v} より早い"
                                 " / 出勤予定時刻を請求勤怠に合わせる（打刻は触らない）"),
                    auto_fix_value=k_in,
                    finalized=finalized,
                    source_file=source_file,
                    **extra,
                ))
                next_id += 1

        # ----- 退勤差異 -----
        diff_out = to_int_diff(krow.get("退勤差分(分)"))
        # 夜勤など深夜跨ぎの退勤は 24時超表記(33:00 等)で表示する（差分は両側同値補正で不変）。
        k_out_disp = overnight_display_value(k_out, k_in, extra["sched_out"], extra["sched_in"])
        j_out_disp = overnight_display_value(j_out, j_in, extra["sched_out"], extra["sched_in"])
        if skip_out:
            pass
        elif diff_out is not None and diff_out != 0:
            level, reason = classify_punch_diff(diff_out, "out")
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=DIFF_KIND_PUNCH_OUT,
                kintai_value=k_out_disp, jinjer_value=j_out_disp, diff_minutes=str(diff_out),
                warn_level=level, warn_reason=reason,
                auto_fix_value=to_jinjer_overnight_punch_out(k_in, k_out),
                finalized=finalized,
                source_file=source_file,
                **extra,
            ))
            next_id += 1
        elif diff_out is None and k_out and not j_out:
            if missing_kind == DIFF_KIND_ABSENCE:
                absence_punch_days.add((emp_id, date_iso))
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=missing_kind or DIFF_KIND_PUNCH_OUT,
                kintai_value=k_out_disp, jinjer_value="", diff_minutes="",
                warn_level=LEVEL_WARN,
                warn_reason=f"jinjer退勤なし{missing_note} / 請求勤怠側に時刻あり",
                auto_fix_value=to_jinjer_overnight_punch_out(k_in, k_out),
                finalized=finalized,
                source_file=source_file,
                punch_side="out",
                **extra,
            ))
            next_id += 1
        elif diff_out is None and j_out and not k_out:
            # 逆向きの片側欠落: jinjer に打刻あり / 請求勤怠側に時刻なし。
            # 請求勤怠が正なら jinjer の打刻は余分なので空で上書きして消す（auto_fix_value=""）。
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=DIFF_KIND_PUNCH_OUT,
                kintai_value="", jinjer_value=j_out_disp, diff_minutes="",
                warn_level=LEVEL_WARN,
                warn_reason="請求勤怠なし / jinjer側に時刻あり",
                auto_fix_value="",
                finalized=finalized,
                source_file=source_file,
                **extra,
            ))
            next_id += 1

        # ----- 月末日跨ぎ: スケジュール退勤を暫定の正として突合 -----
        # jinjerは月をまたぐ範囲の汎用データを出せないため、月末夜勤の後半（翌月1日
        # 0:00〜）を翌月に突合することはできない。運用ルール: 月末の日跨ぎ勤務は
        # スケジュールの勤怠を暫定で働いたとみなし、当月勤怠として計上する。
        # そのため退勤は退勤予定時刻（24時超表記）を暫定の正として jinjer実績と突合し、
        # 採用時はその値を開始日の行へ書き戻す。
        if month_end_sched:
            sched_out_disp = overnight_display_value(
                extra["sched_out"], k_in or j_in, extra["sched_out"], extra["sched_in"]
            )
            sched_min = parse_hhmm(sched_out_disp)
            j_min = parse_hhmm(j_out_disp)
            if sched_min is not None and j_min is not None:
                tokki_note_level = LEVEL_INFO  # 退勤を突合できたので注記は参考扱い
                diff_sched = sched_min - j_min
                if diff_sched != 0:
                    level, reason = classify_punch_diff(diff_sched, "out")
                    rows.append(DiffRow(
                        row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                        kind=DIFF_KIND_PUNCH_OUT,
                        kintai_value=sched_out_disp, jinjer_value=j_out_disp,
                        diff_minutes=str(diff_sched),
                        warn_level=level,
                        warn_reason=f"月末日跨ぎ: スケジュール退勤を暫定の正として突合(当月計上) / {reason}",
                        auto_fix_value=sched_out_disp,
                        finalized=finalized,
                        source_file=source_file,
                        **extra,
                    ))
                    next_id += 1
            elif sched_min is not None and not j_out:
                rows.append(DiffRow(
                    row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                    kind=DIFF_KIND_PUNCH_OUT,
                    kintai_value=sched_out_disp, jinjer_value="", diff_minutes="",
                    warn_level=LEVEL_WARN,
                    warn_reason="月末日跨ぎ: jinjer退勤なし / スケジュール退勤(暫定)あり",
                    auto_fix_value=sched_out_disp,
                    finalized=finalized,
                    source_file=source_file,
                    **extra,
                ))
                next_id += 1
            # スケジュール未設定の場合は注記行(要確認)のみ＝手動確認

        # ----- 総労働時間差異 -----
        # 「休憩」の突合は廃止（総労働時間が正味で突合されるため不要）。
        if jrow is not None and not skip_total:
            k_total_min, k_total = kintai_total_minutes(krow)
            j_total = str(jrow.get(JINJER_HEADERS["total_work"]) or "").strip()
            j_total_min = parse_hhmm(j_total)
            if k_total_min is not None and j_total_min is not None and k_total_min != j_total_min:
                diff_total = k_total_min - j_total_min
                # 出退勤のズレだけで説明がつく総労働差異は行を出さない（出勤/退勤行と
                # 重複アラートになるため）。拘束時間の差（請求−jinjer）が総労働の差と
                # 一致する ＝ 休憩は両者同じで、差はすべて出退勤行に現れている。
                k_elapsed = elapsed_minutes_from_values(k_in, k_out)
                j_elapsed = elapsed_minutes_from_values(j_in, j_out)
                explained_by_punches = (
                    k_elapsed is not None
                    and j_elapsed is not None
                    and diff_total == k_elapsed - j_elapsed
                )
                if not explained_by_punches:
                    level, reason = classify_total_work_diff(diff_total, jrow)
                    rows.append(DiffRow(
                        row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                        kind=DIFF_KIND_TOTAL,
                        kintai_value=k_total, jinjer_value=j_total, diff_minutes=str(diff_total),
                        warn_level=level, warn_reason=reason,
                        auto_fix_value="",  # 総労働時間は自動反映しない
                        finalized=finalized,
                        source_file=source_file,
                        **extra,
                    ))
                    next_id += 1
            else:
                total_warn = classify_total_work(jrow)
                if total_warn:
                    level, reason, j_total = total_warn
                    rows.append(DiffRow(
                        row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                        kind=DIFF_KIND_TOTAL,
                        kintai_value=k_total or "-", jinjer_value=j_total, diff_minutes="",
                        warn_level=level, warn_reason=reason,
                        auto_fix_value="",  # 総労働時間は自動反映しない
                        finalized=finalized,
                        source_file=source_file,
                        **extra,
                    ))
                    next_id += 1

        # ----- 特記の注記行 -----
        # 月末日跨ぎ・前月末後半(書戻し不可)・Fieldglass時刻なし等は、判断材料として
        # 1行だけ注記を出す。総労働時間の行なので手順3で書き戻されることはない。
        if tokki_note:
            if "Fieldglass時刻なし" in tokki_note:
                reason = f"{tokki_note}: 出退勤時刻が取得できないため突合不可。Fieldglassのエントリ/レポートを確認"
            else:
                reason = tokki_note
            k_span = f"{k_in}〜{k_out_disp}" if (k_in or k_out_disp) else ""
            j_span = f"{j_in}〜{j_out_disp}" if (j_in or j_out_disp) else ""
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=DIFF_KIND_TOTAL,
                kintai_value=k_span, jinjer_value=j_span, diff_minutes="",
                warn_level=tokki_note_level, warn_reason=reason,
                auto_fix_value="",
                finalized=finalized,
                source_file=source_file,
                **extra,
            ))
            next_id += 1

    # 氏名→ID 解決不可だが請求勤怠に勤務がある人を「jinjer未登録」要確認行として可視化する。
    # 1人につき1行（勤務日数・期間を集約）。派遣等の正当な除外か、給与対象者の抜けかを人が判断する。
    for uname, info in unresolved.items():
        dates = sorted(info["dates"])
        if not dates:
            continue
        span = dates[0] if len(dates) == 1 else f"{dates[0]}〜{dates[-1]}"
        rows.append(DiffRow(
            row_id=next_id, emp_id="", name=uname, target_date=dates[0],
            kind=DIFF_KIND_UNMATCHED,
            kintai_value=f"勤務{len(dates)}日", jinjer_value="", diff_minutes="",
            warn_level=LEVEL_WARN,
            warn_reason=(
                f"jinjer未登録（氏名→従業員ID解決不可）。請求勤怠に勤務{len(dates)}日"
                f"（{span}）あり＝給与未反映の恐れ。jinjer登録/氏名表記を確認"
            ),
            auto_fix_value="",
            finalized="",
            source_file=info["source"],
        ))
        next_id += 1
        logs.append(LogEntry(
            "WARN",
            f"jinjer未登録の可能性: {uname} 請求勤怠 勤務{len(dates)}日（{span}）→ 差異一覧に要確認行を追加",
        ))

    # ----- jinjer欠勤（勤務状況=1）の日単位転記 -----
    # 欠勤日はjinjer側に打刻が無く、請求勤怠側にも何も無いと突合結果に行が現れない
    # ことがあるため、突合結果ではなく汎用データ全体を走査する。
    # 対象はチェック対象（突合結果に登場し従業員IDを解決できた）の従業員のみ。
    # 打刻行を「欠勤」種別で出した日は情報が重複するため、日単位の行は足さない。
    status_header = extra_cols.get("勤務状況")
    if status_header:
        scope_emp_ids = {e for e, _ in seen_emp_date}
        absence_rows = 0
        for (emp_id, date_iso), jrow in jinjer_index.items():
            if emp_id not in scope_emp_ids:
                continue
            if clean_cell(jrow.get(status_header)) != "1":
                continue
            if (emp_id, date_iso) in absence_punch_days:
                continue  # 打刻行（差異種別=欠勤）で表示済み
            day_state = billing_day.get((emp_id, date_iso)) or {}
            name = clean_cell(jrow.get(JINJER_HEADERS["name"])) or emp_id
            if day_state.get("has_work"):
                level = LEVEL_WARN
                reason = "jinjer欠勤だが請求勤怠に勤務あり（未払いの恐れ）。欠勤登録と請求勤怠を確認"
            else:
                level = LEVEL_INFO
                reason = "jinjer欠勤（請求勤怠に勤務なし）。欠勤で正しいか確認"
            rows.append(DiffRow(
                row_id=next_id, emp_id=emp_id, name=name, target_date=date_iso,
                kind=DIFF_KIND_ABSENCE,
                kintai_value=day_state.get("span") or "",
                jinjer_value="欠勤", diff_minutes="",
                warn_level=level, warn_reason=reason,
                auto_fix_value="",
                finalized=clean_cell(jrow.get(JINJER_HEADERS["finalized"])),
                source_file="",
                **_collect_extras(jrow, emp_id, date_iso),
            ))
            next_id += 1
            absence_rows += 1
        if absence_rows:
            logs.append(LogEntry(
                "INFO", f"jinjer欠勤（勤務状況=1）を差異一覧へ転記: {absence_rows} 件"))

    if suppressed_leave_days or suppressed_holiday_days:
        logs.append(LogEntry(
            "INFO",
            "請求勤怠が記載なし/0のため差異行を出さない日: "
            f"全日休暇(振休/代休/年次有給) {len(suppressed_leave_days)}日 / "
            f"法休・所休 {len(suppressed_holiday_days)}日",
        ))

    # トリアージ: 各差異行を 要確認/自動採用/自動OK に分類し、既定の人間判断を付ける
    _TOKKI_KEYWORDS = ("月末日跨ぎ", "前月末夜勤の後半", "Fieldglass時刻なし", "jinjer側2行分割登録")
    for r in rows:
        # 欠勤の日単位転記行は必ず人が見る（手順3の書き戻し対象外のため提案値も付けない）
        if r.kind == DIFF_KIND_ABSENCE and not r.punch_side:
            r.triage = TRIAGE_NEEDS_CHECK
            r.judge_default = ""
            r.recommend_judge = ""
            continue
        # 特記の注記行は休暇情報等で自動OK/自動採用に紛れさせない。
        # WARN（要対応の可能性）は要確認、INFO（説明だけの注記）は参考のみに固定する。
        if r.kind == DIFF_KIND_TOTAL and any(k in (r.warn_reason or "") for k in _TOKKI_KEYWORDS):
            r.triage = TRIAGE_NEEDS_CHECK if r.warn_level in (LEVEL_DANGER, LEVEL_WARN) else TRIAGE_INFO_ONLY
            r.judge_default = ""
            r.recommend_judge = ""
            continue
        # 差異種別を欠勤/未打刻へ付け替えた打刻行は、従来どおり出勤/退勤として分類・提案する
        if r.punch_side == "in":
            eff_kind = DIFF_KIND_PUNCH_IN
        elif r.punch_side == "out":
            eff_kind = DIFF_KIND_PUNCH_OUT
        else:
            eff_kind = r.kind
        r.triage, r.judge_default = triage_classify(
            kind=eff_kind,
            warn_level=r.warn_level,
            punch_comment=r.punch_comment,
            stamp_comment=r.jinjer_stamp_comment,
            kintai_value=r.kintai_value,
            holiday_name1=r.holiday_name1,
            holiday_name1_type=r.holiday_name1_type,
        )
        # 自動修正提案値（採用ラベル）。新ルール: コメントあり→jinjer勤怠 /
        # 差分≧しきい値or片側欠落→保留 / それ以外→請求勤怠（総労働は対象外）。
        has_comment = bool((r.punch_comment or "").strip() or (r.jinjer_stamp_comment or "").strip())
        r.recommend_judge = recommend_judge_label(
            eff_kind, r.diff_minutes, threshold_minutes, has_comment
        )

    return rows


def classify_punch_diff(diff_minutes: int, side: str) -> tuple[str, str]:
    """出勤/退勤の差分行の警告レベル判定。
    side: 'in' or 'out'

    注意: 実績確定状況（jinjer 側で本人が打刻申請を確定した状態）は警告レベル判定に
    使わない。実績確定済は「本人が確定済み」という意味であって「勤怠が正しい」という
    意味ではないため、それを理由に上書きをブロックしたり DANGER 表示すべきではない。
    実績確定状況は差異一覧シートに参考列として表示するだけ。
    """
    reasons: list[str] = []
    level = LEVEL_INFO

    abs_diff = abs(diff_minutes)
    if abs_diff >= PUNCH_DIFF_WARN_MIN:
        level = LEVEL_WARN
        reasons.append(f"{side}差分 {abs_diff}分 (>={PUNCH_DIFF_WARN_MIN})")

    if not reasons:
        reasons.append("通常差異")
    return level, " / ".join(reasons)


def classify_break(jrow: dict) -> tuple[str, str, str] | None:
    """jinjer 側の休憩時間に問題があるか判定。
    戻り値: (level, reason, jinjer_break_str) or None（警告不要）
    """
    j_break_str = str(jrow.get(JINJER_HEADERS["break_total"]) or "").strip()
    j_total_str = str(jrow.get(JINJER_HEADERS["total_work"]) or "").strip()
    j_break_min = parse_hhmm(j_break_str)
    j_total_min = parse_hhmm(j_total_str)

    reasons: list[str] = []
    level = LEVEL_INFO

    # ① 勤務 6h 超で休憩 0 → 労基法違反疑い (DANGER)
    if j_break_min == 0 and j_total_min is not None and j_total_min >= SHORT_BREAK_AT_LONG_WORK * 60:
        level = LEVEL_DANGER
        reasons.append(f"休憩 0:00 だが総労働 {format_minutes_as_hhmm(j_total_min)} (労基法違反疑い)")

    # ② 休憩 2h 超 → WARN
    if j_break_min is not None and j_break_min > OVER_BREAK_HOURS * 60:
        if level == LEVEL_INFO:
            level = LEVEL_WARN
        reasons.append(f"休憩 {format_minutes_as_hhmm(j_break_min)} 過剰の疑い")

    if not reasons:
        return None
    return level, " / ".join(reasons), j_break_str


def classify_total_work(jrow: dict) -> tuple[str, str, str] | None:
    """jinjer 側の総労働時間に問題があるか判定。
    戻り値: (level, reason, jinjer_total_str) or None

    ※ 長時間労働(>10h)の注意喚起は廃止。差分が無い総労働は行を生成しない。
       ここでは集計不整合（出勤打刻あり/総労働0:00）のみを検出する。
    """
    j_total_str = str(jrow.get(JINJER_HEADERS["total_work"]) or "").strip()
    # 空セルは CSV 読込で NaN になり `or ""` では拾えない（"nan"が打刻あり扱いになる）ため clean_cell で判定
    j_punch_in = clean_cell(jrow.get(JINJER_HEADERS["punch_in_1"]))
    j_total_min = parse_hhmm(j_total_str)

    reasons: list[str] = []
    level = LEVEL_INFO

    if j_total_min is None:
        return None

    # 出勤打刻あり / 総労働 0:00 → 計算不能警告（集計不整合）
    if j_total_min == 0 and j_punch_in:
        level = LEVEL_DANGER
        reasons.append("出勤打刻あり/総労働 0:00 (集計不整合)")

    if not reasons:
        return None
    return level, " / ".join(reasons), j_total_str


def classify_total_work_diff(diff_minutes: int, jrow: dict) -> tuple[str, str]:
    """勤務表とjinjerの総労働時間差異を判定。"""
    abs_diff = abs(diff_minutes)
    reasons = [f"請求勤怠総労働とjinjer総労働の差分 {abs_diff}分"]
    level = LEVEL_INFO
    if abs_diff >= PUNCH_DIFF_WARN_MIN:
        level = LEVEL_WARN

    j_total_str = str(jrow.get(JINJER_HEADERS["total_work"]) or "").strip()
    # 空セルの NaN を「打刻あり」と誤判定しないよう clean_cell で判定（欠勤・未打刻日の誤DANGER防止）
    j_punch_in = clean_cell(jrow.get(JINJER_HEADERS["punch_in_1"]))
    j_total_min = parse_hhmm(j_total_str)
    # 長時間労働(>10h)の昇格は廃止。集計不整合のみ DANGER に昇格する。
    if j_total_min == 0 and j_punch_in:
        level = LEVEL_DANGER
        reasons.append("出勤打刻あり/総労働 0:00 (集計不整合)")

    return level, " / ".join(reasons)


# ----------------------------------------------------------------------
# Excel 出力
# ----------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
# 「人間判断」列ヘッダー強調（入力すべき列だと分かるように）
JUDGE_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
LEVEL_FILL = {
    LEVEL_DANGER: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    LEVEL_WARN: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    LEVEL_INFO: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
# トリアージ区分の色（要確認を目立たせ、自動は淡色に）
TRIAGE_FILL = {
    TRIAGE_NEEDS_CHECK: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    TRIAGE_AUTO_OK: PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    TRIAGE_AUTO_KINTAI: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    TRIAGE_INFO_ONLY: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}
# 行の縞模様（先頭データ行から 緑→白 の繰り返し）。横に長いシートで行を目で
# 追いやすくする（2026-08-13 谷津さん要望）。警告理由・確認区分など意味のある色は
# 縞を塗ったあとに個別セルへ上塗りするので、そちらが優先で残る。
# 緑はトリアージの E2EFDA / INFO の C6EFCE と紛れない薄さにしてある。
STRIPE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
# 文字サイズは全シート 12pt に統一（2026-08-20 谷津さん要望。既定の11ptだと見づらい）
BASE_FONT_SIZE = 12
DATA_FONT = Font(size=BASE_FONT_SIZE)
HEADER_FONT = Font(bold=True, size=BASE_FONT_SIZE)


def write_excel(output_path: Path, diff_rows: list[DiffRow], logs: list[LogEntry], month_label: str) -> None:
    # スケジュール開始合わせは差異一覧のシートに出さない（確認対象を増やさない・
    # 2026-08-13 谷津さん指定）。別シートに参考として記録し、手順3が黙って
    # 出勤予定時刻を合わせる（人間判断は不要）。
    sched_aligns = [r for r in diff_rows if r.kind == DIFF_KIND_SCHED_START]
    diff_rows = [r for r in diff_rows if r.kind != DIFF_KIND_SCHED_START]
    # トリアージ区分で並べ替え（要確認を上へ。同区分内は元の順序を保持＝安定ソート）
    diff_rows = sorted(diff_rows, key=lambda r: TRIAGE_ORDER.get(r.triage, 9))
    needs_cnt = sum(1 for r in diff_rows if r.triage == TRIAGE_NEEDS_CHECK)
    auto_ok_cnt = sum(1 for r in diff_rows if r.triage == TRIAGE_AUTO_OK)
    auto_kintai_cnt = sum(1 for r in diff_rows if r.triage == TRIAGE_AUTO_KINTAI)
    info_only_cnt = sum(1 for r in diff_rows if r.triage == TRIAGE_INFO_ONLY)

    wb = Workbook()
    # サマリ
    ws_sum = wb.active
    ws_sum.title = "サマリ"
    summary_data = [
        ("対象月", month_label),
        ("総差異件数", len(diff_rows)),
        ("★要確認 件数（人が見る）", needs_cnt),
        ("自動OK(jinjer勤怠) 件数", auto_ok_cnt),
        ("自動採用(請求勤怠) 件数", auto_kintai_cnt),
        ("参考のみ 件数（判断不要）", info_only_cnt),
        ("出勤差異件数", sum(1 for r in diff_rows if r.kind == DIFF_KIND_PUNCH_IN)),
        ("スケジュール開始合わせ件数（別シート・自動反映）", len(sched_aligns)),
        ("退勤差異件数", sum(1 for r in diff_rows if r.kind == DIFF_KIND_PUNCH_OUT)),
        ("総労働時間差異件数", sum(1 for r in diff_rows if r.kind == DIFF_KIND_TOTAL)),
        ("欠勤件数", sum(1 for r in diff_rows if r.kind == DIFF_KIND_ABSENCE)),
        ("未打刻件数", sum(1 for r in diff_rows if r.kind == DIFF_KIND_NO_PUNCH)),
        ("DANGER 件数", sum(1 for r in diff_rows if r.warn_level == LEVEL_DANGER)),
        ("WARN 件数", sum(1 for r in diff_rows if r.warn_level == LEVEL_WARN)),
        ("INFO 件数", sum(1 for r in diff_rows if r.warn_level == LEVEL_INFO)),
        ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for r_idx, (label, value) in enumerate(summary_data, start=1):
        c1 = ws_sum.cell(row=r_idx, column=1, value=label)
        c1.font = HEADER_FONT
        ws_sum.cell(row=r_idx, column=2, value=value).font = DATA_FONT
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 28

    # 差異一覧
    ws = wb.create_sheet("差異一覧")
    for col_idx, header in enumerate(DIFF_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        # 「人間判断」列は入力すべき列だと一目で分かるよう、ヘッダーを目立たせる
        c.fill = JUDGE_HEADER_FILL if header == "人間判断" else HEADER_FILL
        c.font = HEADER_FONT
        # ヘッダーは折り返して表示（列幅をデータに合わせて詰めても列名が全部読める）
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32   # 折り返した2行分の高さ
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DIFF_COLUMNS))}1"
    # A〜G列(従業員ID/氏名/対象日付/差異種別/請求勤怠値/jinjer値/差分(分))とヘッダー行を固定。
    # 右へスクロールしても差異の中身ごと見える（2026-08-13 谷津さん要望で E2 → H2 に拡大）。
    ws.freeze_panes = "H2"

    # 「手入力休憩時間」は休憩1・復帰1から自動計算する数式列（手入力させない。下の
    # シート保護でこの列だけロックする）。列文字は列順の変更に追随するよう都度算出する。
    break_start_col = get_column_letter(DIFF_COLUMNS.index("手入力休憩1") + 1)
    break_end_col = get_column_letter(DIFF_COLUMNS.index("手入力復帰1") + 1)
    break_calc_idx = DIFF_COLUMNS.index("手入力休憩時間") + 1

    # データ行（列名→値の対応で書き込み、列順の変更に強くする）
    for r_idx, drow in enumerate(diff_rows, start=2):
        row_values = {
            "従業員ID": drow.emp_id,
            "氏名": drow.name,
            "対象日付": drow.target_date,
            "スケジュール出勤": drow.sched_in,
            "スケジュール退勤": drow.sched_out,
            "休憩予定": drow.sched_break,
            "復帰予定": drow.sched_break_end,
            "有休": drow.yukyu,
            "AM有休": drow.am_yukyu,
            "PM有休": drow.pm_yukyu,
            "休日休暇名1": drow.holiday_name1,
            "休日休暇名1：種別": drow.holiday_name1_type,
            "差異種別": drow.kind,
            "請求勤怠値": drow.kintai_value,
            "jinjer値": drow.jinjer_value,
            "差分(分)": drow.diff_minutes,
            "確認区分": drow.triage,
            "警告理由": drow.warn_reason,
            "打刻時コメント": drow.punch_comment,
            "打刻修正時コメント": drow.jinjer_stamp_comment,
            "人間判断": drow.judge_default,   # トリアージの既定値を事前入力（要確認は空欄）
            "判断メモ": "",
            "手入力休憩1": "",     # 休憩差異を承認して汎用データに反映する場合のみ
            "手入力復帰1": "",
            # 手入力休憩時間は下の数式で自動計算（row_values には入れない）
            "実績確定状況": drow.finalized,  # 参考表示
        }
        # 休憩1・復帰1 が両方入っていれば休憩時間を計算する。MOD で日跨ぎ休憩も正の値になり、
        # TEXT が "h:mm" の文字列を返すので手順3は手入力時と同じ形で読める。
        break_formula = (
            f'=IF(AND({break_start_col}{r_idx}<>"",{break_end_col}{r_idx}<>""),'
            f'IFERROR(TEXT(MOD({break_end_col}{r_idx}-{break_start_col}{r_idx},1),"h:mm"),""),"")'
        )
        for c_idx, header in enumerate(DIFF_COLUMNS, start=1):
            value = break_formula if header == "手入力休憩時間" else row_values.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.font = DATA_FONT
            if r_idx % 2 == 0:   # 先頭データ行(2行目)が緑、次が白…の縞
                cell.fill = STRIPE_FILL
        # 深刻さ（DANGER/WARN/INFO）は「警告理由」セルの色で表す（警告レベル列は廃止）
        fill = LEVEL_FILL.get(drow.warn_level)
        if fill:
            ws.cell(row=r_idx, column=DIFF_COLUMNS.index("警告理由") + 1).fill = fill
        # 確認区分セルに色塗り（要確認を目立たせる）
        tfill = TRIAGE_FILL.get(drow.triage)
        if tfill:
            ws.cell(row=r_idx, column=DIFF_COLUMNS.index("確認区分") + 1).fill = tfill

    # 手入力列に文字列書式('@')を事前設定（人間判断プルダウンと同じデータ行範囲）。
    # 理由は MANUAL_INPUT_TEXT_COLUMNS 定義部のコメント参照（Excel の時刻自動変換防止）。
    for header in MANUAL_INPUT_TEXT_COLUMNS:
        col_idx = DIFF_COLUMNS.index(header) + 1
        for r_idx in range(2, 2 + len(diff_rows)):
            ws.cell(row=r_idx, column=col_idx).number_format = "@"

    # データ検証プルダウン: 人間判断
    judge_col_letter = get_column_letter(DIFF_COLUMNS.index("人間判断") + 1)
    if diff_rows:
        dv = DataValidation(
            type="list",
            formula1='"請求勤怠,jinjer勤怠,保留"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="不正な値",
            error="請求勤怠 / jinjer勤怠 / 保留 のいずれかを選択してください",
        )
        ws.add_data_validation(dv)
        judge_range = f"{judge_col_letter}2:{judge_col_letter}{1 + len(diff_rows)}"
        dv.add(judge_range)

        # 人間判断の選択値に応じてセル色を変える（条件付き書式。選んだ瞬間に反映される）
        # 保留 → 背景赤・太字白 ／ jinjer勤怠 → 背景黄・太字黒
        ws.conditional_formatting.add(judge_range, CellIsRule(
            operator="equal", formula=['"保留"'],
            fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            font=Font(bold=True, color="FFFFFF"),
        ))
        ws.conditional_formatting.add(judge_range, CellIsRule(
            operator="equal", formula=['"jinjer勤怠"'],
            fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            font=Font(bold=True, color="000000"),
        ))

    # 列幅（列名で指定。列順を変えても崩れない）。ヘッダーは折り返し表示なので、
    # 時刻・数値しか入らない列は中身の幅まで詰めて横スクロールを減らす。
    width_map = {
        "従業員ID": 10, "氏名": 16, "対象日付": 11,
        "スケジュール出勤": 8, "スケジュール退勤": 8, "休憩予定": 8, "復帰予定": 8,
        "有休": 8, "AM有休": 8, "PM有休": 8,
        "休日休暇名1": 14, "休日休暇名1：種別": 12,
        "差異種別": 10, "請求勤怠値": 9, "jinjer値": 9, "差分(分)": 7,
        "確認区分": 14, "警告理由": 44,
        "人間判断": 12, "判断メモ": 28,
        "打刻時コメント": 40, "打刻修正時コメント": 40,
        "手入力休憩1": 10, "手入力復帰1": 10, "手入力休憩時間": 10,
        "実績確定状況": 10,
    }
    for i, header in enumerate(DIFF_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(header, 14)

    # シート保護（2026-08-20 谷津さん指定）。「手入力休憩時間」は自動計算の数式なので
    # 手で書き換えられないよう固定する。それ以外のデータセルは今までどおり編集できる
    # ように先に解除しておく（Excel は既定で全セルがロック扱いのため）。
    # ヘッダー行(1行目)はロックのまま残す＝列名を書き換えられると手順3の転記が静かに壊れるため。
    # パスワードは付けないので、必要になれば「校閲→シート保護の解除」でいつでも外せる。
    # ※ ロックされたセルを含む範囲は Excel の仕様で並べ替えができない（フィルタの絞り込みは可）。
    unlocked = Protection(locked=False)
    locked = Protection(locked=True)
    last_data_row = 1 + len(diff_rows)
    for row in ws.iter_rows(min_row=2, max_row=last_data_row + 300,
                            min_col=1, max_col=len(DIFF_COLUMNS) + 4):
        for cell in row:
            cell.protection = unlocked
    for r_idx in range(2, last_data_row + 1):
        ws.cell(row=r_idx, column=break_calc_idx).protection = locked
    prot = ws.protection
    prot.sheet = True        # 保護ON（True=禁止 / False=許可。sheet だけ True=保護ON）
    prot.autoFilter = False  # フィルタの絞り込みは許可
    prot.sort = False
    prot.formatCells = False
    prot.formatColumns = False
    prot.formatRows = False
    prot.insertRows = False

    # スケジュール開始合わせ（参考シート。手順3が読んで出勤予定時刻を自動で合わせる）
    if sched_aligns:
        ws_sa = wb.create_sheet("スケジュール開始合わせ")
        sa_headers = ["従業員ID", "氏名", "対象日付", "現在の予定開始",
                      "新しい開始(請求勤怠)", "説明"]
        for c_idx, h in enumerate(sa_headers, start=1):
            c = ws_sa.cell(row=1, column=c_idx, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        for r_idx, r in enumerate(sched_aligns, start=2):
            for c_idx, v in enumerate(
                [r.emp_id, r.name, r.target_date, r.jinjer_value], start=1
            ):
                ws_sa.cell(row=r_idx, column=c_idx, value=v).font = DATA_FONT
            c5 = ws_sa.cell(row=r_idx, column=5, value=r.kintai_value)
            c5.font = DATA_FONT
            c5.number_format = "@"   # Excelの時刻型変換を防ぐ（手入力列と同じ理由）
            ws_sa.cell(row=r_idx, column=6, value=r.warn_reason).font = DATA_FONT
        for i, w in enumerate([12, 16, 12, 14, 18, 60], start=1):
            ws_sa.column_dimensions[get_column_letter(i)].width = w
        ws_sa.freeze_panes = "A2"
        note = ws_sa.cell(row=len(sched_aligns) + 3, column=1,
                          value="※このシートの確認・編集は不要です。手順3の書き戻しが"
                                "出勤予定時刻だけを上の値に自動で合わせます（打刻は触りません）。"
                                "合わせたくない行は削除してください。")
        note.font = Font(color="808080", size=BASE_FONT_SIZE)

    # 取込ログ
    ws_log = wb.create_sheet("取込ログ")
    ws_log.cell(row=1, column=1, value="severity").font = HEADER_FONT
    ws_log.cell(row=1, column=2, value="message").font = HEADER_FONT
    ws_log.cell(row=1, column=3, value="source").font = HEADER_FONT
    for r_idx, entry in enumerate(logs, start=2):
        for c_idx, v in enumerate([entry.severity, entry.message, entry.source], start=1):
            ws_log.cell(row=r_idx, column=c_idx, value=v).font = DATA_FONT
    ws_log.column_dimensions["A"].width = 10
    ws_log.column_dimensions["B"].width = 80
    ws_log.column_dimensions["C"].width = 32

    # 「サマリ」は wb.active として index 0 に作られる。最後尾へ回し、開いたときに
    # 差異一覧が出るようにする（2026-08-20 谷津さん要望）。offset は相対移動量なので
    # 「サマリが先頭にいる」前提でシート数-1 だけ後ろへずらす。
    wb.move_sheet("サマリ", offset=len(wb.sheetnames) - 1)
    wb.active = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ----------------------------------------------------------------------
# 共通実行関数（CLI / Flask 共用）
# ----------------------------------------------------------------------

@dataclass
class CompareResult:
    ok: bool
    output_path: Path
    diff_count: int = 0
    danger_count: int = 0
    warn_count: int = 0
    info_count: int = 0
    kintai_rows_read: int = 0
    jinjer_rows_read: int = 0
    name_map_size: int = 0
    error: str = ""
    logs: list[LogEntry] = field(default_factory=list)
    # 「jinjer未登録」（請求勤怠に勤務があるが氏名→従業員IDを解決できない）人の氏名。
    # 一括モードが「未処理・未マッチ」シートに一覧として載せる。
    unmatched_names: list[str] = field(default_factory=list)


def run_quick_compare(
    kintai_dir: Path,
    jinjer_dir: Path,
    output_path: Path,
    month_label: str,
    log_func=print,
    application_csv: Path | None = None,
    threshold_minutes: int = DEFAULT_RECOMMEND_THRESHOLD_MIN,
) -> CompareResult:
    """突合結果xlsx + jinjer CSV → 差異一覧xlsx を生成する。CLI と Web UI から共用。

    application_csv: jinjer「申請データ（打刻修正申請）」CSV のパス（任意）。
    指定すると「打刻修正時コメント」列にその「理由」を併記する。未指定なら attendances API
    の打刻コメントにフォールバックする。

    threshold_minutes: 自動修正提案値（採用ラベル）の許容しきい値。出退勤の差分が
    この分数以内なら「請求勤怠」、超過なら「jinjer勤怠」を提案する（差分勤怠チェッカーの選択範囲）。
    """
    logs: list[LogEntry] = []
    result = CompareResult(ok=False, output_path=output_path, logs=logs)

    log_func(f"[start] 勤怠突合結果xlsx/フォルダ: {kintai_dir}")
    log_func(f"[start] jinjer CSV/フォルダ: {jinjer_dir}")
    log_func(f"[start] 出力先: {output_path}")

    kintai_df = load_kintai_results(kintai_dir, logs)
    if kintai_df.empty:
        msg = "突合結果xlsx が読めませんでした"
        log_func(f"[error] {msg}")
        result.error = msg
        write_excel(output_path, [], logs, month_label)
        return result

    jinjer_df = load_jinjer_csvs(jinjer_dir, logs)
    if jinjer_df.empty:
        msg = "jinjer CSV が読めませんでした"
        log_func(f"[error] {msg}")
        result.error = msg
        write_excel(output_path, [], logs, month_label)
        return result

    result.kintai_rows_read = len(kintai_df)
    result.jinjer_rows_read = len(jinjer_df)
    log_func(f"[info] 突合結果 {len(kintai_df)} 行 / jinjer {len(jinjer_df)} 行 読み込み完了")

    name_map = build_name_to_emp_map(jinjer_df)
    jinjer_index = build_jinjer_index(jinjer_df)
    result.name_map_size = len(name_map)
    log_func(f"[info] 氏名→ID マップ {len(name_map)} 件 / jinjer index {len(jinjer_index)} 件")

    # 汎用データから転記する予定・有休 列を解決（見つからない列は空欄転記＋ログ）
    extra_cols = resolve_jinjer_extra_columns(jinjer_df.columns)
    missing_extra = [k for k in JINJER_EXTRA_COLUMNS if k not in extra_cols]
    if missing_extra:
        logs.append(LogEntry(
            "INFO",
            f"汎用データに次の列が見つからないため空欄で転記します: {', '.join(missing_extra)}",
        ))
    log_func(f"[info] 予定/有休 転記列の解決: {extra_cols}")

    # 「打刻修正時コメント」列の元データ。申請データCSVが指定されていればその「理由」を最優先。
    # 未指定なら attendances API の打刻コメントにフォールバック。いずれも取得できなくても続行。
    stamp_comments: dict[tuple[str, str], list[dict]] = {}
    if application_csv is not None:
        stamp_comments = load_stamp_correction_reasons(application_csv, logs)
        log_func(f"[info] 申請データCSVから打刻修正理由: {len(stamp_comments)} (emp,date) 件")
    else:
        checked_emp_ids = sorted({
            eid for n in kintai_df["氏名"].dropna().unique()
            if (eid := resolve_emp_id(str(n).strip(), name_map))
        })
        if checked_emp_ids and re.fullmatch(r"\d{4}-\d{2}", str(month_label or "").strip()):
            try:
                from services.jinjer_api_client import fetch_stamp_comments
                stamp_comments = fetch_stamp_comments(checked_emp_ids, str(month_label).strip())
                log_func(f"[info] jinjer 打刻コメント取得(API): {len(stamp_comments)} (emp,date) 件")
            except Exception as e:  # 認証失敗・通信不可など。差異一覧は続行する。
                logs.append(LogEntry("WARN", f"jinjer 打刻コメント取得に失敗（コメント欄は空で続行）: {e}"))
                log_func(f"[warn] jinjer 打刻コメント取得に失敗: {e}")
        else:
            logs.append(LogEntry(
                "INFO",
                f"打刻コメント取得をスキップ（申請データCSV未指定 / 対象月={month_label} / 対象ID数={len(checked_emp_ids)}）",
            ))

    diff_rows = compute_diffs(
        kintai_df, jinjer_index, name_map, logs, extra_cols, stamp_comments,
        threshold_minutes=threshold_minutes,
    )
    result.diff_count = len(diff_rows)
    log_func(f"[info] 差異・警告 合計 {len(diff_rows)} 件")

    # jinjer未登録（氏名→従業員ID を解決できない）人の氏名。1人1行に集約済みだが
    # 念のため重複を除いて出現順を保つ。一括モードが未処理・未マッチシートに載せる。
    result.unmatched_names = list(dict.fromkeys(
        r.name for r in diff_rows if r.kind == DIFF_KIND_UNMATCHED and r.name
    ))
    if result.unmatched_names:
        log_func(f"[info] jinjer勤怠未登録者 {len(result.unmatched_names)} 名")

    by_level = {LEVEL_DANGER: 0, LEVEL_WARN: 0, LEVEL_INFO: 0}
    for r in diff_rows:
        by_level[r.warn_level] = by_level.get(r.warn_level, 0) + 1
    result.danger_count = by_level[LEVEL_DANGER]
    result.warn_count = by_level[LEVEL_WARN]
    result.info_count = by_level[LEVEL_INFO]
    log_func(f"[info] DANGER={result.danger_count} / WARN={result.warn_count} / INFO={result.info_count}")

    write_excel(output_path, diff_rows, logs, month_label)
    log_func(f"[done] 出力完了: {output_path}")
    result.ok = True
    return result


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="5月本番 MVP 差異一覧生成")
    p.add_argument("--month", required=True, help="対象月 YYYY-MM (例: 2026-05)")
    p.add_argument("--kintai-dir", required=True, help="突合結果xlsx 格納フォルダ")
    p.add_argument("--jinjer-dir", required=True, help="jinjer 汎用データCSV 格納フォルダ")
    p.add_argument("--application-csv", default="", help="jinjer 申請データ(打刻修正申請) CSV（任意）")
    p.add_argument("--output", required=True, help="出力 xlsx パス")
    return p.parse_args()


def _unquote_path(s: str) -> str:
    """前後の空白と、前後が同じクォートで囲まれているときだけ1組を除去する。"""
    s = (s or "").strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        s = s[1:-1].strip()
    return s


def main() -> int:
    args = parse_args()
    app_csv = _unquote_path(args.application_csv)
    result = run_quick_compare(
        kintai_dir=Path(_unquote_path(args.kintai_dir)),
        jinjer_dir=Path(_unquote_path(args.jinjer_dir)),
        output_path=Path(_unquote_path(args.output)),
        month_label=args.month,
        application_csv=Path(app_csv) if app_csv else None,
    )
    return 0 if result.ok else 1


if __name__ == "__main__":
    sys.exit(main())
