# -*- coding: utf-8 -*-
"""健康診断HPM変換マスタ（Excel）を作る

    python tools/build_health_hpm_master.py            # 既定パスへ作成
    python tools/build_health_hpm_master.py --show     # 中身を表示するだけ
    python tools/build_health_hpm_master.py --force    # 既存を上書き（バックアップを取る）

302列ヘッダー・項目マッピング・健診機関コード・健診コースをExcelに外出しし、
HPM側で列や機関が増えてもマスタを直すだけで済むようにする（exe再ビルド不要）。

初期データの出どころ:
  - 302列ヘッダー … HPM取込用CSVの1行目（tools/data/hpm_canonical_header_302.json）
  - 項目名        … health-check-pdf-to-excel が実際に吐く表記（整形済Excelの項目別データ）
  - 機関コード    … 過去に作成したHPM取込用CSVの場所コード
  - コース値      … 同じくCSVの健診コースコード（同友会はコース名テキスト、他は数値）
"""

from __future__ import annotations

import argparse
import datetime
import json
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
DEFAULT_HEADER_JSON = REPO / "tools" / "data" / "hpm_canonical_header_302.json"
DEFAULT_OUT = Path(r"Z:\NMHT総務関係\健康診断\健康診断HPM変換マスタ.xlsx")
DEFAULT_SOURCE_CSV = Path(
    r"Z:\NMHT総務関係\健康診断\2026\2026年度健康診断受診者結果\CSV格納"
    r"\HPM取込用_同友会_20260701-0703_6名.csv"
)

TITLE_FILL = PatternFill("solid", fgColor="FF1F4E79")
HEADER_FILL = PatternFill("solid", fgColor="FFD9E2F3")
WARN_FILL = PatternFill("solid", fgColor="FFFCEBEB")


# --- 健診機関（機関名 / HPM場所コード / 公開医療機関コード参考 / HPM確認済み / 備考） ---
INSTITUTIONS = [
    ("医療法人社団 同友会 春日クリニック", "1310528885", "", "TRUE",
     "2026-07受診分で使用。当初1310528018で作成し修正版で1310528885に訂正した（谷津さん確認済み）"),
    ("ヘルスケアクリニック厚木", "1412910586", "", "TRUE",
     "2026-05受診（岡田さん）で使用"),
    ("桜十字グランフロント大阪クリニック", "13X5035440", "", "TRUE",
     "2026-05受診（山長さん・健保内健診）で使用。コードに英字Xを含むので文字列で扱う"),
    ("医療法人徳洲会 生駒市立病院", "0301619", "", "TRUE",
     "2026-05受診（松井さん・イーウェル人間ドック経由）で使用。前ゼロが落ちないよう文字列。"
     "健診機関一覧には同じ社員番号で石上クリニックの行もあるため、次回利用時に要確認"),
    ("IMS Me-Lifeクリニック千葉", "1220700072", "", "TRUE",
     "2026-05受診（大村さん・IMS千葉）で使用。旧称 千葉ロイヤルクリニック"),
]

# --- 機関別名（整形済Excelやメモで使われる略称 → 正式機関名） ---
ALIASES = [
    ("同友会", "医療法人社団 同友会 春日クリニック"),
    ("春日クリニック", "医療法人社団 同友会 春日クリニック"),
    ("ヘルスケア厚木", "ヘルスケアクリニック厚木"),
    ("健保内健診", "桜十字グランフロント大阪クリニック"),
    ("イーウェル人間ドック", "医療法人徳洲会 生駒市立病院"),
    ("生駒市立病院", "医療法人徳洲会 生駒市立病院"),
    ("IMS千葉", "IMS Me-Lifeクリニック千葉"),
    ("千葉ロイヤルクリニック", "IMS Me-Lifeクリニック千葉"),
]

# --- 健診種別（機関名 / 種別表示名 / HPM出力値） ---
# 同友会はコース名テキストをそのまま、他機関は数値コードを出す（実CSVがそうなっている）
COURSES = [
    ("医療法人社団 同友会 春日クリニック", "人間ドックC 胃カメラ（40歳以上）",
     "人間ドックＣ\u3000胃カメラ（４０歳以上）"),
    ("ヘルスケアクリニック厚木", "定期健康診断", "10"),
    ("桜十字グランフロント大阪クリニック", "人間ドックB（胃X線）", "12"),
    ("医療法人徳洲会 生駒市立病院", "人間ドックC（胃カメラ）", "13"),
    ("IMS Me-Lifeクリニック千葉", "人間ドックC（胃カメラ）", "13"),
]

SETTINGS = [
    ("会場コード", "2", "HPM列21に出す固定値"),
]

# --- 項目マッピング（分類 / 項目名 / 測定回 / HPM列番号 / 値種別 / 検査方式 / 備考） ---
# 項目名は health-check-pdf-to-excel が実際に吐く表記に合わせること。
# 表記が変わったらこのシートを直す（コードは触らない）。
ITEM_MAP = [
    ("身体計測", "身長", 1, 24, "数値", "", ""),
    ("身体計測", "体重", 1, 25, "数値", "", ""),
    ("身体計測", "BMI", 1, 27, "数値", "", "列183も『ＢＭＩ』だがそちらは判定欄なので使わない"),
    ("身体計測", "腹囲", 1, 28, "数値", "", ""),

    ("視力", "視力（左）（裸眼）", 1, 29, "数値", "", ""),
    ("視力", "視力（右）（裸眼）", 1, 30, "数値", "", ""),
    ("視力", "視力（左）（矯正）", 1, 31, "数値", "", ""),
    ("視力", "視力（右）（矯正）", 1, 32, "数値", "", ""),

    ("肺機能", "努力性肺活量", 1, 45, "数値", "", ""),
    ("肺機能", "%肺活量", 1, 46, "数値", "", ""),
    ("肺機能", "1秒率", 1, 47, "数値", "", ""),
    ("肺機能", "1秒量", 1, 48, "数値", "", ""),

    # 血圧は1回目・2回目を別の列に出す。平均は作らない。ここを動かすと読み込みが止まる。
    ("血圧", "収縮期血圧", 1, 50, "数値", "", "1回目。列を変えてはいけない"),
    ("血圧", "拡張期血圧", 1, 51, "数値", "", "1回目。列を変えてはいけない"),
    ("血圧", "収縮期血圧", 2, 52, "数値", "", "2回目。1回目の複製・平均は禁止"),
    ("血圧", "拡張期血圧", 2, 53, "数値", "", "2回目。1回目の複製・平均は禁止"),

    ("血液一般", "赤血球数", 1, 74, "数値", "", ""),
    ("血液一般", "白血球数", 1, 75, "数値", "", ""),
    ("血液一般", "血色素量", 1, 76, "数値", "", ""),
    ("血液一般", "ヘマトクリット", 1, 77, "数値", "", ""),
    ("血液一般", "MCV", 1, 78, "数値", "", ""),
    ("血液一般", "MCH", 1, 79, "数値", "", ""),
    ("血液一般", "MCHC", 1, 80, "数値", "", ""),
    ("血液一般", "血小板数", 1, 82, "数値", "", ""),

    ("肝機能", "総蛋白", 1, 100, "数値", "", ""),
    ("肝機能", "総ビリルビン", 1, 102, "数値", "", ""),
    ("肝機能", "AST", 1, 107, "数値", "", ""),
    ("肝機能", "ALT", 1, 108, "数値", "", ""),
    ("肝機能", "γ-GT", 1, 111, "数値", "", ""),
    ("肝機能", "ChE", 1, 113, "数値", "",
     "列112も同名『コリンエステラーゼ』。過去のCSVは113を使っている"),
    ("肝機能", "ALP", 1, 290, "数値", "IFCC",
     "旧法の列109ではなくIFCC法の290。原票が旧法なら109へ変更する"),
    ("肝機能", "LD", 1, 291, "数値", "IFCC",
     "旧法の列110ではなくIFCC法の291"),
    ("肝機能", "アルブミン", 1, 292, "数値", "BCP改", "列292は『アルブミン［ＢＣＰ改］』"),

    ("腎機能", "尿素窒素", 1, 124, "数値", "", ""),
    ("腎機能", "クレアチニン", 1, 125, "数値", "", ""),
    ("腎機能", "eGFR", 1, 126, "数値", "", ""),

    ("脂質", "総コレステロール", 1, 131, "数値", "", ""),
    ("脂質", "中性脂肪", 1, 132, "数値", "", ""),
    ("脂質", "HDL-C", 1, 133, "数値", "", ""),
    ("脂質", "LDL-C", 1, 134, "数値", "", ""),

    ("痛風", "尿酸", 1, 136, "数値", "", ""),

    ("糖代謝", "空腹時血糖", 1, 138, "数値", "", ""),
    ("糖代謝", "HbA1c", 1, 144, "数値", "", "列144はNGSP値"),

    ("血清反応", "CRP", 1, 148, "数値", "", "定量。定性は列149"),

    # --- ここから定性検査（スキーマv2で取り込む。(-)は陰性としてそのまま出す） ---
    ("尿検査", "尿蛋白", 1, 54, "定性", "", ""),
    ("尿検査", "尿糖", 1, 55, "定性", "", ""),
    ("尿検査", "尿ウロビリノーゲン", 1, 56, "定性", "", ""),
    ("尿検査", "尿潜血", 1, 57, "定性", "", ""),
    ("尿検査", "尿ビリルビン", 1, 58, "定性", "", ""),
    ("尿検査", "尿ケトン体", 1, 61, "定性", "", ""),

    ("便潜血", "便潜血", 1, 70, "定性", "", "1回目の検体"),
    ("便潜血", "便潜血", 2, 71, "定性", "", "2回目の検体。測定回で振り分ける"),

    ("感染症", "HBs抗原", 1, 115, "定性", "MAT", "検査方式が原票に明記されている場合だけ出す"),
    ("感染症", "HBs抗原", 1, 117, "定性", "CLIA", "同上。方式不明なら出さずに警告する"),
    ("感染症", "HBs抗体", 1, 116, "定性", "PHA", ""),
    ("感染症", "HBs抗体", 1, 119, "定性", "CLIA", ""),
    ("感染症", "HCV抗体", 1, 121, "定性", "CLEIA", ""),
    ("感染症", "HCV抗体", 1, 123, "定性", "LPIA", ""),
    ("感染症", "ヘリコバクターピロリ", 1, 179, "定性", "", "列179はH.P判定"),

    ("血清反応", "CRP（定性）", 1, 149, "定性", "", ""),
    ("血清反応", "RPR", 1, 153, "定性", "", ""),
    ("血清反応", "TPHA", 1, 154, "定性", "", ""),

    ("腫瘍マーカー", "AFP（定性）", 1, 161, "定性", "", ""),
]


def _sheet(wb, title, description, headers, rows, widths=None, text_columns=()):
    """1行目タイトル / 2行目説明 / 3行目ヘッダー / 4行目〜データ の共通レイアウト。"""
    ws = wb.create_sheet(title)
    cell = ws.cell(1, 1)
    cell.value = title
    cell.font = Font(name="Meiryo UI", size=12, bold=True, color="FFFFFFFF")
    cell.fill = TITLE_FILL
    ws.cell(2, 1).value = description
    ws.cell(2, 1).font = Font(name="Meiryo UI", size=9, color="FF555555")

    for c, name in enumerate(headers, start=1):
        h = ws.cell(3, c)
        h.value = name
        h.font = Font(name="Meiryo UI", size=9.5, bold=True)
        h.fill = HEADER_FILL
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row in enumerate(rows, start=4):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c)
            cell.value = value
            cell.font = Font(name="Meiryo UI", size=9.5)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c in text_columns:
                # "0301619" の前ゼロ・"13X5035440" の英字を守る
                cell.number_format = "@"

    for c, width in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A4"
    return ws


def build(header_columns: list[str]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet(
        wb, "健診機関",
        "HPMの場所コード。HPM確認済み=TRUE の機関だけCSVを作れます。"
        "新しい施設は https://www.iryokikan.info/ などで人が調べ、HPMで確認してから TRUE にしてください。",
        ["機関名", "HPM場所コード", "公開医療機関コード参考", "HPM確認済み", "備考"],
        INSTITUTIONS,
        widths=[38, 18, 22, 14, 60],
        text_columns=(2, 3),
    )

    _sheet(
        wb, "機関別名",
        "健診結果Excelやメモで使われる略称を正式機関名に読み替えます。増やすほど選択の手間が減ります。",
        ["別名", "正式機関名", "備考"],
        [(a, b, "") for a, b in ALIASES],
        widths=[24, 38, 40],
    )

    _sheet(
        wb, "健診種別",
        "HPM列18（健診コースコード）に出す値。同友会はコース名テキスト、他機関は数値コードという"
        "不統一を、この表で吸収します。",
        ["機関名", "種別表示名", "HPM出力値", "備考"],
        [(a, b, c, "") for a, b, c in COURSES],
        widths=[38, 30, 34, 30],
        text_columns=(3,),
    )

    _sheet(
        wb, "設定",
        "モード全体の固定値。",
        ["キー", "値", "備考"],
        SETTINGS,
        widths=[20, 16, 44],
        text_columns=(2,),
    )

    _sheet(
        wb, "項目マッピング",
        "健診結果Excelの（分類・項目名・測定回）を、HPMの列番号（0始まり）へ対応づけます。"
        "血圧の50〜53と、判定列183〜197・識別列0〜23を指す設定は読み込み時に拒否されます。",
        ["分類", "項目名", "測定回", "HPM列番号", "値種別", "検査方式", "備考"],
        [(a, b, c, d, e, f, g) for a, b, c, d, e, f, g in ITEM_MAP],
        widths=[14, 24, 8, 12, 10, 12, 52],
    )

    ws = _sheet(
        wb, "HPMヘッダー",
        "HPM取込用CSVの302列。列名には重複があるため、対応づけは必ず列番号（0始まり）で行います。"
        "この表を書き換えると出力CSVのヘッダーも変わります。",
        ["列番号", "列名", "用途"],
        [(i, name, _column_role(i)) for i, name in enumerate(header_columns)],
        widths=[10, 40, 34],
    )
    # 血圧と判定列に色を付けて、触ってはいけないことを見て分かるようにする
    for i in range(len(header_columns)):
        role = _column_role(i)
        if role:
            for c in range(1, 4):
                ws.cell(4 + i, c).fill = WARN_FILL

    return wb


def _column_role(index: int) -> str:
    if index in (50, 51):
        return "血圧1回目（平均を作らない）"
    if index in (52, 53):
        return "血圧2回目（1回目を複製しない）"
    if 183 <= index <= 197:
        return "判定列（原票A〜Gは転記しない）"
    if index in (6, 7, 8, 9, 10, 18, 19, 20, 21, 23):
        return "識別列（システムが埋める）"
    return ""


def load_header(path: Path, source_csv: Path | None) -> list[str]:
    """正本CSVがあればそこから、無ければ同梱JSONから302列を得る。"""
    if source_csv and source_csv.exists():
        import csv
        import io
        text = source_csv.read_bytes().decode("cp932")
        columns = next(csv.reader(io.StringIO(text)))
        if len(columns) != 302:
            raise SystemExit(f"正本CSVが302列ではありません（{len(columns)}列）: {source_csv}")
        print(f"ヘッダー取得元: {source_csv}")
        return columns
    data = json.loads(path.read_text(encoding="utf-8"))
    columns = data["columns"]
    if len(columns) != 302:
        raise SystemExit(f"{path} が302列ではありません（{len(columns)}列）")
    print(f"ヘッダー取得元: {path}")
    return columns


def main() -> int:
    ap = argparse.ArgumentParser(description="健康診断HPM変換マスタを作る")
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--source-csv", default=str(DEFAULT_SOURCE_CSV),
                    help="302列ヘッダーの取得元CSV。無ければ同梱JSONを使う")
    ap.add_argument("--header-json", default=str(DEFAULT_HEADER_JSON))
    ap.add_argument("--force", action="store_true", help="既存ファイルを上書きする（バックアップを取る）")
    ap.add_argument("--show", action="store_true", help="作らずに内容だけ表示する")
    args = ap.parse_args()

    header_columns = load_header(Path(args.header_json), Path(args.source_csv))

    if args.show:
        print(f"\n健診機関 {len(INSTITUTIONS)}件 / 別名 {len(ALIASES)}件 / "
              f"健診種別 {len(COURSES)}件 / 項目マッピング {len(ITEM_MAP)}件")
        for name, code, _, confirmed, _note in INSTITUTIONS:
            print(f"  {name} → {code} (HPM確認済み={confirmed})")
        numeric = sum(1 for r in ITEM_MAP if r[4] == "数値")
        print(f"  項目マッピング: 数値{numeric}件 / 定性{len(ITEM_MAP) - numeric}件")
        return 0

    out = Path(args.out)
    if out.exists():
        if not args.force:
            print(f"[中断] 既にあります: {out}\n  上書きするなら --force を付けてください。")
            return 1
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = out.with_name(f"{out.stem}_backup_{stamp}{out.suffix}")
        shutil.copy2(out, backup)
        print(f"バックアップ: {backup.name}")

    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build(header_columns)
    wb.save(out)
    print(f"作成: {out}")

    # 作ったものを読み込みモジュールで検証する（壊れたマスタを置いて帰らない）
    sys.path.insert(0, str(REPO))
    from services.health_hpm_master import load_master

    master = load_master(str(out))
    print(f"検証OK: 機関{len(master.institutions)} / 別名{len(master.aliases)} / "
          f"種別{sum(len(v) for v in master.courses.values())} / "
          f"項目{len(master.item_map)} / ヘッダー{len(master.header)}列 / "
          f"会場コード={master.venue_code}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
