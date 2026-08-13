# -*- coding: utf-8 -*-
"""スケジュール開始合わせ（実績が予定より早い日）のテスト。

2026-08-13 谷津さん指定:
  勤務開始（請求勤怠）がスケジュール開始より早い → 出勤予定時刻を請求勤怠に合わせる。
  遅い（遅刻方向）は動かさない。打刻には触らない。
"""
import sys
from datetime import date
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from quick_compare import (
    DIFF_KIND_PUNCH_IN,
    DIFF_KIND_SCHED_START,
    JINJER_HEADERS,
    LogEntry,
    compute_diffs,
    resolve_jinjer_extra_columns,
)
from services.triage import JUDGE_KINTAI, TRIAGE_AUTO_KINTAI, TRIAGE_NEEDS_CHECK, classify


def _jrow(**overrides):
    row = {
        JINJER_HEADERS["name"]: "上原 奏吾",
        JINJER_HEADERS["emp_id"]: "2018057",
        JINJER_HEADERS["date"]: "2026/4/1",
        JINJER_HEADERS["punch_in_1"]: "8:30",
        JINJER_HEADERS["punch_out_1"]: "18:00",
        JINJER_HEADERS["break_total"]: "",
        JINJER_HEADERS["total_work"]: "",
        JINJER_HEADERS["finalized"]: "",
        "出勤予定時刻": "9:00",
        "退勤予定時刻": "18:00",
    }
    row.update(overrides)
    return row


def _kdf(k_in="8:30", j_in="8:30", diff_in=0):
    return pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": k_in,
        "jinjer_出勤": j_in,
        "出勤差分(分)": diff_in,
        "勤務表_退勤": "18:00",
        "jinjer_退勤": "18:00",
        "退勤差分(分)": 0,
        "_source_file": "x.xlsx",
    }])


def _run(kdf, jrow):
    logs: list[LogEntry] = []
    return compute_diffs(
        kdf,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        resolve_jinjer_extra_columns(list(jrow.keys())),
    )


def test_early_actual_emits_sched_start_row_auto_adopted():
    """実績 8:30 < 予定 9:00（打刻は一致）→ スケジュール開始行が自動採用で出る。"""
    rows = _run(_kdf(), _jrow())
    sched = [r for r in rows if r.kind == DIFF_KIND_SCHED_START]
    assert len(sched) == 1
    r = sched[0]
    assert r.kintai_value == "8:30"      # 合わせ先（請求勤怠の開始）
    assert r.jinjer_value == "9:00"      # 現在のスケジュール開始
    assert r.auto_fix_value == "8:30"
    assert r.diff_minutes == "-30"
    assert r.triage == TRIAGE_AUTO_KINTAI
    assert r.judge_default == JUDGE_KINTAI
    assert "打刻は触らない" in r.warn_reason


def test_late_actual_does_not_move_schedule():
    """実績 9:15 > 予定 9:00（遅刻方向）→ スケジュールは動かさない＝行を出さない。"""
    rows = _run(_kdf(k_in="9:15", j_in="9:15"), _jrow(**{JINJER_HEADERS["punch_in_1"]: "9:15"}))
    assert [r for r in rows if r.kind == DIFF_KIND_SCHED_START] == []


def test_equal_start_no_row():
    rows = _run(_kdf(k_in="9:00", j_in="9:00"), _jrow(**{JINJER_HEADERS["punch_in_1"]: "9:00"}))
    assert [r for r in rows if r.kind == DIFF_KIND_SCHED_START] == []


def test_suppressed_when_punch_diff_row_exists():
    """出勤差異の行が出た日は出さない（打刻採用のおまけ同期に任せる）。"""
    rows = _run(_kdf(k_in="8:30", j_in="8:45", diff_in=-15),
                _jrow(**{JINJER_HEADERS["punch_in_1"]: "8:45"}))
    assert [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert [r for r in rows if r.kind == DIFF_KIND_SCHED_START] == []


def test_no_schedule_or_no_sched_out_no_row():
    # スケジュール未設定
    rows = _run(_kdf(), _jrow(出勤予定時刻="", 退勤予定時刻=""))
    assert [r for r in rows if r.kind == DIFF_KIND_SCHED_START] == []
    # 退勤予定だけ空（出勤予定のみ書くと jinjer が行ごと弾く）
    rows = _run(_kdf(), _jrow(退勤予定時刻=""))
    assert [r for r in rows if r.kind == DIFF_KIND_SCHED_START] == []


def test_triage_comment_forces_needs_check():
    """打刻時コメントがある日は自動採用にせず人が読む。"""
    assert classify(kind=DIFF_KIND_SCHED_START, warn_level="INFO") == \
        (TRIAGE_AUTO_KINTAI, JUDGE_KINTAI)
    triage, judge = classify(kind=DIFF_KIND_SCHED_START, warn_level="INFO",
                             punch_comment="現場都合で早出")
    assert (triage, judge) == (TRIAGE_NEEDS_CHECK, "")


# ---------------------------------------------------------------------------
# write_excel: 差異一覧の行には出さず、別シートに記録する（2026-08-13 谷津さん指定）
# ---------------------------------------------------------------------------

def test_write_excel_moves_sched_rows_to_separate_sheet(tmp_path):
    import openpyxl
    from quick_compare import write_excel

    rows = _run(_kdf(), _jrow())
    assert [r for r in rows if r.kind == DIFF_KIND_SCHED_START]
    out = tmp_path / "diff.xlsx"
    write_excel(out, rows, [], "2026-04")
    wb = openpyxl.load_workbook(out)
    # 差異一覧シートに「スケジュール開始」行は出ない（確認対象を増やさない）
    ws = wb["差異一覧"]
    kinds = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert DIFF_KIND_SCHED_START not in kinds
    # 別シートに記録される
    assert "スケジュール開始合わせ" in wb.sheetnames
    sa = wb["スケジュール開始合わせ"]
    assert sa.cell(row=2, column=1).value == "2018057"
    assert sa.cell(row=2, column=3).value == "2026-04-01"
    assert sa.cell(row=2, column=4).value == "9:00"    # 現在の予定開始
    assert sa.cell(row=2, column=5).value == "8:30"    # 新しい開始(請求勤怠)
    # サマリに件数（別シート・自動反映）
    ws_sum = wb["サマリ"]
    labels = {ws_sum.cell(row=r, column=1).value: ws_sum.cell(row=r, column=2).value
              for r in range(1, ws_sum.max_row + 1)}
    assert labels.get("スケジュール開始合わせ件数（別シート・自動反映）") == 1


def test_write_excel_no_sheet_when_no_aligns(tmp_path):
    import openpyxl
    from quick_compare import write_excel

    rows = _run(_kdf(k_in="9:00", j_in="9:00"), _jrow(**{JINJER_HEADERS["punch_in_1"]: "9:00"}))
    out = tmp_path / "diff.xlsx"
    write_excel(out, rows, [], "2026-04")
    wb = openpyxl.load_workbook(out)
    assert "スケジュール開始合わせ" not in wb.sheetnames


# ---------------------------------------------------------------------------
# 別シートの読み込みと適用（quick_export）
# ---------------------------------------------------------------------------

def _align(new="8:30"):
    return {"emp": "2018057", "date_iso": "2026-04-01", "new_start": new, "name": "上原"}


def _sheet_env():
    from quick_export import Stats, build_jinjer_row_index
    headers = ["*従業員ID", "*年月日", "出勤予定時刻", "退勤予定時刻", "出勤1", "退勤1"]
    rows = [["2018057", "2026/4/1", "9:00", "18:00", "8:30", "18:00"]]
    return headers, rows, build_jinjer_row_index(headers, rows), Stats()


def test_apply_sched_aligns_writes_schedule_only():
    from quick_export import apply_sched_aligns
    headers, rows, idx, stats = _sheet_env()
    changed = set()
    apply_sched_aligns(headers, rows, idx, [_align()], set(), stats, changed)
    assert rows[0][headers.index("出勤予定時刻")] == "8:30"
    assert rows[0][headers.index("出勤1")] == "8:30"          # 打刻は触らない
    assert stats.overwritten_sched_start == 1
    assert ("2018057", "2026-04-01") in changed


def test_apply_sched_aligns_held_day_skipped():
    from quick_export import apply_sched_aligns
    headers, rows, idx, stats = _sheet_env()
    changed = set()
    apply_sched_aligns(headers, rows, idx, [_align()], {("2018057", "2026-04-01")}, stats, changed)
    assert rows[0][headers.index("出勤予定時刻")] == "9:00"   # 保留日は触らない
    assert stats.overwritten_sched_start == 0 and changed == set()


def test_apply_sched_aligns_direction_guard():
    """遅刻方向（新しい開始が予定以降）は適用時にも動かさない（古い差異一覧の保険）。"""
    from quick_export import apply_sched_aligns
    headers, rows, idx, stats = _sheet_env()
    changed = set()
    apply_sched_aligns(headers, rows, idx, [_align(new="9:30")], set(), stats, changed)
    assert rows[0][headers.index("出勤予定時刻")] == "9:00"
    assert stats.overwritten_sched_start == 0


def test_load_sched_aligns_roundtrip(tmp_path):
    """write_excel が書いた別シートを quick_export が読める（Excel日付型も正規化）。"""
    from quick_compare import write_excel
    from quick_export import load_sched_aligns

    rows = _run(_kdf(), _jrow())
    out = tmp_path / "diff.xlsx"
    write_excel(out, rows, [], "2026-04")
    aligns = load_sched_aligns(out)
    assert aligns == [{"emp": "2018057", "date_iso": "2026-04-01",
                       "new_start": "8:30", "name": "上原 奏吾"}]
    # シートが無い旧フォーマットは空リスト
    write_excel(out, [r for r in rows if r.kind != DIFF_KIND_SCHED_START], [], "2026-04")
    assert load_sched_aligns(out) == []


# ---------------------------------------------------------------------------
# 書き戻し（quick_export・差異一覧行としての後方互換）
# ---------------------------------------------------------------------------

def _approved(kind, auto_fix, manual_fix=""):
    from quick_export import ApprovedRow
    return ApprovedRow(emp_id="2018057", target_date_iso="2026-04-01", kind=kind,
                       auto_fix_value=auto_fix, manual_fix_value=manual_fix,
                       manual_break_start="", manual_break_end="", manual_break_total="",
                       name="上原", warn_level="", source_diff_row_id=1)


def test_apply_writes_schedule_only_not_punch():
    from quick_export import DIFF_KIND_SCHED_START as K, Stats, apply_approved_rows, build_jinjer_row_index

    headers = ["*従業員ID", "*年月日", "出勤予定時刻", "退勤予定時刻", "出勤1", "退勤1"]
    rows = [["2018057", "2026/4/1", "9:00", "18:00", "8:30", "18:00"]]
    idx = build_jinjer_row_index(headers, rows)
    stats = Stats()
    apply_approved_rows(headers, rows, idx, [_approved(K, "8:30")], stats)
    assert rows[0][headers.index("出勤予定時刻")] == "8:30"   # 予定だけ動く
    assert rows[0][headers.index("出勤1")] == "8:30"          # 打刻はそのまま
    assert rows[0][headers.index("退勤予定時刻")] == "18:00"
    assert stats.overwritten_sched_start == 1
    assert stats.overwritten_punch_in == 0


def test_apply_skips_when_sched_out_missing():
    """退勤予定が空の行は書かず警告（jinjer が行ごと弾くため）。"""
    from quick_export import DIFF_KIND_SCHED_START as K, Stats, apply_approved_rows, build_jinjer_row_index

    headers = ["*従業員ID", "*年月日", "出勤予定時刻", "退勤予定時刻", "出勤1", "退勤1"]
    rows = [["2018057", "2026/4/1", "9:00", "", "8:30", "18:00"]]
    idx = build_jinjer_row_index(headers, rows)
    stats = Stats()
    apply_approved_rows(headers, rows, idx, [_approved(K, "8:30")], stats)
    assert rows[0][headers.index("出勤予定時刻")] == "9:00"
    assert stats.overwritten_sched_start == 0
    assert any("退勤予定が空" in w for w in stats.warnings)
