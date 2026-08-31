"""BBS（ブロードバンドセキュリティ）勤務表パーサの回帰テスト。

主眼は「行頭が空欄の人だけ1日ズレる」事故の再発防止。
AI読み取り（CSV風テキスト→Claude）は連続カンマを数え違えるため、
列位置で読む構造化パースに寄せた（2026-08-31）。
"""

import calendar
from datetime import date

import openpyxl
import pytest

from services.bbs_shift_parser import (
    is_bbs_shift_xlsx,
    parse_bbs_shift_xlsx,
)
from services.multi_year_shift_parser import parse_structured_files

_WEEKDAY_KANJI = ["月", "火", "水", "木", "金", "土", "日"]
_COL_DAY1 = 4  # D列＝1日


def _make_bbs_book(
    path,
    year=2026,
    month=9,
    employees=(("2", "(N)加藤 英人", {2: "A", 3: "A", 4: "A", 7: "B"}, {}),),
    legend_rows=True,
    weekday_year=None,
):
    """BBS勤務表 xlsx を組み立てる

    employees は (枠, 氏名, {日: 計画コード}, {日: リーダー行の値}) のタプル列。
    weekday_year を渡すと曜日行だけ別年で書ける（曜日不一致テスト用）。
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"{year}-{month:02d}"

    worksheet.cell(1, 1).value = f"令和8年({year})年{month:02d}月勤務表"
    worksheet.cell(2, 1).value = "枠"
    worksheet.cell(2, 2).value = "日･曜日　　　　氏　名"
    for day in range(1, 32):
        worksheet.cell(2, _COL_DAY1 + day - 1).value = day
    worksheet.cell(2, _COL_DAY1 + 31).value = "労働日数"

    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        weekday = date(weekday_year or year, month, day).weekday()
        worksheet.cell(3, _COL_DAY1 + day - 1).value = _WEEKDAY_KANJI[weekday]

    row = 5
    for waku, name, plan, leader in employees:
        worksheet.cell(row, 1).value = waku
        worksheet.cell(row, 2).value = name
        worksheet.cell(row, 3).value = "計画"
        for day, code in plan.items():
            worksheet.cell(row, _COL_DAY1 + day - 1).value = code
        worksheet.cell(row + 1, 3).value = "リーダー"
        for day, value in leader.items():
            worksheet.cell(row + 1, _COL_DAY1 + day - 1).value = value
        worksheet.cell(row + 2, 3).value = "時間"
        row += 3

    if legend_rows:
        row += 2
        for col, value in enumerate(
            ["B", "＝20:00～32:15", "A", "＝8:00～20:15", "E", "＝9:00～18:00"], start=5
        ):
            worksheet.cell(row, col).value = value
        for col, value in enumerate(
            ["有", "＝有給休暇", "24シフト休憩時間＝1.5時間", "通常日勤休憩時間＝1時間"],
            start=5,
        ):
            worksheet.cell(row + 1, col).value = value

    workbook.save(path)
    return str(path)


def _codes_by_day(employee):
    return {int(shift["date"][-2:]): shift["code"] for shift in employee["shifts"]}


def test_leading_blank_day_is_kept_as_a_day(tmp_path):
    """月初が空欄の人でも1日ズレない（AI読み取りで加藤 英人さんが実際にズレた事象）。"""
    path = _make_bbs_book(
        tmp_path / "2026_09NMshift.xlsx",
        employees=(
            # 9/1 が空欄。2〜4日が A、7〜10日が B。
            ("2", "(N)加藤 英人", {2: "A", 3: "A", 4: "A", 7: "B", 8: "B", 9: "B", 10: "B"}, {}),
            # 比較用。9/1 から記号が入っている人はもともとズレていなかった。
            ("7", "(N)河端 桂大", {1: "B", 2: "B", 3: "B"}, {}),
        ),
    )

    assert is_bbs_shift_xlsx(path) is True
    parsed = parse_bbs_shift_xlsx(path, 2026, 9)

    assert parsed["year"] == 2026 and parsed["month"] == 9
    kato = parsed["employees"][0]
    assert kato["name"] == "加藤 英人"
    assert len(kato["shifts"]) == 30          # 9月は30日。31列目は捨てる
    assert kato["shifts"][0] == {"date": "2026-09-01", "code": ""}
    assert kato["shifts"][1] == {"date": "2026-09-02", "code": "A"}
    codes = _codes_by_day(kato)
    assert [codes[d] for d in (1, 2, 3, 4, 5, 6, 7)] == ["", "A", "A", "A", "", "", "B"]

    kawabata = _codes_by_day(parsed["employees"][1])
    assert [kawabata[d] for d in (1, 2, 3, 4)] == ["B", "B", "B", ""]


def test_only_marked_staff_rows_are_imported(tmp_path):
    """「(N)」の付かない行（BBS側の要員）は取り込まない。"""
    path = _make_bbs_book(
        tmp_path / "shift.xlsx",
        employees=(
            ("2", "(N)加藤 英人", {1: "A"}, {}),
            ("3", "BBS 山田 太郎", {1: "A"}, {}),
        ),
    )

    parsed = parse_bbs_shift_xlsx(path, 2026, 9)

    assert [e["name"] for e in parsed["employees"]] == ["加藤 英人"]
    assert parsed["skipped_names"] == ["BBS 山田 太郎"]


def test_no_marked_staff_raises(tmp_path):
    """当社社員が1人も居ない表は解析失敗にする（他社の予定を投入しないため）。"""
    path = _make_bbs_book(
        tmp_path / "shift.xlsx",
        employees=(("3", "BBS 山田 太郎", {1: "A"}, {}),),
    )

    with pytest.raises(ValueError, match="当社社員の行"):
        parse_bbs_shift_xlsx(path, 2026, 9)


def test_sheet_legend_wins_and_break_splits_by_span(tmp_path):
    """シート最下段の凡例を読む。拘束12時間以上は1.5h休憩、それ以外は1h。"""
    path = _make_bbs_book(
        tmp_path / "shift.xlsx",
        employees=(("2", "(N)加藤 英人", {1: "A", 2: "B", 3: "E"}, {}),),
    )

    parsed = parse_bbs_shift_xlsx(path, 2026, 9)
    legend = {e["code"]: e for e in parsed["legend"]}

    assert (legend["A"]["start_time"], legend["A"]["end_time"]) == ("8:00", "20:15")
    assert (legend["B"]["start_time"], legend["B"]["end_time"]) == ("20:00", "32:15")
    assert (legend["E"]["start_time"], legend["E"]["end_time"]) == ("9:00", "18:00")
    assert legend["A"]["break_minutes"] == 90
    assert legend["B"]["break_minutes"] == 90
    assert legend["E"]["break_minutes"] == 60
    assert legend["有"]["is_off"] is True
    assert parsed["legend_filled_from_default"] == []
    assert parsed["unknown_codes"] == []


def test_default_legend_fills_when_sheet_has_none(tmp_path):
    """凡例欄が省かれている月（2026-09 の実ファイル）は既定時刻で補い、画面に出す。"""
    path = _make_bbs_book(
        tmp_path / "shift.xlsx",
        employees=(("2", "(N)加藤 英人", {2: "A", 7: "B"}, {}),),
        legend_rows=False,
    )

    parsed = parse_bbs_shift_xlsx(path, 2026, 9)
    legend = {e["code"]: e for e in parsed["legend"]}

    assert sorted(parsed["legend_filled_from_default"]) == ["A", "B"]
    assert (legend["A"]["start_time"], legend["A"]["end_time"]) == ("8:00", "20:15")
    assert (legend["B"]["start_time"], legend["B"]["end_time"]) == ("20:00", "32:15")


def test_unknown_code_is_reported(tmp_path):
    """既定にもシート凡例にも無い記号は未知として返す（凡例確認画面で指定させる）。"""
    path = _make_bbs_book(
        tmp_path / "shift.xlsx",
        employees=(("2", "(N)加藤 英人", {1: "Z"}, {}),),
        legend_rows=False,
    )

    assert parse_bbs_shift_xlsx(path, 2026, 9)["unknown_codes"] == ["Z"]


def test_leader_row_paid_leave_is_lifted_but_place_note_is_not(tmp_path):
    """計画が空の日のリーダー行メモは、休暇だけ記号として拾う。

    実例: 2026-05 加藤 英人さん 5/12「有給休暇」／2026-04 大坪さん 4/1「新宿」。
    """
    path = _make_bbs_book(
        tmp_path / "shift.xlsx",
        employees=(
            ("2", "(N)加藤 英人", {1: "A"}, {5: "有給休暇", 6: "新宿", 1: "AL"}),
        ),
    )

    parsed = parse_bbs_shift_xlsx(path, 2026, 9)
    codes = _codes_by_day(parsed["employees"][0])

    assert codes[5] == "有給休暇"                 # 休暇は拾う
    assert codes[6] == ""                        # 勤務地メモは拾わない
    assert codes[1] == "A"                       # 計画がある日はリーダー当番に上書きされない
    assert parsed["leader_notes"] == ["加藤 英人 9/6「新宿」"]
    assert {e["code"] for e in parsed["legend"]} >= {"有給休暇"}


def test_target_month_mismatch_raises(tmp_path):
    """画面の対象年月と表の年月が食い違ったら誤投入防止で中止する。"""
    path = _make_bbs_book(tmp_path / "shift.xlsx", year=2026, month=9)

    with pytest.raises(ValueError, match="一致しません"):
        parse_bbs_shift_xlsx(path, 2026, 8)


def test_weekday_row_mismatch_raises(tmp_path):
    """曜日行が対象年月と合わない表は、列のずれを疑って解析失敗にする。"""
    path = _make_bbs_book(tmp_path / "shift.xlsx", year=2026, month=9, weekday_year=2025)

    with pytest.raises(ValueError, match="曜日行"):
        parse_bbs_shift_xlsx(path, 2026, 9)


def test_sniffer_ignores_unrelated_book(tmp_path):
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "ただの表"
    path = tmp_path / "other.xlsx"
    workbook.save(path)

    assert is_bbs_shift_xlsx(str(path)) is False


def test_parse_structured_files_consumes_bbs_book(tmp_path):
    """アップロード経路（構造化パース）で BBS 勤務表が AI に回らないこと。"""
    path = _make_bbs_book(
        tmp_path / "2026_09NMshift.xlsx",
        employees=(("2", "(N)加藤 英人", {2: "A", 3: "A"}, {}),),
    )

    sheets, consumed, warnings = parse_structured_files([path], 2026, 9)

    assert consumed == [path]
    assert len(sheets) == 1
    sheet = sheets[0]
    assert sheet["source"] == "bbs"
    assert sheet["year"] == 2026 and sheet["month"] == 9
    assert sheet["employees"][0]["shifts"][0] == {"date": "2026-09-01", "code": ""}
    assert warnings == []
