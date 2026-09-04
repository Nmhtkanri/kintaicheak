r"""標準報酬月額チェックのユニットテスト。

等級表ローダは in-memory の Workbook を組み立てて検証を確かめる（実データ不要）。
実在の令和8年度Excel（Z:\API連携\標準月額資料）を読むテストだけは、
共有フォルダが無い環境では SkipTest で飛ばす。
"""

import os
import tempfile
import unittest

import openpyxl

from services.shaho_engine import (assess_month, calc_teiji_kettei, collect_remuneration,
                                   payment_base_days)
from services.shaho_master import (GradeRow, ShahoMasterError, load_class_master,
                                   load_grade_table, resolve_class)

REAL_XLSX = r"Z:\API連携\標準月額資料\令和8年度_標準報酬月額表_関東IT健保_協会けんぽ東京比較.xlsx"
OFFICIAL_XLSX = r"Z:\API連携\標準月額資料\標準報酬月額_2026_09.xlsx"

# ミニ等級表: 健保3等級（最終は上限なし）・厚年2等級。料率は本物と同じ。
RATES = {"kenpo": 0.04635, "kodomo": 0.00115, "kaigo": 0.009, "konen": 0.0915}
KENPO_ROWS = [
    # (健保等級, 健保標報, 下限, 上限, 厚年等級, 厚年標報)
    (1, 58000, 0, 63000, 1, 88000),
    (2, 68000, 63000, 73000, 1, 88000),
    (3, 78000, 73000, None, 2, 98000),
]
KONEN_ROWS = [
    (1, 88000, 0, 73000),
    (2, 98000, 73000, None),
]


def build_workbook(kenpo_rows=None, konen_rows=None, year_text="令和8年度 標準報酬月額・本人負担額一覧",
                   rate_rows=None, premium_override=None):
    """検証を通る最小の等級表ブックを組む。引数で壊し方を指定できる。"""
    kenpo_rows = KENPO_ROWS if kenpo_rows is None else kenpo_rows
    konen_rows = KONEN_ROWS if konen_rows is None else konen_rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "標準報酬月額表"
    ws.cell(1, 1).value = year_text
    headers = ["健保等級", "健保標準報酬月額", "報酬月額下限（以上）", "報酬月額上限（未満）",
               "報酬月額の範囲", "厚年等級", "厚年標準報酬月額",
               "ITS健康", "ITS支援金", "ITS健康＋支援金", "ITS介護",
               "協会東京健康", "協会東京支援金", "協会東京健康＋支援金", "協会東京介護",
               "厚生年金", "合計 ITS（40歳未満）", "合計 協会東京（40歳未満）", "差額（協会－ITS）"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c).value = h
    for i, (g, smr, lo, up, kg, ksmr) in enumerate(kenpo_rows):
        r = 5 + i
        prem = {
            "ITS健康": smr * RATES["kenpo"], "ITS支援金": smr * RATES["kodomo"],
            "ITS介護": smr * RATES["kaigo"], "厚生年金": ksmr * RATES["konen"],
            "協会東京健康": smr * 0.04925, "協会東京支援金": smr * 0.00115,
            "協会東京介護": smr * 0.0081,
        }
        if premium_override and g in premium_override:
            prem.update(premium_override[g])
        values = [g, smr, lo, up, "", kg, ksmr,
                  prem["ITS健康"], prem["ITS支援金"], prem["ITS健康"] + prem["ITS支援金"],
                  prem["ITS介護"],
                  prem["協会東京健康"], prem["協会東京支援金"],
                  prem["協会東京健康"] + prem["協会東京支援金"], prem["協会東京介護"],
                  prem["厚生年金"], 0, 0, 0]
        for c, v in enumerate(values, start=1):
            ws.cell(r, c).value = v

    ws2 = wb.create_sheet("設定・出典")
    ws2.cell(1, 1).value = "令和8年度 料率設定・出典"
    for c, h in enumerate(["区分", "保険者", "項目", "全体料率", "本人負担率", "適用・備考", "出典URL"],
                          start=1):
        ws2.cell(4, c).value = h
    default_rates = [
        ("健康保険", "関東IT健保", "健康保険", 0.0927, 0.04635),
        ("健康保険", "関東IT健保", "子ども・子育て支援金", 0.0023, 0.00115),
        ("介護保険", "関東IT健保", "介護保険", 0.018, 0.009),
        ("厚生年金", "共通", "厚生年金", 0.183, 0.0915),
        ("健康保険", "協会けんぽ東京", "健康保険", 0.0985, 0.04925),
        ("健康保険", "協会けんぽ東京", "子ども・子育て支援金", 0.0023, 0.00115),
        ("介護保険", "協会けんぽ東京", "介護保険", 0.0162, 0.0081),
    ]
    for i, row in enumerate(rate_rows if rate_rows is not None else default_rates):
        for c, v in enumerate(row, start=1):
            ws2.cell(5 + i, c).value = v
    for c, h in enumerate(["厚年等級", "標準報酬月額", "報酬月額下限（以上）", "報酬月額上限（未満）"],
                          start=9):
        ws2.cell(4, c).value = h
    for i, (g, smr, lo, up) in enumerate(konen_rows):
        for c, v in enumerate((g, smr, lo, up), start=9):
            ws2.cell(5 + i, c).value = v
    return wb


def load_from_workbook(wb, insurer="its", year=2026):
    """一時ファイルへ保存して load_grade_table に読ませる。"""
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "等級表.xlsx")
        wb.save(path)
        return load_grade_table(path, insurer, year)


class GradeLookupTests(unittest.TestCase):
    def setUp(self):
        self.master = load_from_workbook(build_workbook())

    def test_boundaries_are_half_open(self):
        """下限は含む・上限は含まない。"""
        self.assertEqual(self.master.find_grade(0).kenpo_grade, 1)
        self.assertEqual(self.master.find_grade(62999.99).kenpo_grade, 1)
        self.assertEqual(self.master.find_grade(63000).kenpo_grade, 2)   # 上限ちょうど→次の等級
        self.assertEqual(self.master.find_grade(72999).kenpo_grade, 2)
        self.assertEqual(self.master.find_grade(73000).kenpo_grade, 3)

    def test_top_grade_has_no_upper_bound(self):
        self.assertEqual(self.master.find_grade(99999999).kenpo_grade, 3)

    def test_kenpo_and_konen_differ(self):
        """健保と厚年の標準報酬月額は別物（上限・下限が違う）。"""
        row = self.master.find_grade(70000)
        self.assertEqual(row.kenpo_smr, 68000)
        self.assertEqual(row.konen_smr, 88000)

    def test_negative_amount_is_rejected(self):
        with self.assertRaises(ShahoMasterError):
            self.master.find_grade(-1)

    def test_rates_are_loaded_for_its(self):
        self.assertAlmostEqual(self.master.rates["kenpo"].employee, 0.04635)
        self.assertAlmostEqual(self.master.rates["kaigo"].employee, 0.009)
        self.assertAlmostEqual(self.master.rates["konen"].employee, 0.0915)

    def test_kyokai_tokyo_rates(self):
        m = load_from_workbook(build_workbook(), insurer="kyokai_tokyo")
        self.assertAlmostEqual(m.rates["kenpo"].employee, 0.04925)
        self.assertAlmostEqual(m.rates["kaigo"].employee, 0.0081)


class GradeValidationTests(unittest.TestCase):
    """壊れた等級表は1円も計算せずに止める。"""

    def test_year_mismatch_stops(self):
        with self.assertRaises(ShahoMasterError) as ctx:
            load_from_workbook(build_workbook(), year=2027)
        self.assertIn("年度", str(ctx.exception))

    def test_missing_year_stops(self):
        with self.assertRaises(ShahoMasterError):
            load_from_workbook(build_workbook(year_text="標準報酬月額の表"))

    def test_gap_in_bands_stops(self):
        rows = [(1, 58000, 0, 63000, 1, 88000),
                (2, 68000, 64000, 73000, 1, 88000),      # 63000-64000 に穴
                (3, 78000, 73000, None, 2, 98000)]
        with self.assertRaises(ShahoMasterError) as ctx:
            load_from_workbook(build_workbook(kenpo_rows=rows))
        self.assertIn("繋がって", str(ctx.exception))

    def test_non_consecutive_grade_stops(self):
        rows = [(1, 58000, 0, 63000, 1, 88000),
                (3, 78000, 63000, None, 2, 98000)]
        with self.assertRaises(ShahoMasterError):
            load_from_workbook(build_workbook(kenpo_rows=rows))

    def test_upper_bound_on_last_row_stops(self):
        rows = [(1, 58000, 0, 63000, 1, 88000),
                (2, 68000, 63000, 73000, 1, 88000),
                (3, 78000, 73000, 83000, 2, 98000)]      # 最終行に上限がある
        with self.assertRaises(ShahoMasterError):
            load_from_workbook(build_workbook(kenpo_rows=rows))

    def test_konen_mismatch_stops(self):
        """健保の帯と厚年等級マスタの食い違いを検知する。"""
        rows = [(1, 58000, 0, 63000, 1, 88000),
                (2, 68000, 63000, 73000, 2, 98000),      # 73000未満は厚年1のはず
                (3, 78000, 73000, None, 2, 98000)]
        with self.assertRaises(ShahoMasterError) as ctx:
            load_from_workbook(build_workbook(kenpo_rows=rows))
        self.assertIn("厚年", str(ctx.exception))

    def test_stale_premium_column_stops(self):
        """保険料額列が料率×標報と合わない＝表か料率が古い。"""
        with self.assertRaises(ShahoMasterError) as ctx:
            load_from_workbook(build_workbook(premium_override={2: {"ITS健康": 9999.0}}))
        self.assertIn("合いません", str(ctx.exception))

    def test_missing_rate_stops(self):
        rates = [("健康保険", "関東IT健保", "健康保険", 0.0927, 0.04635),
                 ("厚生年金", "共通", "厚生年金", 0.183, 0.0915)]
        with self.assertRaises(ShahoMasterError) as ctx:
            load_from_workbook(build_workbook(rate_rows=rates))
        self.assertIn("揃って", str(ctx.exception))

    def test_employee_rate_must_be_half_of_total(self):
        rates = [("健康保険", "関東IT健保", "健康保険", 0.0927, 0.05),
                 ("健康保険", "関東IT健保", "子ども・子育て支援金", 0.0023, 0.00115),
                 ("介護保険", "関東IT健保", "介護保険", 0.018, 0.009),
                 ("厚生年金", "共通", "厚生年金", 0.183, 0.0915)]
        with self.assertRaises(ShahoMasterError) as ctx:
            load_from_workbook(build_workbook(rate_rows=rates))
        self.assertIn("半分", str(ctx.exception))

    def test_unknown_insurer_stops(self):
        with self.assertRaises(ShahoMasterError):
            load_from_workbook(build_workbook(), insurer="kokuho")


class ClassMasterTests(unittest.TestCase):
    HEADER = "source_key,salary_system_label,label,class,fixed,適用開始月,適用終了月,note\n"

    def _load(self, body):
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "class.csv")
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                f.write(self.HEADER + body)
            return load_class_master(path)

    def test_label_specific_row_wins_over_common_row(self):
        """allowance2 は体系別名の行が共通行より優先される。"""
        m = self._load(
            "salary_items:allowance2,当月みなし時間外手当,みなし給,対象,1,,,\n"
            "salary_items:allowance2,前月超過勤務,前月超過勤務,対象外,0,,,情報項目\n"
            "salary_items:allowance1,,基本給,対象,1,,,\n")
        self.assertEqual(resolve_class(m, "salary_items:allowance2", "当月みなし時間外手当").cls, "対象")
        self.assertEqual(resolve_class(m, "salary_items:allowance2", "前月超過勤務").cls, "対象外")
        self.assertIsNone(resolve_class(m, "salary_items:allowance2", "謎の名前"))
        self.assertEqual(resolve_class(m, "salary_items:allowance1", "何でも").cls, "対象")

    def test_period_rows_switch_by_month(self):
        """同じラベルでも適用期間で分類が切り替わる（時給制みなし給の再掲→実額）。"""
        m = self._load(
            "salary_items:allowance2,みなし給,みなし給,対象外,,,2026-07,再掲\n"
            "salary_items:allowance2,みなし給,みなし給,対象,1,2026-08,,実額\n")
        self.assertEqual(resolve_class(m, "salary_items:allowance2", "みなし給", "2026-05").cls,
                         "対象外")
        self.assertEqual(resolve_class(m, "salary_items:allowance2", "みなし給", "2026-08").cls,
                         "対象")

    def test_overlapping_periods_stop(self):
        with self.assertRaises(ShahoMasterError):
            self._load(
                "salary_items:allowance2,みなし給,みなし給,対象外,,,2026-08,\n"
                "salary_items:allowance2,みなし給,みなし給,対象,1,2026-08,,\n")

    def test_bad_period_format_stops(self):
        with self.assertRaises(ShahoMasterError):
            self._load("salary_items:allowance1,,基本給,対象,1,2026/08,,\n")

    def test_invalid_class_value_stops(self):
        with self.assertRaises(ShahoMasterError):
            self._load("salary_items:allowance1,,基本給,たいしょう,1,,,\n")

    def test_invalid_fixed_value_stops(self):
        with self.assertRaises(ShahoMasterError):
            self._load("salary_items:allowance1,,基本給,対象,yes,,,\n")

    def test_duplicate_row_stops(self):
        with self.assertRaises(ShahoMasterError):
            self._load("salary_items:allowance1,,基本給,対象,1,,,\n"
                       "salary_items:allowance1,,基本給,対象外,0,,,\n")

    def test_missing_file_stops_with_hint(self):
        with self.assertRaises(ShahoMasterError) as ctx:
            load_class_master(os.path.join(tempfile.gettempdir(), "no_such_class.csv"))
        self.assertIn("build_shaho_class_master", str(ctx.exception))


class _ClassMasterMixin:
    """エンジン系テスト用の最小分類マスタ。"""

    CSV = ("source_key,salary_system_label,label,class,fixed,適用開始月,適用終了月,note\n"
           "salary_items:allowance1,,基本給,対象,1,,,\n"
           "salary_items:allowance2,当月みなし時間外手当,みなし給,対象,1,,,\n"
           "salary_items:allowance2,前月超過勤務,前月超過勤務,対象外,,,,\n"
           "salary_items:allowance2,みなし給,みなし給,対象外,,,2026-07,再掲\n"
           "salary_items:allowance2,みなし給,みなし給,対象,1,2026-08,,実額\n"
           "salary_items:allowance24,,差額調整,対象,0,,,\n"
           "salary_items:allowance35,,非課税通勤費,対象,1,,,\n"
           "salary_items:allowance18,,テレワーク手当,対象,0,,,\n"
           "salary_items:allowance5,,前月実績分,対象外,,,,\n"
           "salary_items:allowance51,,立替金,対象外,,,,\n"
           "salary_items:allowance53,,現物支給,現物,0,,,\n"
           "salary_items:allowance55,,謎の現物枠,未設定,,,,\n")

    @classmethod
    def setUpClass(cls):
        cls._dir = tempfile.TemporaryDirectory()
        path = os.path.join(cls._dir.name, "class.csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(cls.CSV)
        cls.cm = load_class_master(path)

    @classmethod
    def tearDownClass(cls):
        cls._dir.cleanup()


def _pi_shaho(shikyu=None, kintai=None, other=None, labels=None):
    """shaho_engine 用の payroll_info。shikyu={'allowance1': 400000, ...}"""
    labels = labels or {}

    def items(pairs):
        return [{"id": k, "value": v, "salary_system_label": labels.get(k, ""), "label": ""}
                for k, v in (pairs or {}).items()]
    return {"salary_items": items(shikyu),
            "salary_attendance_items": items(kintai),
            "salary_other_items": items(other)}


class RemunerationTests(_ClassMasterMixin, unittest.TestCase):
    def test_gate_passes_when_taishou_equals_gross(self):
        """高岩さん2026-04の形: 対象計=other5。再掲(a5)・立替金・情報項目は乗らない。"""
        pi = _pi_shaho(
            shikyu={"allowance1": 426650, "allowance2": 133350, "allowance24": 70233,
                    "allowance35": 15000, "allowance5": 65565, "allowance51": 10000},
            other={"other5": 645233},
            labels={"allowance2": "当月みなし時間外手当"})
        rem = collect_remuneration(pi, self.cm)
        self.assertEqual(rem.total, 645233)
        self.assertTrue(rem.gate_ok)
        self.assertEqual(rem.unclassified, [])
        self.assertEqual(rem.fixed_total, 426650 + 133350 + 15000)

    def test_gate_breaks_when_master_misses_a_paid_item(self):
        """総支給に乗る項目が対象から漏れたら検算が割れる。"""
        pi = _pi_shaho(shikyu={"allowance1": 400000, "allowance18": 8000},
                       other={"other5": 408000})
        rem = collect_remuneration(pi, self.cm)
        self.assertTrue(rem.gate_ok)
        pi2 = _pi_shaho(shikyu={"allowance1": 400000, "allowance18": 8000,
                                "allowance24": 5000},        # 対象なのにother5に乗せ忘れ
                        other={"other5": 408000})
        self.assertFalse(collect_remuneration(pi2, self.cm).gate_ok)

    def test_unclassified_item_with_amount_is_flagged(self):
        pi = _pi_shaho(shikyu={"allowance1": 400000, "allowance55": 5280, "allowance99": 100},
                       other={"other5": 400000})
        rem = collect_remuneration(pi, self.cm)
        reasons = {u["source_key"]: u["理由"] for u in rem.unclassified}
        self.assertEqual(reasons, {"salary_items:allowance55": "未設定",
                                   "salary_items:allowance99": "マスタに行が無い"})

    def test_genbutsu_counts_as_pay_but_not_in_gate(self):
        """現物支給は報酬に数えるが、現金グロスの外側なので検算からは除く。"""
        pi = _pi_shaho(shikyu={"allowance1": 400000, "allowance53": 3000},
                       other={"other5": 400000})       # 現物はother5に乗らない（実測どおり）
        rem = collect_remuneration(pi, self.cm)
        self.assertTrue(rem.gate_ok)                    # 検算は現金だけで成立
        self.assertEqual(rem.cash_total, 400000)
        self.assertEqual(rem.genbutsu_total, 3000)
        self.assertEqual(rem.total, 403000)             # 報酬＝通貨＋現物
        self.assertEqual(rem.unclassified, [])

    def test_hourly_allowance2_is_excluded_by_label(self):
        pi = _pi_shaho(shikyu={"allowance1": 408000, "allowance2": 692},
                       other={"other5": 408000},
                       labels={"allowance2": "前月超過勤務"})
        rem = collect_remuneration(pi, self.cm)
        self.assertEqual(rem.total, 408000)
        self.assertTrue(rem.gate_ok)

    def test_gross_falls_back_to_other6(self):
        pi = _pi_shaho(shikyu={"allowance1": 300000},
                       other={"other5": 0, "other6": 300000})
        self.assertTrue(collect_remuneration(pi, self.cm).gate_ok)


class BaseDaysTests(unittest.TestCase):
    def test_monthly_uses_calendar_days(self):
        bd = payment_base_days(_pi_shaho(), "月給制1", "2026-04")
        self.assertEqual((bd.days, bd.basis, bd.approx), (30, "暦日", False))
        self.assertEqual(payment_base_days(_pi_shaho(), "月給制2", "2026-05").days, 31)

    def test_hourly_uses_working_plus_paid_leave(self):
        bd = payment_base_days(_pi_shaho(kintai={"kintai10": 15, "kintai13": 3},
                                         labels={"kintai10": "出勤日数",
                                                 "kintai13": "前月有休消化日数"}),
                               "時給制1", "2026-04")
        self.assertEqual((bd.days, bd.approx), (18, False))

    def test_hourly_reads_by_label_not_id(self):
        """項目IDは月で意味が変わる。2026-04 は kintai6=出勤日数・kintai8=前月有給消化日数。"""
        bd = payment_base_days(_pi_shaho(kintai={"kintai6": 23, "kintai8": 0},
                                         labels={"kintai6": "出勤日数",
                                                 "kintai8": "前月有給消化日数"}),
                               "時給制1", "2026-04")
        self.assertEqual(bd.days, 23)

    def test_paid_leave_label_variants(self):
        """「有給」「有休」の表記ゆれを両方拾う。"""
        for label in ("前月有給消化日数", "前月有休消化日数"):
            bd = payment_base_days(_pi_shaho(kintai={"a": 10, "b": 5},
                                             labels={"a": "出勤日数", "b": label}),
                                   "時給制1", "2026-06")
            self.assertEqual(bd.days, 15, label)

    def test_may_2026_slot_override(self):
        """2026-05 の時給制はラベルすら当てにならないので、スロット直読みで補正する。

        実データでは kintai6 のラベルが「内前月実績超過勤怠時間60時間以上」なのに
        中身は出勤日数、kintai10（ラベル=出勤日数）は空、という移行漏れが起きていた。
        """
        pi = _pi_shaho(
            kintai={"kintai6": 21.5, "kintai10": 0, "kintai13": 0.5, "kintai11": 60},
            labels={"kintai6": "内前月実績超過勤怠時間60時間以上", "kintai10": "出勤日数",
                    "kintai13": "前月有休消化日数", "kintai11": "欠勤日数"},
            other={"other5": 480000})
        bd = payment_base_days(pi, "時給制1", "2026-05")
        self.assertEqual(bd.days, 22.0)                 # 21.5 出勤 + 0.5 有休
        self.assertIn("配置ずれ", bd.basis)

    def test_other_months_still_use_labels(self):
        """補正は2026-05だけ。他の月はラベルで引く（同じ形でも kintai6 を出勤日数にしない）。"""
        pi = _pi_shaho(
            kintai={"kintai6": 21.5, "kintai10": 18},
            labels={"kintai6": "内前月実績超過勤怠時間60時間以上", "kintai10": "出勤日数"},
            other={"other5": 480000})
        self.assertEqual(payment_base_days(pi, "時給制1", "2026-06").days, 18)

    def test_zero_attendance_with_pay_is_treated_as_missing(self):
        """出勤日数0なのに報酬が出ている＝未入力を疑い、0日として除外しない。

        2026-05 支給分は時給制49名全員が出勤日数ゼロだった（体系移行月の入力漏れ）。
        """
        bd = payment_base_days(_pi_shaho(kintai={"a": 0}, labels={"a": "出勤日数"},
                                         other={"other5": 480000}),
                               "時給制1", "2026-06")
        self.assertIsNone(bd.days)
        self.assertIn("未入力", bd.basis)

    def test_zero_attendance_without_pay_is_really_zero(self):
        """報酬も0なら本当に稼働ゼロ（休職など）。0日として扱う。"""
        bd = payment_base_days(_pi_shaho(kintai={"a": 0}, labels={"a": "出勤日数"},
                                         other={"other5": 0}),
                               "時給制1", "2026-06")
        self.assertEqual(bd.days, 0)

    def test_monthly_absence_read_by_label(self):
        bd = payment_base_days(_pi_shaho(kintai={"x": 20}, labels={"x": "欠勤日数"}),
                               "月給制3", "2026-06")
        self.assertEqual((bd.days, bd.approx), (10, True))

    def test_hourly_without_kintai_is_unknown(self):
        bd = payment_base_days(_pi_shaho(), "時給制1", "2026-04")
        self.assertIsNone(bd.days)

    def test_unknown_system_is_unknown(self):
        self.assertIsNone(payment_base_days(_pi_shaho(), "月給制9", "2026-04").days)


class TeijiKetteiTests(_ClassMasterMixin, unittest.TestCase):
    def _master(self):
        return load_from_workbook(build_workbook())

    def _assess(self, ym, base, days_kintai=None, system="月給制1", labels=None):
        pi = _pi_shaho(shikyu={"allowance1": base}, kintai=days_kintai,
                       other={"other5": base}, labels=labels)
        return assess_month(ym, pi, system, self.cm, threshold=17)

    def test_average_is_floored_and_graded(self):
        """3か月平均は円未満切捨て → 等級ルックアップ。"""
        a = [self._assess("2026-04", 68000), self._assess("2026-05", 68000),
             self._assess("2026-06", 68001)]
        tk = calc_teiji_kettei(a, self._master())
        self.assertEqual(tk.adopted_n, 3)
        self.assertEqual(tk.average, 68000)          # 204001/3 = 68000.33 → 68000
        self.assertEqual(tk.kenpo_grade, 2)
        self.assertEqual(tk.konen_smr, 88000)

    def test_low_base_days_month_is_excluded(self):
        """17日未満の月は平均から除外（時給14日→除外）。"""
        a = [self._assess("2026-04", 68000),
             self._assess("2026-06", 30000, days_kintai={"k": 14},
                          labels={"k": "出勤日数"}, system="時給制1"),
             self._assess("2026-07", 68000)]
        tk = calc_teiji_kettei(a, self._master())
        self.assertEqual(tk.adopted_n, 2)
        self.assertEqual(tk.average, 68000)
        self.assertIn("14日", a[1].reason)

    def test_short_time_worker_threshold_11(self):
        pi = _pi_shaho(shikyu={"allowance1": 30000}, kintai={"k": 12},
                       other={"other5": 30000}, labels={"k": "出勤日数"})
        ma = assess_month("2026-06", pi, "時給制1", self.cm, threshold=11)
        self.assertTrue(ma.adopted)

    def test_all_months_excluded_gives_no_average(self):
        a = [self._assess("2026-04", 30000, days_kintai={"k": 5},
                          labels={"k": "出勤日数"}, system="時給制1")]
        tk = calc_teiji_kettei(a, self._master())
        self.assertEqual(tk.adopted_n, 0)
        self.assertIsNone(tk.average)

    def test_gate_and_approx_flags_bubble_up(self):
        pi = _pi_shaho(shikyu={"allowance1": 68000}, kintai={"k": 3},
                       labels={"k": "欠勤日数"},
                       other={"other5": 99999})     # ゲート割れ＋欠勤概算
        a = [assess_month("2026-04", pi, "月給制1", self.cm, threshold=17)]
        tk = calc_teiji_kettei(a, self._master())
        self.assertTrue(tk.gate_ng)
        self.assertTrue(tk.approx_used)


class PremiumTests(unittest.TestCase):
    """保険料の丸めと対象区分。"""

    def test_rounding_modes(self):
        from services.shaho_check import round_premium
        self.assertEqual(round_premium(100.5, "50sen"), 100)    # 50銭ちょうど→切捨て
        self.assertEqual(round_premium(100.51, "50sen"), 101)   # 50銭超→切上げ
        self.assertEqual(round_premium(100.49, "50sen"), 100)
        self.assertEqual(round_premium(100.9, "floor"), 100)
        self.assertEqual(round_premium(100.1, "ceil"), 101)
        self.assertEqual(round_premium(100.5, "round"), 101)

    def test_premiums_respect_classifications(self):
        from services.shaho_check import calc_premiums
        master = load_from_workbook(build_workbook())
        bi = {"health_insurance_calculation_classification": {"name": "被保険者（対象）"},
              "care_insurance_calculation_classification": {"name": "第2号被保険者"},
              "employees_pension": {"calculation_classification": {"name": "被保険者（対象）"}}}
        p = calc_premiums(620000, 620000, bi, master, "50sen")
        self.assertEqual(p["kenpo"], 28737)                     # 実測値の再現
        self.assertEqual(p["konen"], 56730)
        self.assertEqual(p["kaigo"], 5580)
        bi2 = dict(bi, care_insurance_calculation_classification={"name": "対象外"},
                   employees_pension={"calculation_classification": {"name": "70歳以上（対象外）"}})
        p2 = calc_premiums(620000, 620000, bi2, master, "50sen")
        self.assertEqual((p2["kaigo"], p2["konen"]), (0, 0))    # 介護外・70歳以上は0
        self.assertEqual(p2["kenpo"], 28737)                    # 健保だけは続く
        bi3 = dict(bi, health_insurance_calculation_classification={"name": "対象外"})
        self.assertEqual(calc_premiums(620000, 620000, bi3, master, "50sen"),
                         {"kenpo": 0, "kodomo": 0, "kaigo": 0, "konen": 0})

    def test_status_priority_merge(self):
        from services.shaho_check import merge_status
        self.assertEqual(merge_status("OK", "DIFFERENCE"), "DIFFERENCE")
        self.assertEqual(merge_status("EXEMPTION_REVIEW", "DIFFERENCE"), "EXEMPTION_REVIEW")
        self.assertEqual(merge_status("NOT_APPLICABLE", "OK"), "OK")
        self.assertEqual(merge_status("PROVISIONAL_OK", "OK"), "PROVISIONAL_OK")


class ClosedMonthTests(unittest.TestCase):
    """給与が未確定の月は値がまだ動くので、その月を含む人は信用しない。"""

    def _map(self, closed_flags):
        return {emp: {"basic_info": {}, "payroll_info": {"is_payroll_closed": c},
                      "n_nonzero": 1}
                for emp, c in closed_flags.items()}

    def test_counts_open_and_closed(self):
        from services.shaho_check import month_closed_stats
        st = month_closed_stats(self._map({"2020001": True, "2020002": False,
                                           "2020003": True}))
        self.assertEqual((st["closed"], st["open"]), (2, 1))
        self.assertEqual(st["open_emps"], ["2020002"])

    def test_all_closed_reports_none_open(self):
        from services.shaho_check import month_closed_stats
        st = month_closed_stats(self._map({"2020001": True, "2020002": True}))
        self.assertEqual((st["closed"], st["open"], st["open_emps"]), (2, 0, []))

    def test_missing_flag_counts_as_open(self):
        """フラグ自体が無い明細は「未確定」に倒す（安全側）。"""
        from services.shaho_check import month_closed_stats
        st = month_closed_stats({"2020001": {"basic_info": {}, "payroll_info": {},
                                             "n_nonzero": 1}})
        self.assertEqual(st["open"], 1)


class SelectTableTests(unittest.TestCase):
    """Codexが毎月1日に置く公式資料（標準報酬月額_YYYY_MM.xlsx）から、対象年月に合うものを選ぶ。

    対象年月以前で最新 → 無ければ同じ年度で最も早いもの → それも無ければ設定の手作りブック。
    """

    def _folder(self, tmp, names, base_name="令和8年度_表.xlsx"):
        base = os.path.join(tmp, base_name)
        for name in [base_name] + list(names):
            with open(os.path.join(tmp, name), "wb") as f:
                f.write(b"x")
        return base

    def test_latest_snapshot_on_or_before_target_month(self):
        from services.shaho_master import select_grade_table
        with tempfile.TemporaryDirectory() as tmp:
            base = self._folder(tmp, ["標準報酬月額_2026_09.xlsx", "標準報酬月額_2026_10.xlsx",
                                      "標準報酬月額_2026_12.xlsx"])
            c = select_grade_table("2026-11", base)
            self.assertEqual(os.path.basename(c.path), "標準報酬月額_2026_10.xlsx")
            self.assertEqual(c.snapshot_ym, "2026-10")
            c = select_grade_table("2026-09", base)
            self.assertEqual(c.snapshot_ym, "2026-09")

    def test_earliest_in_same_fiscal_year_when_none_before(self):
        """9月の資料しか無いのに7月分を計算する: 同じ年度なので9月の資料を使う。"""
        from services.shaho_master import select_grade_table
        with tempfile.TemporaryDirectory() as tmp:
            base = self._folder(tmp, ["標準報酬月額_2026_09.xlsx", "標準報酬月額_2026_10.xlsx"])
            c = select_grade_table("2026-07", base)
            self.assertEqual(c.snapshot_ym, "2026-09")

    def test_falls_back_to_configured_when_no_snapshot_fits(self):
        """前年度（2026-03 は令和7年度）には 2026_09 の資料は使えない → 設定のブック。"""
        from services.shaho_master import select_grade_table
        with tempfile.TemporaryDirectory() as tmp:
            base = self._folder(tmp, ["標準報酬月額_2026_09.xlsx"])
            c = select_grade_table("2026-03", base)
            self.assertEqual(c.path, base)
            self.assertIsNone(c.snapshot_ym)

    def test_other_filenames_are_ignored(self):
        """命名ルール（年4桁_月2桁）から外れるものは拾わない。"""
        from services.shaho_master import select_grade_table
        with tempfile.TemporaryDirectory() as tmp:
            base = self._folder(tmp, ["標準報酬月額_2026_9.xlsx", "メモ.xlsx",
                                      "標準報酬月額表_2026_09.xlsx", "標準報酬月額_2026_13.xlsx"])
            self.assertEqual(select_grade_table("2026-09", base).path, base)

    def test_missing_folder_is_safe(self):
        from services.shaho_master import select_grade_table
        c = select_grade_table("2026-09", "Z:/ない/ない.xlsx")
        self.assertEqual(c.path, "Z:/ない/ない.xlsx")


# ---------------------------------------------------------------------------
# Codex 公式資料（協会けんぽの様式そのまま）のミニ版
# ---------------------------------------------------------------------------
OFFICIAL_RATES = {"kenpo": 0.0985, "kenpo_kaigo": 0.1147, "kodomo": 0.0023, "konen": 0.183}
# (等級欄の表記, 健保標報, 下限, 上限, 厚年標報|None=厚年の額が空欄の行)
# 実物と同じく、厚年の帯の端の等級は括弧も厚年の額も付かない
OFFICIAL_ROWS = [
    (1, 58000, None, 63000, None),
    ("2(1)", 68000, 63000, 73000, 88000),
    ("3(2)", 78000, 73000, None, 98000),
]
OFFICIAL_NAME = "標準報酬月額_2026_09.xlsx"


def build_official_workbook(rows=None, konen_rows=None, rates=None, about_ym="2026年9月",
                            year_text="令和8年度　都道府県別保険料率一覧", listed_override=None,
                            premium_override=None, konen_sheet_pct=18.3):
    """検証を通る最小の公式資料を組む。引数で壊し方を指定できる。"""
    rows = OFFICIAL_ROWS if rows is None else rows
    konen_rows = KONEN_ROWS if konen_rows is None else konen_rows
    r_ = dict(OFFICIAL_RATES, **(rates or {}))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "東京"
    ws["A1"] = "令和８年３月分（４月納付分）からの健康保険・厚生年金保険の保険料額表"
    ws["A3"] = "　・健康保険料率：令和8年3月分～適用"
    ws["E3"] = "　・厚生年金保険料率：平成29年9月分～適用"
    ws["I3"] = "・子ども・子育て支援金率：令和8年4月分（5月納付分）～適用"
    ws["A4"] = "　・介護保険料率：令和8年3月分～適用"
    ws["A5"] = "（東京支部）"
    ws["A6"] = "標  準  報  酬"
    ws["C6"] = "報  酬  月  額"
    ws["F6"] = "全国健康保険協会管掌健康保険料・介護保険料"
    ws["J6"] = "子ども・子育て支援金"
    ws["L6"] = "厚生年金保険料\n（厚生年金基金加入員を除く）"
    ws["F7"] = "介護保険第２号被保険者\nに該当しない場合"
    ws["H7"] = "介護保険第２号被保険者\nに該当する場合"
    ws["A8"] = "等級"
    ws["B8"] = "月  額"
    ws["F8"], ws["H8"], ws["J8"], ws["L8"] = (r_["kenpo"], r_["kenpo_kaigo"],
                                              r_["kodomo"], r_["konen"])
    for c in ("F", "H", "J", "L"):
        ws[f"{c}9"] = "全  額"
    for c in ("G", "I", "K", "M"):
        ws[f"{c}9"] = "折半額"
    ws["C10"], ws["E10"] = "円以上", "円未満"
    for i, (g, smr, lo, up, ksmr) in enumerate(rows):
        r = 11 + i
        prem = {"G": smr * r_["kenpo"] / 2, "I": smr * r_["kenpo_kaigo"] / 2,
                "K": smr * r_["kodomo"] / 2,
                "M": ksmr * r_["konen"] / 2 if ksmr is not None else None}
        if premium_override and g in premium_override:
            prem.update(premium_override[g])
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = g, smr, lo, "～", up
        for col, v in prem.items():
            ws[f"{col}{r}"] = v
    ws[f"A{11 + len(rows) + 1}"] = "◆等級欄の（　）内の数字は、厚生年金保険の標準報酬月額等級です。"

    ws2 = wb.create_sheet("厚生年金_公式表")
    ws2["A1"] = "○令和2年9月分（10月納付分）からの厚生年金保険料額表（令和8年度版）"
    ws2["E2"], ws2["F2"] = "全額保険料率", "折半保険料率"
    ws2["E3"], ws2["F3"] = konen_sheet_pct, konen_sheet_pct / 2
    for c, h in enumerate(["等級", "標準報酬月額", "報酬月額 下限", "報酬月額 上限",
                           "保険料 全額", "保険料 折半額"], start=1):
        ws2.cell(5, c).value = h
    k_rate = konen_sheet_pct / 100          # このシートの額は自分の率で作る（率のズレ検知を試すため）
    for i, (g, smr, lo, up) in enumerate(konen_rows):
        for c, v in enumerate((g, smr, lo or None, up, smr * k_rate, smr * k_rate / 2), start=1):
            ws2.cell(6 + i, c).value = v

    ws3 = wb.create_sheet("都道府県料率一覧")
    ws3["A1"] = year_text
    for c, h in enumerate(["都道府県", "健康保険料率\n（介護非該当）", "健康＋介護保険料率\n（介護該当）",
                           "子ども・子育て\n支援金率", "厚生年金\n保険料率", "参照シート"], start=1):
        ws3.cell(3, c).value = h
    listed = dict(r_, **(listed_override or {}))
    for c, v in enumerate(["東京", listed["kenpo"], listed["kenpo_kaigo"], listed["kodomo"],
                           listed["konen"], "東京"], start=1):
        ws3.cell(4, c).value = v

    ws4 = wb.create_sheet("概要・出典")
    ws4["A1"] = "標準報酬月額・保険料額資料　2026年9月スナップショット"
    ws4["A3"], ws4["B3"] = "対象年月", about_ym
    ws4["A4"], ws4["B4"] = "確認日時（日本時間）", "2026/9/4 9:32:30"
    ws4["A10"], ws4["B10"] = "機関", "公式ページ"
    ws4["A11"], ws4["B11"] = "全国健康保険協会", "https://www.kyoukaikenpo.or.jp/x"
    ws4["A12"], ws4["B12"] = "日本年金機構", "https://www.nenkin.go.jp/x"
    return wb


def load_official(wb, insurer="kyokai_tokyo", year=2026, name=OFFICIAL_NAME, its_wb=None):
    """公式資料（と its 用の手作りブック）を一時フォルダに保存して読む。"""
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, name)
        wb.save(path)
        its_path = os.path.join(d, "令和8年度_表.xlsx")
        if its_wb is not None:
            its_wb.save(its_path)
        return load_grade_table(path, insurer, year, its_rates_xlsx=its_path)


class OfficialFormatTests(unittest.TestCase):
    """Codex の公式資料（協会けんぽ様式＋厚生年金_公式表）を直接読む。"""

    def test_kyokai_tokyo_reads_everything_from_the_file(self):
        m = load_official(build_official_workbook())
        self.assertEqual(m.snapshot_ym, "2026-09")
        self.assertEqual(m.year, 2026)
        self.assertEqual([g.kenpo_smr for g in m.grades], [58000, 68000, 78000])
        self.assertEqual([g.konen_grade for g in m.grades], [1, 1, 2])
        self.assertEqual([g.konen_smr for g in m.grades], [88000, 88000, 98000])
        self.assertIsNone(m.grades[-1].upper)
        self.assertAlmostEqual(m.rates["kenpo"].total, 0.0985)
        self.assertAlmostEqual(m.rates["kenpo"].employee, 0.04925)
        self.assertAlmostEqual(m.rates["kaigo"].total, 0.0162)
        self.assertAlmostEqual(m.rates["kodomo"].employee, 0.00115)
        self.assertAlmostEqual(m.rates["konen"].employee, 0.0915)
        self.assertEqual(m.rates["kenpo"].applies, "令和8年3月分~適用")
        self.assertIn("kyoukaikenpo", m.rates["kenpo"].source_url)
        self.assertIn("nenkin", m.rates["konen"].source_url)
        self.assertIn("2026-09", m.description)

    def test_its_takes_kenpo_and_kaigo_rates_from_handmade_workbook(self):
        """関東IT健保の健康・介護は公式資料に無いので手作りブックから。支援金・厚年は公式資料。"""
        m = load_official(build_official_workbook(), insurer="its", its_wb=build_workbook())
        self.assertAlmostEqual(m.rates["kenpo"].employee, 0.04635)
        self.assertAlmostEqual(m.rates["kaigo"].employee, 0.009)
        self.assertAlmostEqual(m.rates["kodomo"].employee, 0.00115)
        self.assertAlmostEqual(m.rates["konen"].employee, 0.0915)
        self.assertIn("nenkin", m.rates["konen"].source_url)
        self.assertTrue(m.its_rates_path.endswith("令和8年度_表.xlsx"))
        self.assertIn("関東IT健保の料率は", m.description)

    def test_its_without_handmade_workbook_stops(self):
        with self.assertRaisesRegex(ShahoMasterError, "手作りブック"):
            load_official(build_official_workbook(), insurer="its")

    def test_its_stops_when_handmade_common_rate_is_stale(self):
        """公式資料の支援金率が変わったのに手作りブックが古いまま → 止めて直す場所を示す。"""
        with self.assertRaisesRegex(ShahoMasterError, "設定・出典.*子ども・子育て支援金"):
            load_official(build_official_workbook(rates={"kodomo": 0.0025}),
                          insurer="its", its_wb=build_workbook())

    def test_its_stops_when_handmade_workbook_is_last_years(self):
        old = build_workbook(year_text="令和7年度 標準報酬月額・本人負担額一覧")
        with self.assertRaisesRegex(ShahoMasterError, "関東IT健保の料率の元"):
            load_official(build_official_workbook(), insurer="its", its_wb=old)

    def test_year_mismatch_stops(self):
        with self.assertRaisesRegex(ShahoMasterError, "年度"):
            load_official(build_official_workbook(), year=2027)

    def test_filename_month_must_match_about_sheet(self):
        with self.assertRaisesRegex(ShahoMasterError, "対象年月"):
            load_official(build_official_workbook(about_ym="2026年10月"))

    def test_stale_premium_column_stops(self):
        with self.assertRaisesRegex(ShahoMasterError, "折半額"):
            load_official(build_official_workbook(premium_override={"2(1)": {"G": 3300.0}}))

    def test_missing_konen_premium_inside_band_stops(self):
        """厚年の帯の内側なのに厚年の額が空欄 → 表が壊れている。

        端の等級（実物の1〜3・36等級以上）は空欄が正しいので、厚年を3等級にして
        真ん中の等級を空欄にして確かめる。
        """
        konen = [(1, 88000, 0, 63000), (2, 98000, 63000, 73000), (3, 104000, 73000, None)]
        rows = [(1, 58000, None, 63000, None), ("2(2)", 68000, 63000, 73000, None),
                ("3(3)", 78000, 73000, None, 104000)]
        with self.assertRaisesRegex(ShahoMasterError, "空欄"):
            load_official(build_official_workbook(rows=rows, konen_rows=konen))

    def test_paren_grade_must_match_konen_table(self):
        rows = [(1, 58000, None, 63000, None), ("2(2)", 68000, 63000, 73000, 88000),
                ("3(2)", 78000, 73000, None, 98000)]
        with self.assertRaisesRegex(ShahoMasterError, "括弧内"):
            load_official(build_official_workbook(rows=rows))

    def test_rate_list_sheet_must_agree(self):
        with self.assertRaisesRegex(ShahoMasterError, "都道府県料率一覧"):
            load_official(build_official_workbook(listed_override={"kenpo": 0.0990}))

    def test_konen_rate_must_agree_between_sheets(self):
        with self.assertRaisesRegex(ShahoMasterError, "厚生年金の料率"):
            load_official(build_official_workbook(konen_sheet_pct=18.5))

    def test_handmade_workbook_still_loads_as_before(self):
        """作りの見分け: 手作りブックは従来どおり読める。"""
        m = load_from_workbook(build_workbook())
        self.assertIsNone(m.snapshot_ym)
        self.assertIn("手作り", m.description)


class RealFileTests(unittest.TestCase):
    """実在の令和8年度Excelを読む（共有フォルダが無ければスキップ）。"""

    @classmethod
    def setUpClass(cls):
        if not os.path.exists(REAL_XLSX):
            raise unittest.SkipTest(f"見つかりません: {REAL_XLSX}")
        cls.master = load_grade_table(REAL_XLSX, "its", 2026)

    def test_shape(self):
        self.assertEqual(len(self.master.grades), 50)
        self.assertEqual(self.master.grades[0].kenpo_smr, 58000)
        self.assertEqual(self.master.grades[-1].kenpo_smr, 1390000)
        self.assertEqual(len(self.master.konen_bands), 32)
        self.assertEqual(self.master.konen_bands[-1][1], 650000)

    def test_known_premium_reproduces(self):
        """実測値の再現: 標報620,000の健保本人負担 = 28,737円（2026-07 の控除実績）。"""
        row = self.master.find_grade(620000)
        self.assertEqual(row.kenpo_smr, 620000)
        self.assertEqual(round(row.kenpo_smr * self.master.rates["kenpo"].employee), 28737)
        self.assertEqual(round(row.konen_smr * self.master.rates["konen"].employee), 56730)

    def test_konen_caps_at_650000(self):
        """高報酬でも厚年は650,000で頭打ち（健保は1,390,000まで伸びる）。"""
        row = self.master.find_grade(2000000)
        self.assertEqual(row.kenpo_smr, 1390000)
        self.assertEqual(row.konen_smr, 650000)

    def test_kyokai_tokyo_also_loads(self):
        m = load_grade_table(REAL_XLSX, "kyokai_tokyo", 2026)
        self.assertAlmostEqual(m.rates["kenpo"].employee, 0.04925)

    def test_wrong_year_stops(self):
        with self.assertRaises(ShahoMasterError):
            load_grade_table(REAL_XLSX, "its", 2027)


class RealOfficialFileTests(unittest.TestCase):
    """Codex が置いた実在の公式資料（2026年9月）を読む（無ければスキップ）。"""

    @classmethod
    def setUpClass(cls):
        if not (os.path.exists(OFFICIAL_XLSX) and os.path.exists(REAL_XLSX)):
            raise unittest.SkipTest(f"見つかりません: {OFFICIAL_XLSX}")
        cls.master = load_grade_table(OFFICIAL_XLSX, "its", 2026, its_rates_xlsx=REAL_XLSX)

    def test_shape_matches_handmade_workbook(self):
        hand = load_grade_table(REAL_XLSX, "its", 2026)
        self.assertEqual(self.master.snapshot_ym, "2026-09")
        self.assertEqual(self.master.grades, hand.grades)
        self.assertEqual(self.master.konen_bands, hand.konen_bands)
        for key in ("kenpo", "kodomo", "kaigo", "konen"):
            self.assertAlmostEqual(self.master.rates[key].employee, hand.rates[key].employee)

    def test_known_premium_reproduces(self):
        row = self.master.find_grade(620000)
        self.assertEqual(round(row.kenpo_smr * self.master.rates["kenpo"].employee), 28737)
        self.assertEqual(round(row.konen_smr * self.master.rates["konen"].employee), 56730)

    def test_kyokai_tokyo_from_official_only(self):
        m = load_grade_table(OFFICIAL_XLSX, "kyokai_tokyo", 2026)
        self.assertAlmostEqual(m.rates["kenpo"].employee, 0.04925)
        self.assertAlmostEqual(m.rates["kaigo"].employee, 0.0081)
        self.assertEqual(m.find_grade(620000).kenpo_smr, 620000)

    def test_select_picks_the_september_snapshot(self):
        from services.shaho_master import select_grade_table
        c = select_grade_table("2026-09", REAL_XLSX)
        self.assertEqual(os.path.abspath(c.path), os.path.abspath(OFFICIAL_XLSX))


if __name__ == "__main__":
    unittest.main()
