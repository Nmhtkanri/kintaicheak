# -*- coding: utf-8 -*-
"""健診HPMモードのテスト用フィクスチャ（テストではなく共有ヘルパー）

実データ（要配慮個人情報）はテストに置かない。実物の形だけを真似た
最小のブックを openpyxl でその場で組む。

実物のレイアウトに合わせる点:
  - どのデータシートも 1行目タイトル / 2行目説明 / 3行目ヘッダー / 4行目〜データ
  - 受診者一覧の性別は "男性"/"女性"、受診日は "2026-07-01 00:00:00" のような文字列
  - 受診No. は "132"（3桁）で来る＝CSV側で "000132" に埋めるのが正
"""

from __future__ import annotations

import json
import os
from pathlib import Path

import openpyxl

_HERE = Path(__file__).resolve().parent

RECIPIENT_HEADERS = ["PDFページ", "氏名", "年齢", "性別", "受診日", "受診No.", "個人票シート"]
ITEM_HEADERS_V1 = ["PDFページ", "氏名", "受診日", "分類", "項目", "値", "単位",
                   "原票判定", "原票注記", "個人票シート"]
ITEM_HEADERS_V2 = ITEM_HEADERS_V1 + ["測定回", "値種別", "原票表記"]


def canonical_header() -> list[str]:
    """正本CSV由来の302列ヘッダー。"""
    path = _HERE / "sample_data" / "hpm_canonical_header_302.json"
    return json.loads(path.read_text(encoding="utf-8"))["columns"]


def item(category, name, value, *, unit="", occurrence=None, value_type="数値",
         judgement="", note="", original=None):
    """項目別データ1行分。"""
    return {
        "category": category,
        "item": name,
        "value": value,
        "unit": unit,
        "occurrence": occurrence,
        "value_type": value_type,
        "judgement": judgement,
        "note": note,
        "original": original if original is not None else value,
    }


def bp_items(sys1=None, dia1=None, sys2=None, dia2=None, sys3=None, dia3=None,
             judgement=""):
    """血圧の項目行。**平均は作らない**ので、渡した回だけが行になる。"""
    out = []
    for occurrence, (s, d) in enumerate([(sys1, dia1), (sys2, dia2), (sys3, dia3)], start=1):
        if s is not None:
            out.append(item("血圧", "収縮期血圧", s, unit="mmHg",
                            occurrence=occurrence, judgement=judgement))
        if d is not None:
            out.append(item("血圧", "拡張期血圧", d, unit="mmHg",
                            occurrence=occurrence, judgement=judgement))
    return out


def person(name="友納 英彦", *, age=58, gender="男性", exam_date="2026-07-01 00:00:00",
           exam_no="132", sheet=None, items=None):
    return {
        "name": name,
        "age": age,
        "gender": gender,
        "exam_date": exam_date,
        "exam_no": exam_no,
        "sheet": sheet or ("P%02d_%s" % (2, name.replace(" ", ""))),
        "items": items if items is not None else [],
    }


def _write_sheet(wb, title, headers, rows, description=""):
    ws = wb.create_sheet(title)
    ws.cell(1, 1).value = title
    if description:
        ws.cell(2, 1).value = description
    for c, name in enumerate(headers, start=1):
        ws.cell(3, c).value = name
    for r, row in enumerate(rows, start=4):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value
    return ws


def make_v2_workbook(path, persons, *, schema_version="2.0", genpyo=True,
                     bp_kept=True, qual_kept=True, v2_columns=True,
                     info_extra=None):
    """スキーマv2の整形済みExcelを作る。

    v2_columns=False にすると測定回・値種別・原票表記の3列を落とし、
    旧スキーマ（v1）相当のブックになる。
    """
    path = str(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- 変換案内 ---
    info = wb.create_sheet("変換案内")
    info.cell(1, 1).value = "健康診断PDF - Excel変換版"
    info.cell(3, 1).value = "項目"
    info.cell(3, 2).value = "内容"
    pairs = [("原本", r"Z:\dummy\source.pdf"), ("対象", "テスト用")]
    if schema_version is not None:
        pairs.append(("健康診断整形スキーマ版", schema_version))
    pairs += [
        ("原票確認済み", "TRUE" if genpyo else "FALSE"),
        ("血圧測定回保持", "TRUE" if bp_kept else "FALSE"),
        ("定性値保持", "TRUE" if qual_kept else "FALSE"),
    ]
    for key, value in (info_extra or {}).items():
        pairs.append((key, value))
    for r, (key, value) in enumerate(pairs, start=4):
        info.cell(r, 1).value = key
        info.cell(r, 2).value = value

    # --- 受診者一覧 ---
    rec_rows = []
    for i, p in enumerate(persons, start=2):
        rec_rows.append([str(i), p["name"], p["age"], p["gender"],
                         p["exam_date"], p["exam_no"], p["sheet"]])
    _write_sheet(wb, "受診者一覧", RECIPIENT_HEADERS, rec_rows)

    # --- 項目別データ ---
    headers = ITEM_HEADERS_V2 if v2_columns else ITEM_HEADERS_V1
    item_rows = []
    for i, p in enumerate(persons, start=2):
        for it in p["items"]:
            row = [str(i), p["name"], p["exam_date"], it["category"], it["item"],
                   it["value"], it["unit"], it["judgement"], it["note"], p["sheet"]]
            if v2_columns:
                row += [it["occurrence"], it["value_type"], it["original"]]
            item_rows.append(row)
    _write_sheet(wb, "項目別データ", headers, item_rows, description="出典: テスト")

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 変換マスタ
# ---------------------------------------------------------------------------

# 血圧4行は load_master が必須で見る。既定はこれに数項目を足したもの。
DEFAULT_ITEM_MAP = [
    ("身体計測", "身長", 1, 24, "数値", "", ""),
    ("身体計測", "体重", 1, 25, "数値", "", ""),
    ("身体計測", "BMI", 1, 27, "数値", "", ""),
    ("血圧", "収縮期血圧", 1, 50, "数値", "", ""),
    ("血圧", "拡張期血圧", 1, 51, "数値", "", ""),
    ("血圧", "収縮期血圧", 2, 52, "数値", "", ""),
    ("血圧", "拡張期血圧", 2, 53, "数値", "", ""),
    ("尿検査", "尿蛋白", 1, 54, "定性", "", ""),
    ("尿検査", "尿糖", 1, 55, "定性", "", ""),
    ("便潜血", "便潜血", 1, 70, "定性", "", ""),
    ("便潜血", "便潜血", 2, 71, "定性", "", ""),
    ("感染症", "HBs抗原", 1, 115, "定性", "MAT", ""),
    ("感染症", "HBs抗原", 1, 117, "定性", "CLIA", ""),
    ("血清反応", "CRP", 1, 148, "数値", "", ""),
]

DEFAULT_INSTITUTIONS = [
    ("医療法人社団 同友会 春日クリニック", "1310528885", "", "TRUE", ""),
    ("医療法人徳洲会 生駒市立病院", "0301619", "", "TRUE", "前ゼロが落ちないこと"),
    ("未確認クリニック", "9999999999", "", "FALSE", "HPMで未確認"),
]

DEFAULT_ALIASES = [("同友会", "医療法人社団 同友会 春日クリニック")]

DEFAULT_COURSES = [
    ("医療法人社団 同友会 春日クリニック", "人間ドックC 胃カメラ（40歳以上）",
     "人間ドックＣ　胃カメラ（４０歳以上）"),
    ("医療法人社団 同友会 春日クリニック", "定期健康診断", "10"),
    ("医療法人徳洲会 生駒市立病院", "人間ドックC（胃カメラ）", "13"),
    ("未確認クリニック", "定期健康診断", "10"),
]

DEFAULT_SETTINGS = [("会場コード", "2", "")]


def make_master_xlsx(path, *, header=None, institutions=None, aliases=None,
                     courses=None, settings=None, item_map=None, break_rule=None):
    """変換マスタの最小版を作る。break_rule で壊し方を指定できる。"""
    path = str(path)
    header = list(header if header is not None else canonical_header())
    institutions = list(institutions if institutions is not None else DEFAULT_INSTITUTIONS)
    aliases = list(aliases if aliases is not None else DEFAULT_ALIASES)
    courses = list(courses if courses is not None else DEFAULT_COURSES)
    settings = list(settings if settings is not None else DEFAULT_SETTINGS)
    item_map = list(item_map if item_map is not None else DEFAULT_ITEM_MAP)

    if break_rule == "bp_shift":
        # 2回目の収縮期を別の列へずらす（平均のような値になる事故の再現）
        item_map = [r if not (r[1] == "収縮期血圧" and r[2] == 2)
                    else (r[0], r[1], r[2], 60, r[4], r[5], r[6]) for r in item_map]
    elif break_rule == "missing_bp2":
        item_map = [r for r in item_map if not (r[0] == "血圧" and r[2] == 2)]
    elif break_rule == "bp_header_name":
        header = list(header)
        header[52] = "血圧（平均）最高"
    elif break_rule == "header_301":
        header = list(header)[:301]
    elif break_rule == "map_judgement":
        item_map = item_map + [("血圧", "血圧判定", 1, 186, "文字", "", "")]
    elif break_rule == "map_identity":
        item_map = item_map + [("身体計測", "生年月日もどき", 1, 8, "文字", "", "")]
    elif break_rule == "dup_key":
        item_map = item_map + [("身体計測", "身長", 1, 26, "数値", "", "")]
    elif break_rule == "dup_col":
        item_map = item_map + [("身体計測", "身長もどき", 1, 24, "数値", "", "")]
    elif break_rule == "no_venue":
        settings = [s for s in settings if s[0] != "会場コード"]
    elif break_rule == "alias_orphan":
        aliases = aliases + [("よその健診", "存在しないクリニック")]
    elif break_rule == "course_orphan":
        courses = courses + [("存在しないクリニック", "定期健康診断", "10")]
    elif break_rule == "col_out_of_range":
        item_map = item_map + [("身体計測", "範囲外", 1, 999, "数値", "", "")]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "健診機関",
                 ["機関名", "HPM場所コード", "公開医療機関コード参考", "HPM確認済み", "備考"],
                 institutions)
    _write_sheet(wb, "機関別名", ["別名", "正式機関名", "備考"],
                 [(a, b, "") for a, b in aliases])
    _write_sheet(wb, "健診種別", ["機関名", "種別表示名", "HPM出力値", "備考"],
                 [(a, b, c, "") for a, b, c in courses])
    _write_sheet(wb, "設定", ["キー", "値", "備考"], settings)
    _write_sheet(wb, "項目マッピング",
                 ["分類", "項目名", "測定回", "HPM列番号", "値種別", "検査方式", "備考"],
                 item_map)
    _write_sheet(wb, "HPMヘッダー", ["列番号", "列名", "用途"],
                 [(i, name, "") for i, name in enumerate(header)])
    if break_rule == "missing_sheet":
        wb.remove(wb["機関別名"])
    wb.save(path)
    return path


def employees_stub() -> list[dict]:
    """jinjer get_employees() の生dictを模したもの。

    実物のネスト（company / personal）に合わせる。照合ロジックを試すため、
    同姓同名・性別違い・自社形式でない社員番号・生年月日なしを混ぜてある。
    """
    def emp(eid, last, first, last_kana, first_kana, birth, gender):
        return {
            "id": eid,
            "company": {
                "last_name": last,
                "first_name": first,
                "last_name_phonetic": last_kana,
                "first_name_phonetic": first_kana,
            },
            "personal": {
                "date_of_birth": birth,
                "gender": {"name": gender} if gender else None,
            },
        }

    return [
        emp("2018013", "友納", "英彦", "トモノウ", "ヒデヒコ", "1968-04-13", "男性"),
        emp("2019022", "高橋", "和紀", "タカハシ", "カズノリ", "1978-12-25", "男性"),
        emp("2020031", "大坪", "淳一", "オオツボ", "ジュンイチ", "1984-10-24", "男性"),
        emp("2021045", "中澤", "寿代", "ナカザワ", "ヒサヨ", "1970-02-08", "女性"),
        # 同姓同名（自動確定してはいけない）
        emp("2022001", "吉田", "拓矢", "ヨシダ", "タクヤ", "1990-05-05", "男性"),
        emp("2022002", "吉田", "拓矢", "ヨシダ", "タクヤ", "1992-08-08", "男性"),
        # 自社形式でない社員番号（候補から外れる）
        emp("5551234", "派遣", "太郎", "ハケン", "タロウ", "1980-01-01", "男性"),
        emp("3333008", "テスト", "花子", "テスト", "ハナコ", "1985-03-03", "女性"),
        # 生年月日なし
        emp("2023077", "熊崎", "俊輔", "クマザキ", "シュンスケ", None, "男性"),
    ]
