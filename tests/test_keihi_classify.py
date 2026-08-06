# -*- coding: utf-8 -*-
"""経費分類・集計エンジン（マクロ移植 P1b）のユニットテスト。

実データ検証: live 統合一覧表(2401行)で live 集計シートと 173名全員のC〜L完全一致・
集計ログ2457行一致を確認済み（2026-07-16）。ここでは判定順序・エッジケースを固定する。
"""
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.keihi_classify import (  # noqa: E402
    DEFAULT_KEYWORDS, CAT_YAKAN, CAT_TRANS, CAT_NONTAX,
    normalize_str, normalize_id, parse_amount_vba, try_parse_date, hit_any,
    build_emp_key, parse_emp_no, classify_rows, aggregate_by_id,
    build_detail_totals, add_summary_sheet, add_log_sheet, load_keywords,
    B_YAKAN, B_RINK, B_TRANS, B_ETC, B_TW, B_BILL, B_NONTAX,
    COL_EMP, COL_NAME, COL_AMT, COL_MEMO_REQ, COL_TRANS, COL_UCH,
    COL_BILLTYPE, COL_EXPTYPE, COL_FARE, COL_MEMO_LINE, COL_BOOKED,
    COL_SOURCE, COL_ESTAFF,
)
from services.keihi_payroll_import import (  # noqa: E402
    build_import_rows, render_import_csv, IMPORT_HEADERS,
)


def _row(**kw):
    r = [""] * 34
    r[COL_EMP] = kw.get("emp", "2026013")
    r[COL_NAME] = kw.get("name", "川口 祐ノ輔")
    r[COL_AMT] = kw.get("amt", "")
    r[COL_UCH] = kw.get("uch", "")
    r[COL_TRANS] = kw.get("trans", "")
    r[COL_BILLTYPE] = kw.get("billtype_id", "")
    r[COL_EXPTYPE] = kw.get("exptype", "")
    r[COL_MEMO_REQ] = kw.get("memo_req", "")
    r[COL_MEMO_LINE] = kw.get("memo_line", "")
    r[COL_FARE] = kw.get("fare", "")
    r[COL_BOOKED] = kw.get("booked", "")
    r[COL_SOURCE] = kw.get("source", "")
    r[COL_ESTAFF] = kw.get("estaff", "")
    return r


def _bucket(rows):
    cls = classify_rows(rows)
    assert len(cls.agg) == 1
    return next(iter(cls.agg.values())), cls


# ----------------------------------------------------------------------
# VBA互換ヘルパー
# ----------------------------------------------------------------------

def test_normalize_str_vba():
    # CR/LF/Tab/全角空白→半角空白→Trim→LCase→全空白除去
    assert normalize_str(" A B\r\nC　D\t ") == "abcd"
    assert normalize_str(None) == ""


def test_parse_amount_vba_negative_parens():
    assert parse_amount_vba("1,375") == 1375
    assert parse_amount_vba("(123)") == -123
    assert parse_amount_vba("（4,560円）") == -4560
    assert parse_amount_vba("abc") == 0
    assert parse_amount_vba("") == 0


def test_try_parse_date():
    assert try_parse_date("2026/6/30") == date(2026, 6, 30)
    assert try_parse_date("2026-06-01") == date(2026, 6, 1)
    assert try_parse_date("") is None
    assert try_parse_date("30日") is None


def test_hit_any_order_first_wins():
    # 登録順の先勝ち: 交通費リストでは「通勤交通費（実費）」が「交通費」より先
    kw = DEFAULT_KEYWORDS[CAT_TRANS]
    assert hit_any(normalize_str("通勤交通費（実費）"), kw) == "通勤交通費（実費）".lower()
    # 部分一致: "jr西日本" は "JR"（LCase格納）にヒット
    assert hit_any(normalize_str("JR西日本"), kw) == "jr"
    assert hit_any("該当なしのテキスト", DEFAULT_KEYWORDS[CAT_YAKAN]) == ""


def test_emp_key_roundtrip():
    key = build_emp_key("2026013", "かわぐち")
    assert parse_emp_no(key) == "2026013"
    assert parse_emp_no(build_emp_key("", "かわぐち")) == ""


def test_normalize_id_digits_only():
    assert normalize_id("2026013") == "2026013"
    assert normalize_id("該当なし") == ""
    assert normalize_id(" 2026-013 ") == "2026013"


# ----------------------------------------------------------------------
# 分類の優先順位（Collect_From_Source 移植）
# ----------------------------------------------------------------------

def test_priority_0_honsha_to_nontax():
    b, _ = _bucket([_row(amt="500", uch="会議費", source="本社経費")])
    assert b[B_NONTAX] == 500 and b[B_ETC] == 0


def test_priority_1_yakan():
    b, _ = _bucket([_row(amt="2500", uch="夜間当番手当(旧_顧客対応当番手当)")])
    assert b[B_YAKAN] == 2500


def test_priority_2_telework_label_is_J():
    # 集計先はK(bucket4)だがログの判定結果はマクロのまま "J:テレワーク手当"
    b, cls = _bucket([_row(amt="3000", uch="テレワーク手当")])
    assert b[B_TW] == 3000
    assert any(e.result == "J:テレワーク手当" for e in cls.log)


def test_priority_3_rink():
    b, _ = _bucket([_row(amt="1000", uch="RINK日当（平日）1")])
    assert b[B_RINK] == 1000


def test_priority_4_nontax_uses_nontax_text_not_memo():
    # 癖1修正: 備考に旧分類名が残っていても非課税精算に引っ張られない
    b, _ = _bucket([_row(amt="356", trans="通勤交通費（実費）",
                         memo_line="交通費（電車・バス）⇒通勤交通費（実費）に変更")])
    # nonTaxText(内訳+交通機関+費用種別)に非課税KWなし → #5交通費でH
    assert b[B_TRANS] == 356 and b[B_NONTAX] == 0


def test_priority_4_nontax_by_transport():
    b, _ = _bucket([_row(amt="356", trans="交通費（電車・バス）")])
    assert b[B_NONTAX] == 356      # 交通機関が本当に非課税KWの行はI


def test_priority_5_trans_ng_goes_to_etc():
    # 内訳が交通費KWでも交通費除外KW（会議等）に当たればその他(J)
    b, cls = _bucket([_row(amt="800", uch="会議への移動")])
    assert b[B_ETC] == 800 and b[B_TRANS] == 0
    assert any("交通費NG" in e.result for e in cls.log)


def test_priority_6_etc_via_judgetext():
    # judgeText(備考含む)でその他KWにヒット
    b, _ = _bucket([_row(amt="1200", uch="雑品", memo_line="懇親会の飲み物")])
    assert b[B_ETC] == 1200


def test_priority_7_default_etc():
    b, cls = _bucket([_row(amt="999", uch="未知の内訳XYZ")])
    assert b[B_ETC] == 999
    assert any(e.matched_kw == "(該当キーワードなし)" for e in cls.log)


def test_judgetext_uses_billtype_id_column():
    # judgeTextの請求区分成分は請求区分ID列(0-based 8)。ID列に「会議」を置けば#6でヒットする
    b, _ = _bucket([_row(amt="700", uch="不明品目", billtype_id="会議")])
    assert b[B_ETC] == 700


# ----------------------------------------------------------------------
# G先行・estFilledガード・金額フォールバック
# ----------------------------------------------------------------------

def test_g_precheck_requires_emp_no():
    # 社員番号が無い行は顧客請求分(G)に加算されない
    cls = classify_rows([_row(emp="該当なし", name="稲場", estaff="1000")])
    assert all(v[B_BILL] == 0 for v in cls.agg.values()) or not cls.agg


def test_g_precheck_excluded_by_kokyaku_ng():
    # 顧客請求除外KW(夜間当番)に当たる → G加算なし（aggエントリ自体できない）・"G:除外" ログ
    cls = classify_rows([_row(amt="", uch="夜間当番手当", estaff="2500")])
    assert not cls.agg
    assert any(e.result == "G:除外" for e in cls.log)


def test_estfilled_guard_skips_amount_and_date():
    # 顧客請求分あり＆夜間当番でない → 金額分類スキップ（計上日もスキップ=GoTo NextR）
    rows = [_row(amt="1160", uch="客先請求分（交通費）", estaff="1160", booked="2026/6/30")]
    cls = classify_rows(rows)
    b = next(iter(cls.agg.values()))
    assert b[B_BILL] == 1160          # G先行は効く
    assert b[B_TRANS] == 0 and b[B_ETC] == 0   # 金額は分類されない
    assert not cls.max_date            # 計上日もスキップ
    assert any(e.result == "D:顧客請求費ありのため除外" for e in cls.log)


def test_estfilled_zero_string_still_guards():
    # AH="0" でも estFilled=True（文字列非空判定）
    rows = [_row(amt="500", uch="不明", estaff="0")]
    cls = classify_rows(rows)
    assert any(e.result == "D:顧客請求費ありのため除外" for e in cls.log)


def test_yakan_passes_guard_even_with_estaff():
    # estFilledでも内訳が夜間当番ならD計上（ガードを通過）
    b, _ = _bucket([_row(amt="2500", uch="顧客対応当番16", estaff="x")])
    assert b[B_YAKAN] == 2500


def test_amount_fallback_fare_positive_only():
    b, _ = _bucket([_row(amt="0", fare="178", trans="交通費（電車・バス）")])
    assert b[B_NONTAX] == 178          # 合計0→金額(交通費)>0を採用
    cls2 = classify_rows([_row(amt="0", fare="-50", uch="謎")])
    assert not any(e.result.startswith(("D:", "H:", "I:", "J:", "E:")) for e in cls2.log)


def test_booked_date_max():
    rows = [
        _row(amt="100", trans="交通費（電車・バス）", booked="2026/6/1"),
        _row(amt="100", trans="交通費（電車・バス）", booked="2026/6/30"),
        _row(amt="", booked="2026/7/15"),   # 金額0でも計上日は対象
    ]
    cls = classify_rows(rows)
    assert list(cls.max_date.values())[0] == date(2026, 7, 15)


# ----------------------------------------------------------------------
# 再集約・集計シート・新入社員包含
# ----------------------------------------------------------------------

def test_aggregate_merges_same_id_and_drops_name_only():
    rows = [
        _row(emp="2026013", name="川口 祐ノ輔", amt="100", trans="交通費（電車・バス）"),
        _row(emp="2026013", name="川口　祐ノ輔", amt="200", trans="交通費（電車・バス）"),  # 氏名表記ゆれ
        _row(emp="該当なし", name="不明 太郎", amt="300", uch="雑費"),                      # ID無し
    ]
    agg = aggregate_by_id(classify_rows(rows))
    assert agg.by_id["2026013"][B_NONTAX] == 300
    assert agg.unmatched_rows == 1 and agg.unmatched_amount == 300


def test_summary_sheet_includes_new_employee_and_excludes_569():
    from openpyxl import Workbook
    rows = [
        _row(emp="2026013", name="川口 祐ノ輔", amt="100", trans="交通費（電車・バス）", booked="2026/6/30"),
        _row(emp="5000001", name="派遣 太郎", amt="999", uch="雑費"),
    ]
    cls = classify_rows(rows)
    agg = aggregate_by_id(cls)
    wb = Workbook()
    stats = add_summary_sheet(wb, agg, cls.emp_names, build_detail_totals(rows))
    ws = wb["集計"]
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "2026013" in ids            # 新入社員が自動で行になる（事故の根本対策）
    assert "5000001" not in ids        # 5始まりは給与計算対象外
    assert stats["excluded_out_of_scope"] == 1
    # C=合計、L=請求日、M=明細合計、O=判定
    i = ids.index("2026013") + 2
    assert ws.cell(row=i, column=3).value == 100
    assert ws.cell(row=i, column=12).value == date(2026, 6, 30)
    assert ws.cell(row=i, column=15).value == "OK"


def test_detail_totals_d_or_ah():
    rows = [
        _row(emp="2026013", amt="100"),           # D非空 → D
        _row(emp="2026013", amt="", estaff="50"), # D空欄 → AH
    ]
    assert build_detail_totals(rows)["2026013"] == 150


def test_log_sheet_headers():
    from openpyxl import Workbook
    cls = classify_rows([_row(amt="100", trans="交通費（電車・バス）")])
    wb = Workbook()
    add_log_sheet(wb, cls.log)
    ws = wb["集計ログ"]
    assert [c.value for c in ws[1]] == ["行番号", "社員番号", "氏名", "内訳", "金額", "判定結果", "マッチしたキーワード"]


def test_load_keywords_csv(tmp_path):
    import csv as _csv
    p = tmp_path / "kw.csv"
    with open(p, "w", encoding="cp932", newline="") as f:
        w = _csv.writer(f)
        w.writerow(["分類名", "キーワード"])
        w.writerow([CAT_YAKAN, "夜間当番"])
        w.writerow([CAT_YAKAN, ""])                 # 空KWはスキップ（R69対応）
        w.writerow([CAT_YAKAN, "顧客当番", "手当"])   # C列以降は無視（R6対応）
        w.writerow([CAT_NONTAX, "日当"])
    kw = load_keywords(p)
    assert kw[CAT_YAKAN] == ["夜間当番", "顧客当番"]
    assert kw[CAT_NONTAX] == ["日当"]


# ----------------------------------------------------------------------
# jinjer給与インポート行（P2）
# ----------------------------------------------------------------------

def test_build_import_rows_mapping():
    by_id = {
        # [夜間, RINK, 交通費, その他, テレワーク, 顧客請求, 非課税精算]
        "2026013": [2500.0, 1000.0, 5000.0, 300.0, 400.0, 8990.0, 1160.0],
        "5000001": [1.0, 0, 0, 0, 0, 0, 0],   # 対象外
    }
    rows, warnings, _ = build_import_rows(by_id, {"2026013": "川口 祐ノ輔"})
    assert len(rows) == 1
    r = rows[0]
    assert r["夜間当番手当"] == 3500            # D+E（F列相当。live集計R/S実測に基づく）
    assert r["非課税通勤費"] == 5000            # H交通費
    assert r["立替金（顧客請求分）"] == 8990     # G
    assert r["立替金"] == 1160                  # I非課税精算
    assert r["その他"] == 300                   # J
    assert r["定常外業務対応手当"] == 0 and r["その他手当"] == 0 and r["現物支給"] == 0
    assert r["支給過不足調整"] == 0
    assert r["テレワーク手当"] == 400           # K（テンプレにテレワーク手当列があれば搬送される）


# ----------------------------------------------------------------------
# 通勤費の月額上限カット（2026-08-06）
# ----------------------------------------------------------------------

def _by_id(trans):
    """交通費(H)だけ入った集計を作る。"""
    return {"2014013": [0, 0, float(trans), 0, 0, 0, 0]}


def test_commute_limit_cuts_only_the_commute_portion():
    """上限を超えた分だけ減らす。同じ交通費(H)の駐車場代などは切らない。"""
    # 交通費(H)=37,660 の内訳: 通勤費35,090 + 駐車場2,570
    rows, warns, cuts = build_import_rows(
        _by_id(37660), {"2014013": "柴田 和浩"},
        commute_by_id={"2014013": 35090.0}, commute_other_by_id={"2014013": 2570.0},
        commute_limit=30000)

    # 30,000 + 上限対象外2,570 = 32,570（切ったのは 5,090 だけ）
    assert rows[0]["非課税通勤費"] == 32570
    assert len(cuts) == 1
    assert (cuts[0]["カット額"], cuts[0]["うち通勤費"], cuts[0]["うち上限対象外"]) == (5090, 35090, 2570)
    assert any("上限カット" in w for w in warns)
    assert any("通勤系と判定できない" in w for w in warns)


def test_commute_limit_leaves_amounts_within_the_limit_untouched():
    rows, warns, cuts = build_import_rows(
        _by_id(29999), {"2014013": "柴田 和浩"},
        commute_by_id={"2014013": 29999.0}, commute_limit=30000)
    assert rows[0]["非課税通勤費"] == 29999
    assert cuts == [] and warns == []


def test_commute_limit_skips_exempt_members_and_travel_members():
    """免除者と移動交通費対象者は切らない。ただし黙って通さず理由を出す。"""
    for kwargs, kw in (({"limit_exempt": {"2014013": "個別許可"}}, "上限免除者"),
                       ({"travel_members": {"2014013": "柴田 和浩"}}, "移動交通費")):
        rows, warns, cuts = build_import_rows(
            _by_id(35090), {"2014013": "柴田 和浩"},
            commute_by_id={"2014013": 35090.0}, commute_limit=30000, **kwargs)
        assert rows[0]["非課税通勤費"] == 35090
        assert cuts == []
        assert any(kw in w for w in warns), f"{kw} の理由が出ていません"


def test_commute_limit_is_off_when_no_limit_given():
    """上限未指定なら従来どおり素通し（既存の動作を変えない）。"""
    rows, _, cuts = build_import_rows(
        _by_id(35090), {"2014013": "柴田 和浩"}, commute_by_id={"2014013": 35090.0})
    assert rows[0]["非課税通勤費"] == 35090 and cuts == []


def test_commute_totals_from_log_splits_commute_and_other():
    """交通費(H)のうち通勤系KWだけを上限対象として数える。"""
    from services.keihi_classify import LogEntry, ClassifyResult, commute_totals_from_log
    cls = ClassifyResult()
    cls.log = [
        LogEntry(2, "2014013", "柴田和浩", "", 35090.0, "H:交通費", "定期代"),
        LogEntry(3, "2021008", "野田祐介", "", 1800.0, "H:交通費", "駐車場"),
        LogEntry(4, "2021008", "野田祐介", "", 2736.0, "H:交通費", "通勤交通費（実費）"),
        LogEntry(5, "2014013", "柴田和浩", "", 9999.0, "I:非課税精算", "交通費（電車・バス）"),
    ]
    commute, other, _ = commute_totals_from_log(cls)
    assert commute == {"2014013": 35090.0, "2021008": 2736.0}   # I:非課税精算 は入らない
    assert other == {"2021008": 1800.0}


def test_commute_totals_from_log_collects_use_months():
    """定期代を2か月分まとめて申請している人を見分けるため利用月を集める。"""
    from services.keihi_classify import LogEntry, ClassifyResult, commute_totals_from_log
    cls = ClassifyResult()
    cls.log = [
        LogEntry(2, "2025007", "吉田拓矢", "", 20250.0, "H:交通費", "定期代"),
        LogEntry(3, "2025007", "吉田拓矢", "", 20250.0, "H:交通費", "定期代"),
    ]
    rows = [
        ["2025007", "吉田 拓矢", "2026/8/2", "", "", "2026/7/1"],
        ["2025007", "吉田 拓矢", "2026/8/2", "", "", "2026/8/1"],
    ]
    _, _, months = commute_totals_from_log(cls, rows)
    assert months == {"2025007": {"2026-07", "2026-08"}}


def test_commute_limit_does_not_cut_when_spanning_multiple_months():
    """2か月分まとめての申請は月ごとなら上限内かもしれないので自動で切らない。"""
    rows, warns, cuts = build_import_rows(
        {"2025007": [0, 0, 40500.0, 0, 0, 0, 0]}, {"2025007": "吉田 拓矢"},
        commute_by_id={"2025007": 40500.0}, commute_limit=30000,
        commute_months={"2025007": {"2026-07", "2026-08"}})

    assert rows[0]["非課税通勤費"] == 40500     # 切らない（過少支給を避ける）
    assert cuts == []
    assert any("2か月分" in w and "カットしていません" in w for w in warns)


def test_render_import_csv_sjis_and_quoting():
    rows, _, _ = build_import_rows({"2026013": [0, 0, 100.0, 0, 0, 0, 0]}, {"2026013": "川口 祐ノ輔"})
    data = render_import_csv(rows)
    text = data.decode("cp932")
    lines = text.strip().split("\r\n")
    assert lines[0] == ",".join(IMPORT_HEADERS)
    assert lines[1].startswith('"2026013","川口 祐ノ輔",0,0,0,100,')


def test_render_import_csv_follows_template_positions():
    """テンプレの列位置に追従する（jinjerは列位置で取込むため）。

    テンプレ37047の実レイアウト（社員番号,空,空,夜間当番手当,空,空,立替金(顧客),非課税通勤費,その他,テレワーク手当）
    で、夜間当番手当が位置4に、非課税通勤費が位置8に入ることを確認する。
    """
    from services.keihi_payroll_import import _match_column, check_template_coverage
    tpl = ["*社員番号", "", "", "夜間当番手当", "", "",
           "立替金（顧客請求分）", "非課税通勤費", "その他", "テレワーク手当"]
    # by_id: [夜間, RINK, 交通費(→非課税通勤費), その他, テレワーク, 顧客請求, 非課税精算(→立替金)]
    rows, _, _ = build_import_rows(
        {"2026013": [30000.0, 0, 0, 0, 0, 0, 12177.0]}, {"2026013": "川口 祐ノ輔"})
    data = render_import_csv(rows, tpl).decode("cp932")
    line = data.strip().split("\r\n")[1].split(",")
    assert line[0] == '"2026013"'
    assert line[1] == "" and line[2] == ""        # 空スキップ列
    assert line[3] == "30000"                     # 位置4=夜間当番手当 ✓
    assert line[7] == "0"                          # 位置8=非課税通勤費（交通費0）
    # 立替金(12177)はテンプレに列が無い → カバレッジ警告
    cov = check_template_coverage(rows, tpl)
    assert any("立替金" in w and "12,177" in w for w in cov)


def test_render_import_csv_template_44450_covers_all():
    """新テンプレ44450「経費APIインポート用」（2026-07-17 谷津作成・空DL検証済）。

    11列＝社員番号,空(氏名なし),夜間当番手当,定常外業務対応手当,支給過不足調整,
    非課税通勤費,立替金(顧客請求分),立替金,その他,その他手当,現物支給。
    旧37047で欠けていた「立替金」裸列が入り、経費の全支給項目をカバーする
    （テレワーク手当はそもそも給与に乗せないので対象外）。
    """
    from services.keihi_payroll_import import check_template_coverage
    tpl = ["*社員番号", "", "夜間当番手当", "定常外業務対応手当", "支給過不足調整",
           "非課税通勤費", "立替金（顧客請求分）", "立替金", "その他", "その他手当", "現物支給"]
    # 川口: 夜間当番30000 + 立替金(非課税精算)12177
    rows, _, _ = build_import_rows(
        {"2026013": [30000.0, 0, 0, 0, 0, 0, 12177.0]}, {"2026013": "川口 祐ノ輔"})
    line = render_import_csv(rows, tpl).decode("cp932").strip().split("\r\n")[1].split(",")
    assert line[0] == '"2026013"'
    assert line[1] == ""          # 氏名列は無い（空スキップ）
    assert line[2] == "30000"     # 位置3=夜間当番手当 ✓
    assert line[7] == "12177"     # 位置8=立替金 ✓（旧37047では欠落し計上漏れだった）
    # 立替金列があるので計上漏れ警告は出ない（テレワーク手当=0なので警告なし）
    assert check_template_coverage(rows, tpl) == []


def test_parse_manual_allowances_formats():
    """手入力欄のパース。Excelの2列コピペ（タブ区切り）もそのまま通す。"""
    from services.keihi_payroll_import import parse_manual_allowances
    got, errs = parse_manual_allowances(
        "社員番号,金額\n"          # ヘッダー行 → 無視
        "2026012,15000\n"
        "2024050\t8,000\n"          # タブ区切り＋カンマ入り金額（Excelコピペ）
        "２０２６０１３，１０００\n"  # 全角
        "# コメント\n"
        "\n"
        "2026012,5000\n"            # 同一社員 → 加算
        "2026014,0\n"               # 0 は無視
    )
    assert got == {"2026012": 20000, "2024050": 8000, "2026013": 1000}
    assert errs == []


def test_parse_manual_allowances_errors_and_scope():
    from services.keihi_payroll_import import parse_manual_allowances
    got, errs = parse_manual_allowances(
        "2026012,あああ\n"     # 金額が数値でない
        "5000001,3000\n"       # 給与計算対象外(5始まり)
        "2026013\n"            # 列が足りない
    )
    assert got == {}
    assert len(errs) == 3
    assert any("金額を数値として読めません" in e for e in errs)
    assert any("給与計算対象外" in e for e in errs)


def test_build_import_rows_with_manual_allowances():
    """手入力の定常外/その他手当が行に乗り、手当だけの社員も行が出る（計上漏れ防止）。"""
    rows, warns, _ = build_import_rows(
        {"2026013": [30000.0, 0, 0, 0, 0, 0, 12177.0]},
        {"2026013": "川口 祐ノ輔"},
        roster_names={"2026013": "川口 祐ノ輔", "2024050": "加藤 英人"},
        manual={
            # 2024050 は経費ゼロ＝イレギュラー経費だけ
            "定常外業務対応手当": {"2026013": 15000, "2024050": 8000},
            "その他手当": {"2026013": 5000},
        },
    )
    by_id = {r["社員番号"]: r for r in rows}
    assert set(by_id) == {"2026013", "2024050"}      # 手当だけの社員も行が出る
    assert by_id["2026013"]["定常外業務対応手当"] == 15000
    assert by_id["2026013"]["その他手当"] == 5000
    assert by_id["2026013"]["夜間当番手当"] == 30000  # 既存の経費由来も維持
    # 経費が無く手当だけの社員：手当以外は0、氏名はロスターから引く
    assert by_id["2024050"]["定常外業務対応手当"] == 8000
    assert by_id["2024050"]["氏名"] == "加藤 英人"
    assert by_id["2024050"]["夜間当番手当"] == 0 and by_id["2024050"]["立替金"] == 0
    assert warns == []


def test_build_import_rows_manual_allowance_unknown_id_warns():
    """在籍者一覧に無い社員番号は入力ミスの可能性として警告（行は出す）。"""
    rows, warns, _ = build_import_rows(
        {}, {}, roster_names={"2026013": "川口 祐ノ輔"},
        manual={"定常外業務対応手当": {"2029999": 1000}})
    assert [r["社員番号"] for r in rows] == ["2029999"]
    assert any("2029999" in w and "在籍者一覧にありません" in w for w in warns)


def test_manual_allowances_same_employee_added_twice_is_summed():
    """画面で同じ社員を2回追加したら合算する（1件＝1人分を積み上げる運用）。"""
    from services.keihi_payroll_import import parse_manual_allowances
    got, errs = parse_manual_allowances("2026013\t2000\n2026013\t500\n2026012\t-3000")
    assert errs == []
    assert got == {"2026013": 2500, "2026012": -3000}


def test_load_irregular_file_wide(tmp_path):
    """ワイド形式（社員番号＋項目名の列）を自動判別する。マイナス・桁区切り可。"""
    from services.keihi_payroll_import import load_irregular_file
    p = tmp_path / "irr.csv"
    p.write_bytes(
        "社員番号,氏名,現物支給,支給過不足調整,社保調整\r\n"
        "2026012,橘 伸俊,3000,-5000,0\r\n"          # マイナス・0は無視
        "2024050,加藤 英人,0,\"12,000\",2000\r\n"     # 桁区切り
        .encode("cp932"))
    got, errs = load_irregular_file(p)
    assert errs == []
    assert got == {
        "現物支給": {"2026012": 3000},
        "支給過不足調整": {"2026012": -5000, "2024050": 12000},
        "社保調整": {"2024050": 2000},
    }


def test_load_irregular_file_long(tmp_path):
    """ロング形式（社員番号・項目・金額）も読める。同一社員同一項目は加算。"""
    from services.keihi_payroll_import import load_irregular_file
    p = tmp_path / "irr.csv"
    p.write_bytes(
        "社員番号,項目,金額\r\n"
        "2026013,現物支給,3000\r\n"
        "2026013,社保調整,1500\r\n"
        "2026013,社保調整,500\r\n"       # 加算 → 2000
        .encode("cp932"))
    got, errs = load_irregular_file(p)
    assert errs == []
    assert got == {"現物支給": {"2026013": 3000}, "社保調整": {"2026013": 2000}}


def test_load_irregular_file_long_unknown_item_errors(tmp_path):
    """選択できない項目名はエラーとして返す（黙って捨てない）。"""
    from services.keihi_payroll_import import load_irregular_file
    p = tmp_path / "irr.csv"
    p.write_bytes("社員番号,項目,金額\r\n2026013,通信手当,500\r\n".encode("cp932"))
    got, errs = load_irregular_file(p)
    assert got == {}
    assert any("通信手当" in e for e in errs)


def test_load_irregular_file_xlsx_wide(tmp_path):
    from openpyxl import Workbook
    from services.keihi_payroll_import import load_irregular_file
    wb = Workbook()
    ws = wb.active
    ws.append(["社員番号", "社保調整"])
    ws.append([2026012, -1500])     # 数値セル・マイナス
    ws.append(["2024050", 2000])
    p = tmp_path / "irr.xlsx"
    wb.save(p)
    got, errs = load_irregular_file(p)
    assert errs == []
    assert got == {"社保調整": {"2026012": -1500, "2024050": 2000}}


def test_load_irregular_file_undetectable_header(tmp_path):
    """項目が判別できない見出しは、黙って捨てずにエラーで知らせる。"""
    from services.keihi_payroll_import import load_irregular_file
    p = tmp_path / "irr.csv"
    p.write_bytes("社員番号,金額\r\n2026013,3000\r\n".encode("cp932"))
    got, errs = load_irregular_file(p)
    assert got == {}
    assert any("判別できませんでした" in e for e in errs)


def test_build_import_rows_genbutsu_and_kabusoku():
    """現物支給・支給過不足調整が行に乗る（調整だけの社員も行が出る）。"""
    rows, warns, _ = build_import_rows(
        {"2026013": [30000.0, 0, 0, 0, 0, 0, 12177.0]},
        {"2026013": "川口 祐ノ輔"},
        roster_names={"2026013": "川口 祐ノ輔", "2026012": "橘 伸俊"},
        manual={"現物支給": {"2026013": 3000},
                "支給過不足調整": {"2026013": -5000, "2026012": 8000}},
    )
    by_id = {r["社員番号"]: r for r in rows}
    assert by_id["2026013"]["現物支給"] == 3000
    assert by_id["2026013"]["支給過不足調整"] == -5000      # マイナスがそのまま乗る
    assert by_id["2026013"]["夜間当番手当"] == 30000        # 既存項目は不変
    assert by_id["2026012"]["支給過不足調整"] == 8000       # 調整だけの社員も行が出る
    assert warns == []


def test_shaho_chosei_missing_column_warns():
    """社保調整はテンプレ44450(11列)に列が無い → 計上漏れ警告が出る（追加してもらう）。"""
    from services.keihi_payroll_import import check_template_coverage
    tpl = ["*社員番号", "", "夜間当番手当", "定常外業務対応手当", "支給過不足調整",
           "非課税通勤費", "立替金（顧客請求分）", "立替金", "その他", "その他手当", "現物支給"]
    rows, _, _ = build_import_rows({}, {}, manual={"社保調整": {"2026013": 2000}})
    cov = check_template_coverage(rows, tpl)
    assert any("社保調整" in w and "2,000" in w for w in cov)


def test_render_csv_all_manual_items_on_template_44450():
    """手決め項目が テンプレ44450 の正しい列位置に載る（社保調整列を足した想定）。"""
    tpl = ["*社員番号", "", "夜間当番手当", "定常外業務対応手当", "支給過不足調整",
           "非課税通勤費", "立替金（顧客請求分）", "立替金", "その他", "その他手当", "現物支給",
           "社保調整"]
    rows, _, _ = build_import_rows(
        {"2026013": [30000.0, 0, 0, 0, 0, 0, 12177.0]}, {"2026013": "川口 祐ノ輔"},
        manual={"定常外業務対応手当": {"2026013": 15000},
                "その他手当": {"2026013": 5000},
                "現物支給": {"2026013": 3000},
                "支給過不足調整": {"2026013": -5000},
                "社保調整": {"2026013": 2000}})
    line = render_import_csv(rows, tpl).decode("cp932").strip().split("\r\n")[1].split(",")
    assert line[2] == "30000"    # 位置3 夜間当番手当
    assert line[3] == "15000"    # 位置4 定常外業務対応手当
    assert line[4] == "-5000"    # 位置5 支給過不足調整（マイナス）
    assert line[7] == "12177"    # 位置8 立替金
    assert line[9] == "5000"     # 位置10 その他手当
    assert line[10] == "3000"    # 位置11 現物支給
    assert line[11] == "2000"    # 位置12 社保調整（控除）


def test_match_column_normalizes():
    from services.keihi_payroll_import import _match_column
    assert _match_column("*社員番号") == "社員番号"
    assert _match_column("氏名") == "氏名"
    assert _match_column("立替金（顧客請求分）") == "立替金（顧客請求分）"
    assert _match_column("") is None
    assert _match_column("基本給") is None       # 経費項目でない列はスキップ


def test_read_template_header(tmp_path):
    from services.keihi_payroll_import import read_template_header
    p = tmp_path / "tpl.csv"
    p.write_bytes('"*社員番号","","夜間当番手当"\r\n'.encode("cp932"))
    assert read_template_header(p) == ["*社員番号", "", "夜間当番手当"]
