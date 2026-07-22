import os
import sys
from datetime import date

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.multi_year_shift_parser import (
    is_monthly_shift_xlsx,
    parse_monthly_shift_xlsx,
    parse_structured_files,
)


def _make_monthly_shift_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "V1.00"

    ws.cell(8, 1).value = "氏名"
    ws.cell(8, 3).value = "管理No．"
    for day in range(1, 31):
        ws.cell(8, 4 + day).value = date(2026, 6, day)

    # 日付コピー行。従業員として扱ってはいけない。
    ws.cell(9, 1).value = "◎ｼﾌﾄﾁｰﾌ"
    for day in range(1, 31):
        ws.cell(9, 4 + day).value = ws.cell(8, 4 + day).value

    ws.cell(10, 1).value = "高岩　猛"
    ws.cell(10, 3).value = "JP214222"
    for day in range(1, 31):
        ws.cell(10, 4 + day).value = 2 if day <= 5 else "休"

    ws.cell(11, 1).value = "澁谷　勝徳"
    ws.cell(11, 3).value = "JP208167"
    for day in range(1, 31):
        ws.cell(11, 4 + day).value = 5 if day % 2 else "明"

    ws.cell(17, 1).value = "１勤（在勤） 9：00～18：00"
    ws.cell(18, 1).value = "２勤（昼勤） 9：00～18：00"
    ws.cell(19, 1).value = "５勤（夜勤） 21：00～翌9：30"
    ws.cell(19, 38).value = "有"
    ws.cell(19, 39).value = "有給休暇(7.5H)"
    ws.cell(20, 38).value = "前2"
    ws.cell(20, 39).value = "(2勤)前半休(3H)"

    wb.save(path)


def test_parse_monthly_shift_xlsx_uses_workbook_month_when_target_mismatches(tmp_path):
    path = tmp_path / "monthly.xlsx"
    _make_monthly_shift_book(path)

    assert is_monthly_shift_xlsx(str(path)) is True

    parsed = parse_monthly_shift_xlsx(str(path), 2026, 5)

    assert parsed["year"] == 2026
    assert parsed["month"] == 6
    assert [e["name"] for e in parsed["employees"]] == ["高岩　猛", "澁谷　勝徳"]
    assert parsed["employees"][0]["shifts"][0] == {"date": "2026-06-01", "code": "2"}
    assert parsed["employees"][0]["shifts"][5] == {"date": "2026-06-06", "code": "休"}

    legend_by_code = {entry["code"]: entry for entry in parsed["legend"]}
    assert legend_by_code["2"]["start_time"] == "09:00"
    assert legend_by_code["2"]["end_time"] == "18:00"
    assert legend_by_code["2"]["break_minutes"] == 60
    assert legend_by_code["5"]["start_time"] == "21:00"
    assert legend_by_code["5"]["end_time"] == "09:30"
    assert legend_by_code["有"]["is_off"] is True


def test_parse_structured_files_consumes_monthly_shift_xlsx(tmp_path):
    path = tmp_path / "monthly.xlsx"
    _make_monthly_shift_book(path)

    result = parse_structured_files([str(path)], 2026, 5)

    assert result is not None
    sheets, consumed, warnings = result
    assert consumed == [str(path)]
    assert warnings == []
    assert sheets[0]["year"] == 2026
    assert sheets[0]["month"] == 6
    assert len(sheets[0]["employees"]) == 2
    assert any(entry["code"] == "休" and entry["is_off"] for entry in sheets[0]["legend"])


def test_parse_structured_files_xlsx_needs_target_month(tmp_path):
    """対象年月未入力: xlsx はスキップして warning（黙ってAIに落とさない）"""
    path = tmp_path / "monthly.xlsx"
    _make_monthly_shift_book(path)

    result = parse_structured_files([str(path)], None, None)

    assert result is not None
    sheets, consumed, warnings = result
    assert sheets == []
    assert consumed == []
    assert len(warnings) == 1
    assert "対象年月が未入力" in warnings[0]
    assert "monthly.xlsx" in warnings[0]
