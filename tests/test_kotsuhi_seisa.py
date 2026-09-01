"""承認前の交通費精査。

判定ロジック本体の細かいテストは Z:\\API連携\\tests\\test_kotsuhi_seisa.py にある。
ここでは「承認が進むたびに何度も回す」運用で効く部分だけを固める。
"""
import openpyxl
import pytest

from services.kotsuhi_seisa import (
    CommuteMaster,
    apply_diff,
    build_limit_over_rows,
    build_no_commute_rows,
    build_workdays,
    collect_flagged,
    month_mismatch_warning,
    is_company_employee,
    load_limit_exempt_members,
    read_previous_keys,
    row_key,
)


class _Sheet:
    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, min_row=1, values_only=False):
        return iter(self._rows)


class _Book:
    def __init__(self, summary, detail=None):
        self._s = summary
        self._d = detail
        self.sheetnames = ["サマリ"] + (["テレワーク明細"] if detail is not None else [])

    def __getitem__(self, name):
        return _Sheet(self._s if name == "サマリ" else self._d)


# ----------------------------------------------------------------------
# 繰り返し実行の差分
# ----------------------------------------------------------------------

def test_row_key_distinguishes_by_sheet_granularity():
    # 実費は日ごと、マスタ更新漏れは検知区分ごとに1件と数える
    assert row_key("実費突合", {"社員番号": "2024009", "利用日": "2026/7/1"}) \
        != row_key("実費突合", {"社員番号": "2024009", "利用日": "2026/7/2"})
    assert row_key("マスタ更新漏れ", {"社員番号": "2024009", "検知区分": "M2"}) \
        != row_key("マスタ更新漏れ", {"社員番号": "2024009", "検知区分": "M3"})
    assert row_key("定期代突合", {"社員番号": "2024009"}) == "定期代突合|2024009"


def test_apply_diff_marks_new_and_carried_over():
    prev = {"定期代突合": {"定期代突合|2025020"}}
    rows = [
        {"社員番号": "2025020", "区分": "要確認"},   # 前回もあった
        {"社員番号": "2026003", "区分": "要確認"},   # 今回から
    ]
    resolved = apply_diff("定期代突合", rows, prev)
    assert [r["前回比"] for r in rows] == ["継続", "新規"]
    assert resolved == 0


def test_apply_diff_counts_resolved():
    """申請者が直して要確認から外れたら「解消」として数える。"""
    prev = {"定期代突合": {"定期代突合|2025020", "定期代突合|2026003"}}
    rows = [
        {"社員番号": "2025020", "区分": "OK"},       # 直った
        {"社員番号": "2026003", "区分": "要確認"},   # まだ
    ]
    resolved = apply_diff("定期代突合", rows, prev)
    assert rows[0]["前回比"] == "解消"
    assert rows[1]["前回比"] == "継続"
    assert resolved == 1   # 行は残るが要確認から外れたので解消1件


def test_apply_diff_counts_rows_that_disappeared():
    """行ごと消えた（申請が取り下げられた等）ケースも解消として数える。"""
    prev = {"定期代突合": {"定期代突合|2025020"}}
    resolved = apply_diff("定期代突合", [], prev)
    assert resolved == 1


def test_apply_diff_is_noop_on_first_run():
    rows = [{"社員番号": "2025020", "区分": "要確認"}]
    assert apply_diff("定期代突合", rows, {}) == 0
    assert rows[0]["前回比"] == ""


def test_read_previous_keys_ignores_missing_file(tmp_path):
    assert read_previous_keys(tmp_path / "無い.xlsx") == {}


def test_read_previous_keys_picks_up_only_flagged_rows(tmp_path):
    p = tmp_path / "前回.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("定期代突合")
    ws.append(["社員番号", "氏名", "区分"])
    ws.append(["2025020", "戸松 乃天", "要確認"])
    ws.append(["2007001", "菅原 伸", "OK"])
    ws2 = wb.create_sheet("通勤費申請なし")
    # このシートの判定列は「確認要否」（「区分」は支給区分で別物）
    ws2.append(["社員番号", "氏名", "区分", "確認要否"])
    ws2.append(["2024009", "山口 太雅", "通勤費", "要確認"])
    ws2.append(["2007001", "菅原 伸", "通勤定期代", "OK"])
    wb.save(p)

    got = read_previous_keys(p)
    assert got["定期代突合"] == {"定期代突合|2025020"}
    assert got["通勤費申請なし"] == {"通勤費申請なし|2024009"}


def test_read_previous_keys_survives_a_broken_file(tmp_path):
    """前回ファイルが壊れていても実行は止めない（差分が出ないだけ）。"""
    p = tmp_path / "壊れ.xlsx"
    p.write_bytes(b"not an xlsx")
    assert read_previous_keys(p) == {}


# ----------------------------------------------------------------------
# サマリの数式（何度も回すので毎回ここを通る）
# ----------------------------------------------------------------------

def test_build_workdays_counts_from_detail_when_formula_has_no_cached_value():
    """テレワーク日数は COUNTIF の数式。書いたばかりのブックには結果が入っていない。"""
    wb = _Book(
        [("2008002", "伊藤 淳", 20, None, None, "7/1、7/2")],
        [("2008002", "伊藤 淳", "2026-07-%02d" % d, "テレワーク") for d in range(1, 17)],
    )
    got = build_workdays(wb)["2008002"]
    assert got["テレワーク日数"] == 16
    assert got["出社日数"] == 4       # 素直に読むと 20 になる箇所


def test_build_workdays_uses_cached_values_when_present():
    wb = _Book([("2008002", "伊藤 淳", 20, 16, 4, "7/1、7/2")])
    got = build_workdays(wb)["2008002"]
    assert (got["出勤日数"], got["テレワーク日数"], got["出社日数"]) == (20, 16, 4)


@pytest.mark.parametrize("emp,expected", [
    ("2024014", True),
    ("合計", False),      # 通勤費シート末尾の合計行
    ("5000001", False),   # 派遣
    ("3333008", False),   # 勤怠実績があっても給与計算対象外
])
def test_is_company_employee(emp, expected):
    assert is_company_employee(emp) is expected


# ----------------------------------------------------------------------
# 通勤費の上限（月3万円）超過の検出
# ----------------------------------------------------------------------

LIMIT_IDX = {"交通機関": 0, "ステータス": 1, "社員番号": 2, "申請者": 3,
             "所属グループ": 4, "申請書No.": 5, "小計": 6}


def _row(kind, emp, name, amount, no="1", status="承認完了", group="UAL 平和島"):
    return [kind, status, emp, name, group, no, amount]


def test_limit_over_flags_only_people_missing_from_the_exempt_list():
    """許可者はOK、リストに無い人だけ要確認（＝上限の切り忘れ検知）。"""
    details = [
        _row("通勤定期代", "2014013", "柴田 和浩", "35090"),   # 許可なしで超過
        _row("通勤定期代", "2025029", "奥山 昌苗", "34350"),   # 許可あり
        _row("通勤定期代", "2026006", "大村 賢治", "30000"),   # 上限ちょうど＝対象外
    ]
    rows = build_limit_over_rows(details, LIMIT_IDX, {"2025029": "個別許可"}, limit=30000)

    by = {r["社員番号"]: r for r in rows}
    assert set(by) == {"2014013", "2025029"}        # 30,000ちょうどは挙がらない
    assert by["2014013"]["区分"] == "要確認"
    assert by["2014013"]["超過額"] == 5090
    assert by["2014013"]["上限免除"] == ""
    assert by["2025029"]["区分"] == "OK"
    assert by["2025029"]["上限免除"] == "○"
    assert "個別許可" in by["2025029"]["説明"]


def test_limit_over_treats_travel_expense_members_as_no_limit():
    """移動交通費（立替精算）対象者は通勤系で申請されていても上限が掛からない。

    2026-08 の山田大海さん（実費51,067円）がこれで要確認に挙がった。金額からは
    判別できないので、対象者リストを見に行かないと毎月ここで引っかかる。
    """
    details = [
        _row("通勤交通費（実費）", "2025033", "山田 大海", "51067"),
        _row("通勤交通費（実費）", "2018001", "有田 功太郎", "31832"),
    ]
    rows = build_limit_over_rows(details, LIMIT_IDX, {}, limit=30000,
                                 travel_members={"2025033"})

    by = {r["社員番号"]: r for r in rows}
    assert by["2025033"]["区分"] == "OK"
    assert by["2025033"]["上限免除"] == "○"
    assert "移動交通費" in by["2025033"]["説明"]
    assert by["2018001"]["区分"] == "要確認"       # リストに無い人はこれまでどおり


def test_limit_over_sums_split_applications_and_ignores_travel_expense():
    """定期代の分割申請は合算し、移動交通費（上限なし）は混ぜない。"""
    details = [
        _row("通勤定期代", "2021020", "稲場 直哉", "16000", no="1"),
        _row("通勤交通費（実費）", "2021020", "稲場 直哉", "16070", no="2"),
        # 移動交通費は何円あっても上限判定に入れない
        _row("交通費（電車・バス）", "2021020", "稲場 直哉", "99999", no="3"),
        _row("交通費（電車・バス）", "2019048", "阿部 涼平", "80000", no="4"),
    ]
    rows = build_limit_over_rows(details, LIMIT_IDX, {}, limit=30000)

    assert [r["社員番号"] for r in rows] == ["2021020"]
    got = rows[0]
    assert (got["通勤費合計"], got["うち定期代"], got["うち実費"]) == (32070, 16000, 16070)
    assert got["申請書No."] == "1, 2"


def test_limit_over_skips_withdrawn_applications_and_non_employees():
    details = [
        _row("通勤定期代", "2014013", "柴田 和浩", "35090", status="取下げ"),
        _row("通勤定期代", "5000001", "派遣 太郎", "40000"),
    ]
    assert build_limit_over_rows(details, LIMIT_IDX, {}, limit=30000) == []


def test_load_limit_exempt_members_reads_number_and_reason(tmp_path):
    p = tmp_path / "通勤費_上限免除者.csv"
    p.write_text("社員番号,氏名,理由\n2025017,杉原 司,個別許可\n\n", encoding="utf-8-sig")
    assert load_limit_exempt_members(p) == {"2025017": "個別許可"}


def test_load_limit_exempt_members_returns_empty_when_missing(tmp_path):
    """リストが無くても精査は止めない（超過者が全員 要確認 に出るので気づける）。"""
    assert load_limit_exempt_members(tmp_path / "ない.csv") == {}


# ----------------------------------------------------------------------
# 通勤費申請なしリストの抽出条件
# （2026-08-12 拡張: 実費申請日数＋テレワーク日数＝出勤日数 なら通過）
# ----------------------------------------------------------------------

NC_IDX = {"ステータス": 0, "交通機関": 1, "社員番号": 2, "利用日": 3}


def _nc_row(kind, emp, date, status="承認完了"):
    return [status, kind, emp, date]


def _nc_workdays(emp="2020001", name="山田 太郎", work=20, tw=0):
    return {emp: {"氏名": name, "出勤日数": work, "テレワーク日数": tw,
                  "出社日数": work - tw, "テレワーク実施日": set()}}


def _nc_master(*legs):
    """(社員番号, 経路No, 出発, 到着, 交通機関, 支給間隔, 支給金額) から通勤費シートを作る。"""
    rows = [("社員番号",) + ("",) * 14]
    for emp, no, dep, arr, kind, interval, amount in legs:
        rows.append((emp, "氏名", no, dep, arr, "", "", "", kind, interval, "", amount, "", "", ""))
    return CommuteMaster(_Sheet(rows))


def _no_commute(details, workdays, master=None, target_ids=(), excluded=None):
    return build_no_commute_rows(details, NC_IDX, master or _nc_master(),
                                 workdays, set(target_ids), excluded or {})


def test_no_commute_excludes_people_who_applied_for_actual_cost_without_telework():
    """テレワーク0の人は従来どおり、実費申請が1件でもあれば対象外。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", "2026/7/1")]
    assert _no_commute(details, _nc_workdays(work=20, tw=0)) == []


def test_no_commute_excludes_people_who_applied_for_a_commuter_pass():
    """定期代の申請がある人は定期代突合シートの担当なので出さない（テレワーク有無を問わず）。"""
    details = [_nc_row("通勤定期代", "2020001", "2026/7/1")]
    assert _no_commute(details, _nc_workdays(work=20, tw=5)) == []


def test_no_commute_keeps_zero_attendance_people():
    """出勤0でも落とさない（代表取締役・打刻申請なしの人を取りこぼさないため）。"""
    master = _nc_master(("2020001", 1, "自宅", "本社", "公共交通機関", "毎月", 18570))
    got = _no_commute([], _nc_workdays(work=0, tw=0), master)
    assert [(r["判定"], r["確認要否"]) for r in got] == [("マスタから支給", "情報")]


def test_no_commute_flags_people_with_no_master_and_no_application():
    master = _nc_master()
    got = _no_commute([], _nc_workdays(work=20, tw=0), master)
    assert got[0]["判定"] == "支給漏れの疑い"
    assert got[0]["確認要否"] == "要確認"


def test_no_commute_passes_when_actual_days_plus_telework_equals_workdays():
    """実費申請日とテレワーク日で出勤日が全部埋まる人はリストから外す（拡張の本体）。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", f"2026/7/{d}") for d in range(1, 16)]
    assert _no_commute(details, _nc_workdays(work=20, tw=5)) == []


def test_no_commute_counts_round_trip_rows_of_the_same_day_once():
    """往復で2行に分かれていても同じ利用日なら1日と数える。"""
    details = []
    for d in range(1, 16):
        details += [_nc_row("通勤交通費（実費）", "2020001", f"2026/7/{d}")] * 2
    assert _no_commute(details, _nc_workdays(work=20, tw=5)) == []


def test_no_commute_passes_when_workdays_is_a_float():
    """出勤日数が数式由来の float でも式判定が成立する。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", f"2026/7/{d}") for d in range(1, 16)]
    assert _no_commute(details, _nc_workdays(work=20.0, tw=5.0)) == []


def test_no_commute_surfaces_partial_telework_people_who_never_applied():
    """テレワークありでも実費申請ゼロなら、これまで見えなかった支給漏れとして出す。"""
    master = _nc_master()
    got = _no_commute([], _nc_workdays(work=20, tw=5), master)
    assert got[0]["判定"] == "支給漏れの疑い"
    assert "テレワーク5日" in got[0]["説明"]


def test_no_commute_shows_telework_context_for_master_paid_people():
    master = _nc_master(("2020001", 1, "自宅", "本社", "公共交通機関", "毎月", 18570))
    got = _no_commute([], _nc_workdays(work=20, tw=5), master)
    assert got[0]["判定"] == "マスタから支給"
    assert got[0]["区分"] == "通勤定期代"
    assert "テレワーク5日" in got[0]["説明"]


def test_no_commute_flags_actual_days_over_the_equation():
    """実費申請＋テレワークが出勤日数を超える＝テレワーク日にも申請している疑い。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", f"2026/7/{d}") for d in range(1, 19)]
    got = _no_commute(details, _nc_workdays(work=20, tw=5))
    assert got[0]["判定"] == "実費申請の日数不一致"
    assert got[0]["確認要否"] == "要確認"
    assert "出勤20日・テレワーク5日・実費申請18日" in got[0]["説明"]
    assert "重なっている疑い" in got[0]["説明"]


def test_no_commute_flags_actual_days_under_the_equation():
    """出社日の一部しか申請が無い＝申請漏れの可能性。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", f"2026/7/{d}") for d in range(1, 11)]
    got = _no_commute(details, _nc_workdays(work=20, tw=5))
    assert got[0]["判定"] == "実費申請の日数不一致"
    assert "申請漏れの可能性" in got[0]["説明"]


def test_no_commute_ignores_withdrawn_applications():
    """取下げ・否認の申請は申請したことにしない（式にも申請済み判定にも入れない）。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", f"2026/7/{d}", status="取下げ")
               for d in range(1, 16)]
    got = _no_commute(details, _nc_workdays(work=20, tw=5))
    assert got[0]["判定"] == "支給漏れの疑い"   # 申請ゼロ扱い


def test_no_commute_sorts_mismatch_between_missing_and_no_attendance():
    details = [_nc_row("通勤交通費（実費）", "2020002", "2026/7/1")]
    workdays = {
        "2020001": {"氏名": "支給漏れ", "出勤日数": 20, "テレワーク日数": 0,
                    "出社日数": 20, "テレワーク実施日": set()},
        "2020002": {"氏名": "日数不一致", "出勤日数": 20, "テレワーク日数": 5,
                    "出社日数": 15, "テレワーク実施日": set()},
        "2020003": {"氏名": "勤怠なし", "出勤日数": 0, "テレワーク日数": 0,
                    "出社日数": 0, "テレワーク実施日": set()},
        "2020004": {"氏名": "マスタ支給", "出勤日数": 20, "テレワーク日数": 0,
                    "出社日数": 20, "テレワーク実施日": set()},
    }
    master = _nc_master(("2020004", 1, "自宅", "本社", "公共交通機関", "毎月", 18570))
    got = _no_commute(details, workdays, master)
    assert [r["判定"] for r in got] == [
        "支給漏れの疑い", "実費申請の日数不一致", "勤怠実績なし", "マスタから支給"]


def test_no_commute_diff_marks_the_new_judgment_rows():
    """新判定の行も従来どおり前回比が付く（確認要否ベースなので追加改修が要らない）。"""
    details = [_nc_row("通勤交通費（実費）", "2020001", "2026/7/1")]
    rows = _no_commute(details, _nc_workdays(work=20, tw=5))
    apply_diff("通勤費申請なし", rows, {"通勤費申請なし": {"通勤費申請なし|2020001"}}, "確認要否")
    assert rows[0]["前回比"] == "継続"


# ----------------------------------------------------------------------
# 入力ブックの取り違え検知（2026-08-12 谷津さんが精査結果ブックを指定してハマった）
# ----------------------------------------------------------------------

def _csv_for_inputs(tmp_path):
    import csv as _csv
    p = tmp_path / "交通費申請.csv"
    with open(p, "w", encoding="cp932", newline="") as f:
        w = _csv.writer(f)
        w.writerow(["ステータス", "交通機関", "社員番号", "申請者", "所属グループ", "申請書No.",
                    "明細No.", "利用日", "金額", "往復", "小計", "乗車場所", "降車場所", "経路", "目的地"])
    return p


def test_load_seisa_inputs_explains_a_book_without_required_sheets(tmp_path):
    """『Worksheet 通勤費 does not exist.』ではなく、何を指定すべきかを日本語で伝える。"""
    import openpyxl as _op
    import pytest as _pytest
    from services.kotsuhi_seisa import load_seisa_inputs
    wb = _op.Workbook()
    wb.active.title = "無関係"
    book = tmp_path / "別物.xlsx"
    wb.save(book)
    with _pytest.raises(ValueError) as e:
        load_seisa_inputs(_csv_for_inputs(tmp_path), book, "2026-07")
    msg = str(e.value)
    assert "『通勤費』『サマリ』シートがありません" in msg
    assert "集計を実行" in msg


def test_load_seisa_inputs_names_the_review_output_book(tmp_path):
    """精査結果ブック（このモードの出力）を指定した実際の間違いは名指しで案内する。"""
    import openpyxl as _op
    import pytest as _pytest
    from services.kotsuhi_seisa import load_seisa_inputs
    wb = _op.Workbook()
    wb.active.title = "サマリ"
    wb.create_sheet("通勤費申請なし")     # 精査結果ブックにだけあるシート
    book = tmp_path / "交通費精査結果_2026年7月.xlsx"
    wb.save(book)
    with _pytest.raises(ValueError) as e:
        load_seisa_inputs(_csv_for_inputs(tmp_path), book, "2026-07")
    msg = str(e.value)
    assert "『通勤費』シートがありません" in msg
    assert "精査結果のブック" in msg


# ----------------------------------------------------------------------
# 備考の転記（2026-08-19 追加。人が精査で見る3点セット＝経路・金額・備考の欠けを塞ぐ）
# ----------------------------------------------------------------------

from services.kotsuhi_seisa import (  # noqa: E402
    biko_indexes,
    build_actual_rows,
    build_pass_rows,
    build_travel_rows,
)

# 実CSVは73列だがテストは要る列だけの短い行で組む（列位置はこの辞書が定義）
BIKO_IDX = {"交通機関": 0, "ステータス": 1, "社員番号": 2, "申請者": 3, "所属グループ": 4,
            "申請書No.": 5, "小計": 6, "乗車場所": 7, "降車場所": 8, "利用日": 9,
            "金額": 10, "往復": 11, "目的地": 12, "備考(申請)": 13, "備考": 14}


def _biko_row(kind, emp, name, amount, frm, to, date, biko="", app_biko="",
              no="1", status="承認完了"):
    return [kind, status, emp, name, "G", no, amount, frm, to, date,
            amount, "往復", "", app_biko, biko]


def _biko_master(rows):
    """CommuteMaster用のモック。r[0]=社員番号…r[9]=支給間隔, r[11]=支給金額の位置だけ合わせる。"""
    def leg(emp, name, frm, to, interval, amount):
        return [emp, name, 1, frm, to, "", "", f"{frm}→{to}", "電車", interval,
                "", amount, "", "", "2026/4/1"]
    return CommuteMaster(_Sheet([leg(*r) for r in rows]))


def test_biko_indexes_maps_detail_and_application_columns():
    # 2列ある形式は後ろ＝明細側。1列だけの形式はそれを明細側として使う
    assert biko_indexes(["備考", "金額", "備考"]) == {"備考": 2, "備考(申請)": 0}
    assert biko_indexes(["金額", "備考"]) == {"備考": 1}
    assert biko_indexes(["金額"]) == {}


def test_pass_rows_collect_biko_from_both_columns_and_dedupe():
    master = _biko_master([("2024009", "田中", "品川", "大崎", "毎月", 8000)])
    details = [
        _biko_row("通勤定期代", "2024009", "田中", "4000", "品川", "大崎", "2026/7/1",
                  biko="上期分", app_biko="7月分の定期代です"),
        # 同じ申請の2明細目: 申請側備考は同文が繰り返されるので1回に畳まれること
        _biko_row("通勤定期代", "2024009", "田中", "4000", "品川", "大崎", "2026/7/1",
                  app_biko="7月分の定期代です"),
    ]
    rows = build_pass_rows(details, BIKO_IDX, master)
    assert rows[0]["備考"] == "上期分／7月分の定期代です"

    # 備考列が無い旧形式のCSVでも動く（後方互換）
    idx_old = {k: v for k, v in BIKO_IDX.items() if not k.startswith("備考")}
    assert build_pass_rows(details, idx_old, master)[0]["備考"] == ""


def test_actual_rows_carry_biko_on_daily_and_monthly_rows():
    master = _biko_master([
        ("2024001", "日次", "品川", "大崎", "毎日", 500),
        ("2024002", "月次", "綾瀬", "東銀座", "毎月", 8000),
    ])
    details = [
        _biko_row("通勤交通費（実費）", "2024001", "日次", "500", "品川", "大崎",
                  "2026/7/1", biko="オンサイト※自宅から直行"),
        # マスタが定期登録の人の実費申請 → 月合計行に集約されても備考が残ること
        _biko_row("通勤交通費（実費）", "2024002", "月次", "8000", "綾瀬", "東銀座",
                  "2026/7/1", biko="区分を間違えました"),
    ]
    rows = build_actual_rows(details, BIKO_IDX, master, {})
    by = {r["社員番号"]: r for r in rows}
    assert by["2024001"]["備考"] == "オンサイト※自宅から直行"
    assert by["2024002"]["備考"] == "区分を間違えました"


def test_travel_detail_rows_carry_biko():
    # 従来は備考をパースしても列に出しておらず常に空だった（2026-08-19 修正）
    details = [_biko_row("交通費（電車・バス）", "2024009", "田中", "480", "品川", "泉岳寺",
                         "2026/7/2", biko="客先往訪", app_biko="研修のため")]
    _summary, detail_rows = build_travel_rows(details, BIKO_IDX, set())
    assert detail_rows[0]["備考"] == "客先往訪／研修のため"


# ----------------------------------------------------------------------
# 要確認カード（数えた件数と、画面に出す行を一致させる）
# ----------------------------------------------------------------------

class _MultiBook:
    """シート名 -> 行リスト（1行目が見出し）のフェイクブック。"""

    def __init__(self, sheets):
        self._sheets = sheets
        self.sheetnames = list(sheets)

    def __getitem__(self, name):
        return _Sheet(self._sheets[name])


def _flagged_book():
    return _MultiBook({
        # 判定列は確認要否。同じシートの「区分」は交通機関の種別で、意味が違う
        "通勤費申請なし": [
            ("社員番号", "氏名", "支給金額", "区分", "判定", "確認要否", "備考"),
            ("2018003", "梅本 剛史", 0, "通勤定期代", "支給漏れの疑い", "要確認", None),
            ("2019004", "佐藤 花子", 12000, "通勤定期代", "問題なし", "OK", None),
        ],
        "マスタ更新漏れ": [
            ("検知区分", "社員番号", "氏名", "内容", "関係金額", "区分"),
            ("M1", "2026015", "木村 悠樹", "マスタ未登録", 3600, "要確認"),
        ],
        "定期代突合": [
            ("社員番号", "氏名", "差額", "判定", "区分"),
            ("2024014", "千代田 昭広", 17590, "D", "要確認"),
        ],
        # 判定列が無いシートは件数にもカードにも入れない
        "移動交通費": [("社員番号", "氏名"), ("2020001", "山田 太郎")],
    })


def test_collect_flagged_shows_every_sheet_it_counts():
    # サマリの「要確認」の数字とカードの枚数がずれないことがこの機能の肝。
    # 2026-09-01 に、13件と出ているのにカードが0枚という画面になった。
    flagged, rows = collect_flagged(_flagged_book())
    assert flagged == {"通勤費申請なし": 1, "マスタ更新漏れ": 1, "定期代突合": 1}
    assert sum(flagged.values()) == len(rows) == 3
    assert [r["シート"] for r in rows] == ["通勤費申請なし", "マスタ更新漏れ", "定期代突合"]


def test_collect_flagged_drops_judgement_column_but_keeps_same_named_kubun():
    rows = collect_flagged(_flagged_book())[1]
    no_apply, gap = rows[0], rows[1]
    # 確認要否は全行「要確認」なので落とす。名前が同じでも意味の違う区分は残す
    assert "確認要否" not in no_apply
    assert no_apply["区分"] == "通勤定期代"
    assert no_apply["判定"] == "支給漏れの疑い"
    assert no_apply["支給金額"] == 0        # 0円は「値なし」ではないので残す
    assert "備考" not in no_apply           # 空セルは画面に出さない
    # マスタ更新漏れ側では区分が判定列なので落ちる
    assert "区分" not in gap
    assert gap["検知区分"] == "M1"


def test_collect_flagged_skips_sheets_without_judgement_column():
    flagged, rows = collect_flagged(_flagged_book())
    assert "移動交通費" not in flagged
    assert all(r["シート"] != "移動交通費" for r in rows)


# ----------------------------------------------------------------------
# 対象月の取り違え検知
# ----------------------------------------------------------------------

def test_warns_when_every_row_was_dropped_as_out_of_month():
    # 2026-09-01、対象月を 2026-09 で回して 2026-08 の600行が全部捨てられたのに
    # 「進行中の申請なし＝この月は見終わりです」と出て、成功に読めた。
    msg = month_mismatch_warning("2026-09", 0, 600)
    assert msg is not None
    assert "2026-09" in msg and "600行" in msg
    assert "対象月が違いませんか" in msg
    # 要確認が0件でないことを「精査できた」と誤読させない一言を必ず添える
    assert "金額・経路の突合はできていません" in msg


def test_no_warning_when_the_month_actually_has_applications():
    assert month_mismatch_warning("2026-08", 569, 31) is None


def test_no_warning_when_there_is_simply_nothing_to_exclude():
    # 申請が本当に0件の月（除外も0）は取り違えではないので黙る
    assert month_mismatch_warning("2026-08", 0, 0) is None


# ----------------------------------------------------------------------
# 移動交通費: 見るのは「通勤費でも申請していないか」だけ
# ----------------------------------------------------------------------

def _travel(emp, name, date, amount="480"):
    return _biko_row("交通費（電車・バス）", emp, name, amount, "品川", "泉岳寺", date)


def _actual(emp, name, date, amount="500"):
    return _biko_row("通勤交通費（実費）", emp, name, amount, "品川", "大崎", date)


def test_travel_rows_flag_roster_members_who_also_claim_commuting_cost():
    """立替精算で計上すべき人が通勤費側にも出していないか、それだけを見る。

    2026-09-01 以前は対象者リストに載っていないだけで要確認にしていたが、
    リストへの登録漏れは経費の誤りではないので判定から外した。
    """
    details = [
        _travel("2024001", "移動のみ", "2026/8/5"),
        _travel("2024002", "二重申請", "2026/8/5"),
        _actual("2024002", "二重申請", "2026/8/6"),
        _actual("2024002", "二重申請", "2026/8/7"),
    ]
    rows, _ = build_travel_rows(details, BIKO_IDX, {"2024001", "2024002"})
    by = {r["社員番号"]: r for r in rows}
    assert by["2024001"]["区分"] == "OK"
    assert by["2024001"]["説明"] == "通勤費の申請なし"
    assert by["2024002"]["区分"] == "要確認"
    assert "通勤交通費（実費） 2日" in by["2024002"]["説明"]


def test_travel_rows_do_not_flag_people_outside_the_roster():
    """リスト外の人は常駐先へ毎日通いながら時々出張する形が普通。

    ここを要確認にすると、2026年8月の実データで11人が19人に膨らみ、
    見なくていい人が8人混ざる（2026-09-01 谷津さん判断で案Bを採用）。
    """
    details = [
        _travel("2024007", "常駐＋出張", "2026/8/5"),
        _actual("2024007", "常駐＋出張", "2026/8/6"),
    ]
    rows, _ = build_travel_rows(details, BIKO_IDX, set())
    assert rows[0]["対象者リスト"] == "リスト外"
    assert rows[0]["区分"] == "OK"
    # 判定はしないが、通勤費を出している事実は参考として残す
    assert "判定対象外" in rows[0]["説明"]
    assert "通勤交通費（実費） 1日" in rows[0]["説明"]


def test_travel_rows_stay_ok_for_roster_gaps_alone():
    details = [_travel("2024004", "リスト外だけ", "2026/8/5")]
    rows, _ = build_travel_rows(details, BIKO_IDX, set())
    assert rows[0]["対象者リスト"] == "リスト外"
    assert rows[0]["区分"] == "OK"
    assert rows[0]["説明"] == "対象者リスト外のため判定対象外"


def test_travel_rows_count_pass_applications_by_case_not_by_day():
    details = [
        _travel("2024005", "定期も出す", "2026/8/5"),
        _biko_row("通勤定期代", "2024005", "定期も出す", "8000", "綾瀬", "東銀座", "2026/8/1"),
    ]
    rows, _ = build_travel_rows(details, BIKO_IDX, {"2024005"})
    assert rows[0]["区分"] == "要確認"
    assert "通勤定期代 1件" in rows[0]["説明"]


def test_travel_rows_ignore_withdrawn_commuting_applications():
    # 取下げ・否認は精査対象外なので、二重申請にも数えない
    details = [
        _travel("2024006", "取下げ済み", "2026/8/5"),
        _biko_row("通勤交通費（実費）", "2024006", "取下げ済み", "500", "品川", "大崎",
                  "2026/8/6", status="取下げ"),
    ]
    rows, _ = build_travel_rows(details, BIKO_IDX, {"2024006"})
    assert rows[0]["区分"] == "OK"
    assert rows[0]["説明"] == "通勤費の申請なし"


# ----------------------------------------------------------------------
# 月まとめ行の見出しに対象月を入れる
# ----------------------------------------------------------------------

def test_monthly_rollup_label_uses_the_target_month():
    # 2026-09-01 まで "7月" が固定で埋め込まれており、8月の結果に7月と出ていた
    master = _biko_master([("2024002", "月次", "綾瀬", "東銀座", "毎月", 8000)])
    details = [_biko_row("通勤交通費（実費）", "2024002", "月次", "8000", "綾瀬",
                         "東銀座", "2026/8/1")]
    rows = build_actual_rows(details, BIKO_IDX, master, {}, "2026-08")
    assert rows[0]["利用日"] == "(8月合計 1日分)"


def test_monthly_rollup_label_falls_back_when_no_month_given():
    master = _biko_master([("2024002", "月次", "綾瀬", "東銀座", "毎月", 8000)])
    details = [_biko_row("通勤交通費（実費）", "2024002", "月次", "8000", "綾瀬",
                         "東銀座", "2026/8/1")]
    rows = build_actual_rows(details, BIKO_IDX, master, {})
    assert rows[0]["利用日"] == "(当月合計 1日分)"
