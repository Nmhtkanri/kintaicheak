# -*- coding: utf-8 -*-
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.expense_check import (
    summarize, build_telework_workbook, _md,
    read_commute_csv, add_commute_sheet, add_selected_employee_views,
    COMMUTE_OUTPUT_COLUMNS, _build_telework_sheets, EMP_PICK_NAME,
    fetch_active_employees, load_travel_expense_members, _norm_date_str,
    validate_travel_expense_rows, save_travel_expense_members,
)


def test_summarize_dedups_and_filters():
    att = [
        {"date": "2026-05-01", "is_absent": False, "attended_at": "x",
         "stamp_classifications": [{"name": "テレワーク（4時間以上）"}]},
        # 同一日の重複レコード → 1日に集約
        {"date": "2026-05-01", "is_absent": False, "attended_at": "x",
         "stamp_classifications": [{"name": "テレワーク（4時間以上）"}]},
        # 出社（テレワーク区分なし）
        {"date": "2026-05-02", "is_absent": False, "attended_at": "x", "stamp_classifications": []},
        # 欠勤 → 除外
        {"date": "2026-05-03", "is_absent": True, "attended_at": None},
        # 打刻なし → 除外
        {"date": "2026-05-04", "is_absent": False, "attended_at": None},
    ]
    work_days, telework = summarize(att)
    assert work_days == 2                      # 5/1, 5/2 のみ
    assert telework == [("2026-05-01", "テレワーク（4時間以上）")]


def test_summarize_multiple_telework_kinds_joined():
    att = [
        {"date": "2026-05-10", "is_absent": False, "attended_at": "x",
         "stamp_classifications": [{"name": "テレワーク（午前）"}, {"name": "テレワーク（午後）"}]},
    ]
    work_days, telework = summarize(att)
    assert work_days == 1
    assert telework[0][0] == "2026-05-10"
    assert "テレワーク（午前）" in telework[0][1] and "テレワーク（午後）" in telework[0][1]


def test_summarize_empty():
    assert summarize([]) == (0, [])


def test_md_formats():
    assert _md("2026-05-01") == "5/1"
    assert _md("2026/5/9") == "5/9"


def test_norm_date_str():
    assert _norm_date_str("2026-07-31") == "2026-07-31"
    assert _norm_date_str("2026/7/1") == "2026-07-01"
    assert _norm_date_str("") == ""
    assert _norm_date_str(None) == ""
    assert _norm_date_str("退職") == ""


# ----------------------------------------------------------------------
# 従業員取得（前月退職者を含める。2026-08-03 谷津さん依頼）
# ----------------------------------------------------------------------

def _emp(emp_id, last, first, cls_id, retired_on="", joined_on="2020-04-01"):
    return {"id": emp_id, "company": {
        "last_name": last, "first_name": first,
        "enrollment_classification": {"id": cls_id, "name": {0: "在籍", 1: "退職", 2: "休職"}[cls_id]},
        "retirement_date": retired_on,
        "joined_on": joined_on,
    }}


class _FakeEmployeeClient:
    """get_employees の only_active を実APIと同じ意味で再現するダミー。"""
    def __init__(self, employees):
        self._employees = employees
        self.calls = []

    def get_employees(self, only_active=True):
        self.calls.append(only_active)
        if only_active:
            return [e for e in self._employees
                    if (e["company"]["enrollment_classification"]["id"] == 0)]
        return list(self._employees)


_EMPLOYEES = [
    _emp("2020001", "田中", "一郎", 0),                       # 在籍
    _emp("2025001", "森田", "恭介", 0, "2026-08-31"),          # 退職日登録済みだがまだ在籍
    _emp("2023019", "小池", "裕也", 1, "2026-07-31"),          # 対象月末の退職者
    _emp("2025022", "池村", "重里", 1, "2026-07-15"),          # 対象月中の退職者
    _emp("2016012", "守屋", "圭祐", 1, "2026-06-30"),          # 前月より前の退職者
    _emp("2020021", "二神", "啓城", 2),                        # 休職 → 従来どおり対象外
    _emp("2010001", "昔の", "退職者", 1, ""),                   # 退職日なし → 対象外
    _emp("2026016", "髙垣", "和希", 0, "", "2026-08-01"),       # 翌月入社 → 対象外
    _emp("2026019", "一戸", "仁美", 0, "", "2026-07-31"),       # 対象月末入社 → 対象
    _emp("2026021", "矢野", "淳大", 0, "", ""),                 # 入社日が読めない → 絞り込まない
]


def test_fetch_active_employees_default_only_active():
    client = _FakeEmployeeClient(_EMPLOYEES)
    got = fetch_active_employees(client)
    assert [e["id"] for e in got] == ["2020001", "2025001", "2026016", "2026019", "2026021"]
    assert client.calls == [True]      # 従来どおり在籍者のみの取得


def test_fetch_active_employees_includes_recent_retirees():
    from datetime import date
    client = _FakeEmployeeClient(_EMPLOYEES)
    got = fetch_active_employees(client, include_retired_since=date(2026, 7, 1))
    ids = [e["id"] for e in got]
    assert "2023019" in ids            # 7/31 退職 → 対象
    assert "2025022" in ids            # 7/15 退職 → 対象
    assert "2016012" not in ids        # 6/30 退職 → 対象外
    assert "2020021" not in ids        # 休職は含めない
    by = {e["id"]: e["name"] for e in got}
    assert by["2023019"] == "小池 裕也"


def test_fetch_active_employees_excludes_next_month_hires():
    """8月入社の人が7月分に載ると、7月実績を計上する8月給与に通勤費が混ざる。"""
    from datetime import date
    client = _FakeEmployeeClient(_EMPLOYEES)
    got = fetch_active_employees(
        client, include_retired_since=date(2026, 7, 1), joined_on_or_before=date(2026, 7, 31)
    )
    ids = [e["id"] for e in got]
    assert "2026016" not in ids        # 8/1 入社 → 対象外
    assert "2026019" in ids            # 7/31 入社（対象月末）→ 対象
    assert "2026021" in ids            # 入社日が読めない → 絞り込まず残す
    assert "2023019" in ids            # 退職側の判定は従来どおり効く
    assert "2016012" not in ids


def test_fetch_active_employees_hire_filter_is_opt_in():
    """joined_on_or_before を渡さなければ従来どおり入社日で絞らない。"""
    from datetime import date
    client = _FakeEmployeeClient(_EMPLOYEES)
    got = fetch_active_employees(client, include_retired_since=date(2026, 7, 1))
    assert "2026016" in [e["id"] for e in got]


# ----------------------------------------------------------------------
# 移動交通費（立替精算）対象者リスト
# ----------------------------------------------------------------------

def test_load_travel_expense_members_utf8_bom(tmp_path):
    p = tmp_path / "travel.csv"
    p.write_text("社員番号,氏名\n2018017,中村 淳一\n2026001,佐久間歩\n2018017,重複\n\n",
                 encoding="utf-8-sig")
    got = load_travel_expense_members(p)
    assert got == {"2018017": "中村 淳一", "2026001": "佐久間歩"}


def test_load_travel_expense_members_cp932(tmp_path):
    import csv
    p = tmp_path / "travel_sjis.csv"
    with open(p, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(["社員番号", "氏名"])
        w.writerow(["2020008", "佐藤 清"])
    assert load_travel_expense_members(p) == {"2020008": "佐藤 清"}


def test_load_travel_expense_members_missing_file(tmp_path):
    assert load_travel_expense_members(tmp_path / "nai.csv") == {}


def test_validate_travel_expense_rows():
    rows = [
        {"id": " 2018017 ", "name": " 中村 淳一 "},   # 前後空白は吸収
        {"id": 2026013.0, "name": "川口"},            # Excel数値化
        {"id": "", "name": ""},                       # 空行はスキップ
        {"id": "5000001", "name": "派遣さん"},        # 形式外 → 注意
        {"id": "", "name": "番号なし"},               # エラー
        {"id": "20x8017", "name": "非数字"},          # エラー
        {"id": "2018017", "name": "重複"},            # エラー
    ]
    normalized, errors, warnings = validate_travel_expense_rows(rows)
    assert [r["id"] for r in normalized] == ["2018017", "2026013", "5000001"]
    assert normalized[0]["name"] == "中村 淳一"
    assert len(errors) == 3
    assert any("番号なし" in e for e in errors)
    assert any("非数字" in e or "20x8017" in e for e in errors)
    assert any("重複" in e for e in errors)
    assert warnings == ["社員番号 5000001 は自社形式（20YY＋3桁）ではありません"]


def test_save_travel_expense_members_roundtrip_and_backup(tmp_path):
    p = tmp_path / "travel.csv"
    r1 = save_travel_expense_members([{"id": "2018017", "name": "中村 淳一"}], p)
    assert r1["backup"] == ""                       # 新規作成はバックアップなし
    assert load_travel_expense_members(p) == {"2018017": "中村 淳一"}

    r2 = save_travel_expense_members(
        [{"id": "2018017", "name": "中村 淳一"}, {"id": "2026001", "name": "佐久間歩"}], p)
    assert r2["count"] == 2
    assert r2["backup"]                             # 2回目は上書き前バックアップあり
    backup_path = tmp_path / "_backup"
    backups = list(backup_path.glob("travel_*.csv"))
    assert len(backups) == 1
    assert load_travel_expense_members(backups[0]) == {"2018017": "中村 淳一"}   # 旧内容
    assert load_travel_expense_members(p) == {"2018017": "中村 淳一", "2026001": "佐久間歩"}
    # UTF-8 BOM で保存されている（谷津さんの Excel でそのまま開ける）
    assert p.read_bytes().startswith(b"\xef\xbb\xbf")


def test_save_travel_expense_members_replace_failure_keeps_original(tmp_path, monkeypatch):
    """置き換えに失敗（Excelで開いている等）しても元ファイルと一時ファイルを汚さない。"""
    import os
    p = tmp_path / "travel.csv"
    save_travel_expense_members([{"id": "2018017", "name": "中村 淳一"}], p)

    def _fail(src, dst):
        raise PermissionError("locked")
    monkeypatch.setattr(os, "replace", _fail)
    try:
        save_travel_expense_members([{"id": "2026001", "name": "佐久間歩"}], p)
        assert False, "PermissionError になるはず"
    except PermissionError:
        pass
    assert load_travel_expense_members(p) == {"2018017": "中村 淳一"}   # 元のまま
    assert not (tmp_path / "travel.csv.tmp").exists()                   # 一時ファイル掃除済み


def test_build_telework_workbook(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "tw.xlsx"
    build_telework_workbook("2026-05", [
        {"id": "2020001", "name": "田中 一郎", "work_days": 20,
         "telework_days": [("2026-05-01", "テレワーク"), ("2026-05-08", "テレワーク")]},
        {"id": "2020002", "name": "上原 奏吾", "work_days": 18, "telework_days": []},
    ], out)
    assert out.exists()
    wb = load_workbook(out)
    # 連動ビュー（テレワーク明細(選択者)）も追加される
    assert set(wb.sheetnames) == {"サマリ", "テレワーク明細", "テレワーク明細(選択者)"}
    assert wb.sheetnames[0] == "サマリ"  # サマリを先頭に表示
    ws = wb["サマリ"]
    # A:F が既存のサマリ見出し（G以降は選択ダッシュボード用なので先頭6列だけ確認）
    assert [ws.cell(row=1, column=c).value for c in range(1, 7)] == \
        ["社員番号", "氏名", "出勤日数", "テレワーク日数", "出社日数", "テレワーク実施日"]
    # 田中: テレワーク実施日は 5/1、5/8 表記
    assert ws.cell(row=2, column=6).value == "5/1、5/8"
    # テレワーク日数は COUNTIF 数式
    assert str(ws.cell(row=2, column=4).value).startswith("=COUNTIF")
    # 明細は田中の2行
    wd = wb["テレワーク明細"]
    assert wd.max_row == 3  # ヘッダー + 2行


def test_selected_employee_views_link_all_sheets(tmp_path):
    """従業員選択ドロップダウン＋連動シートが正しく作られること。"""
    from openpyxl import Workbook, load_workbook
    rows = [
        {"id": "2020001", "name": "田中 一郎", "work_days": 20,
         "telework_days": [("2026-05-01", "テレワーク"), ("2026-05-08", "テレワーク")]},
        {"id": "2020002", "name": "上原 奏吾", "work_days": 18, "telework_days": []},
    ]
    commute = [
        {"社員番号": "2020001", "氏名": "田中 一郎", "経路No": 1, "出発": "笹塚", "到着": "豊洲",
         "経由1": "", "経由2": "", "利用交通機関": "公共交通機関", "通勤経路": "笹塚 → 豊洲",
         "支給間隔": "毎月", "支給方法": "一括", "支給金額": 18350, "非課税通勤費": 18350,
         "課税通勤費": 0, "支給開始": "2025-06", "片道距離(km)": ""},
    ]
    wb = Workbook()
    _build_telework_sheets(wb, "2026-05", rows)
    add_commute_sheet(wb, commute)
    add_selected_employee_views(wb, employee_count=len(rows), commute_row_count=len(commute))
    out = tmp_path / "views.xlsx"
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)

    rb = load_workbook(out)
    # サマリ先頭 → 連動ビュー → 元データ の順
    assert rb.sheetnames[0] == "サマリ"
    assert "テレワーク明細(選択者)" in rb.sheetnames
    assert "通勤費(選択者)" in rb.sheetnames

    # 定義名 選択社員 = サマリ!$I$2、初期値は先頭従業員
    assert EMP_PICK_NAME in rb.defined_names
    assert rb.defined_names[EMP_PICK_NAME].attr_text == "サマリ!$I$2"
    assert rb["サマリ"]["I2"].value == "2020001"

    # 連動シートは INDEX+MATCH で選択社員の該当行を各セルに引く（配列/スピル非依存）
    tw_view = rb["テレワーク明細(選択者)"]
    tw_a4 = tw_view.cell(row=4, column=1).value
    assert tw_a4.startswith("=IFERROR(INDEX(テレワーク明細!A$2")
    assert "MATCH(ROW()-3" in tw_a4
    # 田中は2件のテレワーク → データ行は2行ぶん生成される
    assert tw_view.cell(row=5, column=1).value is not None
    assert tw_view.cell(row=6, column=1).value is None

    # 順位ヘルパーは元シート側（テレワーク明細 E列）に付き、非表示
    src = rb["テレワーク明細"]
    e2 = src.cell(row=2, column=5).value
    assert "COUNTIF" in e2 and EMP_PICK_NAME in e2
    assert src.column_dimensions["E"].hidden is True

    cm_view = rb["通勤費(選択者)"]
    cm_a4 = cm_view.cell(row=4, column=1).value
    assert cm_a4.startswith("=IFERROR(INDEX(通勤費!A$2") and "MATCH(ROW()-3" in cm_a4

    # サマリのミニ集計は XLOOKUP（内部名 _xlfn.XLOOKUP）
    assert str(rb["サマリ"]["I3"].value).startswith("=_xlfn.XLOOKUP(")


def test_selected_views_skips_when_no_employees():
    """従業員0名なら連動ビューは追加しない（例外にならない）。"""
    from openpyxl import Workbook
    wb = Workbook()
    _build_telework_sheets(wb, "2026-05", [])
    add_selected_employee_views(wb, employee_count=0)
    assert "テレワーク明細(選択者)" not in wb.sheetnames
    assert EMP_PICK_NAME not in wb.defined_names


def test_read_commute_csv_maps_fields(tmp_path):
    import csv
    p = tmp_path / "commute.csv"
    headers = ["社員番号", "職場氏名(氏)", "職場氏名(名)", "非課税通勤費(通勤1)", "課税通勤費(通勤1)",
               "支給金額(通勤1)", "利用開始日(通勤1)", "出発(通勤1)", "到着(通勤1)", "経由1(通勤1)",
               "経由2(通勤1)", "利用交通機関(通勤1)", "経路(通勤1)", "片道距離(km)(通勤1)"]
    with open(p, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerow(["2007002", "横山", "弘樹", "18350", "0", "18,350", "", "笹塚", "豊洲", "市ヶ谷",
                    "", "公共交通機関", "笹塚 → 京王線 → 豊洲", "0.00"])
        w.writerow(["2008002", "伊藤", "淳", "0", "0", "1056", "", "新八柱", "潮見", "",
                    "", "公共交通機関", "新八柱 → JR → 潮見", ""])
    rows = read_commute_csv(p)
    assert len(rows) == 2
    r = rows[0]
    assert r["社員番号"] == "2007002"
    assert r["氏名"] == "横山 弘樹"          # 氏＋名を結合
    assert r["出発"] == "笹塚" and r["到着"] == "豊洲"
    assert r["経由1"] == "市ヶ谷" and r["経由2"] == ""
    assert r["通勤経路"].startswith("笹塚")
    assert r["支給金額"] == 18350            # カンマ除去して int 化
    assert r["非課税通勤費"] == 18350


def test_add_commute_sheet(tmp_path):
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    add_commute_sheet(wb, [
        {"社員番号": "2007002", "氏名": "横山 弘樹", "出発": "笹塚", "到着": "豊洲", "経由1": "市ヶ谷",
         "経由2": "", "利用交通機関": "公共交通機関", "通勤経路": "笹塚 → 豊洲", "支給金額": 18350,
         "非課税通勤費": 18350, "課税通勤費": 0, "利用開始日": "", "片道距離(km)": ""},
        {"社員番号": "2008002", "氏名": "伊藤 淳", "出発": "新八柱", "到着": "潮見", "経由1": "",
         "経由2": "", "利用交通機関": "公共交通機関", "通勤経路": "新八柱 → 潮見", "支給金額": 1056,
         "非課税通勤費": 0, "課税通勤費": 0, "利用開始日": "", "片道距離(km)": ""},
    ])
    out = tmp_path / "c.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    assert "通勤費" in wb2.sheetnames
    ws = wb2["通勤費"]
    assert [c.value for c in ws[1]] == COMMUTE_OUTPUT_COLUMNS
    assert ws.cell(row=2, column=1).value == "2007002"
    assert ws.cell(row=2, column=COMMUTE_OUTPUT_COLUMNS.index("出発") + 1).value == "笹塚"
    # 合計行（支給金額の SUM 数式）
    total_row = 2 + 2
    assert str(ws.cell(row=total_row, column=COMMUTE_OUTPUT_COLUMNS.index("支給金額") + 1).value).startswith("=SUM")


def test_fetch_commute_rows_via_api_maps_fields():
    from services.expense_check import fetch_commute_rows_via_api

    class FakeClient:
        def get_commuting_information(self):
            return [
                {"employee_id": "2007002", "commuting": [
                    {"departure": "笹塚", "arrival": "豊洲", "transit_1": "市ヶ谷", "transit_2": "",
                     "path": "笹塚 → 豊洲", "type": {"name": "公共交通機関"}, "one_way_distance": "0.00",
                     "payment": {"start_date": "2025-06", "interval": {"name": "毎月"},
                                 "method": {"name": "一括"}, "total": "18350",
                                 "tax_exemption_amount": "18350", "taxable_amount": "0"}},
                ]},
                {"employee_id": "2009009", "commuting": [
                    {"departure": "A", "arrival": "B", "transit_1": "", "transit_2": "", "path": "A → B",
                     "type": {"name": "車"}, "one_way_distance": "5.0",
                     "payment": {"start_date": "2026-04", "interval": {"name": "6ヶ月"},
                                 "method": {"name": "月割"}, "total": "30000",
                                 "tax_exemption_amount": "30000", "taxable_amount": "0"}},
                    {"departure": "C", "arrival": "D", "transit_1": "", "transit_2": "", "path": "C → D",
                     "type": {"name": "公共交通機関"}, "one_way_distance": "0.00",
                     "payment": {"start_date": "2026-04", "interval": {"name": "毎月"},
                                 "method": {"name": "一括"}, "total": "5000",
                                 "tax_exemption_amount": "5000", "taxable_amount": "0"}},
                ]},
            ]

    rows = fetch_commute_rows_via_api(FakeClient(), {"2007002": "横山 弘樹", "2009009": "田中 花子"})
    assert len(rows) == 3  # 2007002:1経路 + 2009009:2経路
    r0 = rows[0]
    assert r0["社員番号"] == "2007002" and r0["氏名"] == "横山 弘樹" and r0["経路No"] == 1
    assert r0["出発"] == "笹塚" and r0["到着"] == "豊洲" and r0["経由1"] == "市ヶ谷"
    assert r0["支給間隔"] == "毎月" and r0["支給金額"] == 18350
    assert r0["利用交通機関"] == "公共交通機関" and r0["支給開始"] == "2025-06"
    multi = [r for r in rows if r["社員番号"] == "2009009"]
    assert len(multi) == 2 and multi[0]["経路No"] == 1 and multi[1]["経路No"] == 2
    assert multi[0]["支給間隔"] == "6ヶ月" and multi[0]["支給方法"] == "月割"
