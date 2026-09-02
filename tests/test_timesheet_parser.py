import os
import sys
from datetime import date, time, timedelta

import pandas as pd
from openpyxl import Workbook
from PIL import Image, ImageDraw

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import services.timesheet_parser as timesheet_parser
from services.timesheet_parser import _load_file_for_claude, parse_timesheet_smart


def test_parse_itone_dispatch_timesheet_excel_direct(tmp_path):
    path = tmp_path / "中澤寿代さん(ITone)派遣労働者勤務報告書_202604.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "派遣労働者勤務報告書"
    ws["B12"] = "日付"
    ws["U6"] = "中澤　寿代"
    ws["B13"] = "2026/4/1"
    ws["X13"] = "09:45"
    ws["AC13"] = "19:33"
    ws["CO13"] = timedelta(hours=8, minutes=48)
    ws["L13"] = "JANET更改業務：時間外会議"
    ws["B14"] = "2026/4/2"
    ws["X14"] = "09:45"
    ws["AC14"] = "18:21"
    ws["CO14"] = "8:21"
    wb.save(path)

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    assert df["氏名"].tolist() == ["中澤　寿代", "中澤　寿代"]
    assert df.iloc[0]["日付"] == date(2026, 4, 1)
    assert df.iloc[0]["出勤時刻"] == time(9, 45)
    assert df.iloc[0]["退勤時刻"] == time(19, 33)
    assert df.iloc[0]["コメント"] == "JANET更改業務：時間外会議"
    assert df.iloc[0]["総労働時間(分)"] == 528
    assert df.iloc[1]["総労働時間(分)"] == 501


def test_parse_nmht_work_time_report_excel_direct(tmp_path):
    path = tmp_path / "勤務表（菅原孝）202509_202605.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務時間報告書"
    ws["C2"] = "年"
    ws["E2"] = "月"
    ws["F2"] = "勤務時間報告書Ver3.3.3"
    ws["C3"] = 2026
    ws["E3"] = 5
    ws["C6"] = "氏名"
    ws["E6"] = "菅原　孝"
    ws["C11"] = "日"
    ws["E11"] = "勤務"
    ws["F11"] = "開始時間"
    ws["H11"] = "終了時間"
    ws["T11"] = "備考"
    ws["F12"] = "時"
    ws["G12"] = "分"
    ws["H12"] = "時"
    ws["I12"] = "分"
    ws["C13"] = 1
    ws["E13"] = "出勤"
    ws["F13"] = 9
    ws["G13"] = 0
    ws["H13"] = 19
    ws["I13"] = 45
    ws["P13"] = 9.75
    ws["T13"] = "テレワーク"
    ws["C14"] = 2
    ws["E14"] = "有休(全休)"
    ws["C15"] = 3
    ws["E15"] = "出勤"
    ws["F15"] = 6
    ws["G15"] = 30
    ws["H15"] = 24
    ws["I15"] = 0
    ws["P15"] = "16:30"
    wb.save(path)

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    assert df["氏名"].tolist() == ["菅原　孝", "菅原　孝"]
    assert df.iloc[0]["日付"] == date(2026, 5, 1)
    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[0]["退勤時刻"] == time(19, 45)
    assert df.iloc[0]["コメント"] == "テレワーク"
    assert df.iloc[1]["日付"] == date(2026, 5, 3)
    assert df.iloc[1]["出勤時刻"] == time(6, 30)
    assert df.iloc[1]["退勤時刻"] == time(0, 0)
    assert df.iloc[0]["総労働時間(分)"] == 585
    assert df.iloc[1]["総労働時間(分)"] == 990


def test_parse_employment_record_excel_direct(tmp_path):
    path = tmp_path / "就業記録表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "就業記録表"
    ws["A1"] = "就 業 記 録 表"
    ws["M3"] = "西村　大"
    ws["A12"] = "日付"
    ws["D12"] = "開始時刻"
    ws["E12"] = "終了時刻"
    ws["A13"] = "2026/6/1"
    ws["D13"] = "08:32"
    ws["E13"] = "22:00"
    ws["G13"] = "8:00"
    ws["H13"] = "4:28"
    ws["J13"] = "在宅"
    wb.save(path)

    result = parse_timesheet_smart(str(path))
    row = result["df"].iloc[0]

    assert result["mode"] == "direct"
    assert row["氏名"] == "西村　大"
    assert row["日付"] == date(2026, 6, 1)
    assert row["出勤時刻"] == time(8, 32)
    assert row["退勤時刻"] == time(22, 0)
    assert row["総労働時間(分)"] == 748


def test_parse_work_result_report_excel_direct(tmp_path):
    path = tmp_path / "作業実績報告書.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "作業実績報告書"
    ws["A1"] = "作業実績報告書"
    ws["D7"] = "佐藤　可奈子"
    ws["A14"] = "日付"
    ws["M14"] = "勤怠区分"
    ws["N14"] = "備考"
    ws["P14"] = "労働時間"
    ws["S14"] = "勤怠区分"  # 集計用。明細の勤怠区分と取り違えない
    ws["D15"] = "開始時刻"
    ws["E15"] = "終了時刻"
    ws["P15"] = "実労働"
    ws["A16"] = "2026/6/1"
    ws["D16"] = "08:45"
    ws["E16"] = "17:00"
    ws["M16"] = "出勤"
    ws["N16"] = "現場"
    ws["P16"] = timedelta(hours=7, minutes=30)
    wb.save(path)

    result = parse_timesheet_smart(str(path))
    row = result["df"].iloc[0]

    assert result["mode"] == "direct"
    assert row["氏名"] == "佐藤　可奈子"
    assert row["日付"] == date(2026, 6, 1)
    assert row["総労働時間(分)"] == 450
    assert row["コメント"] == "出勤 / 現場"


def test_parse_work_result_report_v202606_layout_direct(tmp_path):
    """26年度版 v202606（列が1本ずれたテンプレート）もAIなしで直接解析する。

    林広美さん 2026-07 の実例: N列に「勤務内容」が挿入されて備考が O:P へ、
    実労働ブロックが P→Q へずれた。P15 決め打ちの判定では不発になり、
    AI解析へフォールバック → AIが失敗した回に請求勤怠ごと捨てられて
    「jinjer にはいるが請求勤怠が届いていない」未提出者に落ちた。
    """
    path = tmp_path / "林_2026年7月作業報告書（26年度版）_v202606.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "作業実績報告書"
    ws["A1"] = "作業実績報告書"
    ws["D7"] = "林広美"
    ws["A14"] = "日付"
    ws["M14"] = "勤怠区分"
    ws["N14"] = "勤務内容"
    ws["O14"] = "備考"
    ws["Q14"] = "労働時間"
    ws["T14"] = "勤怠区分"  # 集計用。明細の勤怠区分と取り違えない
    ws["D15"] = "開始時刻"
    ws["E15"] = "終了時刻"
    ws["Q15"] = "実労働"
    ws["A16"] = "2026/7/21"
    ws["D16"] = "09:00"
    ws["E16"] = "17:30"
    ws["M16"] = "早退"
    ws["N16"] = "DHD_社内ヘルプデスク"
    ws["O16"] = "自社面談のため"
    ws["Q16"] = timedelta(hours=7, minutes=30)
    ws["T16"] = 1
    wb.save(path)

    result = parse_timesheet_smart(str(path))
    row = result["df"].iloc[0]

    assert result["mode"] == "direct"
    assert row["氏名"] == "林広美"
    assert row["日付"] == date(2026, 7, 21)
    assert row["出勤時刻"] == time(9, 0)
    assert row["退勤時刻"] == time(17, 30)
    assert row["総労働時間(分)"] == 450
    # 毎日同じ値が並ぶ「勤務内容」は差異一覧のノイズになるので拾わない
    assert row["コメント"] == "早退 / 自社面談のため"


def test_ai_total_work_time_is_normalized_to_minutes():
    parsed = {
        "employee_name": "山田太郎",
        "records": [{
            "date": "2026-06-01",
            "start_time": "09:00",
            "end_time": "18:00",
            "total_work_time": "7:30",
            "comment": None,
        }],
    }

    df = timesheet_parser._normalize_records(parsed)

    assert df.iloc[0]["総労働時間(分)"] == 450


def test_duration_to_minutes_accepts_excel_and_text_formats():
    assert timesheet_parser._duration_to_minutes(timedelta(hours=8, minutes=41)) == 521
    assert timesheet_parser._duration_to_minutes(time(7, 30)) == 450
    assert timesheet_parser._duration_to_minutes("8時間30分") == 510
    assert timesheet_parser._duration_to_minutes(7.5) == 450


def test_nmht_name_can_fall_back_to_filename():
    assert timesheet_parser._extract_nmht_name_from_filename("勤務表（土屋）202606.xlsx") == "土屋"


def test_parse_sap_timesheet_excel_direct(tmp_path):
    path = tmp_path / "sap_timesheet.xlsx"
    pd.DataFrame([
        {
            "スタッフ": "上原, 奏吾",
            "食事休憩 1 開始": "12:00:00",
            "食事休憩 1 終了": "13:00:00",
            "終了時刻": "17:30:00",
            "出勤時刻": "09:00:00",
            "タイムシートステータス": "請求済み",
            "時間エントリ日": "2026-04-01 00:00:00",
        },
        {
            "スタッフ": "二宮, 正年",
            "終了時刻": "18:00:00",
            "出勤時刻": "10:00:00",
            "タイムシートステータス": "請求済み",
            "時間エントリ日": "2026-04-02 00:00:00",
        },
    ]).to_excel(path, index=False)

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    assert df["氏名"].tolist() == ["上原, 奏吾", "二宮, 正年"]
    assert df.iloc[0]["日付"] == date(2026, 4, 1)
    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[0]["退勤時刻"] == time(17, 30)
    assert df["コメント"].isna().all()


def test_parse_sap_timesheet_csv_direct(tmp_path):
    path = tmp_path / "sap_timesheet.csv"
    pd.DataFrame([
        {
            "スタッフ": "上原, 奏吾",
            "終了時刻": "17:30",
            "出勤時刻": "9:00",
            "時間エントリ日": "2026/4/1",
        },
        {
            "スタッフ": "及川, 航平",
            "終了時刻": "09:30",
            "出勤時刻": "00:00",
            "時間エントリ日": "2026/4/4",
        },
    ]).to_csv(path, index=False, encoding="utf-8-sig")

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    assert df.iloc[0]["氏名"] == "上原, 奏吾"
    assert df.iloc[0]["日付"] == date(2026, 4, 1)
    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[1]["出勤時刻"] == time(0, 0)
    assert df.iloc[1]["退勤時刻"] == time(9, 30)


def test_parse_estaffing_timesheet_csv_direct(tmp_path):
    path = tmp_path / "estaffing.csv"
    pd.DataFrame([
        {
            "e-staffing契約No": "C100887997-040",
            "スタッフ氏名": "寺山 枝美",
            "就業年月日": "2026/4/1",
            "日々勤怠状況": "承認済",
            "区分": "通常",
            "開始時刻": "9:00",
            "終了時刻": "24:00:00",
            "休憩時間": "1:30",
            "備考コメント": "",
        },
        {
            "e-staffing契約No": "C100887997-040",
            "スタッフ氏名": "寺山 枝美",
            "就業年月日": "2026/4/2",
            "日々勤怠状況": "承認済",
            "区分": "通常",
            "開始時刻": "9:00",
            "終了時刻": "22:30",
            "休憩時間": "2:00",
            "備考コメント": "テレワーク",
        },
    ]).to_csv(path, index=False, encoding="cp932")

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    assert df.iloc[0]["氏名"] == "寺山 枝美"
    assert df.iloc[0]["日付"] == date(2026, 4, 1)
    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[0]["退勤時刻"] == time(0, 0)
    assert pd.isna(df.iloc[0]["コメント"])
    assert df.iloc[1]["退勤時刻"] == time(22, 30)
    assert df.iloc[1]["コメント"] == "テレワーク"


def test_estaffing_break_includes_night_break(tmp_path):
    """e-staffingの実休憩は 休憩時間(AD列) + 深夜休憩時間(AE列) の合算。

    夜勤者は深夜休憩が別列のため、AD列だけだと正味労働が過大になる
    （河端さんの90分・加藤英人さんの30分の偽差異の原因）。
    """
    path = tmp_path / "estaffing.csv"
    pd.DataFrame([
        {   # 夜勤: 20:00〜翌8:15 拘束735分 − (休憩60 + 深夜休憩30) = 645分
            "スタッフ氏名": "加藤 英人",
            "就業年月日": "2026/6/3",
            "開始時刻": "20:00",
            "終了時刻": "32:15:00",
            "休憩時間": "1:00",
            "深夜休憩時間": "0:30",
            "備考コメント": "",
        },
        {   # 深夜休憩のみ: 拘束735分 − 90分 = 645分
            "スタッフ氏名": "河端 桂大",
            "就業年月日": "2026/6/1",
            "開始時刻": "20:00",
            "終了時刻": "32:15:00",
            "休憩時間": "0:00",
            "深夜休憩時間": "1:30",
            "備考コメント": "",
        },
    ]).to_csv(path, index=False, encoding="cp932")

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert df.iloc[0]["総労働時間(分)"] == 645
    assert df.iloc[1]["総労働時間(分)"] == 645


def test_sap_placeholder_row_has_no_time_and_tokki(tmp_path):
    """Fieldglassの「00:00〜00:00・24h」プレースホルダ行は時刻なし＋特記付きで取り込む。

    実打刻ではないため、00:00打刻や総労働24:00として突合してはならない。
    """
    path = tmp_path / "sap_timesheet.csv"
    pd.DataFrame([
        {
            "スタッフ": "福家, 寛昭",
            "出勤時刻": "00:00",
            "終了時刻": "00:00",
            "エントリ日の労働時間 (ブレークダウンなし)": "24.000",
            "時間エントリ日": "2026-06-30",
        },
        {
            "スタッフ": "福家, 寛昭",
            "出勤時刻": "08:30",
            "終了時刻": "17:00",
            "エントリ日の労働時間 (ブレークダウンなし)": "7.500",
            "時間エントリ日": "2026-06-29",
        },
    ]).to_csv(path, index=False, encoding="utf-8-sig")

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    placeholder = df[df["日付"] == date(2026, 6, 30)].iloc[0]
    assert placeholder["出勤時刻"] is None or pd.isna(placeholder["出勤時刻"])
    assert placeholder["退勤時刻"] is None or pd.isna(placeholder["退勤時刻"])
    assert placeholder["総労働時間(分)"] is None or pd.isna(placeholder["総労働時間(分)"])
    assert "Fieldglass時刻なし" in str(placeholder["特記"])
    normal = df[df["日付"] == date(2026, 6, 29)].iloc[0]
    assert normal["出勤時刻"] == time(8, 30)
    assert str(normal["特記"] or "") in ("", "nan")


def test_parse_estaffing_timesheet_text_direct(tmp_path):
    path = tmp_path / "estaffing.txt"
    lines = [
        "H\tC102350238-017\t2026/01/01\t2026/03/31\tAHIG\t東 昭夫",
        "D\t2026/03/02\t1\t9:00\t16:30\t1:00\t\t*[テレワーク、場所：自宅]\t0\t6:30\t6:30",
        "D\t2026/03/03\t\t\t\t\t\t\t\t\t",
        "D\t2026/03/04\t1\t16:45\t33:30\t1:00\t\t夜勤\t0\t15:45\t15:45",
    ]
    path.write_text("\n".join(lines), encoding="cp932")

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert len(df) == 2
    assert df.iloc[0]["氏名"] == "東 昭夫"
    assert df.iloc[0]["日付"] == date(2026, 3, 2)
    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[0]["退勤時刻"] == time(16, 30)
    assert df.iloc[0]["コメント"] == "*[テレワーク、場所：自宅]"
    assert df.iloc[1]["出勤時刻"] == time(16, 45)
    assert df.iloc[1]["退勤時刻"] == time(9, 30)


def test_parse_fieldglass_pdf_direct(monkeypatch, tmp_path):
    path = tmp_path / "ts_uploadid_timesheet_ERCSTS01161606奈良.pdf"
    pdf_text = "\n".join([
        "Time Sheet",
        "ID ERCSTS01161606 Worker Nara, Takahiro(ERCSWK00144174)",
        "Period 2026-04-01 to 2026-04-30 Job Posting Support Engineer|JP|Job Stage",
        "Time in/time out",
        "Day 3-30 Mon 3-31 Tue 4-01 Wed 4-02 Thu 4-03 Fri 4-04 Sat 4-05 Sun Total",
        "Time In 09:00 09:00 09:00 00:00 00:00",
        "Meal Break 1 12:00 - 13:00 12:00 - 13:00 12:00 - 13:00",
        "Time Out 20:30 18:00 21:30 00:00 00:00",
        "Total 10h 30m 8h 0m 11h 30m 0h 0m 0h 0m 30h 0m",
    ])
    monkeypatch.setattr(
        timesheet_parser,
        "_pdf_to_text_or_bytes",
        lambda filepath: (pdf_text, None),
    )

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert df["氏名"].tolist() == ["奈良", "奈良", "奈良"]
    assert df["日付"].tolist() == [date(2026, 4, 1), date(2026, 4, 2), date(2026, 4, 3)]
    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[0]["退勤時刻"] == time(20, 30)
    assert df.iloc[2]["退勤時刻"] == time(21, 30)


def test_parse_fieldglass_pdf_12h_ampm(monkeypatch, tmp_path):
    """Fieldglass PDF の 12時間制(AM/PM) を 24時間制に正しく変換する。

    実データ(MAHARJAN/奈良)の不具合の再現: "6:00 PM"(=18:00) の PM を落として
    6:00 と誤読し、出勤(9:00)より早い→夜勤跨ぎ+24h で 30:00 に化けていた。
    夜勤(太田)は Time In "8:45 PM"=20:45、Time Out 翌9:00 AM も確認する。
    """
    path = tmp_path / "timesheet_ERCSTS_ampm.pdf"
    pdf_text = "\n".join([
        "Time Sheet",
        "ID ERCSTS01182072 Worker Ramita, Maharjan(ERCSWK00144175)",
        "Period 2026-06-01 to 2026-06-30 Job Posting Support Engineer|JP|Job Stage",
        "Time in/time out",
        "Day 6-01 Mon 6-02 Tue 6-03 Wed 6-06 Sat 6-07 Sun Total",
        "Time In 9:00 AM 9:00 AM 9:00 AM 12:00 AM 12:00 AM",
        "Time Out 6:00 PM 8:00 PM 7:00 PM 12:00 AM 12:00 AM",
        "Total 8h 0m 10h 0m 9h 0m 0h 0m 0h 0m 27h 0m",
    ])
    monkeypatch.setattr(
        timesheet_parser,
        "_pdf_to_text_or_bytes",
        lambda filepath: (pdf_text, None),
    )

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert df.iloc[0]["出勤時刻"] == time(9, 0)
    assert df.iloc[0]["退勤時刻"] == time(18, 0)   # 6:00 PM → 18:00（30:00 に化けない）
    assert df.iloc[1]["退勤時刻"] == time(20, 0)   # 8:00 PM → 20:00
    assert df.iloc[2]["退勤時刻"] == time(19, 0)   # 7:00 PM → 19:00


def test_fieldglass_12h_to_24h_conversion():
    """AM/PM → 24時間制の境界値（12:00 AM=00:00, 12:00 PM=12:00）。"""
    f = timesheet_parser._fieldglass_12h_to_24h
    assert f(12, 0, "AM") == "00:00"   # 深夜0時
    assert f(12, 0, "PM") == "12:00"   # 正午
    assert f(6, 0, "PM") == "18:00"
    assert f(8, 45, "PM") == "20:45"
    assert f(9, 0, "AM") == "09:00"
    assert f(20, 30, None) == "20:30"  # AM/PM無し＝既に24時間制


def test_parse_fieldglass_pdf_katakana_filename_uses_romaji_worker_name(monkeypatch, tmp_path):
    """ファイル名がカタカナ通称（ラミタ）でも、PDF本文のローマ字氏名で突合できる。

    jinjer側はローマ字（MAHARJAN RAMITA）で登録されるため、カタカナ通称のままでは
    一致しない。PDF本文の "Worker Ramita, Maharjan" を採用して MAHARJAN RAMITA とする。
    """
    path = tmp_path / "timesheet_ERCSTS0ラミタさん.pdf"
    pdf_text = "\n".join([
        "Time Sheet",
        "ID ERCSTS01173544 Worker Ramita, Maharjan(ERCSWK00144175)",
        "Period 2026-05-01 to 2026-05-31 Job Posting Support Engineer|JP|Job Stage",
        "Time in/time out",
        "Day 5-01 Thu 5-02 Fri 5-03 Sat Total",
        "Time In 09:00 09:00 00:00",
        "Time Out 22:30 20:00 00:00",
        "Total 12h 30m 10h 0m 0h 0m 22h 30m",
    ])
    monkeypatch.setattr(
        timesheet_parser,
        "_pdf_to_text_or_bytes",
        lambda filepath: (pdf_text, None),
    )

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert df["氏名"].tolist() == ["MAHARJAN RAMITA", "MAHARJAN RAMITA"]
    assert df.iloc[0]["日付"] == date(2026, 5, 1)


def test_extract_fieldglass_name_from_filename_variants(tmp_path):
    """ファイル名からの氏名の取り方。

    Fieldglassが出す元の名前（timesheet_ERCSTS…）と、谷津さんが付け替える
    「氏名_社員番号」の両方を読む。社員番号が無い名前は氏名とみなさない
    （PDF本文のローマ字氏名にフォールバックさせる）。
    """
    f = timesheet_parser._extract_fieldglass_name_from_filename
    assert f(str(tmp_path / "太田 琢也_2025030.pdf")) == "太田 琢也"
    assert f(str(tmp_path / "奈良 隆宏_2022013.pdf")) == "奈良 隆宏"
    assert f(str(tmp_path / "MAHARJAN RAMITA_2022002.pdf")) == "MAHARJAN RAMITA"
    # Fieldglassの元ファイル名は従来どおり
    assert f(str(tmp_path / "timesheet_ERCSTS0ラミタさん.pdf")) == "ラミタ"
    assert f(str(tmp_path / "timesheet_ERCSTS01199793奈良 隆宏.pdf")) == "奈良 隆宏"
    # 氏名が付いていない元ファイル名・社員番号の無い名前は拾わない
    assert f(str(tmp_path / "timesheet_ERCSTS01200802.pdf")) is None
    assert f(str(tmp_path / "2026年8月分.pdf")) is None


def test_parse_fieldglass_pdf_name_id_filename_wins_over_romaji(monkeypatch, tmp_path):
    """★回帰: 「氏名_社員番号.pdf」に付け替えた勤務表は漢字氏名で突合する。

    FieldglassのPDF本文は "Worker Ohta, Takuya" のローマ字しか持たず、jinjerは
    「太田 琢也」の漢字で登録されている。本文由来のローマ字を採ると突合できず
    「jinjer未提出者」に落ちていた（2026-09 実データ: 太田 琢也_2025030 /
    奈良 隆宏_2022013。給与計算中に発覚）。
    """
    path = tmp_path / "太田 琢也_2025030.pdf"
    pdf_text = "\n".join([
        "Time Sheet",
        "ID ERCSTS01200802 Worker Ohta, Takuya(ERCSWK00145993)",
        "Period 2026-08-01 to 2026-08-31 Job Posting Automated Operations Engineer|JP",
        "Time in/time out",
        "Day 8-03 Mon 8-04 Tue Total",
        "Time In 9:00 AM 9:00 AM",
        "Time Out 6:00 PM 6:00 PM",
        "Total 8h 0m 8h 0m 16h 0m",
    ])
    monkeypatch.setattr(
        timesheet_parser,
        "_pdf_to_text_or_bytes",
        lambda filepath: (pdf_text, None),
    )

    result = parse_timesheet_smart(str(path))
    df = result["df"]

    assert result["mode"] == "direct"
    assert df["氏名"].tolist() == ["太田 琢也", "太田 琢也"], "本文のTAKUYA OHTAではない"
    assert df.iloc[0]["日付"] == date(2026, 8, 3)


def test_image_only_pdf_is_sent_as_image_for_ai_fallback(tmp_path):
    path = tmp_path / "scanned_timesheet.pdf"
    image = Image.new("RGB", (480, 640), "white")
    draw = ImageDraw.Draw(image)
    draw.text((40, 40), "タイムシート", fill="black")
    draw.text((40, 90), "氏名 小林 拓己", fill="black")
    image.save(path, "PDF")

    content, file_type, media_type = _load_file_for_claude(str(path))

    assert file_type == "image"
    assert media_type == "image/png"
    assert content.startswith(b"\x89PNG")


def test_parse_timesheet_smart_falls_back_when_legend_parser_returns_empty(monkeypatch, tmp_path):
    path = tmp_path / "timesheet.png"
    Image.new("RGB", (320, 240), "white").save(path)

    monkeypatch.setattr(
        timesheet_parser,
        "parse_with_legend_extraction",
        lambda file_content, file_type, media_type=None: {"mode": "direct", "data": []},
    )
    monkeypatch.setattr(
        timesheet_parser,
        "_parse_with_claude",
        lambda file_content, file_type, media_type=None: {
            "employee_name": "小林 拓己",
            "records": [
                {
                    "date": "2026-04-01",
                    "start_time": "09:00",
                    "end_time": "17:30",
                    "comment": None,
                }
            ],
        },
    )

    result = parse_timesheet_smart(str(path))

    assert result["mode"] == "direct"
    assert len(result["df"]) == 1
    assert result["df"].iloc[0]["氏名"] == "小林 拓己"
    assert result["df"].iloc[0]["出勤時刻"] == time(9, 0)
