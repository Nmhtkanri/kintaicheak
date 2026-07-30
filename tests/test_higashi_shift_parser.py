"""東さん形式の月次シフト希望表パーサ回帰テスト。"""

from datetime import date, timedelta

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from services.higashi_shift_parser import (
    is_higashi_shift_xlsx,
    parse_higashi_shift_xlsx,
)
from services.multi_year_shift_parser import parse_structured_files


def _make_higashi_book(path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "2025 1"

    worksheet["B1"] = date(2026, 8, 1)
    worksheet["T4"] = "2026年8月エヌエム・ヒューマテック要員シフト希望表"
    worksheet["B5"] = "勤務時間帯"
    worksheet["C7"] = "日：09:00-16:30　（終日支店・テレワーク）休憩：12：00-13：00"
    worksheet["C8"] = "前：09:00-12:00　休憩：なし"
    worksheet["C9"] = "後：12:00-16:30　休憩：なし"

    worksheet["B12"] = "拠\n点"
    worksheet["C12"] = "名前"
    worksheet["D12"] = "役割"
    worksheet["E12"] = "備考\n(所属)"
    for col, code in zip(range(54, 59), ("日", "前", "後", "自", "○")):
        worksheet.cell(12, col).value = code

    # 実ファイル同様、可視対象月 N:AR の前後に非表示の補助日付列を置く。
    raw_start = date(2021, 12, 25)
    for offset, col in enumerate(range(7, 54)):  # G:BA
        worksheet.cell(12, col).value = raw_start + timedelta(days=offset)
    for col in list(range(7, 14)) + list(range(45, 54)):  # G:M, AS:BA
        worksheet.column_dimensions[get_column_letter(col)].hidden = True

    worksheet["B14"] = "G"
    worksheet["C14"] = "東"
    worksheet["D14"] = "メンバー"

    # 可視 N:AR は8月1〜31日に対応。3/5/7日だけ勤務にする。
    for day in (3, 5, 7):
        worksheet.cell(14, 13 + day).value = "日"

    # 非表示の翌月列に残った値。修正前は day-of-month 後勝ちで
    # 8/1(土)・8/2(日)などへ上書きされていた。
    for col, code in zip(range(45, 50), ("日", "外", "日", "外", "日")):
        worksheet.cell(14, col).value = code

    workbook.save(path)


def test_higashi_parser_reads_only_visible_target_month_columns(tmp_path):
    path = tmp_path / "2026年8月シフト希望表.xlsx"
    _make_higashi_book(path)

    assert is_higashi_shift_xlsx(str(path)) is True
    parsed = parse_higashi_shift_xlsx(str(path), 2026, 8)

    assert parsed["year"] == 2026
    assert parsed["month"] == 8
    assert parsed["section_info"]["start_col"] == 14  # N
    assert parsed["section_info"]["end_col"] == 44    # AR
    assert parsed["section_info"]["visible_columns_only"] is True

    employee = parsed["employees"][0]
    assert employee["name"] == "東"
    assert len(employee["shifts"]) == 31
    by_day = {int(shift["date"][-2:]): shift["code"] for shift in employee["shifts"]}
    assert by_day[1] == ""
    assert by_day[2] == ""
    assert by_day[3] == "日"
    assert by_day[4] == ""
    assert by_day[5] == "日"
    assert by_day[7] == "日"
    assert "外" not in by_day.values()

    legend = {entry["code"]: entry for entry in parsed["legend"]}
    assert legend["日"]["start_time"] == "09:00"
    assert legend["日"]["end_time"] == "16:30"
    assert legend["日"]["break_minutes"] == 60


def test_parse_structured_files_consumes_higashi_book_without_ai(tmp_path):
    path = tmp_path / "2026年8月シフト希望表.xlsx"
    _make_higashi_book(path)

    result = parse_structured_files([str(path)], 2026, 8)

    assert result is not None
    sheets, consumed, warnings = result
    assert consumed == [str(path)]
    assert warnings == []
    assert len(sheets) == 1
    shifts = sheets[0]["employees"][0]["shifts"]
    assert shifts[0] == {"date": "2026-08-01", "code": ""}
    assert shifts[1] == {"date": "2026-08-02", "code": ""}
    assert shifts[2] == {"date": "2026-08-03", "code": "日"}


def test_higashi_parser_rejects_other_titles(tmp_path):
    path = tmp_path / "other.xlsx"
    _make_higashi_book(path)
    workbook = openpyxl.load_workbook(path)
    workbook.active["T4"] = "2026年8月 別会社シフト表"
    workbook.save(path)

    assert is_higashi_shift_xlsx(str(path)) is False
    assert parse_structured_files([str(path)], 2026, 8) is None


def test_higashi_parser_rejects_target_month_mismatch(tmp_path):
    path = tmp_path / "2026年8月シフト希望表.xlsx"
    _make_higashi_book(path)

    with pytest.raises(ValueError, match="一致しません"):
        parse_higashi_shift_xlsx(str(path), 2026, 7)
