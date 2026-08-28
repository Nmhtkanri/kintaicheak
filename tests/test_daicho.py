# -*- coding: utf-8 -*-
"""派遣元管理台帳ジェネレーターのテスト（合成データ。実データは使わない）。

実行: python -X utf8 -m pytest tests/test_daicho.py -q   （kintai-checker 直下から）
"""
from __future__ import annotations

import csv
import datetime as dt
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from services.daicho import config  # noqa: E402
from services.daicho.estaffing import Contract, contracts_in_quarter, load_contracts, parse_date, split_time  # noqa: E402
from services.daicho.records import build_record  # noqa: E402
from services.daicho.roster import Person, Roster, load_roster, normalize_name  # noqa: E402
from services.daicho.template import CELL_MAP, EXTRA_ROWS, build_template  # noqa: E402
from services.daicho.writer import estimate_row_height, sheet_title, write_quarter  # noqa: E402

CMD = "事業所の名称及び所在地その他派遣就業場所"


def _write_cp932_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    with path.open("w", encoding="cp932", newline="") as fp:
        w = csv.writer(fp)
        w.writerow(header)
        w.writerows(rows)


def _tc_row(no: str, sei: str, mei: str, start: str, end: str, **over) -> dict:
    d = {
        "契約No": no, "契約区分": "2", "ステータス": "確定", "就業先企業名": "テスト商事",
        "就業先事業所": "本社", "就業先住所": "東京都千代田区1-1", "就業先正式事業所名称": "テスト商事株式会社　本社",
        "就業先正式部署": "第一部", "スタッフ姓（日本語）": sei, "スタッフ名（日本語）": mei,
        "スタッフ姓（カナ）": "テスト", "スタッフ名（カナ）": "タロウ", "スタッフ性別": "男",
        "契約開始日": start, "契約終了日": end, "シフト": "0", "休日（その他）": "",
        "組織単位": "第一部", "組織の長の職名": "部長", "事業所抵触日": "2027/10/1",
        "派遣元での雇用形態": "0", "法定休日": "法定休日:日曜日", "契約確定日": "2025/6/25",
        "便宜供与：診療施設": "0", "便宜供与：給食施設": "1", "便宜供与：休憩室": "1", "便宜供与：更衣室": "0",
        "安全及び衛生": "安全衛生の定め", "期間制限の対象外理由：無期雇用派遣労働者": "1",
        "期間制限の対象外理由：60歳以上派遣労働者": "0", "期間制限の対象外理由：有期プロジェクト業務": "0",
        "期間制限の対象外理由：日数限定業務": "0", "期間制限の対象外理由：産前産後、育児休業、介護休業等の代替要員": "0",
        "契約書備考": "備考", "苦情処理結果": "", "36協定": "", "36協定#2": "",
    }
    d.update(over)
    return d


def _cpi_row(no: str, name: str, **over) -> dict:
    d = {
        "e-staffing契約No": no, "契約番号": "TT-TT", "派遣先企業 名称": "テスト商事株式会社",
        f"{CMD} 部署名称": "第一部", f"{CMD} 事業所の所在地及び就業場所": "東京都千代田区1-1",
        f"{CMD} 事業所の名称": "本社", f"{CMD} 指揮命令者部署": "第一部", f"{CMD} 指揮命令者役職": "課長",
        f"{CMD} 指揮命令者氏名": "指揮 太郎", f"{CMD} 部署TEL": "03-0000-0000",
        "職種": "", "業務内容": "サーバ運用", "責任の程度": "権限なし", "派遣期間 開始日": "2025/07/01",
        "派遣期間 終了日": "2025/09/30", "勤務日": "月 火 水 木 金", "就業時間 開始時間": "09:00",
        "就業時間 終了時間": "17:30", "就業時間 就業時間": "07:30", "休憩時間1 開始時間": "12:00",
        "休憩時間1 終了時間": "13:00", "休憩時間1 時間": "01:00", "休憩時間2 開始時間": "--:--",
        "休憩時間2 終了時間": "--:--", "休憩時間2 時間": "--:--", "休日": "土曜日 日曜日 祝日", "休日労働": "有",
        "派遣先責任者 部署": "第一部", "派遣先責任者 TEL": "03-1111-1111", "派遣先責任者 役職": "部長",
        "派遣先責任者 氏名": "責任 花子", "派遣先苦情申出先 部署": "総務部", "派遣先苦情申出先 TEL": "03-2222-2222",
        "派遣先苦情申出先 役職": "", "派遣先苦情申出先 氏名": "苦情 次郎", "派遣元責任者 部署": "営業部",
        "派遣元責任者 TEL": "03-3333-3333", "派遣元責任者 役職": "", "派遣元責任者 氏名": "元 責任",
        "派遣元苦情申出先 部署": "営業部", "派遣元苦情申出先 TEL": "03-3333-3333", "派遣元苦情申出先 役職": "",
        "派遣元苦情申出先 氏名": "元 苦情", "派遣元企業 名称": "株式会社エヌエム・ヒューマテック",
        "派遣元企業 派遣許可番号": "派13-301312", "36協定1 時間外労働、休日労働": "1日15時間…",
        "労働者氏名": name, "協定対象派遣労働者に該当するか否かの別": "該当する", "健康保険": "有", "健康保険 補足": "",
        "厚生年金": "有", "厚生年金 補足": "", "雇用保険": "無(加入対象外)", "雇用保険 補足": "時間不足：1週20時間未満の契約のため",
        "派遣元での雇用形態": "無期雇用契約", "個人抵触日": "--/--/-- (期間制限の対象外)",
        "期間制限の対象外理由": "無期雇用派遣労働者", "スタッフコード": "TT",
    }
    d.update(over)
    return d


@pytest.fixture
def inputs(tmp_path: Path):
    tc = [_tc_row("C100000001-001", "試験", "太郎", "2025/07/01", "2025/09/30"),
          _tc_row("C100000002-001", "試験", "花子", "2025/04/01", "2025/06/30"),          # 期外
          _tc_row("C100000003-001", "髙橋", "次郎", "2025/09/01", "2025/12/31",            # 跨ぎ・旧字体・60歳
                  **{"期間制限の対象外理由：60歳以上派遣労働者": "1", "派遣元での雇用形態": "1"})]
    cpi = [_cpi_row("C100000001-001", "試験 太郎"),
           _cpi_row("C100000002-001", "試験 花子", **{"派遣期間 開始日": "2025/04/01", "派遣期間 終了日": "2025/06/30"}),
           _cpi_row("C100000003-001", "髙橋 次郎", **{"派遣期間 開始日": "2025/09/01", "派遣期間 終了日": "2025/12/31",
                                                   "期間制限の対象外理由": "60歳以上派遣労働者", "派遣元での雇用形態": "有期雇用契約",
                                                   "就業時間 開始時間": "09:00 別途シフト表に定める"})]
    tc_p, cpi_p = tmp_path / "TCnmht_test.csv", tmp_path / "CPInmht_test.csv"
    _write_cp932_csv(tc_p, list(tc[0].keys()), [list(r.values()) for r in tc])
    _write_cp932_csv(cpi_p, list(cpi[0].keys()), [list(r.values()) for r in cpi])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "従業員一覧_test"
    ws.append(["No", "＊従業員ID", "パスワード", "パスワード確認用", "メールアドレス(任意)", "＊氏", "＊名", "所属グループ",
               "＊打刻グループ", "＊メイン打刻グループ", "退職日", "休職開始日", "休職終了日", "勤務先", "所属部署(就業先情報)",
               None, "所属グループ", "勤務先", "所属部署(就業先情報)"])
    ws.append([1, "2020001", None, None, "a@x", "試験", "太郎", "OT：その他", "A", "A", None, None, None, "旧社名", "", None, "OT：UAL（首都圏）", "テスト商事株式会社", ""])
    ws.append([2, "2010002", None, None, "b@x", "高橋", "次郎", "OT：その他", "A", "A", None, None, None, "テスト商事株式会社", "", None, "", "", ""])
    d = wb.create_sheet("jinjer_勤務先_9999_20260821")
    d.append(["社員番号", "職場氏名(氏)", "職場氏名(名)", "入社年月日", "勤続期間", "在籍区分", "雇用区分", "性別", "生年月日", "年齢"])
    d.append(["2020001", "試験", "太郎", dt.datetime(2020, 4, 1), "", "在籍", "正社員", "男", dt.datetime(1990, 5, 5), "35歳"])
    d.append(["2010002", "高橋", "次郎", dt.datetime(2010, 4, 1), "", "在籍", "契約社員", "男", dt.datetime(1964, 1, 1), "61歳"])
    d.append(["2005003", "退職", "済子", dt.datetime(2005, 4, 1), "", "退職", "正社員", "女", dt.datetime(1980, 1, 1), "46歳"])
    roster_p = tmp_path / "従業員一覧_test.xlsx"
    wb.save(roster_p)
    return tc_p, cpi_p, roster_p


def test_quarter_range():
    assert config.quarter_range("2025Q3") == (dt.date(2025, 7, 1), dt.date(2025, 9, 30))
    assert config.quarter_range("2026q1") == (dt.date(2026, 1, 1), dt.date(2026, 3, 31))
    assert config.quarter_range("2025Q4")[1] == dt.date(2025, 12, 31)
    assert config.quarter_label("2025Q3") == "2025年7-9月期"
    with pytest.raises(ValueError):
        config.quarter_range("2025-07")


def test_helpers():
    assert parse_date("2025/07/01") == dt.date(2025, 7, 1)
    assert parse_date("") is None
    assert split_time("09:00 別途シフト表に定める") == ("09:00", True)
    assert split_time("--:--") == ("", False)
    assert split_time("9:30") == ("09:30", False)
    assert normalize_name("髙橋", " 次郎") == normalize_name("高橋　次郎")
    assert estimate_row_height("短い") == 15.0
    assert estimate_row_height("あ" * 200) > 40


def test_load_and_join(inputs):
    tc_p, cpi_p, _ = inputs
    contracts, warnings = load_contracts(tc_p, cpi_p)
    assert len(contracts) == 3 and warnings == []
    hit = contracts_in_quarter(contracts, dt.date(2025, 7, 1), dt.date(2025, 9, 30))
    assert [c.contract_no for c in hit] == ["C100000001-001", "C100000003-001"]   # 期外の花子は落ちる・跨ぎは入る
    assert hit[0].client_name == "テスト商事株式会社"


def test_join_warns_when_cpi_missing(tmp_path, inputs):
    tc_p, cpi_p, _ = inputs
    cpi_only_one = tmp_path / "cpi1.csv"
    rows = list(csv.reader(open(cpi_p, encoding="cp932")))
    _write_cp932_csv(cpi_only_one, rows[0], rows[1:2])
    contracts, warnings = load_contracts(tc_p, cpi_only_one)
    assert len(contracts) == 3
    assert sum("契約書・通知書データ(CPI)に無い" in w for w in warnings) == 2


def test_roster_and_record(inputs):
    tc_p, cpi_p, roster_p = inputs
    roster = load_roster(roster_p)
    assert set(roster.people) == {"2020001", "2010002", "2005003"}
    p, state = roster.find("試験", "太郎")
    assert state == "ok" and p.emp_id == "2020001" and p.birth == dt.date(1990, 5, 5)
    assert p.workplace == "テスト商事株式会社" and p.group == "OT：UAL（首都圏）"   # 右側（更新後）の列を優先
    assert roster.find("髙橋", "次郎")[1] == "ok"                                   # 旧字体でも一致
    assert roster.find("居ない", "人")[1] == "none"

    contracts, _ = load_contracts(tc_p, cpi_p)
    q = (dt.date(2025, 7, 1), dt.date(2025, 9, 30))
    c1 = next(c for c in contracts if c.contract_no == "C100000001-001")
    rec = build_record(c1, p, "ok", *q, generated_at=dt.datetime(2026, 8, 21, 12, 0))
    f = rec.fields
    assert f["社員番号"] == "2020001" and f["氏名"] == "試験 太郎" and f["年齢"] == "35歳"
    assert f["60歳以上の者であるか否かの別"] == "60歳未満"
    assert f["協定対象派遣労働者であるか否かの別"] == "協定対象派遣労働者"
    assert f["有期か無期かの別"] == "無期雇用派遣労働者"
    assert f["就業時間"] == "09:00～17:30（実働 07:30）" and f["休憩時間1"] == "休憩1: 12:00～13:00（01:00）"
    assert f["派遣先責任者_役職氏名"] == "部長 責任 花子" and f["指揮命令者_役職氏名"] == "課長 指揮 太郎"
    assert f["雇用保険_加入"] == "無" and "1週20時間未満" in f["資格取得届_雇用保険"]
    assert f["健康保険_加入"] == "有" and f["資格取得届_健康保険"].endswith("提出済（加入）")
    assert f["組織単位"] == "第一部（組織の長の職名：部長）"
    assert f["便宜供与"] == "給食施設、休憩室"
    assert f["事業所単位の抵触日"] == "2027/10/01"
    assert "無期雇用派遣労働者" in f["個人単位の抵触日"]
    assert rec.warnings == []

    # 60歳・有期・シフト・jinjer不一致の検算
    c3 = next(c for c in contracts if c.contract_no == "C100000003-001")
    p3, _ = roster.find(*c3.worker_sei_mei)
    rec3 = build_record(c3, p3, "ok", *q)
    assert rec3.fields["60歳以上の者であるか否かの別"] == "60歳以上"
    assert rec3.fields["有期か無期かの別"] == "有期雇用派遣労働者"
    assert "別途シフト表に定める" in rec3.fields["就業時間"]
    assert rec3.warnings == []   # 契約書=60歳以上 / jinjer 61歳 → 一致、契約社員=有期 → 一致

    rec_none = build_record(c1, None, "none", *q)
    assert rec_none.fields["社員番号"] == "" and any("見つからない" in w for w in rec_none.warnings)


def test_fieldglass_name_candidates():
    from services.daicho.fieldglass import name_candidates
    assert name_candidates("上原, 奏吾") == [normalize_name("上原奏吾"), normalize_name("奏吾上原")]
    assert name_candidates("髙橋, 次郎")[0] == normalize_name("高橋次郎")
    assert name_candidates("単名") == [normalize_name("単名")]
    assert name_candidates("") == []


def test_contract_from_workorder(inputs):
    from services.daicho.estaffing import load_contracts
    from services.daicho.fieldglass import WorkOrder, contract_from_workorder
    tc_p, cpi_p, roster_p = inputs
    contracts, _ = load_contracts(tc_p, cpi_p)
    src = next(c for c in contracts if c.contract_no == "C100000001-001")
    person, _ = load_roster(roster_p).find("試験", "太郎")
    wo = WorkOrder(wo_id="WOJP00001", revision=1, staff_raw="試験, 太郎", staff_fg_id="F1",
                   start=dt.date(2026, 4, 1), end=None, site="本社ビル",
                   business_unit="新しい部", supervisors=["監督 一郎", "監督 二郎"])
    vc = contract_from_workorder(wo, src, dt.date(2026, 6, 30), person)
    assert vc.start == dt.date(2026, 4, 1) and vc.end == dt.date(2026, 6, 30)   # 終了日空→期末
    assert vc.contract_no == "FG-WOJP00001(改訂1)"
    assert vc.t("組織単位") == "新しい部"
    rec = build_record(vc, person, "ok", dt.date(2026, 4, 1), dt.date(2026, 6, 30))
    assert rec.fields["指揮命令者_役職氏名"] == "監督 一郎、監督 二郎"       # WOの現行値
    assert rec.fields["派遣先責任者_役職氏名"] == "部長 責任 花子"           # 引き継ぎ値
    assert rec.fields["責任の程度"] == "権限なし" and rec.fields["健康保険_加入"] == "有"
    assert rec.fields["契約期間"] == "2026/04/01～2026/06/30"
    # 引き継ぎ元なし
    vc2 = contract_from_workorder(wo, None, dt.date(2026, 6, 30), None)
    rec2 = build_record(vc2, None, "none", dt.date(2026, 4, 1), dt.date(2026, 6, 30))
    assert rec2.fields["氏名"] == "試験 太郎" and rec2.fields["派遣先名称"] == "ユニアデックス株式会社"
    assert rec2.fields["責任の程度"] == ""

    # SAP詳細で上書き（空欄の項目は引き継ぎのまま）
    detail = {
        "clientResponsible": {"name": "堀 史", "title": "部長", "department": "マネ部", "phone": "050-0000-0000"},
        "complaintRecipient": {"name": "古賀 正英", "title": "本部長", "department": "マネ本部", "phone": "050-1111-1111"},
        "responsibilityDegree": "一般業務 役職を有さない",
        "businessContent": "営業支援/構築支援",
        "workdaysHolidays": "月曜 勤務-火曜 勤務-水曜 勤務-木曜 勤務-金曜 勤務-土曜 非勤務-日曜 非勤務",
        "workdaysHolidaysNotes": "休日は土日祝および年末年始12/29～1/3",
        "workStart": "08:45", "workEnd": "17:45", "breakStart": "12:00", "breakEnd": "13:00",
        "jobPostingId": "36828JP00001732",
    }
    vc3 = contract_from_workorder(wo, src, dt.date(2026, 6, 30), person, detail)
    rec3 = build_record(vc3, person, "ok", dt.date(2026, 4, 1), dt.date(2026, 6, 30))
    assert rec3.fields["派遣先責任者_役職氏名"] == "部長 堀 史" and rec3.fields["派遣先責任者_部署"] == "マネ部"
    assert rec3.fields["苦情申出先_役職氏名"] == "本部長 古賀 正英"
    assert rec3.fields["責任の程度"] == "一般業務 役職を有さない"
    assert rec3.fields["業務内容"] == "営業支援/構築支援"
    assert rec3.fields["就業曜日"] == "月 火 水 木 金" and rec3.fields["休日"] == "土 日"
    assert "年末年始" in rec3.fields["休日備考"]
    assert rec3.fields["就業時間"] == "08:45～17:45" and "12:00～13:00" in rec3.fields["休憩時間1"]
    assert rec3.fields["指揮命令者_役職氏名"] == "監督 一郎、監督 二郎"     # 指揮命令者はWOのまま
    assert rec3.fields["健康保険_加入"] == "有"                            # 保険は引き継ぎのまま
    # 就業時間が空の詳細 → 引き継ぎ値を保持
    detail2 = {"clientResponsible": {"name": "堀 史", "title": "部長", "department": "マネ部", "phone": ""},
               "responsibilityDegree": "", "businessContent": "", "workdaysHolidays": "",
               "workStart": "", "workEnd": "", "breakStart": "", "breakEnd": ""}
    vc4 = contract_from_workorder(wo, src, dt.date(2026, 6, 30), person, detail2)
    rec4 = build_record(vc4, person, "ok", dt.date(2026, 4, 1), dt.date(2026, 6, 30))
    assert rec4.fields["就業時間"] == "09:00～17:30（実働 07:30）"          # 引き継ぎ
    assert rec4.fields["責任の程度"] == "権限なし"                          # 引き継ぎ
    assert rec4.fields["派遣先責任者_役職氏名"] == "部長 堀 史"             # 上書き


def test_defaults_and_person_fill(inputs):
    from services.daicho.estaffing import load_contracts
    from services.daicho.fieldglass import WorkOrder, apply_defaults, contract_from_workorder, derive_defaults, fill_from_person
    tc_p, cpi_p, roster_p = inputs
    contracts, _ = load_contracts(tc_p, cpi_p)
    defaults = derive_defaults(contracts, dt.date(2025, 7, 1), dt.date(2025, 9, 30))
    assert defaults["cpi"]["健康保険"] == "有" and defaults["cpi"]["派遣元責任者 氏名"] == "元 責任"
    assert defaults["cpi"]["協定対象派遣労働者に該当するか否かの別"] == "該当する"
    assert defaults["tc"]["事業所抵触日"] == "2027/10/1"

    wo = WorkOrder(wo_id="WOJP00002", revision=0, staff_raw="新人, 太郎", staff_fg_id="F9",
                   start=dt.date(2026, 5, 18), end=None, site="本社ビル", business_unit="新部署", supervisors=["監督"])
    vc = contract_from_workorder(wo, None, dt.date(2026, 6, 30), None)
    filled = apply_defaults(vc, defaults)
    assert "健康保険" in filled and vc.c("36協定1 時間外労働、休日労働")
    seishain = Person(emp_id="2026009", sei="新人", mei="太郎", birth=dt.date(2000, 1, 1), employment_type="正社員")
    fill_from_person(vc, seishain, wo.start)
    rec = build_record(vc, seishain, "ok", dt.date(2026, 4, 1), dt.date(2026, 6, 30))
    assert rec.fields["有期か無期かの別"] == "無期雇用派遣労働者"
    assert "無期雇用派遣労働者" in rec.fields["期間制限の対象外理由"]
    assert rec.fields["健康保険_加入"] == "有" and rec.fields["派遣元責任者_役職氏名"] == "元 責任"
    # 有期かつ60歳未満 → 受入開始から3年の抵触日（要確認つき）
    vc2 = contract_from_workorder(wo, None, dt.date(2026, 6, 30), None)
    apply_defaults(vc2, defaults)
    keiyaku = Person(emp_id="2026010", sei="別", mei="人", birth=dt.date(1990, 1, 1), employment_type="契約社員")
    fill_from_person(vc2, keiyaku, wo.start)
    assert vc2.c("個人抵触日") == "2029/05/18（受入開始から3年・要確認）"
    assert vc2.c("派遣元での雇用形態") == "有期雇用契約"


def test_jinjer_api_merge(tmp_path, inputs):
    from services.daicho.jinjer_api import load_cache, merge_into_roster, refresh_cache  # noqa: F401
    from services.daicho.jinjer_api import _person_from_dict, _person_to_dict
    _, _, roster_p = inputs
    roster = load_roster(roster_p)
    assert roster.find("既存", "無person")[1] == "none"
    api_people = [
        Person(emp_id="2001001", sei="退職済", mei="花子", birth=dt.date(1970, 2, 2), sex="女",
               employment_type="正社員", status="退職", retired=dt.date(2026, 5, 31)),
        Person(emp_id="2020001", sei="試験", mei="太郎", birth=dt.date(1990, 5, 5), sex="男",
               employment_type="正社員", status="在籍"),
    ]
    note = merge_into_roster(roster, api_people)
    assert "追加 1" in note
    p, state = roster.find("退職済", "花子")
    assert state == "ok" and p.emp_id == "2001001" and p.status == "退職"
    p2, _ = roster.find("試験", "太郎")
    assert p2.workplace == "テスト商事株式会社"      # xlsx 側の情報は保持
    # キャッシュのシリアライズ往復
    d = _person_to_dict(api_people[0])
    back = _person_from_dict(d)
    assert back.emp_id == "2001001" and back.birth == dt.date(1970, 2, 2) and back.retired == dt.date(2026, 5, 31)


def test_parse_workdays():
    from services.daicho.fieldglass import parse_workdays
    w, o = parse_workdays("月曜 勤務-火曜 勤務-水曜 勤務-木曜 非勤務-金曜 非勤務-土曜 非勤務-日曜 非勤務")
    assert w == "月 火 水" and o == "木 金 土 日"
    assert parse_workdays("") == ("", "")


def test_sheet_title_unique():
    used: set[str] = set()
    from services.daicho.records import LedgerRecord
    r = LedgerRecord(fields={"社員番号": "2020001", "氏名": "試験 太郎"})
    assert sheet_title(r, used) == "2020001_試験太郎"
    assert sheet_title(r, used) == "2020001_試験太郎_2"
    r2 = LedgerRecord(fields={"社員番号": "", "氏名": "A/B:C"})
    assert sheet_title(r2, used) == "X_ABC"


@pytest.mark.skipif(not config.SOURCE_FORM_XLSM.exists(), reason="旧フォーム（NAS）が無い環境")
def test_template_and_write(tmp_path, inputs):
    tpl = build_template(config.SOURCE_FORM_XLSM, tmp_path / "tpl.xlsx")
    wb = openpyxl.load_workbook(tpl)
    ws = wb[config.TEMPLATE_SHEET]
    assert ws["B5"].value == "派遣先名称" and ws["D5"].value is None          # ラベルは写す・値は写さない
    assert ws["E2"].value is None and ws["D12"].value is None
    assert ws["I2"].value == "健康保険" and ws["B57"].value == EXTRA_ROWS[0][0]
    assert "D5:N5" in {str(r) for r in ws.merged_cells.ranges}
    assert ws.page_setup.fitToHeight == 1 and ws.page_setup.fitToWidth == 1
    for key, coord in CELL_MAP.items():
        assert ws[coord].value is None or key in ("契約期間",), f"{key} → {coord} に値が残っている"

    tc_p, cpi_p, roster_p = inputs
    contracts, _ = load_contracts(tc_p, cpi_p)
    roster = load_roster(roster_p)
    q = (dt.date(2025, 7, 1), dt.date(2025, 9, 30))
    recs = []
    for c in contracts_in_quarter(contracts, *q):
        p, st = roster.find(*c.worker_sei_mei)
        recs.append(build_record(c, p, st, *q))
    paths = write_quarter(recs, "2025Q3", tpl, tmp_path / "out", ["全体テスト"])
    out = openpyxl.load_workbook(paths["xlsx"])
    assert out.sheetnames[:2] == ["目次", "警告"] and len(out.sheetnames) == 4
    s = out["2020001_試験太郎"]
    assert s["D5"].value == "テスト商事株式会社" and s["E17"].value == "2025/07/01～2025/09/30"
    assert s["D24"].value == "月 火 水 木 金" and s["D26"].value == "土曜日 日曜日 祝日"    # 行ズレ修正の確認
    assert s["D32"].value == "部長 責任 花子" and s["D41"].value == "課長 指揮 太郎"         # 責任者/指揮命令者の取り違え修正
    assert s["M3"].value == "無"
    assert out["目次"]["B5"].hyperlink is not None
    rows = list(csv.DictReader(open(paths["csv"], encoding="utf-8-sig")))
    assert len(rows) == 2 and rows[0]["契約No"]
