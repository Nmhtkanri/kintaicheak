# -*- coding: utf-8 -*-
"""健診PDFの読み取り結果 → PersonRecord 変換のテスト

Claude API には触らない。読み取り結果（JSON）を模した dict を入れて、
Excel経路と同じ絶対ルールが守られているかを見る:

  1. 血圧は測定回ごとに別々のまま。平均を作らない。測定回不明なら値ごと捨てる
  2. (-) は陰性として残し、空欄や裸のハイフンを (-) にしない
  3. 原票判定A〜Gは判定欄にだけ入れ、検査値に混ぜない
  4. 監査用Excelを書いて読み直すと、チェッカーがエラー0件で受け取れる
"""

from __future__ import annotations

import os
import sys
from datetime import date

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.health_hpm_excel import parse_health_workbook  # noqa: E402
from services.health_hpm_master import load_master  # noqa: E402
from services.health_hpm_pdf import (  # noqa: E402
    BMI_TOLERANCE,
    PdfReadError,
    analyze_health_pdf,
    build_item_lookup,
    crop_band,
    page_images_for_reading,
    readings_to_person,
    render_pdf_pages,
    write_audit_workbook,
)
from tests.health_hpm_fixtures import make_master_xlsx  # noqa: E402


@pytest.fixture
def item_lookup(tmp_path):
    return build_item_lookup(load_master(make_master_xlsx(tmp_path / "master.xlsx")))


def reading(**overrides):
    """読み取り結果1ページ分。実際にClaudeが返した形に合わせてある。"""
    base = {
        "identity": {"氏名": "友納 英彦", "年齢": 58, "性別": "男性",
                     "受診日": "2026-07-01", "受診No": "132"},
        "blood_pressure": [{"occurrence": 1, "systolic": 132, "diastolic": 86},
                           {"occurrence": 2, "systolic": 118, "diastolic": 72}],
        "metrics": [{"category": "身体計測", "item": "身長", "value": "181.1"},
                    {"category": "身体計測", "item": "体重", "value": "89.3"}],
        "qualitative": [{"category": "尿検査", "item": "尿蛋白", "value": "(-)",
                         "method": None}],
        "judgements": {"身体計測": "B", "血圧": "A"},
        "needs_check": [],
    }
    base.update(overrides)
    return base


def codes(person, level=None):
    return {i.code for i in person.issues if level is None or i.level == level}


def values_of(person, item):
    return [m.value for m in person.metrics if m.item == item]


def metric_of(person, item):
    """血圧の行が先に入るので、項目名で引く（metrics[0] は使わない）。"""
    return next(m for m in person.metrics if m.item == item)


# ---------------------------------------------------------------------------
# 血圧（最重要）
# ---------------------------------------------------------------------------

class TestBloodPressure:
    def test_two_rounds_stay_separate(self, item_lookup):
        person = readings_to_person(reading(), 2, item_lookup)
        assert person.blood_pressure() == {1: {"sys": "132", "dia": "86"},
                                           2: {"sys": "118", "dia": "72"}}

    def test_no_average_value_created(self, item_lookup):
        """132と118の平均125、86と72の平均79がどこにも出てこないこと。"""
        person = readings_to_person(reading(), 2, item_lookup)
        stored = {m.value for m in person.metrics}
        assert "125" not in stored and "79" not in stored

    def test_single_measurement_not_duplicated(self, item_lookup):
        person = readings_to_person(
            reading(blood_pressure=[{"occurrence": 1, "systolic": 120, "diastolic": 61}]),
            2, item_lookup)
        assert person.blood_pressure() == {1: {"sys": "120", "dia": "61"}}
        assert 2 not in person.blood_pressure()
        assert "BP_SINGLE" not in codes(person), "この段階では付けない（analyze側で付ける）"

    def test_occurrence_null_is_error_and_value_dropped(self, item_lookup):
        person = readings_to_person(
            reading(blood_pressure=[{"occurrence": None, "systolic": 120,
                                     "diastolic": 61}]), 2, item_lookup)
        assert "BP_OCCURRENCE_UNKNOWN" in codes(person, "error")
        assert person.blood_pressure() == {}, "測定回が分からない値は取り込まない"

    def test_occurrence_garbage_is_error(self, item_lookup):
        person = readings_to_person(
            reading(blood_pressure=[{"occurrence": "?", "systolic": 120,
                                     "diastolic": 61}]), 2, item_lookup)
        assert "BP_OCCURRENCE_UNKNOWN" in codes(person, "error")
        assert person.blood_pressure() == {}

    def test_only_second_round_is_kept_as_second(self, item_lookup):
        person = readings_to_person(
            reading(blood_pressure=[{"occurrence": 2, "systolic": 118, "diastolic": 72}]),
            2, item_lookup)
        assert person.blood_pressure() == {2: {"sys": "118", "dia": "72"}}

    def test_blank_second_measurement_is_skipped(self, item_lookup):
        person = readings_to_person(
            reading(blood_pressure=[{"occurrence": 1, "systolic": 120, "diastolic": 61},
                                    {"occurrence": 2, "systolic": None,
                                     "diastolic": None}]), 2, item_lookup)
        assert person.blood_pressure() == {1: {"sys": "120", "dia": "61"}}

    def test_judgement_goes_to_judgement_field_only(self, item_lookup):
        person = readings_to_person(reading(), 2, item_lookup)
        bp = [m for m in person.metrics if m.category == "血圧"]
        assert all(m.source_judgement == "A" for m in bp)
        assert all(m.value not in ("A", "B") for m in person.metrics)


# ---------------------------------------------------------------------------
# 識別情報
# ---------------------------------------------------------------------------

class TestIdentity:
    def test_basic_fields(self, item_lookup):
        person = readings_to_person(reading(), 3, item_lookup)
        assert person.key == "p03"
        assert person.sheet == "P03_友納英彦"
        assert person.name == "友納 英彦"
        assert person.age == 58
        assert person.exam_date == date(2026, 7, 1)
        assert person.exam_no == "000132", "受診No.は6桁ゼロ埋め"

    def test_invalid_exam_date(self, item_lookup):
        person = readings_to_person(
            reading(identity={"氏名": "友納 英彦", "年齢": 58, "性別": "男性",
                              "受診日": "よめない", "受診No": "132"}), 2, item_lookup)
        assert "EXAM_DATE_INVALID" in codes(person, "error")

    def test_invalid_exam_no(self, item_lookup):
        person = readings_to_person(
            reading(identity={"氏名": "友納 英彦", "年齢": 58, "性別": "男性",
                              "受診日": "2026-07-01", "受診No": ""}), 2, item_lookup)
        assert "EXAM_NO_MISSING" in codes(person, "error")

    def test_long_name_sheet_is_truncated(self, item_lookup):
        person = readings_to_person(
            reading(identity={"氏名": "あ" * 40, "年齢": 50, "性別": "男性",
                              "受診日": "2026-07-01", "受診No": "1"}), 2, item_lookup)
        assert len(person.sheet) <= 31


# ---------------------------------------------------------------------------
# 項目の正規化
# ---------------------------------------------------------------------------

class TestItems:
    def test_fecal_blood_split_by_occurrence(self, item_lookup):
        person = readings_to_person(reading(qualitative=[
            {"category": "大腸", "item": "便潜血①", "value": "(-)"},
            {"category": "大腸", "item": "便潜血②", "value": "(+)"},
        ]), 2, item_lookup)
        got = {(m.item, m.occurrence): m.value for m in person.qualitative()}
        assert got == {("便潜血", 1): "(-)", ("便潜血", 2): "(+)"}

    def test_category_comes_from_master(self, item_lookup):
        """読み取り側が「大腸」と言っても、分類はマスタの登録が正。"""
        person = readings_to_person(reading(qualitative=[
            {"category": "大腸", "item": "便潜血①", "value": "(-)"}]), 2, item_lookup)
        assert person.qualitative()[0].category == "便潜血"

    def test_urobilinogen_alias(self, item_lookup):
        person = readings_to_person(reading(qualitative=[
            {"category": "尿検査", "item": "ウロビリノーゲン", "value": "(-)"}]),
            2, item_lookup)
        assert person.qualitative()[0].item == "尿ウロビリノーゲン"

    def test_dropped_duplicate_item(self, item_lookup):
        person = readings_to_person(reading(qualitative=[
            {"category": "糖代謝", "item": "尿糖（糖代謝）", "value": "(-)"}]),
            2, item_lookup)
        assert person.qualitative() == [], "別欄の重複印字は取り込まない"

    def test_unknown_item_keeps_reading_category(self, item_lookup):
        person = readings_to_person(reading(metrics=[
            {"category": "謎の検査", "item": "知らない項目", "value": "1.23"}]),
            2, item_lookup)
        metric = metric_of(person, "知らない項目")
        assert metric.category == "謎の検査" and metric.value == "1.23"

    @pytest.mark.parametrize("raw,expect", [("＊291", "291"), ("291＊", "291"),
                                            ("*74", "74"), ("291", "291")])
    def test_asterisk_moved_to_note(self, item_lookup, raw, expect):
        person = readings_to_person(reading(metrics=[
            {"category": "肝機能", "item": "LD", "value": raw}]), 2, item_lookup)
        metric = metric_of(person, "LD")
        assert metric.value == expect
        assert metric.source_note == ("＊" if raw != expect else "")

    def test_note_field_is_kept(self, item_lookup):
        person = readings_to_person(reading(metrics=[
            {"category": "肝機能", "item": "LD", "value": "291", "note": "＊"}]),
            2, item_lookup)
        assert metric_of(person, "LD").source_note == "＊"

    def test_method_kept_in_original_display(self, item_lookup):
        person = readings_to_person(reading(qualitative=[
            {"category": "感染症", "item": "HBs抗原", "value": "(-)", "method": "CLIA"}]),
            2, item_lookup)
        assert "CLIA" in person.qualitative()[0].original_display


# ---------------------------------------------------------------------------
# 定性値
# ---------------------------------------------------------------------------

class TestQualitative:
    @pytest.mark.parametrize("raw", ["(-)", "（－）", "陰性"])
    def test_negative_variants(self, item_lookup, raw):
        person = readings_to_person(reading(qualitative=[
            {"category": "尿検査", "item": "尿蛋白", "value": raw}]), 2, item_lookup)
        assert person.qualitative()[0].value == "(-)"

    def test_plus_minus_preserved(self, item_lookup):
        person = readings_to_person(reading(qualitative=[
            {"category": "尿検査", "item": "尿蛋白", "value": "（±）"}]), 2, item_lookup)
        assert person.qualitative()[0].value == "(±)"

    @pytest.mark.parametrize("raw", ["", None, "   "])
    def test_blank_not_stored(self, item_lookup, raw):
        person = readings_to_person(reading(qualitative=[
            {"category": "尿検査", "item": "尿蛋白", "value": raw}]), 2, item_lookup)
        assert person.qualitative() == [], "空欄を (-) にしない"
        assert person.issues == []

    def test_bare_hyphen_is_error_and_dropped(self, item_lookup):
        person = readings_to_person(reading(qualitative=[
            {"category": "尿検査", "item": "尿蛋白", "value": "-"}]), 2, item_lookup)
        assert "QUAL_BARE_HYPHEN" in codes(person, "error")
        assert person.qualitative() == []


# ---------------------------------------------------------------------------
# BMI照合・needs_check
# ---------------------------------------------------------------------------

class TestChecks:
    def _with_bmi(self, bmi):
        return reading(metrics=[
            {"category": "身体計測", "item": "身長", "value": "163.5"},
            {"category": "身体計測", "item": "体重", "value": "54.3"},
            {"category": "身体計測", "item": "BMI", "value": bmi},
        ])

    def test_bmi_consistent_no_warning(self, item_lookup):
        person = readings_to_person(self._with_bmi("20.3"), 2, item_lookup)
        assert "BMI_MISMATCH" not in codes(person)

    def test_bmi_mismatch_warns(self, item_lookup):
        person = readings_to_person(self._with_bmi("28.2"), 2, item_lookup)
        assert "BMI_MISMATCH" in codes(person, "warning")

    def test_bmi_boundary(self, item_lookup):
        """計算値20.31との差が許容内なら出さない。"""
        person = readings_to_person(self._with_bmi("20.2"), 2, item_lookup)
        assert "BMI_MISMATCH" not in codes(person)
        assert BMI_TOLERANCE == 0.15

    def test_bmi_missing_parts_skips_check(self, item_lookup):
        person = readings_to_person(reading(metrics=[
            {"category": "身体計測", "item": "BMI", "value": "99.9"}]), 2, item_lookup)
        assert "BMI_MISMATCH" not in codes(person)

    def test_needs_check_is_info_only(self, item_lookup):
        person = readings_to_person(reading(needs_check=["ZTT: 空欄", "RF: 空欄"]),
                                    2, item_lookup)
        infos = [i for i in person.issues if i.code == "READ_NEEDS_CHECK"]
        assert len(infos) == 2
        assert all(i.level == "info" for i in infos)


# ---------------------------------------------------------------------------
# 画像化と全体制御
# ---------------------------------------------------------------------------

def make_pdf(path, pages=2):
    """テスト用の白紙PDF（中身は読ませないのでレンダリングできればよい）。"""
    from PIL import Image

    images = [Image.new("RGB", (595, 842), "white") for _ in range(pages)]
    images[0].save(str(path), save_all=True, append_images=images[1:])
    return str(path)


class TestRendering:
    def test_render_pages(self, tmp_path):
        pngs = render_pdf_pages(make_pdf(tmp_path / "a.pdf", 3), dpi=72)
        assert len(pngs) == 3
        assert all(p[:8] == b"\x89PNG\r\n\x1a\n" for p in pngs)

    def test_render_rejects_unreadable(self, tmp_path):
        bad = tmp_path / "bad.pdf"
        bad.write_bytes(b"not a pdf")
        with pytest.raises(PdfReadError):
            render_pdf_pages(str(bad))

    def test_render_rejects_too_many_pages(self, tmp_path, monkeypatch):
        from config import Config
        monkeypatch.setattr(Config, "HEALTH_HPM_PDF_MAX_PAGES", 2)
        with pytest.raises(PdfReadError) as e:
            render_pdf_pages(make_pdf(tmp_path / "many.pdf", 3), dpi=72)
        assert "分けて" in str(e.value)

    def test_crop_band_height(self, tmp_path):
        from PIL import Image
        png = render_pdf_pages(make_pdf(tmp_path / "a.pdf", 1), dpi=72)[0]
        full = Image.open(__import__("io").BytesIO(png))
        band = Image.open(__import__("io").BytesIO(crop_band(png, 0.10, 0.55)))
        assert band.width == full.width
        assert abs(band.height - int(full.height * 0.45)) <= 1

    def test_three_images_are_sent(self, tmp_path):
        png = render_pdf_pages(make_pdf(tmp_path / "a.pdf", 1), dpi=72)[0]
        assert len(page_images_for_reading(png)) == 3


class TestAnalyze:
    def _run(self, tmp_path, monkeypatch, readings, pages=2):
        import services.health_hpm_pdf as mod
        master = load_master(make_master_xlsx(tmp_path / "master.xlsx"))
        monkeypatch.setattr(mod, "render_pdf_pages",
                            lambda path, dpi=200: [b"png"] * pages)
        monkeypatch.setattr(mod, "page_images_for_reading", lambda png: [png])
        calls = {"n": 0}

        def fake_call(client, images):
            calls["n"] += 1
            value = readings[calls["n"] - 1]
            if isinstance(value, Exception):
                raise value
            return value

        monkeypatch.setattr(mod, "_call_claude_for_page", fake_call)
        events = list(analyze_health_pdf("dummy.pdf", master,
                                         source_filename="健診.pdf", client=object()))
        return events

    def test_progress_then_result(self, tmp_path, monkeypatch):
        events = self._run(tmp_path, monkeypatch,
                           [reading(), reading(identity={
                               "氏名": "高橋 和紀", "年齢": 47, "性別": "男性",
                               "受診日": "2026-07-01", "受診No": "186"})])
        kinds = [k for k, _ in events]
        assert kinds[-1] == "result"
        assert kinds.count("progress") >= 3
        result = events[-1][1]
        assert [p.name for p in result.parse.persons] == ["友納 英彦", "高橋 和紀"]
        assert result.pages == {"p01": 1, "p02": 2}
        assert set(result.page_pngs) == {1, 2}

    def test_schema_flags_and_no_genpyo_error(self, tmp_path, monkeypatch):
        result = self._run(tmp_path, monkeypatch, [reading(), reading()])[-1][1]
        parse = result.parse
        assert parse.schema_version == "2.0"
        assert parse.genpyo_confirmed is False
        assert parse.bp_occurrence_kept is True and parse.qualitative_kept is True
        assert "GENPYO_FLAG_MISSING" not in {i.code for i in parse.errors()}, \
            "画面のチェックが確認ゲートなので、ここでは出さない"

    def test_blood_pressure_warnings_match_excel_route(self, tmp_path, monkeypatch):
        single = reading(blood_pressure=[{"occurrence": 1, "systolic": 120,
                                          "diastolic": 61}])
        result = self._run(tmp_path, monkeypatch, [single, single])[-1][1]
        assert "BP_SINGLE" in {i.code for i in result.parse.warnings()}

    def test_non_person_page_is_skipped(self, tmp_path, monkeypatch):
        cover = {"identity": {}, "metrics": [], "qualitative": [], "needs_check": []}
        result = self._run(tmp_path, monkeypatch, [cover, reading()])[-1][1]
        assert len(result.parse.persons) == 1
        assert result.pages == {"p02": 2}

    def test_page_failure_continues(self, tmp_path, monkeypatch):
        from services.health_hpm_pdf import PageReadError
        result = self._run(tmp_path, monkeypatch,
                           [PageReadError("読めない"), reading()])[-1][1]
        assert "PDF_PAGE_READ_FAILED" in {i.code for i in result.parse.errors()}
        assert len(result.parse.persons) == 1, "残りのページは読み続ける"

    def test_no_persons_is_error(self, tmp_path, monkeypatch):
        cover = {"identity": {}, "metrics": [], "qualitative": [], "needs_check": []}
        result = self._run(tmp_path, monkeypatch, [cover, cover])[-1][1]
        assert "PDF_NO_PERSONS" in {i.code for i in result.parse.errors()}


# ---------------------------------------------------------------------------
# 監査用Excel
# ---------------------------------------------------------------------------

class TestAuditWorkbook:
    def _build(self, tmp_path, item_lookup, readings=None):
        from services.health_hpm_excel import WorkbookParseResult, _check_blood_pressure
        readings = readings or [reading()]
        parse = WorkbookParseResult(schema_version="2.0", genpyo_confirmed=False,
                                    bp_occurrence_kept=True, qualitative_kept=True,
                                    source_filename="健診.pdf")
        pages = {}
        for page_no, raw in enumerate(readings, start=1):
            person = readings_to_person(raw, page_no, item_lookup)
            _check_blood_pressure(person)
            parse.persons.append(person)
            pages[person.key] = page_no
        pngs = {n: _tiny_png() for n in pages.values()}
        return parse, pages, pngs

    def test_roundtrip_parses_without_errors(self, tmp_path, item_lookup):
        parse, pages, pngs = self._build(tmp_path, item_lookup)
        out = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                   pdf_name="健診.pdf", confirmed_at="2026-08-10 13:00")

        back = parse_health_workbook(out, os.path.basename(out))
        assert back.errors() == [], [i.message for i in back.errors()]
        assert back.schema_version == "2.0"
        assert back.genpyo_confirmed is True
        assert len(back.persons) == len(parse.persons)

    def test_roundtrip_keeps_values(self, tmp_path, item_lookup):
        parse, pages, pngs = self._build(tmp_path, item_lookup)
        out = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                   pdf_name="健診.pdf", confirmed_at="2026-08-10 13:00")
        back = parse_health_workbook(out)
        before, after = parse.persons[0], back.persons[0]

        assert after.name == before.name
        assert after.exam_no == "000132"
        assert after.exam_date == before.exam_date
        assert after.blood_pressure() == before.blood_pressure()
        assert {(m.item, m.value) for m in after.qualitative()} == \
               {(m.item, m.value) for m in before.qualitative()}

    def test_confirmed_stamp_recorded(self, tmp_path, item_lookup):
        import openpyxl
        parse, pages, pngs = self._build(tmp_path, item_lookup)
        out = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                   pdf_name="健診.pdf", confirmed_at="2026-08-10 13:00")
        wb = openpyxl.load_workbook(out)
        info = {row[0]: row[1] for row in wb["変換案内"].iter_rows(values_only=True)
                if row and row[0]}
        wb.close()
        assert info["原票確認済み"] == "TRUE"
        assert "2026-08-10 13:00" in info["確認方法"]
        assert "チェッカー画面" in info["確認方法"]

    def test_page_images_embedded(self, tmp_path, item_lookup):
        import openpyxl
        parse, pages, pngs = self._build(tmp_path, item_lookup)
        out = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                   pdf_name="健診.pdf", confirmed_at="2026-08-10 13:00")
        wb = openpyxl.load_workbook(out)
        assert "P01_友納英彦" in wb.sheetnames
        assert len(wb["P01_友納英彦"]._images) == 1
        wb.close()

    def test_needs_check_written_to_memo(self, tmp_path, item_lookup):
        import openpyxl
        parse, pages, pngs = self._build(tmp_path, item_lookup,
                                         [reading(needs_check=["ZTTが空欄"])])
        out = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                   pdf_name="健診.pdf", confirmed_at="2026-08-10 13:00")
        wb = openpyxl.load_workbook(out)
        text = " ".join(str(r[0]) for r in wb["要確認メモ"].iter_rows(values_only=True) if r[0])
        wb.close()
        assert "ZTTが空欄" in text

    def test_existing_file_gets_suffix(self, tmp_path, item_lookup):
        parse, pages, pngs = self._build(tmp_path, item_lookup)
        first = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                     pdf_name="健診.pdf", confirmed_at="t")
        before = open(first, "rb").read()
        second = write_audit_workbook(str(tmp_path / "audit.xlsx"), parse, pages, pngs,
                                      pdf_name="健診.pdf", confirmed_at="t")
        assert second.endswith("audit_2.xlsx")
        assert open(first, "rb").read() == before, "既存ファイルを壊さない"


def _tiny_png() -> bytes:
    import io as _io

    from PIL import Image
    buf = _io.BytesIO()
    Image.new("RGB", (60, 80), "white").save(buf, format="PNG")
    return buf.getvalue()
