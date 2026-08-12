# -*- coding: utf-8 -*-
"""経費統合一覧表の生成（マクロ移植 P1a）のユニットテスト。

各ソース変換・SAP税抜補正・顧客請求分の寄せ・経路突合の判定・Excel出力を検証する。
実データ（Rev5）との全行一致は別途 谷津さん提供の突合セットで確認する。
"""
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.keihi_summary import (  # noqa: E402
    excel_coerce, norm_date_slash, norm_date_pad, normkey, parse_amount,
    _round_half_up, is_night_duty_text, in_company_scope,
    transform_jinjer, transform_estaffing, transform_sap, transform_freee,
    evaluate_route_check, add_integrated_sheet, run_keihi_integration,
    build_integrated_rows, add_roster_entry,
    INTEGRATED_HEADERS, NCOL,
    C_EMP, C_NAME, C_APPDATE, C_TOTAL, C_MEMO_REQ, C_USEDATE, C_TRANS, C_DETAIL,
    C_BILLTYPE, C_FARE, C_AMOUNT, C_SUBTOTAL, C_MEMO_LINE, C_ENTRY_TYPE,
    C_BOARD, C_ALIGHT, C_CUSTOMER_BILL,
)


# ----------------------------------------------------------------------
# ヘルパー関数
# ----------------------------------------------------------------------

def test_excel_coerce_strips_leading_zeros_and_dates():
    # 仕訳No. の先頭ゼロは Excel 取込で落ちる
    assert excel_coerce("00000236") == "236"
    # 日付はゼロ埋めなしで再表示
    assert excel_coerce("2026/06/01") == "2026/6/1"
    assert excel_coerce("2026/6/30") == "2026/6/30"
    # 社員番号・金額はそのまま
    assert excel_coerce("2026013") == "2026013"
    assert excel_coerce("356") == "356"
    # 8桁の計上日(yyyymmdd) は数値化されてもそのまま
    assert excel_coerce("20260630") == "20260630"
    # 文字列はそのまま
    assert excel_coerce("交通費（電車・バス）") == "交通費（電車・バス）"
    assert excel_coerce("") == ""


def test_norm_dates():
    assert norm_date_slash("2026/06/01") == "2026/6/1"
    assert norm_date_slash("2026-6-1") == "2026/6/1"
    assert norm_date_pad("2026/6/1") == "2026/06/01"
    assert norm_date_pad("2026/4/30 12:17") == "2026/04/30"   # 時刻切捨て


def test_normkey_and_night_duty():
    assert normkey(" 山田　太郎 ") == "山田太郎"
    assert is_night_duty_text("顧客対応当番16") is True
    assert is_night_duty_text("深夜出動手当") is True
    assert is_night_duty_text("交通費") is False


def test_excel_coerce_jp_date():
    from services.keihi_summary import excel_coerce_jp_date
    # Excel が OpenText で「6月5日」を日付化する挙動（年は year_hint から）
    assert excel_coerce_jp_date("6月5日", "2026/06/05") == "2026/06/05"
    assert excel_coerce_jp_date("6/17", "2026/06/17") == "2026/06/17"
    assert excel_coerce_jp_date("2026/6/4", "") == "2026/06/04"
    # 日付でない説明は原文そのまま（前後空白も保持）
    assert excel_coerce_jp_date("東京から潮見の往復", "2026/06/13") == "東京から潮見の往復"
    assert excel_coerce_jp_date("その他の費用(10%) 6/4 ", "2026/06/04") == "その他の費用(10%) 6/4 "
    assert excel_coerce_jp_date("", "2026/06/04") == ""


def test_parse_amount_and_round():
    assert parse_amount("1,375") == "1375"
    assert parse_amount("￥2,750円") == "2750"
    assert parse_amount("") == ""
    # SAP 税抜（÷1.1・四捨五入）
    assert _round_half_up(2750 / 1.1) == 2500
    assert _round_half_up(1375 / 1.1) == 1250
    assert _round_half_up(5500 / 1.1) == 5000


def test_in_company_scope():
    assert in_company_scope("2026013") is True
    assert in_company_scope("5000001") is False   # 派遣
    assert in_company_scope("9000001") is False
    assert in_company_scope("") is False


# ----------------------------------------------------------------------
# ソース別変換
# ----------------------------------------------------------------------

def _jinjer_row(**kw):
    r = [""] * 33
    r[C_EMP] = kw.get("emp", "2026013")
    r[C_NAME] = kw.get("name", "川口 祐ノ輔")
    r[C_TOTAL] = kw.get("total", "356")
    r[C_USEDATE] = kw.get("use", "2026/06/01")
    r[C_TRANS] = kw.get("trans", "交通費（電車・バス）")
    r[C_DETAIL] = kw.get("detail", "")
    r[C_BILLTYPE] = kw.get("billtype", "")
    return r


def test_transform_jinjer_identity_and_coerce():
    rows = transform_jinjer([], [_jinjer_row(use="2026/06/01", total="356")])
    assert len(rows) == 1
    row = rows[0]
    assert row[C_EMP] == "2026013"
    assert row[C_USEDATE] == "2026/6/1"       # ゼロ埋め落ち
    assert row[C_TOTAL] == "356"
    assert row[C_CUSTOMER_BILL] == ""          # 通常行は 34列目 空


def test_transform_jinjer_moves_customer_bill_to_col34():
    # 請求区分=顧客請求・夜間当番でない → 金額を 34列目へ、合計等はクリア
    r = _jinjer_row(billtype="顧客請求", total="8990", detail="客先請求分（交通費）")
    row = transform_jinjer([], [r])[0]
    assert row[C_CUSTOMER_BILL] == "8990"
    assert row[C_TOTAL] == ""
    assert row[C_SUBTOTAL] == ""
    assert row[C_FARE] == ""
    assert row[C_AMOUNT] == ""


def test_transform_jinjer_night_duty_customer_bill_stays():
    # 請求区分=顧客請求 でも夜間当番系なら 34列目へ寄せない（合計に残す）
    r = _jinjer_row(billtype="顧客請求", total="2500", detail="夜間当番手当")
    row = transform_jinjer([], [r])[0]
    assert row[C_CUSTOMER_BILL] == ""
    assert row[C_TOTAL] == "2500"


def test_transform_estaffing_all_to_customer_bill():
    # 立替金CSV 15列: [0]TC番号...[4]スタッフ名 [5]就業年月日 [8]出発地 [9]到着地 [10]交通手段 ...[14]金額
    r = [""] * 15
    r[4], r[5], r[8], r[9], r[10], r[14] = "衣笠 博陽", "2026/6/3", "大阪", "大津", "JR西日本", "960"
    roster = {}
    add_roster_entry(roster, "衣笠 博陽", "2021001")
    row = transform_estaffing([], [r], roster)[0]
    assert row[C_EMP] == "2021001"
    assert row[C_NAME] == "衣笠 博陽"
    assert row[C_TRANS] == "JR西日本"
    assert row[C_BOARD] == "大阪" and row[C_ALIGHT] == "大津"
    assert row[C_AMOUNT] == "960" and row[C_FARE] == "960"
    assert row[C_CUSTOMER_BILL] == "960"       # e-staffing は全額客先請求


def _sap_row(sei="太田", mei="裕一", amt="398", vendor="京葉線", desc="", cc="", sid=""):
    r = [""] * 13
    r[0], r[1], r[2], r[3], r[5], r[8], r[10] = sei, mei, amt, vendor, desc, cc, sid
    r[4] = "2026/6/13"   # 費用エントリ日
    r[6] = "2026/6/30"   # 承認日
    return r


_SAP_HEADERS_13 = ["姓", "名", "費用合計", "業者名", "費用エントリ日", "説明",
                   "費用シート承認日", "事業単位", "コストセンター", "通貨",
                   "費用シート ID", "勤務地", "費用シートのステータス"]
# 2026-08-06 SAP側がAPI連携用に6列追加（13列→19列）。増えたのは11列目以降だったので
# 位置決め打ちのままでも壊れなかったが、前の方に入ったら黙って別の列を読んでしまう。
_SAP_HEADERS_19 = ["姓", "名", "費用合計", "業者名", "費用エントリ日", "説明",
                   "費用シート承認日", "事業単位", "コストセンター", "通貨",
                   "費用シート ID", "費用コード", "費用名", "勤務地",
                   "費用シートのステータス", "スタッフ ID", "費用シート提出日",
                   "発注者", "作業オーダー ID"]


class TestTransformSapColumnLayout:
    """SAP生CSVの列が増減・並べ替えされても壊れないこと（列名で位置を決める）"""

    def _base(self):
        return transform_sap(_SAP_HEADERS_13, [_sap_row(cc="CC1", sid="EXP-1")], {})

    def test_19col_layout_matches_13col(self):
        """新19列レイアウト（末尾寄りに6列追加）でも結果が変わらない"""
        r13 = _sap_row(cc="CC1", sid="EXP-1")
        # 11列目までは同じ、以降に 費用コード/費用名 を差し込み 勤務地・ステータスが後ろへ
        r19 = r13[:11] + ["CODE", "費用名X"] + [r13[11], r13[12]] + ["S-1", "2026/6/28", "発注者A", "WO-1"]
        assert transform_sap(_SAP_HEADERS_19, [r19], {}) == self._base()

    def test_column_inserted_at_front_does_not_shift_data(self):
        """先頭に列が入っても列名で引くのでズレない（旧実装ならここで全部ズレた）"""
        r13 = _sap_row(cc="CC1", sid="EXP-1")
        got = transform_sap(["作業オーダー ID"] + _SAP_HEADERS_13, [["WO-1"] + r13], {})
        assert got == self._base()

    def test_reordered_columns(self):
        """並べ替えにも耐える"""
        order = [10, 0, 1, 8, 2, 3, 4, 5, 6, 7, 9, 11, 12]
        headers = [_SAP_HEADERS_13[i] for i in order]
        r13 = _sap_row(cc="CC1", sid="EXP-1")
        assert transform_sap(headers, [[r13[i] for i in order]], {}) == self._base()

    def test_falls_back_to_positions_without_header_names(self):
        """ヘッダー名が読めないCSVは従来の13列レイアウトの位置で読む（後方互換）"""
        r13 = _sap_row(cc="CC1", sid="EXP-1")
        assert transform_sap([], [r13], {}) == self._base()
        assert transform_sap([""] * 13, [r13], {}) == self._base()

    def test_approved_date_alias(self):
        """承認日の列名が「承認日」でも「費用シート承認日」でも同じ列を読む"""
        alias = list(_SAP_HEADERS_13)
        alias[6] = "承認日"
        r13 = _sap_row(cc="CC1", sid="EXP-1")
        assert transform_sap(alias, [r13], {}) == self._base()


def test_transform_sap_normal_to_customer_bill():
    row = transform_sap([], [_sap_row(vendor="京葉線", amt="398")], {})[0]
    assert row[C_DETAIL] == "京葉線"
    assert row[C_CUSTOMER_BILL] == "398"
    assert row[C_TOTAL] == ""                 # 通常行は D:合計 に入れない


def test_transform_sap_night_duty_tax_strip():
    # 業者名=顧客対応当番16・費用合計=2750（税込）→ ÷1.1=2500 を D:合計へ
    roster = {}
    add_roster_entry(roster, "大北 将司", "2024008")
    r = _sap_row(sei="将司", mei="大北", amt="2750", vendor="顧客対応当番16")
    row = transform_sap([], [r], roster)[0]
    assert row[C_TOTAL] == "2500"             # 税抜補正＋夜間当番→合計
    assert row[C_CUSTOMER_BILL] == ""
    assert row[C_EMP] == "2024008"            # 姓名逆でも照合


def test_transform_sap_8a_copies_desc_to_vendor():
    # 夜間当番KWが説明(F)のみ・業者名(D)空 → 8aでD←F、8bで税抜、夜間当番判定も効く
    r = _sap_row(sei="小池", mei="裕也", amt="2750", vendor="", desc="6/12 顧客当番16")
    row = transform_sap([], [r], {})[0]
    assert row[C_DETAIL] == "6/12 顧客当番16"   # 内訳=業者名(=説明の転記)
    assert row[C_TOTAL] == "2500"


def test_transform_sap_memo_coerces_date_desc():
    # 説明が「6月5日」の行 → Excel日付化を再現し 備考(明細)は費用エントリ日で始まる
    r = _sap_row(sei="将司", mei="大北", amt="2750", vendor="顧客対応当番16",
                 desc="6月5日", cc="ＵＡＬＨＷ保守全般", sid="36828ES00003651")
    row = transform_sap([], [r], {})[0]
    # 費用エントリ日=2026/6/13 → 年は2026、6月5日→2026/06/05
    assert row[C_MEMO_LINE] == "2026/06/05 / CC: ＵＡＬＨＷ保守全般 / ID: 36828ES00003651"
    # 説明が日付でない通常行は原文（前後空白保持）
    r2 = _sap_row(vendor="京葉線", amt="398", desc="東京から潮見の往復 ", cc="X", sid="Y")
    row2 = transform_sap([], [r2], {})[0]
    assert row2[C_MEMO_LINE] == "東京から潮見の往復  / CC: X / ID: Y"


def test_transform_freee_preserves_newline_in_memo():
    # freee 内容の埋め込み改行はそのまま保持する（マクロ CStr 相当）
    head = [""] * 32
    head[2], head[3], head[12], head[13] = "2026/7/3", "稲場　直哉", "6月分", "199"
    head[18], head[22], head[23], head[25], head[29] = "2026/6/22", "交通費", "入：新橋駅\n出：品川駅", "199", ""
    rows = transform_freee([], [head], roster={})
    assert rows[0][C_MEMO_LINE] == "入：新橋駅\n出：品川駅"
    assert rows[0][C_USEDATE] == "2026/06/22"


def _freee_rows():
    # 32列。前方フィル対象: 申請日(2) 申請者(3) 申請タイトル(12) 合計金額(13)
    head = [""] * 32
    head[2], head[3], head[12], head[13] = "2026/7/2", "kousei.shiokawa@nmht.co.jp", "6月_経費精算", "2958"
    head[18], head[22], head[23], head[25], head[29] = "2026/6/30", "会議接待費", "森田との面談", "1100", "承認願います"
    cont = [""] * 32
    cont[18], cont[22], cont[23], cont[25] = "2026/6/26", "交通費", "UAL定例会議", "448"
    return [head, cont]


def test_transform_freee_forward_fill_and_mapping():
    roster = {}
    add_roster_entry(roster, "塩川 浩生", "2019047")
    rows = transform_freee([], _freee_rows(), roster)
    assert len(rows) == 2
    r0, r1 = rows
    # 名称変換（メール→氏名）＋社員番号照合
    assert r0[C_NAME] == "塩川 浩生" and r0[C_EMP] == "2019047"
    # 内訳=経費科目、D=金額、仕訳区分=本社経費、利用日=日付（Append_本社経費 標準経路）
    assert r0[C_DETAIL] == "会議接待費" and r0[C_TOTAL] == "1100"
    assert r0[C_ENTRY_TYPE] == "本社経費"
    assert r0[C_USEDATE] == "2026/06/30"       # F:利用日 ← 日付(ゼロ埋め)
    assert r0[C_MEMO_REQ] == "6月_経費精算"
    assert "森田との面談" in r0[C_MEMO_LINE] and "承認願います" in r0[C_MEMO_LINE]
    # 2行目は申請者/申請日/タイトルが前方フィルされる
    assert r1[C_NAME] == "塩川 浩生"
    assert r1[C_DETAIL] == "交通費" and r1[C_TOTAL] == "448"


def test_freee_unmatched_name_marks_該当なし():
    rows = transform_freee([], _freee_rows(), roster={})
    assert rows[0][C_EMP] == "該当なし"


# ----------------------------------------------------------------------
# 経路突合チェック
# ----------------------------------------------------------------------

def _integrated_transit(emp, board, alight, kikan):
    row = [""] * NCOL
    row[C_EMP], row[C_NAME] = emp, "テスト 太郎"
    row[C_USEDATE] = "2026/6/1"
    row[C_TRANS] = kikan
    row[C_BOARD], row[C_ALIGHT] = board, alight
    row[C_TOTAL] = "300"
    return row


def test_route_check_flags_on_route_non_commute():
    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    rows = [
        _integrated_transit("2020001", "東京", "品川", "交通費（電車・バス）"),  # ★経路内なのに通勤系以外
        _integrated_transit("2020001", "東京", "品川", "通勤定期代"),            # OK 通勤系
        _integrated_transit("2020001", "渋谷", "新宿", "通勤定期代"),            # △通勤系だが経路外
    ]
    res = evaluate_route_check(rows, commute)
    verdicts = [r["判定"] for r in res]
    assert verdicts[0].startswith("★")
    assert verdicts[1].startswith("OK")
    assert verdicts[2].startswith("△")


def test_route_check_skips_non_transit_rows():
    # 交通機関も乗降車も無い行（夜間当番手当等）はスキップ
    row = [""] * NCOL
    row[C_EMP], row[C_DETAIL], row[C_TOTAL] = "2020001", "夜間当番手当", "2500"
    assert evaluate_route_check([row], []) == []


def test_route_check_travel_members():
    """移動交通費（立替精算）対象者は経路の有無によらず立替精算が正。"""
    travel = {"2018017": "中村 淳一"}
    rows = [
        # 対象者が立替系で申請 → OK（通勤経路登録なしでも △ にしない）
        _integrated_transit("2018017", "東京", "品川", "交通費（電車・バス）"),
        # 対象者が通勤系を選択 → △ で知らせる
        _integrated_transit("2018017", "東京", "品川", "通勤定期代"),
        # 対象外の人は従来どおり（経路未登録＋通勤系 → △）
        _integrated_transit("2020001", "東京", "品川", "通勤定期代"),
    ]
    res = evaluate_route_check(rows, [], travel)
    assert res[0]["判定"] == "OK（移動交通費対象者＝立替精算）"
    assert res[1]["判定"] == "△逆要確認（移動交通費（立替精算）対象者が通勤系を選択）"
    assert res[2]["判定"] == "△逆要確認（通勤系なのに登録経路と不一致）"


def test_route_check_travel_member_overrides_on_route_flag():
    """対象者は経路内一致でも ★（経路内なのに通勤系以外）にしない。"""
    commute = [{"社員番号": "2018017", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    rows = [_integrated_transit("2018017", "東京", "品川", "交通費（電車・バス）")]
    res = evaluate_route_check(rows, commute, {"2018017": "中村 淳一"})
    assert res[0]["判定"] == "OK（移動交通費対象者＝立替精算）"
    assert res[0]["一致"] == "経路内"   # 一致欄は事実をそのまま残す


# ----------------------------------------------------------------------
# 経路突合レビュー（人が行ごとに計上先を選ぶ二段階実行・2026-08-07）
# ----------------------------------------------------------------------

def _review_rows():
    """★1行・△1行・OK1行・対象者の△1行 の経路突合結果を作る。"""
    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    rows = [
        _integrated_transit("2020001", "東京", "品川", "交通費（電車・バス）"),  # ★
        _integrated_transit("2020001", "東京", "品川", "通勤定期代"),            # OK
        _integrated_transit("2020001", "渋谷", "新宿", "通勤定期代"),            # △
        _integrated_transit("2018017", "東京", "品川", "通勤定期代"),            # △（対象者）
    ]
    return evaluate_route_check(rows, commute, {"2018017": "中村 淳一"})


def test_route_choice_keys_cover_only_reviewable_rows():
    """★△だけにキーを振り、移動交通費対象者の行は人に選ばせない（自動計上のため）。"""
    from services.keihi_summary import build_route_choice_keys
    pairs = build_route_choice_keys(_review_rows(), {"2018017": "中村 淳一"})
    assert [r["判定"][0] for _k, r in pairs] == ["★", "△"]
    assert all("2018017" not in k for k, _r in pairs)


def test_route_choice_keys_number_duplicate_rows():
    """同じ内容の行が並んでも連番で区別でき、何度呼んでも同じキーになる。"""
    from services.keihi_summary import build_route_choice_keys
    rows = [_integrated_transit("2020001", "渋谷", "新宿", "通勤定期代")] * 2
    res = evaluate_route_check(rows, [])
    keys1 = [k for k, _r in build_route_choice_keys(res)]
    keys2 = [k for k, _r in build_route_choice_keys(res)]
    assert keys1 == keys2
    assert keys1[0].endswith("|1") and keys1[1].endswith("|2")
    assert len(set(keys1)) == 2


def test_match_route_choices_accepts_exact_set():
    from services.keihi_summary import build_route_choice_keys, match_route_choices
    res = _review_rows()
    pairs = build_route_choice_keys(res, {"2018017": "中村 淳一"})
    choices = [{"key": k, "choice": "通勤費"} for k, _r in pairs]
    matched, errors = match_route_choices(res, choices, {"2018017": "中村 淳一"})
    assert errors == []
    assert [c for _r, c in matched] == ["通勤費", "通勤費"]


def test_match_route_choices_rejects_any_mismatch():
    """過不足・不正な選択・重複はすべてエラー。黙って一部だけ反映しない。"""
    from services.keihi_summary import build_route_choice_keys, match_route_choices
    res = _review_rows()
    travel = {"2018017": "中村 淳一"}
    pairs = build_route_choice_keys(res, travel)
    full = [{"key": k, "choice": "通勤費"} for k, _r in pairs]

    for label, choices in (
        ("不足", full[:1]),
        ("過剰", full + [{"key": "9999999|2026/6/1|通勤定期代|300||　|1", "choice": "通勤費"}]),
        ("不正な値", [{"key": c["key"], "choice": "その他"} for c in full]),
        ("重複", full + [full[0]]),
    ):
        matched, errors = match_route_choices(res, choices, travel)
        assert errors, f"{label} がエラーになっていません"
        assert matched == []


def test_route_preview_defaults_and_writes_no_file(tmp_path):
    """プレビューは既定＝システムの判定結果で返し、ファイルを1つも作らない。"""
    import csv
    from services.keihi_summary import run_keihi_route_preview
    jcsv = tmp_path / "jinjer.csv"
    with open(jcsv, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(["申請者社員番号", "申請者名"] + [f"c{i}" for i in range(31)])
        # ★経路内なのに通勤系以外 → 分類は「I:非課税精算」＝移動交通費側
        w.writerow(_jinjer_row(emp="2020001", trans="交通費（電車・バス）", total="300"))
        # △通勤系だが経路外 → 分類は「H:交通費」＝通勤費側
        w.writerow(_jinjer_row(emp="2020001", trans="通勤定期代", total="8000",
                               use="2026/07/01"))
    before = sorted(p.name for p in tmp_path.iterdir())

    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    # 乗降場所は CSV に無いので経路突合は「通勤経路登録なし」→ △ になる行だけ拾う
    res = run_keihi_route_preview(jinjer_csv=jcsv, log_func=lambda m: None,
                                  commute_rows=commute, travel_members={}, roster={})
    assert res.ok is True
    assert sorted(p.name for p in tmp_path.iterdir()) == before   # ファイルを作らない

    by_kikan = {r["交通機関"]: r for r in res.review_rows}
    assert by_kikan["通勤定期代"]["計上先"] == "通勤費"
    assert by_kikan["通勤定期代"]["付替可"] is True
    # △（通勤系なのに登録経路と不一致）＝システムは通勤でないとみている → 移動交通費
    assert by_kikan["通勤定期代"]["判定"].startswith("△")
    assert by_kikan["通勤定期代"]["既定選択"] == "移動交通費"
    assert by_kikan["通勤定期代"]["移動額"] == 8000


def test_default_route_choice_follows_system_verdict():
    """プルダウンの初期値はシステム判定に合わせる（2026-08-07 谷津さん要望）。

    ★は毎回135行あり、現状維持を初期値にすると全部選び直しになっていた。
    """
    from services.keihi_summary import default_route_choice
    # ★経路内なのに通勤系以外 → 通勤費が妥当（現状の計上先が移動交通費でも通勤費を既定に）
    assert default_route_choice("★要確認（経路内なのに通勤系以外を選択）", "移動交通費", True) == "通勤費"
    # △通勤系なのに登録経路と不一致 → 通勤ではなさそう（現状が通勤費でも移動交通費を既定に）
    assert default_route_choice("△逆要確認（通勤系なのに登録経路と不一致）", "通勤費", True) == "移動交通費"
    # 計上先を動かせない行は対象外で固定
    assert default_route_choice("★要確認（経路内なのに通勤系以外を選択）",
                                "対象外（顧客請求分）", False) == "対象外"


# ----------------------------------------------------------------------
# 車通勤は金額一致でOK（2026-08-12 谷津さん決定）
# ----------------------------------------------------------------------

def _car_commute(emp="2007001", amount=20000, interval="毎月"):
    return {"社員番号": emp, "出発": "", "到着": "", "経由1": "", "経由2": "",
            "通勤経路": "", "利用交通機関": "車", "支給間隔": interval, "支給金額": amount}


def test_car_commuter_with_matching_amount_is_ok_and_not_reviewable():
    """車通勤は駅の突合ができず毎月△に落ちていた。金額がマスタと合えばOKで一覧に出さない。"""
    from services.keihi_summary import build_route_choice_keys
    commute = [_car_commute()]
    row = _integrated_transit("2007001", "", "", "通勤定期代")
    row[C_TOTAL] = "20000"
    res = evaluate_route_check([row], commute)
    assert res[0]["判定"] == "OK（車通勤・マスタ金額一致）"
    assert build_route_choice_keys(res) == []


def test_car_commuter_with_wrong_amount_stays_reviewable():
    """金額がおかしい車通勤者は従来どおり△で出す（見えなくしない）。"""
    commute = [_car_commute(amount=20000)]
    row = _integrated_transit("2007001", "", "", "通勤定期代")
    row[C_TOTAL] = "12345"
    res = evaluate_route_check([row], commute)
    assert res[0]["判定"].startswith("△")


def test_car_commuter_matches_the_sum_of_multiple_legs():
    """乗継で経路が分かれている場合は合算も一致とみなす。"""
    commute = [_car_commute(amount=12000), _car_commute(amount=8000)]
    row = _integrated_transit("2007001", "", "", "通勤定期代")
    row[C_TOTAL] = "20000"
    res = evaluate_route_check([row], commute)
    assert res[0]["判定"] == "OK（車通勤・マスタ金額一致）"


def test_car_commuter_non_commute_application_keeps_current_behavior():
    """車通勤の非通勤系申請は従来どおり（経路外＝もともとOK）。金額ルールは通勤系だけ。"""
    commute = [_car_commute()]
    row = _integrated_transit("2007001", "", "", "交通費（電車・バス）")
    row[C_TOTAL] = "20000"
    res = evaluate_route_check([row], commute)
    assert res[0]["判定"] == "OK（経路外）"


def test_car_rule_does_not_swallow_star_for_mixed_mode_people():
    """車＋公共の両方を持つ人が登録経路内を非通勤系で申請 → ★のまま。"""
    commute = [_car_commute(amount=20000),
               {"社員番号": "2007001", "出発": "東京", "到着": "品川", "経由1": "",
                "経由2": "", "通勤経路": "", "利用交通機関": "公共交通機関",
                "支給間隔": "毎月", "支給金額": 15000}]
    row = _integrated_transit("2007001", "東京", "品川", "交通費（電車・バス）")
    row[C_TOTAL] = "300"
    res = evaluate_route_check([row], commute)
    assert res[0]["判定"].startswith("★")


def test_car_rule_works_with_csv_fallback_rows_without_interval():
    """CSVフォールバック由来の行は支給間隔が空。それでも金額突合が働く。"""
    commute = [_car_commute(interval="")]
    row = _integrated_transit("2007001", "", "", "通勤交通費（実費）")
    row[C_TOTAL] = "20000"
    res = evaluate_route_check([row], commute)
    assert res[0]["判定"] == "OK（車通勤・マスタ金額一致）"


def test_car_travel_member_precedence_is_unchanged():
    """移動交通費対象者は車通勤でも従来どおり対象者ルールが先に効く。"""
    commute = [_car_commute(emp="2018017")]
    row = _integrated_transit("2018017", "", "", "通勤定期代")
    row[C_TOTAL] = "20000"
    res = evaluate_route_check([row], commute, {"2018017": "中村 淳一"})
    assert res[0]["判定"].startswith("△逆要確認（移動交通費")


# ----------------------------------------------------------------------
# 片側一致（▲）もレビューに出す（2026-08-12 谷津さん決定）
# ----------------------------------------------------------------------

def _side_match_rows():
    """乗車だけ／降車だけが登録経路上の行を作る（通勤系・非通勤系の両方）。"""
    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    rows = [
        _integrated_transit("2020001", "東京", "新宿", "交通費（電車・バス）"),  # 乗車のみ一致
        _integrated_transit("2020001", "渋谷", "品川", "通勤定期代"),            # 降車のみ一致
    ]
    return evaluate_route_check(rows, commute)


def test_route_check_marks_one_side_match_as_reviewable():
    """片側一致は「参考」ではなく▲要確認。通勤系で申請していても同じ扱い。"""
    res = _side_match_rows()
    assert [r["判定"] for r in res] == ["▲要確認（片側のみ一致）"] * 2
    assert [r["一致"] for r in res] == ["片側一致"] * 2


def test_route_choice_keys_include_one_side_matches():
    from services.keihi_summary import build_route_choice_keys
    pairs = build_route_choice_keys(_side_match_rows())
    assert len(pairs) == 2
    assert all(r["判定"].startswith("▲") for _k, r in pairs)


def test_default_route_choice_keeps_applied_side_for_one_side_match():
    """▲はシステムに判断材料が無いので勝手に動かさない＝申請どおりのまま。"""
    from services.keihi_summary import default_route_choice
    assert default_route_choice("▲要確認（片側のみ一致）", "通勤費", True) == "通勤費"
    assert default_route_choice("▲要確認（片側のみ一致）", "移動交通費", True) == "移動交通費"
    assert default_route_choice("▲要確認（片側のみ一致）", "対象外（顧客請求分）", False) == "対象外"


def test_match_route_choices_requires_one_side_match_rows_too():
    """▲行の選択が抜けていたら確定させない（★△と同じ全件必須ルール）。"""
    from services.keihi_summary import build_route_choice_keys, match_route_choices
    res = _side_match_rows()
    pairs = build_route_choice_keys(res)
    matched, errors = match_route_choices(res, [{"key": pairs[0][0], "choice": "通勤費"}])
    assert errors
    assert matched == []


def test_route_check_sheets_put_one_side_matches_last_and_color_them(tmp_path):
    from openpyxl import Workbook
    from services.keihi_summary import add_route_check_sheets
    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    rows = [
        _integrated_transit("2020001", "東京", "新宿", "交通費（電車・バス）"),  # ▲
        _integrated_transit("2020001", "東京", "品川", "交通費（電車・バス）"),  # ★
        _integrated_transit("2020001", "渋谷", "新宿", "通勤定期代"),            # △
    ]
    res = evaluate_route_check(rows, commute)
    wb = Workbook()
    summary = add_route_check_sheets(wb, res)
    assert (summary["flagged_rows"], summary["rev_rows"], summary["side_rows"]) == (1, 1, 1)

    ws = wb["要確認(経路突合)"]
    verdicts = [ws.cell(row=r, column=14).value for r in range(2, 5)]
    assert [v[0] for v in verdicts] == ["★", "△", "▲"]      # ★→△→▲の順
    assert ws.cell(row=4, column=1).fill.fgColor.rgb.endswith("DDEBF7")


def test_route_preview_counts_marks_separately(tmp_path):
    """▲を△に混ぜて数えない（引き算でrev_rowsを出していた箇所の回帰）。"""
    import csv
    from services.keihi_summary import run_keihi_route_preview
    jcsv = tmp_path / "jinjer.csv"
    with open(jcsv, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(["申請者社員番号", "申請者名"] + [f"c{i}" for i in range(31)])
        w.writerow(_jinjer_row(emp="2020001", trans="通勤定期代", total="8000",
                               use="2026/07/01"))
    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    res = run_keihi_route_preview(jinjer_csv=jcsv, log_func=lambda m: None,
                                  commute_rows=commute, travel_members={}, roster={})
    assert res.ok is True
    s = res.summary
    assert s["flagged_rows"] + s["rev_rows"] + s["side_rows"] == s["review_rows"]


# ----------------------------------------------------------------------
# 経路突合・分類をオフにしたときの警告（2026-08-12 谷津さん決定）
# ----------------------------------------------------------------------

def _minimal_jinjer_csv(tmp_path):
    import csv
    jcsv = tmp_path / "jinjer.csv"
    with open(jcsv, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(["申請者社員番号", "申請者名"] + [f"c{i}" for i in range(31)])
        w.writerow(_jinjer_row(emp="2020001", trans="通勤定期代", total="8000"))
    return jcsv


@pytest.mark.parametrize("route_check,classify,expected", [
    (False, True, ["経路突合がオフなので、経路内の移動交通費申請は立替金のままです"]),
    (True, False, ["分類・集計がオフなのでレビュー・付け替えは行われません"]),
    (False, False, ["経路突合がオフなので、経路内の移動交通費申請は立替金のままです",
                    "分類・集計がオフなのでレビュー・付け替えは行われません"]),
    (True, True, []),
])
def test_integration_warns_when_review_is_skipped(tmp_path, route_check, classify, expected):
    """止めはしないが、付け替えが起きないことは必ず画面に出す。"""
    from services.keihi_summary import run_keihi_integration
    res = run_keihi_integration(
        output_path=tmp_path / "out.xlsx", jinjer_csv=_minimal_jinjer_csv(tmp_path),
        route_check=route_check, classify=classify, log_func=lambda m: None)
    assert res.warnings == expected


# ----------------------------------------------------------------------
# Excel 出力・統合
# ----------------------------------------------------------------------

def test_add_integrated_sheet(tmp_path):
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    rows = [transform_jinjer([], [_jinjer_row()])[0]]
    add_integrated_sheet(wb, rows)
    out = tmp_path / "t.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws = wb2["経費統合一覧表"]
    assert ws.max_column == NCOL
    assert [c.value for c in ws[1]] == INTEGRATED_HEADERS
    assert ws.cell(row=2, column=1).value == "2026013"
    wb2.close()


def test_build_integrated_rows_order_and_counts(tmp_path):
    # jinjer CSV を作って build_integrated_rows（jinjer→e-staffing→SAP→freee 順）
    import csv
    jcsv = tmp_path / "jinjer.csv"
    with open(jcsv, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        header = ["申請者社員番号", "申請者名"] + [f"c{i}" for i in range(31)]
        w.writerow(header)
        w.writerow(_jinjer_row(emp="2026013", total="356"))
    rows, counts = build_integrated_rows(jinjer_csv=jcsv, log_func=lambda m: None)
    assert counts["jinjer"] == 1
    assert counts["estaffing"] is None      # 未取込
    assert rows[0][C_EMP] == "2026013"


def _route_choice_csv(tmp_path):
    """★（経路内なのに立替系）と △（通勤系だが経路外）が1行ずつ出る jinjer CSV。"""
    import csv
    star = _jinjer_row(emp="2020001", trans="交通費（電車・バス）", total="300")
    star[C_BOARD], star[C_ALIGHT] = "東京", "品川"
    rev = _jinjer_row(emp="2020001", trans="通勤定期代", total="8000", use="2026/07/01")
    rev[C_BOARD], rev[C_ALIGHT] = "渋谷", "新宿"
    jcsv = tmp_path / "jinjer.csv"
    with open(jcsv, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(["申請者社員番号", "申請者名"] + [f"c{i}" for i in range(31)])
        w.writerow(star)
        w.writerow(rev)
    commute = [{"社員番号": "2020001", "出発": "東京", "到着": "品川",
                "経由1": "", "経由2": "", "通勤経路": "", "利用交通機関": "電車"}]
    return jcsv, commute


def _offline_route_check(monkeypatch, commute):
    """jinjer API を叩かずに経路突合を回す（通勤経路だけ差し込む）。"""
    import services.keihi_summary as ks
    monkeypatch.setattr(ks, "fetch_roster_and_commute",
                        lambda client, route_check, log_func: ({}, {}, commute))
    monkeypatch.setattr(ks, "load_travel_members", lambda log_func=print: {})


def test_run_keihi_integration_applies_route_choices(tmp_path, monkeypatch):
    """人が選んだ計上先どおりに非課税通勤費と立替金を入れ替え、証跡を残す。"""
    from openpyxl import load_workbook
    from services.keihi_summary import run_keihi_route_preview
    jcsv, commute = _route_choice_csv(tmp_path)
    _offline_route_check(monkeypatch, commute)

    pre = run_keihi_route_preview(jinjer_csv=jcsv, log_func=lambda m: None,
                                  commute_rows=commute, travel_members={}, roster={})
    assert {r["交通機関"]: r["計上先"] for r in pre.review_rows} == {
        "交通費（電車・バス）": "移動交通費", "通勤定期代": "通勤費"}
    # ★は通勤費へ、△は移動交通費へ（＝両者を入れ替える）
    choices = [{"key": r["key"],
                "choice": "移動交通費" if r["交通機関"] == "通勤定期代" else "通勤費"}
               for r in pre.review_rows]

    out = tmp_path / "integrated.xlsx"
    res = run_keihi_integration(output_path=out, jinjer_csv=jcsv, route_check=True,
                                classify=True, route_choices=choices,
                                log_func=lambda m: None)
    assert res.ok is True
    row = res.import_preview[0]
    assert (row["非課税通勤費"], row["立替金"]) == (300, 8000)   # 8000 と 300 が入れ替わる
    assert sum(v for k, v in row.items() if isinstance(v, int)) == 8300   # 合計は不変
    assert sorted(m["変更"] for m in res.route_moves) == [
        "立替金 → 非課税通勤費", "非課税通勤費 → 立替金"]

    wb = load_workbook(out)
    # シートの並びは従来どおり（経路突合は分類より後に書くが位置は変えない）
    assert wb.sheetnames == ["経費統合一覧表", "要確認(経路突合)", "全交通費行(経路突合)",
                             "集計", "集計ログ"]
    ws = wb["要確認(経路突合)"]
    header = [c.value for c in ws[1]]
    assert header[-2:] == ["人間判定", "計上先変更"]
    judged = {ws.cell(row=r, column=4).value: ws.cell(row=r, column=len(header)).value
              for r in (2, 3)}
    assert judged["交通費（電車・バス）"] == "立替金 → 非課税通勤費 300円"
    assert judged["通勤定期代"] == "非課税通勤費 → 立替金 8,000円"
    wb.close()


def test_run_keihi_integration_stops_when_choices_do_not_match(tmp_path, monkeypatch):
    """プレビュー後に入力が変わった場合は生成せず止める（人が見ていない行を通さない）。"""
    jcsv, commute = _route_choice_csv(tmp_path)
    _offline_route_check(monkeypatch, commute)
    out = tmp_path / "integrated.xlsx"
    res = run_keihi_integration(
        output_path=out, jinjer_csv=jcsv, route_check=True, classify=True,
        route_choices=[{"key": "9999999|2026/6/1|通勤定期代|300|||1", "choice": "通勤費"}],
        log_func=lambda m: None)
    assert res.ok is False
    assert "プレビュー" in res.error
    assert not out.exists()
    assert not list(tmp_path.glob("*インポート.csv"))


def test_run_keihi_integration_requires_classify_for_choices(tmp_path, monkeypatch):
    jcsv, commute = _route_choice_csv(tmp_path)
    _offline_route_check(monkeypatch, commute)
    res = run_keihi_integration(
        output_path=tmp_path / "x.xlsx", jinjer_csv=jcsv, route_check=True,
        classify=False, route_choices=[{"key": "a", "choice": "通勤費"}],
        log_func=lambda m: None)
    assert res.ok is False and "分類・集計" in res.error


def test_run_keihi_integration_jinjer_only(tmp_path):
    # route_check=False かつ jinjer のみ → API 不要でオフライン実行できる
    import csv
    from openpyxl import load_workbook
    jcsv = tmp_path / "jinjer.csv"
    with open(jcsv, "w", encoding="cp932", newline="") as f:
        w = csv.writer(f)
        w.writerow(["申請者社員番号", "申請者名"] + [f"c{i}" for i in range(31)])
        w.writerow(_jinjer_row(emp="2026013", total="356"))
        w.writerow(_jinjer_row(emp="2026013", billtype="顧客請求", total="8990",
                               detail="客先請求分（交通費）"))
    out = tmp_path / "integrated.xlsx"
    res = run_keihi_integration(output_path=out, jinjer_csv=jcsv, route_check=False,
                                log_func=lambda m: None)
    assert res.ok is True
    assert res.integrated_rows == 2
    assert res.source_counts["jinjer"] == 2
    assert out.exists()
    wb = load_workbook(out)
    assert "経費統合一覧表" in wb.sheetnames
    wb.close()
