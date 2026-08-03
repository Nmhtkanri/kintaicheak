# -*- coding: utf-8 -*-
"""通勤費申請なしリストの自動整備（経費チェックモードの改修案 P0）のテスト。"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook, load_workbook

from services.expense_check import (
    _norm_emp_id, read_no_commute_list, build_no_commute_rows, add_no_commute_sheet,
    NO_COMMUTE_HEADERS, NO_COMMUTE_SHEET,
)


def _summary(emp_id, name, work, tw_count):
    return {"id": emp_id, "name": name, "work_days": work,
            "telework_days": [("2026-06-01", "テレワーク")] * tw_count}


def test_norm_emp_id():
    assert _norm_emp_id(2026013) == "2026013"
    assert _norm_emp_id(2026013.0) == "2026013"   # Excelの数値セル
    assert _norm_emp_id(" 2026013 ") == "2026013"
    assert _norm_emp_id(None) == ""


def test_read_no_commute_list(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["社員番号", "氏名"])
    ws.append([2026013, "川口 祐ノ輔"])   # 数値セル（改修案エッジ1: 型不一致）
    ws.append(["2020008", "佐藤 清"])
    ws.append([None, None])              # 空行はスキップ
    ws.append([2026013, "川口 祐ノ輔"])   # 重複IDはスキップ
    p = tmp_path / "list.xlsx"
    wb.save(p)
    entries = read_no_commute_list(p)
    assert entries == [
        {"id": "2026013", "name": "川口 祐ノ輔"},
        {"id": "2020008", "name": "佐藤 清"},
    ]


def test_read_no_commute_list_prefers_named_sheet(tmp_path):
    wb = Workbook()
    wb.active.append(["9999999", "別シートの人"])
    ws = wb.create_sheet(NO_COMMUTE_SHEET)
    ws.append(["社員番号", "氏名"])
    ws.append(["2020001", "対象の人"])
    p = tmp_path / "list.xlsx"
    wb.save(p)
    assert read_no_commute_list(p) == [{"id": "2020001", "name": "対象の人"}]


def test_step2_full_remote_removed_but_zero_attendance_kept():
    entries = [{"id": "1", "name": "完全在宅"}, {"id": "2", "name": "出勤ゼロ"},
               {"id": "3", "name": "出社あり"}, {"id": "9", "name": "勤怠なし"}]
    summary = [_summary("1", "完全在宅", 20, 20),
               _summary("2", "出勤ゼロ", 0, 0),
               _summary("3", "出社あり", 20, 5)]
    kept, removed = build_no_commute_rows(entries, summary, [])
    assert [r["id"] for r in removed] == ["1"]          # 完全在宅のみ除外
    assert [r["id"] for r in kept] == ["2", "3", "9"]   # 出勤0・サマリ外は残す
    by = {r["id"]: r for r in kept}
    assert by["2"]["remark"] == "出勤0"
    assert by["3"]["remark"] == ""
    assert by["9"]["remark"] == "勤怠データなし"


def test_step3_amount_sum_and_interval_mapping():
    entries = [{"id": "10", "name": ""}, {"id": "11", "name": "毎日さん"},
               {"id": "12", "name": "未登録"}, {"id": "13", "name": "金額ゼロ"},
               {"id": "14", "name": "間隔対象外"}]
    summary = [_summary(i, n, 20, 5) for i, n in
               [("10", "定期さん"), ("11", "毎日さん"), ("12", "未登録"),
                ("13", "金額ゼロ"), ("14", "間隔対象外")]]
    commute = [
        # 複数経路 → 金額は合算、支給間隔は先頭経路（経路No最小）を採用
        {"社員番号": "10", "経路No": 2, "支給間隔": "毎月", "支給金額": 5000, "利用交通機関": "電車"},
        {"社員番号": "10", "経路No": 1, "支給間隔": "毎月", "支給金額": 8000, "利用交通機関": "電車"},
        {"社員番号": "11", "経路No": 1, "支給間隔": "毎日", "支給金額": 400, "利用交通機関": "バス"},
        # 支給金額0で登録あり → 金額0だが区分は間隔どおり（改修案エッジ3）
        {"社員番号": "13", "経路No": 1, "支給間隔": "毎月", "支給金額": 0, "利用交通機関": "電車"},
        # 毎月・毎日以外の間隔 → 区分空欄
        {"社員番号": "14", "経路No": 1, "支給間隔": "3ヶ月", "支給金額": 30000, "利用交通機関": "電車"},
    ]
    kept, _ = build_no_commute_rows(entries, summary, commute)
    by = {r["id"]: r for r in kept}
    assert by["10"]["amount"] == 13000 and by["10"]["kubun"] == "通勤定期代"
    assert by["10"]["name"] == "定期さん"                       # 氏名はサマリから補完
    assert by["11"]["amount"] == 400 and by["11"]["kubun"] == "通勤費"
    assert by["12"]["amount"] == 0 and by["12"]["kubun"] == ""   # 未登録（エッジ2）
    assert by["12"]["kikan"] == ""
    assert by["13"]["amount"] == 0 and by["13"]["kubun"] == "通勤定期代"
    assert by["14"]["amount"] == 30000 and by["14"]["kubun"] == ""


def test_step3_amount_skips_blank_values():
    """支給金額が空文字（_to_amount の空）でも落ちずに合算する。"""
    entries = [{"id": "20", "name": "空金額"}]
    summary = [_summary("20", "空金額", 20, 5)]
    commute = [
        {"社員番号": "20", "経路No": 1, "支給間隔": "毎月", "支給金額": "", "利用交通機関": "電車"},
        {"社員番号": "20", "経路No": 2, "支給間隔": "毎月", "支給金額": 1200, "利用交通機関": "電車"},
    ]
    kept, _ = build_no_commute_rows(entries, summary, commute)
    assert kept[0]["amount"] == 1200
    assert kept[0]["kubun"] == "通勤定期代"


def test_travel_members_get_kubun_and_auto_added():
    """移動交通費（立替精算）対象者は区分を固定し、リスト漏れは自動追加する。"""
    entries = [{"id": "2018017", "name": "中村 淳一"}, {"id": "30", "name": "普通の人"}]
    summary = [_summary("2018017", "中村 淳一", 20, 0),
               _summary("30", "普通の人", 20, 5),
               _summary("2026001", "佐久間歩", 15, 0)]
    # 中村さんに通勤費登録があっても（毎月＝通勤定期代のはず）移動交通費が優先される
    commute = [{"社員番号": "2018017", "経路No": 1, "支給間隔": "毎月",
                "支給金額": 5000, "利用交通機関": "電車"}]
    travel = {"2018017": "中村 淳一", "2026001": "佐久間歩"}
    kept, removed = build_no_commute_rows(entries, summary, commute, travel)
    by = {r["id"]: r for r in kept}
    assert by["2018017"]["kubun"] == "移動交通費"
    assert by["2018017"]["amount"] == 5000            # 金額は事実として転記
    assert by["2018017"]["remark"] == "立替精算対象"
    assert by["30"]["kubun"] == ""                    # 対象外の人は従来どおり
    # 手動リストに無い対象者 2026001 は自動追加され、氏名はサマリから補完
    assert "2026001" in by
    assert by["2026001"]["kubun"] == "移動交通費"
    assert by["2026001"]["name"] == "佐久間歩"
    assert removed == []


def test_travel_member_full_remote_still_removed():
    """移動交通費対象者でも完全在宅（STEP2）はこれまでどおり除外される。"""
    entries = [{"id": "2020008", "name": "佐藤 清"}]
    summary = [_summary("2020008", "佐藤 清", 20, 20)]
    kept, removed = build_no_commute_rows(entries, summary, [], {"2020008": "佐藤 清"})
    assert kept == []
    assert [r["id"] for r in removed] == ["2020008"]


def test_travel_member_remark_joins_attendance_note():
    """勤怠なしの対象者は備考が「勤怠データなし・立替精算対象」になる。"""
    kept, _ = build_no_commute_rows([], [], [], {"2020021": "二神 啓城"})
    assert kept[0]["id"] == "2020021"
    assert kept[0]["remark"] == "勤怠データなし・立替精算対象"
    assert kept[0]["kubun"] == "移動交通費"


def test_add_no_commute_sheet(tmp_path):
    wb = Workbook()
    add_no_commute_sheet(wb, [
        {"id": "2020001", "name": "田中 一郎", "amount": 8000, "kubun": "通勤定期代",
         "kikan": "電車", "remark": ""},
        {"id": "2020002", "name": "上原 奏吾", "amount": 0, "kubun": "",
         "kikan": "", "remark": "出勤0"},
    ])
    p = tmp_path / "out.xlsx"
    wb.save(p)
    ws = load_workbook(p)[NO_COMMUTE_SHEET]
    assert [c.value for c in ws[1]] == NO_COMMUTE_HEADERS
    # 空文字セルは再読込時に None になる
    assert [c.value for c in ws[2]] == ["2020001", "田中 一郎", 8000, "通勤定期代", "電車", None]
    assert [c.value for c in ws[3]] == ["2020002", "上原 奏吾", 0, None, None, "出勤0"]
    assert ws.cell(row=4, column=3).value == "=SUM(C2:C3)"   # 合計行
