# -*- coding: utf-8 -*-
"""sap_duplicate_filter のテスト ＋ run_keihi_integration へのSAP重複除外統合テスト"""
import csv

import pytest
from openpyxl import load_workbook

from services.sap_duplicate_filter import (
    collect_past_files,
    parse_past_inputs,
    run_sap_dedup,
)
from services.keihi_summary import run_keihi_integration

SAP_HEADERS = ["姓", "名", "費用合計", "業者名", "費用エントリ日", "説明", "承認日",
               "事業単位", "コストセンター", "通貨", "費用シート ID", "勤務地", "ステータス"]


def _sap_row(sheet_id, sei="山田", mei="太郎", total="1000"):
    return [sei, mei, total, "業者A", "2026/06/01", "説明X", "2026/06/02",
            "BU", "CC1", "JPY", sheet_id, "東京", "承認済"]


def _write_sap_csv(path, rows, headers=None, encoding="cp932"):
    with open(path, "w", encoding=encoding, newline="") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow(headers if headers is not None else SAP_HEADERS)
        w.writerows(rows)
    return path


class _NoApiClient:
    """roster 構築の fetch_active_employees を失敗させてオフライン続行させるダミー"""


class TestRunSapDedup:
    def test_basic_dedup(self, tmp_path):
        past = _write_sap_csv(tmp_path / "past.csv", [_sap_row("EXP-1"), _sap_row("EXP-2")])
        cur = _write_sap_csv(tmp_path / "current.csv",
                             [_sap_row("EXP-2"), _sap_row("EXP-3"), _sap_row("EXP-4")])
        out = tmp_path / "out"
        r = run_sap_dedup([past], cur, out)
        assert r.removed_row_count == 1
        assert r.kept_row_count == 2
        assert r.remaining_match_count == 0
        assert r.removed_rows[0]["照合キー"] == "EXP-2"
        assert "past.csv" in r.removed_rows[0]["重複元ファイル"]
        # 除外済みCSVには EXP-3 / EXP-4 だけが残る
        with open(r.clean_path, encoding="cp932", newline="") as f:
            rows = list(csv.DictReader(f))
        assert [row["費用シート ID"] for row in rows] == ["EXP-3", "EXP-4"]
        # 除外一覧CSVには付加列がある
        with open(r.removed_path, encoding="cp932", newline="") as f:
            removed = list(csv.DictReader(f))
        assert removed[0]["除外理由"].startswith("過去SAP経費と")
        assert removed[0]["重複元件数"] == "1"

    def test_empty_id_never_removed(self, tmp_path):
        past = _write_sap_csv(tmp_path / "past.csv", [_sap_row(""), _sap_row("EXP-1")])
        cur = _write_sap_csv(tmp_path / "current.csv", [_sap_row(""), _sap_row("EXP-9")])
        r = run_sap_dedup([past], cur, tmp_path / "out")
        assert r.removed_row_count == 0
        assert r.kept_row_count == 2

    def test_multi_line_sheet_all_removed(self, tmp_path):
        """1つの費用シートに複数明細 → そのIDの当月行はすべて除外"""
        past = _write_sap_csv(tmp_path / "past.csv", [_sap_row("EXP-5")])
        cur = _write_sap_csv(tmp_path / "current.csv",
                             [_sap_row("EXP-5", total="100"), _sap_row("EXP-5", total="200"),
                              _sap_row("EXP-6")])
        r = run_sap_dedup([past], cur, tmp_path / "out")
        assert r.removed_row_count == 2
        assert r.kept_row_count == 1

    def test_id_whitespace_stripped(self, tmp_path):
        past = _write_sap_csv(tmp_path / "past.csv", [_sap_row(" EXP-7 ")])
        cur = _write_sap_csv(tmp_path / "current.csv", [_sap_row("EXP-7")])
        r = run_sap_dedup([past], cur, tmp_path / "out")
        assert r.removed_row_count == 1

    def test_folder_recursive_and_generated_skipped(self, tmp_path):
        """フォルダ指定でサブフォルダも走査し、生成物CSVは過去として読まない"""
        pastdir = tmp_path / "R8年"
        sub = pastdir / "6月" / "経費確認精査データ"
        sub.mkdir(parents=True)
        _write_sap_csv(sub / "SAP_経費月次出力 (3).csv", [_sap_row("EXP-A")])
        # 生成物（読み飛ばし対象）: もし読まれたら EXP-B も除外されテストが落ちる
        _write_sap_csv(sub / "SAP_経費月次出力 (3)_重複除外済.csv", [_sap_row("EXP-B")])
        _write_sap_csv(sub / "経費統合一覧表_SAP除外一覧.csv", [_sap_row("EXP-C")])
        cur = _write_sap_csv(tmp_path / "current.csv",
                             [_sap_row("EXP-A"), _sap_row("EXP-B"), _sap_row("EXP-C")])
        r = run_sap_dedup([pastdir], cur, tmp_path / "out")
        assert r.past_file_count == 1
        assert r.removed_row_count == 1
        assert {row["照合キー"] for row in r.removed_rows} == {"EXP-A"}

    def test_current_csv_not_matched_against_itself(self, tmp_path):
        """当月CSVが過去フォルダ内にあっても自分自身とは突合しない"""
        folder = tmp_path / "data"
        folder.mkdir()
        cur = _write_sap_csv(folder / "current.csv", [_sap_row("EXP-X")])
        with pytest.raises(ValueError, match="過去SAP CSVが1つも"):
            collect_past_files([folder], cur, recursive=True)

    def test_missing_id_column(self, tmp_path):
        headers = [h for h in SAP_HEADERS if h != "費用シート ID"]
        past = _write_sap_csv(tmp_path / "past.csv",
                              [[c for c in _sap_row("EXP-1")][:10] + ["東京", "承認済"]],
                              headers=headers)
        cur = _write_sap_csv(tmp_path / "current.csv", [_sap_row("EXP-1")])
        with pytest.raises(ValueError, match="必要な列がありません"):
            run_sap_dedup([past], cur, tmp_path / "out")

    def test_utf8_bom_csv(self, tmp_path):
        past = _write_sap_csv(tmp_path / "past.csv", [_sap_row("EXP-1")], encoding="utf-8-sig")
        cur = _write_sap_csv(tmp_path / "current.csv",
                             [_sap_row("EXP-1"), _sap_row("EXP-2")], encoding="utf-8-sig")
        r = run_sap_dedup([past], cur, tmp_path / "out")
        assert r.removed_row_count == 1

    def test_output_stem(self, tmp_path):
        past = _write_sap_csv(tmp_path / "past.csv", [_sap_row("EXP-1")])
        cur = _write_sap_csv(tmp_path / "current.csv", [_sap_row("EXP-2")])
        r = run_sap_dedup([past], cur, tmp_path / "out", output_stem="経費統合一覧表_2026年07月")
        assert r.clean_path.name == "経費統合一覧表_2026年07月_SAP重複除外済.csv"
        assert r.removed_path.name == "経費統合一覧表_2026年07月_SAP除外一覧.csv"


class TestParsePastInputs:
    def test_semicolon_and_newline(self):
        got = parse_past_inputs('Y:\\a.csv; "Y:\\b フォルダ"\nY:\\c.csv;')
        assert got == ["Y:\\a.csv", "Y:\\b フォルダ", "Y:\\c.csv"]

    def test_empty(self):
        assert parse_past_inputs("") == []
        assert parse_past_inputs("  \n ; ") == []


class TestKeihiIntegrationWithLedger:
    """run_keihi_integration が取込済み費用シート台帳と突合することの確認（2026-08-06〜）"""

    def _ledger(self, tmp_path, rows):
        from services.sap_import_ledger import (
            Ledger, append_provisional, confirm_month, save_ledger)
        led = Ledger(path=tmp_path / "台帳.csv", rows=[])
        append_provisional(led, rows, "2026-07", "past.csv", user="tester")
        confirm_month(led, "2026-07", user="tester")
        save_ledger(led)
        return led.path

    def _cur(self, tmp_path, rows):
        return _write_sap_csv(tmp_path / "current.csv", rows)

    def test_ledger_applied_before_transform(self, tmp_path):
        """明細まで一致＝除外／IDだけ一致＝要確認／台帳に無い＝取込"""
        ledger_csv = self._ledger(tmp_path, [dict(zip(SAP_HEADERS, _sap_row("EXP-1")))])
        cur = self._cur(tmp_path, [
            _sap_row("EXP-1"),                    # 明細まで一致 → 除外
            _sap_row("EXP-1", total="9999"),      # IDだけ一致   → 要確認
            _sap_row("EXP-2"),                    # 台帳に無い   → 取込
        ])
        out = tmp_path / "integrated.xlsx"
        res = run_keihi_integration(
            output_path=out, sap_csv=cur, route_check=False, classify=False,
            sap_ledger_csv=str(ledger_csv), client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.ok is True
        assert res.sap_dedup["excluded_rows"] == 1
        assert res.sap_dedup["review_rows"] == 1
        assert res.sap_dedup["kept_rows"] == 1
        # 統合一覧表に載るのは取込対象の1行だけ（要確認は取り込まない）
        assert res.source_counts["sap"] == 1

        wb = load_workbook(out)
        assert "SAP除外(支給済み)" in wb.sheetnames
        assert "SAP要確認" in wb.sheetnames
        ws = wb["SAP除外(支給済み)"]
        headers = [c.value for c in ws[1]]
        assert "費用シート ID" in headers and "判定" in headers and "理由" in headers
        assert ws.cell(row=2, column=headers.index("費用シート ID") + 1).value == "EXP-1"
        # 3種のCSVが出力フォルダに残る（証跡）
        assert (tmp_path / "integrated_SAP取込対象.csv").exists()
        assert (tmp_path / "integrated_SAP除外.csv").exists()
        assert (tmp_path / "integrated_SAP要確認.csv").exists()

    def test_without_ledger_behaves_as_before(self, tmp_path):
        cur = self._cur(tmp_path, [_sap_row("EXP-1"), _sap_row("EXP-2")])
        out = tmp_path / "integrated.xlsx"
        res = run_keihi_integration(
            output_path=out, sap_csv=cur, route_check=False, classify=False,
            sap_ledger_csv=None, client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.ok is True
        assert res.sap_dedup == {}
        assert res.source_counts["sap"] == 2
        assert "SAP除外(支給済み)" not in load_workbook(out).sheetnames

    def test_missing_ledger_file_is_not_an_error(self, tmp_path):
        """台帳がまだ無いとき（初回運用）は除外なしで通す"""
        cur = self._cur(tmp_path, [_sap_row("EXP-1")])
        res = run_keihi_integration(
            output_path=tmp_path / "integrated.xlsx", sap_csv=cur,
            route_check=False, classify=False,
            sap_ledger_csv=str(tmp_path / "まだ無い台帳.csv"),
            client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.ok is True
        assert res.sap_dedup["excluded_rows"] == 0
        assert res.sap_dedup["kept_rows"] == 1

    def test_record_ledger_writes_provisional(self, tmp_path):
        from services.sap_import_ledger import load_ledger
        ledger_csv = tmp_path / "台帳.csv"
        cur = self._cur(tmp_path, [_sap_row("EXP-1"), _sap_row("EXP-2")])
        res = run_keihi_integration(
            output_path=tmp_path / "integrated.xlsx", sap_csv=cur,
            route_check=False, classify=False,
            sap_ledger_csv=str(ledger_csv), sap_import_month="2026-08",
            sap_record_ledger=True, client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.ok is True
        assert res.sap_dedup["recorded_month"] == "2026-08"
        assert res.sap_dedup["recorded_rows"] == 2
        led = load_ledger(ledger_csv)
        assert led.provisional_months == ["2026-08"]
        # 暫定は判定に使われないので、確定しない限り翌回も除外されない
        assert led.confirmed_rows == []

    def test_skip_reason_is_reported(self, tmp_path):
        """記録しなかったことを黙って進まない（確定漏れ＝翌月の二重計上につながるため）"""
        cur = self._cur(tmp_path, [_sap_row("EXP-1")])
        res = run_keihi_integration(
            output_path=tmp_path / "integrated.xlsx", sap_csv=cur,
            route_check=False, classify=False,
            sap_ledger_csv=str(tmp_path / "台帳.csv"),
            sap_record_ledger=False,
            sap_record_skip_reason="台帳への書き込みは 谷津さん・平良さん のみです。",
            client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.ok is True
        assert res.sap_dedup["record_skipped"] is True
        assert "谷津さん・平良さん" in res.sap_dedup["record_skip_reason"]
        # 記録していないので台帳ファイルは作られない
        assert not (tmp_path / "台帳.csv").exists()

    def test_no_skip_flag_when_recorded(self, tmp_path):
        cur = self._cur(tmp_path, [_sap_row("EXP-1")])
        res = run_keihi_integration(
            output_path=tmp_path / "integrated.xlsx", sap_csv=cur,
            route_check=False, classify=False,
            sap_ledger_csv=str(tmp_path / "台帳.csv"),
            sap_import_month="2026-08", sap_record_ledger=True,
            client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.sap_dedup.get("record_skipped") is None
        assert res.sap_dedup["recorded_rows"] == 1

    def test_ledger_failure_aborts(self, tmp_path):
        """台帳が壊れていて読めないときは、黙って重複入りで進めず中止する"""
        broken = tmp_path / "壊れた台帳.csv"
        broken.write_bytes(b"\xff\xfe\x00\x00not a csv")
        cur = self._cur(tmp_path, [_sap_row("EXP-1")])
        res = run_keihi_integration(
            output_path=tmp_path / "integrated.xlsx", sap_csv=cur,
            route_check=False, classify=False, sap_ledger_csv=str(broken),
            client=_NoApiClient(), log_func=lambda m: None,
        )
        assert res.ok is False
        assert "SAP台帳との突合に失敗" in res.error
