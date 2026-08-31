# -*- coding: utf-8 -*-
"""有休ブックの読み手（services/paid_leave_mail.py）のテスト。

実ブック（有休取得4日以下_*.xlsx）と同じ形の合成ブックで、
グレー塗りの除外・取得日数の合計・不足日数・取得期限を確かめる。
"""
import os
import sys
import tempfile
import unittest
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.mail_draft import STATUS_OK, build_plans_for  # noqa: E402
from services.paid_leave_mail import (  # noqa: E402
    deadline_for,
    japanese_date,
    load_paid_leave_table,
    read_paid_leave_report,
)

GRAY = PatternFill("solid", fgColor="D9D9D9")


def make_book(path, *, gray_rows=(), as_of=date(2026, 8, 31)):
    """実ブックと同じ3シート構成の合成ブックを作る。

    対象者一覧は見出しが5行目、データが6行目から（実ブックと同じ並び）。
    """
    workbook = Workbook()
    target = workbook.active
    target.title = "対象者一覧"
    detail = workbook.create_sheet("有休明細")
    condition = workbook.create_sheet("集計条件")

    target["A1"] = "有休取得4日以下"
    target.append([])  # 2行目
    target.append([])
    target.append([])
    target.append(["社員番号", "氏名", "取得日数", "全日回数", "半休回数",
                   "時間休回数", "雇用区分", "所属"])
    people = [("2024001", "山田太郎"), ("2024002", "鈴木花子"), ("2024003", "佐藤次郎")]
    for employee_id, name in people:
        target.append([employee_id, name, "", "", "", "", "正社員", "本社"])

    # 行全体（A〜H）をグレーで塗った人は対象外
    for row_number in gray_rows:
        for column in range(1, 9):
            target.cell(row_number, column).fill = GRAY

    detail.append(["社員番号", "氏名", "日付", "種別", "換算日数"])
    detail.append(["2024001", "山田太郎", date(2026, 5, 1), "全日", 1.0])
    detail.append(["2024001", "山田太郎", date(2026, 6, 2), "半休", 0.5])
    detail.append(["2024002", "鈴木花子", date(2026, 7, 3), "半休", 0.5])
    detail.append(["2024002", "鈴木花子", date(2026, 7, 4), "時間休", 0.0])
    # 2024003 は取得なし（明細に行が無い）

    condition["A1"] = "対象開始日"
    condition["B1"] = date(2026, 4, 1)
    condition["A2"] = "対象終了日"
    condition["B2"] = as_of
    condition["A3"] = "法定5日到達基準"
    condition["B3"] = 5
    workbook.save(path)


class PaidLeaveTableTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.path = os.path.join(self.tmp.name, "有休ブック.xlsx")

    def test_columns_match_the_shared_template(self):
        """列名は共有テンプレート「有休取得のお願い」の差し込み名と一致させる。"""
        make_book(self.path)
        headers, rows, id_key, name_key, meta = load_paid_leave_table(self.path)
        self.assertEqual(headers, ["社員番号", "氏名", "取得日数", "不足日数", "取得期限"])
        self.assertEqual((id_key, name_key), ("社員番号", "氏名"))
        self.assertEqual(len(rows), 3)

    def test_taken_days_are_summed_from_detail(self):
        make_book(self.path)
        _, rows, _, _, _ = load_paid_leave_table(self.path)
        taken = {r["社員番号"]: (r["取得日数"], r["不足日数"]) for r in rows}
        self.assertEqual(taken["2024001"], ("1.5", "3.5"))   # 全日1.0 + 半休0.5
        self.assertEqual(taken["2024002"], ("0.5", "4.5"))   # 時間休は0.0で数えない
        self.assertEqual(taken["2024003"], ("0.0", "5.0"))   # 明細なし

    def test_gray_rows_are_excluded(self):
        make_book(self.path, gray_rows=(7,))  # 6行目=2024001, 7行目=2024002
        _, rows, _, _, meta = load_paid_leave_table(self.path)
        self.assertEqual([r["社員番号"] for r in rows], ["2024001", "2024003"])
        self.assertEqual([p["社員番号"] for p in meta["gray_excluded"]], ["2024002"])
        self.assertIn("対象外 1人", meta["note"])

    def test_deadline_defaults_to_next_march(self):
        make_book(self.path)
        _, rows, _, _, meta = load_paid_leave_table(self.path)
        self.assertEqual(rows[0]["取得期限"], "2027年3月31日")
        self.assertEqual(meta["deadline"], "2027-03-31")
        self.assertEqual(deadline_for(date(2026, 8, 31)), date(2027, 3, 31))

    def test_deadline_can_be_given(self):
        make_book(self.path)
        _, rows, _, _, _ = load_paid_leave_table(self.path, deadline="2026-12-31")
        self.assertEqual(rows[0]["取得期限"], "2026年12月31日")

    def test_deadline_before_as_of_is_rejected(self):
        make_book(self.path)
        with self.assertRaises(ValueError) as raised:
            load_paid_leave_table(self.path, deadline="2026-08-01")
        self.assertIn("集計基準日", str(raised.exception))

    def test_all_gray_is_an_error_not_an_empty_mail_run(self):
        make_book(self.path, gray_rows=(6, 7, 8))
        with self.assertRaises(ValueError) as raised:
            load_paid_leave_table(self.path)
        self.assertIn("対象者がいません", str(raised.exception))

    def test_missing_sheet_is_reported(self):
        make_book(self.path)
        from openpyxl import load_workbook
        workbook = load_workbook(self.path)
        del workbook["有休明細"]
        workbook.save(self.path)
        with self.assertRaises(ValueError) as raised:
            read_paid_leave_report(self.path)
        self.assertIn("有休明細", str(raised.exception))

    def test_report_keeps_the_period(self):
        make_book(self.path)
        report = read_paid_leave_report(self.path)
        self.assertEqual(report.start_date, date(2026, 4, 1))
        self.assertEqual(report.as_of_date, date(2026, 8, 31))
        self.assertEqual(report.target_days, 5.0)
        self.assertEqual(japanese_date(report.as_of_date), "2026年8月31日")


class PaidLeavePlansTest(unittest.TestCase):
    """メール下書きモードの共通処理に、そのまま流れることを確かめる。"""

    def test_build_plans_for_paid_leave_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            book_path = os.path.join(tmp, "有休ブック.xlsx")
            make_book(book_path, gray_rows=(8,))
            address_path = os.path.join(tmp, "台帳.xlsm")
            address = Workbook()
            sheet = address.active
            sheet.title = "メール送信"
            sheet.append(["", "社員番号", "氏名", "社用", "就業先", "個人"])
            sheet.append(["", "2024001", "山田太郎", "yamada@nmht.co.jp", "", ""])
            sheet.append(["", "2024002", "鈴木花子", "suzuki@nmht.co.jp", "", ""])
            address.save(address_path)

            template = {"subject": "【ご確認】年次有給休暇について",
                        "body": "{{氏名}}さん\n取得日数は{{取得日数}}日です。\n"
                                "{{取得期限}}までにあと{{不足日数}}日お願いします。",
                        "cc": "kanri@nmht.co.jp", "bcc_mode": "to_only",
                        "importance": "high"}
            plans, meta = build_plans_for(book_path, address_path, template,
                                          source="paid_leave")
            self.assertEqual(meta["source"], "paid_leave")
            self.assertEqual(meta["counts"]["total"], 2)      # 佐藤次郎はグレーで対象外
            self.assertEqual(meta["counts"]["ok"], 2)
            self.assertEqual(len(meta["source_meta"]["gray_excluded"]), 1)
            first = [p for p in plans if p["employee_id"] == "2024001"][0]
            self.assertEqual(first["status"], STATUS_OK)
            self.assertEqual(first["to"], ["yamada@nmht.co.jp"])
            self.assertEqual(first["cc"], ["kanri@nmht.co.jp"])
            self.assertIn("取得日数は1.5日です。", first["body"])
            self.assertIn("2027年3月31日までにあと3.5日", first["body"])


if __name__ == "__main__":
    unittest.main()
