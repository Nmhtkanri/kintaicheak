import os
import sys
from datetime import date

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from quick_compare import (  # noqa: E402
    DIFF_COLUMNS,
    DIFF_KIND_ABSENCE,
    DIFF_KIND_PUNCH_IN,
    DIFF_KIND_PUNCH_OUT,
    DIFF_KIND_TOTAL,
    JINJER_HEADERS,
    LogEntry,
    classify_total_work,
    clean_punch_comment,
    compute_diffs,
    format_stamp_comments,
    kintai_total_minutes,
    load_stamp_correction_reasons,
    normalize_kintai_result_columns,
    overnight_display_value,
    recommend_judge_label,
    resolve_jinjer_extra_columns,
    strip_punch_noise_words,
    to_jinjer_overnight_punch_out,
)


def _jinjer_row(**overrides):
    row = {
        JINJER_HEADERS["name"]: "上原 奏吾",
        JINJER_HEADERS["emp_id"]: "2018057",
        JINJER_HEADERS["date"]: "2026/4/1",
        JINJER_HEADERS["punch_in_1"]: "",
        JINJER_HEADERS["punch_out_1"]: "",
        JINJER_HEADERS["break_total"]: "",
        JINJER_HEADERS["total_work"]: "",
        JINJER_HEADERS["finalized"]: "",
    }
    row.update(overrides)
    return row


def test_compute_diffs_creates_punch_rows_when_jinjer_punches_are_blank():
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00",
        "jinjer_出勤": "",
        "出勤差分(分)": None,
        "勤務表_退勤": "18:00",
        "jinjer_退勤": "",
        "退勤差分(分)": None,
        "_source_file": "勤怠突合結果.xlsx",
    }])
    logs: list[LogEntry] = []

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row()},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )

    punch_rows = [row for row in rows if row.kind in {DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT}]
    assert [row.kind for row in punch_rows] == [DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT]
    assert [row.auto_fix_value for row in punch_rows] == ["9:00", "18:00"]
    assert punch_rows[0].warn_reason == "jinjer出勤なし / 請求勤怠側に時刻あり"
    assert punch_rows[1].warn_reason == "jinjer退勤なし / 請求勤怠側に時刻あり"


def test_compute_diffs_creates_punch_rows_when_kintai_is_blank():
    """逆向きの片側欠落: jinjer に打刻あり / 請求勤怠側に時刻なし → 要確認行を生成。"""
    from services.triage import TRIAGE_NEEDS_CHECK

    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "",
        "jinjer_出勤": "9:00",
        "出勤差分(分)": None,
        "勤務表_退勤": "",
        "jinjer_退勤": "18:00",
        "退勤差分(分)": None,
        "_source_file": "勤怠突合結果.xlsx",
    }])
    logs: list[LogEntry] = []

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row(
            **{JINJER_HEADERS["punch_in_1"]: "9:00", JINJER_HEADERS["punch_out_1"]: "18:00"})},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )

    punch_rows = [row for row in rows if row.kind in {DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT}]
    assert [row.kind for row in punch_rows] == [DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT]
    # jinjer値が表示され、請求勤怠値は空。承認時は空で上書き（jinjer打刻を消す）。
    assert [row.jinjer_value for row in punch_rows] == ["9:00", "18:00"]
    assert [row.kintai_value for row in punch_rows] == ["", ""]
    assert [row.auto_fix_value for row in punch_rows] == ["", ""]
    assert punch_rows[0].warn_reason == "請求勤怠なし / jinjer側に時刻あり"
    assert punch_rows[1].warn_reason == "請求勤怠なし / jinjer側に時刻あり"
    # コメント・休暇なしなら要確認（人が見る）に分類される
    assert punch_rows[0].triage == TRIAGE_NEEDS_CHECK
    assert punch_rows[0].judge_default == ""


def test_compute_diffs_kintai_blank_fullday_holiday_is_auto_ok():
    """全日休暇日の jinjer 打刻（請求勤怠なし）は triage が自動OK(jinjer)へ回す。"""
    from services.triage import TRIAGE_AUTO_OK, JUDGE_JINJER

    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "", "jinjer_出勤": "9:00", "出勤差分(分)": None,
        "勤務表_退勤": "", "jinjer_退勤": "18:00", "退勤差分(分)": None,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "9:00", JINJER_HEADERS["punch_out_1"]: "18:00",
        "休日休暇名1": "有給休暇", "休日休暇名1：種別": "全日",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        extra_cols,
    )
    punch_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert punch_rows
    assert punch_rows[0].triage == TRIAGE_AUTO_OK
    assert punch_rows[0].judge_default == JUDGE_JINJER


def test_compute_diffs_surfaces_jinjer_unregistered_worker():
    """jinjerに氏名が無い(ID解決不可)が請求勤怠に勤務がある人は、捨てずに
    「jinjer未登録」要確認行(1人1行)として可視化する。"""
    from quick_compare import DIFF_KIND_UNMATCHED
    from services.triage import TRIAGE_NEEDS_CHECK

    kintai_df = pd.DataFrame([
        {"氏名": "幽霊 太郎", "日付": date(2026, 5, 1),
         "勤務表_出勤": "9:00", "jinjer_出勤": "", "出勤差分(分)": None,
         "勤務表_退勤": "18:00", "jinjer_退勤": "", "退勤差分(分)": None,
         "_source_file": "月給制1.xlsx"},
        {"氏名": "幽霊 太郎", "日付": date(2026, 5, 2),
         "勤務表_出勤": "9:00", "jinjer_出勤": "", "出勤差分(分)": None,
         "勤務表_退勤": "17:30", "jinjer_退勤": "", "退勤差分(分)": None,
         "_source_file": "月給制1.xlsx"},
        # 請求勤怠も空の日（休等）は未払い候補にしない＝行に数えない
        {"氏名": "幽霊 太郎", "日付": date(2026, 5, 3),
         "勤務表_出勤": "", "jinjer_出勤": "", "出勤差分(分)": None,
         "勤務表_退勤": "", "jinjer_退勤": "", "退勤差分(分)": None,
         "_source_file": "月給制1.xlsx"},
    ])
    logs: list[LogEntry] = []

    # name_map に「幽霊 太郎」は無い → 解決不可
    rows = compute_diffs(
        kintai_df,
        {},  # jinjer index 空
        {"別人": "2018999"},
        logs,
    )

    unmatched = [r for r in rows if r.kind == DIFF_KIND_UNMATCHED]
    assert len(unmatched) == 1, "未登録者は1人1行に集約"
    r = unmatched[0]
    assert r.name == "幽霊 太郎"
    assert r.emp_id == ""
    assert r.kintai_value == "勤務2日"   # 勤務がある日のみ計上（5/3は除外）
    assert r.triage == TRIAGE_NEEDS_CHECK
    assert r.judge_default == ""
    assert r.recommend_judge == ""       # 採用ラベルは付けない
    assert "jinjer未登録" in r.warn_reason


def test_compute_diffs_unresolved_with_no_work_emits_nothing():
    """ID解決不可でも請求勤怠に勤務が無ければ未登録行は出さない（休のみの人）。"""
    from quick_compare import DIFF_KIND_UNMATCHED
    kintai_df = pd.DataFrame([{
        "氏名": "休 のみ子", "日付": date(2026, 5, 1),
        "勤務表_出勤": "", "jinjer_出勤": "", "出勤差分(分)": None,
        "勤務表_退勤": "", "jinjer_退勤": "", "退勤差分(分)": None,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(kintai_df, {}, {"別人": "2018999"}, logs)
    assert [r for r in rows if r.kind == DIFF_KIND_UNMATCHED] == []


def test_normalize_kintai_result_columns_accepts_seikyu_kintai_headers():
    df = pd.DataFrame([{
        "請求勤怠_出勤": "9:00",
        "請求勤怠_退勤": "18:00",
        "請求勤怠_総労働": "9:00",
    }])

    normalized = normalize_kintai_result_columns(df)

    assert normalized.iloc[0]["勤務表_出勤"] == "9:00"
    assert normalized.iloc[0]["勤務表_退勤"] == "18:00"
    assert normalized.iloc[0]["勤務表_総労働"] == "9:00"


def test_to_jinjer_overnight_punch_out():
    # 翌朝退勤（出勤 > 退勤）→ 24時超表記へ変換
    assert to_jinjer_overnight_punch_out("21:00", "08:15") == "32:15"
    assert to_jinjer_overnight_punch_out("17:00", "01:30") == "25:30"
    # 通常勤務（出勤 < 退勤）→ そのまま
    assert to_jinjer_overnight_punch_out("9:00", "18:00") == "18:00"
    # すでに24時超表記なら再変換しない（冪等）
    assert to_jinjer_overnight_punch_out("21:00", "32:15") == "32:15"
    # 出退勤のどちらかが空/不正ならそのまま
    assert to_jinjer_overnight_punch_out("", "08:15") == "08:15"
    assert to_jinjer_overnight_punch_out("21:00", "") == ""


def test_compute_diffs_converts_overnight_punch_out_to_jinjer_format():
    """夜勤（請求勤怠 出勤21:00→退勤翌08:15）の退勤提案値が 32:15 になる。"""
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "21:00",
        "jinjer_出勤": "21:00",
        "出勤差分(分)": 0,
        "勤務表_退勤": "08:15",
        "jinjer_退勤": "",
        "退勤差分(分)": None,
        "_source_file": "勤怠突合結果.xlsx",
    }])
    logs: list[LogEntry] = []

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row(**{JINJER_HEADERS["punch_in_1"]: "21:00"})},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )

    out_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_OUT]
    assert len(out_rows) == 1
    # 夜勤の退勤は表示も 24時超表記に揃える（請求勤怠値=32:15）。自動修正提案値も 32:15。
    assert out_rows[0].kintai_value == "32:15"
    assert out_rows[0].auto_fix_value == "32:15"


def test_diff_columns_include_manual_review_fields():
    assert "手入力休憩1" in DIFF_COLUMNS
    assert "手入力復帰1" in DIFF_COLUMNS
    assert "手入力休憩時間" in DIFF_COLUMNS
    # 2026-08-20 廃止（打刻修正の旧名「手入力修正値」も含めて出さない）
    assert "打刻修正" not in DIFF_COLUMNS
    assert "手入力修正値" not in DIFF_COLUMNS
    assert "自動修正提案値" not in DIFF_COLUMNS


def test_diff_columns_include_schedule_and_leave_fields():
    for col in ("スケジュール出勤", "スケジュール退勤", "休憩予定", "復帰予定",
                "休日休暇名1", "休日休暇名1：種別"):
        assert col in DIFF_COLUMNS
    # 旧名（jinjer 側の内部キー名と同じ）では出力しない
    assert "出勤予定" not in DIFF_COLUMNS
    assert "退勤予定" not in DIFF_COLUMNS
    # 有休/AM有休/PM有休 は汎用データ194列に該当列が無く常に空のため出力しない
    for col in ("有休", "AM有休", "PM有休"):
        assert col not in DIFF_COLUMNS


def test_resolve_jinjer_extra_columns_exact_and_partial():
    cols = [
        "名前", "*従業員ID", "*年月日",
        "出勤予定時刻", "退勤予定時刻", "休憩予定時間",
        "有休", "AM有休", "PM有休",
    ]
    resolved = resolve_jinjer_extra_columns(cols)
    assert resolved["出勤予定"] == "出勤予定時刻"
    assert resolved["退勤予定"] == "退勤予定時刻"
    assert resolved["休憩予定"] == "休憩予定時間"
    assert resolved["有休"] == "有休"
    assert resolved["AM有休"] == "AM有休"
    assert resolved["PM有休"] == "PM有休"


def test_resolve_jinjer_extra_columns_leave_is_exact_only():
    # 「有休」列が無く AM有休/PM有休 だけある場合、「有休」を部分一致で誤ヒットさせない
    cols = ["名前", "AM有休", "PM有休"]
    resolved = resolve_jinjer_extra_columns(cols)
    assert "有休" not in resolved
    assert resolved.get("AM有休") == "AM有休"
    assert resolved.get("PM有休") == "PM有休"


def test_resolve_jinjer_extra_columns_missing_returns_no_key():
    resolved = resolve_jinjer_extra_columns(["名前", "*従業員ID", "*年月日"])
    assert resolved == {}


def test_compute_diffs_transcribes_schedule_and_leave():
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00",
        "jinjer_出勤": "",
        "出勤差分(分)": None,
        "勤務表_退勤": "18:00",
        "jinjer_退勤": "",
        "退勤差分(分)": None,
        "_source_file": "勤怠突合結果.xlsx",
    }])
    logs: list[LogEntry] = []
    jrow = _jinjer_row(**{
        "出勤予定時刻": "9:00",
        "退勤予定時刻": "18:00",
        "休憩予定時刻1": "12:00",
        "復帰予定時刻1": "13:00",
        "有休": "",
        "AM有休": "1",
        "PM有休": "",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        extra_cols,
    )

    assert rows, "差異行が生成されること"
    r = rows[0]
    assert r.sched_in == "9:00"
    assert r.sched_out == "18:00"
    assert r.sched_break == "12:00"     # 休憩予定時刻1（部分一致）
    assert r.sched_break_end == "13:00"  # 復帰予定時刻1（新規）
    assert r.am_yukyu == "1"
    assert r.yukyu == ""
    assert r.pm_yukyu == ""


def test_kintai_total_minutes_prefers_actual_work():
    """請求勤怠の正味労働(勤務表_実働時間)が拘束時間より優先される。"""
    row = pd.Series({
        "勤務表_実働時間": "7:00",     # 正味（休憩控除後）
        "勤務表_総労働時間": "9:00",   # 拘束（退勤−出勤）
        "勤務表_出勤": "9:00",
        "勤務表_退勤": "18:00",
    })
    minutes, hhmm = kintai_total_minutes(row)
    assert minutes == 420
    assert hhmm == "7:00"


def test_kintai_total_minutes_fallback_when_actual_blank():
    """実働列が空なら従来の拘束時間にフォールバックする（後方互換）。"""
    row = pd.Series({"勤務表_実働時間": "", "勤務表_総労働時間": "9:00"})
    minutes, _ = kintai_total_minutes(row)
    assert minutes == 540


def test_compute_diffs_total_compares_net_vs_net():
    """総労働差異は請求勤怠の正味(実働) vs jinjer 総労働 で突合される。

    拘束時間9:00ではなく実働8:00で比較されるため、jinjer総労働8:00とは差異なし。
    """
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:00", "出勤差分(分)": 0,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "勤務表_実働時間": "8:00",  # 正味（休憩1h控除後）
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    jrow = _jinjer_row(**{
        JINJER_HEADERS["total_work"]: "8:00",
        JINJER_HEADERS["break_total"]: "1:00",
        JINJER_HEADERS["punch_in_1"]: "9:00",
        JINJER_HEADERS["punch_out_1"]: "18:00",
    })
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )
    total_rows = [r for r in rows if r.kind == DIFF_KIND_TOTAL]
    assert total_rows == []  # 実働8:00 = jinjer総労働8:00 → 差異なし


def test_format_stamp_comments():
    """打刻コメント list → 種別[方法] コメント / ... の整形。"""
    items = [
        {"type": "出勤", "method": "打刻修正申請", "comment": "KDX出社"},
        {"type": "退勤", "method": "PC", "comment": "私用のため早退"},
    ]
    assert format_stamp_comments(items) == "出勤[打刻修正申請] KDX出社 / 退勤[PC] 私用のため早退"
    assert format_stamp_comments([]) == ""
    assert format_stamp_comments([{"comment": "", "type": "出勤", "method": "PC"}]) == ""


def test_compute_diffs_attaches_stamp_comment():
    """打刻修正申請の従業員コメントが当日の差異行に併記される。"""
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:30", "出勤差分(分)": 30,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    stamp_comments = {
        ("2018057", "2026-04-01"): [
            {"type": "出勤", "method": "打刻修正申請", "comment": "KDX出社"},
        ],
    }
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row()},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        None,
        stamp_comments,
    )
    punch_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert punch_rows
    assert punch_rows[0].jinjer_stamp_comment == "出勤[打刻修正申請] KDX出社"


def test_comment_columns_left_of_judgment_and_adjacent():
    """打刻時コメント／打刻修正時コメントは隣同士で、人間判断より左（判断材料）に置く。"""
    assert "打刻時コメント" in DIFF_COLUMNS
    assert "打刻修正時コメント" in DIFF_COLUMNS
    assert "jinjer打刻コメント" not in DIFF_COLUMNS
    # 隣同士（打刻時コメント → 打刻修正時コメント）
    assert DIFF_COLUMNS.index("打刻時コメント") + 1 == DIFF_COLUMNS.index("打刻修正時コメント")
    # どちらも人間判断より左
    assert DIFF_COLUMNS.index("打刻修正時コメント") < DIFF_COLUMNS.index("人間判断")


def test_diff_columns_layout_identity_first():
    """横スクロール対策: 識別4列(従業員ID/氏名/対象日付/差異種別)が先頭。手入力は右へ。"""
    assert DIFF_COLUMNS[:4] == ["従業員ID", "氏名", "対象日付", "差異種別"]
    # 従業員IDは先頭(A列)へ移動済み
    assert DIFF_COLUMNS.index("従業員ID") == 0
    # 手入力欄は人間判断より右
    assert DIFF_COLUMNS.index("手入力休憩1") > DIFF_COLUMNS.index("人間判断")
    # 行ID・元突合結果ファイルは削除済み
    assert "行ID" not in DIFF_COLUMNS
    assert "元突合結果ファイル" not in DIFF_COLUMNS


def test_clean_punch_comment():
    """汎用データ#96『打刻時コメント』の整形（空ラベルのゴミは除去）。"""
    from quick_compare import clean_punch_comment
    assert clean_punch_comment("出勤: KDX出社 , 退勤:  , ") == "出勤: KDX出社"
    assert clean_punch_comment("出勤:  , 退勤:  , ") == ""
    assert clean_punch_comment("出勤:  , 退勤: テレワーク , ") == "退勤: テレワーク"
    assert clean_punch_comment("") == ""
    assert clean_punch_comment(None) == ""


def test_triage_column_present_and_placed():
    """確認区分（旧トリアージ区分）は差分とコメント列の間に置く。警告レベル列は廃止。"""
    assert "確認区分" in DIFF_COLUMNS
    assert "トリアージ区分" not in DIFF_COLUMNS
    assert "警告レベル" not in DIFF_COLUMNS
    assert DIFF_COLUMNS.index("差分(分)") < DIFF_COLUMNS.index("確認区分")
    assert DIFF_COLUMNS.index("確認区分") < DIFF_COLUMNS.index("打刻時コメント")


def test_schedule_leave_columns_left_of_judgment():
    """有休も判断材料なので、予定/休日休暇は人間判断より左に置く。"""
    for col in ("スケジュール出勤", "スケジュール退勤", "休憩予定", "復帰予定",
                "休日休暇名1", "休日休暇名1：種別"):
        assert DIFF_COLUMNS.index(col) < DIFF_COLUMNS.index("人間判断")
    # 復帰予定は 休憩予定 と 休日休暇名1 の間に挿入
    assert DIFF_COLUMNS.index("休憩予定") < DIFF_COLUMNS.index("復帰予定") < DIFF_COLUMNS.index("休日休暇名1")


def test_human_judgment_conditional_formatting(tmp_path):
    """人間判断の値に応じた条件付き書式（保留/jinjer勤怠）が差異一覧に入る。"""
    import html
    import zipfile
    from quick_compare import write_excel, DiffRow

    drow = DiffRow(
        row_id=1, emp_id="1001", name="山田 太郎", target_date="2026-05-01", kind="出勤",
        kintai_value="9:00", jinjer_value="9:30", diff_minutes="30",
        warn_level="INFO", warn_reason="", auto_fix_value="9:00",
        finalized="", source_file="x.xlsx",
    )
    out = tmp_path / "d.xlsx"
    write_excel(out, [drow], [], "2026-05")

    with zipfile.ZipFile(out) as z:
        sheets = " ".join(
            z.read(n).decode("utf-8") for n in z.namelist()
            if n.startswith("xl/worksheets/sheet")
        )
    assert "conditionalFormatting" in sheets
    # openpyxl は数式内の日本語を数値文字参照にエスケープするので unescape して判定
    decoded = html.unescape(sheets)
    assert '"保留"' in decoded
    assert '"jinjer勤怠"' in decoded
    # cellIs ルールが2件（保留 / jinjer勤怠）入っている
    assert decoded.count('type="cellIs"') >= 2


def _sample_diff_rows(count=2):
    from quick_compare import DiffRow
    return [
        DiffRow(
            row_id=i, emp_id=f"100{i}", name="山田 太郎", target_date=f"2026-05-0{i}", kind="出勤",
            kintai_value="9:00", jinjer_value="9:30", diff_minutes="30",
            warn_level="INFO", warn_reason="", auto_fix_value="9:00",
            finalized="", source_file="x.xlsx",
        )
        for i in range(1, count + 1)
    ]


def test_manual_input_columns_have_text_format(tmp_path):
    """手入力列（手入力休憩1/手入力復帰1）は文字列書式('@')。

    既定の書式だと Excel が「31:00」を経過時間、「7:00」を時刻型へ自動変換して保存し、
    openpyxl 経由で timedelta/time になって quick_export の転記が壊れる（2026-07-09 の事故）。
    生成時に '@' を設定しておけば入力値がそのまま文字列で保存される。
    手入力休憩時間は数式列なので '@' の対象外（付けると数式が文字列表示に化ける）。
    """
    from openpyxl import load_workbook
    from quick_compare import write_excel, MANUAL_INPUT_TEXT_COLUMNS

    drows = _sample_diff_rows()
    out = tmp_path / "d.xlsx"
    write_excel(out, drows, [], "2026-05")

    wb = load_workbook(out)
    ws = wb["差異一覧"]
    assert MANUAL_INPUT_TEXT_COLUMNS == ["手入力休憩1", "手入力復帰1"]
    for header in MANUAL_INPUT_TEXT_COLUMNS:
        col_idx = DIFF_COLUMNS.index(header) + 1
        # 全データ行（人間判断プルダウンと同じ範囲 = 2行目〜1+件数行目）に設定される
        for r_idx in (2, 3):
            assert ws.cell(row=r_idx, column=col_idx).number_format == "@", (header, r_idx)
        # ヘッダー行は対象外
        assert ws.cell(row=1, column=col_idx).number_format == "General"
    # 手入力列以外は文字列書式にしない（時刻・数値の見た目を変えない）
    judge_col = DIFF_COLUMNS.index("人間判断") + 1
    assert ws.cell(row=2, column=judge_col).number_format == "General"
    # 既存のデータ検証（人間判断プルダウン）が壊れていないこと
    dvs = ws.data_validations.dataValidation
    assert len(dvs) == 1
    assert dvs[0].formula1 == '"請求勤怠,jinjer勤怠,保留"'
    judge_letter = ws.cell(row=1, column=judge_col).column_letter
    assert str(dvs[0].sqref) == f"{judge_letter}2:{judge_letter}3"


def test_manual_break_total_is_formula(tmp_path):
    """手入力休憩時間は休憩1・復帰1から計算する数式（手入力させない）。"""
    from openpyxl import load_workbook
    from quick_compare import write_excel

    out = tmp_path / "d.xlsx"
    write_excel(out, _sample_diff_rows(), [], "2026-05")

    wb = load_workbook(out)
    ws = wb["差異一覧"]
    start_letter = ws.cell(row=1, column=DIFF_COLUMNS.index("手入力休憩1") + 1).column_letter
    end_letter = ws.cell(row=1, column=DIFF_COLUMNS.index("手入力復帰1") + 1).column_letter
    total_col = DIFF_COLUMNS.index("手入力休憩時間") + 1

    for r_idx in (2, 3):
        cell = ws.cell(row=r_idx, column=total_col)
        assert cell.value.startswith("=IF(AND(")
        # 自分の行の休憩1・復帰1を参照する
        assert f"{start_letter}{r_idx}" in cell.value
        assert f"{end_letter}{r_idx}" in cell.value
        # 数式列に '@' を付けると Excel で文字列表示に化けるため General のまま
        assert cell.number_format == "General"


def test_manual_break_total_locked_and_sheet_protected(tmp_path):
    """手入力休憩時間だけロックし、他のセルは今までどおり編集できる。"""
    from openpyxl import load_workbook
    from quick_compare import write_excel

    out = tmp_path / "d.xlsx"
    write_excel(out, _sample_diff_rows(), [], "2026-05")

    wb = load_workbook(out)
    ws = wb["差異一覧"]
    total_col = DIFF_COLUMNS.index("手入力休憩時間") + 1
    judge_col = DIFF_COLUMNS.index("人間判断") + 1

    assert ws.protection.sheet is True
    assert ws.protection.autoFilter is False   # False=許可（フィルタの絞り込みは使える）
    for r_idx in (2, 3):
        assert ws.cell(row=r_idx, column=total_col).protection.locked is True
        # 人間判断・手入力休憩1 は入力できる
        assert ws.cell(row=r_idx, column=judge_col).protection.locked is False
        assert ws.cell(row=r_idx, column=DIFF_COLUMNS.index("手入力休憩1") + 1).protection.locked is False
    # ヘッダー行はロックのまま（列名を書き換えられると手順3の転記が壊れる）
    assert ws.cell(row=1, column=1).protection.locked is True


def test_sheet_order_and_header_style(tmp_path):
    """サマリは最後尾。開いた時は差異一覧。ヘッダーは12pt・折り返し。"""
    from openpyxl import load_workbook
    from quick_compare import write_excel

    out = tmp_path / "d.xlsx"
    write_excel(out, _sample_diff_rows(), [], "2026-05")

    wb = load_workbook(out)
    assert wb.sheetnames[0] == "差異一覧"
    assert wb.sheetnames[-1] == "サマリ"
    assert wb.active.title == "差異一覧"

    ws = wb["差異一覧"]
    head = ws.cell(row=1, column=1)
    assert head.font.size == 12
    assert head.font.bold is True
    assert head.alignment.wrap_text is True
    assert ws.row_dimensions[1].height == 32
    assert ws.cell(row=2, column=1).font.size == 12


def test_recommend_judge_label_new_rule():
    """新ルール: コメントあり→jinjer勤怠 / 差分≧しきい値or片側欠落→保留 / それ以外→請求勤怠。"""
    # コメントなし & 差分がしきい値(10)未満 → 請求勤怠
    assert recommend_judge_label(DIFF_KIND_PUNCH_IN, "9", 10) == "請求勤怠"
    assert recommend_judge_label(DIFF_KIND_PUNCH_OUT, "-9", 10) == "請求勤怠"
    # コメントなし & 差分がしきい値以上 → 保留（10ちょうども「以上」で保留）
    assert recommend_judge_label(DIFF_KIND_PUNCH_IN, "10", 10) == "保留"
    assert recommend_judge_label(DIFF_KIND_PUNCH_OUT, "11", 10) == "保留"
    assert recommend_judge_label(DIFF_KIND_PUNCH_IN, "-30", 10) == "保留"
    # 片側欠落（差分なし）→ 保留（安全側）
    assert recommend_judge_label(DIFF_KIND_PUNCH_IN, "", 10) == "保留"
    # コメントあり → jinjer勤怠（差分の大小に関わらず）
    assert recommend_judge_label(DIFF_KIND_PUNCH_IN, "5", 10, has_comment=True) == "jinjer勤怠"
    assert recommend_judge_label(DIFF_KIND_PUNCH_OUT, "120", 10, has_comment=True) == "jinjer勤怠"
    # 総労働 → 対象外＝jinjer勤怠（コメント有無に関わらず変更なし）
    assert recommend_judge_label(DIFF_KIND_TOTAL, "5", 10) == "jinjer勤怠"
    assert recommend_judge_label(DIFF_KIND_TOTAL, "5", 10, has_comment=True) == "jinjer勤怠"


def test_recommend_judge_label_set_on_rows():
    """compute_diffs が各差異行に recommend_judge を設定する（新ルール）。"""
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:05", "jinjer_出勤": "9:00", "出勤差分(分)": 5,
        "勤務表_退勤": "20:00", "jinjer_退勤": "18:00", "退勤差分(分)": 120,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row(
            **{JINJER_HEADERS["punch_in_1"]: "9:00", JINJER_HEADERS["punch_out_1"]: "18:00"})},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        threshold_minutes=10,
    )
    in_row = next(r for r in rows if r.kind == DIFF_KIND_PUNCH_IN)
    out_row = next(r for r in rows if r.kind == DIFF_KIND_PUNCH_OUT)
    assert in_row.recommend_judge == "請求勤怠"   # 差分5分 < 10 / コメントなし
    assert out_row.recommend_judge == "保留"      # 差分120分 ≥ 10 / コメントなし


def test_recommend_judge_label_comment_makes_jinjer():
    """打刻時コメントがある出退勤差異は、差分が大きくても jinjer勤怠 を提案する。"""
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "9:30", JINJER_HEADERS["punch_out_1"]: "18:00",
        "打刻時コメント": "出勤: 客先直行のため",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾", "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:30", "出勤差分(分)": 30,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs, extra_cols, threshold_minutes=10,
    )
    in_row = next(r for r in rows if r.kind == DIFF_KIND_PUNCH_IN)
    assert in_row.punch_comment  # コメントが転記されている
    assert in_row.recommend_judge == "jinjer勤怠"  # コメントあり → jinjer勤怠


def test_long_work_over_10h_no_longer_warns():
    """>10h の長時間労働注意喚起は廃止。差分なしの総労働行は生成しない。集計不整合は残す。"""
    # 純粋な >10h（差分なし）→ 警告なし（None）
    assert classify_total_work(_jinjer_row(**{
        JINJER_HEADERS["total_work"]: "10:30",
        JINJER_HEADERS["punch_in_1"]: "9:00",
    })) is None
    # 集計不整合（出勤打刻あり / 総労働0:00）は引き続き DANGER
    res = classify_total_work(_jinjer_row(**{
        JINJER_HEADERS["total_work"]: "0:00",
        JINJER_HEADERS["punch_in_1"]: "9:00",
    }))
    assert res is not None and res[0] == "DANGER"

    # 請求=jinjer（差分なし）で >10h の行は総労働行を生成しない
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:00", "出勤差分(分)": 0,
        "勤務表_退勤": "20:30", "jinjer_退勤": "20:30", "退勤差分(分)": 0,
        "勤務表_実働時間": "10:30",
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row(**{
            JINJER_HEADERS["total_work"]: "10:30",
            JINJER_HEADERS["punch_in_1"]: "9:00",
            JINJER_HEADERS["punch_out_1"]: "20:30",
        })},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )
    assert rows == []  # >10h 単独では行を作らない


def test_compute_diffs_sets_triage_default():
    """コメント無し・小差分の出勤差異 → 自動採用(請求勤怠) が既定セットされる。"""
    from services.triage import TRIAGE_AUTO_KINTAI

    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾", "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:05", "jinjer_出勤": "9:00", "出勤差分(分)": 5,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): _jinjer_row()},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )
    punch = [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert punch
    assert punch[0].triage == TRIAGE_AUTO_KINTAI
    assert punch[0].judge_default == "請求勤怠"


def test_load_stamp_correction_reasons(tmp_path):
    """申請データCSVの「理由」を (従業員ID,日付ISO)→コメント に読み込む。理由空は除外。"""
    import csv as _csv
    from pathlib import Path

    p = tmp_path / "app.csv"
    with open(p, "w", encoding="cp932", newline="") as f:
        w = _csv.writer(f)
        w.writerow([
            "No.", "従業員ID", "従業員名", "所属グループ", "打刻グループ名", "日付",
            "申請種別", "申請前", "申請内容", "理由", "ステータス", "承認者コメント",
            "申請日時", "対応日時",
        ])
        w.writerow(["1", "2018012", "加藤 昌", "本部", "T2課", "2026年05月07日",
                    "打刻修正申請", "なし", "出退勤 15:30~18:30", "午前休", "申請中", "", "", ""])
        w.writerow(["2", "2018012", "加藤 昌", "本部", "T2課", "2026年05月16日",
                    "打刻修正申請", "なし", "出退勤 09:00~17:30", "研修", "承認済", "", "", ""])
        w.writerow(["3", "2018012", "加藤 昌", "本部", "T2課", "2026年05月18日",
                    "打刻修正申請", "なし", "出退勤", "", "申請中", "", "", ""])  # 理由空→除外

    logs: list[LogEntry] = []
    d = load_stamp_correction_reasons(Path(p), logs)
    assert d[("2018012", "2026-05-07")][0]["comment"] == "午前休（申請中）"
    assert d[("2018012", "2026-05-16")][0]["comment"] == "研修（承認済）"
    assert ("2018012", "2026-05-18") not in d
    assert format_stamp_comments(d[("2018012", "2026-05-07")]) == "午前休（申請中）"


def test_holiday_columns_transcribed_from_jinjer():
    """汎用データの 休日休暇名1 / 休日休暇名1：種別 が差異行へ転記される。"""
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:30", "出勤差分(分)": 30,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    jrow = _jinjer_row(**{"休日休暇名1": "有給休暇", "休日休暇名1：種別": "1"})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    assert extra_cols.get("休日休暇名1") == "休日休暇名1"
    assert extra_cols.get("休日休暇名1：種別") == "休日休暇名1：種別"

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        extra_cols,
    )
    punch_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert punch_rows
    assert punch_rows[0].holiday_name1 == "有給休暇"
    assert punch_rows[0].holiday_name1_type == "1"


def test_punch_comment_transcribed_from_jinjer():
    """汎用データの『打刻時コメント』(#96) が整形のうえ差異行へ転記される。"""
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:30", "出勤差分(分)": 30,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    jrow = _jinjer_row(**{"打刻時コメント": "出勤: KDX出社のため、10:00時差出勤 , 退勤:  , "})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    assert extra_cols.get("打刻時コメント") == "打刻時コメント"

    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
        extra_cols,
    )
    punch_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert punch_rows
    assert punch_rows[0].punch_comment == "出勤: KDX出社のため、10:00時差出勤"


# =============================================================================
# 打刻コメントの定型語除外（打刻忘れ／打刻漏れ／打刻修正）
# =============================================================================

def test_strip_punch_noise_words_removes_only_keywords():
    # 語そのものだけなら空に
    assert strip_punch_noise_words("打刻忘れ") == ""
    assert strip_punch_noise_words("打刻漏れ") == ""
    assert strip_punch_noise_words("打刻修正") == ""
    # 文中の語だけ除去し、残りは保持
    assert strip_punch_noise_words("打刻忘れ 客先で終日対応") == "客先で終日対応"
    assert strip_punch_noise_words("客先直行のため打刻修正") == "客先直行のため"
    # 空・NaN
    assert strip_punch_noise_words("") == ""
    assert strip_punch_noise_words(None) == ""
    assert strip_punch_noise_words("nan") == ""


def test_clean_punch_comment_strips_noise_per_side():
    # 出勤側が定型語のみ → 出勤ラベルは落ち、退勤側の本文は残る
    out = clean_punch_comment("出勤: 打刻忘れ , 退勤: 私用で早退")
    assert out == "退勤: 私用で早退"
    # 退勤側が定型語のみ → 退勤ラベルは落ち、出勤側は残る
    out2 = clean_punch_comment("出勤: 客先直行 , 退勤: 打刻修正")
    assert out2 == "出勤: 客先直行"
    # 両方本文あり（定型語なし）はそのまま
    out3 = clean_punch_comment("出勤: 直行 , 退勤: 直帰")
    assert out3 == "出勤: 直行 / 退勤: 直帰"


def test_format_stamp_comments_strips_noise():
    # 除外はコメント本文に対して行う（打刻方法ラベル [打刻修正申請] 等のメタ情報は対象外）。
    items = [
        {"type": "出勤", "method": "PC", "comment": "打刻忘れ KDX出社"},
        {"type": "退勤", "method": "PC", "comment": "打刻修正"},
    ]
    out = format_stamp_comments(items)
    # 1件目: 本文「KDX出社」が残り、定型語は消える
    assert "KDX出社" in out
    assert "打刻忘れ" not in out
    # 2件目: 定型語のみ → その項目ごと消える（退勤の項目が出ない）
    assert "退勤" not in out


# =============================================================================
# 夜勤の退勤を 24時超表記(33:00 等)で表示
# =============================================================================

def test_overnight_display_value_by_actual_in():
    # 実出勤 20:45 → 退勤 09:00 は翌朝 ＝ 33:00
    assert overnight_display_value("09:00", "20:45") == "33:00"
    assert overnight_display_value("01:00", "22:00") == "25:00"
    # 通常勤務（退勤 > 出勤）は変換しない
    assert overnight_display_value("18:00", "09:00") == "18:00"
    # すでに24時超表記は冪等
    assert overnight_display_value("33:00", "20:45") == "33:00"
    # 空・不正はそのまま空
    assert overnight_display_value("", "20:45") == ""


def test_overnight_display_value_by_schedule_when_in_missing():
    # 実出勤が無くても、退勤予定が24時超(夜勤)で退勤が翌朝側なら +24h
    assert overnight_display_value("09:00", "", sched_out="33:00", sched_in="20:45") == "33:00"
    # 退勤予定が通常(日勤)なら変換しない
    assert overnight_display_value("09:00", "", sched_out="17:30", sched_in="09:00") == "09:00"


def test_compute_diffs_night_shift_punch_out_shown_as_over24():
    """夜勤(出勤20:45/退勤予定33:00)の退勤09:00 は差異一覧で 33:00 表示。差分は不変。"""
    kintai_df = pd.DataFrame([{
        "氏名": "田中 一郎", "日付": date(2026, 5, 7),
        "勤務表_出勤": "20:45", "jinjer_出勤": "20:45", "出勤差分(分)": 0,
        "勤務表_退勤": "09:00", "jinjer_退勤": "09:01", "退勤差分(分)": 1,
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "20:45", JINJER_HEADERS["punch_out_1"]: "09:01",
        "出勤予定時刻": "20:45", "退勤予定時刻": "33:00",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-05-07"): jrow},
        {"田中 一郎": "2018057", "田中一郎": "2018057"},
        logs, extra_cols, threshold_minutes=10,
    )
    out_row = next(r for r in rows if r.kind == DIFF_KIND_PUNCH_OUT)
    assert out_row.kintai_value == "33:00"
    assert out_row.jinjer_value == "33:01"
    assert out_row.diff_minutes == "1"  # 差分は不変


def test_total_diff_explained_by_punches_is_suppressed():
    """出退勤のズレだけで説明がつく総労働時間差異は行を出さない（重複アラート防止）。

    例: 退勤183分差 → 総労働も183分差。総労働行は退勤行と同じ情報のため冗長。
    """
    kintai_df = pd.DataFrame([{
        "氏名": "畑中 竜哉", "日付": date(2026, 6, 9),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:00", "出勤差分(分)": 0,
        "勤務表_退勤": "20:30", "jinjer_退勤": "23:33", "退勤差分(分)": 183,
        "勤務表_実働時間": "10:30",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "9:00",
        JINJER_HEADERS["punch_out_1"]: "23:33",
        JINJER_HEADERS["total_work"]: "13:33",
    })
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-09"): jrow},
        {"畑中 竜哉": "2018057", "畑中竜哉": "2018057"},
        logs,
    )
    kinds = [r.kind for r in rows]
    assert DIFF_KIND_PUNCH_OUT in kinds
    assert DIFF_KIND_TOTAL not in kinds  # 退勤行で説明がつくため総労働行は出さない


def test_total_diff_from_break_difference_still_reported():
    """出退勤が一致していて総労働だけ違う（＝休憩の違い）場合は従来どおり行を出す。"""
    kintai_df = pd.DataFrame([{
        "氏名": "上原 奏吾", "日付": date(2026, 6, 10),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:00", "出勤差分(分)": 0,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
        "勤務表_実働時間": "8:00",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "9:00",
        JINJER_HEADERS["punch_out_1"]: "18:00",
        JINJER_HEADERS["total_work"]: "8:30",
    })
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-10"): jrow},
        {"上原 奏吾": "2018057", "上原奏吾": "2018057"},
        logs,
    )
    total_rows = [r for r in rows if r.kind == DIFF_KIND_TOTAL]
    assert len(total_rows) == 1
    assert total_rows[0].diff_minutes == "-30"


def test_month_end_tokki_keeps_punch_in_but_skips_out_and_total():
    """月末日跨ぎの特記行: 出勤は通常突合、退勤・総労働は出さず注記行を1行出す。"""
    kintai_df = pd.DataFrame([{
        "氏名": "山口 太雅", "日付": date(2026, 6, 30),
        "勤務表_出勤": "16:45", "jinjer_出勤": "17:15", "出勤差分(分)": 30,
        "勤務表_退勤": "00:00", "jinjer_退勤": "09:33", "退勤差分(分)": None,
        "勤務表_実働時間": "6:30",
        "特記": "月末日跨ぎ勤務(請求側の翌月分未取得のため退勤・総労働は翌月確認)",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "17:15",
        JINJER_HEADERS["punch_out_1"]: "33:33",
        JINJER_HEADERS["total_work"]: "15:03",
    })
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-30"): jrow},
        {"山口 太雅": "2018057", "山口太雅": "2018057"},
        logs,
    )
    kinds = [r.kind for r in rows]
    assert kinds.count(DIFF_KIND_PUNCH_IN) == 1
    assert DIFF_KIND_PUNCH_OUT not in kinds  # 24:00は打切りであり実退勤ではない
    note_rows = [r for r in rows if r.kind == DIFF_KIND_TOTAL]
    assert len(note_rows) == 1
    assert "月末日跨ぎ" in note_rows[0].warn_reason
    assert note_rows[0].auto_fix_value == ""


def test_prev_month_tail_matched_emits_overnight_autofix():
    """前月末日で突合した夜勤後半: 退勤の書き戻し値は24時超表記（例 33:30）になる。"""
    kintai_df = pd.DataFrame([{
        "氏名": "大堀 広智", "日付": date(2026, 5, 31),
        "勤務表_出勤": "16:45", "jinjer_出勤": "16:45", "出勤差分(分)": 0,
        "勤務表_退勤": "09:30", "jinjer_退勤": "09:33", "退勤差分(分)": 3,
        "特記": "前月末夜勤の後半(前月末日で突合)",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["date"]: "2026/5/31",
        JINJER_HEADERS["punch_in_1"]: "16:45",
        JINJER_HEADERS["punch_out_1"]: "33:33",
        JINJER_HEADERS["total_work"]: "15:03",
    })
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-05-31"): jrow},
        {"大堀 広智": "2018057", "大堀広智": "2018057"},
        logs,
    )
    kinds = [r.kind for r in rows]
    assert DIFF_KIND_TOTAL not in kinds  # 後半のみの実働のため総労働は突合しない
    out_row = next(r for r in rows if r.kind == DIFF_KIND_PUNCH_OUT)
    assert out_row.target_date == "2026-05-31"  # 書き戻し先は勤務開始日
    assert out_row.auto_fix_value == "33:30"    # jinjerインポートは24時超表記


def test_prev_month_tail_unmatched_tokki_note_only():
    """前月末日のjinjer行が無い孤児行: 片側欠落行を出さず、注記行のみ出す。"""
    kintai_df = pd.DataFrame([{
        "氏名": "大堀 広智", "日付": date(2026, 6, 1),
        "勤務表_出勤": "00:00", "jinjer_出勤": "", "出勤差分(分)": None,
        "勤務表_退勤": "09:30", "jinjer_退勤": "", "退勤差分(分)": None,
        "特記": "前月末夜勤の後半(jinjer側は前月末日行に記録・自動書戻し不可。jinjerダウンロード範囲に前月末日を含めると突合可能)",
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-01"): _jinjer_row()},
        {"大堀 広智": "2018057", "大堀広智": "2018057"},
        logs,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_TOTAL]
    assert "自動書戻し不可" in rows[0].warn_reason
    assert rows[0].auto_fix_value == ""


def test_fieldglass_no_time_tokki_note_only():
    """Fieldglass時刻なし行: 00:00打刻や総労働24:00の偽差異を出さず、注記行のみ出す。"""
    kintai_df = pd.DataFrame([{
        "氏名": "福家 寛昭", "日付": date(2026, 6, 30),
        "勤務表_出勤": "", "jinjer_出勤": "08:30", "出勤差分(分)": None,
        "勤務表_退勤": "", "jinjer_退勤": "17:12", "退勤差分(分)": None,
        "特記": "Fieldglass時刻なし(00:00-00:00プレースホルダ行)",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "08:30",
        JINJER_HEADERS["punch_out_1"]: "17:12",
        JINJER_HEADERS["total_work"]: "07:30",
    })
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-30"): jrow},
        {"福家 寛昭": "2018057", "福家寛昭": "2018057"},
        logs,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_TOTAL]
    assert "Fieldglass時刻なし" in rows[0].warn_reason
    assert "Fieldglass" in rows[0].warn_reason
    assert rows[0].jinjer_value == "08:30〜17:12"
    assert rows[0].auto_fix_value == ""
    # 注記行は休暇情報等に関わらず必ず要確認（自動OKに紛れて見落とされない）
    from services.triage import TRIAGE_NEEDS_CHECK
    assert rows[0].triage == TRIAGE_NEEDS_CHECK
    assert rows[0].recommend_judge == ""


def test_month_end_sched_deemed_punch_out_comparison():
    """月末日跨ぎ: 退勤予定（スケジュール）を暫定の正としてjinjer実績と突合する。

    運用ルール: 月末の日跨ぎ勤務はスケジュール勤怠を暫定で働いたとみなし当月計上。
    jinjerは月をまたぐ範囲をダウンロードできないため、翌月送りにはしない。
    """
    from services.triage import TRIAGE_INFO_ONLY

    kintai_df = pd.DataFrame([{
        "氏名": "山口 太雅", "日付": date(2026, 6, 30),
        "勤務表_出勤": "16:45", "jinjer_出勤": "17:15", "出勤差分(分)": 30,
        "勤務表_退勤": "00:00", "jinjer_退勤": "09:33", "退勤差分(分)": None,
        "勤務表_実働時間": "6:30",
        "特記": "月末日跨ぎ勤務(請求側の翌月分は取得不可。スケジュール勤怠を暫定の正として当月計上)",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "17:15",
        JINJER_HEADERS["punch_out_1"]: "33:33",
        JINJER_HEADERS["total_work"]: "15:03",
        "出勤予定時刻": "16:45",
        "退勤予定時刻": "33:30",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-30"): jrow},
        {"山口 太雅": "2018057", "山口太雅": "2018057"},
        logs, extra_cols,
    )
    # 出勤は通常突合（30分差）
    in_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_IN]
    assert len(in_rows) == 1
    # 退勤はスケジュール(33:30) vs jinjer実績(33:33) の暫定突合
    out_rows = [r for r in rows if r.kind == DIFF_KIND_PUNCH_OUT]
    assert len(out_rows) == 1
    assert out_rows[0].kintai_value == "33:30"   # 承認するとこの値が書き戻される
    assert out_rows[0].jinjer_value == "33:33"
    assert out_rows[0].diff_minutes == "-3"
    assert "スケジュール退勤を暫定の正" in out_rows[0].warn_reason
    # 注記行は参考のみ（退勤を突合できたため判断不要）
    note_rows = [r for r in rows if r.kind == DIFF_KIND_TOTAL]
    assert len(note_rows) == 1
    assert note_rows[0].triage == TRIAGE_INFO_ONLY


def test_month_end_sched_equal_emits_no_punch_out_row():
    """月末日跨ぎでスケジュール退勤とjinjer実績が一致していれば退勤行は出ない。"""
    kintai_df = pd.DataFrame([{
        "氏名": "及川 航平", "日付": date(2026, 6, 30),
        "勤務表_出勤": "16:45", "jinjer_出勤": "16:45", "出勤差分(分)": 0,
        "勤務表_退勤": "00:00", "jinjer_退勤": "09:30", "退勤差分(分)": None,
        "特記": "月末日跨ぎ勤務(請求側の翌月分は取得不可。スケジュール勤怠を暫定の正として当月計上)",
        "_source_file": "x.xlsx",
    }])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "16:45",
        JINJER_HEADERS["punch_out_1"]: "33:30",
        "出勤予定時刻": "16:45",
        "退勤予定時刻": "33:30",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-06-30"): jrow},
        {"及川 航平": "2018057", "及川航平": "2018057"},
        logs, extra_cols,
    )
    kinds = [r.kind for r in rows]
    assert DIFF_KIND_PUNCH_IN not in kinds
    assert DIFF_KIND_PUNCH_OUT not in kinds  # スケジュールどおり → 差異なし
    assert kinds.count(DIFF_KIND_TOTAL) == 1  # 参考の注記のみ


def test_prev_month_tail_note_is_info_only():
    """前月末夜勤の後半（前月分で暫定計上済み）の注記は参考のみ＝判断不要。"""
    from services.triage import TRIAGE_INFO_ONLY

    kintai_df = pd.DataFrame([{
        "氏名": "大堀 広智", "日付": date(2026, 7, 1),
        "勤務表_出勤": "00:00", "jinjer_出勤": "", "出勤差分(分)": None,
        "勤務表_退勤": "09:30", "jinjer_退勤": "", "退勤差分(分)": None,
        "特記": "前月末夜勤の後半(前月分でスケジュール勤怠により暫定計上済み・自動書戻し不可)",
        "_source_file": "x.xlsx",
    }])
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-07-01"): _jinjer_row()},
        {"大堀 広智": "2018057", "大堀広智": "2018057"},
        logs,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_TOTAL]
    assert rows[0].triage == TRIAGE_INFO_ONLY
    assert rows[0].recommend_judge == ""


# ---- 欠勤転記・全日休暇/法休所休の行抑制（2026-07-10 谷津指定） ----

# jinjer 汎用データの実ヘッダー（注記付きの長い名前）
ABSENCE_HEADER = "勤務状況（0:未打刻1:欠勤）"
HOLIDAY_KIND_HEADER = (
    "休日（0:法定休日1:所定休日2:法休(振替休出)3:所休(振替休出)"
    "4:法休(時間外休出)5:所休(時間外休出)）"
)

_NAME_MAP = {"上原 奏吾": "2018057", "上原奏吾": "2018057"}


def _kintai_row(**overrides):
    row = {
        "氏名": "上原 奏吾",
        "日付": date(2026, 4, 1),
        "勤務表_出勤": "", "jinjer_出勤": "", "出勤差分(分)": None,
        "勤務表_退勤": "", "jinjer_退勤": "", "退勤差分(分)": None,
        "_source_file": "x.xlsx",
    }
    row.update(overrides)
    return row


def test_resolve_jinjer_extra_columns_absence_and_holiday_kind():
    """勤務状況(CT列)・休日区分(U列)は注記付きの実ヘッダーに部分一致で解決する。"""
    resolved = resolve_jinjer_extra_columns(["名前", ABSENCE_HEADER, HOLIDAY_KIND_HEADER])
    assert resolved["勤務状況"] == ABSENCE_HEADER
    assert resolved["休日区分"] == HOLIDAY_KIND_HEADER


def test_compute_diffs_transcribes_jinjer_absence_row():
    """勤務状況=1（欠勤）の日は、突合結果に行が無くても欠勤行として転記する。"""
    from services.triage import TRIAGE_NEEDS_CHECK

    # 突合結果には 4/2 の差異なし行だけがあり、欠勤日 4/1 の行は無い
    kintai_df = pd.DataFrame([_kintai_row(**{
        "日付": date(2026, 4, 2),
        "勤務表_出勤": "9:00", "jinjer_出勤": "9:00", "出勤差分(分)": 0,
        "勤務表_退勤": "18:00", "jinjer_退勤": "18:00", "退勤差分(分)": 0,
    })])
    absent = _jinjer_row(**{ABSENCE_HEADER: "1"})
    normal = _jinjer_row(**{
        JINJER_HEADERS["date"]: "2026/4/2",
        JINJER_HEADERS["punch_in_1"]: "9:00",
        JINJER_HEADERS["punch_out_1"]: "18:00",
        ABSENCE_HEADER: "0",
    })
    extra_cols = resolve_jinjer_extra_columns(list(absent.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df,
        {("2018057", "2026-04-01"): absent, ("2018057", "2026-04-02"): normal},
        _NAME_MAP, logs, extra_cols,
    )
    ab = [r for r in rows if r.kind == DIFF_KIND_ABSENCE]
    assert len(ab) == 1
    assert ab[0].target_date == "2026-04-01"
    assert ab[0].jinjer_value == "欠勤"
    assert ab[0].warn_level == "INFO"  # 請求勤怠にも勤務なし＝整合
    assert ab[0].triage == TRIAGE_NEEDS_CHECK  # ただし欠勤の妥当性は人が見る
    assert ab[0].recommend_judge == ""
    assert ab[0].auto_fix_value == ""


def test_compute_diffs_absence_with_billing_work_relabels_punch_rows():
    """欠勤なのに請求勤怠に勤務がある日: 出退勤の片側欠落行の差異種別（D列）が「欠勤」になる。

    日単位の転記行は重複するため追加しない。書き戻し先は警告理由と punch_side で保持する。
    """
    from services.triage import TRIAGE_NEEDS_CHECK, JUDGE_HOLD

    kintai_df = pd.DataFrame([_kintai_row(**{
        "勤務表_出勤": "9:00", "勤務表_退勤": "18:00",
    })])
    jrow = _jinjer_row(**{ABSENCE_HEADER: "1"})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_ABSENCE, DIFF_KIND_ABSENCE]
    in_row, out_row = rows
    assert (in_row.punch_side, out_row.punch_side) == ("in", "out")
    assert in_row.warn_reason.startswith("jinjer出勤なし(欠勤・未払いの恐れ)")
    assert out_row.warn_reason.startswith("jinjer退勤なし(欠勤・未払いの恐れ)")
    assert (in_row.kintai_value, out_row.kintai_value) == ("9:00", "18:00")
    # 書き戻し値・トリアージ・提案は出勤/退勤の片側欠落と同じ扱い
    assert (in_row.auto_fix_value, out_row.auto_fix_value) == ("9:00", "18:00")
    assert all(r.triage == TRIAGE_NEEDS_CHECK for r in rows)
    assert all(r.recommend_judge == JUDGE_HOLD for r in rows)


def test_compute_diffs_no_punch_status_zero_relabels_punch_rows():
    """勤務状況=0（未打刻）の日: 出退勤の片側欠落行の差異種別（D列）が「未打刻」になる。"""
    from quick_compare import DIFF_KIND_NO_PUNCH
    from services.triage import JUDGE_HOLD

    kintai_df = pd.DataFrame([_kintai_row(**{
        "勤務表_出勤": "9:00", "勤務表_退勤": "18:00",
    })])
    jrow = _jinjer_row(**{ABSENCE_HEADER: "0"})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_NO_PUNCH, DIFF_KIND_NO_PUNCH]
    assert rows[0].warn_reason.startswith("jinjer出勤なし(未打刻)")
    assert rows[1].warn_reason.startswith("jinjer退勤なし(未打刻)")
    assert all(r.recommend_judge == JUDGE_HOLD for r in rows)
    # 欠勤の日単位転記行は出ない
    assert [r for r in rows if r.kind == DIFF_KIND_ABSENCE] == []


def test_compute_diffs_blank_status_keeps_punch_kinds():
    """勤務状況が空（打刻あり等）の片側欠落は従来どおり出勤/退勤のまま。"""
    kintai_df = pd.DataFrame([_kintai_row(**{
        "勤務表_出勤": "9:00", "勤務表_退勤": "18:00",
    })])
    jrow = _jinjer_row(**{ABSENCE_HEADER: ""})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT]
    assert rows[0].warn_reason == "jinjer出勤なし / 請求勤怠側に時刻あり"


def test_compute_diffs_suppresses_fullday_leave_when_billing_zero():
    """振休/代休/年次有給の全日で請求勤怠が0/記載なしの日は差異行を一切出さない。"""
    for leave_name in ("振休", "代休", "年次有給"):
        # SAP等が休暇日に出す 0:00 行（従来は「jinjer出勤なし/請求勤怠側に時刻あり」が出ていた）
        kintai_df = pd.DataFrame([_kintai_row(**{
            "勤務表_出勤": "00:00", "勤務表_退勤": "00:00", "勤務表_実働時間": "00:00",
        })])
        jrow = _jinjer_row(**{"休日休暇名1": leave_name, "休日休暇名1：種別": "全日"})
        extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
        logs: list[LogEntry] = []
        rows = compute_diffs(
            kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
        )
        assert rows == [], f"{leave_name}(全日)・請求勤怠0の日は行を出さない"


def test_compute_diffs_keeps_rows_when_fullday_leave_but_billing_has_work():
    """全日休暇でも請求勤怠に勤務があれば従来どおり差異行を出す（未払い防止）。"""
    kintai_df = pd.DataFrame([_kintai_row(**{
        "勤務表_出勤": "9:00", "勤務表_退勤": "18:00",
    })])
    jrow = _jinjer_row(**{"休日休暇名1": "年次有給", "休日休暇名1：種別": "全日"})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert [r.kind for r in rows] == [DIFF_KIND_PUNCH_IN, DIFF_KIND_PUNCH_OUT]


def test_compute_diffs_keeps_rows_for_other_fullday_leave_names():
    """指定外の全日休暇（生理休暇等）は従来どおり行を出す（triageの自動OKに任せる）。"""
    from services.triage import TRIAGE_AUTO_OK

    kintai_df = pd.DataFrame([_kintai_row(**{
        "jinjer_出勤": "9:00", "jinjer_退勤": "18:00",
    })])
    jrow = _jinjer_row(**{
        JINJER_HEADERS["punch_in_1"]: "9:00", JINJER_HEADERS["punch_out_1"]: "18:00",
        "休日休暇名1": "生理休暇", "休日休暇名1：種別": "全日",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert rows
    assert all(r.triage == TRIAGE_AUTO_OK for r in rows if r.kind == DIFF_KIND_PUNCH_IN)


def test_compute_diffs_suppresses_legal_and_scheduled_holiday_when_billing_zero():
    """法休(0)/所休(1)で請求勤怠が0/記載なしの日は差異行を出さない（SAPの休日0:00行対策）。"""
    for code in ("0", "1"):
        kintai_df = pd.DataFrame([_kintai_row(**{
            "勤務表_出勤": "00:00", "勤務表_退勤": "00:00", "勤務表_実働時間": "00:00",
        })])
        jrow = _jinjer_row(**{HOLIDAY_KIND_HEADER: code})
        extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
        logs: list[LogEntry] = []
        rows = compute_diffs(
            kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
        )
        assert rows == [], f"休日区分={code}・請求勤怠0の日は行を出さない"


def test_compute_diffs_suppresses_holiday_jinjer_punch_when_billing_empty():
    """法休/所休の日のjinjer打刻（請求勤怠なし）も行を出さない（谷津指定）。"""
    kintai_df = pd.DataFrame([_kintai_row(**{
        "jinjer_出勤": "9:00", "jinjer_退勤": "17:45",
    })])
    jrow = _jinjer_row(**{
        HOLIDAY_KIND_HEADER: "1",
        JINJER_HEADERS["punch_in_1"]: "9:00", JINJER_HEADERS["punch_out_1"]: "17:45",
        JINJER_HEADERS["total_work"]: "7:45",
    })
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert rows == []


def test_compute_diffs_keeps_holiday_work_codes():
    """振替休出(2,3)・時間外休出(4,5)は働く日のため通常どおり突合する。"""
    kintai_df = pd.DataFrame([_kintai_row(**{
        "勤務表_出勤": "00:00", "勤務表_退勤": "00:00", "勤務表_実働時間": "00:00",
    })])
    jrow = _jinjer_row(**{HOLIDAY_KIND_HEADER: "3"})
    extra_cols = resolve_jinjer_extra_columns(list(jrow.keys()))
    logs: list[LogEntry] = []
    rows = compute_diffs(
        kintai_df, {("2018057", "2026-04-01"): jrow}, _NAME_MAP, logs, extra_cols,
    )
    assert rows, "休日区分=3（所休・振替休出）は抑制しない"
