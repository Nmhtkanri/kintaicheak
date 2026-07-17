"""経費統合一覧表の生成（経費マクロ移植 P1a）

経費一覧表マクロ（`経費利用履歴 RevN.xlsm` の VBA 群）の「4ソース取込 → 経費統合一覧表(34列)生成」
を Python へ移植したもの。jinjer / e-staffing / SAP Fieldglass / freee の各生CSVを取り込み、
社員番号照合・SAP税抜補正などの前処理をして 34 列の統合一覧表を組み立て、Excel 出力する。
同じブックに「経路突合チェック」シート（通勤経路上の移動なのに通勤系以外で申請された交通費の検出）も出す。

移植の合格ライン: 出力する統合一覧表が現行マクロの「経費統合一覧表」シートと全行・全列一致すること。
2026-06 実データ（`経費利用履歴 Rev5.csv`）に対し jinjer 1947/1947・e-staffing 56/56 の完全一致を確認済み。

対応する VBA:
  - jinjer     … E一覧表作成マクロ.Append_jinjer_CSV_to_経費統合一覧表 ＋ MoveJinjerCustomerBillToAH
  - e-staffing … E一覧表作成マクロ.Append_e_staffing_出力_to_経費統合一覧表
  - SAP        … hSAP経費貼り付けマクロ.Paste_SAP経費_From_File(8a/8b) ＋ hSAP経費取り込みマクロ.Append_SAP経費_to_経費統合一覧表
  - freee      … GFREE整形マクロ ＋ b経費統合一覧表取込マクロ.Append_本社経費_to_経費統合一覧表

参考: 経費一覧表マクロフォルダの `README_勤怠チェッカー組み込み仕様.md` / `AI解析ガイド.md`。
重複削除(A/D/F)は現行マクロで無効化されており Rev5 も未削除のため、本移植でも既定で行わない（2026-07-15 谷津確定）。
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Meiryo UI"

# 経費統合一覧表の 34 列ヘッダー（Rev5 実測。1-based の列位置がそのまま index+1）
INTEGRATED_HEADERS = [
    "社員番号", "氏名", "申請日(yyyy/mm/dd)", "合計", "備考(申請書データ)",
    "利用日(yyyy/mm/dd)", "交通機関", "内訳", "請求区分ID", "請求区分",
    "費用種別", "費用種別ID", "小計", "金額(交通費)", "往復", "金額",
    "単価", "数量", "人数", "備考(明細)", "計上日(yyyy/mm/dd)", "計上日(yyyymmdd)",
    "借方：税率", "貸方：税率", "仕訳No.", "仕訳区分", "企業名", "出張区分",
    "出張先", "支払方法", "乗車場所", "降車場所", "経路", "顧客請求分",
]
NCOL = len(INTEGRATED_HEADERS)  # 34

# 0-based 列インデックス（読みやすさのため）
C_EMP, C_NAME, C_APPDATE, C_TOTAL, C_MEMO_REQ = 0, 1, 2, 3, 4
C_USEDATE, C_TRANS, C_DETAIL, C_BILLTYPE_ID, C_BILLTYPE = 5, 6, 7, 8, 9
C_EXPTYPE, C_EXPTYPE_ID, C_SUBTOTAL, C_FARE, C_ROUNDTRIP, C_AMOUNT = 10, 11, 12, 13, 14, 15
C_MEMO_LINE = 19
C_ENTRY_TYPE = 25  # 仕訳区分
C_BOARD, C_ALIGHT, C_ROUTE, C_CUSTOMER_BILL = 30, 31, 32, 33

# 夜間当番手当キーワード（jinjer/SAP 共通。IsJinjerNightDutyText / IsNightDutyVendor と同一）
NIGHT_DUTY_KEYWORDS = (
    "夜間当番", "24時間準直当番", "準直当番", "深夜出動",
    "顧客当番", "顧客対応当番", "オンコール",
)
# SAP 税抜補正の対象判定キーワード（部分一致。IsNightDutyExpenseRow と同一）
SAP_TAXSTRIP_KEYWORDS = ("顧客対応", "顧客当番")

# freee の申請者名 → 正式氏名 変換（GFREE整形マクロ のハードコードと同一）
FREEE_NAME_CONV = {
    "tomono": "友納 英彦",
    "maki.murayama": "村山 真紀",
    "kazushi.mitani": "三谷 一志",
    "kousei.shiokawa": "塩川 浩生",
    "rina.hirano": "平野 梨奈",
}

# 経路突合チェック: 通勤系とみなす交通機関
COMMUTE_TYPES = {"通勤定期代", "通勤交通費（実費）"}
_ROUTE_LINE_PAT = re.compile(r"(線$|線・|ライン|行$|バス$|快速|各停|快特|特急|急行|新幹線|徒歩)")


# ======================================================================
# 汎用ヘルパー
# ======================================================================

def read_csv_any_enc(path: "str | Path") -> tuple[list[str], list[list[str]]]:
    """CSV を cp932 → utf-8-sig → utf-8 の順で読み、(ヘッダー, データ行) を返す。"""
    path = Path(path)
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc, newline="") as f:
                rows = list(csv.reader(f))
            if not rows:
                return [], []
            return rows[0], rows[1:]
        except UnicodeDecodeError:
            continue
    raise ValueError(f"CSV の文字コードを判別できませんでした: {path}")


def excel_coerce(v) -> str:
    """Excel が `Workbooks.Open(csv, Local:=True)` で CSV を開いたときの型変換を再現する。

    jinjer 取込は CSV を Excel で開いてセル値を読むため、数字らしい文字列は数値化されて
    先頭ゼロが落ち（"00000236"→"236"）、日付らしい文字列は日付シリアル化されて
    ゼロ埋めなしで再表示される（"2026/06/01"→"2026/6/1"）。これを文字列処理で模す。
    """
    s = "" if v is None else str(v)
    t = s.strip()
    if t == "":
        return ""
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", t)
    if m:
        y, mo, d = (int(x) for x in m.groups())
        return f"{y}/{mo}/{d}"
    if re.fullmatch(r"-?\d+", t):
        try:
            return str(int(t))
        except ValueError:
            return s
    if re.fullmatch(r"-?\d+\.\d+", t):
        f = float(t)
        return str(int(f)) if f == int(f) else str(f)
    return s


def norm_date_slash(v) -> str:
    """"2026/06/01" / "2026-6-1" → "2026/6/1"（ゼロ埋めなし）。日付でなければそのまま。"""
    t = ("" if v is None else str(v)).strip().replace("-", "/")
    if " " in t:
        t = t.split(" ")[0]
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", t)
    if m:
        y, mo, d = (int(x) for x in m.groups())
        return f"{y}/{mo}/{d}"
    return t


def norm_date_pad(v) -> str:
    """"2026/4/22 12:17" → "2026/04/22"（ゼロ埋め・時刻切捨て）。SAP の NormalizeDateStr と同じ。"""
    t = ("" if v is None else str(v)).strip()
    if " " in t:
        t = t.split(" ")[0]
    parts = t.replace("-", "/").split("/")
    if len(parts) == 3:
        y, mo, d = parts
        if len(mo) == 1:
            mo = "0" + mo
        if len(d) == 1:
            d = "0" + d
        return f"{y}/{mo}/{d}"
    return t


def normkey(s) -> str:
    """氏名などの照合キー。前後・全角/半角スペースをすべて除去する（NormKey 相当）。"""
    return re.sub(r"[\s　]", "", "" if s is None else str(s))


def excel_coerce_jp_date(text: str, year_hint: str = "") -> str:
    """Excel が OpenText で日付らしい文字列を日付シリアル化する挙動を再現する。

    SAP CSV を `Workbooks.OpenText` で開くと、説明(F)の「6月5日」「6/5」等が日付に自動変換され、
    yyyy/mm/dd（ゼロ埋め）で再表示される（→備考(明細)に入る）。これを文字列処理で模す。
    年の無い「M月D日」「M/D」は year_hint（同行の費用エントリ日など）の年を使う。
    日付でなければ元の文字列を返す。
    """
    if text is None:
        return ""
    d = str(text).strip()
    if d:
        year = year_hint.split("/")[0] if "/" in (year_hint or "") else ""
        m = re.fullmatch(r"(\d{1,2})月(\d{1,2})日", d)
        if m and year:
            return f"{year}/{int(m.group(1)):02d}/{int(m.group(2)):02d}"
        m = re.fullmatch(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", d)
        if m:
            return f"{int(m.group(1))}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
        m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})", d)
        if m and year:
            return f"{year}/{int(m.group(1)):02d}/{int(m.group(2)):02d}"
    return str(text)   # 非日付は原文（前後空白も）そのまま（マクロ SafeStr=CStr 相当）


def parse_amount(v) -> str:
    """金額文字列を正規化（カンマ/￥/円を除去し整数文字列に）。数値でなければそのまま。"""
    s = "" if v is None else str(v)
    s = s.replace(",", "").replace("\\", "").replace("￥", "").replace("円", "").strip()
    if s == "":
        return ""
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def _round_half_up(x: float) -> int:
    """Excel の Application.WorksheetFunction.Round（四捨五入・切り上げ側）と同じ丸め。"""
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def is_night_duty_text(s: str) -> bool:
    s = s or ""
    return any(k in s for k in NIGHT_DUTY_KEYWORDS)


def _is_sap_taxstrip_row(vendor: str, descr: str = "") -> bool:
    s = (vendor or "") + " " + (descr or "")
    return any(k in s for k in SAP_TAXSTRIP_KEYWORDS)


def in_company_scope(emp_no: str) -> bool:
    """20YY 始まり＝自社社員（給与計算対象）。5/6/9 始まりは派遣・テストで対象外。"""
    e = (emp_no or "").strip()
    return bool(e) and e[0] not in ("5", "6", "9")


def _blank_row() -> list[str]:
    return [""] * NCOL


# ======================================================================
# 社員番号ロスター（氏名 → 社員番号）
# ======================================================================

def add_roster_entry(roster: dict, name: str, emp_no: str) -> None:
    """氏名→社員番号を姓名そのまま・姓名逆順の両方で登録する（先勝ち）。"""
    emp_no = (emp_no or "").strip()
    if not name or not emp_no:
        return
    k = normkey(name)
    if k:
        roster.setdefault(k, emp_no)
    parts = re.split(r"[\s　]+", str(name).strip())
    if len(parts) >= 2:
        roster.setdefault(normkey(parts[-1] + parts[0]), emp_no)


def build_roster_from_jinjer_csv(headers: list[str], rows: list[list[str]]) -> dict:
    """jinjer 仕訳データ CSV の（社員番号・氏名）ペアからロスターを作る（API 不通時のフォールバック）。"""
    roster: dict = {}
    for r in rows:
        emp = (r[C_EMP].strip() if len(r) > C_EMP and r[C_EMP] else "")
        nm = r[C_NAME] if len(r) > C_NAME else ""
        add_roster_entry(roster, nm, emp)
    return roster


def build_roster_from_api(employees: list[dict]) -> dict:
    """jinjer API の在籍者一覧（[{id, name}]）からロスターを作る。集計シートの役割。"""
    roster: dict = {}
    for e in employees or []:
        add_roster_entry(roster, e.get("name", ""), str(e.get("id", "")))
    return roster


# ======================================================================
# ソース別変換（各 34 列の行リストを返す）
# ======================================================================

def transform_jinjer(headers: list[str], rows: list[list[str]]) -> list[list[str]]:
    """jinjer 仕訳データ CSV → 経費統合一覧表(34列)。

    仕訳データの 33 列は統合一覧表の 1〜33 列と 1:1（申請者社員番号→社員番号, 申請者名→氏名）。
    Excel 取込の型変換（excel_coerce）を通し、請求区分=="顧客請求" かつ夜間当番系でない行は
    金額を 34 列目（顧客請求分）へ寄せ、合計/小計/金額(交通費)/金額 をクリアする
    （E一覧表作成マクロ.MoveJinjerCustomerBillToAH と同一）。
    """
    out: list[list[str]] = []
    for r in rows:
        row = _blank_row()
        for i in range(min(33, len(r))):
            row[i] = excel_coerce(r[i])
        # MoveJinjerCustomerBillToAH: 請求区分(10列目)が「顧客請求」の行だけ対象
        bill_type = (r[C_BILLTYPE].strip() if len(r) > C_BILLTYPE and r[C_BILLTYPE] else "")
        if bill_type == "顧客請求":
            judge = " ".join([
                _cell(r, C_DETAIL), _cell(r, C_MEMO_REQ), _cell(r, C_EXPTYPE), _cell(r, C_MEMO_LINE),
            ])
            if not is_night_duty_text(judge):
                bill_amt = ""
                for idx in (C_TOTAL, C_AMOUNT, C_SUBTOTAL, C_FARE):
                    if len(r) > idx and (r[idx] or "").strip():
                        bill_amt = excel_coerce(r[idx])
                        break
                if bill_amt != "":
                    row[C_CUSTOMER_BILL] = bill_amt
                    row[C_TOTAL] = ""
                    row[C_SUBTOTAL] = ""
                    row[C_FARE] = ""
                    row[C_AMOUNT] = ""
        out.append(row)
    return out


def transform_estaffing(headers: list[str], rows: list[list[str]], roster: dict) -> list[list[str]]:
    """e-staffing 立替金 CSV → 経費統合一覧表(34列)。全額を顧客請求分(34列)へ入れる。

    立替金 CSV(15列): スタッフ名(5) / 就業年月日(6) / 出発地(9) / 到着地(10) / 交通手段(11) /
    その他立替金内容(12) / 非課税立替金内容(13) / 課税対象外内容(14) / 金額(15)。
    社員番号は氏名照合（ロスター）で付与する。
    """
    out: list[list[str]] = []
    for r in rows:
        nm = _cell(r, 4)
        amt = parse_amount(_cell(r, 14))
        if not (nm or amt):
            continue
        detail = _cell(r, 11) or _cell(r, 12) or _cell(r, 13)
        row = _blank_row()
        row[C_EMP] = roster.get(normkey(nm), "")
        row[C_NAME] = nm
        row[C_USEDATE] = norm_date_slash(_cell(r, 5))
        row[C_TRANS] = _cell(r, 10)   # 交通機関 ← 交通手段
        row[C_DETAIL] = detail        # 内訳
        row[C_FARE] = amt             # 金額(交通費)
        row[C_AMOUNT] = amt           # 金額
        row[C_BOARD] = _cell(r, 8)    # 乗車場所 ← 出発地
        row[C_ALIGHT] = _cell(r, 9)   # 降車場所 ← 到着地
        row[C_CUSTOMER_BILL] = amt    # 顧客請求分（e-staffing は全額客先請求）
        out.append(row)
    return out


def transform_sap(headers: list[str], rows: list[list[str]], roster: dict) -> list[list[str]]:
    """SAP Fieldglass 経費 生CSV → 経費統合一覧表(34列)。

    生CSV(13列): 姓(1) 名(2) 費用合計(3) 業者名(4) 費用エントリ日(5) 説明(6) 承認日(7)
    事業単位(8) コストセンター(9) 通貨(10) 費用シートID(11) 勤務地(12) ステータス(13)。
    前処理 8a/8b（Paste_SAP経費_From_File）→ 34列マッピング（Append_SAP経費）を行う。
      8a: 夜間当番KWが説明(F)のみにあり業者名(D)に無い行は D←F を転記
      8b: 業者名(D)or説明(F)に「顧客対応/顧客当番」を含む行は 費用合計(C)を ÷1.1（四捨五入）で税抜化
    夜間当番系業者名の行は金額を D:合計（夜間当番手当）へ、その他は 34列目:顧客請求分 へ入れる。
    """
    work = [list(r) + [""] * (13 - len(r)) for r in rows]
    # 8a
    for w in work:
        if _is_sap_taxstrip_row(w[5]) and not _is_sap_taxstrip_row(w[3]):
            w[3] = w[5]
    # 8b
    for w in work:
        if _is_sap_taxstrip_row(w[3], w[5]):
            raw = str(w[2]).replace(",", "").replace("\\", "").replace("￥", "").replace("円", "").strip()
            try:
                w[2] = str(_round_half_up(float(raw) / 1.1))
            except ValueError:
                pass

    out: list[list[str]] = []
    for w in work:
        sei, mei = (w[0] or "").strip(), (w[1] or "").strip()
        amt = parse_amount(w[2])
        vendor = (w[3] or "").strip()
        entry_date = norm_date_pad(w[4])       # 費用エントリ日 → 利用日
        # 説明(F): Excel が OpenText で「6月5日」等を日付化するため、その挙動を再現して備考に入れる。
        # 非日付は原文そのまま（前後空白も保持＝マクロ CStr 相当）
        desc = excel_coerce_jp_date(w[5], entry_date)
        cc = (w[8] or "").strip()
        sap_id = (w[10] or "").strip()

        memo = desc
        if cc:
            memo += " / CC: " + cc
        if sap_id:
            memo += " / ID: " + sap_id

        row = _blank_row()
        row[C_EMP] = roster.get(normkey(sei + mei)) or roster.get(normkey(mei + sei)) or ""
        row[C_NAME] = (sei + " " + mei).strip()
        row[C_APPDATE] = norm_date_pad(w[6])   # 申請日 ← 承認日
        row[C_USEDATE] = entry_date            # 利用日 ← 費用エントリ日
        row[C_DETAIL] = vendor[:80]            # 内訳 ← 業者名（80文字制限）
        row[C_MEMO_LINE] = memo                # 備考(明細)
        if any(k in vendor for k in NIGHT_DUTY_KEYWORDS):
            row[C_TOTAL] = amt                 # D: 合計（夜間当番手当として集計）
        else:
            row[C_CUSTOMER_BILL] = amt         # 34: 顧客請求分（SAP 通常行）
        out.append(row)
    return out


# freee 経費精算 CSV の列インデックス（0-based, 32列固定）
_FREEE_COLS = {
    "申請日": 2, "申請者": 3, "申請タイトル": 12, "合計金額": 13,
    "日付": 18, "経費科目": 22, "内容": 23, "金額": 25, "備考": 29,
}


def transform_freee(headers: list[str], rows: list[list[str]], roster: dict) -> list[list[str]]:
    """freee 経費精算 CSV（本社経費）→ 経費統合一覧表(34列)。

    申請ヘッダ＋明細の縦持ち構造なので、申請日/申請者/申請タイトル/合計金額 を前方フィルする。
    名称変換（メール→氏名, ハードコード変換）後、社員番号を氏名照合で付与し、Append_本社経費 相当で
    34列へ落とす（D=金額, F:利用日=日付, H:内訳=経費科目, T:備考=内容/備考, Z:仕訳区分="本社経費"）。
    """
    C = _FREEE_COLS
    filled = [list(r) + [""] * (32 - len(r)) for r in rows]
    last: dict[int, str] = {}
    for r in filled:
        for c in (C["申請日"], C["申請者"], C["申請タイトル"], C["合計金額"]):
            if (r[c] or "").strip() == "":
                r[c] = last.get(c, "")
            else:
                last[c] = r[c]

    out: list[list[str]] = []
    for r in filled:
        applicant = (r[C["申請者"]] or "").strip()
        if "@" in applicant:
            applicant = applicant.split("@")[0]
        for kw, nm in FREEE_NAME_CONV.items():
            if kw in applicant.lower():
                applicant = nm
                break
        emp_no = roster.get(normkey(applicant), "該当なし")
        amt = parse_amount(r[C["金額"]])
        # 内容・備考は原文そのまま（改行・前後空白も保持＝マクロ CStr 相当。内容に改行を含むセルあり）
        cont = r[C["内容"]] or ""
        memo = r[C["備考"]] or ""

        row = _blank_row()
        row[C_EMP] = emp_no
        row[C_NAME] = applicant
        row[C_APPDATE] = norm_date_pad(r[C["申請日"]])
        row[C_USEDATE] = norm_date_pad(r[C["日付"]])         # F: 利用日 ← 日付（Append_本社経費）
        row[C_TOTAL] = amt                                   # D: 合計 ← 金額
        row[C_MEMO_REQ] = (r[C["申請タイトル"]] or "").strip()   # E: 備考(申請書) ← 申請タイトル
        row[C_DETAIL] = (r[C["経費科目"]] or "").strip()        # H: 内訳 ← 経費科目
        row[C_AMOUNT] = amt                                  # P: 金額
        row[C_MEMO_LINE] = (cont + " / " + memo) if (cont and memo) else (cont or memo)
        row[C_ENTRY_TYPE] = "本社経費"                        # Z: 仕訳区分
        out.append(row)
    return out


def _cell(r: list, i: int) -> str:
    return (r[i].strip() if i < len(r) and r[i] is not None else "")


# ======================================================================
# 経路突合チェック（通勤経路上の移動なのに通勤系以外で申請された交通費の検出）
# ======================================================================

def _norm_station(s) -> str:
    if s is None:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKC", str(s)).strip()
    s = re.sub(r"[（(].*?[）)]", "", s)   # 括弧内の路線名等を除去
    s = s.replace(" ", "").replace("　", "")
    s = re.sub(r"駅$", "", s)
    return s


def _stations_from_route_text(text) -> list[str]:
    if not text:
        return []
    toks = [t.strip() for t in re.split(r"→|⇒|〜|~", str(text)) if t.strip()]
    out = []
    for t in toks:
        if _ROUTE_LINE_PAT.search(t):
            continue
        n = _norm_station(t)
        if n:
            out.append(n)
    return out


def build_commute_station_sets(commute_rows: list[dict]) -> dict:
    """通勤費（API 取得）から社員番号 → {駅集合, 経路表現} を作る。"""
    commute: dict = {}
    for c in commute_rows or []:
        emp = str(c.get("社員番号") or "").strip()
        if not emp:
            continue
        ent = commute.setdefault(emp, {"stations": set(), "routes": []})
        sts = [_norm_station(c.get(k)) for k in ("出発", "到着", "経由1", "経由2")]
        sts = [s for s in sts if s]
        sts += _stations_from_route_text(c.get("通勤経路"))
        ent["stations"].update(sts)
        rep = "→".join(str(c.get(k)) for k in ("出発", "経由1", "経由2", "到着") if c.get(k))
        if not rep:
            rep = str(c.get("通勤経路") or "")
        if rep:
            kikan = c.get("利用交通機関") or ""
            ent["routes"].append(rep + (f"[{kikan}]" if kikan else ""))
    return commute


def evaluate_route_check(integrated_rows: list[list[str]], commute_rows: list[dict]) -> list[dict]:
    """統合一覧表の交通費行を通勤経路と突合し、判定付きの行リストを返す。"""
    commute = build_commute_station_sets(commute_rows)
    results: list[dict] = []
    for r in integrated_rows:
        kikan = _cell(r, C_TRANS)
        board = _norm_station(_cell(r, C_BOARD))
        alight = _norm_station(_cell(r, C_ALIGHT))
        if not kikan and not (board or alight):
            continue  # 交通費行以外（夜間当番手当等）はスキップ
        emp = _cell(r, C_EMP)
        ent = commute.get(emp)
        stations = ent["stations"] if ent else set()
        routes = " / ".join(ent["routes"]) if ent else ""

        if not ent or not stations:
            match = "通勤経路登録なし"
        elif board and alight and board in stations and alight in stations:
            match = "経路内"
        elif (board and board in stations) or (alight and alight in stations):
            match = "片側一致"
        else:
            match = "一致なし"

        if match == "経路内":
            verdict = "OK（通勤系を選択）" if kikan in COMMUTE_TYPES else "★要確認（経路内なのに通勤系以外を選択）"
        elif match in ("一致なし", "通勤経路登録なし") and kikan in COMMUTE_TYPES:
            verdict = "△逆要確認（通勤系なのに登録経路と不一致）"
        elif match == "片側一致":
            verdict = "参考（片側のみ一致）"
        else:
            verdict = "OK（経路外）"

        amount = _cell(r, C_TOTAL) or _cell(r, C_FARE) or _cell(r, C_CUSTOMER_BILL)
        results.append({
            "社員番号": emp, "氏名": _cell(r, C_NAME), "利用日": _cell(r, C_USEDATE),
            "交通機関": kikan, "内訳": _cell(r, C_DETAIL), "乗車場所": _cell(r, C_BOARD),
            "降車場所": _cell(r, C_ALIGHT), "経路": _cell(r, C_ROUTE), "金額": amount,
            "往復": _cell(r, C_ROUNDTRIP), "備考(明細)": _cell(r, C_MEMO_LINE),
            "登録通勤経路": routes, "一致": match, "判定": verdict,
        })
    return results


# ======================================================================
# Excel 出力
# ======================================================================

def add_integrated_sheet(wb: Workbook, rows: list[list[str]]) -> None:
    """「経費統合一覧表」シートを追加する（34列・全セル文字列）。"""
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    body_font = Font(name=FONT)

    ws = wb.active
    ws.title = "経費統合一覧表"
    ws.append(INTEGRATED_HEADERS)
    for row in rows:
        ws.append([("" if v is None else str(v)) for v in row])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    n = len(rows)
    for row in ws.iter_rows(min_row=2, max_row=max(n + 1, 2)):
        for cell in row:
            cell.font = body_font
    widths = [10, 14, 15, 9, 24, 15, 20, 22, 9, 10, 14, 10, 9, 12, 6, 9,
              8, 6, 6, 30, 15, 13, 9, 9, 12, 10, 26, 9, 12, 12, 14, 14, 40, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if n:
        ws.auto_filter.ref = f"A1:{get_column_letter(NCOL)}{n + 1}"


_ROUTE_COLS = ["社員番号", "氏名", "利用日", "交通機関", "内訳", "乗車場所", "降車場所",
               "経路", "金額", "往復", "備考(明細)", "登録通勤経路", "一致", "判定"]


def _write_route_sheet(ws, rows: list[dict]) -> None:
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    ws.append(_ROUTE_COLS)
    for c in ws[1]:
        c.font = Font(name=FONT, bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in _ROUTE_COLS])
        verdict = r.get("判定", "")
        if verdict.startswith("★"):
            for c in ws[ws.max_row]:
                c.fill = red
        elif verdict.startswith("△"):
            for c in ws[ws.max_row]:
                c.fill = yellow
    widths = [9, 12, 10, 18, 14, 12, 12, 40, 8, 6, 20, 40, 12, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions


def add_route_check_sheets(wb: Workbook, route_results: list[dict]) -> dict:
    """「要確認」「全交通費行」シートを追加し、集計サマリを返す。"""
    flagged = [x for x in route_results if x["判定"].startswith("★")]
    rev_flagged = [x for x in route_results if x["判定"].startswith("△")]
    ws1 = wb.create_sheet("要確認(経路突合)")
    _write_route_sheet(ws1, flagged + rev_flagged)
    ws2 = wb.create_sheet("全交通費行(経路突合)")
    _write_route_sheet(ws2, route_results)

    def _emp_count(rows):
        return len({(r["社員番号"], r["氏名"]) for r in rows})

    def _sum(rows):
        total = 0
        for r in rows:
            try:
                total += int(float(str(r.get("金額") or 0)))
            except ValueError:
                pass
        return total

    return {
        "flagged_rows": len(flagged), "flagged_emps": _emp_count(flagged), "flagged_amount": _sum(flagged),
        "rev_rows": len(rev_flagged), "rev_emps": _emp_count(rev_flagged),
        "total_rows": len(route_results),
    }


# ======================================================================
# 共通実行関数（CLI / Flask 共用）
# ======================================================================

@dataclass
class KeihiResult:
    ok: bool
    output_path: Path
    integrated_rows: int = 0
    source_counts: dict = field(default_factory=dict)      # {jinjer: n, ...}
    unmatched_emp: int = 0                                  # 社員番号が引けなかった行数
    route_summary: dict = field(default_factory=dict)
    classify_summary: dict = field(default_factory=dict)   # 分類・集計の統計（P1b）
    import_preview: list = field(default_factory=list)     # インポート行プレビュー（人間チェック用）
    import_warnings: list = field(default_factory=list)
    import_csv_name: str = ""                               # 出力したインポートCSVのファイル名
    error: str = ""
    logs: list[str] = field(default_factory=list)


def build_integrated_rows(
    jinjer_csv: "str | Path | None" = None,
    estaffing_csv: "str | Path | None" = None,
    sap_csv: "str | Path | None" = None,
    freee_csv: "str | Path | None" = None,
    roster: "dict | None" = None,
    log_func=print,
) -> tuple[list[list[str]], dict]:
    """指定された生CSVを取り込み、34列統合一覧表の行リストと各ソース件数を返す。

    順序は jinjer → e-staffing → SAP → freee（Rev5 実測に合わせた既定）。
    roster が None のとき、jinjer CSV から氏名→社員番号ロスターを自動構築する。
    """
    roster = dict(roster or {})
    rows: list[list[str]] = []
    counts: dict = {}

    if jinjer_csv:
        # フォルダ指定なら直下の *.csv を名前順で全部取り込む（後からCSVを追加する運用。
        # 同じデータのCSVを2つ置くと二重計上になるので注意。重複削除はしない=Rev5準拠）
        jp = Path(jinjer_csv)
        jinjer_files = sorted(jp.glob("*.csv")) if jp.is_dir() else [jp]
        if not jinjer_files:
            raise ValueError(f"jinjer経費CSVフォルダに .csv がありません: {jp}")
        total_j = 0
        for jf in jinjer_files:
            jh, jr = read_csv_any_enc(jf)
            if len(jh) < 33:
                raise ValueError(
                    f"jinjer仕訳データCSVの列数が不足しています（{len(jh)}列）: {jf.name}")
            # jinjer CSV の（社員番号・氏名）でロスターを補完（API 不通時のフォールバック）
            for k, v in build_roster_from_jinjer_csv(jh, jr).items():
                roster.setdefault(k, v)
            tj = transform_jinjer(jh, jr)
            rows += tj
            total_j += len(tj)
            if len(jinjer_files) > 1:
                log_func(f"[info]   {jf.name}: {len(tj)} 行")
        counts["jinjer"] = total_j
        log_func(f"[info] jinjer 仕訳データ: {total_j} 行（{len(jinjer_files)} ファイル）")
    else:
        counts["jinjer"] = None

    if estaffing_csv:
        eh, er = read_csv_any_enc(estaffing_csv)
        te = transform_estaffing(eh, er, roster)
        rows += te
        counts["estaffing"] = len(te)
        log_func(f"[info] e-staffing 立替金: {len(te)} 行")
    else:
        counts["estaffing"] = None

    if sap_csv:
        sh, sr = read_csv_any_enc(sap_csv)
        ts = transform_sap(sh, sr, roster)
        rows += ts
        counts["sap"] = len(ts)
        log_func(f"[info] SAP Fieldglass: {len(ts)} 行")
    else:
        counts["sap"] = None

    if freee_csv:
        fh, fr = read_csv_any_enc(freee_csv)
        tf = transform_freee(fh, fr, roster)
        rows += tf
        counts["freee"] = len(tf)
        log_func(f"[info] freee 本社経費: {len(tf)} 行")
    else:
        counts["freee"] = None

    return rows, counts


def run_keihi_integration(
    output_path: Path,
    jinjer_csv: "str | Path | None" = None,
    estaffing_csv: "str | Path | None" = None,
    sap_csv: "str | Path | None" = None,
    freee_csv: "str | Path | None" = None,
    route_check: bool = True,
    classify: bool = True,
    keywords_file: "str | Path | None" = None,
    import_template_csv: "str | Path | None" = None,
    teijo_allowances: "dict | None" = None,
    sonota_allowances: "dict | None" = None,
    log_func=print,
    client=None,
) -> KeihiResult:
    """4ソース生CSV → 経費統合一覧表(34列) ＋ 経路突合チェック ＋ 集計/集計ログ の Excel を出力する。

    Args:
        output_path: 出力 xlsx パス
        *_csv: 各ソースの生CSVパス（未指定ソースは「未取込」）。jinjer_csv はフォルダ指定可
               （直下の *.csv を全部取り込む）
        route_check: True なら jinjer API(commuting-information) から通勤経路を取り、経路突合シートを追加
        classify: True なら分類・集計して「集計」「集計ログ」シートとインポートプレビューを出す（P1b）
        keywords_file: 分類キーワード設定（設定シート形式 xlsx/CSV）。未指定は内蔵デフォルト
        teijo_allowances: 定常外業務対応手当 {社員番号: 金額}（画面手入力。経費から導けないため）
        sonota_allowances: その他手当 {社員番号: 金額}（同上）
        client: jinjer API クライアント（省略時は route_check/ロスターのため内部生成）
    """
    result = KeihiResult(ok=False, output_path=output_path)

    if not any([jinjer_csv, estaffing_csv, sap_csv, freee_csv]):
        result.error = "少なくとも1つのソースCSVを指定してください。"
        log_func(f"[error] {result.error}")
        return result

    # --- ロスター（氏名→社員番号）を jinjer API から構築（e-staffing/SAP/freee 用） ---
    roster: dict = {}
    roster_id_to_name: dict = {}
    commute_rows: list[dict] = []
    if route_check or estaffing_csv or sap_csv or freee_csv:
        try:
            from services.expense_check import fetch_active_employees, fetch_commute_rows_via_api
            from services.jinjer_api_client import JinjerClient
            if client is None:
                client = JinjerClient()
                client.authenticate()
            employees = fetch_active_employees(client)
            roster = build_roster_from_api(employees)
            roster_id_to_name = {str(e["id"]): e["name"] for e in employees}
            log_func(f"[info] jinjer 在籍者ロスター: {len(employees)} 名")
            if route_check:
                commute_rows = fetch_commute_rows_via_api(client, roster_id_to_name)
                log_func(f"[info] 通勤経路(API): {len(commute_rows)} 経路")
        except Exception as e:  # noqa: BLE001 — API 不通でも jinjer CSV から最低限のロスターで続行
            log_func(f"[warn] jinjer API に接続できませんでした（CSV由来のロスターで続行）: {e}")

    # --- 統合一覧表を組み立て ---
    try:
        rows, counts = build_integrated_rows(
            jinjer_csv=jinjer_csv, estaffing_csv=estaffing_csv,
            sap_csv=sap_csv, freee_csv=freee_csv, roster=roster, log_func=log_func,
        )
    except Exception as e:  # noqa: BLE001
        result.error = f"統合一覧表の生成に失敗しました: {e}"
        log_func(f"[error] {result.error}")
        return result

    result.source_counts = counts
    result.integrated_rows = len(rows)
    result.unmatched_emp = sum(
        1 for r in rows if not r[C_EMP] or r[C_EMP] == "該当なし"
    )

    # --- Excel 出力 ---
    wb = Workbook()
    add_integrated_sheet(wb, rows)
    if route_check:
        try:
            route_results = evaluate_route_check(rows, commute_rows)
            result.route_summary = add_route_check_sheets(wb, route_results)
            log_func(
                f"[info] 経路突合: ★要確認 {result.route_summary['flagged_rows']}行/"
                f"{result.route_summary['flagged_emps']}名（約{result.route_summary['flagged_amount']:,}円）・"
                f"△逆要確認 {result.route_summary['rev_rows']}行/{result.route_summary['rev_emps']}名"
            )
        except Exception as e:  # noqa: BLE001 — 経路突合が失敗しても統合一覧表は出す
            log_func(f"[warn] 経路突合チェックの作成に失敗（スキップ）: {e}")

    # --- 分類・集計（P1b）＋ jinjer給与インポートのプレビュー ---
    if classify:
        try:
            from services.keihi_classify import classify_and_summarize, load_keywords
            from services.keihi_payroll_import import build_import_rows, write_import_csv

            keywords = None
            if keywords_file:
                keywords = load_keywords(keywords_file)
                log_func(f"[info] 分類キーワードを読込: {keywords_file}")
            stats = classify_and_summarize(rows, wb, keywords, roster_id_to_name)
            agg = stats.pop("_agg")
            emp_names = stats.pop("_emp_names")
            result.classify_summary = stats
            log_func(
                f"[info] 分類・集計: 処理 {stats['classified_hits']} 件 / 集計 {stats['summary_employees']} 名 / "
                f"未照合(氏名のみ) {stats['unmatched_rows']} 件 / 判定差分あり {stats['diff_ng']} 名"
            )
            if stats["excluded_out_of_scope"]:
                log_func(f"[info] 給与計算対象外(5/6/9始まり)を集計から除外: {stats['excluded_out_of_scope']} 名")

            # インポート行（人間チェック用のプレビューとCSV）
            import_rows, warnings = build_import_rows(
                agg.by_id, emp_names, roster_id_to_name,
                teijo=teijo_allowances, sonota=sonota_allowances)
            result.import_preview = import_rows
            n_teijo = sum(1 for a in (teijo_allowances or {}).values() if a)
            n_sonota = sum(1 for a in (sonota_allowances or {}).values() if a)
            if n_teijo or n_sonota:
                log_func(f"[info] 手入力手当を反映: 定常外業務対応手当 {n_teijo}名 "
                         f"({sum((teijo_allowances or {}).values()):,}円) / "
                         f"その他手当 {n_sonota}名 ({sum((sonota_allowances or {}).values()):,}円)")

            # jinjerテンプレCSVが指定されていれば、その列並びに追従してCSVを組み立てる
            # （jinjerのテンプレインポートは列位置対応。並びがズレると値が別項目に入る）
            template_header = None
            if import_template_csv:
                from services.keihi_payroll_import import read_template_header, check_template_coverage
                try:
                    template_header = read_template_header(import_template_csv)
                    log_func(f"[info] インポートテンプレに追従: {len(template_header)}列 "
                             f"({import_template_csv})")
                    cov = check_template_coverage(import_rows, template_header)
                    warnings = cov + warnings
                except Exception as e:  # noqa: BLE001
                    log_func(f"[warn] テンプレCSVを読めませんでした（既定ヘッダーで出力）: {e}")

            result.import_warnings = warnings
            csv_name = output_path.stem + "_jinjerインポート.csv"
            write_import_csv(import_rows, output_path.parent / csv_name, template_header)
            result.import_csv_name = csv_name
            log_func(f"[info] jinjer給与インポートCSVを出力: {csv_name}（{len(import_rows)} 名）")
            for w in warnings:
                log_func(f"[warn] {w}")
        except Exception as e:  # noqa: BLE001 — 分類が失敗しても統合一覧表は出す
            log_func(f"[warn] 分類・集計の作成に失敗（スキップ）: {e}")

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    result.ok = True
    got = [k for k, v in counts.items() if v is not None]
    missing = [k for k, v in counts.items() if v is None]
    log_func(
        f"[done] 出力完了: {output_path} / 統合一覧表 {len(rows)} 行 / "
        f"取込: {', '.join(got) or 'なし'}" + (f" / 未取込: {', '.join(missing)}" if missing else "")
    )
    return result
