import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from config import Config


# 判定ごとの背景色
COLORS = {
    "OK": "C6EFCE",
    "NG": "FFC7CE",
    "要確認": "FFEB9C",
    "データ欠損": "BDD7EE",
}

HEADER_COLOR = "D9D9D9"


def _format_time(t):
    if t is None:
        return ""
    try:
        return t.strftime("%H:%M")
    except Exception:
        return str(t)


def _format_date(d):
    if d is None:
        return ""
    try:
        return d.strftime("%Y/%m/%d")
    except Exception:
        return str(d)


def export_to_excel(result_df, threshold_minutes, output_folder=None, unsubmitted_names=None):
    """
    突合結果DataFrameをExcelに出力する

    Args:
        result_df: 突合結果DataFrame
        threshold_minutes: 許容差分（分）
        output_folder: 出力先フォルダ（省略時はConfig.OUTPUT_FOLDER）
        unsubmitted_names: 勤務表未提出の社員名リスト（省略可）

    Returns:
        str: 出力ファイルパス
    """
    if output_folder is None:
        output_folder = Config.OUTPUT_FOLDER
    os.makedirs(output_folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"勤怠突合結果_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "突合結果"

    headers = [
        "氏名", "日付",
        "勤務表_出勤", "jinjer_出勤", "出勤差分(分)",
        "勤務表_退勤", "jinjer_退勤", "退勤差分(分)",
        "判定", "詳細",
        "jinjer_コメント", "勤務表_コメント"
    ]

    col_map = {
        "氏名": "氏名",
        "日付": "日付",
        "勤務表_出勤": "勤務表_出勤時刻",
        "jinjer_出勤": "jinjer_出勤時刻",
        "出勤差分(分)": "出勤差分(分)",
        "勤務表_退勤": "勤務表_退勤時刻",
        "jinjer_退勤": "jinjer_退勤時刻",
        "退勤差分(分)": "退勤差分(分)",
        "判定": "判定",
        "詳細": "詳細",
        "jinjer_コメント": "jinjer_コメント",
        "勤務表_コメント": "勤務表_コメント",
    }

    # ヘッダー行
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # フィルター設定
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # データ行
    for row_idx, row in result_df.iterrows():
        excel_row = row_idx + 2
        values = []
        for header in headers:
            df_col = col_map[header]
            val = row.get(df_col)
            if df_col == "日付":
                val = _format_date(val)
            elif "_出勤時刻" in df_col or "_退勤時刻" in df_col:
                val = _format_time(val)
            elif val is None or (hasattr(val, '__class__') and val.__class__.__name__ == 'float' and str(val) == 'nan'):
                val = ""
            values.append(val)

        judgment = str(row.get("判定", ""))
        fill_color = COLORS.get(judgment)

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val if val is not None else "")
            cell.alignment = Alignment(vertical="center")
            # 判定列（9列目）に色付け
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # 列幅の自動調整
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header) * 2
        for row_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(cell_val) * 1.5)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # サマリーシート
    ws_summary = wb.create_sheet(title="サマリー")
    counts = result_df["判定"].value_counts().to_dict()

    summary_data = [
        ("全件数", len(result_df)),
        ("OK件数", counts.get("OK", 0)),
        ("NG件数", counts.get("NG", 0)),
        ("要確認件数", counts.get("要確認", 0)),
        ("データ欠損件数", counts.get("データ欠損", 0)),
        ("許容差分設定値(分)", threshold_minutes),
    ]

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        cell = ws_summary.cell(row=r_idx, column=2, value=value)
        # 件数セルに色付け
        if label == "NG件数" and value > 0:
            cell.fill = PatternFill(start_color=COLORS["NG"], end_color=COLORS["NG"], fill_type="solid")
        elif label == "要確認件数" and value > 0:
            cell.fill = PatternFill(start_color=COLORS["要確認"], end_color=COLORS["要確認"], fill_type="solid")
        elif label == "OK件数":
            cell.fill = PatternFill(start_color=COLORS["OK"], end_color=COLORS["OK"], fill_type="solid")

    # 未提出者リストシート
    if unsubmitted_names:
        ws_unsub = wb.create_sheet(title="未提出者リスト")
        unsub_header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

        ws_unsub.cell(row=1, column=1, value="No.").fill = unsub_header_fill
        ws_unsub.cell(row=1, column=1).font = Font(bold=True)
        ws_unsub.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws_unsub.cell(row=1, column=2, value="氏名").fill = unsub_header_fill
        ws_unsub.cell(row=1, column=2).font = Font(bold=True)
        ws_unsub.cell(row=1, column=2).alignment = Alignment(horizontal="center")

        for idx, name in enumerate(unsubmitted_names, start=1):
            ws_unsub.cell(row=idx + 1, column=1, value=idx)
            ws_unsub.cell(row=idx + 1, column=2, value=name)

        ws_unsub.column_dimensions["A"].width = 8
        ws_unsub.column_dimensions["B"].width = 25

        # サマリーに未提出者数を追加
        next_row = len(summary_data) + 2
        ws_summary.cell(row=next_row, column=1, value="勤務表未提出者数").font = Font(bold=True)
        cell = ws_summary.cell(row=next_row, column=2, value=len(unsubmitted_names))
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    wb.save(filepath)
    return filepath
