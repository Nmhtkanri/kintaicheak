# -*- coding: utf-8 -*-
"""メール下書きモード — 一覧表 × テンプレート × メール台帳から Outlook 下書きを一括作成する。

2026-07-28 決定（Z:\\API連携\\docs\\システム全体マップと機能追加ルール.md 3-1）:
- 作るのは「下書きまで」だった。2026-08-31 に、作った下書きをまとめて送る機能を足した（下記）。

2026-08-31 更新（谷津さん決定・同 3-1 を改訂）:
- **下書きを作ったあと、その下書きをまとめて送れる**ようにした（send_batch）。送るのは
  create_drafts が作った下書きそのもの＝ Outlook で人が直した内容がそのまま送られる。
  新しいメールを組み立て直すことはしない。
- 歯止めは5つ。①送れるのは自分が作ったバッチだけ（EntryID は本人の Outlook
  プロファイルでしか解決できない） ②プレビュー必須 ③件数入りの確認文字「SEND n」
  ④要確認の人はそもそもバッチに入らない ⑤全件を下書き作成ログに残す。
- 差し込みは {{列名}}。列が無い・値が空欄の人は「要確認」にして下書きを作らない。
- 宛先はメール台帳（B=社員番号 / C=氏名 / D=社用 / E=就業先 / F=個人）と社員番号で突合。
  To=社用優先、就業先・個人は本人BCC。台帳に無い人・氏名相違・アドレス無しは自動で対象外。
- テンプレートは共有フォルダの JSON（Config.MAIL_TEMPLATES_JSON）。exe 再ビルド不要で追加・修正できる。
- Outlook COM (classic Outlook) に触るのは create_drafts() だけ。他は純ロジックでテスト可能。

移植元: Z:\\API連携\\create_paid_leave_outlook_mail.py（有休案内専用ツール）の宛先突合・検証ロジック。
"""
from __future__ import annotations

import csv
import json
import os
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from config import Config
from services.sap_import_ledger import current_user

ADDRESS_SHEET_NAMES = ("メール送信", "メール送信シート")
ID_HEADER_CANDIDATES = ("社員番号", "従業員番号")
NAME_HEADER_CANDIDATES = ("氏名", "名前", "社員名", "従業員名", "スタッフ名")
ADDRESS_SPLITTER = re.compile(r"[;,\n、]+")
EMAIL_PATTERN = re.compile(r"^[^@\s;]+@[^@\s;]+\.[^@\s;]+$")
PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*([^{}]+?)\s*\}\}")
IMPORTANCE_LEVELS = {"normal": 1, "high": 2}
STATUS_OK = "OK"
STATUS_NG = "要確認"
LOG_FILENAME = "下書き作成ログ.csv"
LOG_HEADERS = ["処理日時", "社員番号", "氏名", "To", "BCC", "件名", "結果"]

# 送信バッチ＝下書きを作ったときの控え。EntryID をここに残しておき、あとで
# send_batch がその下書きを引き当てて送る（新しいメールは作らない）。
SEND_BATCH_PREFIX = "送信バッチ_"
SEND_BATCH_GLOB = SEND_BATCH_PREFIX + "*.json"
SEND_STATE_DRAFT = "draft"
SEND_STATE_SENT = "sent"


# ---------------------------------------------------------------------------
# 正規化・検証（移植）
# ---------------------------------------------------------------------------

def normalize_employee_id(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"[\s\u3000]+", "", text).casefold()


def split_addresses(value: Any) -> tuple[str, ...]:
    parts = [part.strip() for part in ADDRESS_SPLITTER.split(str(value or ""))]
    unique: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if not part:
            continue
        key = part.casefold()
        if key not in seen:
            unique.append(part)
            seen.add(key)
    return tuple(unique)


def invalid_addresses(addresses: Iterable[str]) -> list[str]:
    return [address for address in addresses if not EMAIL_PATTERN.fullmatch(address)]


def value_to_text(value: Any) -> str:
    """差し込み用の文字列化。Excel由来の 3.0 → 3、日付 → 2027年3月31日 に整える。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return f"{value.year}年{value.month}月{value.day}日"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, float):
        return f"{round(value, 2):g}"
    return str(value).strip()


# ---------------------------------------------------------------------------
# 一覧表（Excel/CSV）の読み込み
# ---------------------------------------------------------------------------

def _rows_from_csv(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "cp932"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                return [row for row in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    raise ValueError(f"CSVの文字コードを判定できません（UTF-8/Shift-JIS以外）: {path}")


def _rows_from_excel(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def load_recipients_table(path: str | Path) -> tuple[list[str], list[dict[str, Any]], str, str | None]:
    """一覧表を読み、(ヘッダー一覧, 行dictのリスト, 社員番号列名, 氏名列名orNone) を返す。

    ヘッダー行は先頭30行から「社員番号 / 従業員番号」を含む行を探す。
    同名ヘッダーが複数ある場合は右側の列が勝つ（差し込みには一意な列名を推奨）。
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"一覧表が見つかりません: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw_rows = _rows_from_csv(path)
    elif suffix in (".xlsx", ".xlsm"):
        raw_rows = _rows_from_excel(path)
    else:
        raise ValueError(f"一覧表は .xlsx / .xlsm / .csv を指定してください: {path.name}")

    header_index = None
    for index, row in enumerate(raw_rows[:30]):
        cells = {str(cell or "").strip() for cell in row}
        if cells & set(ID_HEADER_CANDIDATES):
            header_index = index
            break
    if header_index is None:
        raise ValueError(
            "一覧表に「社員番号」（または従業員番号）のヘッダー行が見つかりません（先頭30行を確認）")

    header_row = raw_rows[header_index]
    headers = [str(cell or "").strip() for cell in header_row]
    id_key = next(h for h in headers if h in ID_HEADER_CANDIDATES)
    name_key = next((h for h in headers if h in NAME_HEADER_CANDIDATES), None)

    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1:]:
        row_map: dict[str, Any] = {}
        for column, header in enumerate(headers):
            if not header:
                continue
            row_map[header] = raw[column] if column < len(raw) else None
        if normalize_employee_id(row_map.get(id_key)):
            rows.append(row_map)
    if not rows:
        raise ValueError("一覧表にデータ行がありません（社員番号が入った行が必要です）")
    return [h for h in headers if h], rows, id_key, name_key


# ---------------------------------------------------------------------------
# メール台帳（移植: B=社員番号 / C=氏名 / D=社用 / E=就業先 / F=個人）
# ---------------------------------------------------------------------------

def load_address_book(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"メール台帳が見つかりません: {path}")
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        sheet_name = next((name for name in ADDRESS_SHEET_NAMES if name in workbook.sheetnames), None)
        if sheet_name is None:
            raise ValueError(
                f"メール台帳に「{' / '.join(ADDRESS_SHEET_NAMES)}」シートがありません")
        sheet = workbook[sheet_name]
        header_row = None
        for row_number in range(1, min(sheet.max_row, 30) + 1):
            if str(sheet.cell(row_number, 2).value or "").strip() in ID_HEADER_CANDIDATES:
                header_row = row_number
                break
        if header_row is None:
            raise ValueError("メール台帳の社員番号ヘッダー（B列）を特定できません")

        entries: dict[str, list[dict[str, Any]]] = {}
        for row_number in range(header_row + 1, sheet.max_row + 1):
            employee_id = normalize_employee_id(sheet.cell(row_number, 2).value)
            if not employee_id:
                continue
            entries.setdefault(employee_id, []).append({
                "employee_id": employee_id,
                "name": str(sheet.cell(row_number, 3).value or "").strip(),
                "company": split_addresses(sheet.cell(row_number, 4).value),
                "client": split_addresses(sheet.cell(row_number, 5).value),
                "personal": split_addresses(sheet.cell(row_number, 6).value),
            })
        return entries
    finally:
        workbook.close()


def _address_signature(entry: dict[str, Any]) -> tuple:
    return (entry["company"], entry["client"], entry["personal"])


def _deduplicate_addresses(*groups: Iterable[str]) -> tuple[str, ...]:
    result: list[str] = []
    seen: set[str] = set()
    for group in groups:
        for address in group:
            key = address.casefold()
            if key not in seen:
                result.append(address)
                seen.add(key)
    return tuple(result)


def select_recipients(entry: dict[str, Any]) -> tuple[tuple[str, ...], tuple[str, ...], str]:
    """To=社用優先（無ければ就業先→個人）、残りは本人BCC。内訳ラベルも返す。"""
    if entry["company"]:
        to_addresses, to_label = entry["company"], "社用"
        other_groups = (entry["client"], entry["personal"])
    elif entry["client"]:
        to_addresses, to_label = entry["client"], "就業先"
        other_groups = (entry["personal"],)
    else:
        to_addresses, to_label = entry["personal"], "個人"
        other_groups = ()

    bcc_addresses = _deduplicate_addresses(*other_groups)
    to_keys = {address.casefold() for address in to_addresses}
    bcc_addresses = tuple(a for a in bcc_addresses if a.casefold() not in to_keys)
    bcc_keys = {address.casefold() for address in bcc_addresses}
    bcc_labels = []
    if any(a.casefold() in bcc_keys for a in entry["client"]):
        bcc_labels.append("就業先")
    if any(a.casefold() in bcc_keys for a in entry["personal"]):
        bcc_labels.append("個人")
    breakdown = f"To:{to_label}"
    if bcc_labels:
        breakdown += " / BCC:" + "・".join(bcc_labels)
    return to_addresses, bcc_addresses, breakdown


# ---------------------------------------------------------------------------
# 差し込み
# ---------------------------------------------------------------------------

def extract_placeholders(*texts: str) -> list[str]:
    found: list[str] = []
    for text in texts:
        for match in PLACEHOLDER_PATTERN.finditer(text or ""):
            key = match.group(1).strip()
            if key not in found:
                found.append(key)
    return found


def render_template_text(text: str, row: dict[str, Any], headers: set[str]) -> tuple[str, list[str]]:
    """{{列名}} を row の値で置換する。列なし・空欄は missing に列挙し、置換しない。"""
    missing: list[str] = []

    def _substitute(match: re.Match) -> str:
        key = match.group(1).strip()
        if key not in headers:
            missing.append(f"{{{{{key}}}}}（列がありません）")
            return match.group(0)
        rendered = value_to_text(row.get(key))
        if rendered == "":
            missing.append(f"{{{{{key}}}}}（空欄）")
        return rendered

    return PLACEHOLDER_PATTERN.sub(_substitute, text or ""), missing


# ---------------------------------------------------------------------------
# 計画づくり（純ロジックの中心）
# ---------------------------------------------------------------------------

def build_mail_plans(
    rows: list[dict[str, Any]],
    id_key: str,
    name_key: str | None,
    headers: list[str],
    address_book: dict[str, list[dict[str, Any]]],
    template: dict[str, Any],
) -> list[dict[str, Any]]:
    importance = str(template.get("importance") or "normal")
    if importance not in IMPORTANCE_LEVELS:
        raise ValueError(f"重要度は normal / high を指定してください: {importance}")
    # to_only: Toだけに送る（既定・2026-07-29に谷津さん指定で変更） / bcc: 就業先・個人を本人BCCに入れる
    bcc_mode = str(template.get("bcc_mode") or "to_only")
    if bcc_mode not in ("bcc", "to_only"):
        bcc_mode = "to_only"
    cc_addresses = split_addresses(template.get("cc"))
    bad_cc = invalid_addresses(cc_addresses)
    if bad_cc:
        raise ValueError(f"CCのアドレス形式が不正です: {', '.join(bad_cc)}")

    header_set = set(headers)
    plans: list[dict[str, Any]] = []
    for row in rows:
        employee_id = normalize_employee_id(row.get(id_key))
        table_name = value_to_text(row.get(name_key)) if name_key else ""
        problems: list[str] = []
        notes: list[str] = []

        entries = address_book.get(employee_id, [])
        entry = None
        if not entries:
            problems.append("社員番号に対応するメール台帳行がありません")
        elif len(entries) == 1:
            entry = entries[0]
        else:
            signatures = {_address_signature(item) for item in entries}
            if len(signatures) > 1:
                problems.append("同じ社員番号に異なるメールアドレスが複数あります")
            else:
                entry = entries[0]
                notes.append("メール台帳に同一社員番号の重複行があります")

        to_addresses: tuple[str, ...] = ()
        bcc_addresses: tuple[str, ...] = ()
        breakdown = ""
        if entry is not None:
            if table_name and entry["name"] and normalize_name(table_name) != normalize_name(entry["name"]):
                problems.append(f"氏名相違（一覧表: {table_name} / 台帳: {entry['name']}）")
            all_addresses = _deduplicate_addresses(
                entry["company"], entry["client"], entry["personal"])
            if invalid_addresses(all_addresses):
                problems.append("メールアドレス形式エラー")
            if not all_addresses:
                problems.append("利用できるメールアドレスがありません")
            elif not problems:
                to_addresses, bcc_addresses, breakdown = select_recipients(entry)
                if bcc_mode == "to_only":
                    bcc_addresses = ()
                    breakdown = breakdown.split(" / BCC:")[0]

        subject, missing_subject = render_template_text(template.get("subject", ""), row, header_set)
        body, missing_body = render_template_text(template.get("body", ""), row, header_set)
        for item in missing_subject + missing_body:
            message = f"差し込みできません: {item}"
            if message not in problems:
                problems.append(message)

        plans.append({
            "employee_id": employee_id,
            "name": table_name or (entry["name"] if entry else ""),
            "to": list(to_addresses),
            "bcc": list(bcc_addresses),
            "cc": list(cc_addresses),
            "breakdown": breakdown,
            "subject": subject,
            "body": body,
            "importance": IMPORTANCE_LEVELS[importance],
            "status": STATUS_NG if problems else STATUS_OK,
            "issues": problems + notes,
        })
    return plans


def build_plans_for(
    table_path: str | Path,
    address_book_path: str | Path,
    template: dict[str, Any],
    *,
    source: str = "table",
    source_options: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """一覧表と台帳を読み、(plans, meta) を返す。ルートからの入口。

    source で読み手だけを差し替える。差し込み・宛先突合・要確認の判定は共通。

    - ``table``（既定） … Excel/CSV の一覧表をそのまま読む
    - ``paid_leave`` … 有休ブック（対象者一覧／有休明細／集計条件）から
      社員番号・氏名・取得日数・不足日数・取得期限の5列を組み立てる
    """
    source_meta: dict[str, Any] = {}
    if source == "paid_leave":
        from services.paid_leave_mail import load_paid_leave_table
        headers, rows, id_key, name_key, source_meta = load_paid_leave_table(
            table_path, **(source_options or {}))
    elif source == "table":
        headers, rows, id_key, name_key = load_recipients_table(table_path)
    else:
        raise ValueError(f"一覧表の種類が不正です: {source}")
    address_book = load_address_book(address_book_path)
    plans = build_mail_plans(rows, id_key, name_key, headers, address_book, template)
    ok_count = sum(plan["status"] == STATUS_OK for plan in plans)
    meta = {
        "columns": headers,
        "id_column": id_key,
        "name_column": name_key,
        "placeholders": extract_placeholders(template.get("subject", ""), template.get("body", "")),
        "counts": {"total": len(plans), "ok": ok_count, "warn": len(plans) - ok_count},
        "source": source,
        "source_meta": source_meta,
    }
    return plans, meta


# ---------------------------------------------------------------------------
# テンプレート（共有フォルダの JSON・再ビルド不要）
# ---------------------------------------------------------------------------

def load_templates(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as exc:
        raise ValueError(f"テンプレートJSONを読み取れません（{path}）: {exc}") from exc
    templates = data.get("templates") if isinstance(data, dict) else None
    if not isinstance(templates, list):
        raise ValueError(f"テンプレートJSONの形式が不正です（templates配列が必要）: {path}")
    return [t for t in templates if isinstance(t, dict) and str(t.get("name", "")).strip()]


def save_template(path: str | Path, template: dict[str, Any], *, delete: bool = False) -> list[dict[str, Any]]:
    path = Path(path)
    name = str(template.get("name", "")).strip()
    if not name:
        raise ValueError("テンプレート名を入力してください")
    templates = load_templates(path) if path.exists() else []
    templates = [t for t in templates if str(t.get("name", "")).strip() != name]
    if not delete:
        record = {
            "name": name,
            "subject": str(template.get("subject", "")),
            "body": str(template.get("body", "")),
            "cc": str(template.get("cc", "")).strip(),
            "bcc_mode": str(template.get("bcc_mode", "to_only")),
            "importance": str(template.get("importance", "normal")),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        if record["importance"] not in IMPORTANCE_LEVELS:
            record["importance"] = "normal"
        if record["bcc_mode"] not in ("bcc", "to_only"):
            record["bcc_mode"] = "bcc"
        templates.append(record)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"version": 1, "templates": templates}
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)
    return templates


# ---------------------------------------------------------------------------
# Outlook 下書き作成（COM に触るのはここだけ。送信機能は存在しない）
# ---------------------------------------------------------------------------

class OutlookMailer:
    def __init__(self) -> None:
        try:
            import win32com.client  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "Outlook連携に必要な pywin32 がありません（pip install pywin32）") from exc
        try:
            self.application = win32com.client.Dispatch("Outlook.Application")
        except Exception as exc:
            raise RuntimeError(f"Outlook を起動できません（classic Outlook が必要）: {exc}") from exc

    def create_draft(self, *, to: str, cc: str, bcc: str, subject: str, body: str,
                     importance: int) -> dict:
        """下書きを保存し、あとで引き当てるための EntryID / StoreID を返す。

        Save() 後の EntryID は、人が Outlook で本文を直して保存し直しても変わらない
        （2026-08-31 に実機で確認）。send_saved() はこの ID で同じ下書きを引き当てる。
        """
        mail = self.application.CreateItem(0)
        mail.To = to
        if cc:
            mail.CC = cc
        if bcc:
            mail.BCC = bcc
        mail.Subject = subject
        mail.Body = body
        mail.Importance = importance
        mail.Save()
        try:
            store_id = str(mail.Parent.StoreID or "")
        except Exception:
            store_id = ""
        return {"entry_id": str(mail.EntryID or ""), "store_id": store_id}

    def send_saved(self, entry_id: str, store_id: str = "") -> None:
        """保存済みの下書きを EntryID で引き当てて送る。**新しいメールは作らない。**

        引き当てられない（人が消した・別フォルダへ動かした）ときは COM が例外を投げる。
        呼び出し側は「送らずスキップ」に倒すこと。
        """
        namespace = self.application.GetNamespace("MAPI")
        item = (namespace.GetItemFromID(entry_id, store_id) if store_id
                else namespace.GetItemFromID(entry_id))
        item.Send()


def _append_log(log_path: Path, rows: list[list[str]]) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    is_new = not log_path.exists()
    # 新規作成時だけ BOM 付き UTF-8（谷津さんの Excel が BOM 無しを文字化けさせるため）。
    encoding = "utf-8-sig" if is_new else "utf-8"
    with open(log_path, "a", encoding=encoding, newline="") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow(LOG_HEADERS)
        writer.writerows(rows)


def create_drafts(
    plans: list[dict[str, Any]],
    *,
    only_ids: list[str],
    log_dir: str | Path | None = None,
    mailer: Any | None = None,
    template_name: str = "",
) -> dict[str, Any]:
    """選択された OK 行だけ Outlook 下書きを作成し、ログと送信バッチを残す。

    - status が要確認の行は選択されていても作らない（skipped に数える）。
    - 3件連続で失敗したら中断する（Outlook 側の異常を疑う）。
    - 作れた下書きの EntryID を**送信バッチ**に控える。あとで send_batch が
      この控えを使って、同じ下書きを引き当てて送る。
    """
    selected = {str(item) for item in only_ids}
    targets = [p for p in plans if p["employee_id"] in selected and p["status"] == STATUS_OK]
    skipped = len([p for p in plans if p["employee_id"] in selected]) - len(targets)
    if not targets:
        return {"processed": 0, "skipped": skipped, "failed": 0, "results": [],
                "log_path": "", "batch_id": "", "batch_path": "", "sendable": 0}

    log_dir = Path(log_dir or Config.MAIL_OUTPUT_DIR)
    log_path = log_dir / LOG_FILENAME

    own_com = mailer is None
    if own_com:
        try:
            import pythoncom  # type: ignore
            pythoncom.CoInitialize()
        except ImportError:
            pythoncom = None
    try:
        mailer = mailer or OutlookMailer()
        processed = 0
        failed = 0
        consecutive_failures = 0
        results: list[dict[str, str]] = []
        log_rows: list[list[str]] = []
        batch_items: list[dict[str, Any]] = []
        for plan in targets:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            to = "; ".join(plan["to"])
            bcc = "; ".join(plan["bcc"])
            cc = "; ".join(plan["cc"])
            try:
                saved = mailer.create_draft(
                    to=to,
                    cc=cc,
                    bcc=bcc,
                    subject=plan["subject"],
                    body=plan["body"],
                    importance=plan["importance"],
                )
                result_text = "下書き作成済"
                processed += 1
                consecutive_failures = 0
                # 差し替え可能なメーラー（テスト用）は戻り値を返さないことがある。
                # その場合 entry_id が空になり、あとで送信対象から自然に外れる。
                info = saved if isinstance(saved, dict) else {}
                batch_items.append({
                    "employee_id": plan["employee_id"],
                    "name": plan["name"],
                    "to": to,
                    "cc": cc,
                    "bcc": bcc,
                    "subject": plan["subject"],
                    "entry_id": str(info.get("entry_id") or ""),
                    "store_id": str(info.get("store_id") or ""),
                    "state": SEND_STATE_DRAFT,
                    "created_at": now,
                })
            except Exception as exc:
                result_text = f"エラー: {exc}"[:250]
                failed += 1
                consecutive_failures += 1
            results.append({"employee_id": plan["employee_id"], "name": plan["name"],
                            "result": result_text})
            log_rows.append([now, plan["employee_id"], plan["name"], to, bcc,
                             plan["subject"], result_text])
            if consecutive_failures >= 3:
                _append_log(log_path, log_rows)
                raise RuntimeError(
                    "Outlook の下書き作成が3件連続で失敗したため中断しました。"
                    f"ログを確認してください: {log_path}")
        _append_log(log_path, log_rows)
        batch_path = ""
        batch_id = ""
        if batch_items:
            batch = {
                "batch_id": _new_batch_id(),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "user": current_user(),
                "template_name": template_name,
                "items": batch_items,
            }
            saved_path = save_send_batch(batch, log_dir=log_dir)
            batch_path = str(saved_path)
            batch_id = batch["batch_id"]
        return {"processed": processed, "skipped": skipped, "failed": failed,
                "results": results, "log_path": str(log_path),
                "batch_id": batch_id, "batch_path": batch_path,
                "sendable": len([i for i in batch_items if i["entry_id"]])}
    finally:
        if own_com:
            try:
                import pythoncom  # type: ignore
                pythoncom.CoUninitialize()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# 送信バッチ — 作った下書きを、あとでまとめて送るための控え
# ---------------------------------------------------------------------------
# 送るときは控えた EntryID で同じ下書きを引き当てるだけなので、Outlook 側で人が
# 本文や宛先を直していれば、直した内容がそのまま送られる（2026-08-31 実機確認）。
# EntryID は作った本人の Outlook プロファイルでしか解決できない＝他人が作った
# バッチは実質送れない。これが「誰でも押せる」ことに対する自然な歯止めになる。


def _new_batch_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_user(user: Any) -> str:
    """ファイル名に使えるユーザー名にする。"""
    cleaned = re.sub(r'[<>:"/|?*\\]', "_", str(user or "")).strip()
    return cleaned or "unknown"


def send_batch_dir(log_dir: str | Path | None = None) -> Path:
    return Path(log_dir or Config.MAIL_OUTPUT_DIR)


def _write_send_batch(path: str | Path, batch: dict[str, Any]) -> None:
    """壊れたJSONを残さないよう、一時ファイルへ書いてから置き換える。"""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(batch, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def save_send_batch(batch: dict[str, Any], *, log_dir: str | Path | None = None) -> Path:
    path = (send_batch_dir(log_dir)
            / f"{SEND_BATCH_PREFIX}{batch['batch_id']}_{_safe_user(batch.get('user'))}.json")
    _write_send_batch(path, batch)
    return path


def load_send_batch(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"送信バッチが見つかりません: {path}")
    try:
        with open(path, encoding="utf-8") as f:
            batch = json.load(f)
    except json.JSONDecodeError as exc:
        raise ValueError(f"送信バッチを読み取れません（{path}）: {exc}") from exc
    if not isinstance(batch, dict) or not isinstance(batch.get("items"), list):
        raise ValueError(f"送信バッチの形式が不正です（itemsの配列が必要）: {path}")
    return batch


def sendable_items(batch: dict[str, Any],
                   only_ids: list[str] | None = None) -> list[dict[str, Any]]:
    """まだ送っていない・EntryIDを控えてある行だけを返す。"""
    selected = {str(item) for item in only_ids} if only_ids is not None else None
    targets: list[dict[str, Any]] = []
    for item in batch.get("items") or []:
        if item.get("state") == SEND_STATE_SENT:
            continue
        if not str(item.get("entry_id") or ""):
            continue
        if selected is not None and str(item.get("employee_id") or "") not in selected:
            continue
        targets.append(item)
    return targets


def list_send_batches(*, user: str | None = None, log_dir: str | Path | None = None,
                      limit: int = 20) -> list[dict[str, Any]]:
    """自分が作ったバッチの概要を新しい順に返す（読み取りのみ）。

    他人のバッチは出さない。EntryID は作った本人の Outlook でしか解決できないので、
    出しても送れず、誤操作のもとになるだけのため。
    """
    folder = send_batch_dir(log_dir)
    if not folder.exists():
        return []
    who = current_user() if user is None else user
    summaries: list[dict[str, Any]] = []
    for path in sorted(folder.glob(SEND_BATCH_GLOB),
                       key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            batch = load_send_batch(path)
        except (ValueError, OSError):
            continue
        if who and normalize_name(str(batch.get("user") or "")) != normalize_name(who):
            continue
        items = batch.get("items") or []
        summaries.append({
            "batch_id": str(batch.get("batch_id") or path.stem),
            "path": str(path),
            "created_at": str(batch.get("created_at") or ""),
            "user": str(batch.get("user") or ""),
            "template_name": str(batch.get("template_name") or ""),
            "subject": str(items[0].get("subject") or "") if items else "",
            "total": len(items),
            "sendable": len(sendable_items(batch)),
            "sent": len([i for i in items if i.get("state") == SEND_STATE_SENT]),
        })
        if len(summaries) >= limit:
            break
    return summaries


def send_batch(
    batch_path: str | Path,
    *,
    only_ids: list[str] | None = None,
    confirm_text: str = "",
    log_dir: str | Path | None = None,
    mailer: Any | None = None,
) -> dict[str, Any]:
    """控えた下書きを引き当てて送る。**新しいメールは組み立て直さない。**

    - 送る前に「SEND 件数」の確認文字を要求する（件数はこの関数が数え直す）。
    - 1件送るたびに控えを書き戻すので、途中で落ちても同じ人へ二重に送らない。
    - 引き当てられない下書き（消された・移動された）は送らずスキップして記録する。
    - 3件連続で失敗したら中断する（Outlook 側の異常を疑う）。
    """
    path = Path(batch_path)
    batch = load_send_batch(path)
    targets = sendable_items(batch, only_ids)
    required = f"SEND {len(targets)}"
    if str(confirm_text or "").strip() != required:
        raise ValueError(
            f"確認文字が一致しません。送信するには「{required}」と入力してください")
    if not targets:
        return {"processed": 0, "failed": 0, "results": [], "log_path": "",
                "batch_path": str(path)}

    log_path = send_batch_dir(log_dir) / LOG_FILENAME
    own_com = mailer is None
    if own_com:
        try:
            import pythoncom  # type: ignore
            pythoncom.CoInitialize()
        except ImportError:
            pythoncom = None
    try:
        mailer = mailer or OutlookMailer()
        processed = 0
        failed = 0
        consecutive_failures = 0
        results: list[dict[str, str]] = []
        log_rows: list[list[str]] = []
        for item in targets:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                mailer.send_saved(str(item.get("entry_id") or ""),
                                  str(item.get("store_id") or ""))
                item["state"] = SEND_STATE_SENT
                item["sent_at"] = now
                item.pop("last_error", None)
                result_text = "送信済"
                processed += 1
                consecutive_failures = 0
            except Exception as exc:
                result_text = f"送信できません（下書きが見つからない可能性）: {exc}"[:250]
                item["last_error"] = result_text
                failed += 1
                consecutive_failures += 1
            # 1件ごとに控えを書き戻す。途中で落ちても、再実行で同じ人へ二重に送らない。
            _write_send_batch(path, batch)
            results.append({"employee_id": str(item.get("employee_id") or ""),
                            "name": str(item.get("name") or ""),
                            "result": result_text})
            log_rows.append([now, str(item.get("employee_id") or ""),
                             str(item.get("name") or ""), str(item.get("to") or ""),
                             str(item.get("bcc") or ""), str(item.get("subject") or ""),
                             result_text])
            if consecutive_failures >= 3:
                _append_log(log_path, log_rows)
                raise RuntimeError(
                    "Outlook への送信が3件連続で失敗したため中断しました。"
                    f"ログを確認してください: {log_path}")
        _append_log(log_path, log_rows)
        return {"processed": processed, "failed": failed, "results": results,
                "log_path": str(log_path), "batch_path": str(path)}
    finally:
        if own_com:
            try:
                import pythoncom  # type: ignore
                pythoncom.CoUninitialize()
            except Exception:
                pass
