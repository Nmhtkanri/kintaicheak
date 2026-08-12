# -*- coding: utf-8 -*-
"""交通費申請 × 通勤費マスタ 突合精査（承認前チェック）

経費申請が承認される前に、金額と経路をjinjer通勤費マスタと突き合わせる。
既存の経路突合（経費統合モード）は仕訳データ＝承認済みしか見られないため、
差し戻せるうちに誤りを見つけるにはこちらが要る。

jinjer APIは叩かないので、承認が進むたびに何度でも軽く回せる。
出力先に前回の結果があれば読み、各行に「前回比（新規/継続/解消）」を入れる。

入力:
  - 交通費申請_9637_YYYYMMDD.csv   (jinjer経費エクスポート / CP932。進行中も含む)
  - 経費チェックYYYY年M月.xlsx      (勤怠チェッカー経費モード出力)
出力:
  - 交通費精査結果_YYYY年M月.xlsx
  - 通勤費未登録者_メール対象_YYYY年M月.xlsx （メール下書きモードの宛先表）

判定コード:
  A 全経路合算一致 / B 一部経路一致 / C 往復・片道の差 / D 金額不一致 / E 区分相違
"""
from __future__ import annotations

import csv
import io
import itertools
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- 判定に使う定数 -------------------------------------------------------
ACTIVE_STATUS = ("承認完了", "進行中")          # 取下げ・否認は精査対象外
KIND_PASS = "通勤定期代"                        # マスタ 支給間隔=毎月 と突合
KIND_ACTUAL = "通勤交通費（実費）"              # マスタ 支給間隔=毎日 と突合
KIND_TRAVEL = ("交通費（電車・バス）", "交通費（特急・新幹線）")  # 移動交通費=一覧化のみ

MAX_COMBO_LEGS = 12          # 経路数がこれを超える人は部分集合の全探索をしない
MIN_DAYS_FOR_DOMINANCE = 3   # 実費申請がこの日数未満だと常用経路を判定できない

# 通勤費（定期代＋実費）の月額上限。移動交通費（KIND_TRAVEL）には適用しない。
# 既定値は Config.KOTSUHI_MONTHLY_LIMIT から渡される。
COMMUTE_MONTHLY_LIMIT = 30000

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
NG_FILL = PatternFill("solid", fgColor="FCE4E4")     # 要確認
INFO_FILL = PatternFill("solid", fgColor="F2F2F2")   # 情報のみ
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")   # 注意
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --- 共通ユーティリティ ---------------------------------------------------
def to_int(value) -> int:
    """'8,970' / '8970.0' / '' → int（変換不能は0）"""
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return 0


def norm_station(value) -> str:
    """駅名の表記ゆれを吸収する。祖師ヶ谷大蔵/祖師ケ谷大蔵、東雲/東雲(東京都) を同一視。"""
    if value is None:
        return ""
    s = unicodedata.normalize("NFKC", str(value)).strip()
    s = re.sub(r"[（(].*?[）)]", "", s)          # 括弧内の地域名を落とす
    s = re.sub(r"[「」『』\[\]]", "", s)          # 入力時に紛れ込む括弧を落とす
    s = s.replace("ヶ", "ケ").replace("ヵ", "ケ")
    return s.replace("　", "").replace(" ", "")


def _load_member_reason_csv(path: Path | None) -> dict[str, str]:
    """「社員番号, 氏名, 理由」形式の名簿CSVを読み {社員番号: 理由} を返す。

    谷津さんが直す運用ファイルなので exe に同梱せず共有フォルダを読む。
    無い・壊れている場合は空 dict を返し、精査そのものは止めない。
    """
    if not path or not path.exists():
        return {}
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "cp932"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        return {}
    out = {}
    for row in list(csv.reader(io.StringIO(text)))[1:]:
        if row and row[0].strip():
            out[row[0].strip()] = (row[2].strip() if len(row) > 2 else "")
    return out


def load_excluded_members(path: Path | None) -> dict[str, str]:
    """通勤費の精査対象外リストを読み {社員番号: 理由} を返す。

    勤怠データからは判別できない事情で通勤費が発生しない人を落とすための表。
    全テレワークでも勤怠上テレワーク0で登録されている人がいるので、
    条件式では拾えず明示リストが要る。理由も持たせて根拠を残す。
    """
    return _load_member_reason_csv(path)


def load_limit_exempt_members(path: Path | None) -> dict[str, str]:
    """通勤費の上限免除者リストを読み {社員番号: 理由} を返す。

    通勤費は月 COMMUTE_MONTHLY_LIMIT 円が上限で超過分は基本的に切るが、
    個別に実費全額を認めている人がいる（2026-08-06 谷津さん指定）。
    金額からは判別できないので明示リストで持つ。
    """
    return _load_member_reason_csv(path)


def is_company_employee(emp: str) -> bool:
    """当社で給与計算する自社社員か。20YY始まりの7桁だけが該当する。

    落ちるもの:
      - シート末尾の「合計」行
      - 5/6/9 始まり（派遣・テスト番号）
      - 3333 始まり（2026-08-05 谷津さん確認。3333003=使用していない捨て垢、
        3333008=派遣社員。勤怠実績があっても当社の給与計算対象ではない）
    """
    return bool(re.fullmatch(r"20\d{5}", str(emp).strip()))


def read_cp932_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    text = path.read_bytes().decode("cp932")
    rows = list(csv.reader(io.StringIO(text)))
    return rows[0], rows[1:]


def col_index(header: list[str], name: str, nth: int = 0) -> int:
    """CSVヘッダは計上部門などが申請/明細で重複するので、出現順で引く。"""
    seen = 0
    for i, h in enumerate(header):
        if h.strip() == name:
            if seen == nth:
                return i
            seen += 1
    raise KeyError(f"列が見つかりません: {name}")


# --- マスタ ---------------------------------------------------------------
class CommuteMaster:
    """経費チェックブックの「通勤費」シート = jinjer 通勤情報。

    1人が複数経路を持ち、意味は2通りある。
      (a) 乗継の区間分割 → 合計が1日/1か月の支給額
      (b) 勤務地パターンの選択制 → いずれか1経路が正
    どちらか判別できないので、合計一致と部分集合一致の両方を当てにいく。
    """

    def __init__(self, ws):
        self.rows: dict[str, list[dict]] = defaultdict(list)
        self.skipped: list[str] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is None or not str(r[0]).strip():
                continue
            emp = str(r[0]).strip()
            if not is_company_employee(emp):
                self.skipped.append(emp)
                continue
            self.rows[emp].append(
                {
                    "氏名": r[1],
                    "経路No": r[2],
                    "出発": r[3],
                    "到着": r[4],
                    "通勤経路": r[7],
                    "利用交通機関": r[8],
                    "支給間隔": str(r[9] or ""),
                    "支給金額": to_int(r[11]),
                    "支給開始": r[14],
                }
            )

    def legs(self, emp: str, interval: str) -> list[dict]:
        return [m for m in self.rows.get(emp, []) if m["支給間隔"] == interval and m["支給金額"] > 0]

    def total(self, emp: str, interval: str) -> int:
        return sum(m["支給金額"] for m in self.legs(emp, interval))

    def subset_sums(self, emp: str, interval: str) -> set[int]:
        amounts = [m["支給金額"] for m in self.legs(emp, interval)]
        if len(amounts) > MAX_COMBO_LEGS:
            return {sum(amounts)}
        sums = set()
        for n in range(1, len(amounts) + 1):
            for combo in itertools.combinations(amounts, n):
                sums.add(sum(combo))
        return sums

    def leg_pairs(self, emp: str, interval: str) -> set[tuple[str, str]]:
        return {(norm_station(m["出発"]), norm_station(m["到着"])) for m in self.legs(emp, interval)}

    def route_text(self, emp: str, interval: str) -> str:
        return " / ".join(
            f"{m['出発']}→{m['到着']}({m['支給金額']:,})" for m in self.legs(emp, interval)
        )

    def other_interval_total(self, emp: str, interval: str) -> int:
        other = "毎日" if interval == "毎月" else "毎月"
        return self.total(emp, other)


# --- 判定 -----------------------------------------------------------------
def judge_amount(applied: int, master: CommuteMaster, emp: str, interval: str) -> tuple[str, str]:
    """(判定コード, 説明) を返す。"""
    if not master.legs(emp, interval):
        other = master.other_interval_total(emp, interval)
        if other:
            kind = "定期" if interval == "毎日" else "実費"
            return "E", f"マスタは{kind}登録（{other:,}円）で申請区分と相違"
        return "E", "通勤費マスタに該当区分の登録なし"

    total = master.total(emp, interval)
    if applied == total:
        return "A", "全経路の合算と一致"
    if applied in master.subset_sums(emp, interval):
        return "B", "登録経路の一部と一致"
    if applied * 2 in master.subset_sums(emp, interval):
        return "C", "片道のみ申請の疑い（×2でマスタ一致）"
    if applied % 2 == 0 and applied // 2 in master.subset_sums(emp, interval):
        return "C", "往復で二重計上の疑い（÷2でマスタ一致）"
    diff = applied - total
    sign = "過大" if diff > 0 else "過少"
    return "D", f"マスタ計{total:,}円と{sign} {abs(diff):,}円の差"


def station_eq(a: str, b: str) -> bool:
    """CSVエクスポートで駅名が途中で切れることがある（'…※旧' と '…※旧ＣＨＯ'）ので前方一致も許す。"""
    if not a or not b:
        return False
    if a == b:
        return True
    short, long_ = (a, b) if len(a) <= len(b) else (b, a)
    # 「上宿」と「上宿/立川バス」はjinjer上の同じ停留所。区切りが入るなら長さは問わない
    if long_.startswith(short) and long_[len(short):].startswith("/"):
        return True
    return len(short) >= 4 and long_.startswith(short)


def pair_in(pair: tuple[str, str], pairs: set[tuple[str, str]]) -> bool:
    return any(station_eq(pair[0], p[0]) and station_eq(pair[1], p[1]) for p in pairs)


def judge_route(applied_pairs: set[tuple[str, str]], master_pairs: set[tuple[str, str]]) -> str:
    if not master_pairs:
        return "マスタ経路なし"
    applied_pairs = {p for p in applied_pairs if p[0] or p[1]}
    if not applied_pairs:
        return "申請経路なし"
    hit = [pair_in(p, master_pairs) for p in applied_pairs]
    if all(hit):
        return "一致"
    if any(hit):
        return "一部相違"
    return "不一致"


def severity(code: str, applied: int, master_total: int) -> str:
    """要確認 / 情報 / OK。過少申請は会社に損害が無いので情報止まり。"""
    if code in ("A", "B"):
        return "OK"
    if code == "D" and applied < master_total:
        return "情報"
    return "要確認"


# --- 集計 -----------------------------------------------------------------
def build_pass_rows(details, idx, master: CommuteMaster) -> list[dict]:
    """定期代は分割申請があるので人単位で合算してから突合する。"""
    per = defaultdict(lambda: {"金額": 0, "件数": 0, "申請書": [], "pairs": set(), "経路": []})
    for r in details:
        if r[idx["交通機関"]] != KIND_PASS or r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        emp = r[idx["社員番号"]]
        acc = per[(emp, r[idx["申請者"]], r[idx["所属グループ"]])]
        acc["金額"] += to_int(r[idx["小計"]])
        acc["件数"] += 1
        acc["申請書"].append(r[idx["申請書No."]])
        acc["ステータス"] = r[idx["ステータス"]]
        frm, to = r[idx["乗車場所"]], r[idx["降車場所"]]
        acc["pairs"].add((norm_station(frm), norm_station(to)))
        acc["経路"].append(f"{frm}→{to}")

    out = []
    for (emp, name, group), acc in sorted(per.items()):
        code, note = judge_amount(acc["金額"], master, emp, "毎月")
        m_total = master.total(emp, "毎月")
        out.append(
            {
                "社員番号": emp,
                "氏名": name,
                "所属グループ": group,
                "ステータス": acc["ステータス"],
                "申請書No.": ", ".join(sorted(set(acc["申請書"]))),
                "明細件数": acc["件数"],
                "申請額(合計)": acc["金額"],
                "マスタ支給額": m_total,
                "差額": acc["金額"] - m_total,
                "判定": code,
                "区分": severity(code, acc["金額"], m_total),
                "経路判定": judge_route(acc["pairs"], master.leg_pairs(emp, "毎月")),
                "申請経路": " / ".join(dict.fromkeys(acc["経路"])),
                "マスタ経路": master.route_text(emp, "毎月"),
                "説明": note,
            }
        )
    return out


def build_limit_over_rows(details, idx, exempt: dict[str, str],
                          limit: int = COMMUTE_MONTHLY_LIMIT,
                          travel_members: set[str] | None = None) -> list[dict]:
    """通勤費（定期代＋実費）の月合計が上限を超えた人を洗い出す。

    当社の通勤費は月3万円が上限で、超過分は基本的に切る。上限は定期代だけでなく
    **実費の月累計にも掛かる**（2026-08-06 谷津さん確認）。移動交通費（立替精算）は
    上限が無いので合算しない。

    上限が掛からない人が2種類いるので、どちらも OK 扱いにして理由を書き分ける。
      - 上限免除者リストの人 … 会社が個別に実費全額を認めている
      - 移動交通費（立替精算）対象者 … 通勤費ではなく立替精算で計上する人。
        jinjer 上は通勤系で申請されるので、金額だけでは判別できない
        （2026-08 山田大海さんがこれで要確認に挙がった）
    **どちらのリストにも無いのに超過している人だけ**を要確認に落とす
    （＝上限を切り忘れて満額払ってしまう事故の検知が目的）。

    定期代は分割申請、実費は日ごとの複数明細で出てくるので、どちらも人単位で
    合算してから判定する。
    """
    travel_members = travel_members or set()
    per = defaultdict(lambda: {"定期": 0, "実費": 0, "申請書": set()})
    for r in details:
        kind = r[idx["交通機関"]]
        if kind not in (KIND_PASS, KIND_ACTUAL) or r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        emp = r[idx["社員番号"]]
        if not is_company_employee(emp):
            continue
        acc = per[(emp, r[idx["申請者"]], r[idx["所属グループ"]])]
        acc["定期" if kind == KIND_PASS else "実費"] += to_int(r[idx["小計"]])
        acc["申請書"].add(r[idx["申請書No."]])

    hint = "免除リストに無いので切るか、許可済みなら docs の 通勤費_上限免除者.csv に追加してください"
    out = []
    for (emp, name, group), acc in sorted(per.items()):
        total = acc["定期"] + acc["実費"]
        if total <= limit:
            continue
        reason = exempt.get(emp)
        is_travel = emp in travel_members
        is_exempt = emp in exempt or is_travel
        if is_travel:
            note = "移動交通費（立替精算）対象者。通勤費ではなく立替精算で計上するため上限なし"
        elif reason:
            note = "上限免除者。" + reason
        elif is_exempt:
            note = "上限免除者（実費全額）"
        else:
            note = f"上限 {limit:,}円 を {total - limit:,}円 超過。" + hint
        out.append(
            {
                "社員番号": emp,
                "氏名": name,
                "所属グループ": group,
                "申請書No.": ", ".join(sorted(acc["申請書"])),
                "通勤費合計": total,
                "うち定期代": acc["定期"],
                "うち実費": acc["実費"],
                "上限": limit,
                "超過額": total - limit,
                "上限免除": "○" if is_exempt else "",
                "区分": "OK" if is_exempt else "要確認",
                "説明": note,
            }
        )
    return out


def build_actual_rows(details, idx, master: CommuteMaster, workdays: dict) -> list[dict]:
    """実費は1日を乗換ごとに複数明細で申請するため、日単位で合算する。

    ただしマスタが定期(毎月)登録の人は、実費カテゴリで月額をまとめて申請してくる。
    この人たちを日単位で毎日額(=0)と比べると全額が過大に見えるので、
    月合計を1行にまとめて毎月額と突合する。
    """
    per = defaultdict(lambda: {"金額": 0, "件数": 0, "pairs": set(), "経路": [], "申請書": set()})
    for r in details:
        if r[idx["交通機関"]] != KIND_ACTUAL or r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        emp = r[idx["社員番号"]]
        key = (emp, r[idx["申請者"]], r[idx["所属グループ"]], r[idx["利用日"]])
        acc = per[key]
        acc["金額"] += to_int(r[idx["小計"]])
        acc["件数"] += 1
        acc["申請書"].add(r[idx["申請書No."]])
        acc["ステータス"] = r[idx["ステータス"]]
        frm, to = r[idx["乗車場所"]], r[idx["降車場所"]]
        acc["pairs"].add((norm_station(frm), norm_station(to)))
        acc["経路"].append(f"{frm}→{to}({to_int(r[idx['金額']]):,}/{r[idx['往復']]})")

    daily_emps = {k[0] for k in per if master.legs(k[0], "毎日")}
    out = []

    for (emp, name, group, date), acc in sorted(per.items(), key=lambda x: (x[0][0], _date_key(x[0][3]))):
        if emp not in daily_emps:
            continue
        code, note = judge_amount(acc["金額"], master, emp, "毎日")
        m_total = master.total(emp, "毎日")
        out.append(
            {
                "社員番号": emp,
                "氏名": name,
                "所属グループ": group,
                "利用日": date,
                "ステータス": acc["ステータス"],
                "申請書No.": ", ".join(sorted(acc["申請書"])),
                "明細件数": acc["件数"],
                "申請額(日計)": acc["金額"],
                "マスタ日額": m_total,
                "差額": acc["金額"] - m_total,
                "判定": code,
                "区分": severity(code, acc["金額"], m_total),
                "経路判定": judge_route(acc["pairs"], master.leg_pairs(emp, "毎日")),
                "申請経路": " / ".join(acc["経路"]),
                "マスタ経路": master.route_text(emp, "毎日"),
                "説明": note,
            }
        )

    # マスタが定期登録の人 → 月合計を1行にまとめて毎月額と突合
    monthly = defaultdict(lambda: {"金額": 0, "件数": 0, "日": set(), "pairs": set(),
                                   "経路": [], "申請書": set(), "ステータス": ""})
    for (emp, name, group, date), acc in per.items():
        if emp in daily_emps:
            continue
        m = monthly[(emp, name, group)]
        m["金額"] += acc["金額"]
        m["件数"] += acc["件数"]
        m["日"].add(date)
        m["pairs"] |= acc["pairs"]
        m["経路"].extend(acc["経路"])
        m["申請書"] |= acc["申請書"]
        m["ステータス"] = acc["ステータス"]

    for (emp, name, group), m in sorted(monthly.items()):
        m_total = master.total(emp, "毎月")
        applied = m["金額"]
        if not m_total:
            code, note, kubun = "E", "通勤費マスタに登録なし（新入社員・経路未登録の疑い）", "要確認"
        elif applied == m_total or applied in master.subset_sums(emp, "毎月"):
            code = "E"
            note = f"金額はマスタ定期額({m_total:,}円)と一致。申請区分が実費になっているだけ"
            kubun = "情報"
        else:
            diff = applied - m_total
            code = "E"
            note = f"マスタは定期({m_total:,}円)。区分相違かつ金額も{'過大' if diff > 0 else '過少'} {abs(diff):,}円"
            kubun = "要確認" if diff > 0 else "情報"
        out.append(
            {
                "社員番号": emp,
                "氏名": name,
                "所属グループ": group,
                "利用日": f"(7月合計 {len(m['日'])}日分)",
                "ステータス": m["ステータス"],
                "申請書No.": ", ".join(sorted(m["申請書"])),
                "明細件数": m["件数"],
                "申請額(日計)": applied,
                "マスタ日額": m_total,
                "差額": applied - m_total,
                "判定": code,
                "区分": kubun,
                "経路判定": judge_route(m["pairs"], master.leg_pairs(emp, "毎月")),
                "申請経路": " / ".join(dict.fromkeys(m["経路"]))[:400],
                "マスタ経路": master.route_text(emp, "毎月"),
                "説明": note,
            }
        )

    # 日数の整合と、テレワーク実施日への交通費申請の検出
    days = defaultdict(set)
    for (emp, _n, _g, date) in per:
        days[emp].add(date)
    for row in out:
        emp = row["社員番号"]
        info = workdays.get(emp)
        applied = len(days[emp])
        shussha = info["出社日数"] if info else "-"
        row["申請日数/出社日数"] = f"{applied} / {shussha}"

        tw_days = info["テレワーク実施日"] if info else set()
        if row["利用日"].startswith("("):
            # 月合計行は、その人の申請日のうちテレワーク日と重なる日を列挙
            hit = sorted(d for d in days[emp] if _md(d) in tw_days)
        else:
            hit = [row["利用日"]] if _md(row["利用日"]) in tw_days else []
        if hit:
            row["テレワーク重複"] = f"{len(hit)}日: " + "、".join(_md(d) for d in hit[:8])
            row["説明"] = (row["説明"] + " ／ " if row["説明"] else "") + \
                "テレワーク実施日に交通費申請"
            if row["区分"] == "OK":
                row["区分"] = "要確認"
        else:
            row["テレワーク重複"] = ""
    return out


def _md(value: str) -> str:
    """'2026/7/1' → '7/1'（勤怠サマリのテレワーク実施日と同じ表記に揃える）"""
    m = re.match(r"\d{4}/(\d{1,2})/(\d{1,2})", str(value))
    return f"{int(m.group(1))}/{int(m.group(2))}" if m else str(value)


def _date_key(value: str):
    m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", str(value))
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else (0, 0, 0)


def year_month(value) -> str:
    m = re.match(r"(\d{4})/(\d{1,2})/", str(value))
    return f"{m.group(1)}-{int(m.group(2)):02d}" if m else ""


def build_no_commute_rows(details, idx, master: CommuteMaster, workdays: dict,
                          target_ids: set[str], excluded: dict[str, str]) -> list[dict]:
    """通勤費申請なしリストを自動抽出する（2026-08-05 谷津さん決定=C案）。

    条件は「当月の通勤費申請なし」。
    定期支給者は毎月申請しないのが正常で、この人たちの通勤費はマスタから支給する。
    だから申請が無いこと自体は異常ではなく、リストは支給用の一覧として使う。
    そのうえで**マスタにも登録が無い人**は1円も出ていない可能性があるので印を付ける。

    出勤実績0でも落とさない（2026-08-05 谷津さん指摘）。代表取締役や打刻申請の
    無い人がいて、その人たちも経費対象者。出勤0を在職の代用にすると取りこぼす。
    対象月に在職しているかは経費モード側の入社日・退職日で絞り込む前提。

    テレワーク実施者の扱い（2026-08-12 谷津さん決定＝抽出条件の拡張）:
    以前は「テレワークが1日でもあれば対象外」だったが、それだとテレワークと実費申請が
    混在する人が丸ごと抜け落ちていた。実費申請者はテレワーク日には申請しないのが正常で、
    **実費申請日数 ＋ テレワーク日数 ＝ 出勤日数** なら全日に説明が付くので通過させる。
    式が合わない人は日数の内訳を添えて要確認に出す。
    """
    # 定期代の申請がある人は定期代突合シートの担当なのでこのリストからは外す。
    # 実費は「何日申請したか」が判定に要るので、ユニークな利用日を数える
    # （build_actual_rows と同じ数え方＝往復で2行あっても同じ日なら1日）。
    applied_pass: set[str] = set()
    actual_days: dict[str, set[str]] = defaultdict(set)
    for r in details:
        if r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        if r[idx["交通機関"]] == KIND_PASS:
            applied_pass.add(r[idx["社員番号"]])
        elif r[idx["交通機関"]] == KIND_ACTUAL:
            actual_days[r[idx["社員番号"]]].add(r[idx["利用日"]])

    rows = []
    for emp, info in workdays.items():
        if not is_company_employee(emp) or emp in applied_pass or emp in excluded:
            continue
        work = info["出勤日数"] if isinstance(info["出勤日数"], (int, float)) else 0
        tw = info["テレワーク日数"] if isinstance(info["テレワーク日数"], (int, float)) else 0
        actual = len(actual_days.get(emp, ()))
        if tw <= 0:
            # テレワークなし: 実費の申請が1件でもあれば申請済み＝このリストの対象外
            if actual:
                continue
        elif int(round(actual + tw)) == int(round(work)):
            # 実費申請日とテレワーク日で出勤日が全部埋まる＝支給漏れではない
            continue

        legs = sorted(master.rows.get(emp, []), key=lambda m: m["経路No"] or 0)
        amount = sum(m["支給金額"] for m in legs)
        interval = legs[0]["支給間隔"] if legs else ""
        kubun = {"毎月": "通勤定期代", "毎日": "通勤費"}.get(interval, "")
        remark = ""
        if emp in target_ids:
            kubun = "移動交通費"
            remark = "立替精算対象"
        state = "マスタ未登録" if not legs else "マスタの支給金額が0"
        tw_note = f"（テレワーク{int(tw)}日）" if tw > 0 else ""
        if actual > 0:
            # ここに来るのはテレワークありで式が合わない人だけ（テレワーク無しの実費申請者は
            # 上で除外済み）。どちらに振れているかで意味が変わるので向きを書き分ける。
            judge = "実費申請の日数不一致"
            direction = ("実費申請がテレワーク日と重なっている疑い"
                         if actual + tw > work else "出社日の一部しか実費申請が無い（申請漏れの可能性）")
            note = (f"出勤{int(work)}日・テレワーク{int(tw)}日・実費申請{actual}日で"
                    f"一致しない（{direction}）")
            kubun_judge = "要確認"
        elif work <= 0:
            # 代表取締役や打刻申請の無い人も出勤0になる。支給漏れとは断定できない。
            # 対象月より後の入社もここに落ちるが、それは経費モード側の入社日判定で消える。
            judge = "マスタから支給" if amount else "勤怠実績なし"
            note = (f"対象月の勤怠実績なし。マスタに{interval}{amount:,}円の登録あり"
                    if amount else f"対象月の勤怠実績がなく、{state}。入社日・在籍状況を確認")
            kubun_judge = "情報"
        elif amount == 0:
            judge = "支給漏れの疑い"
            note = f"{int(work - tw)}日出社しているが、{state}かつ通勤費の申請も無い{tw_note}"
            kubun_judge = "要確認"
        else:
            judge = "マスタから支給"
            note = f"申請は無いがマスタに{interval}{amount:,}円の登録あり{tw_note}"
            kubun_judge = "OK"

        rows.append({
            "社員番号": emp, "氏名": info["氏名"],
            "出勤日数": int(work), "出社日数": int(work - tw),
            "支給金額": amount, "区分": kubun,
            "利用交通機関": legs[0]["利用交通機関"] if legs else "",
            "備考": remark,
            "判定": judge,
            "説明": note,
            "確認要否": kubun_judge,
        })
    order = {"支給漏れの疑い": 0, "実費申請の日数不一致": 1, "勤怠実績なし": 2, "マスタから支給": 3}
    rows.sort(key=lambda r: (order.get(r["判定"], 9), r["社員番号"]))
    return rows


def build_workdays(wb) -> dict[str, dict]:
    """経費チェックブックの「サマリ」から勤務日数を読む。

    テレワーク日数(D列)と出社日数(E列)は数式（`COUNTIF(テレワーク明細!A:A,…)` と `C-D`）で、
    経費モードが書いたばかりのブックには**計算結果が保存されていない**。
    素直に読むと全員テレワーク0＝全日出社に見えてしまうので、
    値が無いときは「テレワーク明細」（1日1行の生データ）から数え直す。
    Excelで一度開いて保存されたブックには計算結果が入るため、そちらはそのまま使える。
    """
    tw_detail: dict[str, int] = defaultdict(int)
    if "テレワーク明細" in wb.sheetnames:
        for r in wb["テレワーク明細"].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                tw_detail[str(r[0]).strip()] += 1

    workdays: dict[str, dict] = {}
    for r in wb["サマリ"].iter_rows(min_row=2, values_only=True):
        if r[0] is None or not str(r[0]).strip():
            continue
        emp = str(r[0]).strip()
        tw = {d.strip() for d in re.split(r"[、,]", str(r[5] or "")) if d.strip()}
        work = r[2] if isinstance(r[2], (int, float)) else 0
        tw_days = r[3] if isinstance(r[3], (int, float)) else tw_detail.get(emp, len(tw))
        workdays[emp] = {
            "氏名": r[1], "出勤日数": work, "テレワーク日数": tw_days,
            "出社日数": work - tw_days, "テレワーク実施日": tw,
        }
    return workdays


def build_mail_targets(details, idx, master: CommuteMaster, workdays: dict,
                       excluded: dict[str, str], target_ym: str) -> list[dict]:
    """通勤費が1円も出ていない可能性がある人を、メール下書きモードの宛先表にする。

    条件は「当月の通勤費申請なし」かつ「jinjer通勤費マスタに金額の登録なし」。
    マスタに登録がある人は申請が無くてもそこから支給されるので対象にしない。

    メール下書きモードの一覧表は**先頭シートしか読まない**ため別ファイルで出す。
    テンプレートは {{列名}} で差し込むので、本文で使う値を列として持たせる。
    """
    applied: set[str] = set()
    for r in details:
        if r[idx["ステータス"]] in ACTIVE_STATUS and r[idx["交通機関"]] in (KIND_PASS, KIND_ACTUAL):
            applied.add(r[idx["社員番号"]])

    y, m = target_ym.split("-")
    rows = []
    for emp, info in workdays.items():
        if not is_company_employee(emp) or emp in applied or emp in excluded:
            continue
        legs = master.rows.get(emp, [])
        if sum(x["支給金額"] for x in legs):
            continue
        work = info["出勤日数"] if isinstance(info["出勤日数"], (int, float)) else 0
        tw = info["テレワーク日数"] if isinstance(info["テレワーク日数"], (int, float)) else 0
        rows.append({
            "社員番号": emp,
            "氏名": info["氏名"] or "",
            "対象月": f"{int(m)}月",
            "出社日数": int(work - tw),
            "状況": "マスタ未登録" if not legs else "登録はあるが金額が0",
            # 出勤実績が無い人は入社したばかりの可能性がある。分けて見えるようにする
            "区分": "当月の勤務実績あり" if work > 0 else "当月の勤務実績なし（入社直後の可能性）",
        })
    rows.sort(key=lambda r: (r["区分"], r["社員番号"]))
    return rows


def write_mail_target_book(rows: list[dict], path: Path) -> Path:
    """宛先表を単独ブックで書き出す（メール下書きモードは先頭シートしか読まない）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "通勤費未登録者"
    cols = ["社員番号", "氏名", "対象月", "出社日数", "状況", "区分"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for row in rows:
        ws.append([row.get(c, "") for c in cols])
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).border = BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(ws.max_row, 1)}"
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = {
            "社員番号": 12, "氏名": 16, "対象月": 9, "出社日数": 10,
            "状況": 20, "区分": 34}.get(name, 12)
    wb.save(path)
    return path


def _dominant_legs(details, idx) -> dict[str, set[tuple[str, str]]]:
    """その人が月を通して常用している通勤経路だけを返す。

    申請日数の半分以上に出てくる区間を「常用」とみなす。
    たまの出張・寄り道は閾値未満になるので落ちる。
    """
    legs: dict[str, dict[tuple[str, str], set]] = defaultdict(lambda: defaultdict(set))
    days: dict[str, set] = defaultdict(set)
    has_pass: set[str] = set()
    for r in details:
        if r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        if r[idx["交通機関"]] not in (KIND_PASS, KIND_ACTUAL):
            continue
        emp, date = r[idx["社員番号"]], r[idx["利用日"]]
        pair = (norm_station(r[idx["乗車場所"]]), norm_station(r[idx["降車場所"]]))
        if not pair[0] and not pair[1]:
            continue
        if r[idx["交通機関"]] == KIND_PASS:
            has_pass.add(emp)
        days[emp].add(date)
        legs[emp][pair].add(date)

    out = {}
    for emp, pairs in legs.items():
        total = len(days[emp]) or 1
        if emp in has_pass:
            # 定期代の申請区間はその人の通勤経路そのもの。日数によらず常用とみなす
            out[emp] = set(pairs)
        elif total >= MIN_DAYS_FOR_DOMINANCE:
            out[emp] = {p for p, ds in pairs.items() if len(ds) / total >= 0.5}
        else:
            # 申請が数日しかない人は、たまたま乗った区間と常用区間を区別できない
            out[emp] = set()
    return out


def build_master_gap_rows(details, idx, master: CommuteMaster, workdays: dict,
                          pass_rows: list[dict], actual_rows: list[dict],
                          excluded: dict[str, str]) -> tuple[list[dict], list[dict]]:
    """jinjer通勤費マスタ側の更新漏れを検知する。

    通勤費マスタは給与の支給元なので、ここが古いまま=誤支給に直結する。
    申請との突合で出た食い違い(M1-M4)に、マスタ単独で見て怪しい状態(M5,M6,M8)を足す。
    在籍/退職はこのブックから判別できないため、7月に勤怠実績が無い人は
    「棚卸し」として別シートに分け、断定はしない。
    """
    applied_kinds: dict[str, set[str]] = defaultdict(set)
    for r in details:
        if r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        kind = r[idx["交通機関"]]
        if kind in (KIND_PASS, KIND_ACTUAL):
            applied_kinds[r[idx["社員番号"]]].add(kind)

    gaps: list[dict] = []

    def add(emp, name, code, title, detail, action, amount=0, priority=2):
        gaps.append({
            "社員番号": emp, "氏名": name or (workdays.get(emp, {}).get("氏名") or ""),
            "検知区分": code, "内容": title, "detail": detail,
            "対応": action, "関係金額": amount, "_p": priority,
            "区分": "要確認",
        })

    # --- M1〜M4: 申請との突合で出た食い違いをマスタ目線に読み替える ---
    # M2は「その月ずっとその経路で通っている」ときだけ拾う。
    # 出張の寄り道を1〜2回混ぜただけの人まで経路変更として挙げると使い物にならない。
    dominant = _dominant_legs(details, idx)

    for row in pass_rows + actual_rows:
        if row["区分"] == "OK":
            continue
        emp, name = row["社員番号"], row["氏名"]
        code = row["判定"]
        mpairs = master.leg_pairs(emp, "毎月") | master.leg_pairs(emp, "毎日")
        route_ng = any(not pair_in(p, mpairs) for p in dominant.get(emp, set())) and bool(mpairs)
        if code == "E" and "登録なし" in row["説明"]:
            add(emp, name, "M1", "マスタ未登録",
                f"通勤費の申請があるがマスタに経路が無い（申請 {row.get('申請額(合計)') or row.get('申請額(日計)'):,}円）",
                "jinjer通勤情報に経路を登録する", row.get("申請額(合計)") or row.get("申請額(日計)"), 1)
        elif code == "E":
            add(emp, name, "M4", "支給区分の相違",
                row["説明"].split(" ／ ")[0],
                "マスタの支給間隔（毎月=定期／毎日=実費）を実態に合わせる", abs(row["差額"]), 2)
        elif route_ng:
            common = "、".join(f"{a}→{b}" for a, b in sorted(dominant.get(emp, set()))
                              if not pair_in((a, b), mpairs))
            add(emp, name, "M2", "経路の相違（勤務地変更の未反映）",
                f"常用 {common} ／ マスタ {row['マスタ経路']}",
                "勤務地変更ならマスタの経路と金額を更新する", abs(row["差額"]), 1)
        elif code == "D" and row["差額"] > 0:
            add(emp, name, "M3", "金額の相違（運賃改定の未反映）",
                f"経路は一致。申請が {row['差額']:+,}円（マスタ {row.get('マスタ支給額') or row.get('マスタ日額'):,}円）",
                "運賃改定ならマスタの支給金額を更新する", row["差額"], 2)

    # 同一人・同一区分の重複をまとめる（実費は日単位で出るため）
    merged: dict[tuple, dict] = {}
    for g in gaps:
        key = (g["社員番号"], g["検知区分"])
        if key in merged:
            merged[key]["_n"] = merged[key].get("_n", 1) + 1
            merged[key]["関係金額"] += g["関係金額"]
            continue
        g["_n"] = 1
        merged[key] = g
    gaps = list(merged.values())
    for g in gaps:
        n = g.pop("_n", 1)
        g["内訳"] = g.pop("detail")
        if n > 1:
            g["内訳"] = f"[{n}件] " + g["内訳"]

    # --- M5: 出社実績があるのに通勤費の登録も申請も無い（支給漏れの疑い） ---
    for emp, info in workdays.items():
        if not is_company_employee(emp) or emp in excluded:
            continue
        shussha = info["出社日数"] if isinstance(info["出社日数"], (int, float)) else 0
        if shussha <= 0:
            continue
        if master.total(emp, "毎月") or master.total(emp, "毎日") or emp in applied_kinds:
            continue
        state = "マスタ未登録" if emp not in master.rows else "マスタの支給金額が0"
        gaps.append({
            "社員番号": emp, "氏名": info["氏名"], "検知区分": "M5",
            "内容": "通勤費が未設定（支給漏れの疑い）",
            "内訳": f"7月に{int(shussha)}日出社しているが、{state}かつ通勤費の申請も無い",
            "対応": "通勤手段を確認し、マスタ登録または実費申請を促す",
            "関係金額": 0, "区分": "要確認", "_p": 1,
        })

    # --- M6a: 申請側で乗車/降車が空欄（jinjer入力の不備） ---
    blank = defaultdict(lambda: [0, ""])
    for r in details:
        if r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        if r[idx["交通機関"]] not in (KIND_PASS, KIND_ACTUAL):
            continue
        if not str(r[idx["乗車場所"]]).strip() and not str(r[idx["降車場所"]]).strip():
            blank[r[idx["社員番号"]]][0] += 1
            blank[r[idx["社員番号"]]][1] = r[idx["申請者"]]
    for emp, (cnt, name) in blank.items():
        gaps.append({
            "社員番号": emp, "氏名": name, "検知区分": "M6",
            "内容": "申請の経路が空欄",
            "内訳": f"乗車場所・降車場所が未入力の明細が{cnt}件。経路突合ができない",
            "対応": "申請者に経路の入力を依頼する",
            "関係金額": 0, "区分": "要確認", "_p": 2,
        })

    # --- M6: マスタの登録が不完全（金額0・支給間隔未選択） ---
    for emp, legs in master.rows.items():
        info = workdays.get(emp)
        for m in legs:
            issues = []
            if m["支給間隔"] in ("未選択", "None", ""):
                issues.append("支給間隔が未選択")
            if m["支給金額"] == 0 and (m["出発"] or m["到着"]):
                issues.append("支給金額が0")
            if not issues:
                continue
            gaps.append({
                "社員番号": emp, "氏名": (info or {}).get("氏名") or m["氏名"] or "",
                "検知区分": "M6", "内容": "マスタの登録が不完全",
                "内訳": f"{'・'.join(issues)}（経路 {m['出発']}→{m['到着']} / {m['支給金額']:,}円）",
                "対応": "jinjer通勤情報の必須項目を埋める",
                "関係金額": m["支給金額"],
                "区分": "要確認" if info else "情報", "_p": 2 if info else 3,
            })

    # --- M8: テレワーク中心なのに定期を支給し続けている ---
    for emp, info in workdays.items():
        teiki = master.total(emp, "毎月")
        tw = info["テレワーク日数"] if isinstance(info["テレワーク日数"], (int, float)) else 0
        shussha = info["出社日数"] if isinstance(info["出社日数"], (int, float)) else 0
        if teiki > 0 and tw > 0 and shussha <= 5:
            gaps.append({
                "社員番号": emp, "氏名": info["氏名"], "検知区分": "M8",
                "内容": "テレワーク中心だが定期を支給中",
                "内訳": f"7月は出社{int(shussha)}日・テレワーク{int(tw)}日。定期 {teiki:,}円/月",
                "対応": "実費支給への切替を検討する",
                "関係金額": teiki, "区分": "要確認", "_p": 2,
            })

    gaps.sort(key=lambda g: (g["_p"], g["検知区分"], g["社員番号"]))
    for g in gaps:
        g.pop("_p", None)

    # --- 棚卸し: 7月の勤怠実績が無いのにマスタに金額が残っている ---
    stock = []
    for emp, legs in master.rows.items():
        if emp in workdays:
            continue
        amount = sum(m["支給金額"] for m in legs)
        if amount <= 0:
            continue
        stock.append({
            "社員番号": emp, "氏名": legs[0]["氏名"] or "",
            "マスタ支給額": amount,
            "支給間隔": "・".join(sorted({m["支給間隔"] for m in legs})),
            "支給開始": min(str(m["支給開始"]) for m in legs),
            "経路": master.route_text(emp, "毎月") or master.route_text(emp, "毎日"),
            "区分": "要確認",
            "確認事項": "7月の勤怠実績なし。退職・休職なら経路の削除要否を確認する",
        })
    stock.sort(key=lambda s: -s["マスタ支給額"])
    return gaps, stock


def build_travel_rows(details, idx, target_ids: set[str]) -> tuple[list[dict], list[dict]]:
    """移動交通費は通勤費マスタに相手がいないので金額突合はしない。一覧化＋リスト外フラグ。"""
    per = defaultdict(lambda: {"金額": 0, "件数": 0, "日": set()})
    detail_rows = []
    for r in details:
        if r[idx["交通機関"]] not in KIND_TRAVEL or r[idx["ステータス"]] not in ACTIVE_STATUS:
            continue
        emp = r[idx["社員番号"]]
        amount = to_int(r[idx["小計"]])
        acc = per[(emp, r[idx["申請者"]], r[idx["所属グループ"]])]
        acc["金額"] += amount
        acc["件数"] += 1
        acc["日"].add(r[idx["利用日"]])
        detail_rows.append(
            {
                "社員番号": emp,
                "氏名": r[idx["申請者"]],
                "利用日": r[idx["利用日"]],
                "ステータス": r[idx["ステータス"]],
                "申請書No.": r[idx["申請書No."]],
                "交通機関": r[idx["交通機関"]],
                "目的地": r[idx["目的地"]],
                "乗車場所": r[idx["乗車場所"]],
                "降車場所": r[idx["降車場所"]],
                "往復": r[idx["往復"]],
                "金額": to_int(r[idx["金額"]]),
                "小計": amount,
                "対象者リスト": "○" if emp in target_ids else "リスト外",
                "備考": r[idx["備考", 1]] if ("備考", 1) in idx else "",
            }
        )

    summary = []
    for (emp, name, group), acc in sorted(per.items(), key=lambda x: -x[1]["金額"]):
        summary.append(
            {
                "社員番号": emp,
                "氏名": name,
                "所属グループ": group,
                "明細件数": acc["件数"],
                "利用日数": len(acc["日"]),
                "金額合計": acc["金額"],
                "対象者リスト": "○" if emp in target_ids else "リスト外",
                "区分": "OK" if emp in target_ids else "要確認",
                "説明": "" if emp in target_ids else "移動交通費対象者リストに未登録",
            }
        )
    return summary, detail_rows


# --- Excel 出力 -----------------------------------------------------------
def write_sheet(wb, title: str, rows: list[dict], columns: list[str], widths: dict | None = None,
                severity_key: str = "区分"):
    ws = wb.create_sheet(title)
    ws.append(columns)
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row in rows:
        ws.append([row.get(c, "") for c in columns])
        r = ws.max_row
        kubun = row.get(severity_key, "")
        fill = {"要確認": NG_FILL, "情報": INFO_FILL}.get(kubun)
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if columns[c - 1] in ("申請額(合計)", "申請額(日計)", "マスタ支給額", "マスタ日額",
                                  "差額", "金額合計", "金額", "小計",
                                  "通勤費合計", "うち定期代", "うち実費", "上限", "超過額"):
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row, 1)}"
    widths = widths or {}
    for i, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, _default_width(name))
    return ws


def _default_width(name: str) -> int:
    if name in ("申請経路", "マスタ経路", "説明"):
        return 46
    if name in ("所属グループ", "申請書No."):
        return 18
    if name in ("氏名", "利用日", "経路判定", "対象者リスト"):
        return 13
    return 11


def write_summary(wb, pass_rows, actual_rows, travel_rows, gap_rows, stock_rows,
                  no_commute_rows, meta: dict):
    ws = wb.create_sheet("サマリ", 0)
    ws["A1"] = "交通費 精査結果サマリ"
    ws["A1"].font = Font(bold=True, size=14)
    r = 3
    for k, v in meta.items():
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    def block(title, rows, amount_key):
        nonlocal r
        r += 1
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        for c, h in enumerate(["区分", "件数", "金額合計"], start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill, cell.font = HDR_FILL, HDR_FONT
            cell.alignment = Alignment(horizontal="center")
        r += 1
        agg = defaultdict(lambda: [0, 0])
        for row in rows:
            a = agg[row.get("区分", "")]
            a[0] += 1
            a[1] += row.get(amount_key, 0) or 0
        for kubun in ("要確認", "情報", "OK"):
            if kubun not in agg:
                continue
            cnt, amt = agg[kubun]
            ws.cell(row=r, column=1, value=kubun)
            ws.cell(row=r, column=2, value=cnt)
            cell = ws.cell(row=r, column=3, value=amt)
            cell.number_format = "#,##0"
            if kubun == "要確認":
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = NG_FILL
            r += 1
        ws.cell(row=r, column=1, value="合計").font = Font(bold=True)
        ws.cell(row=r, column=2, value=len(rows)).font = Font(bold=True)
        cell = ws.cell(row=r, column=3, value=sum(x.get(amount_key, 0) or 0 for x in rows))
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        r += 1

    block("■ 通勤定期代（人単位）", pass_rows, "申請額(合計)")
    block("■ 通勤交通費・実費（日単位）", actual_rows, "申請額(日計)")
    block("■ 移動交通費（人単位・金額突合なし）", travel_rows, "金額合計")

    # マスタ更新漏れは検知区分ごとに件数を出す
    r += 1
    ws.cell(row=r, column=1, value="■ マスタ更新漏れ（検知区分別）").font = Font(bold=True, size=11)
    r += 1
    for c, h in enumerate(["検知区分", "件数", "内容"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    r += 1
    gap_agg = defaultdict(lambda: [0, ""])
    for g in gap_rows:
        gap_agg[g["検知区分"]][0] += 1
        gap_agg[g["検知区分"]][1] = g["内容"]
    for code in sorted(gap_agg):
        cnt, title = gap_agg[code]
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=title)
        r += 1
    ws.cell(row=r, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(gap_rows)).font = Font(bold=True)
    r += 2
    ws.cell(row=r, column=1, value="■ 通勤費申請なし（自動抽出）").font = Font(bold=True, size=11)
    r += 1
    # 判定ごとに数える。支給漏れの疑いと日数不一致はどちらも確認要否=要確認なので、
    # 確認要否で数えると2種類が混ざってしまう。
    by_judge: dict[str, int] = defaultdict(int)
    for x in no_commute_rows:
        by_judge[x["判定"]] += 1
    for label, note in [
        ("支給漏れの疑い", "出社しているが、マスタ登録も当月の申請も無い"),
        ("実費申請の日数不一致", "実費申請日数＋テレワーク日数が出勤日数と合わない"),
        ("勤怠実績なし", "対象月の勤怠実績が無い（入社日・在籍状況の確認が必要）"),
        ("マスタから支給", "申請は無いがマスタに登録あり＝この金額で支給する"),
    ]:
        n = by_judge.get(label, 0)
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=note)
        if label in ("支給漏れの疑い", "実費申請の日数不一致") and n:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = NG_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ マスタ棚卸し").font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="対象")
    ws.cell(row=r, column=2, value=len(stock_rows))
    ws.cell(row=r, column=3, value="7月の勤怠実績なし × マスタに支給額あり（在籍状況の確認が必要）")
    r += 1

    r += 1
    ws.cell(row=r, column=1, value="判定コードの意味").font = Font(bold=True, size=11)
    r += 1
    for code, desc in [
        ("A", "登録経路の合算と完全一致（乗継を区間分割して登録しているケース）"),
        ("B", "登録経路の一部と一致（複数勤務地パターンを登録しているケース）"),
        ("C", "往復・片道の取り違えが疑われる（×2 または ÷2 でマスタと一致）"),
        ("D", "金額不一致。過大は要確認、過少は情報のみ（会社に損害なし）"),
        ("E", "区分相違。マスタが定期なのに実費申請、または通勤費マスタ未登録"),
        ("", ""),
        ("M1", "マスタ未登録。申請はあるがjinjer通勤情報に経路が無い"),
        ("M2", "経路の相違。勤務地変更がマスタに反映されていない疑い"),
        ("M3", "金額の相違。運賃改定がマスタに反映されていない疑い"),
        ("M4", "支給区分の相違。マスタの毎月/毎日が実態と合っていない"),
        ("M5", "出社実績があるのに通勤費の登録も申請も無い（支給漏れの疑い）"),
        ("M6", "マスタの登録が不完全（支給金額0・支給間隔が未選択）"),
        ("M8", "テレワーク中心だが定期を支給中。実費への切替検討"),
    ]:
        ws.cell(row=r, column=1, value=code).font = Font(bold=True)
        ws.cell(row=r, column=2, value=desc)
        r += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 16
    return ws


# --- メイン ---------------------------------------------------------------
def row_key(sheet: str, row: dict) -> str:
    """前回実行との突き合わせに使う行キー。同じ指摘を同じものと見なせる粒度にする。"""
    emp = str(row.get("社員番号") or "")
    if sheet == "実費突合":
        return f"{sheet}|{emp}|{row.get('利用日')}"
    if sheet == "マスタ更新漏れ":
        return f"{sheet}|{emp}|{row.get('検知区分')}"
    return f"{sheet}|{emp}"


def read_previous_keys(path: Path) -> dict[str, set[str]]:
    """前回の出力ブックから「要確認だった行」のキーを読む。

    承認が進むたびに再実行する運用なので、毎回ゼロから読み直すのは辛い。
    前回どれが挙がっていたかを持っておき、新規／継続／解消を出せるようにする。
    壊れていても実行は止めない（差分が出ないだけ）。
    """
    if not path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return {}
    out: dict[str, set[str]] = {}
    try:
        for sheet in ("通勤費申請なし", "マスタ更新漏れ", "定期代突合", "実費突合",
                      "通勤費上限超過", "移動交通費"):
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            head = [str(c or "") for c in rows[0]]
            sev = "確認要否" if "確認要否" in head else ("区分" if "区分" in head else None)
            keys = set()
            for r in rows[1:]:
                d = dict(zip(head, r))
                if not d.get("社員番号"):
                    continue
                if sev and str(d.get(sev) or "") != "要確認":
                    continue
                keys.add(row_key(sheet, d))
            out[sheet] = keys
    finally:
        wb.close()
    return out


def apply_diff(sheet: str, rows: list[dict], prev: dict[str, set[str]],
               severity_key: str = "区分") -> int:
    """各行に「前回比」を入れ、前回あって今回消えた件数（解消）を返す。"""
    if not prev:
        for r in rows:
            r["前回比"] = ""
        return 0
    before = prev.get(sheet, set())
    now = set()
    for r in rows:
        key = row_key(sheet, r)
        flagged = str(r.get(severity_key) or "") == "要確認"
        if flagged:
            now.add(key)
            r["前回比"] = "継続" if key in before else "新規"
        else:
            r["前回比"] = "解消" if key in before else ""
    return len(before - now)


@dataclass
class SeisaInputs:
    """精査の入力一式。main と仕訳追記（shiwake_teiki_append）で共用する。"""
    details: list                  # 対象月に絞った申請明細
    idx: dict                      # 列名 → CSVの列位置
    master: CommuteMaster
    workdays: dict
    target_ids: set                # 移動交通費（立替精算）対象者
    excluded: dict                 # 精査対象外者 {社員番号: 理由}
    limit_exempt: dict             # 通勤費の上限免除者 {社員番号: 理由}
    out_of_month: int = 0          # 対象月外の有効明細数（警告用）


def load_seisa_inputs(csv_path: Path, check_path: Path, target_ym: str,
                      target_list: Path | None = None,
                      excluded_list: Path | None = None,
                      limit_exempt_list: Path | None = None) -> SeisaInputs:
    """交通費申請CSVと経費チェックブックを読み、精査に要る入力をまとめて返す。"""
    header, all_details = read_cp932_csv(csv_path)
    names = ["ステータス", "交通機関", "社員番号", "申請者", "所属グループ", "申請書No.",
             "明細No.", "利用日", "金額", "往復", "小計", "乗車場所", "降車場所", "経路", "目的地"]
    idx = {n: col_index(header, n) for n in names}

    # 定期代は翌月分を同じ申請書にまとめて出す運用があるため、利用日の年月で対象月に絞る。
    # これを外すと「7月分＋8月分」が二重申請に見えてしまう。
    details = [r for r in all_details if year_month(r[idx["利用日"]]) == target_ym]
    out_of_month = sum(
        1 for r in all_details
        if r[idx["ステータス"]] in ACTIVE_STATUS and year_month(r[idx["利用日"]]) != target_ym
    )

    wb_src = openpyxl.load_workbook(check_path, data_only=True)
    try:
        master = CommuteMaster(wb_src["通勤費"])
        workdays = build_workdays(wb_src)
    finally:
        wb_src.close()

    target_ids = set()
    if target_list and target_list.exists():
        raw = target_list.read_bytes()
        for enc in ("utf-8-sig", "cp932"):
            try:
                txt = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        for row in list(csv.reader(io.StringIO(txt)))[1:]:
            if row and row[0].strip():
                target_ids.add(row[0].strip())

    return SeisaInputs(
        details=details, idx=idx, master=master, workdays=workdays,
        target_ids=target_ids,
        excluded=load_excluded_members(excluded_list),
        limit_exempt=load_limit_exempt_members(limit_exempt_list),
        out_of_month=out_of_month,
    )


def main(csv_path: Path, check_path: Path, out_path: Path,
         target_list: Path | None = None, target_ym: str = "2026-07",
         excluded_list: Path | None = None,
         limit_exempt_list: Path | None = None,
         monthly_limit: int = COMMUTE_MONTHLY_LIMIT):
    src = load_seisa_inputs(csv_path, check_path, target_ym,
                            target_list=target_list, excluded_list=excluded_list,
                            limit_exempt_list=limit_exempt_list)
    details, idx, master, workdays = src.details, src.idx, src.master, src.workdays
    target_ids, excluded, limit_exempt = src.target_ids, src.excluded, src.limit_exempt
    out_of_month = src.out_of_month

    pass_rows = build_pass_rows(details, idx, master)
    actual_rows = build_actual_rows(details, idx, master, workdays)
    limit_rows = build_limit_over_rows(details, idx, limit_exempt, monthly_limit, target_ids)
    travel_rows, travel_details = build_travel_rows(details, idx, target_ids)
    gap_rows, stock_rows = build_master_gap_rows(details, idx, master, workdays, pass_rows, actual_rows, excluded)
    no_commute_rows = build_no_commute_rows(details, idx, master, workdays, target_ids, excluded)

    # 承認が進むたびに再実行する運用。前回どれが挙がっていたかと突き合わせる
    prev = read_previous_keys(out_path)
    resolved = 0
    resolved += apply_diff("通勤費申請なし", no_commute_rows, prev, "確認要否")
    resolved += apply_diff("マスタ更新漏れ", gap_rows, prev)
    resolved += apply_diff("定期代突合", pass_rows, prev)
    resolved += apply_diff("実費突合", actual_rows, prev)
    resolved += apply_diff("通勤費上限超過", limit_rows, prev)
    resolved += apply_diff("移動交通費", travel_rows, prev)
    new_count = sum(1 for rows in (no_commute_rows, gap_rows, pass_rows, actual_rows,
                                   limit_rows, travel_rows)
                    for r in rows if r.get("前回比") == "新規")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_sheet(wb, "通勤費申請なし", no_commute_rows,
                ["社員番号", "氏名", "出勤日数", "出社日数", "支給金額", "区分",
                 "利用交通機関", "判定", "確認要否", "前回比", "説明", "備考"],
                {"説明": 52, "判定": 16, "確認要否": 10}, severity_key="確認要否")
    write_sheet(wb, "マスタ更新漏れ", gap_rows,
                ["検知区分", "社員番号", "氏名", "内容", "内訳", "対応", "関係金額", "区分", "前回比"],
                {"内容": 30, "内訳": 62, "対応": 40})
    write_sheet(wb, "マスタ棚卸し", stock_rows,
                ["社員番号", "氏名", "マスタ支給額", "支給間隔", "支給開始", "経路", "確認事項", "区分"],
                {"経路": 46, "確認事項": 46})
    write_sheet(wb, "定期代突合", pass_rows,
                ["社員番号", "氏名", "所属グループ", "ステータス", "申請書No.", "明細件数",
                 "申請額(合計)", "マスタ支給額", "差額", "判定", "区分", "前回比", "経路判定",
                 "申請経路", "マスタ経路", "説明"])
    write_sheet(wb, "実費突合", actual_rows,
                ["社員番号", "氏名", "所属グループ", "利用日", "ステータス", "申請書No.", "明細件数",
                 "申請額(日計)", "マスタ日額", "差額", "判定", "区分", "前回比", "経路判定",
                 "申請日数/出社日数", "テレワーク重複", "申請経路", "マスタ経路", "説明"])
    write_sheet(wb, "通勤費上限超過", limit_rows,
                ["社員番号", "氏名", "所属グループ", "申請書No.", "通勤費合計", "うち定期代",
                 "うち実費", "上限", "超過額", "上限免除", "区分", "前回比", "説明"],
                {"説明": 62, "上限免除": 9})
    write_sheet(wb, "移動交通費", travel_rows,
                ["社員番号", "氏名", "所属グループ", "明細件数", "利用日数", "金額合計",
                 "対象者リスト", "区分", "前回比", "説明"])
    write_sheet(wb, "移動交通費明細", travel_details,
                ["社員番号", "氏名", "利用日", "ステータス", "申請書No.", "交通機関", "目的地",
                 "乗車場所", "降車場所", "往復", "金額", "小計", "対象者リスト"])

    total_rows = sum(1 for r in details if r[idx["ステータス"]] in ACTIVE_STATUS)
    approved = sum(1 for r in details if r[idx["ステータス"]] == "承認完了")
    pending = sum(1 for r in details if r[idx["ステータス"]] == "進行中")
    y, m = target_ym.split("-")
    write_summary(wb, pass_rows, actual_rows, travel_rows, gap_rows, stock_rows, no_commute_rows, {
        "実行日時": _now_text(),
        "対象月": f"{y}年{int(m)}月",
        # 承認が進むたびに再実行する。進行中が0になったらこの月は見終わり
        "承認状況": f"承認完了 {approved}行 / 進行中 {pending}行",
        "前回からの変化": (f"新規 {new_count}件 / 解消 {resolved}件" if prev else "初回実行"),
        "交通費申請CSV": csv_path.name,
        "経費チェックブック": check_path.name,
        "対象明細行(承認完了+進行中)": total_rows,
        "除外(取下げ・否認)": len(details) - total_rows,
        "除外(対象月外の利用日)": out_of_month,
        "通勤費マスタ": f"{len(master.rows)}名 / {sum(len(v) for v in master.rows.values())}経路",
        "マスタ対象外(合計行・非20YY番号)": "、".join(master.skipped) or "なし",
        "精査対象外リスト": f"{len(excluded)}名（docs\通勤費_精査対象外者.csv）" if excluded else "なし",
        "通勤費の上限": f"月 {monthly_limit:,}円（移動交通費は対象外）",
        "上限免除リスト": (f"{len(limit_exempt)}名（docs\通勤費_上限免除者.csv）"
                     if limit_exempt else "なし（リスト未配置＝超過者は全員 要確認になります）"),
        "上限超過": (f"{len(limit_rows)}名（うち免除 "
                 f"{sum(1 for r in limit_rows if r['上限免除'])}名 / "
                 f"要確認 {sum(1 for r in limit_rows if r['区分'] == '要確認')}名）"
                 if limit_rows else "なし"),
    })

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # メール下書きモード用の宛先表（先頭シートしか読まれないので別ファイル）
    mail_rows = build_mail_targets(details, idx, master, workdays, excluded, target_ym)
    mail_path = out_path.with_name(f"通勤費未登録者_メール対象_{y}年{int(m)}月.xlsx")
    write_mail_target_book(mail_rows, mail_path)

    return out_path, pass_rows, actual_rows, travel_rows, mail_path, mail_rows


def _now_text() -> str:
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M")


@dataclass
class PreReviewResult:
    ok: bool
    output_path: Path
    mail_path: "Path | None" = None
    month: str = ""
    error: "str | None" = None
    approved_rows: int = 0
    pending_rows: int = 0
    new_count: int = 0
    resolved_count: int = 0
    flagged: "dict[str, int] | None" = None
    mail_targets: int = 0
    first_run: bool = True


def run_pre_approval_review(
    csv_path: "str | Path",
    check_xlsx: "str | Path",
    output_path: "str | Path",
    month: str,
    log_func=print,
) -> PreReviewResult:
    """承認前の交通費精査。jinjer APIは叩かないので何度でも軽く回せる。

    経費申請は承認が進むたびに状態が変わるため、同じ月で繰り返し実行する前提。
    出力先に前回の結果があれば読み、各行に「前回比（新規/継続/解消）」を入れる。
    サマリの「承認状況」で進行中が0になれば、その月は見終わったと判断できる。
    """
    from config import Config

    csv_path, check_xlsx = Path(csv_path), Path(check_xlsx)
    output_path = Path(output_path)
    result = PreReviewResult(ok=False, output_path=output_path, month=month)

    if not csv_path.exists():
        result.error = f"交通費申請CSVが見つかりません: {csv_path}"
        return result
    if not check_xlsx.exists():
        result.error = f"経費チェックのブックが見つかりません: {check_xlsx}"
        return result

    result.first_run = not output_path.exists()
    travel_csv = Path(getattr(Config, "KEIHI_TRAVEL_EXPENSE_MEMBERS_CSV", "")) \
        if getattr(Config, "KEIHI_TRAVEL_EXPENSE_MEMBERS_CSV", "") else None
    excluded_csv = Path(getattr(Config, "KOTSUHI_EXCLUDED_MEMBERS_CSV", "")) \
        if getattr(Config, "KOTSUHI_EXCLUDED_MEMBERS_CSV", "") else None
    exempt_csv = Path(getattr(Config, "KOTSUHI_LIMIT_EXEMPT_MEMBERS_CSV", "")) \
        if getattr(Config, "KOTSUHI_LIMIT_EXEMPT_MEMBERS_CSV", "") else None
    monthly_limit = int(getattr(Config, "KOTSUHI_MONTHLY_LIMIT", COMMUTE_MONTHLY_LIMIT))
    if exempt_csv and not exempt_csv.exists():
        # 上限超過の判定自体は続ける。免除者が全員 要確認 に出るので気づける。
        log_func(f"[warn] 通勤費の上限免除者リストが見つかりません（免除なしで続行）: {exempt_csv}")

    try:
        out, pass_rows, actual_rows, travel_rows, mail_path, mail_rows = main(
            csv_path, check_xlsx, output_path,
            target_list=travel_csv, target_ym=month, excluded_list=excluded_csv,
            limit_exempt_list=exempt_csv, monthly_limit=monthly_limit,
        )
    except Exception as e:  # noqa: BLE001
        log_func(f"[error] 精査に失敗しました: {e}")
        result.error = str(e)
        return result

    wb = openpyxl.load_workbook(out, data_only=True, read_only=True)
    try:
        flagged = {}
        for sheet in ("通勤費申請なし", "マスタ更新漏れ", "定期代突合", "実費突合",
                      "通勤費上限超過", "移動交通費"):
            if sheet not in wb.sheetnames:
                continue
            rows = list(wb[sheet].iter_rows(values_only=True))
            head = [str(c or "") for c in rows[0]] if rows else []
            sev = "確認要否" if "確認要否" in head else ("区分" if "区分" in head else None)
            if not sev:
                continue
            n = sum(1 for r in rows[1:] if dict(zip(head, r)).get(sev) == "要確認")
            flagged[sheet] = n
        summary = {str(r[0] or ""): r[1] for r in wb["サマリ"].iter_rows(values_only=True) if r and r[0]}
    finally:
        wb.close()

    status = str(summary.get("承認状況") or "")
    m_ap = re.search(r"承認完了\s*(\d+)", status)
    m_pd = re.search(r"進行中\s*(\d+)", status)
    change = str(summary.get("前回からの変化") or "")
    m_new = re.search(r"新規\s*(\d+)", change)
    m_res = re.search(r"解消\s*(\d+)", change)

    result.ok = True
    result.output_path = out
    result.mail_path = mail_path
    result.approved_rows = int(m_ap.group(1)) if m_ap else 0
    result.pending_rows = int(m_pd.group(1)) if m_pd else 0
    result.new_count = int(m_new.group(1)) if m_new else 0
    result.resolved_count = int(m_res.group(1)) if m_res else 0
    result.flagged = flagged
    result.mail_targets = len(mail_rows)

    log_func(f"[info] 承認状況: 承認完了 {result.approved_rows}行 / 進行中 {result.pending_rows}行")
    if result.first_run:
        log_func("[info] 初回実行（前回比なし）")
    else:
        log_func(f"[info] 前回から 新規 {result.new_count}件 / 解消 {result.resolved_count}件")
    for sheet, n in flagged.items():
        if n:
            log_func(f"[info] 要確認 {sheet}: {n}件")
    log_func(f"[info] 通勤費未登録（メール対象）: {result.mail_targets}名")
    if result.pending_rows == 0:
        log_func("[done] 進行中の申請はありません。この月の精査は完了です")
    return result
