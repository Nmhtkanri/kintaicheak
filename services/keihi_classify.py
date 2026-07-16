"""経費分類・集計エンジン（経費マクロ移植 P1b）

経費一覧表マクロの `c設定.bas::Run_経費集計_設定シート版`（分類→社員別集計）を Python へ忠実移植する。
入力は keihi_summary が生成する 34 列の経費統合一覧表、出力は「集計」「集計ログ」シート。

移植元との対応:
  - LoadKeywordsFromSetting → DEFAULT_KEYWORDS（live 設定シート全量スナップショット）＋ load_keywords()
  - Collect_From_Source     → classify_rows()（判定順序・judgeText/nonTaxText・ガードまで同一）
  - Rewrite_Output          → aggregate_by_id() ＋ add_summary_sheet()
  - iチェック表作成マクロ    → build_detail_totals()（M/N/O チェック列）

マクロとの意図的な違い（2026-07 新入社員計上漏れ事故の根本対策）:
  - マクロは集計シートA列の既存社員だけ更新し未登録社員を無言スキップするが、
    本実装は**データに出てきた全社員**を集計シートに出力する（20YY始まりのみ。5/6/9始まりはログ集計）。

忠実移植の要点（c設定.bas live 実測 2026-07-16。live=フォルダ.bas diff 0 確認済み）:
  - judgeText の「請求区分」成分は FindCol("請求区分") が **請求区分ID列（9列目）に先ヒット**するため ID 列の値。
  - estFilled ガード（顧客請求分あり＆夜間当番不一致→スキップ）は GoTo NextR のため**計上日の集計もスキップ**される。
  - テレワーク手当の判定結果ラベルはマクロのまま "J:テレワーク手当"（実際の集計先は K 列）。
  - キーワード照合は LCase 格納＋vbTextCompare（大小・全半角同一視）→ NFKC + casefold で再現。
  - 金額 0 の行は分類されない（G先行と計上日のみ処理）。負数括弧 `(123)` は -123。
"""

from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.keihi_summary import in_company_scope

FONT = "Meiryo UI"

# ----------------------------------------------------------------------
# 統合一覧表の列（0-based）。FindCol（部分一致・左から先勝ち）の着地を固定化したもの。
# ----------------------------------------------------------------------
COL_EMP = 0        # 社員番号
COL_NAME = 1       # 氏名
COL_AMT = 3        # 合計（FindCol("合計")）
COL_MEMO_REQ = 4   # 備考(申請書データ)
COL_TRANS = 6      # 交通機関
COL_UCH = 7        # 内訳
COL_BILLTYPE = 8   # ★請求区分ID（FindCol("請求区分") が 10列目より先にヒット）
COL_EXPTYPE = 10   # 費用種別
COL_FARE = 13      # 金額(交通費)
COL_MEMO_LINE = 19 # 備考(明細)
COL_BOOKED = 20    # 計上日(yyyy/mm/dd)
COL_SOURCE = 25    # 仕訳区分
COL_ESTAFF = 33    # 顧客請求分（FindCol不一致→34列フォールバック）

# bucket 配列の意味（agg 値 = 7要素リスト）
B_YAKAN, B_RINK, B_TRANS, B_ETC, B_TW, B_BILL, B_NONTAX = 0, 1, 2, 3, 4, 5, 6

# 分類名（設定シート A 列と一致させる。マクロの CAT_* 定数と同一）
CAT_YAKAN = "夜間当番手当"
CAT_RINK = "RINK手当"
CAT_TRANS = "交通費"
CAT_TW = "テレワーク手当"
CAT_NONTAX = "非課税精算(立替金)"
CAT_ETC = "その他(会議費・消耗品など)"
CAT_TRANS_NG = "交通費除外"
CAT_KOKYAKU_NG = "顧客請求除外"

# live「設定」シート全量スナップショット（2026-07-16 実測・行順保持。順序=先勝ちマッチに影響）
DEFAULT_KEYWORDS: dict[str, list[str]] = {
    CAT_YAKAN: [
        "夜間当番", "24時間準直当番", "準直当番", "深夜出動", "顧客当番", "顧客対応当番", "オンコール",
    ],
    CAT_RINK: ["RINK"],
    CAT_TW: ["テレワーク", "在宅"],
    CAT_TRANS: [
        "通勤交通費（実費）", "交通費", "電車", "バス", "タクシー", "地下鉄", "鉄道", "新幹線",
        "モノレール", "JR", "私鉄", "有料列車", "飛行機", "航空券", "定期券", "定期代",
        "ガソリン", "燃料", "駐車場", "パーキング", "高速", "ETC", "車両", "レンタカー",
        "移動", "旅費", "出張", "宿泊", "ホテル", "北総鉄道北総線", "通勤定期代",
        "多摩都市モノレール", "日ごとの通勤費",
    ],
    CAT_NONTAX: [
        "交通費（電車・バス）", "交通費（特急・新幹線）", "交通費（タクシー）", "交通費（航空機）",
        "旅費", "出張", "宿泊", "ホテル", "日当", "交通費（船舶）",
    ],
    CAT_TRANS_NG: [
        "会議", "交際", "接待", "飲食", "手土産", "福利厚生", "親睦", "定期健康診断", "健康診断",
    ],
    CAT_KOKYAKU_NG: ["夜間当番", "RINK", "顧客当番", "顧客対応当番", "オンコール"],
    CAT_ETC: [
        "会議接待費", "会議交際費", "会議費", "交際費", "接待", "飲食", "懇親", "手土産",
        "福利厚生", "消耗品", "消耗品費", "備品", "事務用品", "その他経費", "その他",
    ],
}
ALL_CATEGORIES = list(DEFAULT_KEYWORDS.keys())


def load_keywords(path: "str | Path") -> dict[str, list[str]]:
    """設定シート形式（A=分類名, B=キーワード）の xlsx / CSV を読み、キーワード辞書を返す。

    マクロの LoadKeywordsFromSetting と同じ規約:
      - 1行目はヘッダーとして読み飛ばす
      - B列（キーワード）が空の行はスキップ（live R69 の空欄対応）
      - C列以降は無視（live R6 の「手当」混入対応）
      - 未知の分類名はスキップ
    行順を保持する（先勝ちマッチに影響するため）。
    """
    path = Path(path)
    rows: list[tuple[str, str]] = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["設定"] if "設定" in wb.sheetnames else wb.worksheets[0]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                cat = str(row[0]).strip() if row and row[0] is not None else ""
                kw = str(row[1]).strip() if row and len(row) > 1 and row[1] is not None else ""
                rows.append((cat, kw))
        finally:
            wb.close()
    else:
        for enc in ("cp932", "utf-8-sig", "utf-8"):
            try:
                with open(path, encoding=enc, newline="") as f:
                    raw = list(csv.reader(f))
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"キーワード設定の文字コードを判別できませんでした: {path}")
        for r in raw[1:]:
            cat = (r[0] or "").strip() if r else ""
            kw = (r[1] or "").strip() if len(r) > 1 else ""
            rows.append((cat, kw))

    result: dict[str, list[str]] = {c: [] for c in ALL_CATEGORIES}
    for cat, kw in rows:
        if kw and cat in result:
            result[cat].append(kw)
    if not any(result.values()):
        raise ValueError(f"キーワード設定にキーワードが1件もありません: {path}")
    return result


# ----------------------------------------------------------------------
# VBA互換ヘルパー
# ----------------------------------------------------------------------

def normalize_str(v) -> str:
    """c設定.NormalizeStr: CR/LF/Tab/全角空白→半角空白 → Trim → LCase → 全空白除去。"""
    s = "" if v is None else str(v)
    for ch in ("\r", "\n", "\t", "　"):
        s = s.replace(ch, " ")
    s = s.strip().lower()
    return s.replace(" ", "")


def normalize_id(v) -> str:
    """c設定.NormalizeId: 数字のみ抽出。"""
    return re.sub(r"[^0-9]", "", "" if v is None else str(v))


def normalize_name(v) -> str:
    """c設定.NormalizeName: 全空白除去＋LCase。"""
    s = "" if v is None else str(v)
    return s.replace("　", "").replace(" ", "").lower()


def parse_amount_vba(v) -> float:
    """c設定.ParseAmount: カンマ/￥/円除去、括弧（全半角）→負号。非数値は0。"""
    s = "" if v is None else str(v)
    s = s.replace(",", "").replace("￥", "").replace("円", "")
    s = s.replace("(", "-").replace(")", "").replace("（", "-").replace("）", "")
    s = s.strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def try_parse_date(v) -> "date | None":
    """c設定.TryParseDate 相当。日付にパースできなければ None。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _fold(s: str) -> str:
    """vbTextCompare の大小・全半角同一視を NFKC + casefold で再現する。"""
    return unicodedata.normalize("NFKC", s).casefold()


def hit_any(text: str, keywords: list[str]) -> str:
    """c設定.HitAnyCollection: 登録順の先勝ち部分一致。ヒットしたキーワード（LCase）を返す。

    VBA は LCase 格納キーワード × LCase テキストを vbTextCompare（大小・全半角同一視）で
    照合するため、両辺を NFKC + casefold して比較する。
    """
    t = _fold(text)
    for kw in keywords:
        if _fold(kw) in t:
            return kw.lower()
    return ""


def build_emp_key(emp_no: str, emp_nm: str) -> str:
    """c設定.BuildEmpKey と同一のキー表現。"""
    return f"ID:{emp_no}|NM:{emp_nm}" if emp_no else f"NM:{emp_nm}"


def parse_emp_no(key: str) -> str:
    """c設定.ParseEmpNo: キーから数字のみIDを取り出す。ID無しキーは空文字。"""
    m = re.match(r"^ID:(.*?)\|NM:", key)
    if m:
        return normalize_id(m.group(1))
    if key.startswith("ID:"):
        return normalize_id(key[3:])
    return ""


# ----------------------------------------------------------------------
# 分類（Collect_From_Source の移植）
# ----------------------------------------------------------------------

@dataclass
class LogEntry:
    row_no: int          # 統合一覧表の行番号（2始まり）
    emp_no: str
    emp_nm: str          # 正規化後氏名（マクロと同じ）
    desc: str            # 内訳の生セル値（マクロと同じ）
    amount: float
    result: str          # 判定結果（"D:夜間当番手当" 等。マクロと同一文字列）
    matched_kw: str


@dataclass
class ClassifyResult:
    agg: dict = field(default_factory=dict)        # key -> [7 floats]
    max_date: dict = field(default_factory=dict)   # key -> date
    emp_names: dict = field(default_factory=dict)  # 数字ID -> 生氏名（初出）
    log: list = field(default_factory=list)        # list[LogEntry]
    hits: int = 0


def classify_rows(rows: list[list[str]], keywords: "dict | None" = None) -> ClassifyResult:
    """統合一覧表の全行を分類し、社員キー単位の bucket 集計とログを返す。"""
    kw = keywords or DEFAULT_KEYWORDS
    kw_yakan = kw.get(CAT_YAKAN, [])
    kw_rink = kw.get(CAT_RINK, [])
    kw_trans = kw.get(CAT_TRANS, [])
    kw_tw = kw.get(CAT_TW, [])
    kw_nontax = kw.get(CAT_NONTAX, [])
    kw_etc = kw.get(CAT_ETC, [])
    kw_trans_ng = kw.get(CAT_TRANS_NG, [])
    kw_kokyaku_ng = kw.get(CAT_KOKYAKU_NG, [])

    res = ClassifyResult()

    def cell(r, i):
        return r[i] if i < len(r) and r[i] is not None else ""

    for idx, r in enumerate(rows):
        row_no = idx + 2  # シート上の行番号（1行目ヘッダー）
        emp_no = normalize_id(cell(r, COL_EMP))
        emp_nm = normalize_name(cell(r, COL_NAME))
        if not emp_no and not emp_nm:
            continue

        key = build_emp_key(emp_no, emp_nm)
        if emp_no and emp_no not in res.emp_names:
            raw_name = str(cell(r, COL_NAME)).strip()
            if raw_name:
                res.emp_names[emp_no] = raw_name

        desc = normalize_str(cell(r, COL_UCH))
        trans = normalize_str(cell(r, COL_TRANS))
        bill_type = normalize_str(cell(r, COL_BILLTYPE))
        expense_type = normalize_str(cell(r, COL_EXPTYPE))
        memo_req = normalize_str(cell(r, COL_MEMO_REQ))
        memo_line = normalize_str(cell(r, COL_MEMO_LINE))
        source_mark = normalize_str(cell(r, COL_SOURCE))
        judge_text = " ".join([desc, trans, bill_type, expense_type, memo_req, memo_line, source_mark])
        est_filled = str(cell(r, COL_ESTAFF)).strip() != ""
        desc_raw = str(cell(r, COL_UCH))

        # --- G列: 顧客請求分（先行処理・社員番号必須・除外KWチェック）---
        if emp_no:
            est_amt = parse_amount_vba(cell(r, COL_ESTAFF))
            if est_amt != 0:
                matched = hit_any(desc, kw_kokyaku_ng)
                if matched == "":
                    bucket = res.agg.setdefault(key, [0.0] * 7)
                    bucket[B_BILL] += est_amt
                    res.hits += 1
                    res.log.append(LogEntry(row_no, emp_no, emp_nm, desc_raw, est_amt, "G:顧客請求分", ""))
                else:
                    res.log.append(LogEntry(row_no, emp_no, emp_nm, desc_raw, est_amt, "G:除外", matched))

        # --- 金額取得（合計、0なら金額(交通費)>0のみ代用）---
        amt = parse_amount_vba(cell(r, COL_AMT))
        if amt == 0:
            fa = parse_amount_vba(cell(r, COL_FARE))
            if fa > 0:
                amt = fa

        if amt != 0:
            # ガード: 顧客請求分あり かつ 夜間当番でない → スキップ
            # （マクロは GoTo NextR のため、この行は計上日の集計もされない）
            if est_filled and hit_any(desc, kw_yakan) == "":
                res.log.append(LogEntry(
                    row_no, emp_no, emp_nm, desc_raw, amt,
                    "D:顧客請求費ありのため除外", "AH優先（二重計上防止）"))
                continue

            bucket = res.agg.setdefault(key, [0.0] * 7)
            result_cat = ""
            matched = ""

            while True:  # GoTo Decided の代替（breakで確定）
                # 0. 本社経費は全て非課税精算(立替金)
                if "本社経費" in _fold(source_mark):
                    bucket[B_NONTAX] += amt
                    result_cat, matched = "I:非課税精算", "本社経費"
                    break
                # 1. 夜間当番手当
                matched = hit_any(desc, kw_yakan)
                if matched:
                    bucket[B_YAKAN] += amt
                    result_cat = "D:夜間当番手当"
                    break
                # 2. テレワーク手当（ログのラベルはマクロのまま J: 表記）
                matched = hit_any(desc, kw_tw)
                if matched:
                    if not est_filled:
                        bucket[B_TW] += amt
                    result_cat = "J:テレワーク手当"
                    break
                # 3. RINK手当
                matched = hit_any(desc, kw_rink)
                if matched:
                    bucket[B_RINK] += amt
                    result_cat = "E:RINK手当"
                    break
                # 4. 非課税精算(立替金) — 照合は nonTaxText（備考を含めない。2026-06-30 癖1修正）
                non_tax_text = " ".join([desc, trans, expense_type])
                matched = hit_any(non_tax_text, kw_nontax)
                if matched:
                    if not est_filled:
                        bucket[B_NONTAX] += amt
                    result_cat = "I:非課税精算"
                    break
                # 5. 交通費（内訳→交通機関の順。交通費除外KWは内訳で判定）
                matched = hit_any(desc, kw_trans)
                if not matched:
                    matched = hit_any(trans, kw_trans)
                if matched:
                    ng = hit_any(desc, kw_trans_ng)
                    if ng == "":
                        if not est_filled:
                            bucket[B_TRANS] += amt
                        result_cat = "H:交通費"
                    else:
                        if not est_filled:
                            bucket[B_ETC] += amt
                        result_cat = f"J:その他（交通費NG: {ng}）"
                        matched = f"{matched} → NG:{ng}"
                    break
                # 6. その他(会議費・消耗品など) — judgeText 全体で照合
                matched = hit_any(judge_text, kw_etc)
                if matched:
                    bucket[B_ETC] += amt
                    result_cat = "J:その他"
                    break
                # 7. デフォルト（無条件でその他）
                bucket[B_ETC] += amt
                result_cat = "J:その他"
                matched = "(該当キーワードなし)"
                break

            res.hits += 1
            res.log.append(LogEntry(row_no, emp_no, emp_nm, desc_raw, amt, result_cat, matched))

        # --- 計上日の最大値（金額0の行も対象。ガードスキップ行は到達しない）---
        d = try_parse_date(cell(r, COL_BOOKED))
        if d is not None:
            cur = res.max_date.get(key)
            if cur is None or d > cur:
                res.max_date[key] = d

    return res


# ----------------------------------------------------------------------
# 社員番号単位の再集約（Rewrite_Output 前半）
# ----------------------------------------------------------------------

@dataclass
class AggregateResult:
    by_id: dict = field(default_factory=dict)       # 数字ID -> [7 floats]
    date_by_id: dict = field(default_factory=dict)  # 数字ID -> date
    unmatched_rows: int = 0                          # ID無しキー（氏名のみ）の件数
    unmatched_amount: float = 0.0
    excluded_ids: list = field(default_factory=list)  # 5/6/9始まり（給与計算対象外）


def aggregate_by_id(cls_result: ClassifyResult) -> AggregateResult:
    """社員キー -> 数字ID に再集約する。ID の無いキー（氏名のみ）は未照合として回収。"""
    out = AggregateResult()
    for key, bucket in cls_result.agg.items():
        emp_id = parse_emp_no(key)
        if not emp_id:
            out.unmatched_rows += 1
            out.unmatched_amount += sum(bucket)
            continue
        cur = out.by_id.setdefault(emp_id, [0.0] * 7)
        for i in range(7):
            cur[i] += bucket[i]
    for key, d in cls_result.max_date.items():
        emp_id = parse_emp_no(key)
        if not emp_id or emp_id not in out.by_id:
            continue
        cur = out.date_by_id.get(emp_id)
        if cur is None or d > cur:
            out.date_by_id[emp_id] = d
    # 給与計算対象外（5/6/9始まり）を分離
    for emp_id in list(out.by_id.keys()):
        if not in_company_scope(emp_id):
            out.excluded_ids.append(emp_id)
    return out


def build_detail_totals(rows: list[list[str]]) -> dict[str, float]:
    """iチェック表作成マクロ.BuildDetailTotals: 社員IDごとの明細合計（D、D空欄時はAH）。"""
    totals: dict[str, float] = {}
    for r in rows:
        emp_id = normalize_id(r[COL_EMP] if len(r) > COL_EMP else "")
        if not emp_id:
            continue
        d_raw = str(r[COL_AMT]).strip() if len(r) > COL_AMT and r[COL_AMT] is not None else ""
        if d_raw == "":
            v = parse_amount_vba(r[COL_ESTAFF] if len(r) > COL_ESTAFF else "")
        else:
            v = parse_amount_vba(d_raw)
        totals[emp_id] = totals.get(emp_id, 0.0) + v
    return totals


# ----------------------------------------------------------------------
# Excel 出力
# ----------------------------------------------------------------------

SUMMARY_HEADERS = [
    "社員番号", "氏名", "合計", "夜間当番手当", "RINK手当", "手当2（夜間＋RINK）",
    "顧客請求分", "交通費", "非課税精算(立替金)", "その他(会議費・消耗品など)",
    "テレワーク手当", "請求日", "経費明細合計（D+D空欄時AH）", "差分（C - M）", "判定",
]


def _num(v: float):
    """金額セル値。整数なら int（マクロの Double 書込は Excel 上整数表示になるため揃える）。"""
    return int(v) if float(v) == int(v) else float(v)


def add_summary_sheet(
    wb: Workbook,
    agg: AggregateResult,
    emp_names: dict,
    detail_totals: "dict | None" = None,
    roster_names: "dict | None" = None,
) -> dict:
    """「集計」シートを追加する。行 = データに出てきた全社員（20YY始まりのみ・昇順）。

    新入社員も自動で行になる（マクロの計上漏れ事故の根本対策）。
    Returns: 統計 dict（社員数・対象外除外数など）。
    """
    detail_totals = detail_totals or {}
    roster_names = roster_names or {}
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="7030A0")
    body_font = Font(name=FONT)
    right = Alignment(horizontal="right")
    fill_ng = PatternFill("solid", start_color="C00000")
    fill_ok = PatternFill("solid", start_color="C6EFCE")
    fill_na = PatternFill("solid", start_color="D9D9D9")

    ids = sorted(
        {i for i in agg.by_id if in_company_scope(i)}
        | {i for i in detail_totals if in_company_scope(i)}
    )

    ws = wb.create_sheet("集計")
    ws.append(SUMMARY_HEADERS)
    for cell_ in ws[1]:
        cell_.font = header_font
        cell_.fill = header_fill
        cell_.alignment = Alignment(horizontal="center", wrap_text=True)

    for emp_id in ids:
        vals = agg.by_id.get(emp_id, [0.0] * 7)
        total = vals[B_YAKAN] + vals[B_RINK] + vals[B_BILL] + vals[B_TRANS] \
            + vals[B_NONTAX] + vals[B_ETC] + vals[B_TW]
        detail = detail_totals.get(emp_id, 0.0)
        diff = total - detail
        if detail == 0:
            verdict = "明細なし"
        elif diff == 0:
            verdict = "OK"
        else:
            verdict = "差分あり"
        name = emp_names.get(emp_id) or roster_names.get(emp_id, "")
        row = [
            emp_id, name, _num(total), _num(vals[B_YAKAN]), _num(vals[B_RINK]),
            _num(vals[B_YAKAN] + vals[B_RINK]), _num(vals[B_BILL]), _num(vals[B_TRANS]),
            _num(vals[B_NONTAX]), _num(vals[B_ETC]), _num(vals[B_TW]),
            agg.date_by_id.get(emp_id), _num(detail), _num(diff), verdict,
        ]
        ws.append(row)
        r = ws.max_row
        for c in ws[r]:
            c.font = body_font
        ws.cell(row=r, column=12).number_format = "yyyy/m/d"
        for ci in (3, 4, 5, 6, 7, 8, 9, 10, 11, 13, 14):
            ws.cell(row=r, column=ci).alignment = right
        vc = ws.cell(row=r, column=15)
        if verdict == "差分あり":
            vc.fill = fill_ng
            vc.font = Font(name=FONT, color="FFFFFF", bold=True)
        elif verdict == "OK":
            vc.fill = fill_ok
        else:
            vc.fill = fill_na

    widths = [10, 16, 10, 12, 10, 14, 11, 10, 14, 18, 12, 11, 16, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if ids:
        ws.auto_filter.ref = f"A1:O{len(ids) + 1}"

    return {
        "summary_employees": len(ids),
        "excluded_out_of_scope": len(agg.excluded_ids),
        "unmatched_rows": agg.unmatched_rows,
        "unmatched_amount": int(agg.unmatched_amount),
        "diff_ng": sum(
            1 for emp_id in ids
            if detail_totals.get(emp_id, 0.0) != 0
            and (sum(agg.by_id.get(emp_id, [0.0] * 7)) - detail_totals.get(emp_id, 0.0)) != 0
        ),
    }


LOG_HEADERS = ["行番号", "社員番号", "氏名", "内訳", "金額", "判定結果", "マッチしたキーワード"]


def add_log_sheet(wb: Workbook, entries: list) -> None:
    """「集計ログ」シートを追加する（マクロ互換7列）。"""
    header_font = Font(name=FONT, bold=True)
    ws = wb.create_sheet("集計ログ")
    ws.append(LOG_HEADERS)
    for c in ws[1]:
        c.font = header_font
    for e in entries:
        ws.append([e.row_no, e.emp_no, e.emp_nm, e.desc, _num(e.amount), e.result, e.matched_kw])
    for col, w in zip("ABCDEFG", [8, 10, 14, 26, 10, 30, 24]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    if entries:
        ws.auto_filter.ref = f"A1:G{len(entries) + 1}"


# ----------------------------------------------------------------------
# オーケストレータ（keihi_summary.run_keihi_integration から呼ばれる）
# ----------------------------------------------------------------------

def classify_and_summarize(
    integrated_rows: list[list[str]],
    wb: Workbook,
    keywords: "dict | None" = None,
    roster_names: "dict | None" = None,
) -> dict:
    """統合一覧表の行を分類・集計し、wb に「集計」「集計ログ」シートを追加して統計を返す。"""
    cls = classify_rows(integrated_rows, keywords)
    agg = aggregate_by_id(cls)
    detail = build_detail_totals(integrated_rows)
    stats = add_summary_sheet(wb, agg, cls.emp_names, detail, roster_names)
    add_log_sheet(wb, cls.log)
    stats["classified_hits"] = cls.hits
    stats["log_rows"] = len(cls.log)
    stats["_agg"] = agg           # 後段（インポート行生成）用
    stats["_emp_names"] = cls.emp_names
    return stats
