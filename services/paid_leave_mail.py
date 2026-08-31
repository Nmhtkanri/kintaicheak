# -*- coding: utf-8 -*-
"""有休ブックを、メール下書きモードの「一覧表」に変える読み手。

入力は抽出exe（``jinjer有休4日以下抽出.exe``）が作り、谷津さんがグレー塗りで
対象外を落としたブック。3シート（対象者一覧／有休明細／集計条件）を読み、
差し込みに使う5列を組み立てて返す。

    社員番号 / 氏名 / 取得日数 / 不足日数 / 取得期限

この5つは共有テンプレート「有休取得のお願い」の ``{{列名}}`` にそのまま対応する。
差し込み・宛先突合・要確認の判定・下書き作成は mail_draft 側の共通処理を通る。

移植元は ``Z:\\API連携\\create_paid_leave_outlook_mail.py``（有休案内専用ツール）の
``read_paid_leave_report`` / ``is_gray_excluded_row`` / ``deadline_for``。
**同じブックから同じ人数・同じ日数が出ること**を回帰テストで担保している。

2026-08-31 時点の運用: 抽出（jinjer API）は exe のまま。ハブは編集済みブックを読む。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.mail_draft import normalize_employee_id

TARGET_SHEET = "対象者一覧"
DETAIL_SHEET = "有休明細"
CONDITION_SHEET = "集計条件"

# 年5日の取得義務。集計条件シートに値があればそちらを優先する。
DEFAULT_TARGET_DAYS = 5.0
# 取得期限の既定＝集計基準日の翌年3月31日（年度末）。
DEFAULT_DEADLINE_MONTH = 3
DEFAULT_DEADLINE_DAY = 31
# グレー塗りの判定に見る列の範囲（A〜H）。移植元と同じ。
GRAY_FIRST_COL = 1
GRAY_LAST_COL = 8


@dataclass
class PaidLeavePerson:
    employee_id: str
    name: str
    taken_days: float
    remaining_days: float
    source_row: int


@dataclass
class PaidLeaveReport:
    source_path: Path
    start_date: date
    as_of_date: date
    target_days: float
    included: list[PaidLeavePerson] = field(default_factory=list)
    gray_excluded: list[PaidLeavePerson] = field(default_factory=list)


def as_date(value: Any, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError as exc:
        raise ValueError(f"{label}を日付として読み取れません: {value!r}") from exc


def japanese_date(value: date) -> str:
    return f"{value.year}年{value.month}月{value.day}日"


def deadline_for(as_of_date: date) -> date:
    return date(as_of_date.year + 1, DEFAULT_DEADLINE_MONTH, DEFAULT_DEADLINE_DAY)


def _fill_signature(cell: Any) -> tuple[Any, ...]:
    fill = cell.fill
    if fill.fill_type is None:
        return ("none",)
    color = fill.fgColor
    return (
        fill.fill_type,
        color.type,
        color.rgb,
        color.indexed,
        color.theme,
        round(float(color.tint or 0), 6),
    )


def is_gray_excluded_row(sheet: Any, row_number: int,
                         first_col: int = GRAY_FIRST_COL,
                         last_col: int = GRAY_LAST_COL) -> bool:
    """利用者が行全体を塗った「対象外」の行かどうか。

    A〜H列がすべて同じ塗りなら対象外とみなす（塗りが無い行は対象）。
    """
    signatures = [_fill_signature(sheet.cell(row=row_number, column=column))
                  for column in range(first_col, last_col + 1)]
    return signatures[0] != ("none",) and len(set(signatures)) == 1


def _find_header_row(sheet: Any, first_header: str, max_rows: int = 30) -> int:
    for row_number in range(1, min(sheet.max_row, max_rows) + 1):
        if str(sheet.cell(row_number, 1).value or "").strip() == first_header:
            return row_number
    raise ValueError(f"{sheet.title}シートに「{first_header}」ヘッダーがありません")


def read_paid_leave_report(path: str | Path) -> PaidLeaveReport:
    """有休ブックを読む。取得日数は明細の換算日数を社員ごとに合計する。"""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"有休ブックが見つかりません: {path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"有休ブックは .xlsx / .xlsm を指定してください: {path.name}")

    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        missing = {TARGET_SHEET, DETAIL_SHEET, CONDITION_SHEET} - set(workbook.sheetnames)
        if missing:
            raise ValueError(
                "有休ブックに必要なシートがありません: "
                f"{'、'.join(sorted(missing))}（抽出exeが作ったブックを指定してください）")

        target_sheet = workbook[TARGET_SHEET]
        detail_sheet = workbook[DETAIL_SHEET]
        condition_sheet = workbook[CONDITION_SHEET]

        conditions: dict[str, Any] = {}
        for row_number in range(1, condition_sheet.max_row + 1):
            label = str(condition_sheet.cell(row_number, 1).value or "").strip()
            if label:
                conditions[label] = condition_sheet.cell(row_number, 2).value
        start_date = as_date(conditions.get("対象開始日"), "対象開始日")
        as_of_date = as_date(conditions.get("対象終了日"), "対象終了日")
        target_days = float(conditions.get("法定5日到達基準") or DEFAULT_TARGET_DAYS)

        detail_header = _find_header_row(detail_sheet, "社員番号")
        detail_columns = {
            str(detail_sheet.cell(detail_header, column).value or "").strip(): column
            for column in range(1, detail_sheet.max_column + 1)}
        id_column = detail_columns.get("社員番号")
        weight_column = detail_columns.get("換算日数")
        if not id_column or not weight_column:
            raise ValueError("有休明細シートに社員番号または換算日数の列がありません")
        taken: dict[str, float] = {}
        for row_number in range(detail_header + 1, detail_sheet.max_row + 1):
            employee_id = normalize_employee_id(detail_sheet.cell(row_number, id_column).value)
            if not employee_id:
                continue
            taken[employee_id] = taken.get(employee_id, 0.0) + float(
                detail_sheet.cell(row_number, weight_column).value or 0)

        target_header = _find_header_row(target_sheet, "社員番号")
        target_columns = {
            str(target_sheet.cell(target_header, column).value or "").strip(): column
            for column in range(1, target_sheet.max_column + 1)}
        target_id_column = target_columns.get("社員番号")
        target_name_column = target_columns.get("氏名")
        if not target_id_column or not target_name_column:
            raise ValueError("対象者一覧シートに社員番号または氏名の列がありません")

        included: list[PaidLeavePerson] = []
        gray_excluded: list[PaidLeavePerson] = []
        for row_number in range(target_header + 1, target_sheet.max_row + 1):
            employee_id = normalize_employee_id(
                target_sheet.cell(row_number, target_id_column).value)
            if not employee_id:
                continue
            taken_days = round(taken.get(employee_id, 0.0), 1)
            person = PaidLeavePerson(
                employee_id=employee_id,
                name=str(target_sheet.cell(row_number, target_name_column).value or "").strip(),
                taken_days=taken_days,
                remaining_days=round(max(0.0, target_days - taken_days), 1),
                source_row=row_number,
            )
            if is_gray_excluded_row(target_sheet, row_number):
                gray_excluded.append(person)
            else:
                included.append(person)

        if not included and not gray_excluded:
            raise ValueError("対象者一覧シートにデータ行がありません")
        return PaidLeaveReport(
            source_path=path,
            start_date=start_date,
            as_of_date=as_of_date,
            target_days=target_days,
            included=included,
            gray_excluded=gray_excluded,
        )
    finally:
        workbook.close()


def load_paid_leave_table(
    path: str | Path,
    *,
    deadline: str | date | None = None,
) -> tuple[list[str], list[dict[str, Any]], str, str, dict[str, Any]]:
    """有休ブックから (ヘッダー, 行, 社員番号列, 氏名列, 補足) を返す。

    グレー塗りの行はここで落とす。数値は移植元と同じ体裁（小数1桁・和暦なしの
    「2027年3月31日」）にそろえ、既存テンプレートの文面がそのまま通るようにする。
    """
    report = read_paid_leave_report(path)
    if deadline in (None, ""):
        limit = deadline_for(report.as_of_date)
    else:
        limit = as_date(deadline, "取得期限")
    if limit <= report.as_of_date:
        raise ValueError(
            f"取得期限は集計基準日（{japanese_date(report.as_of_date)}）より"
            "あとの日付にしてください")

    headers = ["社員番号", "氏名", "取得日数", "不足日数", "取得期限"]
    rows = [{
        "社員番号": person.employee_id,
        "氏名": person.name,
        "取得日数": f"{person.taken_days:.1f}",
        "不足日数": f"{person.remaining_days:.1f}",
        "取得期限": japanese_date(limit),
    } for person in report.included]
    if not rows:
        raise ValueError(
            "対象者がいません（対象者一覧の行がすべてグレー塗りで対象外になっています）")

    meta = {
        "start_date": report.start_date.isoformat(),
        "as_of_date": report.as_of_date.isoformat(),
        "deadline": limit.isoformat(),
        "target_days": report.target_days,
        "gray_excluded": [{"社員番号": p.employee_id, "氏名": p.name,
                           "取得日数": f"{p.taken_days:.1f}"}
                          for p in report.gray_excluded],
        "note": (
            f"{japanese_date(report.start_date)}〜{japanese_date(report.as_of_date)}の集計。"
            f"対象 {len(rows)}人／グレー塗りで対象外 {len(report.gray_excluded)}人。"
            f"取得期限は {japanese_date(limit)}。"),
    }
    return headers, rows, "社員番号", "氏名", meta


def default_deadline_text(path: str | Path) -> str:
    """画面の初期値に出す取得期限（ブックの集計基準日から決める）。"""
    report = read_paid_leave_report(path)
    return deadline_for(report.as_of_date).isoformat()
