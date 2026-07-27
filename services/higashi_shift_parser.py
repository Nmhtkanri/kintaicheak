"""東さん形式の NMHT 要員シフト希望表を確定的に解析する。

この形式は、画面上の対象月31列の前後に、非表示の補助日付列を持つ。
補助列には同じ曜日配置の過去年月や翌月のシフト値が残ることがあるため、
生の日付や全列の位置ではなく、タイトル/B1 の対象年月と可視31列だけを使う。

ほかの月次表へ誤適用しないよう、タイトル・固定ヘッダー・勤務説明・集計
ヘッダー・前後の非表示日付列・曜日配置がすべて一致した場合だけ認識する。
"""

from __future__ import annotations

import calendar
import os
import re
from datetime import date, datetime, timedelta

import openpyxl


_EXCEL_EPOCH = date(1899, 12, 30)
_FULLWIDTH_DIGIT_TRANS = str.maketrans("０１２３４５６７８９", "0123456789")

_TITLE_RE = re.compile(
    r"(\d{4})\s*年\s*(\d{1,2})\s*月.*エヌエム・ヒューマテック要員シフト希望表"
)
_TIME_RE = re.compile(
    r"([0-9０-９]{1,2})[:：]([0-9０-９]{2})\s*[～~-]\s*"
    r"(?:翌)?\s*([0-9０-９]{1,2})[:：]([0-9０-９]{2})"
)
_HEADER_LABELS = ("拠点", "名前", "役割")
_SUMMARY_CODES = ("日", "前", "後", "自", "○")
_OFF_LABELS = {
    "○": "週休",
    "◎": "代休",
    "●": "希望休日",
    "★": "有給休暇",
    "▲": "前半休",
    "▼": "後半休",
    "♪": "夏休み",
    "祝": "祝日扱い",
}
_FULL_DAY_ALIASES = {
    "自": "自社勤務日",
    "出": "出張",
    "研": "研修",
    "支": "支店勤務",
    "東": "東京勤務",
    "応": "応援勤務",
    "外": "外勤",
}


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _compact(value) -> str:
    return re.sub(r"[\s　]+", "", _clean(value))


def _normalize_digits(value) -> str:
    return _clean(value).translate(_FULLWIDTH_DIGIT_TRANS)


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            serial = int(value)
        except (TypeError, ValueError):
            return None
        if not 1 <= serial <= 60_000:
            return None
        try:
            return _EXCEL_EPOCH + timedelta(days=serial)
        except (OverflowError, ValueError):
            return None
    return None


def _format_time(hour: str, minute: str) -> str:
    return f"{int(_normalize_digits(hour)):02d}:{int(_normalize_digits(minute)):02d}"


def _minutes(hhmm: str) -> int:
    hour, minute = map(int, hhmm.split(":"))
    return hour * 60 + minute


def _default_break_minutes(start: str, end: str) -> int:
    start_min = _minutes(start)
    end_min = _minutes(end)
    if end_min <= start_min:
        end_min += 24 * 60
    return 60 if end_min - start_min > 6 * 60 else 0


def _is_hidden_column(ws, col_1based: int) -> bool:
    """範囲指定された ColumnDimension を含めて非表示列を判定する。"""
    for dimension in ws.column_dimensions.values():
        if not dimension.hidden:
            continue
        start = int(dimension.min or 0)
        end = int(dimension.max or start)
        if start <= col_1based <= end:
            return True
    return False


def _find_title_month(ws) -> tuple[int, int] | None:
    title_month: tuple[int, int] | None = None
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 10),
        values_only=True,
    ):
        for value in row:
            if not isinstance(value, str):
                continue
            match = _TITLE_RE.search(_compact(value))
            if match:
                title_month = (int(match.group(1)), int(match.group(2)))
                break
        if title_month:
            break
    if title_month is None:
        return None

    anchor_date = _to_date(ws.cell(1, 2).value)  # B1: 対象月初
    if (
        anchor_date is None
        or anchor_date.day != 1
        or (anchor_date.year, anchor_date.month) != title_month
    ):
        return None
    return title_month


def _find_header_row(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        labels = tuple(
            _compact(ws.cell(row_idx, col).value)
            for col in range(2, 5)
        )
        note_label = _compact(ws.cell(row_idx, 5).value)
        if labels == _HEADER_LABELS and note_label.startswith("備考"):
            return row_idx
    return None


def _has_identifying_labels(ws, header_row: int) -> bool:
    top_texts = [
        _compact(cell.value)
        for row in ws.iter_rows(min_row=1, max_row=min(header_row, 12))
        for cell in row
        if cell.value is not None
    ]
    required_starts = (
        "勤務時間帯",
        "日：09:00-16:30",
        "前：09:00-12:00",
        "後：12:00-16:30",
    )
    if not all(
        any(text.startswith(prefix) for text in top_texts)
        for prefix in required_starts
    ):
        return False

    row_values = [
        _compact(ws.cell(header_row, col).value)
        for col in range(1, ws.max_column + 1)
    ]
    width = len(_SUMMARY_CODES)
    return any(
        tuple(row_values[idx:idx + width]) == _SUMMARY_CODES
        for idx in range(0, len(row_values) - width + 1)
    )


def _detect_layout(ws) -> dict | None:
    target_month = _find_title_month(ws)
    header_row = _find_header_row(ws)
    if target_month is None or header_row is None:
        return None
    if not _has_identifying_labels(ws, header_row):
        return None

    year, month = target_month
    days_in_month = calendar.monthrange(year, month)[1]
    if days_in_month != 31:
        # 現在確認済みの東さん形式は可視31列の月だけ。
        return None

    date_columns: dict[int, date] = {}
    for col in range(1, ws.max_column + 1):
        parsed = _to_date(ws.cell(header_row, col).value)
        if parsed is not None:
            date_columns[col] = parsed

    visible_columns = [
        col for col in date_columns
        if not _is_hidden_column(ws, col)
    ]
    if len(visible_columns) != days_in_month:
        return None
    first_col = visible_columns[0]
    if visible_columns != list(range(first_col, first_col + days_in_month)):
        return None

    # 固有レイアウト: 対象月の前後に非表示の補助日付列がある。
    hidden_before = any(
        col < first_col and _is_hidden_column(ws, col)
        for col in date_columns
    )
    hidden_after = any(
        col > visible_columns[-1] and _is_hidden_column(ws, col)
        for col in date_columns
    )
    if not hidden_before or not hidden_after:
        return None

    # 生の日付は代用年月でも、対象月と曜日配置は完全一致する必要がある。
    weekday_matched = sum(
        1
        for day, col in enumerate(visible_columns, start=1)
        if date_columns[col].weekday() == date(year, month, day).weekday()
    )
    if weekday_matched != days_in_month:
        return None

    return {
        "year": year,
        "month": month,
        "header_row": header_row,
        "visible_columns": visible_columns,
        "weekday_matched": weekday_matched,
    }


def _parse_work_legend(ws, seen_codes: set[str]) -> list[dict]:
    legend: list[dict] = []
    by_code: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12)):
        for cell in row:
            text = _normalize_digits(cell.value)
            if not text:
                continue
            code_match = re.match(r"^(日|前|後)\s*[：:]", text)
            time_match = _TIME_RE.search(text)
            if not code_match or not time_match:
                continue
            code = code_match.group(1)
            start = _format_time(time_match.group(1), time_match.group(2))
            end = _format_time(time_match.group(3), time_match.group(4))
            by_code[code] = {
                "code": code,
                "label": f"{code}勤務",
                "start_time": start,
                "end_time": end,
                "break_minutes": _default_break_minutes(start, end),
                "is_off": False,
            }

    for code in ("日", "前", "後"):
        if code in by_code:
            legend.append(by_code[code])

    day_entry = by_code.get("日")
    if day_entry:
        for code, label in _FULL_DAY_ALIASES.items():
            if code not in seen_codes:
                continue
            legend.append({
                **day_entry,
                "code": code,
                "label": label,
            })

    for code, label in _OFF_LABELS.items():
        if code not in seen_codes:
            continue
        legend.append({
            "code": code,
            "label": label,
            "start_time": "",
            "end_time": "",
            "break_minutes": 0,
            "is_off": True,
        })
    return legend


def parse_higashi_shift_xlsx(
    filepath: str,
    target_year: int | None = None,
    target_month: int | None = None,
) -> dict:
    """可視31列だけを対象年月の1〜31日として解析する。"""
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    if len(workbook.worksheets) != 1:
        raise ValueError("東さん形式は単一シートのみ対応です")
    worksheet = workbook.active
    layout = _detect_layout(worksheet)
    if layout is None:
        raise ValueError("東さん形式のシフト希望表ではありません")

    year = layout["year"]
    month = layout["month"]
    if target_year and target_month and (target_year, target_month) != (year, month):
        raise ValueError(
            f"シフト表の年月 {year}年{month}月 が対象 "
            f"{target_year}年{target_month}月 と一致しません"
        )

    visible_columns = layout["visible_columns"]
    employees: list[dict] = []
    seen_codes: set[str] = set()
    for row_idx in range(layout["header_row"] + 1, worksheet.max_row + 1):
        name = _clean(worksheet.cell(row_idx, 3).value)  # C: 名前
        if not name or name == "名前":
            continue
        codes = [
            _normalize_digits(worksheet.cell(row_idx, col).value)
            for col in visible_columns
        ]
        if not any(codes):
            continue
        seen_codes.update(code for code in codes if code)
        employees.append({
            "name": name,
            "shifts": [
                {
                    "date": date(year, month, day).isoformat(),
                    "code": code,
                }
                for day, code in enumerate(codes, start=1)
            ],
        })

    if not employees:
        raise ValueError("東さん形式の従業員シフト行が見つかりません")

    return {
        "year": year,
        "month": month,
        "filename": os.path.basename(filepath),
        "legend": _parse_work_legend(worksheet, seen_codes),
        "employees": employees,
        "off_markers": [""] + list(_OFF_LABELS),
        "section_info": {
            "section_index": None,
            "start_col": visible_columns[0],
            "end_col": visible_columns[-1],
            "weekday_match_ratio": 1.0,
            "weekday_matched": layout["weekday_matched"],
            "weekday_total": len(visible_columns),
            "total_sections": 1,
            "visible_columns_only": True,
        },
    }


def is_higashi_shift_xlsx(filepath: str) -> bool:
    """東さん形式だけを厳密に判定する。"""
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        if len(workbook.worksheets) != 1:
            return False
        return _detect_layout(workbook.active) is not None
    except Exception:
        return False
