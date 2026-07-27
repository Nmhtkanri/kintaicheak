r"""経理モード C-4: 給与の「その他」(allowance52) を freee の費用品目へ分解する。

Z:\API連携\scripts\keiri_keihi_tenki.py から移植（2026-07-27）。本体はこちらが正。
レビュー資料（draftD.md/.html）の生成は Z:\API連携 側の開発ツールが本モジュールを呼ぶ。

jinjer の salary_items:allowance52「その他」は経費一覧表マクロが出した集計値で、
freee では 消耗品費／社員育成費／会議接待費／通信費／福利厚生費／雑費 に分かれて計上される。
分解の材料は同じマクロブックの中にある:

  経費利用履歴 RevN.xlsm
    ├ 集計ログ         … 1明細=1行。判定結果が「J:その他」の行が allowance52 の中身
    ├ 経費統合一覧表   … 集計ログの「行番号」で引ける。備考(明細)を取るために使う
    └ 集計             … 「その他(会議費・消耗品など)」列が allowance52 と一致する（検算用）

品目の割り当てはマッピング表（既定 Z:\API連携\docs\経理モード_経費転記マッピング_draftD.csv）
駆動。**備考(明細)の部分一致を先に見て、当たらなければ内訳の完全一致、それも無ければ既定**の順。
備考のほうが具体的な情報を持つため（例: 内訳「郵便料金」でも備考に『書類』があれば雑費）。

安全弁: 分解した合計が allowance52 と1円でも違う人は**行を作らず要確認へ回す**。
  実績で不一致になるのは以下のような正当な理由があるとき（自動では判断できない）:
    - 経理が jinjer 投入前にマクロ出力から行を抜いた（2026-06 築城2018047 の引っ越し代仮払い）
    - 前月分をまとめて計上した（2026-06 田中2018037 の CCNA 2回分 93,720 = 46,860×2）
    - マクロ側に重複行がある（2026-05 塚本2024032 の手数料等 23,350×2）
    - そもそも経費申請ではなく手入力（2026-07 稲田2026009・柳場2026010 の CCNA 46,860）
"""

from __future__ import annotations

import csv
import glob
import os
import re
import unicodedata
from collections import defaultdict

from config import Config

MAPPING_CSV = Config.KEIRI_KEIHI_MAPPING_CSV
BOOK_BASE = Config.KEIRI_KEIHI_BOOK_DIR
SONOTA_JUDGE = "J:その他"          # 集計ログの判定結果（マクロの CAT_ETC）

# 集計ログの列（0-based。2026-07 Rev5 実測）
L_ROWNO, L_EMP, L_NAME, L_UCHIWAKE, L_AMOUNT, L_JUDGE, L_KEYWORD = 0, 1, 2, 3, 4, 5, 6
I_MEMO_LINE = 19                    # 経費統合一覧表の 備考(明細)


def normalize(s):
    """キーワード照合用の正規化（全半角・大小・空白を吸収）。"""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(s or ""))).casefold()


def load_mapping(path=MAPPING_CSV):
    """マッピング表を (備考ルール, 内訳ルール, 既定) に分けて返す。order 昇順＝先勝ち。"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = sorted(csv.DictReader(f), key=lambda r: int(r["order"]))
    biko = [r for r in rows if r["match_field"] == "備考"]
    uchiwake = [r for r in rows if r["match_field"] == "内訳"]
    default = next((r for r in rows if r["match_field"] == "既定"), None)
    if default is None:
        raise SystemExit(f"マッピング表に既定行がありません: {path}")
    return biko, uchiwake, default


def classify(uchiwake, memo, mapping):
    """1明細 → (マッピング行, 判定根拠の文字列)。"""
    biko_rules, uchi_rules, default = mapping
    nmemo = normalize(memo)
    for r in biko_rules:
        if nmemo and normalize(r["keyword"]) in nmemo:
            return r, f"備考に『{r['keyword']}』"
    nuchi = normalize(uchiwake)
    for r in uchi_rules:
        if nuchi and nuchi == normalize(r["keyword"]):
            return r, f"内訳『{r['keyword']}』"
    return default, "既定（キーワード未一致）"


def find_book(month, base=BOOK_BASE):
    """支給月フォルダの 経費利用履歴 RevN.xlsm を探す（バックアップ・一時ファイルは除く）。"""
    pattern = os.path.join(base, f"{int(month[5:7])}月", "経費利用履歴*.xlsm")
    cands = [p for p in glob.glob(pattern)
             if not os.path.basename(p).startswith("~$") and "backup" not in os.path.basename(p).lower()]
    return max(cands, key=os.path.getmtime) if cands else None


def load_details(book_path):
    """{社員番号: [(内訳, 金額, 備考), ...]} を返す（集計ログの J:その他 行のみ）。"""
    from openpyxl import load_workbook
    wb = load_workbook(book_path, read_only=True, data_only=True, keep_links=False)
    if "集計ログ" not in wb.sheetnames:
        raise SystemExit(f"集計ログシートがありません: {book_path}")
    log = list(wb["集計ログ"].iter_rows(values_only=True))
    integrated = list(wb["経費統合一覧表"].iter_rows(values_only=True))

    def memo_of(rowno):
        """集計ログの行番号は経費統合一覧表の 1-based 行番号（1行目=ヘッダー）。"""
        try:
            idx = int(rowno) - 1
        except (TypeError, ValueError):
            return ""
        if 0 < idx < len(integrated) and len(integrated[idx]) > I_MEMO_LINE:
            return str(integrated[idx][I_MEMO_LINE] or "").strip()
        return ""

    out = defaultdict(list)
    for r in log[1:]:
        if len(r) <= L_JUDGE or str(r[L_JUDGE] or "").strip() != SONOTA_JUDGE:
            continue
        amount = r[L_AMOUNT]
        if not isinstance(amount, (int, float)) or not amount:
            continue
        emp = str(r[L_EMP] or "").strip()
        out[emp].append((str(r[L_UCHIWAKE] or "").strip(), float(amount), memo_of(r[L_ROWNO])))
    return out


def decompose(details, mapping):
    """1人分の明細 → ({(品目, 勘定科目, 税区分): 金額}, [判定根拠], 合計)。

    備考は品目ごとに代表1件＋件数にまとめる（最終CSVの備考は経理が書き直す前提）。
    """
    by_item = defaultdict(float)
    memos = defaultdict(list)
    reasons = []
    for uchiwake, amount, memo in details:
        rule, why = classify(uchiwake, memo, mapping)
        key = (rule["freee_item"], rule["freee_account"], rule["freee_tax"])
        by_item[key] += amount
        if memo:
            memos[key].append(memo)
        reasons.append({"内訳": uchiwake, "金額": amount, "備考": memo,
                        "品目": rule["freee_item"], "根拠": why, "status": rule["status"]})
    biko = {k: (v[0] if len(v) == 1 else f"{v[0]} ほか{len(v) - 1}件") for k, v in memos.items()}
    return by_item, biko, reasons, sum(a for _u, a, _m in details)
