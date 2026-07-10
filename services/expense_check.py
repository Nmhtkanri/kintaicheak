"""経費チェックモード — テレワーク日数・出社日数の月次集計

データソース: GET /v1/employees/attendances?employee-id=...&month=YYYY-MM
- テレワーク日 = その日の stamp_classifications に「テレワーク」を含む区分が選択されている日
- 出勤日   = is_absent=false かつ出勤打刻がある日
- 出社日数 = 出勤日数 - テレワーク日数（Excel 数式）

社員番号・氏名ごとに Excel（サマリ＋テレワーク明細）へ出力する。
既存 Z:\\API連携\\export_telework_days.py のロジックを kintai-checker 用に移植。
"""

from __future__ import annotations

import time as _time
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from services.jinjer_api_client import JinjerClient, JinjerAPIError, _safe_get

TELEWORK_KEYWORD = "テレワーク"
PACING_SEC = 0.3
MAX_RETRY = 3
FONT = "Meiryo UI"


# ----------------------------------------------------------------------
# jinjer 取得
# ----------------------------------------------------------------------

def fetch_active_employees(client: JinjerClient) -> list[dict]:
    """在籍中の従業員を [{id, name}] で返す（社員番号順）"""
    employees = client.get_employees(only_active=True)
    result = []
    for emp in employees:
        if not isinstance(emp, dict):
            continue
        emp_id = emp.get("id") or emp.get("employee_id")
        if emp_id is None:
            continue
        last = str(_safe_get(emp, "company", "last_name") or "").strip()
        first = str(_safe_get(emp, "company", "first_name") or "").strip()
        name = f"{last} {first}".strip()
        result.append({"id": str(emp_id).strip(), "name": name})
    result.sort(key=lambda e: e["id"])
    return result


def fetch_attendances(client: JinjerClient, employee_id: str, month: str) -> list[dict]:
    """1名分の日次勤怠を取得。429 はリトライ、該当なし(404)は空リスト。"""
    url = f"{client.base_url}/v1/employees/attendances"
    headers = client._auth_headers()
    params = {"employee-id": employee_id, "month": month}
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=60)
        except requests.RequestException as e:
            raise JinjerAPIError(f"勤怠取得に失敗 employee_id={employee_id}: {e}") from e
        if r.status_code == 200:
            return (r.json().get("data") or {}).get("attendances", []) or []
        if r.status_code == 429:
            _time.sleep(5 * attempt)
            continue
        if r.status_code == 404:
            return []
        raise JinjerAPIError(
            f"勤怠取得に失敗 employee_id={employee_id}: {r.status_code} {r.text[:200]}"
        )
    raise JinjerAPIError(f"勤怠取得リトライ上限 employee_id={employee_id}")


def summarize(attendances: list[dict]) -> tuple[int, list[tuple[str, str]]]:
    """(出勤日数, テレワーク日リスト[(日付, 区分名)]) を返す。

    API が同一日のレコードを重複して返すことがあるため、日付単位でユニーク化して集計する。
    """
    worked_dates: set[str] = set()
    telework_by_date: dict[str, set[str]] = {}
    for day in attendances or []:
        if not isinstance(day, dict):
            continue
        if day.get("is_absent"):
            continue
        if not day.get("attended_at"):
            continue
        date_str = str(day.get("date") or "")
        if not date_str:
            continue
        worked_dates.add(date_str)
        names = [str((c or {}).get("name") or "") for c in (day.get("stamp_classifications") or [])]
        tw_names = [n for n in names if TELEWORK_KEYWORD in n]
        if tw_names:
            telework_by_date.setdefault(date_str, set()).update(tw_names)
    telework_days = [(d, "、".join(sorted(ks))) for d, ks in sorted(telework_by_date.items())]
    return len(worked_dates), telework_days


# ----------------------------------------------------------------------
# Excel 出力
# ----------------------------------------------------------------------

def _md(date_str: str) -> str:
    """'2026-05-01' / '2026/5/1' → '5/1'"""
    s = str(date_str).replace("/", "-")
    try:
        _, m, d = s.split("-")
        return f"{int(m)}/{int(d)}"
    except ValueError:
        return date_str


def _build_telework_sheets(wb: Workbook, month: str, rows: list[dict]) -> None:
    """wb に サマリ / テレワーク明細 シートを作る（保存はしない）。"""
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")
    body_font = Font(name=FONT)
    center = Alignment(horizontal="center")

    # ---- 明細シート（先に作って COUNTIF の参照先にする） ----
    ws_d = wb.create_sheet("テレワーク明細")
    ws_d.append(["社員番号", "氏名", "テレワーク日", "打刻区分"])
    for r in rows:
        for date_str, kubun in r["telework_days"]:
            ws_d.append([r["id"], r["name"], date_str, kubun])
    for col, width in zip("ABCD", [12, 18, 14, 24]):
        ws_d.column_dimensions[col].width = width
    for cell in ws_d[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row in ws_d.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
    ws_d.freeze_panes = "A2"
    ws_d.auto_filter.ref = f"A1:D{max(ws_d.max_row, 2)}"

    # ---- サマリシート ----
    ws = wb.active
    ws.title = "サマリ"
    ws.append(["社員番号", "氏名", "出勤日数", "テレワーク日数", "出社日数", "テレワーク実施日"])
    for i, r in enumerate(rows, start=2):
        dates = "、".join(_md(d) for d, _ in r["telework_days"])
        ws.cell(row=i, column=1, value=r["id"])
        ws.cell(row=i, column=2, value=r["name"])
        ws.cell(row=i, column=3, value=r["work_days"])
        ws.cell(row=i, column=4, value=f'=COUNTIF(テレワーク明細!A:A,A{i})')
        ws.cell(row=i, column=5, value=f"=C{i}-D{i}")
        ws.cell(row=i, column=6, value=dates)
    last = len(rows) + 1
    total_row = last + 1
    ws.cell(row=total_row, column=2, value="合計").font = Font(name=FONT, bold=True)
    for col in (3, 4, 5):
        letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{last})")
        c.font = Font(name=FONT, bold=True)

    for col, width in zip("ABCDEF", [12, 18, 10, 14, 10, 60]):
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row in ws.iter_rows(min_row=2, max_row=last):
        for cell in row:
            if cell.font.bold is not True:
                cell.font = body_font
        for col in (3, 4, 5):
            row[col - 1].alignment = center
    ws.freeze_panes = "A2"
    if last >= 1:
        ws.auto_filter.ref = f"A1:F{last}"


def build_telework_workbook(month: str, rows: list[dict], out_path: Path) -> None:
    """テレワークのみの Excel を保存する（CLI / テスト用）。"""
    wb = Workbook()
    _build_telework_sheets(wb, month, rows)
    add_selected_employee_views(wb, employee_count=len(rows))
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ----------------------------------------------------------------------
# 通勤費シート（jinjer 人事データ出力CSVから）
# ----------------------------------------------------------------------

# 出力列（キャノニカル）→ jinjer 通勤費CSVの候補列名。
# jinjer のエクスポートは「○○(通勤1)」形式。複数経路(通勤2..)は当面 通勤1 のみ対象。
COMMUTE_FIELDS: list[tuple[str, list[str]]] = [
    ("社員番号", ["社員番号"]),
    ("出発", ["出発(通勤1)", "出発"]),
    ("到着", ["到着(通勤1)", "到着"]),
    ("経由1", ["経由1(通勤1)", "経由1"]),
    ("経由2", ["経由2(通勤1)", "経由2"]),
    ("利用交通機関", ["利用交通機関(通勤1)", "利用交通機関"]),
    ("通勤経路", ["経路(通勤1)", "通勤経路", "経路"]),
    ("支給金額", ["支給金額(通勤1)", "支給金額"]),
    ("非課税通勤費", ["非課税通勤費(通勤1)", "非課税通勤費"]),
    ("課税通勤費", ["課税通勤費(通勤1)", "課税通勤費"]),
    ("支給開始", ["支給開始(通勤1)", "利用開始日(通勤1)", "支給開始", "利用開始日"]),
    ("片道距離(km)", ["片道距離(km)(通勤1)", "片道距離(km)", "片道距離"]),
]
COMMUTE_AMOUNT_COLS = {"支給金額", "非課税通勤費", "課税通勤費"}
# 出力シートの列順。ユーザー要望（出発・到着・経由1・経由2・通勤経路・支給間隔・支給金額）を含む。
# 支給間隔/支給方法/支給開始/経路No は API（commuting-information）からのみ取得。
COMMUTE_OUTPUT_COLUMNS = [
    "社員番号", "氏名", "経路No", "出発", "到着", "経由1", "経由2", "通勤経路",
    "利用交通機関", "支給間隔", "支給方法", "支給金額", "非課税通勤費", "課税通勤費",
    "支給開始", "片道距離(km)",
]


def _to_amount(v):
    """通勤費の金額文字列を int に。空・非数値はそのまま（空は空文字）。"""
    s = str(v or "").strip().replace(",", "")
    if not s:
        return ""
    try:
        return int(float(s))
    except ValueError:
        return v


def read_commute_csv(path: str | Path) -> list[dict]:
    """jinjer「人事データ出力（通勤費）」CSV を読み、社員ごとの通勤情報 dict を返す。

    CP932 優先で読み込み、列名でマッピングする。氏名は 職場氏名(氏)＋(名) を結合。
    出発・到着・経由1・経由2・通勤経路(経路)・支給金額 等を取り出す。
    """
    path = Path(path)
    rows_raw: list[list[str]] = []
    headers: list[str] | None = None
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            import csv as _csv
            with open(path, encoding=enc, newline="") as f:
                reader = _csv.reader(f)
                headers = next(reader)
                rows_raw = [r for r in reader]
            break
        except UnicodeDecodeError:
            continue
    if headers is None:
        raise ValueError(f"通勤費CSV の文字コードを判別できませんでした: {path}")

    hidx = {str(h).strip(): i for i, h in enumerate(headers)}

    def col(cands: list[str]) -> int | None:
        for c in cands:
            if c in hidx:
                return hidx[c]
        return None

    field_idx = {canon: col(cands) for canon, cands in COMMUTE_FIELDS}
    last_i = hidx.get("職場氏名(氏)")
    first_i = hidx.get("職場氏名(名)")
    name_i = hidx.get("氏名")

    result: list[dict] = []
    for r in rows_raw:
        def get(i):
            return r[i].strip() if (i is not None and i < len(r) and r[i] is not None) else ""
        rec: dict = {}
        for canon, _ in COMMUTE_FIELDS:
            v = get(field_idx.get(canon))
            rec[canon] = _to_amount(v) if canon in COMMUTE_AMOUNT_COLS else v
        if name_i is not None:
            rec["氏名"] = get(name_i)
        else:
            rec["氏名"] = (get(last_i) + " " + get(first_i)).strip()
        # CSV には無い列（API のみ）は空で埋める。1行=1経路なので経路No=1。
        rec.setdefault("経路No", 1)
        rec.setdefault("支給間隔", "")
        rec.setdefault("支給方法", "")
        if rec.get("社員番号"):
            result.append(rec)
    result.sort(key=lambda x: str(x.get("社員番号") or ""))
    return result


def fetch_commute_rows_via_api(client: JinjerClient, id_to_name: dict | None = None) -> list[dict]:
    """jinjer API（commuting-information）から通勤情報を取得し、シート用の行に整形する。

    複数経路は経路ごとに1行（経路No=1,2,…）。氏名は ``id_to_name`` で補完（無ければ空）。
    出発・到着・経由1・経由2・通勤経路・利用交通機関・支給間隔・支給方法・支給金額・
    非課税/課税通勤費・支給開始・片道距離 を取り出す。
    """
    items = client.get_commuting_information()
    id_to_name = id_to_name or {}
    rows: list[dict] = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        eid = str(item.get("employee_id") or "").strip()
        for i, route in enumerate(item.get("commuting") or [], start=1):
            if not isinstance(route, dict):
                continue
            pay = route.get("payment") or {}
            rtype = route.get("type") or {}
            rows.append({
                "社員番号": eid,
                "氏名": id_to_name.get(eid, ""),
                "経路No": i,
                "出発": route.get("departure") or "",
                "到着": route.get("arrival") or "",
                "経由1": route.get("transit_1") or "",
                "経由2": route.get("transit_2") or "",
                "通勤経路": route.get("path") or "",
                "利用交通機関": (rtype.get("name") if isinstance(rtype, dict) else "") or "",
                "支給間隔": ((pay.get("interval") or {}).get("name")) or "",
                "支給方法": ((pay.get("method") or {}).get("name")) or "",
                "支給金額": _to_amount(pay.get("total")),
                "非課税通勤費": _to_amount(pay.get("tax_exemption_amount")),
                "課税通勤費": _to_amount(pay.get("taxable_amount")),
                "支給開始": pay.get("start_date") or "",
                "片道距離(km)": route.get("one_way_distance") or "",
            })
    rows.sort(key=lambda x: (str(x.get("社員番号") or ""), x.get("経路No") or 0))
    return rows


def add_commute_sheet(wb: Workbook, commute_rows: list[dict]) -> None:
    """wb に「通勤費」シートを追加する（社員番号・氏名・出発・到着・経由・通勤経路・支給金額…）。"""
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="548235")
    body_font = Font(name=FONT)
    right = Alignment(horizontal="right")

    ws = wb.create_sheet("通勤費")
    ws.append(COMMUTE_OUTPUT_COLUMNS)
    for rec in commute_rows:
        ws.append([rec.get(c, "") for c in COMMUTE_OUTPUT_COLUMNS])

    n = len(commute_rows)
    # 金額の合計行
    if n:
        total_row = n + 2
        ws.cell(row=total_row, column=1, value="合計").font = Font(name=FONT, bold=True)
        for col_name in ("支給金額", "非課税通勤費", "課税通勤費"):
            ci = COMMUTE_OUTPUT_COLUMNS.index(col_name) + 1
            letter = get_column_letter(ci)
            c = ws.cell(row=total_row, column=ci, value=f"=SUM({letter}2:{letter}{n + 1})")
            c.font = Font(name=FONT, bold=True)
            c.alignment = right

    widths = {
        "社員番号": 12, "氏名": 16, "経路No": 7, "出発": 14, "到着": 14, "経由1": 12,
        "経由2": 12, "通勤経路": 60, "利用交通機関": 14, "支給間隔": 10, "支給方法": 8,
        "支給金額": 12, "非課税通勤費": 12, "課税通勤費": 12, "支給開始": 10, "片道距離(km)": 11,
    }
    for i, name in enumerate(COMMUTE_OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    amount_cols = {COMMUTE_OUTPUT_COLUMNS.index(c) + 1 for c in ("支給金額", "非課税通勤費", "課税通勤費")}
    for row in ws.iter_rows(min_row=2, max_row=max(n + 1, 2)):
        for cell in row:
            if cell.font.bold is not True:
                cell.font = body_font
            if cell.column in amount_cols:
                cell.alignment = right
    ws.freeze_panes = "A2"
    if n:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COMMUTE_OUTPUT_COLUMNS))}{n + 1}"


# ----------------------------------------------------------------------
# 通勤費申請なしリスト（手動Excelをインポートして自動整備）
# ----------------------------------------------------------------------

NO_COMMUTE_SHEET = "通勤費申請なし"
NO_COMMUTE_HEADERS = ["社員番号", "氏名", "通勤費", "区分", "利用交通機関(参考)", "備考"]
# 支給間隔 → 区分のマッピング。毎月＝定期代支給、毎日＝出社日ごとの実費。
# それ以外（3ヶ月・6ヶ月等）は空欄にして目視判断に委ねる。
_INTERVAL_TO_KUBUN = {"毎月": "通勤定期代", "毎日": "通勤費"}


def _norm_emp_id(v) -> str:
    """社員番号を文字列に正規化する（Excelの数値セル 2026013.0 → '2026013'）。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def read_no_commute_list(path: "str | Path") -> list[dict]:
    """手動整備の「通勤費申請なし」Excelブックを読み [{id, name}] を返す。

    シート名「通勤費申請なし」があればそれを、無ければ先頭シートを使う。
    A列=社員番号・B列=氏名。「社員番号」ヘッダー行・空行・重複IDはスキップする。
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[NO_COMMUTE_SHEET] if NO_COMMUTE_SHEET in wb.sheetnames else wb.worksheets[0]
        entries: list[dict] = []
        seen: set[str] = set()
        for row in ws.iter_rows(values_only=True):
            emp = _norm_emp_id(row[0] if row else None)
            if not emp or emp == "社員番号" or emp in seen:
                continue
            seen.add(emp)
            name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            entries.append({"id": emp, "name": name})
        return entries
    finally:
        wb.close()


def build_no_commute_rows(
    entries: list[dict],
    summary_rows: list[dict],
    commute_rows: list[dict],
) -> tuple[list[dict], list[dict]]:
    """「通勤費申請なし」リストを自動整備する（経費チェックモードの改修案 STEP2/3）。

    STEP2: 出勤日数=テレワーク日数（＝出社日数0＝完全在宅）の社員をリストから除外。
           ただし出勤0（テレワーク0＝出社0の両方ゼロ）は実質非該当のため残す。
    STEP3: 通勤費の経路行から支給金額を合算し（複数経路は合算）、
           先頭経路（経路No最小）の支給間隔を 毎月→通勤定期代／毎日→通勤費 に
           マッピングする。通勤費未登録は金額0・区分空欄、金額0登録ありは区分のみ入る。

    Returns:
        (整備済み行, 除外した行)。行キー: id / name / amount / kubun / kikan / remark
    """
    att = {str(r["id"]).strip(): (r["work_days"], len(r["telework_days"])) for r in summary_rows}
    names = {str(r["id"]).strip(): r["name"] for r in summary_rows}
    routes_by_emp: dict[str, list[dict]] = {}
    for c in commute_rows:
        routes_by_emp.setdefault(_norm_emp_id(c.get("社員番号")), []).append(c)

    kept: list[dict] = []
    removed: list[dict] = []
    for e in entries:
        emp = e["id"]
        name = e["name"] or names.get(emp, "")
        work_tw = att.get(emp)
        if work_tw is not None:
            work, tw = work_tw
            if work > 0 and work == tw:
                removed.append({"id": emp, "name": name})
                continue
            remark = "出勤0" if work == 0 else ""
        else:
            remark = "勤怠データなし"

        routes = sorted(routes_by_emp.get(emp, []), key=lambda c: c.get("経路No") or 0)
        amount = 0
        for c in routes:
            v = c.get("支給金額")
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                amount += int(v)
        interval = str(routes[0].get("支給間隔") or "").strip() if routes else ""
        kept.append({
            "id": emp,
            "name": name,
            "amount": amount,
            "kubun": _INTERVAL_TO_KUBUN.get(interval, ""),
            "kikan": str(routes[0].get("利用交通機関") or "") if routes else "",
            "remark": remark,
        })
    return kept, removed


def add_no_commute_sheet(wb: Workbook, rows: list[dict]) -> None:
    """wb に整備済み「通勤費申請なし」シートを追加する。"""
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="C55A11")
    body_font = Font(name=FONT)
    right = Alignment(horizontal="right")

    ws = wb.create_sheet(NO_COMMUTE_SHEET)
    ws.append(NO_COMMUTE_HEADERS)
    for r in rows:
        ws.append([r["id"], r["name"], r["amount"], r["kubun"], r["kikan"], r["remark"]])
    n = len(rows)
    if n:
        total_row = n + 2
        ws.cell(row=total_row, column=2, value="合計").font = Font(name=FONT, bold=True)
        c = ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{n + 1})")
        c.font = Font(name=FONT, bold=True)
        c.alignment = right

    for col, width in zip("ABCDEF", [12, 18, 12, 14, 16, 16]):
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2, max_row=max(n + 1, 2)):
        for cell in row:
            if cell.font.bold is not True:
                cell.font = body_font
            if cell.column == 3:
                cell.alignment = right
    ws.freeze_panes = "A2"
    if n:
        ws.auto_filter.ref = f"A1:F{n + 1}"


# ----------------------------------------------------------------------
# 従業員フィルタ連動ビュー（Excel 365 の FILTER で全シート連動）
# ----------------------------------------------------------------------

# 選択中の従業員（社員番号）を保持する定義名。全シートがこれを参照する。
EMP_PICK_NAME = "選択社員"

# openpyxl は新関数名をそのまま書き出すため、Excel が認識する内部名で保存する。
# XLOOKUP は単一値なので "_xlfn." 付きでそのまま動く（サマリのミニ集計で使用）。
# ※ 連動ビューの明細は FILTER（動的配列）だと openpyxl がメタデータを書けず
#   スピルしないため、INDEX+AGGREGATE 方式（_fill_linked_rows）を採用している。
_XLOOKUP = "_xlfn.XLOOKUP"


def _detail_last_row(wb: Workbook) -> int:
    """テレワーク明細シートの最終データ行（合計行なし）。ヘッダーのみなら 1。"""
    ws = wb["テレワーク明細"]
    return max(ws.max_row, 1)


def _max_rows_per_id(ws, last_data_row: int, id_col: int = 1, first_row: int = 2) -> int:
    """元シートで、同一社員番号が最大何行あるか（連動ビューの行数決定用）。"""
    from collections import Counter
    counter: Counter = Counter()
    for r in range(first_row, last_data_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is not None and str(v).strip():
            counter[str(v).strip()] += 1
    return max(counter.values()) if counter else 0


def _add_source_rank_helper(ws, helper_col: int, last_data_row: int) -> None:
    """元シートに「選択社員の何番目の行か（1,2,3…）」を返す順位ヘルパー列を足す。

    選択社員に一致する行にだけ 1,2,3… の連番を振り、それ以外は ""。
    配列数式を使わない純スカラー式（COUNTIF/IF）なので、Excel の実装差や
    暗黙的交差の影響を受けない。連動ビューはこの列を MATCH で引く。
    """
    letter = get_column_letter(helper_col)
    for r in range(2, last_data_row + 1):
        ws.cell(row=r, column=helper_col).value = (
            f'=IF($A{r}={EMP_PICK_NAME},COUNTIF($A$2:$A{r},{EMP_PICK_NAME}),"")'
        )
    ws.column_dimensions[letter].hidden = True


def _fill_linked_rows(
    vs, source_sheet: str, ncols: int, src_helper_letter: str,
    last_data_row: int, n_rows: int,
) -> None:
    """選択社員の該当行だけを MATCH+INDEX で引いて各セルに書き込む。

    元シートの順位ヘルパー列（src_helper_letter）に 1,2,3… が振られているので、
    ビューの i 行目は「順位 i の行」を MATCH で探して INDEX で各列を取得する。
    使う関数は INDEX/MATCH/IFERROR/ROW のみ＝配列評価もダイナミック配列も不要で、
    どの Excel でも確実にスピルなしで表示できる。範囲は実データ行までに限定する
    （全列参照 A:A は Excel でも計算エンジンでも重いため）。
    """
    if n_rows <= 0 or last_data_row < 2:
        vs["A4"] = "（該当データなし）"
        vs["A4"].font = Font(name=FONT, italic=True, color="808080")
        return

    hcol = src_helper_letter
    match_range = f"{source_sheet}!${hcol}$2:${hcol}${last_data_row}"
    for i in range(1, n_rows + 1):
        r = 3 + i
        for c in range(1, ncols + 1):
            cl = get_column_letter(c)
            data_range = f"{source_sheet}!{cl}$2:{cl}${last_data_row}"
            vs.cell(row=r, column=c).value = (
                f'=IFERROR(INDEX({data_range},MATCH(ROW()-3,{match_range},0)),"")'
            )


def add_selected_employee_views(
    wb: Workbook,
    employee_count: int,
    commute_row_count: int = 0,
) -> None:
    """サマリに「従業員選択」ドロップダウンを付け、選択した従業員だけを表示する
    連動シート（テレワーク明細(選択者)/通勤費(選択者)）を追加する。

    1つのドロップダウン（定義名 ``選択社員`` = サマリ!$I$2）を全シートが FILTER で
    参照するため、選択を変えると連動シートが同じ従業員に一斉に切り替わる。
    Excel 365 / 2021 の FILTER・XLOOKUP を使う（マクロ不要）。

    Args:
        wb: サマリ／テレワーク明細（＋任意で通勤費）が既にあるブック
        employee_count: サマリのデータ行数（従業員数）
        commute_row_count: 通勤費のデータ行数（0 なら通勤費連動シートは作らない）
    """
    if employee_count <= 0 or "サマリ" not in wb.sheetnames:
        return

    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    pick_fill = PatternFill("solid", start_color="FFC000")   # 選択セル＝目立つ黄色
    label_font = Font(name=FONT, bold=True)
    body_font = Font(name=FONT)
    center = Alignment(horizontal="center")

    ws = wb["サマリ"]
    s_first, s_last = 2, employee_count + 1   # サマリのデータ行範囲
    id_range = f"サマリ!$A${s_first}:$A${s_last}"
    name_range = f"サマリ!$B${s_first}:$B${s_last}"

    # ---- サマリ右側（H:I）に選択ダッシュボードを置く（既存 A:F 表は触らない） ----
    ws["H1"] = "🔎 従業員を選択"
    ws["H1"].font = label_font
    labels = [
        ("H2", "社員番号", "I2", None),
        ("H3", "氏名", "I3", f'={_XLOOKUP}({EMP_PICK_NAME},{id_range},{name_range},"—")'),
        ("H4", "出勤日数", "I4", f'={_XLOOKUP}({EMP_PICK_NAME},{id_range},サマリ!$C${s_first}:$C${s_last},"—")'),
        ("H5", "テレワーク日数", "I5", f'={_XLOOKUP}({EMP_PICK_NAME},{id_range},サマリ!$D${s_first}:$D${s_last},"—")'),
        ("H6", "出社日数", "I6", f'={_XLOOKUP}({EMP_PICK_NAME},{id_range},サマリ!$E${s_first}:$E${s_last},"—")'),
    ]
    for lcell, ltext, vcell, formula in labels:
        ws[lcell] = ltext
        ws[lcell].font = label_font
        if formula is not None:
            ws[vcell] = formula
            ws[vcell].font = body_font
    ws["H7"] = "↑ここで選ぶと「(選択者)」シートが連動します"
    ws["H7"].font = Font(name=FONT, italic=True, size=9, color="808080")

    # 選択セル I2：初期値＝先頭従業員、黄色ハイライト、プルダウン、定義名
    first_id = ws.cell(row=s_first, column=1).value
    ws["I2"] = first_id
    ws["I2"].fill = pick_fill
    ws["I2"].font = Font(name=FONT, bold=True)
    dv = DataValidation(type="list", formula1=f"={id_range}", allow_blank=False)
    dv.error = "一覧にある社員番号を選んでください"
    dv.prompt = "社員番号を選択"
    ws.add_data_validation(dv)
    dv.add(ws["I2"])
    # 既存の定義名があれば消してから登録（再生成に強くする）
    try:
        if EMP_PICK_NAME in wb.defined_names:
            del wb.defined_names[EMP_PICK_NAME]
    except (KeyError, TypeError):
        pass
    wb.defined_names.add(DefinedName(EMP_PICK_NAME, attr_text="サマリ!$I$2"))

    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 22

    def _style_view(vs, headers, banner_label):
        """連動シートの共通スタイル（バナー＋ヘッダー）を整える。"""
        vs["A1"] = "選択中 ▶"
        vs["A1"].font = label_font
        vs["B1"] = f"={EMP_PICK_NAME}"
        vs["B1"].font = Font(name=FONT, bold=True)
        vs["B1"].fill = pick_fill
        vs["C1"] = f'={_XLOOKUP}({EMP_PICK_NAME},{id_range},{name_range},"—")'
        vs["C1"].font = label_font
        vs["E1"] = "※選択は「サマリ」シートの黄色セルで変更"
        vs["E1"].font = Font(name=FONT, italic=True, size=9, color="808080")
        for ci, h in enumerate(headers, start=1):
            c = vs.cell(row=3, column=ci, value=h)
            c.font = header_font
            c.fill = PatternFill("solid", start_color="4472C4")
            c.alignment = center

    # ---- テレワーク明細(選択者) ----
    d_last = _detail_last_row(wb)
    tw_headers = ["社員番号", "氏名", "テレワーク日", "打刻区分"]
    # 元シートに順位ヘルパー列（データ4列の右＝E列）を足す
    tw_helper_col = len(tw_headers) + 1        # E
    _add_source_rank_helper(wb["テレワーク明細"], tw_helper_col, d_last)
    vs = wb.create_sheet("テレワーク明細(選択者)")
    _style_view(vs, tw_headers, "テレワーク")
    tw_n = _max_rows_per_id(wb["テレワーク明細"], d_last)
    _fill_linked_rows(vs, "テレワーク明細", len(tw_headers), get_column_letter(tw_helper_col), d_last, tw_n)
    for col, width in zip("ABCD", [12, 18, 14, 24]):
        vs.column_dimensions[col].width = width
    vs.freeze_panes = "A4"

    # ---- 通勤費(選択者) ----（通勤費シートがある場合のみ） ----
    if commute_row_count > 0 and "通勤費" in wb.sheetnames:
        c_last = commute_row_count + 1          # 合計行(n+2)は含めない
        cm_helper_col = len(COMMUTE_OUTPUT_COLUMNS) + 1   # Q
        _add_source_rank_helper(wb["通勤費"], cm_helper_col, c_last)
        cv = wb.create_sheet("通勤費(選択者)")
        _style_view(cv, COMMUTE_OUTPUT_COLUMNS, "通勤費")
        cm_n = _max_rows_per_id(wb["通勤費"], c_last)
        _fill_linked_rows(cv, "通勤費", len(COMMUTE_OUTPUT_COLUMNS), get_column_letter(cm_helper_col), c_last, cm_n)
        widths = {
            "社員番号": 12, "氏名": 16, "経路No": 7, "出発": 14, "到着": 14, "経由1": 12,
            "経由2": 12, "通勤経路": 60, "利用交通機関": 14, "支給間隔": 10, "支給方法": 8,
            "支給金額": 12, "非課税通勤費": 12, "課税通勤費": 12, "支給開始": 10, "片道距離(km)": 11,
        }
        for i, name in enumerate(COMMUTE_OUTPUT_COLUMNS, start=1):
            cv.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)
        cv.freeze_panes = "A4"

    # ---- シート順：サマリ → 連動ビュー → 元データ ----
    order = ["サマリ", "テレワーク明細(選択者)"]
    if "通勤費(選択者)" in wb.sheetnames:
        order.append("通勤費(選択者)")
    order += [n for n in ("テレワーク明細", "通勤費") if n in wb.sheetnames]
    order += [n for n in wb.sheetnames if n not in order]
    wb._sheets.sort(key=lambda s: order.index(s.title))


# ----------------------------------------------------------------------
# 共通実行関数（CLI / Flask 共用）
# ----------------------------------------------------------------------

@dataclass
class TeleworkResult:
    ok: bool
    output_path: Path
    month: str
    employee_count: int = 0
    telework_total: int = 0
    no_data_count: int = 0
    commute_count: int = 0
    no_commute_kept: int = 0
    no_commute_removed: int = 0
    error: str = ""
    logs: list[str] = field(default_factory=list)


def run_telework_export(
    month: str,
    output_path: Path,
    log_func=print,
    limit: int | None = None,
    commute_csv: "str | Path | None" = None,
    no_commute_xlsx: "str | Path | None" = None,
) -> TeleworkResult:
    """指定月のテレワーク日数・出社日数を集計して Excel を出力する。

    month: "YYYY-MM"
    limit: 動作確認用に先頭 N 名だけ処理する（本番は None）。
    commute_csv: jinjer「人事データ出力（通勤費）」CSV のパス（任意）。指定すると
                 同じブックに「通勤費」シートを追加する。
    no_commute_xlsx: 手動整備の「通勤費申請なし」リスト Excel のパス（任意）。
                 指定すると完全在宅者を除外し通勤費・区分を転記した
                 「通勤費申請なし」シートを同じブックに追加する。
    """
    result = TeleworkResult(ok=False, output_path=output_path, month=month)

    client = JinjerClient()
    try:
        client.authenticate()
    except JinjerAPIError as e:
        result.error = f"jinjer 認証に失敗しました: {e}"
        log_func(f"[error] {result.error}")
        return result

    try:
        employees = fetch_active_employees(client)
    except JinjerAPIError as e:
        result.error = f"従業員一覧の取得に失敗しました: {e}"
        log_func(f"[error] {result.error}")
        return result

    if limit:
        employees = employees[:limit]
    total = len(employees)
    log_func(f"[start] 対象月 {month} / 在籍 {total} 名（1名ずつ取得するため数分かかります）")

    rows: list[dict] = []
    skipped = 0
    for idx, emp in enumerate(employees, 1):
        try:
            attendances = fetch_attendances(client, emp["id"], month)
        except JinjerAPIError as e:
            result.error = f"勤怠取得に失敗しました（{emp['name']} {emp['id']}）: {e}"
            log_func(f"[error] {result.error}")
            return result
        work_days, telework_days = summarize(attendances)
        if work_days == 0 and not telework_days:
            skipped += 1
        rows.append({**emp, "work_days": work_days, "telework_days": telework_days})
        if idx % 25 == 0 or idx == total:
            log_func(f"[進捗] {idx}/{total} 名 取得済み")
        _time.sleep(PACING_SEC)

    # ブックを組み立て: テレワーク（サマリ＋明細）＋ 通勤費シート。
    # 通勤費は既定で jinjer API（commuting-information）から取得。commute_csv 指定時はCSV優先。
    wb = Workbook()
    _build_telework_sheets(wb, month, rows)
    commute_count = 0
    commute_rows: list[dict] = []
    try:
        if commute_csv:
            commute_rows = read_commute_csv(commute_csv)
            log_func(f"[info] 通勤費CSV 読込: {len(commute_rows)} 行 → 通勤費シート追加")
        else:
            id_to_name = {e["id"]: e["name"] for e in employees}
            commute_rows = fetch_commute_rows_via_api(client, id_to_name)
            log_func(f"[info] 通勤情報をAPIから取得: {len(commute_rows)} 行 → 通勤費シート追加")
        add_commute_sheet(wb, commute_rows)
        commute_count = len(commute_rows)
        result.commute_count = commute_count
    except Exception as e:
        # 通勤費が取れなくてもテレワーク集計は出す
        log_func(f"[warn] 通勤費シートの作成に失敗（スキップ）: {e}")
        commute_rows = []

    # 通勤費申請なしリスト（手動Excel）を読み込み、整備済みシートを追加
    if no_commute_xlsx:
        try:
            entries = read_no_commute_list(no_commute_xlsx)
            kept, removed = build_no_commute_rows(entries, rows, commute_rows)
            add_no_commute_sheet(wb, kept)
            result.no_commute_kept = len(kept)
            result.no_commute_removed = len(removed)
            log_func(
                f"[info] 通勤費申請なしリスト: {len(entries)} 名読込 → "
                f"完全在宅 {len(removed)} 名を除外 / {len(kept)} 名を整備"
            )
            if removed:
                log_func("[info] 除外（完全在宅）: " + "、".join(
                    f"{r['id']} {r['name']}".strip() for r in removed))
        except Exception as e:
            # リストが読めなくてもテレワーク集計・通勤費は出す
            log_func(f"[warn] 通勤費申請なしシートの作成に失敗（スキップ）: {e}")

    # 従業員選択で全シート連動するビューを追加（サマリのプルダウン＋(選択者)シート）
    try:
        add_selected_employee_views(wb, employee_count=len(rows), commute_row_count=commute_count)
    except Exception as e:
        log_func(f"[warn] 従業員選択ビューの追加に失敗（スキップ）: {e}")

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    result.employee_count = total
    result.telework_total = sum(len(r["telework_days"]) for r in rows)
    result.no_data_count = skipped
    result.ok = True
    log_func(
        f"[done] 出力完了: {output_path} / テレワーク延べ {result.telework_total} 日 / "
        f"勤怠データなし {skipped} 名"
    )
    return result
