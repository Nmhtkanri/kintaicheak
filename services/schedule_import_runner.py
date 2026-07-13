# -*- coding: utf-8 -*-
"""schedule_import_runner — シフト表グリッドCSVを jinjer API でスケジュール投入する

スケジュールアップロードモードの後工程。従来はグリッドCSVを生成して
ユーザーがjinjer画面から月次スケジュールインポートしていたが、
(a) 一部従業員にサイレントに反映されない (b) 打刻グループ混在で全行エラー
(c) 反映検証手段がない、という問題が実測されたため、
手順③（kintai_import_runner）と同じ思想で API 直接投入に置き換える。

流れ:
    1. 生成済みグリッドCSV群（氏名/従業員ID + 日別セル: 雛形ID or 所/法/休み）を読込
    2. スケジュール雛形一覧CSVで 雛形ID → 出勤/退勤/休憩予定1〜5 を解決
    3. jinjer現状取得（work-schedules・先頭採用 / requested-day-offs / affiliations）
    4. 差分プラン構築: 食い違う日だけ書く。休みセルに予定が残る日は削除不可→要手動一覧
    5. dry-run: 承認用Excel + fingerprint → [ユーザー承認]
    6. execute: プラン再計算＋fingerprint一致ガード → 194列CSV生成(4900行分割) →
       POST kintai-imports(種別5) → ポーリング → 再取得検証（休憩なしまで厳密） → Excelレポート

設計: docs/PLAN_スケジュールAPI投入.md
実証元: Z:\\API連携\\scripts\\reinject_july_grid_schedules.py（2026-07-13・11名159行ドライラン）

手順③との違い（重要）:
    - validate_upload_csv は使わない。スケジュール登録は翌月分（未来月）が本命のため。
      代替ガード = month明示必須／グリッドヘッダー年月一致必須／日付は自前生成（Excel経由なし）
    - 休暇ガードはCSVの休暇列ではなく requested-day-offs API で行う（グリッドに休暇列が無い）
    - 検証は「日丸ごと置換」の仕様に合わせ、期待休憩が空なら現物も空であることまで見る
"""
from __future__ import annotations

import calendar
import csv
import hashlib
import io
import re
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import Config
from services.jinjer_api_client import (
    JinjerAPIError,
    JinjerClient,
    pick_attendance_group_at,
)
from services.jinjer_template_matcher import _tpl_get, load_jinjer_templates
from services.kintai_import_runner import (
    MAX_ROWS_PER_IMPORT,
    poll_import_status,
    t2m,
)

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
# 汎用データインポート（種別5）の194列ヘッダー。
# 出所: 汎用データテンプレート\汎用データ(まるめ適用後)ダウンロード_9637_20260522150030.csv
# 実行時のファイル依存を無くすため定数化（テストで同梱CSVとの一致をドリフト検知する）。
# ※「スケジュール雛形ID」列（5列目）は種別5での動作が未実証のため使わず、
#   実証済みの出勤予定/退勤予定/休憩予定の時刻直書きに統一する。
GENERIC_IMPORT_HEADER: tuple[str, ...] = (
    '名前', '*従業員ID', '*年月日', '*打刻グループID', '所属グループ名', 'スケジュール雛形ID',
    '出勤予定時刻', '退勤予定時刻', '休憩予定時刻1', '復帰予定時刻1', '休憩予定時刻2', '復帰予定時刻2',
    '休憩予定時刻3', '復帰予定時刻3', '休憩予定時刻4', '復帰予定時刻4', '休憩予定時刻5', '復帰予定時刻5',
    'スケジュール外休憩予定時刻', 'スケジュール外復帰予定時刻', '休日（0:法定休日1:所定休日2:法休(振替休出)3:所休(振替休出)4:法休(時間外休出)5:所休(時間外休出)）', '出勤1', '退勤1', '出勤2',
    '退勤2', '出勤3', '退勤3', '出勤4', '退勤4', '出勤5',
    '退勤5', '出勤6', '退勤6', '出勤7', '退勤7', '出勤8',
    '退勤8', '出勤9', '退勤9', '出勤10', '退勤10', '休憩1',
    '復帰1', '休憩2', '復帰2', '休憩3', '復帰3', '休憩4',
    '復帰4', '休憩5', '復帰5', '休憩6', '復帰6', '休憩7',
    '復帰7', '休憩8', '復帰8', '休憩9', '復帰9', '休憩10',
    '復帰10', '食事1開始', '食事1終了', '食事2開始', '食事2終了', '外出1',
    '再入1', '外出2', '再入2', '外出3', '再入3', '外出4',
    '再入4', '外出5', '再入5', '外出6', '再入6', '外出7',
    '再入7', '外出8', '再入8', '外出9', '再入9', '外出10',
    '再入10', '休日休暇名1', '休日休暇名1：種別', '休日休暇名1：開始時間', '休日休暇名1：終了時間', '休日休暇名1：理由',
    '休日休暇名2', '休日休暇名2：種別', '休日休暇名2：開始時間', '休日休暇名2：終了時間', '休日休暇名2：理由', '打刻時コメント',
    '管理者備考', '勤務状況（0:未打刻1:欠勤）', '遅刻取消処理の有無（0:無1:有）', '早退取消処理の有無（0:無1:有）', '遅刻（0:有1:無）', '早退（0:有1:無）',
    '直行1', '直帰1', '直行2', '直帰2', '直行3', '直帰3',
    '直行4', '直帰4', '直行5', '直帰5', '直行6', '直帰6',
    '直行7', '直帰7', '直行8', '直帰8', '直行9', '直帰9',
    '直行10', '直帰10', '打刻区分ID:1', '打刻区分ID:2', '打刻区分ID:3', '打刻区分ID:4',
    '打刻区分ID:5', '打刻区分ID:6', '打刻区分ID:7', '打刻区分ID:8', '打刻区分ID:9', '打刻区分ID:10',
    '打刻区分ID:11', '打刻区分ID:12', '打刻区分ID:13', '打刻区分ID:14', '打刻区分ID:15', '打刻区分ID:16',
    '打刻区分ID:17', '打刻区分ID:18', '打刻区分ID:19', '打刻区分ID:20', '打刻区分ID:21', '打刻区分ID:22',
    '打刻区分ID:23', '打刻区分ID:24', '打刻区分ID:25', '打刻区分ID:26', '打刻区分ID:27', '打刻区分ID:28',
    '打刻区分ID:29', '打刻区分ID:30', '打刻区分ID:31', '打刻区分ID:32', '打刻区分ID:33', '打刻区分ID:34',
    '打刻区分ID:35', '打刻区分ID:36', '打刻区分ID:37', '打刻区分ID:38', '打刻区分ID:39', '打刻区分ID:40',
    '打刻区分ID:41', '打刻区分ID:42', '打刻区分ID:43', '打刻区分ID:44', '打刻区分ID:45', '打刻区分ID:46',
    '打刻区分ID:47', '打刻区分ID:48', '打刻区分ID:49', '打刻区分ID:50', '未打刻', '欠勤',
    '休日打刻', '休暇打刻', '実績確定状況', '総労働時間', '実労働時間', '休憩時間',
    '総残業時間', '法定内残業時間（スケジュール軸）', '法定内残業時間（労働時間軸）', '法定外残業時間', '深夜時間', '不足労働時間数（スケジュール軸）',
    '不足労働時間数（労働時間軸）', '申請承認済総残業時間', '申請承認済法定内残業時間', '申請承認済法定外残業時間', '出勤乖離時間（出勤時刻ー入館時刻）', '退勤乖離時間（退館時刻ー退勤時刻）',
    '出勤乖離時間（出勤時刻ーPC起動時刻）', '退勤乖離時間（PC停止時刻ー退勤時刻）',
)

# グリッドCSVで「休み」を表すセル値（スケジュールを書かない）
# "0" は手作りグリッドの明け休表記（夜勤N20の翌日セル等。KDX 140-160の2026-07実データで確認。
# 雛形IDに "0" は存在しないため休み扱いで安全）
REST_MARKERS = {"", "所", "法", "休み", "休", "0"}
WEEKDAY_KANJI = ["月", "火", "水", "木", "金", "土", "日"]
BREAK_PAIRS = [(f"休憩予定時刻{i}", f"復帰予定時刻{i}") for i in range(1, 6)]

# 要手動確認の区分
KUBUN_DELETE_NEEDED = "休日に予定残存"          # APIでは削除不可 → 画面から手動削除
KUBUN_DAYOFF_REST = "休暇登録日（削除しないで確認）"  # 半休の可能性 → 誤削除防止
KUBUN_DAYOFF_SKIP = "休暇登録日スキップ"        # jinjerが書込を無視するため投入対象外
KUBUN_NO_GROUP = "打刻グループ不明"
KUBUN_NO_EMPID = "従業員ID空欄"
KUBUN_VERIFY_NG = "検証NG"
KUBUN_IMPORT_FAIL = "インポート失敗"
KUBUN_ABORT = "中止（送信前チェック）"

_MONTH_RE = re.compile(r"^(\d{4})-(\d{2})$")


# ---------------------------------------------------------------------------
# 純粋関数（テスト対象・I/Oなし）
# ---------------------------------------------------------------------------
def norm_time(v) -> str:
    """'09:00:00'→'9:00'。24h超（'33:30'）は維持。不正は空文字"""
    s = str(v or "").strip()
    m = re.match(r"^(\d{1,3}):(\d{2})(?::\d{2})?$", s)
    return f"{int(m.group(1))}:{m.group(2)}" if m else ""


def norm_cell(v) -> str:
    """グリッドのセル値・雛形IDを正規化する（NFKC=全角英数→半角、前後空白除去）。

    手作りグリッドに全角の雛形ID（'１' 等）が混ざる実データがあるため
    （UAL時給制 2026-07 で確認）、突合前に表記を揃える。所/法/休み等の漢字は不変。
    """
    return unicodedata.normalize("NFKC", str(v or "").strip())


def fmt_breaks(brs) -> str:
    return " ".join(f"{s}-{e}" for s, e in brs) if brs else "(なし)"


def parse_grid_rows(rows: list[list[str]], *, filename: str) -> dict:
    """グリッドCSVの生 rows（csv.reader 出力）を解析する。

    フォーマット（jinjer月次スケジュール登録CSV = 当アプリの exporter 出力）:
        行1: "{Y}年", "{M}月", 1, 2, ..., 31
        行2: "氏名", "従業員ID", 曜日...
        行3〜: 氏名, 従業員ID, セル値（雛形ID / 所 / 法 / 休み）...

    Returns:
        {"year": int, "month": int,
         "employees": {emp_id: {"name": str, "file": str, "days": {day(int): cell(str)}}},
         "no_empid_names": [氏名, ...]}   # 従業員ID空欄の行

    Raises:
        ValueError: 年月ヘッダーが読めない場合
    """
    if len(rows) < 3:
        raise ValueError(f"{filename}: 行数が不足しています（ヘッダー2行+データ行が必要）")
    h1 = [str(c or "").strip() for c in rows[0]]
    m_y = re.match(r"^(\d{4})年$", h1[0] if len(h1) > 0 else "")
    m_m = re.match(r"^(\d{1,2})月$", h1[1] if len(h1) > 1 else "")
    if not m_y or not m_m:
        raise ValueError(f"{filename}: 年月ヘッダーが不正です（先頭2セル={h1[:2]}）")
    year, month = int(m_y.group(1)), int(m_m.group(1))
    day_of_col = {j: int(h1[j]) for j in range(2, len(h1)) if h1[j].isdigit()}

    employees: dict[str, dict] = {}
    no_empid: list[str] = []
    for r in rows[2:]:
        if len(r) < 2:
            continue
        name = str(r[0] or "").strip()
        emp = norm_cell(r[1])  # 全角の従業員IDも許容
        if not name and not emp:
            continue
        if not emp:
            no_empid.append(name or "(名無し)")
            continue
        days = {d: (norm_cell(r[j]) if j < len(r) else "")
                for j, d in day_of_col.items()}
        employees[emp] = {"name": name, "file": filename, "days": days}
    return {"year": year, "month": month,
            "employees": employees, "no_empid_names": no_empid}


def merge_grids(grids: list[dict]) -> tuple[dict[str, dict], list[str]]:
    """複数グリッドの employees をマージする。

    同一従業員が複数ファイルに居る場合、内容（days）が同一なら先勝ちで許容、
    異なればエラー（投入中止の判断材料）。

    Returns:
        (employees, errors)
    """
    merged: dict[str, dict] = {}
    errors: list[str] = []
    for g in grids:
        for emp, info in (g.get("employees") or {}).items():
            if emp in merged:
                if merged[emp]["days"] != info["days"]:
                    errors.append(
                        f"従業員 {emp}({info['name']}) が複数グリッドに異なる内容で存在: "
                        f"{merged[emp]['file']} / {info['file']} → どちらが正か確認してください")
                continue
            merged[emp] = info
    return merged, errors


def _tpl_break_pair(row: dict, i: int) -> tuple[str, str]:
    """雛形CSVの休憩i列を取得する。列名の癖: 休憩1だけ「休憩開始時間1」、2〜5は「休憩時間N」"""
    start_keys = ([f"休憩開始時間{i}(0:00~47:59)", f"休憩時間{i}(0:00~47:59)"]
                  if i == 1 else
                  [f"休憩時間{i}(0:00~47:59)", f"休憩開始時間{i}(0:00~47:59)"])
    bs = ""
    for k in start_keys:
        v = str(row.get(k) or "").strip()
        if v:
            bs = v
            break
    be = str(row.get(f"復帰時間{i}(0:00~47:59)") or "").strip()
    return norm_time(bs), norm_time(be)


def build_template_index(templates: list[dict]) -> dict[str, dict]:
    """雛形一覧CSVの行リスト（load_jinjer_templates出力）→ {雛形ID: 定義} 索引。

    Returns:
        {雛形ID: {"name": str, "start": "9:00", "end": "18:00",
                  "breaks": [("12:00","13:00"), ...]}}   # breaks は開始時刻順
    """
    index: dict[str, dict] = {}
    for row in templates or []:
        tid = norm_cell(_tpl_get(row, "＊スケジュール雛形ID"))
        if not tid:
            continue
        start = norm_time(_tpl_get(row, "＊出勤時間(0:00~47:59)"))
        end = norm_time(_tpl_get(row, "＊退勤時間(0:00~47:59)"))
        breaks = []
        for i in range(1, 6):
            bs, be = _tpl_break_pair(row, i)
            if t2m(bs) is not None and t2m(be) is not None:
                breaks.append((bs, be))
        breaks.sort(key=lambda x: t2m(x[0]))
        index[tid] = {"name": str(_tpl_get(row, "＊スケジュール雛形名") or "").strip(),
                      "start": start, "end": end, "breaks": breaks}
    return index


def find_unknown_cells(employees: dict[str, dict], tpl_index: dict[str, dict]) -> list[tuple[str, str]]:
    """休みマーカーでも雛形IDでもないセル値を列挙する（1件でもあれば投入前に中止）。

    Returns:
        [(セル値, 従業員ID), ...]  ソート済み・重複なし
    """
    unknown = {(cell, emp)
               for emp, info in (employees or {}).items()
               for cell in info["days"].values()
               if cell not in REST_MARKERS and cell not in tpl_index}
    return sorted(unknown)


def _breaks_minutes(brs) -> list[tuple[int, int]]:
    return sorted((t2m(s), t2m(e)) for s, e in (brs or []) if t2m(s) is not None)


@dataclass
class PlanBuildResult:
    plan: list[dict] = field(default_factory=list)
    manual: list[dict] = field(default_factory=list)     # 要手動確認（従業員番号/氏名/日付/区分/備考）
    matched: dict[str, int] = field(default_factory=dict)  # emp → 一致日数
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def build_diff_plan(
    employees: dict[str, dict],
    tpl_index: dict[str, dict],
    current: dict[str, dict[str, dict]],
    dayoffs: dict[str, dict[str, str]],
    groups: dict[str, tuple[str, str]],
    *,
    year: int,
    month: int,
    exclude_emps: set[str] | None = None,
) -> PlanBuildResult:
    """グリッド（あるべき姿）と jinjer 現状の差分から投入プランを作る（純粋関数）。

    判定（reinject_july_grid_schedules.py の実績ロジック）:
        - 休みセル × 予定残存 → 要手動（休暇登録があれば「削除しないで確認」区分。
          グリッドは半休を 所/法 で表現するため、休暇登録日の予定は正しい可能性がある）
        - 勤務セル × 現状一致（開始/終了/休憩を分単位比較）→ matched（書かない）
        - 勤務セル × 休暇登録日 → 要手動（jinjerが書込を無視するため投入しない）
        - 勤務セル × 差分 → plan 行（現状行なし=新規 / あり=修正）
    """
    result = PlanBuildResult()
    exclude_emps = exclude_emps or set()
    days_in_month = calendar.monthrange(year, month)[1]

    for emp in sorted(employees):
        info = employees[emp]
        if emp in exclude_emps:
            result.warnings.append(f"{emp} {info['name']}: 個別除外指定によりスキップ")
            continue
        gid, gname = groups.get(emp, ("", ""))
        if not gid:
            result.manual.append({
                "従業員番号": emp, "氏名": info["name"], "日付": "",
                "区分": KUBUN_NO_GROUP,
                "備考": "所属履歴から打刻グループを解決できず → この従業員は投入不可"})
            continue
        cur_rows = current.get(emp, {})
        offs = dayoffs.get(emp, {})
        for day in range(1, days_in_month + 1):
            d_iso = f"{year:04d}-{month:02d}-{day:02d}"
            cell = info["days"].get(day, "")
            cur = cur_rows.get(d_iso)
            cur_has = bool(cur and (t2m(cur.get("start")) is not None
                                    or t2m(cur.get("end")) is not None))
            off = offs.get(d_iso)
            if cell in REST_MARKERS:
                if cur_has:
                    cur_txt = f"{cur.get('start')}-{cur.get('end')}"
                    if off:
                        result.manual.append({
                            "従業員番号": emp, "氏名": info["name"], "日付": d_iso,
                            "区分": KUBUN_DAYOFF_REST,
                            "備考": f"グリッド={cell or '(空欄)'} 現予定={cur_txt} "
                                    f"休暇登録={off} → 半休の予定かもしれないため削除前に要確認"})
                    else:
                        result.manual.append({
                            "従業員番号": emp, "氏名": info["name"], "日付": d_iso,
                            "区分": KUBUN_DELETE_NEEDED,
                            "備考": f"グリッド={cell or '(空欄)'} 現予定={cur_txt} "
                                    "→ APIではスケジュール削除不可・jinjer画面から手動削除"})
                continue
            tpl = tpl_index.get(cell)
            if tpl is None:
                # find_unknown_cells で事前に弾いている前提の防御
                result.errors.append(f"{emp} {d_iso}: 未知のセル値 {cell!r}")
                continue
            same = (cur_has
                    and t2m(cur.get("start")) == t2m(tpl["start"])
                    and t2m(cur.get("end")) == t2m(tpl["end"])
                    and _breaks_minutes(cur.get("breaks")) == _breaks_minutes(tpl["breaks"]))
            if same:
                result.matched[emp] = result.matched.get(emp, 0) + 1
                continue
            cur_txt = (f"{cur.get('start')}-{cur.get('end')} "
                       f"休憩[{fmt_breaks(cur.get('breaks') or [])}]" if cur_has else "(行なし)")
            if off:
                result.manual.append({
                    "従業員番号": emp, "氏名": info["name"], "日付": d_iso,
                    "区分": KUBUN_DAYOFF_SKIP,
                    "備考": f"雛形{cell}({tpl['start']}-{tpl['end']})にしたいが休暇登録あり"
                            f"({off}) 現={cur_txt} → jinjerが書込を無視するため投入対象外"})
                continue
            result.plan.append({
                "emp": emp, "name": info["name"],
                "date_iso": d_iso, "day": day,
                "youbi": WEEKDAY_KANJI[date(year, month, day).weekday()],
                "cell": cell, "tpl_name": tpl["name"],
                "start": tpl["start"], "end": tpl["end"], "breaks": list(tpl["breaks"]),
                "cur": cur_txt,
                "kind": "修正" if cur_has else "新規",
                "store_id": gid, "store_name": gname,
            })
    return result


def plan_fingerprint(plan: list[dict]) -> str:
    """プランの内容ハッシュ。dry-run承認後にjinjer側が変わっていないかの検知に使う。"""
    lines = sorted(
        f"{p['emp']}|{p['date_iso']}|{p['start']}|{p['end']}|"
        f"{','.join(f'{s}-{e}' for s, e in p['breaks'])}|{p['store_id']}"
        for p in plan or []
    )
    return hashlib.sha1("\n".join(lines).encode("utf-8")).hexdigest()


def build_import_rows(
    plan: list[dict],
    header: Sequence[str] = GENERIC_IMPORT_HEADER,
) -> list[list[str]]:
    """プラン行 → 汎用データインポート形式の行リスト。

    値を入れるのは 名前/*従業員ID/*年月日/*打刻グループID/出勤予定時刻/退勤予定時刻/
    休憩予定時刻1〜5/復帰予定時刻1〜5 のみ。他列は空欄＝jinjer仕様「無処理」
    （打刻・休暇には一切触れない）。日付は jinjer形式 YYYY/M/D（ゼロ埋めなし）。
    """
    ci = {c: header.index(c) for c in
          ["名前", "*従業員ID", "*年月日", "*打刻グループID",
           "出勤予定時刻", "退勤予定時刻"] + [c for p in BREAK_PAIRS for c in p]}
    rows: list[list[str]] = []
    for p in sorted(plan, key=lambda x: (x["emp"], x["date_iso"])):
        y, m, d = (int(x) for x in p["date_iso"].split("-"))
        row = [""] * len(header)
        row[ci["名前"]] = p["name"]
        row[ci["*従業員ID"]] = p["emp"]
        row[ci["*年月日"]] = f"{y}/{m}/{d}"
        row[ci["*打刻グループID"]] = p["store_id"]
        row[ci["出勤予定時刻"]] = p["start"]
        row[ci["退勤予定時刻"]] = p["end"]
        for (bs_col, be_col), (bs, be) in zip(BREAK_PAIRS, p["breaks"]):
            row[ci[bs_col]] = bs
            row[ci[be_col]] = be
        rows.append(row)
    return rows


def rows_to_csv_bytes(header: Sequence[str], rows: list[list[str]]) -> bytes:
    """CP932 / CRLF のCSVバイト列（jinjer kintai-imports の入力仕様）"""
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow(list(header))
    w.writerows(rows)
    return buf.getvalue().encode("cp932")


def verify_plan_rows(
    plan: list[dict],
    after: dict[str, dict[str, dict]],
) -> tuple[list[dict], list[dict]]:
    """投入後の反映を検証する（純粋関数）。

    スケジュール書込は日単位の丸ごと置換のため、開始・終了に加えて
    **期待休憩が空なら現物も空であること** まで厳密に比較する
    （kintai_import_runner.compare_row の「空欄=無検証」とは意図的に異なる）。

    Returns:
        (検証結果全行 [{従業員番号,氏名,日付,判定,詳細}],
         NG行の要手動リスト [{従業員番号,氏名,日付,区分,備考}])
    """
    verify_rows: list[dict] = []
    ng_manual: list[dict] = []
    for p in sorted(plan, key=lambda x: (x["emp"], x["date_iso"])):
        got = (after.get(p["emp"]) or {}).get(p["date_iso"])
        want_brs = _breaks_minutes(p["breaks"])
        good = bool(
            got
            and t2m(got.get("start")) == t2m(p["start"])
            and t2m(got.get("end")) == t2m(p["end"])
            and _breaks_minutes(got.get("breaks")) == want_brs
        )
        if good:
            verify_rows.append({"従業員番号": p["emp"], "氏名": p["name"],
                                "日付": p["date_iso"], "判定": "OK", "詳細": ""})
            continue
        got_txt = (f"{got.get('start')}-{got.get('end')} "
                   f"休憩[{fmt_breaks(got.get('breaks') or [])}]" if got else "(行なし)")
        want_txt = f"{p['start']}-{p['end']} 休憩[{fmt_breaks(p['breaks'])}]"
        detail = f"期待 {want_txt} → 実際 {got_txt}"
        verify_rows.append({"従業員番号": p["emp"], "氏名": p["name"],
                            "日付": p["date_iso"], "判定": "NG", "詳細": detail})
        ng_manual.append({"従業員番号": p["emp"], "氏名": p["name"],
                          "日付": p["date_iso"], "区分": KUBUN_VERIFY_NG, "備考": detail})
    return verify_rows, ng_manual


# ---------------------------------------------------------------------------
# 実行結果
# ---------------------------------------------------------------------------
@dataclass
class ScheduleImportResult:
    ok: bool = False
    dry_run: bool = True
    month: str = ""
    plan_rows: int = 0
    matched_rows: int = 0
    plan: list[dict] = field(default_factory=list)
    manual: list[dict] = field(default_factory=list)
    fingerprint: str = ""
    submitted_rows: int = 0
    import_statuses: list[str] = field(default_factory=list)
    verified_ok: int = 0
    verified_ng: int = 0
    report_path: str = ""
    log: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# I/O
# ---------------------------------------------------------------------------
def load_grid_file(path: Path) -> dict:
    """グリッドCSV（CP932）を読み込んで parse_grid_rows する"""
    with open(path, encoding="cp932", newline="") as f:
        rows = list(csv.reader(f))
    return parse_grid_rows(rows, filename=path.name)


def _with_retry(fn, what: str, log, attempts: int = 3, wait_sec: float = 8.0):
    """jinjer APIの一時的な失敗（429等）をリトライする。最終試行の例外はそのまま送出"""
    for attempt in range(1, attempts + 1):
        try:
            return fn()
        except JinjerAPIError as e:
            if attempt >= attempts:
                raise
            log(f"  [WARN] {what} 失敗（{e}）→ {wait_sec:.0f}秒後にリトライ "
                f"{attempt}/{attempts - 1}")
            time.sleep(wait_sec * attempt)


def run_schedule_api_import(
    grid_csvs: list[Path],
    output_dir: Path,
    month: str,
    executor_id: str = "",
    dry_run: bool = True,
    expected_fingerprint: str = "",
    expected_rows: int | None = None,
    exclude_emps: set[str] | None = None,
    template_csv: str = "",
    log_func: Callable[[str], None] | None = None,
    client: JinjerClient | None = None,
) -> ScheduleImportResult:
    """グリッドCSV群を差分プラン化し、jinjer API でスケジュール投入する。

    Args:
        grid_csvs: グリッドCSVファイルのパス群（当アプリの exporter 出力 or 手作りの同形式）
        output_dir: レポート出力先
        month: 対象月 "YYYY-MM"（必須・グリッドのヘッダー年月と一致必須。未来月OK）
        executor_id: kintai-imports の実行者社員番号（既定""=未指定運用。
            勤怠管理者権限の無いIDを指定するとサイレント破棄されるため注意）
        dry_run: True なら差分プランとExcelのみ（jinjerへ送信しない）
        expected_fingerprint: execute時、dry-runで得た fingerprint と再計算値が
            一致しなければ投入せず中止（承認後にjinjer側が変わった場合の保険）
        expected_rows: execute時、書込行数がこの値と一致しなければ中止（CLI --expect）
        exclude_emps: 従業員個別除外
        template_csv: 雛形一覧CSVパス（空なら Config.get_jinjer_template_csv_path()）
        log_func: 進捗ログ出力先（未指定なら print）
        client: テスト時に差し替え可能な JinjerClient
    """
    result = ScheduleImportResult(dry_run=dry_run, month=month)

    def log(msg: str) -> None:
        result.log.append(msg)
        (log_func or print)(msg)

    def abort(msgs: list[str]) -> ScheduleImportResult:
        for m in msgs:
            log(f"[中止] {m}")
            result.manual.append({"従業員番号": "", "氏名": "", "日付": "",
                                  "区分": KUBUN_ABORT, "備考": m})
        log("送信前チェックNGのため、jinjerへは送信していません。")
        result.ok = False
        result.report_path = _write_schedule_report(output_dir, result, [], [], dry_run=dry_run)
        return result

    # ---- 1. 対象月・グリッド読込 ----
    m_month = _MONTH_RE.match(month or "")
    if not m_month:
        return abort([f"対象月の形式が不正です: {month!r}（YYYY-MM で指定してください）"])
    year, mon = int(m_month.group(1)), int(m_month.group(2))

    grids: list[dict] = []
    errs: list[str] = []
    for p in grid_csvs:
        p = Path(p)
        try:
            g = load_grid_file(p)
        except FileNotFoundError:
            errs.append(f"グリッドCSVが見つかりません: {p}")
            continue
        except (ValueError, UnicodeDecodeError) as e:
            errs.append(f"グリッドCSVを読めません: {p.name}: {e}")
            continue
        if (g["year"], g["month"]) != (year, mon):
            errs.append(f"{p.name}: グリッドの年月 {g['year']}-{g['month']:02d} が"
                        f"対象月 {month} と一致しません")
            continue
        for nm in g["no_empid_names"]:
            result.manual.append({"従業員番号": "", "氏名": nm, "日付": "",
                                  "区分": KUBUN_NO_EMPID,
                                  "備考": f"{p.name}: 従業員ID空欄のため投入対象外"})
        grids.append(g)
    if errs:
        return abort(errs)
    if not grids:
        return abort(["読み込めたグリッドCSVがありません"])

    employees, merge_errors = merge_grids(grids)
    if merge_errors:
        return abort(merge_errors)
    if not employees:
        return abort(["グリッドCSVに従業員行がありません"])
    # 日付列が月の日数に満たないグリッドは、無い日を休み扱いにしてしまうため警告
    last_day_of_month = calendar.monthrange(year, mon)[1]
    for g in grids:
        if not g["employees"]:
            continue
        covered = set(next(iter(g["employees"].values()))["days"].keys())
        miss = [d for d in range(1, last_day_of_month + 1) if d not in covered]
        if miss:
            fname = next(iter(g["employees"].values()))["file"]
            log(f"[注意] {fname}: 日付列が {len(miss)}日分ありません"
                f"（{miss[:5]}{'…' if len(miss) > 5 else ''}）→ 無い日は休み扱いになります。"
                "既存予定があると「休日に予定残存」として列挙されます")
    log(f"グリッド読込: {len(grid_csvs)}ファイル / 対象 {len(employees)}名 / 対象月 {month}")
    log("※jinjer画面からの月次スケジュールインポートは併用しないでください（このAPI投入に一本化）")

    # ---- 2. 雛形解決 ----
    tpl_path = template_csv or Config.get_jinjer_template_csv_path()
    templates = load_jinjer_templates(tpl_path)
    if not templates:
        return abort([f"スケジュール雛形一覧CSVを読めません: {tpl_path}"])
    tpl_index = build_template_index(templates)
    log(f"雛形定義: {len(tpl_index)}件 （{Path(tpl_path).name}）")
    unknown = find_unknown_cells(employees, tpl_index)
    if unknown:
        return abort(
            [f"雛形一覧に無いセル値: {cell!r}（従業員 {emp}）" for cell, emp in unknown]
            + ["→ 新規雛形をjinjerに登録し、雛形一覧CSVを更新してから再実行してください"])

    # ---- 3. jinjer 現状取得 ----
    cli = client or JinjerClient()
    emps = sorted(employees)
    current: dict[str, dict[str, dict]] = {}
    dayoffs: dict[str, dict[str, str]] = {}
    log(f"jinjer現状取得中（work-schedules / requested-day-offs × {len(emps)}名）…")
    for i, emp in enumerate(emps, 1):
        try:
            current[emp] = _with_retry(
                lambda e=emp: cli.get_work_schedules(e, month),
                f"work-schedules取得(emp={emp})", log)
        except JinjerAPIError as e:
            # 現状が取れないまま差分を作ると全日「新規」と誤判定するため中止する
            return abort([f"スケジュール現状の取得に失敗（emp={emp}）: {e}"])
        time.sleep(0.15)
        dayoffs[emp] = cli.get_requested_day_offs(emp, month)  # 失敗は空dict（検証NGで顕在化）
        time.sleep(0.15)
        if i % 10 == 0 or i == len(emps):
            log(f"  … {i}/{len(emps)} 名")
    n_off = sum(len(v) for v in dayoffs.values())
    if n_off:
        log(f"休暇登録: {n_off}件")
        for emp in emps:
            for d_iso, info in sorted(dayoffs.get(emp, {}).items()):
                log(f"    {emp} {employees[emp]['name']} {d_iso} {info}")

    # 打刻グループ: 既存スケジュール行のstoreは誤レイヤーを含むことがあるため
    # 所属履歴（affiliations）の現行グループを正とする（能美・及川 2026-07 実測）
    try:
        affs_map = _with_retry(lambda: cli.get_affiliations(emps), "所属履歴取得", log)
    except JinjerAPIError as e:
        return abort([f"所属履歴の取得に失敗: {e}"])
    last_day = calendar.monthrange(year, mon)[1]
    target_d = date(year, mon, last_day)
    groups: dict[str, tuple[str, str]] = {}
    for emp in emps:
        affs = affs_map.get(emp, [])
        groups[emp] = pick_attendance_group_at(affs, target_d)
        for a in affs:
            di = str(a.get("date_of_issue") or "")
            if f"{year:04d}-{mon:02d}-01" < di <= f"{year:04d}-{mon:02d}-{last_day:02d}":
                log(f"  [注意] {emp} {employees[emp]['name']}: 月中 {di} に所属変更あり"
                    f"（投入グループは月末時点 {groups[emp][0]} を使用）")

    # ---- 4. 差分プラン ----
    built = build_diff_plan(employees, tpl_index, current, dayoffs, groups,
                            year=year, month=mon, exclude_emps=exclude_emps)
    if built.errors:
        return abort(built.errors)
    for w in built.warnings:
        log(f"  {w}")
    result.plan = built.plan
    result.manual.extend(built.manual)
    result.plan_rows = len(built.plan)
    result.matched_rows = sum(built.matched.values())
    result.fingerprint = plan_fingerprint(built.plan)

    summary_rows = []
    for emp in emps:
        rows_ = [p for p in built.plan if p["emp"] == emp]
        summary_rows.append({
            "従業員番号": emp, "氏名": employees[emp]["name"],
            "グリッドファイル": employees[emp]["file"],
            "グループ": f"{groups.get(emp, ('', ''))[0]}({groups.get(emp, ('', ''))[1]})",
            "新規": sum(1 for p in rows_ if p["kind"] == "新規"),
            "修正": sum(1 for p in rows_ if p["kind"] == "修正"),
            "一致": built.matched.get(emp, 0),
            "要手動確認": sum(1 for m in built.manual if m["従業員番号"] == emp),
            "休暇登録": len(dayoffs.get(emp, {})),
        })

    log(f"\n=== 差分プラン: 書込 {result.plan_rows}行 / "
        f"{len({p['emp'] for p in built.plan})}名 （一致 {result.matched_rows}日は書かない） ===")
    for s in summary_rows:
        if s["新規"] or s["修正"] or s["要手動確認"]:
            log(f"  {s['従業員番号']} {s['氏名']}: 新規{s['新規']} 修正{s['修正']} "
                f"一致{s['一致']} 要手動{s['要手動確認']} グループ{s['グループ']}")
    if result.manual:
        log(f"要手動確認: {len(result.manual)}件（レポートの「要手動確認」シート参照）")
    log(f"プランfingerprint: {result.fingerprint}")

    if dry_run:
        result.ok = True
        result.report_path = _write_schedule_report(
            output_dir, result, [], summary_rows, dry_run=True)
        log(f"dry-run 完了（jinjerへは送信していません）。承認用レポート: {result.report_path}")
        return result

    # ---- 5. 投入前ガード ----
    if expected_rows is not None and result.plan_rows != expected_rows:
        return abort([f"書込行数 {result.plan_rows} が承認済みの {expected_rows} と一致しません。"
                      "dry-run からやり直してください"])
    if expected_fingerprint and result.fingerprint != expected_fingerprint:
        return abort(["差分プランが承認時から変化しています（jinjer側が更新された可能性）。"
                      "dry-run からやり直してください"])
    if not built.plan:
        result.ok = True
        log("書込対象がありません（すべて一致または要手動確認）")
        result.report_path = _write_schedule_report(
            output_dir, result, [], summary_rows, dry_run=False)
        return result

    # ---- 6. 投入（4900行分割・同時予約1件のため直列） ----
    import_rows = build_import_rows(built.plan)
    ts = datetime.now().strftime("%H%M%S")
    chunks = [import_rows[i:i + MAX_ROWS_PER_IMPORT]
              for i in range(0, len(import_rows), MAX_ROWS_PER_IMPORT)]
    for n, chunk in enumerate(chunks, 1):
        csv_bytes = rows_to_csv_bytes(GENERIC_IMPORT_HEADER, chunk)
        file_name = f"スケジュールAPI投入_{month.replace('-', '')}_{ts}_{n}.csv"
        log(f"投入 {n}/{len(chunks)}: {file_name} （{len(chunk)}行, {len(csv_bytes):,} bytes）"
            f" executor={executor_id or '(未指定=マスタ)'}")
        resp = cli.post_kintai_import(csv_bytes, file_name, executor_id=executor_id or None)
        log(f"  POST応答: executor={resp.get('executor')} type={resp.get('type')}")
        result.submitted_rows += len(chunk)
        status = poll_import_status(cli, file_name, log)
        result.import_statuses.append(status)
        if status != "1":
            log(f"[ERROR] インポートが成功しませんでした (status={status})")
            ci_emp = GENERIC_IMPORT_HEADER.index("*従業員ID")
            ci_date = GENERIC_IMPORT_HEADER.index("*年月日")
            ci_name = GENERIC_IMPORT_HEADER.index("名前")
            for r in chunk:
                result.manual.append({
                    "従業員番号": r[ci_emp], "氏名": r[ci_name],
                    "日付": r[ci_date], "区分": KUBUN_IMPORT_FAIL,
                    "備考": f"インポートstatus={status}。通知メールを確認して再実行"})
            result.report_path = _write_schedule_report(
                output_dir, result, [], summary_rows, dry_run=False)
            return result

    # ---- 7. 反映検証（開始・終了・休憩を厳密比較） ----
    log("反映検証中（work-schedules API 再取得）…")
    after: dict[str, dict[str, dict]] = {}
    verify_emps = sorted({p["emp"] for p in built.plan})
    for i, emp in enumerate(verify_emps, 1):
        try:
            after[emp] = _with_retry(
                lambda e=emp: cli.get_work_schedules(e, month),
                f"検証取得(emp={emp})", log)
        except JinjerAPIError as e:
            log(f"[WARN] 検証取得失敗 emp={emp}: {e}")
            after[emp] = {}
        time.sleep(0.15)
        if i % 10 == 0 or i == len(verify_emps):
            log(f"  … {i}/{len(verify_emps)} 名検証済み")
    verify_rows, ng_manual = verify_plan_rows(built.plan, after)
    result.verified_ok = sum(1 for v in verify_rows if v["判定"] == "OK")
    result.verified_ng = len(ng_manual)
    result.manual.extend(ng_manual)
    result.ok = result.verified_ng == 0
    log(f"検証結果: 反映OK {result.verified_ok} / NG {result.verified_ng}")

    result.report_path = _write_schedule_report(
        output_dir, result, verify_rows, summary_rows, dry_run=False)
    log(f"レポート出力: {result.report_path}")
    return result


# ---------------------------------------------------------------------------
# Excel レポート
# ---------------------------------------------------------------------------
def _write_schedule_report(
    output_dir: Path,
    result: ScheduleImportResult,
    verify_rows: list[dict],
    summary_rows: list[dict],
    dry_run: bool,
) -> str:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    tag = "ドライラン" if dry_run else "結果"
    path = output_dir / (f"スケジュールAPI投入_{tag}_{(result.month or 'unknown').replace('-', '')}"
                         f"_{datetime.now():%H%M%S}.xlsx")

    wb = Workbook()
    fill = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")

    def _sheet(ws, cols: list[str], rows: list[list], widths: list[int]) -> None:
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=j, value=c)
            cell.font = Font(bold=True)
            cell.fill = fill
        for i, r in enumerate(rows, 2):
            for j, v in enumerate(r, 1):
                ws.cell(row=i, column=j, value=v)
        for j, wd in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = wd
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"

    ws1 = wb.active
    ws1.title = "書込予定"
    _sheet(
        ws1,
        ["従業員番号", "氏名", "日付", "曜", "区分", "グリッド値", "雛形名",
         "新予定", "新休憩", "現状", "グループID", "グループ名"],
        [[p["emp"], p["name"], p["date_iso"], p["youbi"], p["kind"], p["cell"],
          p["tpl_name"], f"{p['start']}-{p['end']}", fmt_breaks(p["breaks"]),
          p["cur"], p["store_id"], p["store_name"]]
         for p in sorted(result.plan, key=lambda x: (x["emp"], x["date_iso"]))],
        [12, 12, 12, 4, 6, 10, 16, 14, 18, 34, 10, 14],
    )

    ws2 = wb.create_sheet("要手動確認")
    _sheet(
        ws2,
        ["従業員番号", "氏名", "日付", "区分", "備考"],
        [[m.get("従業員番号", ""), m.get("氏名", ""), m.get("日付", ""),
          m.get("区分", ""), m.get("備考", "")]
         for m in sorted(result.manual,
                         key=lambda x: (x.get("区分", ""), x.get("従業員番号", ""),
                                        x.get("日付", "")))],
        [12, 12, 12, 26, 90],
    )

    ws3 = wb.create_sheet("サマリ")
    _sheet(
        ws3,
        ["従業員番号", "氏名", "グリッドファイル", "グループ",
         "新規", "修正", "一致", "要手動確認", "休暇登録"],
        [[s["従業員番号"], s["氏名"], s["グリッドファイル"], s["グループ"],
          s["新規"], s["修正"], s["一致"], s["要手動確認"], s["休暇登録"]]
         for s in summary_rows],
        [12, 12, 48, 20, 6, 6, 6, 10, 8],
    )

    if verify_rows:
        ws4 = wb.create_sheet("検証結果")
        _sheet(
            ws4,
            ["従業員番号", "氏名", "日付", "判定", "詳細"],
            [[v["従業員番号"], v["氏名"], v["日付"], v["判定"], v["詳細"]]
             for v in verify_rows],
            [12, 12, 12, 6, 80],
        )

    ws5 = wb.create_sheet("取込ログ")
    ws5.cell(row=1, column=1, value="ログ").font = Font(bold=True)
    ws5.cell(row=1, column=1).fill = fill
    for i, line in enumerate(result.log, 2):
        ws5.cell(row=i, column=1, value=line)
    ws5.column_dimensions["A"].width = 120

    wb.save(path)
    return str(path)
