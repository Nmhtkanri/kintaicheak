"""手順1(突合)＋手順2(差異一覧)を1アクションで実行するオーケストレーション（機能3）。

入力（パス指定。ブラウザの1ファイルずつアップロードを廃止）:
  - 請求勤怠フォルダ（or 単一ファイル）: 配下を再帰的に集め、形式は自動判定で解析。
  - jinjer 汎用データCSV（or フォルダ）: 出勤1/退勤1 を手順1の打刻ソースにも流用し、手順2の汎用データにも使う。
  - 申請データCSV（任意）: 打刻修正時コメント用。
  - 対象月。

出力: 差異一覧xlsx（トリアージ済み）＋「未処理・未マッチ」シート（解析できなかった請求勤怠・未提出者）。
グループはまたがない（複数グループは従来どおり別々に実行）。
"""
from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from services.timesheet_parser import parse_timesheet_smart
from services.jinjer_parser import parse_jinjer_csv
from services.matcher import match
from services.excel_exporter import export_to_excel
from quick_compare import run_quick_compare, LogEntry, _input_files

# 請求勤怠として扱う拡張子
_TIMESHEET_EXTS = (".xlsx", ".xls", ".xlsb", ".csv", ".txt", ".pdf", ".png", ".jpg", ".jpeg")
# 集計系フォルダ（日別の請求勤怠ではないので一括解析の対象外）
_EXCLUDE_DIR_NAMES = ("月の総労働時間", "日の合計勤務時間")


def gather_timesheet_files(timesheet_dir: Path) -> list[Path]:
    """請求勤怠フォルダ配下を再帰的に集める（単一ファイル指定にも対応）。

    一時ファイル(~$)・集計系フォルダ配下は除外する。
    """
    if timesheet_dir.is_file():
        return [timesheet_dir]
    files: list[Path] = []
    for p in sorted(timesheet_dir.rglob("*")):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in _TIMESHEET_EXTS:
            continue
        if any(part in _EXCLUDE_DIR_NAMES for part in p.parts):
            continue
        files.append(p)
    return files


def _append_unprocessed_sheet(
    output_path: Path, skipped: list[tuple[str, str]], unsubmitted: list[str]
) -> None:
    """差異一覧xlsx に「未処理・未マッチ」シートを追記する。

    skipped: 解析できなかった請求勤怠ファイル [(ファイル名, 理由)]。
    unsubmitted: jinjer にいるが請求勤怠が無い（＝未提出の）氏名。
    """
    if not output_path.exists():
        return
    try:
        wb = load_workbook(output_path)
    except Exception:
        return
    ws = wb.create_sheet("未処理・未マッチ")
    bold = Font(bold=True)
    head_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    r = 1
    ws.cell(row=r, column=1, value="● 解析できなかった請求勤怠ファイル（手順1で個別確認が必要）").font = bold
    r += 1
    ws.cell(row=r, column=1, value="ファイル名").fill = head_fill
    ws.cell(row=r, column=1).font = bold
    ws.cell(row=r, column=2, value="理由").fill = head_fill
    ws.cell(row=r, column=2).font = bold
    r += 1
    for name, reason in skipped:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=reason)
        r += 1
    if not skipped:
        ws.cell(row=r, column=1, value="（なし）")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="● 請求勤怠 未提出者（jinjer にはいるが請求勤怠が届いていない）").font = bold
    r += 1
    ws.cell(row=r, column=1, value="氏名").fill = head_fill
    ws.cell(row=r, column=1).font = bold
    r += 1
    for name in unsubmitted:
        ws.cell(row=r, column=1, value=name)
        r += 1
    if not unsubmitted:
        ws.cell(row=r, column=1, value="（なし）")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    wb.save(output_path)


def run_batch_compare(
    timesheet_dir: Path,
    jinjer_dir: Path,
    output_path: Path,
    month_label: str,
    application_csv: Path | None = None,
    threshold_minutes: int = 10,
    log_func=print,
):
    """請求勤怠フォルダ＋汎用データ＋申請データ → 突合→差異一覧 を一括実行する。

    Returns:
        (result: CompareResult, skipped: list[(name,reason)], unsubmitted: list[str])
        result.ok=False の場合は error にメッセージが入る。
    """
    logs: list[LogEntry] = []
    skipped: list[tuple[str, str]] = []
    unsubmitted: list[str] = []

    # 1. 請求勤怠の解析（フォルダ配下を再帰、形式自動判定）
    ts_files = gather_timesheet_files(Path(timesheet_dir))
    log_func(f"[batch] 請求勤怠ファイル {len(ts_files)} 件を解析します")
    dfs = []
    for f in ts_files:
        try:
            res = parse_timesheet_smart(str(f))
        except Exception as e:
            skipped.append((f.name, f"解析失敗: {e}"))
            log_func(f"[warn] {f.name} 解析失敗: {e}")
            continue
        mode = res.get("mode")
        if mode == "direct" and res.get("df") is not None and not res["df"].empty:
            dfs.append(res["df"])
            log_func(f"[batch] {f.name}: {len(res['df'])} 行")
        elif mode == "code":
            skipped.append((f.name, "記号式シフト表（凡例確認が必要なため一括対象外。手順1で個別処理してください）"))
        else:
            skipped.append((f.name, "出退勤データを抽出できませんでした"))

    if not dfs:
        return _fail("請求勤怠を1件も解析できませんでした", output_path, logs, month_label, skipped, unsubmitted, log_func)

    timesheet_df = pd.concat(dfs, ignore_index=True)

    # 2. jinjer 打刻（汎用データCSVの出勤1/退勤1 を流用）
    jinjer_csvs = [c for c in _input_files(Path(jinjer_dir), ["*.csv", "*.xlsx"]) if not c.name.startswith("~$")]
    jdfs = []
    for c in jinjer_csvs:
        try:
            jdfs.append(parse_jinjer_csv(str(c)))
        except Exception as e:
            log_func(f"[warn] jinjer 汎用データ解析失敗 {c.name}: {e}")
    if not jdfs:
        return _fail("jinjer 汎用データCSV を読めませんでした", output_path, logs, month_label, skipped, unsubmitted, log_func)
    jinjer_df = pd.concat(jdfs, ignore_index=True)

    # 3. 突合（手順1相当）
    result_df, unsubmitted = match(jinjer_df, timesheet_df, threshold_minutes)
    log_func(f"[batch] 突合 {len(result_df)} 行 / 未提出 {len(unsubmitted)} 名")

    # 4. 突合結果を一時フォルダへ出力 → 5. 手順2（差異一覧＋トリアージ）
    tmpdir = Path(tempfile.mkdtemp(prefix="batch_kintai_"))
    try:
        kintai_path = export_to_excel(
            result_df, threshold_minutes, output_folder=str(tmpdir), unsubmitted_names=unsubmitted
        )
        result = run_quick_compare(
            kintai_dir=Path(kintai_path),
            jinjer_dir=Path(jinjer_dir),
            output_path=Path(output_path),
            month_label=month_label,
            log_func=log_func,
            application_csv=Path(application_csv) if application_csv else None,
        )
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    # 6. 未処理・未マッチシートを追記
    if result.ok:
        _append_unprocessed_sheet(Path(output_path), skipped, unsubmitted)
    return result, skipped, unsubmitted


def _fail(msg, output_path, logs, month_label, skipped, unsubmitted, log_func):
    """請求勤怠/jinjer が読めない等の早期失敗。空の差異一覧＋未処理シートを出す。"""
    from quick_compare import write_excel, CompareResult
    log_func(f"[error] {msg}")
    logs.append(LogEntry("ERROR", msg))
    write_excel(Path(output_path), [], logs, month_label)
    _append_unprocessed_sheet(Path(output_path), skipped, unsubmitted)
    result = CompareResult(ok=False, output_path=Path(output_path), logs=logs, error=msg)
    return result, skipped, unsubmitted
