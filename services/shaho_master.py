r"""標準報酬月額チェックのマスタ読み込み（等級表Excel・料率・報酬分類CSV）。

等級表は2つの作りを読める（`load_grade_table` がシート名で見分ける）:

(A) Codex の月次公式資料: Z:\API連携\標準月額資料\標準報酬月額_YYYY_MM.xlsx
  Codex が毎月1日に協会けんぽ・日本年金機構の公式Excelを取得して置く（上書きしない）。
  - 都道府県シート（「東京」など47枚）: 協会けんぽの保険料額表そのまま。
    A列 等級「4(1)」= 健保4等級（厚年1等級）、B 月額、C 下限（円以上）、E 上限（円未満）、
    F/G 健康 全額/折半、H/I 健康＋介護、J/K 支援金、L/M 厚年。8行目の同じ列に料率。
  - 「厚生年金_公式表」: 厚年32等級（等級/標準報酬月額/下限/上限/全額/折半額）と料率。
  - 「都道府県料率一覧」: 支部ごとの料率（都道府県シートの8行目と突き合わせる）。
  - 「概要・出典」: 対象年月・確認日時・出典URL。
  関東IT健保（its）の健康・介護の料率は公式資料に無い（健保組合なので）。
  そこだけは (B) の「設定・出典」から取り、支援金・厚年の料率が公式資料と一致することを確かめる。
  対象年月に合う資料は `select_grade_table` が選ぶ。

(B) 手作りの比較ブック: Z:\API連携\標準月額資料\令和8年度_標準報酬月額表_関東IT健保_協会けんぽ東京比較.xlsx
  - シート「標準報酬月額表」: 健保等級1〜50。ヘッダは4行目、データは5行目〜。
    下限「以上」・上限「未満」の半開区間で、最終等級だけ上限が空欄（＝上限なし）。
  - シート「設定・出典」: 料率8行（区分/保険者/項目/全体料率/本人負担率/適用・備考/出典URL）
    ＋ I列以降に厚生年金の等級マスタ（1〜32、88,000〜650,000）。
  (A) が1つも無いときの等級表であり、(A) を使うときも ITS 料率の元として読む。

**マスタが壊れている限り1円も計算しない。** 列がズレたまま等級を引くと、全員の保険料が
静かに間違う。そのため読み込み時に総当たりで検証し、少しでも違えば ShahoMasterError を
投げる（health_hpm_master.py と同じ思想）:
  - 見出し名でパースする（列位置の決め打ちをしない。翌年度にExcelを差し替えても動く）
  - 等級の連番・下限昇順・半開区間の連続性（前の行の上限 = 次の行の下限）
  - 健保の各帯が厚年等級マスタと矛盾しないこと
  - 保険料額列 = 標準報酬月額 × 本人負担率 が全行で一致すること（±0.51円）
  - 年度表記（令和N年度）と実行側の対象年度の照合。ズレていたら停止

保険者は関東IT健保（its）が既定。協会けんぽ東京（kyokai_tokyo）は設定で切替できる
（実測: 控除 28,737÷620,000 = 4.635% = ITS本人負担率と完全一致。freee仕訳の取引先も関東IT健保）。
"""

from __future__ import annotations

import csv
import io
import os
import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl

SHEET_GRADES = "標準報酬月額表"
SHEET_RATES = "設定・出典"
GRADE_HEADER_ROW = 4
RATE_HEADER_ROW = 4

# 保険者キー → 等級表シートの保険料額列のプレフィックス
INSURERS = {"its": "ITS", "kyokai_tokyo": "協会東京"}
# 「設定・出典」の保険者名 → 保険者キー（「共通」は全保険者）
RATE_INSURER_NAMES = {"関東IT健保": "its", "協会けんぽ東京": "kyokai_tokyo", "共通": "共通"}
# 「設定・出典」の項目名 → 内部キー
RATE_ITEM_KEYS = {"健康保険": "kenpo", "子ども・子育て支援金": "kodomo",
                  "介護保険": "kaigo", "厚生年金": "konen", "厚生年金（参照）": "konen"}

REIWA_BASE_YEAR = 2018            # 令和N年 = 西暦 N + 2018
PREMIUM_TOLERANCE = 0.51          # 保険料額列の検証誤差（端数処理前の生floatのため）


class ShahoMasterError(ValueError):
    """マスタが読めない・壊れている。計算は必ず止める。"""


@dataclass(frozen=True)
class GradeRow:
    """等級表の1行。健保の帯（半開区間）と、対応する健保・厚年の標準報酬月額。"""
    kenpo_grade: int
    kenpo_smr: int                 # 健保 標準報酬月額（円）
    lower: int                     # 報酬月額下限（以上）
    upper: int | None              # 報酬月額上限（未満）。最終等級は None（上限なし）
    konen_grade: int
    konen_smr: int                 # 厚年 標準報酬月額（円）

    def contains(self, amount: float) -> bool:
        return amount >= self.lower and (self.upper is None or amount < self.upper)


@dataclass(frozen=True)
class Rate:
    """本人負担率（全体率の半分）。適用・備考と出典はレポートの「使用設定・出典」に出す。"""
    total: float                   # 全体料率
    employee: float                # 本人負担率
    applies: str = ""              # 適用・備考（例: 令和8年4月分～）
    source_url: str = ""


@dataclass
class ShahoMaster:
    path: str
    year: int                      # 対象年度（西暦。令和8年度 = 2026）
    insurer: str                   # its / kyokai_tokyo
    grades: list[GradeRow] = field(default_factory=list)
    rates: dict[str, Rate] = field(default_factory=dict)      # kenpo/kodomo/kaigo/konen
    konen_bands: list[tuple[int, int, int, int | None]] = field(default_factory=list)
    #                (厚年等級, 標準報酬月額, 下限, 上限None=∞)
    snapshot_ym: str | None = None # Codex 公式資料なら対象年月 "2026-09"。手作りブックは None
    description: str = ""          # レポート・画面向けの一言（何を元に計算したか）
    its_rates_path: str = ""       # ITS 料率を取った手作りブック（公式資料＋its のときだけ）

    def find_grade(self, amount: float) -> GradeRow:
        """報酬月額 → 等級行（下限以上・上限未満の半開区間）。"""
        if amount < 0:
            raise ShahoMasterError(f"報酬月額がマイナスです: {amount}")
        for row in self.grades:
            if row.contains(amount):
                return row
        raise ShahoMasterError(f"報酬月額 {amount} がどの等級にも入りません")


def _norm(v) -> str:
    return unicodedata.normalize("NFKC", str(v if v is not None else "")).strip()


def _header_map(ws, row: int, sheet: str, required: tuple, first_col: int = 1,
                last_col: int | None = None) -> dict:
    """見出し行 → {見出し名: 列番号}。必須見出しが無ければ ShahoMasterError。

    照合は NFKC 正規化どうしで行う（全角括弧と半角括弧の揺れを吸収する）。
    呼び出し側が書いたままの名前でも引けるよう、required の原文も別名で入れておく。
    """
    cols = {}
    for c in range(first_col, (last_col or ws.max_column) + 1):
        name = _norm(ws.cell(row, c).value)
        if name and name not in cols:
            cols[name] = c
    missing = [h for h in required if _norm(h) not in cols]
    if missing:
        raise ShahoMasterError(
            f"シート「{sheet}」の{row}行目に必須の見出しがありません: {'、'.join(missing)}"
            "（列名を変えた場合はコード側の対応も必要です）")
    for h in required:
        cols[h] = cols[_norm(h)]
    return cols


def _to_int(v, what: str):
    if v is None or _norm(v) == "":
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        raise ShahoMasterError(f"{what} が数値ではありません: {v!r}")
    if abs(f - round(f)) > 1e-6:
        raise ShahoMasterError(f"{what} が整数ではありません: {v!r}")
    return int(round(f))


def _parse_reiwa_year(text: str) -> int | None:
    """'令和8年度...' → 2026。見つからなければ None。"""
    t = _norm(text)
    if "令和" not in t:
        return None
    digits = ""
    for ch in t.split("令和", 1)[1]:
        if ch.isdigit():
            digits += ch
        else:
            break
    return (int(digits) + REIWA_BASE_YEAR) if digits else None


# ---------------------------------------------------------------------------
# 等級表の選択（Codex の月次公式資料 → 無ければ設定の手作りブック）
# ---------------------------------------------------------------------------
# Codex が毎月1日に取得して置く公式資料のファイル名（標準報酬月額_YYYY_MM.xlsx）。
# 既存を上書きしない運用なので、月が進むほどファイルが増える。
CODEX_TABLE_RE = re.compile(r"^標準報酬月額_(\d{4})_(\d{2})\.xlsx$")


def fiscal_year(ym: str) -> int:
    """保険料の年度（4月始まり）。'2026-03' → 2025、'2026-04' → 2026。"""
    return int(ym[:4]) if int(ym[5:7]) >= 4 else int(ym[:4]) - 1


@dataclass(frozen=True)
class TableChoice:
    path: str
    snapshot_ym: str | None        # Codex 公式資料なら "2026-09"。手作りブックなら None
    reason: str                    # なぜそれを選んだか（画面・ログ向け）


def list_snapshots(folder: str) -> list[tuple[str, str]]:
    """フォルダにある Codex 公式資料を [(対象年月 'YYYY-MM', パス)] で古い順に返す。"""
    try:
        names = os.listdir(folder)
    except OSError:
        return []
    found = []
    for name in names:
        m = CODEX_TABLE_RE.match(name)
        if m and 1 <= int(m.group(2)) <= 12:
            found.append((f"{m.group(1)}-{m.group(2)}", os.path.join(folder, name)))
    return sorted(found)


def select_grade_table(ym: str, configured_path: str) -> TableChoice:
    """対象年月 ym の計算に使う等級表を決める。

    1. 設定の等級表と同じフォルダにある Codex 公式資料のうち、対象年月以前で最新のもの
       （料率は「N月分から」変わるので、対象月の時点で最新の資料が正しい）
    2. 無ければ同じ年度で最も早いもの（年度初めの資料がまだ無く、後の月の資料だけある場合。
       年度内の料率は同じなので、その年度の資料なら使える）
    3. それも無ければ設定の等級表（手作りの比較ブック）

    年度が食い違う資料を掴んでも `load_grade_table` の年度チェックが止める。
    """
    snapshots = list_snapshots(os.path.dirname(configured_path))
    before = [s for s in snapshots if s[0] <= ym]
    if before:
        sym, path = before[-1]
        return TableChoice(path, sym, f"対象年月 {ym} 以前で最新の公式資料（{sym}）")
    same_fy = [s for s in snapshots if fiscal_year(s[0]) == fiscal_year(ym)]
    if same_fy:
        sym, path = same_fy[0]
        return TableChoice(path, sym,
                           f"対象年月 {ym} 以前の公式資料が無いため、同じ年度で最も早い公式資料（{sym}）")
    return TableChoice(configured_path, None, "Codex の公式資料が無いため、設定の等級表")


# ---------------------------------------------------------------------------
# 等級表の読み込み
# ---------------------------------------------------------------------------
def load_grade_table(path: str, insurer: str, expected_year: int,
                     its_rates_xlsx: str | None = None) -> ShahoMaster:
    """等級表Excelを読み、総当たり検証してから ShahoMaster を返す。

    シート名で作りを見分ける。Codex の公式資料（「厚生年金_公式表」がある）なら
    `_load_official_table`、それ以外は手作りの比較ブックとして読む。
    its_rates_xlsx は公式資料で its を使うときの ITS 料率の元（既定: 設定の手作りブック）。
    """
    if insurer not in INSURERS:
        raise ShahoMasterError(
            f"保険者 '{insurer}' は未対応です（使えるのは {' / '.join(INSURERS)}）")
    if not os.path.exists(path):
        raise ShahoMasterError(f"等級表Excelが見つかりません: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    if is_official_workbook(wb):
        return _load_official_table(wb, path, insurer, expected_year, its_rates_xlsx)
    for sheet in (SHEET_GRADES, SHEET_RATES):
        if sheet not in wb.sheetnames:
            raise ShahoMasterError(f"シート「{sheet}」がありません: {path}")

    ws = wb[SHEET_GRADES]

    # --- 年度検証（タイトル行の「令和N年度」と対象年度を突き合わせる） ---
    year = None
    for r in range(1, GRADE_HEADER_ROW):
        year = _parse_reiwa_year(ws.cell(r, 1).value)
        if year:
            break
    if year is None:
        raise ShahoMasterError(
            f"シート「{SHEET_GRADES}」の冒頭に年度表記（令和N年度）が見つかりません")
    if year != expected_year:
        raise ShahoMasterError(
            f"等級表の年度（令和{year - REIWA_BASE_YEAR}年度={year}）と対象年度（{expected_year}）が"
            "一致しません。翌年度の等級表に差し替えるか、対象年度を確認してください")

    prefix = INSURERS[insurer]
    required = ("健保等級", "健保標準報酬月額", "報酬月額下限（以上）", "報酬月額上限（未満）",
                "厚年等級", "厚年標準報酬月額", "厚生年金",
                f"{prefix}健康", f"{prefix}支援金", f"{prefix}介護")
    cols = _header_map(ws, GRADE_HEADER_ROW, SHEET_GRADES, required)

    grades: list[GradeRow] = []
    premiums: list[dict] = []          # 保険料額列の検証用（行と同順）
    for r in range(GRADE_HEADER_ROW + 1, ws.max_row + 1):
        g = _to_int(ws.cell(r, cols["健保等級"]).value, f"{r}行目の健保等級")
        if g is None:
            break                      # 空行が来たら終わり
        row = GradeRow(
            kenpo_grade=g,
            kenpo_smr=_to_int(ws.cell(r, cols["健保標準報酬月額"]).value, f"{r}行目の健保標準報酬月額"),
            lower=_to_int(ws.cell(r, cols["報酬月額下限（以上）"]).value, f"{r}行目の下限") or 0,
            upper=_to_int(ws.cell(r, cols["報酬月額上限（未満）"]).value, f"{r}行目の上限"),
            konen_grade=_to_int(ws.cell(r, cols["厚年等級"]).value, f"{r}行目の厚年等級"),
            konen_smr=_to_int(ws.cell(r, cols["厚年標準報酬月額"]).value, f"{r}行目の厚年標準報酬月額"),
        )
        grades.append(row)
        premiums.append({
            "kenpo": ws.cell(r, cols[f"{prefix}健康"]).value,
            "kodomo": ws.cell(r, cols[f"{prefix}支援金"]).value,
            "kaigo": ws.cell(r, cols[f"{prefix}介護"]).value,
            "konen": ws.cell(r, cols["厚生年金"]).value,
        })
    if not grades:
        raise ShahoMasterError(f"シート「{SHEET_GRADES}」にデータ行がありません")

    _check_kenpo_bands(grades)

    # --- 料率と厚年等級マスタ ---
    rates = _load_rates(wb, insurer)
    konen_bands = _load_konen_bands(wb)
    _check_konen_consistency(grades, konen_bands)

    # --- 保険料額列 = 標準報酬月額 × 本人負担率 の総当たり検証 ---
    for row, prem in zip(grades, premiums):
        checks = (("kenpo", row.kenpo_smr), ("kodomo", row.kenpo_smr),
                  ("kaigo", row.kenpo_smr), ("konen", row.konen_smr))
        for key, smr in checks:
            cell = prem[key]
            if cell is None:
                raise ShahoMasterError(
                    f"健保等級{row.kenpo_grade}: {key} の保険料額が空欄です")
            expected = smr * rates[key].employee
            if abs(float(cell) - expected) > PREMIUM_TOLERANCE:
                raise ShahoMasterError(
                    f"健保等級{row.kenpo_grade}: {key} の保険料額({cell})が "
                    f"標準報酬月額×本人負担率({expected:.2f})と合いません。"
                    "料率か表のどちらかが古い可能性があります")

    return ShahoMaster(path=path, year=year, insurer=insurer,
                       grades=grades, rates=rates, konen_bands=konen_bands,
                       description=f"手作りの比較ブック・令和{year - REIWA_BASE_YEAR}年度・検証済み")


def _load_rates(wb, insurer: str) -> dict[str, Rate]:
    """「設定・出典」の料率ブロックから、指定保険者＋共通の4本を取る。"""
    ws = wb[SHEET_RATES]
    cols = _header_map(ws, RATE_HEADER_ROW, SHEET_RATES,
                       ("区分", "保険者", "項目", "全体料率", "本人負担率"),
                       first_col=1, last_col=8)
    url_col = cols.get("出典URL")
    biko_col = cols.get("適用・備考")
    # 照合は NFKC 正規化どうし（全角括弧の揺れを吸収）
    item_keys = {_norm(k): v for k, v in RATE_ITEM_KEYS.items()}
    ref_item = _norm("厚生年金（参照）")
    rates: dict[str, Rate] = {}
    for r in range(RATE_HEADER_ROW + 1, ws.max_row + 1):
        ins = _norm(ws.cell(r, cols["保険者"]).value)
        item = _norm(ws.cell(r, cols["項目"]).value)
        if not ins and not item:
            break                      # 料率ブロックの終わり（空行）
        ins_key = RATE_INSURER_NAMES.get(ins)
        item_key = item_keys.get(item)
        if ins_key is None or item_key is None:
            raise ShahoMasterError(
                f"「{SHEET_RATES}」{r}行目: 保険者「{ins}」項目「{item}」を解釈できません")
        if ins_key not in ("共通", insurer):
            continue
        if item_key in rates and item != ref_item:
            raise ShahoMasterError(f"「{SHEET_RATES}」: {item} の料率が二重に定義されています")
        total = ws.cell(r, cols["全体料率"]).value
        emp = ws.cell(r, cols["本人負担率"]).value
        if not isinstance(total, (int, float)) or not isinstance(emp, (int, float)):
            raise ShahoMasterError(f"「{SHEET_RATES}」{r}行目: 料率が数値ではありません")
        if not 0 < emp <= total <= 0.5:
            raise ShahoMasterError(
                f"「{SHEET_RATES}」{r}行目: 料率が不自然です（全体{total}・本人{emp}）")
        if abs(emp - total / 2) > 1e-9:
            raise ShahoMasterError(
                f"「{SHEET_RATES}」{r}行目: 本人負担率({emp})が全体率({total})の半分ではありません")
        if item_key not in rates:
            rates[item_key] = Rate(
                total=float(total), employee=float(emp),
                applies=_norm(ws.cell(r, biko_col).value) if biko_col else "",
                source_url=_norm(ws.cell(r, url_col).value) if url_col else "")
    missing = [k for k in ("kenpo", "kodomo", "kaigo", "konen") if k not in rates]
    if missing:
        raise ShahoMasterError(
            f"「{SHEET_RATES}」に保険者の料率が揃っていません（不足: {'、'.join(missing)}）")
    return rates


def _load_konen_bands(wb) -> list[tuple[int, int, int, int | None]]:
    """「設定・出典」I列以降の厚生年金 等級マスタ（1〜32）。"""
    ws = wb[SHEET_RATES]
    cols = _header_map(ws, RATE_HEADER_ROW, SHEET_RATES,
                       ("厚年等級", "標準報酬月額", "報酬月額下限（以上）", "報酬月額上限（未満）"),
                       first_col=9)
    bands = []
    for r in range(RATE_HEADER_ROW + 1, ws.max_row + 1):
        g = _to_int(ws.cell(r, cols["厚年等級"]).value, f"厚年等級マスタ{r}行目の等級")
        if g is None:
            break
        bands.append((
            g,
            _to_int(ws.cell(r, cols["標準報酬月額"]).value, f"厚年等級{g}の標準報酬月額"),
            _to_int(ws.cell(r, cols["報酬月額下限（以上）"]).value, f"厚年等級{g}の下限") or 0,
            _to_int(ws.cell(r, cols["報酬月額上限（未満）"]).value, f"厚年等級{g}の上限"),
        ))
    if not bands:
        raise ShahoMasterError("厚生年金の等級マスタが読めません")
    for i, (g, _smr, lo, up) in enumerate(bands):
        if g != i + 1:
            raise ShahoMasterError(f"厚年等級が連番ではありません: {i + 1} のはずが {g}")
        if i > 0 and bands[i - 1][3] != lo:
            raise ShahoMasterError(f"厚年等級{g}: 前の等級と区間が繋がっていません")
    if bands[-1][3] is not None:
        raise ShahoMasterError("厚年の最終等級の上限は空欄（上限なし）のはずです")
    return bands


def _check_kenpo_bands(grades: list[GradeRow]) -> None:
    """健保等級の連番・半開区間の連続性・最終等級だけ上限なし、を確かめる。"""
    for i, row in enumerate(grades):
        if row.kenpo_grade != i + 1:
            raise ShahoMasterError(
                f"健保等級が連番ではありません: {i + 1} のはずが {row.kenpo_grade}")
        if row.upper is not None and row.upper <= row.lower:
            raise ShahoMasterError(
                f"健保等級{row.kenpo_grade}: 上限({row.upper}) ≤ 下限({row.lower}) です")
        if i > 0 and grades[i - 1].upper != row.lower:
            raise ShahoMasterError(
                f"健保等級{row.kenpo_grade}: 前の等級の上限({grades[i - 1].upper})と"
                f"下限({row.lower})が繋がっていません")
    if grades[0].lower != 0:
        raise ShahoMasterError(f"最初の等級の下限が0ではありません: {grades[0].lower}")
    if grades[-1].upper is not None:
        raise ShahoMasterError("最終等級の上限は空欄（上限なし）のはずです")
    for row in grades[:-1]:
        if row.upper is None:
            raise ShahoMasterError(
                f"健保等級{row.kenpo_grade}: 最終等級以外で上限が空欄です")


def _konen_band_at(konen_bands: list, amount: float) -> tuple[int, int]:
    """報酬月額 → (厚年等級, 厚年標準報酬月額)。"""
    for g, smr, lo, up in konen_bands:
        if amount >= lo and (up is None or amount < up):
            return g, smr
    raise ShahoMasterError(f"報酬月額 {amount} が厚年等級マスタのどの帯にも入りません")


def _check_konen_consistency(grades: list[GradeRow], konen_bands: list) -> None:
    """健保の各帯（下限と上限-1）が、厚年等級マスタの標準報酬月額と矛盾しないこと。"""
    for row in grades:
        probes = [row.lower] if row.upper is None else [row.lower, row.upper - 1]
        for p in probes:
            _g, expected = _konen_band_at(konen_bands, p)
            if expected != row.konen_smr:
                raise ShahoMasterError(
                    f"健保等級{row.kenpo_grade}（報酬月額{p}）の厚年標準報酬月額が"
                    f"厚年等級マスタと食い違います: 表={row.konen_smr} マスタ={expected}")


def _check_konen_bands(bands: list, sheet: str) -> None:
    """厚年等級の連番・区間の連続性・最終等級だけ上限なし。"""
    if not bands:
        raise ShahoMasterError(f"「{sheet}」: 厚生年金の等級マスタが読めません")
    for i, (g, _smr, lo, up) in enumerate(bands):
        if g != i + 1:
            raise ShahoMasterError(f"「{sheet}」: 厚年等級が連番ではありません: {i + 1} のはずが {g}")
        if i > 0 and bands[i - 1][3] != lo:
            raise ShahoMasterError(f"「{sheet}」: 厚年等級{g}: 前の等級と区間が繋がっていません")
    if bands[-1][3] is not None:
        raise ShahoMasterError(f"「{sheet}」: 厚年の最終等級の上限は空欄（上限なし）のはずです")


# ---------------------------------------------------------------------------
# Codex 公式資料（協会けんぽ保険料額表 ＋ 日本年金機構 厚生年金保険料額表）の読み込み
# ---------------------------------------------------------------------------
OFFICIAL_SHEET_KONEN = "厚生年金_公式表"
OFFICIAL_SHEET_RATES = "都道府県料率一覧"
OFFICIAL_SHEET_ABOUT = "概要・出典"
# 健保の等級の帯は全国共通（法定）なので、帯はこのシートから取る。
# 保険者ごとに違うのは健康・介護の料率だけ。
OFFICIAL_BAND_SHEET = "東京"
# 協会けんぽの保険者キー → 支部シート名（健康・介護の料率をそのシートの料率行から取る）
INSURER_PREFECTURE = {"kyokai_tokyo": "東京"}
_GRADE_PAIR_RE = re.compile(r"^(\d+)\((\d+)\)$")      # 「4(1)」= 健保4等級・厚年1等級
RATE_MATCH_TOLERANCE = 1e-9                             # 同じ資料内の料率の突合（float の誤差だけ許す）


def is_official_workbook(wb) -> bool:
    """Codex の公式資料か（「厚生年金_公式表」があり、手作りブックの「標準報酬月額表」が無い）。"""
    return OFFICIAL_SHEET_KONEN in wb.sheetnames and SHEET_GRADES not in wb.sheetnames


def _is_official_file(path: str) -> bool:
    """シート名だけ見て Codex の公式資料か判定する（read_only は閉じないとファイルを掴み続ける）。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return is_official_workbook(wb)
    finally:
        wb.close()


def _compact(v) -> str:
    """NFKC 正規化して空白を全部落とす（「全  額」「報酬月額 下限」の揺れを吸収）。"""
    return _norm(v).replace(" ", "")


def _find_cell(ws, needle: str, rows, exact: bool = False):
    """rows の範囲で needle を含む（exact なら一致する）最初のセルの (行, 列)。無ければ None。"""
    for r in rows:
        for c in range(1, ws.max_column + 1):
            t = _compact(ws.cell(r, c).value)
            if t and ((t == needle) if exact else (needle in t)):
                return r, c
    return None


def _official_year(wb) -> int:
    """「都道府県料率一覧」（無ければ「厚生年金_公式表」）のタイトルの 令和N年度 → 西暦。"""
    for sheet in (OFFICIAL_SHEET_RATES, OFFICIAL_SHEET_KONEN):
        if sheet in wb.sheetnames:
            y = _parse_reiwa_year(wb[sheet].cell(1, 1).value)
            if y:
                return y
    raise ShahoMasterError(
        f"公式資料のシート「{OFFICIAL_SHEET_RATES}」「{OFFICIAL_SHEET_KONEN}」の"
        "冒頭に年度表記（令和N年度）が見つかりません")


def _official_about(wb) -> dict:
    """「概要・出典」→ {snapshot_ym: 'YYYY-MM' | None, checked_at, urls: {機関: 公式ページURL}}。"""
    info = {"snapshot_ym": None, "checked_at": "", "urls": {}}
    if OFFICIAL_SHEET_ABOUT not in wb.sheetnames:
        return info
    ws = wb[OFFICIAL_SHEET_ABOUT]
    for r in range(1, ws.max_row + 1):
        key = _compact(ws.cell(r, 1).value)
        val = _norm(ws.cell(r, 2).value)
        if key == "対象年月":
            m = re.fullmatch(r"(\d{4})年(\d{1,2})月", val)
            if m:
                info["snapshot_ym"] = f"{m.group(1)}-{int(m.group(2)):02d}"
        elif key.startswith("確認日時"):
            info["checked_at"] = val
        elif key in ("全国健康保険協会", "日本年金機構") and val.startswith("http"):
            info["urls"][key] = val
    return info


def _read_prefecture_sheet(ws, sheet: str) -> dict:
    """協会けんぽの都道府県シート（保険料額表そのまま）を読む。

    戻り値:
      rows     : [(健保等級, 括弧内の厚年等級|None, 健保標準報酬月額, 下限, 上限|None)]
      premiums : 行ごとの折半額 {kenpo, kenpo_kaigo, kodomo, konen}（厚年の外側の等級は None）
      rates    : 全体料率 {kenpo, kaigo, kenpo_kaigo, kodomo, konen}（料率行の値）
      applies  : 「令和8年3月分～適用」などの適用表記 {kenpo, kaigo, kodomo, konen}
    列は見出しの文言で探す（協会けんぽの様式が変わったら止まる）。
    """
    def fail(msg):
        raise ShahoMasterError(f"公式資料のシート「{sheet}」: {msg}")

    pos = _find_cell(ws, "等級", range(1, 16), exact=True)
    if pos is None:
        fail("見出し「等級」が見つかりません")
    rate_row = pos[0]                          # 「等級」と同じ行の各列に料率が入っている
    head_rows = range(max(1, rate_row - 3), rate_row)

    def group_col(needle: str, what: str) -> int:
        p = _find_cell(ws, needle, head_rows)
        if p is None:
            fail(f"見出し「{what}」が見つかりません")
        return p[1]

    groups = {
        "kenpo": group_col("管掌健康保険料", "全国健康保険協会管掌健康保険料"),
        "kenpo_kaigo": group_col("該当する場合", "介護保険第2号被保険者に該当する場合"),
        "kodomo": group_col("子ども・子育て支援金", "子ども・子育て支援金"),
        "konen": group_col("厚生年金保険料", "厚生年金保険料"),
    }
    p = _find_cell(ws, "該当しない場合", head_rows)
    if p is None or p[1] != groups["kenpo"]:
        fail("「介護保険第2号被保険者に該当しない場合」の列が健康保険料の列と揃っていません")
    for key, c in groups.items():
        if (_compact(ws.cell(rate_row + 1, c).value) != "全額"
                or _compact(ws.cell(rate_row + 1, c + 1).value) != "折半額"):
            fail(f"{key} の列（{c}列目）の下に「全額」「折半額」の見出しがありません")
    p = _find_cell(ws, "月額", [rate_row], exact=True)
    if p is None:
        fail(f"{rate_row}行目に見出し「月額」がありません")
    col_smr = p[1]
    unit_row = rate_row + 2
    lo = _find_cell(ws, "円以上", [unit_row], exact=True)
    up = _find_cell(ws, "円未満", [unit_row], exact=True)
    if lo is None or up is None:
        fail(f"{unit_row}行目に「円以上」「円未満」の見出しがありません")
    col_lower, col_upper = lo[1], up[1]

    rates = {}
    for key, c in groups.items():
        v = ws.cell(rate_row, c).value
        if not isinstance(v, (int, float)) or not 0 < v < 0.5:
            fail(f"{key} の料率（{rate_row}行目{c}列目）が読めません: {v!r}")
        rates[key] = round(float(v), 6)        # 0.09849999999999999 のような float の尻尾を落とす
    if rates["kenpo_kaigo"] <= rates["kenpo"]:
        fail("健康＋介護の料率が健康保険の料率以下です")
    rates["kaigo"] = round(rates["kenpo_kaigo"] - rates["kenpo"], 6)

    applies = {}
    for key, needle in (("kenpo", "健康保険料率"), ("kaigo", "介護保険料率"),
                        ("kodomo", "支援金率"), ("konen", "厚生年金保険料率")):
        p = _find_cell(ws, needle, range(1, rate_row))
        if p:
            text = _norm(ws.cell(p[0], p[1]).value)
            parts = re.split(r"[:：]", text, maxsplit=1)
            applies[key] = (parts[1] if len(parts) == 2 else text).strip()

    rows, premiums = [], []
    for r in range(unit_row + 1, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None or _norm(raw) == "":
            break                              # 表の終わり（この下は注記）
        m = _GRADE_PAIR_RE.match(_compact(raw))
        if m:
            g, kg = int(m.group(1)), int(m.group(2))
        else:
            g, kg = _to_int(raw, f"{sheet} {r}行目の等級"), None
        rows.append((g, kg,
                     _to_int(ws.cell(r, col_smr).value, f"{sheet} {r}行目の標準報酬月額"),
                     _to_int(ws.cell(r, col_lower).value, f"{sheet} {r}行目の下限") or 0,
                     _to_int(ws.cell(r, col_upper).value, f"{sheet} {r}行目の上限")))
        premiums.append({key: ws.cell(r, c + 1).value for key, c in groups.items()})
    if not rows:
        fail("等級の行がありません")
    return {"rows": rows, "premiums": premiums, "rates": rates, "applies": applies}


def _read_official_konen(wb) -> tuple[list, float, list]:
    """「厚生年金_公式表」→ (等級マスタ [(等級, 標報, 下限, 上限)], 全体料率, 行ごとの折半額)。"""
    ws = wb[OFFICIAL_SHEET_KONEN]

    def fail(msg):
        raise ShahoMasterError(f"公式資料のシート「{OFFICIAL_SHEET_KONEN}」: {msg}")

    pos = _find_cell(ws, "等級", range(1, 12), exact=True)
    if pos is None:
        fail("見出し「等級」が見つかりません")
    hr = pos[0]
    cols = {}
    for c in range(1, ws.max_column + 1):
        t = _compact(ws.cell(hr, c).value)
        if t and t not in cols:
            cols[t] = c
    missing = [h for h in ("等級", "標準報酬月額", "報酬月額下限", "報酬月額上限", "保険料折半額")
               if h not in cols]
    if missing:
        fail(f"{hr}行目に必須の見出しがありません: {'、'.join(missing)}")
    p = _find_cell(ws, "全額保険料率", range(1, hr))
    if p is None:
        fail("「全額保険料率」が見つかりません")
    pct = ws.cell(p[0] + 1, p[1]).value           # 見出しの真下に % 表記（18.3）
    if not isinstance(pct, (int, float)) or not 0 < pct < 50:
        fail(f"全額保険料率が読めません: {pct!r}")
    rate = round(float(pct) / 100, 6)

    bands, halves = [], []
    for r in range(hr + 1, ws.max_row + 1):
        g = _to_int(ws.cell(r, cols["等級"]).value, f"厚年等級マスタ{r}行目の等級")
        if g is None:
            break
        bands.append((
            g,
            _to_int(ws.cell(r, cols["標準報酬月額"]).value, f"厚年等級{g}の標準報酬月額"),
            _to_int(ws.cell(r, cols["報酬月額下限"]).value, f"厚年等級{g}の下限") or 0,
            _to_int(ws.cell(r, cols["報酬月額上限"]).value, f"厚年等級{g}の上限"),
        ))
        halves.append(ws.cell(r, cols["保険料折半額"]).value)
    _check_konen_bands(bands, OFFICIAL_SHEET_KONEN)
    return bands, rate, halves


def _official_listed_rates(wb, prefecture: str) -> dict | None:
    """「都道府県料率一覧」の該当支部の行 → {kenpo, kenpo_kaigo, kodomo, konen}。シートが無ければ None。"""
    if OFFICIAL_SHEET_RATES not in wb.sheetnames:
        return None
    ws = wb[OFFICIAL_SHEET_RATES]
    head = _find_cell(ws, "都道府県", range(1, 10), exact=True)
    if head is None:
        return None
    hr = head[0]
    cols = {}
    for c in range(1, ws.max_column + 1):
        t = _compact(ws.cell(hr, c).value)
        if t:
            cols[t] = c

    def col_of(needle: str):
        return next((c for t, c in cols.items() if needle in t), None)

    want = {"kenpo": col_of("介護非該当"), "kenpo_kaigo": col_of("介護該当"),
            "kodomo": col_of("支援金"), "konen": col_of("厚生年金")}
    if any(c is None for c in want.values()):
        raise ShahoMasterError(f"「{OFFICIAL_SHEET_RATES}」の見出しを解釈できません")
    for r in range(hr + 1, ws.max_row + 1):
        if _compact(ws.cell(r, 1).value) == prefecture:
            return {k: ws.cell(r, c).value for k, c in want.items()}
    raise ShahoMasterError(f"「{OFFICIAL_SHEET_RATES}」に {prefecture} の行がありません")


def _load_official_table(wb, path: str, insurer: str, expected_year: int,
                         its_rates_xlsx: str | None) -> ShahoMaster:
    """Codex の公式資料を読み、総当たり検証してから ShahoMaster を返す。

    検証:
      - 年度（都道府県料率一覧のタイトル）と対象年度の照合。ファイル名の年月と「概要・出典」の対象年月の照合
      - 健保50等級の連番・半開区間の連続性、厚年32等級の連番・連続性
      - 健保の各帯と厚年等級マスタの整合、等級欄の「4(1)」の括弧内が厚年表と一致すること
      - 都道府県シートの折半額列 = 標準報酬月額 × 全体率 ÷ 2（±0.51円）を全行・全列で
      - 厚生年金_公式表の折半額 = 標報 × 率 ÷ 2、その率が都道府県シートの厚年率と一致
      - 都道府県シートの料率行と「都道府県料率一覧」の一致
    its のときは健康・介護の料率を手作りブック（its_rates_xlsx）の「設定・出典」から取り、
    そのブック自身の検証（年度含む）も通す。支援金・厚年の率が公式資料と違えば止める。
    """
    year = _official_year(wb)
    if year != expected_year:
        raise ShahoMasterError(
            f"公式資料の年度（令和{year - REIWA_BASE_YEAR}年度={year}）と対象年度（{expected_year}）が"
            f"一致しません: {path}")
    about = _official_about(wb)
    m = CODEX_TABLE_RE.match(os.path.basename(path))
    file_ym = f"{m.group(1)}-{m.group(2)}" if m else None
    if about["snapshot_ym"] and file_ym and about["snapshot_ym"] != file_ym:
        raise ShahoMasterError(
            f"公式資料のファイル名の年月（{file_ym}）と「{OFFICIAL_SHEET_ABOUT}」の対象年月"
            f"（{about['snapshot_ym']}）が食い違います: {path}")
    snapshot_ym = about["snapshot_ym"] or file_ym or "?"

    band_sheet = INSURER_PREFECTURE.get(insurer, OFFICIAL_BAND_SHEET)
    if band_sheet not in wb.sheetnames:
        raise ShahoMasterError(f"公式資料にシート「{band_sheet}」がありません: {path}")
    pref = _read_prefecture_sheet(wb[band_sheet], band_sheet)
    konen_bands, konen_rate, konen_halves = _read_official_konen(wb)

    # --- 厚生年金_公式表: 折半額 = 標報 × 率 ÷ 2。率は都道府県シートの厚年率と同じこと ---
    for (g, smr, _lo, _up), half in zip(konen_bands, konen_halves):
        if half is None or abs(float(half) - smr * konen_rate / 2) > PREMIUM_TOLERANCE:
            raise ShahoMasterError(
                f"「{OFFICIAL_SHEET_KONEN}」厚年等級{g}: 折半額({half})が "
                f"標準報酬月額×率÷2({smr * konen_rate / 2:.2f})と合いません")
    if abs(konen_rate - pref["rates"]["konen"]) > RATE_MATCH_TOLERANCE:
        raise ShahoMasterError(
            f"厚生年金の料率が「{band_sheet}」({pref['rates']['konen']})と"
            f"「{OFFICIAL_SHEET_KONEN}」({konen_rate})で違います")

    # --- 都道府県シートの料率行 = 都道府県料率一覧 ---
    listed = _official_listed_rates(wb, band_sheet)
    if listed:
        for key in ("kenpo", "kenpo_kaigo", "kodomo", "konen"):
            v = listed[key]
            if not isinstance(v, (int, float)) or abs(float(v) - pref["rates"][key]) > RATE_MATCH_TOLERANCE:
                raise ShahoMasterError(
                    f"{key} の料率が「{band_sheet}」({pref['rates'][key]})と"
                    f"「{OFFICIAL_SHEET_RATES}」({v!r})で違います")

    # --- 健保の帯 → 厚年等級を厚年表から引く。括弧内の等級とも突き合わせる ---
    grades: list[GradeRow] = []
    for g, kg, smr, lower, upper in pref["rows"]:
        konen_grade, konen_smr = _konen_band_at(konen_bands, lower)
        if kg is not None and kg != konen_grade:
            raise ShahoMasterError(
                f"「{band_sheet}」健保等級{g}: 等級欄の括弧内の厚年等級({kg})が"
                f"「{OFFICIAL_SHEET_KONEN}」から引いた等級({konen_grade})と食い違います")
        grades.append(GradeRow(kenpo_grade=g, kenpo_smr=smr, lower=lower, upper=upper,
                               konen_grade=konen_grade, konen_smr=konen_smr))
    _check_kenpo_bands(grades)
    _check_konen_consistency(grades, konen_bands)

    # --- 折半額列 = 標準報酬月額 × 全体率 ÷ 2 の総当たり（資料そのものの料率で） ---
    r = pref["rates"]
    first_konen, last_konen = konen_bands[0][0], konen_bands[-1][0]
    for row, prem in zip(grades, pref["premiums"]):
        checks = (("kenpo", row.kenpo_smr * r["kenpo"] / 2),
                  ("kenpo_kaigo", row.kenpo_smr * r["kenpo_kaigo"] / 2),
                  ("kodomo", row.kenpo_smr * r["kodomo"] / 2))
        for key, expected in checks:
            cell = prem[key]
            if cell is None or abs(float(cell) - expected) > PREMIUM_TOLERANCE:
                raise ShahoMasterError(
                    f"「{band_sheet}」健保等級{row.kenpo_grade}: {key} の折半額({cell})が "
                    f"標準報酬月額×率÷2({expected:.2f})と合いません")
        cell = prem["konen"]
        if cell is None:
            # 厚年の等級幅の外側（健保1〜3等級・36等級以上）は厚年の額が空欄
            if row.konen_grade not in (first_konen, last_konen):
                raise ShahoMasterError(
                    f"「{band_sheet}」健保等級{row.kenpo_grade}: 厚生年金の折半額が空欄です")
        else:
            expected = row.konen_smr * konen_rate / 2
            if abs(float(cell) - expected) > PREMIUM_TOLERANCE:
                raise ShahoMasterError(
                    f"「{band_sheet}」健保等級{row.kenpo_grade}: 厚生年金の折半額({cell})が "
                    f"標準報酬月額×率÷2({expected:.2f})と合いません")

    # --- 料率の確定 ---
    kyokai_url = about["urls"].get("全国健康保険協会", "")
    nenkin_url = about["urls"].get("日本年金機構", "")
    applies = pref["applies"]
    common = {
        "kodomo": Rate(r["kodomo"], r["kodomo"] / 2, applies.get("kodomo", ""), kyokai_url),
        "konen": Rate(konen_rate, konen_rate / 2, applies.get("konen", ""), nenkin_url),
    }
    reiwa = f"令和{year - REIWA_BASE_YEAR}年度"
    if insurer in INSURER_PREFECTURE:
        rates = {"kenpo": Rate(r["kenpo"], r["kenpo"] / 2, applies.get("kenpo", ""), kyokai_url),
                 "kaigo": Rate(r["kaigo"], r["kaigo"] / 2, applies.get("kaigo", ""), kyokai_url),
                 **common}
        return ShahoMaster(
            path=path, year=year, insurer=insurer, grades=grades, rates=rates,
            konen_bands=konen_bands, snapshot_ym=snapshot_ym,
            description=f"Codex公式資料 {snapshot_ym}（協会けんぽ{band_sheet}支部・{reiwa}・検証済み）")

    # 関東IT健保: 健康・介護の料率は公式資料に無いので手作りブックの「設定・出典」から。
    # そのブック自身の検証（年度・保険料額列）も通すので、年度が変わって古いままなら止まる。
    if its_rates_xlsx is None:
        from config import Config
        its_rates_xlsx = Config.SHAHO_GRADE_TABLE_XLSX
    if not its_rates_xlsx or not os.path.exists(its_rates_xlsx):
        raise ShahoMasterError(
            "関東IT健保の健康・介護の料率は公式資料に載っていないため、料率を書いた手作りブック"
            f"（「{SHEET_RATES}」シート）が必要です。見つかりません: {its_rates_xlsx}")
    if os.path.abspath(its_rates_xlsx) == os.path.abspath(path) or _is_official_file(its_rates_xlsx):
        raise ShahoMasterError(
            f"関東IT健保の料率の元に公式資料は使えません（「{SHEET_RATES}」シートが無い）: {its_rates_xlsx}")
    try:
        its_master = load_grade_table(its_rates_xlsx, "its", expected_year)
    except ShahoMasterError as e:
        raise ShahoMasterError(
            f"関東IT健保の料率の元（{os.path.basename(its_rates_xlsx)}）が読めません: {e}\n"
            "手作りブックの「設定・出典」の料率と年度を最新にしてください") from e
    for key, label in (("kodomo", "子ども・子育て支援金"), ("konen", "厚生年金")):
        if abs(its_master.rates[key].total - common[key].total) > RATE_MATCH_TOLERANCE:
            raise ShahoMasterError(
                f"{os.path.basename(its_rates_xlsx)} の「{SHEET_RATES}」の {label} の全体料率"
                f"({its_master.rates[key].total})が公式資料({common[key].total})と違います。"
                "手作りブックの料率が古い可能性があります（関東IT健保の健康・介護の料率も"
                "改定されていないか確認してください）")
    rates = {"kenpo": its_master.rates["kenpo"], "kaigo": its_master.rates["kaigo"], **common}
    return ShahoMaster(
        path=path, year=year, insurer=insurer, grades=grades, rates=rates,
        konen_bands=konen_bands, snapshot_ym=snapshot_ym, its_rates_path=its_rates_xlsx,
        description=(f"Codex公式資料 {snapshot_ym}（協会けんぽ・日本年金機構・{reiwa}・検証済み）"
                     f"＋関東IT健保の料率は {os.path.basename(its_rates_xlsx)}"))


# ---------------------------------------------------------------------------
# 報酬分類マスタ（CSV）
# ---------------------------------------------------------------------------
# 適用開始月・適用終了月（YYYY-MM・両端含む・空=無期限）を持つのは、**同じ項目・同じ
# 体系別名でも月によって意味が変わる**実例があるため。時給制の「みなし給」(allowance2) は
# 2026-05〜07 支給分では差額調整(a24)への再掲（総支給に乗らない情報項目）だが、
# 2026-08 支給分からは実額の支給になった（経理モードの ZANTEI_LABELS_FROM と同じ事象）。
CLASS_MASTER_COLS = ["source_key", "salary_system_label", "label", "class", "fixed",
                     "適用開始月", "適用終了月", "note"]
# 対象   = 通貨の報酬（検算ゲートの基準＝雇用保険対象額に入っている）
# 現物   = 現物給与。報酬には数えるが、現金の総支給額の**外側**なので検算からは除く
#          （実測: allowance53 は other5 に不算入）
# 対象外 = 報酬に数えない（実費弁償・再掲・情報項目）
# 未設定 = 判断待ち。金額が出たらその人は要確認へ落ちる
CLASS_VALUES = ("対象", "現物", "対象外", "未設定")


@dataclass(frozen=True)
class ClassRule:
    source_key: str
    salary_system_label: str       # 正規化済み。空=全体系共通
    label: str
    cls: str                       # 対象 / 対象外 / 未設定
    fixed: str                     # "1"=固定的賃金 / "0"=非固定 / ""=未設定
    ym_from: str = ""              # 適用開始月 YYYY-MM（空=最初から）
    ym_to: str = ""                # 適用終了月 YYYY-MM（空=無期限）
    note: str = ""

    def covers(self, ym: str) -> bool:
        return ((not self.ym_from or self.ym_from <= ym)
                and (not self.ym_to or ym <= self.ym_to))


def load_class_master(path: str) -> dict:
    """報酬分類マスタを読む。

    戻り値 dict:
      lookup  : {(source_key, 正規化体系別名): ClassRule}（体系別名空の行は ("key","")）
      rules   : 全行
      path    : 読んだパス
    分類の解決は (source_key, 体系別名) 完全一致 → (source_key, 空) の順で
    resolve_class() を使う。
    """
    if not os.path.exists(path):
        raise ShahoMasterError(
            f"報酬分類マスタが見つかりません: {path}\n"
            "先に tools/build_shaho_class_master.py で初期版を作ってください")
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "cp932", "utf-8"):
        try:
            rows = list(csv.DictReader(io.StringIO(raw.decode(enc))))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ShahoMasterError(f"報酬分類マスタの文字コードを判別できません: {path}")
    if not rows:
        raise ShahoMasterError(f"報酬分類マスタが空です: {path}")
    missing = [c for c in CLASS_MASTER_COLS if c not in rows[0]]
    if missing:
        raise ShahoMasterError(
            f"報酬分類マスタに列が足りません: {'、'.join(missing)}（{path}）")

    def norm_ym(v, i, what):
        t = _norm(v)
        if t and (len(t) != 7 or t[4] != "-" or not (t[:4] + t[5:]).isdigit()):
            raise ShahoMasterError(
                f"報酬分類マスタ{i}行目: {what} '{t}' は YYYY-MM 形式で書いてください")
        return t

    lookup: dict[tuple[str, str], list[ClassRule]] = {}
    rules: list[ClassRule] = []
    for i, r in enumerate(rows, start=2):
        key = _norm(r.get("source_key"))
        if not key or key.startswith("#"):
            continue
        if ":" not in key:
            raise ShahoMasterError(
                f"報酬分類マスタ{i}行目: source_key '{key}' は "
                "'salary_items:allowance1' の形式で書いてください")
        cls = _norm(r.get("class"))
        if cls not in CLASS_VALUES:
            raise ShahoMasterError(
                f"報酬分類マスタ{i}行目: class '{cls}' は "
                f"{' / '.join(CLASS_VALUES)} のどれかにしてください")
        fixed = _norm(r.get("fixed"))
        if fixed not in ("", "0", "1"):
            raise ShahoMasterError(
                f"報酬分類マスタ{i}行目: fixed '{fixed}' は 1 / 0 / 空 のどれかです")
        rule = ClassRule(
            source_key=key,
            salary_system_label=_norm(r.get("salary_system_label")).replace(" ", ""),
            label=_norm(r.get("label")),
            cls=cls, fixed=fixed,
            ym_from=norm_ym(r.get("適用開始月"), i, "適用開始月"),
            ym_to=norm_ym(r.get("適用終了月"), i, "適用終了月"),
            note=_norm(r.get("note")))
        if rule.ym_from and rule.ym_to and rule.ym_from > rule.ym_to:
            raise ShahoMasterError(
                f"報酬分類マスタ{i}行目: 適用開始月({rule.ym_from})が"
                f"終了月({rule.ym_to})より後です")
        pair = (rule.source_key, rule.salary_system_label)
        for other in lookup.get(pair, []):
            lo1, hi1 = rule.ym_from or "0000-00", rule.ym_to or "9999-99"
            lo2, hi2 = other.ym_from or "0000-00", other.ym_to or "9999-99"
            if lo1 <= hi2 and lo2 <= hi1:
                raise ShahoMasterError(
                    f"報酬分類マスタ{i}行目: {pair} の適用期間が別の行と重なっています")
        lookup.setdefault(pair, []).append(rule)
        rules.append(rule)
    return {"lookup": lookup, "rules": rules, "path": path}


def resolve_class(master: dict, source_key: str, system_label: str,
                  ym: str = "9999-99") -> ClassRule | None:
    """(source_key, 体系別名) 完全一致 → (source_key, 空) の順で、支給月 ym に
    効いている行を引く。無ければ None。"""
    norm_label = _norm(system_label).replace(" ", "")
    for pair in ((source_key, norm_label), (source_key, "")):
        for rule in master["lookup"].get(pair, []):
            if rule.covers(ym):
                return rule
    return None
