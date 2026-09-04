r"""標準報酬月額チェックのレポート出力（Excel複数シート＋JSON）。

出力先はローカル（outputs/shaho/{YYYY}/）。氏名・給与額・保険料を含む個人情報なので、
共有フォルダへコピーしない（サマリーシートにも注意書きを焼き込む）。
"""

from __future__ import annotations

import datetime
import json
import os

from openpyxl import Workbook

from services.kotsuhi_seisa import write_sheet
from services.shaho_check import REVIEW_STATUSES, STATUS_JA

MONEY_COLS = {"平均報酬月額", "計算健保標報", "計算厚年標報", "登録健保標報", "登録厚年標報",
              "報酬計", "通貨報酬", "現物報酬", "計算健保", "実績健保", "計算介護", "実績介護",
              "計算厚年", "実績厚年", "計算支援金", "実績支援金", "社保調整", "差合計"}


def _kubun(status: str) -> str:
    if status in REVIEW_STATUSES:
        return "要確認"
    if status in ("PROVISIONAL_OK", "NOT_APPLICABLE"):
        return "情報"
    return ""


def _sheet(wb, title, rows, columns):
    write_sheet(wb, title, rows, columns)
    ws = wb[title]
    for ci, name in enumerate(columns, start=1):
        if name in MONEY_COLS:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, ci).number_format = "#,##0"


def build_workbook(check: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    results = check["results"]
    master = check["master"]

    # --- サマリー ---
    counts = {}
    for r in results:
        counts[r.total_status] = counts.get(r.total_status, 0) + 1
    up = sum(1 for r in results if r.teiji and r.teiji.kenpo_smr
             and r.teiji.kenpo_smr > r.reg_kenpo > 0)
    down = sum(1 for r in results if r.teiji and r.teiji.kenpo_smr
               and 0 < r.teiji.kenpo_smr < r.reg_kenpo)
    rows = [{"項目": k, "値": v} for k, v in [
        ("実行日時", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("算定", f"{check['year']}年4〜6月支給 → {check['year']}年9月適用予定"),
        ("突合月", f"{check['check_month']}（本人控除＝{check['cfg']['lag']}か月前の分）"),
        ("保険者", {"its": "関東IT健保", "kyokai_tokyo": "協会けんぽ東京"}[check["insurer"]]),
        ("等級表", f"{master.path}（{master.description}）"),
        ("分類マスタ", f"{check['class_master']['path']}（{len(check['class_master']['rules'])}行）"),
        ("丸め", check["cfg"]["rounding"]), ("許容差", f"±{check['cfg']['tolerance']}円"),
        ("対象者", f"{len(results)}名"),
        ("等級 上がる予定/下がる予定", f"{up}名 / {down}名"),
        ("期中改定の検知", f"{len(check['revisions'])}件"),
        ("", ""),
        ("⚠ このファイルは個人情報（氏名・給与・保険料）を含む",
         "共有フォルダに置かない。社労士へ渡すときは経路に注意"),
        ("⚠ 9月適用前の「差異」は「改定予定」の意味",
         "9月のjinjer更新後にもう一度実行して答え合わせする"),
    ]]
    # 給与が未確定の月（あるとスナップショットがまだ動くので結果を信用しない）
    for ym, st in sorted(check.get("month_status", {}).items()):
        rows.append({
            "項目": f"給与の確定状況: {ym}",
            "値": (f"確定 {st['closed']}名"
                  + (f" / ⚠未確定 {st['open']}名（{'、'.join(st['open_emps'][:5])}）"
                     if st["open"] else "")),
            "区分": "要確認" if st["open"] else "",
        })
    for st in [s for s in STATUS_JA if counts.get(s)]:
        rows.append({"項目": f"総合判定: {STATUS_JA[st]}（{st}）", "値": f"{counts[st]}名",
                     "区分": _kubun(st)})
    _sheet(wb, "サマリー", rows, ["項目", "値", "区分"])

    # --- 社員別判定 ---
    rows = []
    for r in results:
        tk = r.teiji
        rows.append({
            "社員番号": r.emp, "氏名": r.name, "給与体系": r.system,
            "採用月数": tk.adopted_n if tk else "",
            "平均報酬月額": tk.average if tk and tk.average is not None else "",
            "計算健保等級": tk.kenpo_grade if tk else "", "計算健保標報": tk.kenpo_smr if tk else "",
            "計算厚年標報": tk.konen_smr if tk else "",
            "登録健保標報": r.reg_kenpo or "", "登録厚年標報": r.reg_konen or "",
            "標報判定": STATUS_JA[r.teiji_status], "控除判定": STATUS_JA[r.check_status],
            "総合": STATUS_JA[r.total_status], "区分": _kubun(r.total_status),
            "備考": "／".join(r.notes),
        })
    _sheet(wb, "社員別判定", rows,
           ["社員番号", "氏名", "給与体系", "採用月数", "平均報酬月額",
            "計算健保等級", "計算健保標報", "計算厚年標報", "登録健保標報", "登録厚年標報",
            "標報判定", "控除判定", "総合", "区分", "備考"])

    # --- 算定明細（4〜6月） ---
    rows = []
    for r in results:
        if not r.teiji:
            continue
        for a in r.teiji.months:
            rows.append({
                "社員番号": r.emp, "氏名": r.name, "支給月": a.ym,
                "基礎日数": a.base_days.days if a.base_days.days is not None else "",
                "根拠": a.base_days.basis, "採用": "採用" if a.adopted else "除外",
                "除外理由": a.reason, "報酬計": a.rem.total,
                "通貨報酬": a.rem.cash_total, "現物報酬": a.rem.genbutsu_total,
                "検算": "OK" if a.rem.gate_ok else f"差{a.rem.gate_diff:+,.0f}円",
                "区分": "" if a.rem.gate_ok else "要確認",
            })
    _sheet(wb, "算定明細", rows,
           ["社員番号", "氏名", "支給月", "基礎日数", "根拠", "採用", "除外理由",
            "報酬計", "通貨報酬", "現物報酬", "検算", "区分"])

    # --- 保険料突合 ---
    rows = []
    for r in results:
        if not r.premiums:
            continue
        diff = sum(r.actuals.values()) - sum(r.premiums.values())
        rows.append({
            "社員番号": r.emp, "氏名": r.name,
            "計算健保": r.premiums["kenpo"], "実績健保": r.actuals["kenpo"],
            "計算介護": r.premiums["kaigo"], "実績介護": r.actuals["kaigo"],
            "計算厚年": r.premiums["konen"], "実績厚年": r.actuals["konen"],
            "計算支援金": r.premiums["kodomo"], "実績支援金": r.actuals["kodomo"],
            "社保調整": r.chosei or "", "差合計": diff,
            "控除判定": STATUS_JA[r.check_status], "区分": _kubun(r.check_status),
        })
    _sheet(wb, "保険料突合", rows,
           ["社員番号", "氏名", "計算健保", "実績健保", "計算介護", "実績介護",
            "計算厚年", "実績厚年", "計算支援金", "実績支援金", "社保調整", "差合計",
            "控除判定", "区分"])

    # --- 要確認 ---
    rows = [{"社員番号": r.emp, "氏名": r.name, "総合": STATUS_JA[r.total_status],
             "理由": "／".join(r.notes), "区分": "要確認"}
            for r in results if r.total_status in REVIEW_STATUSES]
    _sheet(wb, "要確認", rows, ["社員番号", "氏名", "総合", "理由", "区分"])

    # --- 期中改定検知 ---
    _sheet(wb, "期中改定検知", check["revisions"],
           ["社員番号", "氏名", "検知区間", "健保標報", "厚年標報", "区分"])

    # --- 使用設定・出典 ---
    rows = []
    for key, rate in check["master"].rates.items():
        rows.append({"項目": key, "全体料率": rate.total, "本人負担率": rate.employee,
                     "適用": rate.applies, "出典": rate.source_url})
    _sheet(wb, "使用設定・出典", rows, ["項目", "全体料率", "本人負担率", "適用", "出典"])
    return wb


def write_reports(check: dict) -> dict:
    out_dir = os.path.join(check["out_base"], str(check["year"]))
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx = os.path.join(out_dir, f"標準報酬チェック_{check['year']}年9月適用_{stamp}.xlsx")
    build_workbook(check).save(xlsx)

    payload = {
        "meta": {"year": check["year"], "check_month": check["check_month"],
                 "insurer": check["insurer"], "generated_at": stamp,
                 "grade_table": check["master"].path,
                 "class_master": check["class_master"]["path"],
                 "month_status": check.get("month_status", {})},
        "employees": [{
            "emp": r.emp, "name": r.name, "system": r.system,
            "teiji_status": r.teiji_status, "check_status": r.check_status,
            "total_status": r.total_status, "notes": r.notes,
            "average": r.teiji.average if r.teiji else None,
            "calc_kenpo_smr": r.teiji.kenpo_smr if r.teiji else None,
            "calc_konen_smr": r.teiji.konen_smr if r.teiji else None,
            "reg_kenpo_smr": r.reg_kenpo, "reg_konen_smr": r.reg_konen,
            "premiums": r.premiums, "actuals": r.actuals, "chosei": r.chosei,
        } for r in check["results"]],
        "revisions": check["revisions"],
    }
    jsn = xlsx.replace(".xlsx", ".json")
    with open(jsn, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    review_n = sum(1 for r in check["results"] if r.total_status in REVIEW_STATUSES)
    open_months = {ym: st["open"] for ym, st in check.get("month_status", {}).items()
                   if st["open"]}
    return {"xlsx": xlsx, "json": jsn, "review_n": review_n,
            "n": len(check["results"]), "open_months": open_months}
