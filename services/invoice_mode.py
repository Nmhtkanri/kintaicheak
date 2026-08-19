"""請求書 PDF から freee 取引インポート CSV を作るサービス。

外部サービスへは書き込まず、共有フォルダを読み取り、確認済み行だけを
ローカル出力フォルダへ CSV と実行ログとして保存する。
"""

from __future__ import annotations

import calendar
import csv
import hashlib
import io
import json
import math
import os
import re
import tempfile
import unicodedata
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence


CSV_COLUMNS = [
    "収支区分", "管理番号", "発生日", "支払期日", "取引先", "勘定科目",
    "税区分", "金額", "税計算区分", "税額", "備考", "品目", "部門",
    "メモタグ（複数指定可、カンマ区切り）", "従業員",
]

DEFAULT_TARGET_ROOTS = (
    r"Z:\NetMarks以外(常駐）\A&A（新井）",
    r"Z:\NetMarks以外(常駐）\BBS（岩崎・河端・加藤・瀧澤）",
    r"Z:\NetMarks以外(常駐）\BCS（矢野）",
    r"Z:\NetMarks以外(常駐）\CCT（千代田）",
    r"Z:\NetMarks以外(常駐）\DTS（岡崎）",
    r"Z:\NetMarks以外(常駐）\EXA（有田）",
    r"Z:\NetMarks以外(常駐）\IXナレッジ（小島・八橋）",
    r"Z:\NetMarks以外(常駐）\JID（藤田）",
    r"Z:\NetMarks以外(常駐）\NICソフト（熊崎・岡崎）",
    r"Z:\NetMarks以外(常駐）\NICパートナーズ（大場・小林）",
    r"Z:\NetMarks以外(常駐）\NTTデータ（西村）",
    r"Z:\NetMarks以外(常駐）\S＆I（出澤・石島）",
    r"Z:\NetMarks以外(常駐）\UBS（田中）",
    r"Z:\NetMarks以外(常駐）\アイ・ティー・ワン（中澤）",
    r"Z:\NetMarks以外(常駐）\アイテック（佐野）",
    r"Z:\NetMarks以外(常駐）\アクシスITパートナーズ（細川・渡会）",
    r"Z:\NetMarks以外(常駐）\エリクソン・ジャパン（ラミタ・奈良）",
    r"Z:\NetMarks以外(常駐）\オリゾンシステムズ（清水石）",
    r"Z:\NetMarks以外(常駐）\コムチュア（田村）",
    r"Z:\NetMarks以外(常駐）\ジャステック（阿部・田村・藤川）",
    r"Z:\NetMarks以外(常駐）\フォーカスシステムズ（柳）",
    r"Z:\NetMarks以外(常駐）\ペンギンソリューションズ（菅原）",
    r"Z:\NetMarks以外(常駐）\マンパワーグループ（福永）",
    r"Z:\NetMarks以外(常駐）\ミネルバ・スピード・コミュニケーション（岩元）",
    r"Z:\NetMarks以外(常駐）\ルミナス・ビー・ジャパン（マーティン）",
    r"Z:\NetMarks以外(常駐）\永和情報（住吉）",
    r"Z:\NetMarks以外(常駐）\東陽テクニカ（塚本・野田）",
    r"Z:\NetMarks以外(常駐）\日本ディクス（佐藤賢吾・福島・佐々木）",
)

_DATE_TOKEN = r"(\d{4})\s*[年/\-.]\s*(\d{1,2})\s*[月/\-.]\s*(\d{1,2})\s*日?"
_MONEY_TOKEN = r"[¥￥]?\s*\(?-?\d[\d,]*\)?"
_REVISION_MARKERS = ("修正版", "修正", "差替", "差し替え", "再発行")


class InvoiceModeError(ValueError):
    """画面で修正できる入力・検証エラー。"""


def _nfkc(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def normalize_name(value: Any) -> str:
    value = _nfkc(value).strip("()（）[]【】")
    return re.sub(r"[\s　・･.,，_\-]", "", value).casefold()


def _clean_display_name(value: Any) -> str:
    value = _nfkc(value).strip("()（）[]【】 ")
    value = re.sub(r"\s+(?:スタッフコード|就業期間|\d|[¥￥])(?:.|\n)*$", "", value)
    return re.sub(r"[\s　,，]+", "", value)


def _company_key(value: Any) -> str:
    value = normalize_name(value)
    for token in ("株式会社", "有限会社", "合同会社", "(請求先)", "請求先"):
        value = value.replace(normalize_name(token), "")
    for before, after in (("アイティーワン", "itone"), ("エリクソンジャパン", "ericsson"),
                          ("ixナレッジ", "iki"), ("アイエックスナレッジ", "iki")):
        value = value.replace(before, after)
    return value


def _parse_money(value: Any) -> int | None:
    if value is None:
        return None
    raw = (_nfkc(value).replace("¥", "").replace("￥", "")
           .replace(",", "").replace(" ", ""))
    negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()")
    if not re.fullmatch(r"-?\d+", raw):
        return None
    amount = int(raw)
    return -abs(amount) if negative else amount


def _money_values(text: str, patterns: Sequence[str]) -> list[int]:
    values: list[int] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.MULTILINE | re.IGNORECASE):
            amount = _parse_money(match.group(1))
            if amount is not None:
                values.append(amount)
    return values


def _last_money_on_lines(text: str, labels: Sequence[str]) -> list[int]:
    values: list[int] = []
    for line in text.splitlines():
        normalized = _nfkc(line)
        if not any(label in normalized for label in labels):
            continue
        parsed = [_parse_money(token) for token in re.findall(_MONEY_TOKEN, normalized)]
        parsed = [value for value in parsed if value is not None]
        if parsed:
            values.append(parsed[-1])
    return values


def _format_date_match(match: re.Match[str] | None) -> str:
    if not match:
        return ""
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
    except ValueError:
        return ""


def _extract_labeled_date(text: str, labels: Sequence[str]) -> str:
    for label in labels:
        match = re.search(rf"{label}\s*[:：]?\s*{_DATE_TOKEN}", text, re.IGNORECASE)
        if match:
            return _format_date_match(match)
    return ""


def _extract_employee_name(text: str, filename: str) -> str:
    patterns = (
        r"作業担当者\s*[:：]\s*([^\n]+)",
        r"スタッフ\s+([^,\n]+,\s*[^\s\n]+)(?=\s+提出者)",
        r"スタッフ氏名\s*[:：]\s*([^\n]+)",
        r"スタッフ名\s+([^\n]+)",
        r"請求書明細\s*\n\s*([^\n]+)",
        r"\bWorker\s+([A-Za-z'\-]+\s*,\s*[A-Za-z'\-]+)",
        r"\bWorker\s+(?:Name\s*)?(.+?)(?=\s+(?:Submit|Remit|Site|Purchase|Business|Job|Status)|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            continue
        candidate = _clean_display_name(match.group(1))
        if candidate and not any(token in candidate for token in ("株式会社", "御中", "請求")):
            return candidate
    bracket = re.search(r"【([^】]+)】", filename)
    if bracket:
        candidate = _clean_display_name(bracket.group(1))
        if (candidate and not candidate.endswith("様")
                and not any(token in candidate for token in ("請求", "A&A"))):
            return candidate
    return ""


def _extract_partner(text: str) -> str:
    for pattern in (
        r"発注者\s+(.+?)\s+サプライヤ",
        r"請求先\s+(.+?)(?:\s+差出人|\n)",
        r"\bBuyer\s+([^\s\n]+)",
    ):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            partner = _nfkc(match.group(1)).strip()
            if (partner and partner not in {"送金先", "請求元"}
                    and "エヌエム・ヒューマテック" not in partner):
                return partner
    for line in text.splitlines()[:30]:
        normalized = _nfkc(line)
        if "御中" not in normalized:
            continue
        partner = re.sub(r"\s*御中.*$", "", normalized).strip()
        partner = re.sub(r"^\(請求先\)", "", partner).strip()
        if partner and "エヌエム・ヒューマテック" not in partner:
            return partner
    return ""


def _extract_invoice_number(text: str) -> str:
    match = re.search(r"請求書(?:番号|コード)\s*[:：]?\s*([^\s\n]+)", text)
    return _nfkc(match.group(1)) if match else ""


def extract_pdf_text(path: os.PathLike[str] | str, max_pages: int = 2) -> str:
    """PDF先頭ページ群を文字列化する。画像PDFは空文字となり画面補完へ回す。"""
    import pdfplumber

    chunks: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages[:max_pages]:
            chunks.append(page.extract_text() or "")
    return "\n".join(chunks)


def parse_invoice_text(text: str, filename: str = "") -> dict[str, Any]:
    """抽出済みテキストを標準化する。欠損値は推測せず空欄で返す。"""
    normalized = _nfkc(text)
    commute_only = any(token in filename for token in ("立替", "交通費")) and (
        "請求" in filename or "invoice" in filename.casefold())
    issue_date = _extract_labeled_date(
        normalized, (r"請求年月日", r"請求日", r"発行日", r"請求確定日時", r"Invoice\s*Date",
                     r"End\s*Date"))   # End Date は Fieldglass(英語)の請求対象期間の末日
    due_date = _extract_labeled_date(
        normalized, (r"入金期日", r"お支払\s*期\s*日", r"支払\s*期\s*日", r"Due\s*Date"))
    tax_values = _money_values(normalized, (
        rf"10\s*[%％]\s*消費税\s*({_MONEY_TOKEN})",
        rf"^\s*消費税(?:額)?\s+({_MONEY_TOKEN})",
        rf"Consumption\s*Tax\s*[:：]?\s*({_MONEY_TOKEN})",
        rf"消費税\s+({_MONEY_TOKEN})",
    ))
    tax_values.extend(_last_money_on_lines(
        normalized, ("消費税_10%", "調整(消費税等)", "消費税(10%)", "合計税額")))
    tax = max(tax_values) if tax_values else None
    total_values = _money_values(normalized, (
        rf"合計金額\s*[:：]?\s*({_MONEY_TOKEN})",
        rf"御請求金額総計\s*({_MONEY_TOKEN})",
        rf"^\s*請求合計\s+({_MONEY_TOKEN})",
        rf"Total\s*Amount\s*Due\s*[:：]?\s*({_MONEY_TOKEN})",
        rf"請求総額\s*\(税込み\)\s*({_MONEY_TOKEN})",
        rf"請求金額\s*({_MONEY_TOKEN})\s*円?\s*\(税込",
        rf"^\s*総合計\s+({_MONEY_TOKEN})",
        rf"^\s*支払金額\s+({_MONEY_TOKEN})",
        rf"^\s*総計\s+({_MONEY_TOKEN})",
        rf"^\s*合計\s+({_MONEY_TOKEN})",
    ))
    total = max(total_values) if total_values else None

    if commute_only:
        gross = total
        if gross is None:
            values = _last_money_on_lines(normalized, ("立替金(交通費等)", "総合計金額"))
            gross = max(values) if values else None
        return {
            "kind": "commute", "employee_name": _extract_employee_name(normalized, filename),
            "partner": _extract_partner(normalized), "issue_date": issue_date,
            "due_date": due_date, "invoice_number": _extract_invoice_number(normalized),
            "main_amount": None, "main_tax": None, "commute_amount": gross,
            "commute_tax": tax, "text": normalized,
        }

    subtotal_values = _money_values(normalized, (
        rf"^\s*小計\s+({_MONEY_TOKEN})\s*$",
        rf"^\s*請求小計\s+({_MONEY_TOKEN})",
        rf"Subtotal\s*[:：]?\s*({_MONEY_TOKEN})",
        rf"税抜合計額\s*({_MONEY_TOKEN})",
        rf"10\s*%\s*対象\s+税抜金額\s*({_MONEY_TOKEN})",
        rf"明細の小計\s*({_MONEY_TOKEN})",
        rf"^\s*10\s*%\s*対象\s+({_MONEY_TOKEN})",
    ))
    subtotal = max(subtotal_values) if subtotal_values else None
    commute_values = _last_money_on_lines(
        normalized, ("立替金(交通費等)", "交通費相当額小計"))
    commute = max(commute_values) if commute_values else 0
    if subtotal is not None and tax is not None:
        main_amount = subtotal + tax
    elif total is not None:
        main_amount = total - max(commute, 0)
    else:
        main_amount = None
    return {
        "kind": "main", "employee_name": _extract_employee_name(normalized, filename),
        "partner": _extract_partner(normalized), "issue_date": issue_date,
        "due_date": due_date, "invoice_number": _extract_invoice_number(normalized),
        "main_amount": main_amount, "main_tax": tax, "commute_amount": commute,
        "commute_tax": math.floor(commute / 11) if commute > 0 else 0,
        "text": normalized,
    }


def parse_invoice_pdf(path: os.PathLike[str] | str) -> dict[str, Any]:
    path = Path(path)
    parsed = parse_invoice_text(extract_pdf_text(path), path.name)
    parsed["source_file"] = str(path)
    parsed["revised"] = any(marker in path.stem for marker in _REVISION_MARKERS)
    return parsed


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return list(csv.DictReader(io.StringIO(path.read_text(encoding=encoding))))
        except UnicodeDecodeError:
            continue
    raise InvoiceModeError(f"CSVの文字コードを判定できません: {path}")


def load_target_roots(csv_path: os.PathLike[str] | str | None = None) -> list[str]:
    """外部CSVがあれば優先し、無ければ指定済みの28フォルダを使う。"""
    if csv_path and Path(csv_path).exists():
        roots: list[str] = []
        for row in _read_csv_rows(Path(csv_path)):
            enabled = _nfkc(row.get("対象", row.get("有効", "1"))).casefold()
            if enabled in {"0", "false", "no", "対象外", "無効"}:
                continue
            raw = row.get("フォルダパス") or row.get("Path") or row.get("path") or ""
            if _nfkc(raw):
                roots.append(str(raw).strip().strip('"'))
        if roots:
            return roots
    return list(DEFAULT_TARGET_ROOTS)


def _submission_bases(root: Path) -> list[Path]:
    try:
        children = [child for child in root.iterdir() if child.is_dir()]
    except OSError:
        return [root]
    matched = [child for child in children if "提出データ" in _nfkc(child.name)]
    return matched or [root]


def _belongs_to_month(path: Path, year: int, month: int) -> bool:
    ym = f"{year}{month:02d}"
    full = _nfkc(str(path))
    return (ym in _nfkc(path.name) or f"{year}年{month}月" in full
            or f"{year}年{month:02d}月" in full or f"{year}-{month:02d}" in full)


def _is_invoice_candidate(path: Path) -> bool:
    name = _nfkc(path.name).casefold()
    return (path.suffix.casefold() == ".pdf" and ("請求" in name or "invoice" in name)
            and "承認メール" not in name)


def _revision_key(path: Path) -> str:
    stem = _nfkc(path.stem).casefold()
    for marker in _REVISION_MARKERS:
        stem = stem.replace(_nfkc(marker).casefold(), "")
    stem = re.sub(r"(?:[_\-\s]*(?:rev|ver|v)\s*\d+(?:\.\d+)*)$", "", stem)
    stem = re.sub(r"[_\-\s（）()]+", "", stem)
    return f"{_nfkc(str(path.parent)).casefold()}|{stem}"


def _revision_rank(path: Path) -> tuple[int, int, int]:
    revised = int(any(marker in path.stem for marker in _REVISION_MARKERS))
    version_match = re.search(r"(?:rev|ver|v)\s*(\d+(?:\.\d+)*)", path.stem, re.IGNORECASE)
    version = int(float(version_match.group(1)) * 100) if version_match else 0
    try:
        mtime = path.stat().st_mtime_ns
    except OSError:
        mtime = 0
    return revised, version, mtime


def find_invoice_files(month: str, roots: Sequence[os.PathLike[str] | str] | None = None) -> dict[str, Any]:
    """指定月の請求書候補を探し、同一名の修正版は最新版だけを採用する。"""
    match = re.fullmatch(r"(\d{4})-(\d{2})", _nfkc(month))
    if not match:
        raise InvoiceModeError("対象月は YYYY-MM 形式で指定してください")
    year, month_no = int(match.group(1)), int(match.group(2))
    if not 1 <= month_no <= 12:
        raise InvoiceModeError("対象月が不正です")
    candidates: list[Path] = []
    missing_roots: list[str] = []
    scan_errors: list[str] = []
    for raw_root in roots or DEFAULT_TARGET_ROOTS:
        root = Path(raw_root)
        if not root.exists():
            missing_roots.append(str(root))
            continue
        for base in _submission_bases(root):
            try:
                for path in base.rglob("*.pdf"):
                    if _belongs_to_month(path, year, month_no) and _is_invoice_candidate(path):
                        candidates.append(path)
            except OSError as exc:
                scan_errors.append(f"{base}: {exc}")
    unique = {os.path.normcase(os.path.abspath(str(path))): path for path in candidates}
    grouped: OrderedDict[str, list[Path]] = OrderedDict()
    for path in sorted(unique.values(), key=lambda item: _nfkc(str(item)).casefold()):
        grouped.setdefault(_revision_key(path), []).append(path)
    selected: list[Path] = []
    ignored: list[dict[str, str]] = []
    for paths in grouped.values():
        winner = max(paths, key=_revision_rank)
        selected.append(winner)
        for path in paths:
            if path != winner:
                ignored.append({"file": str(path), "reason": f"修正版を優先: {winner.name}"})
    return {"selected": [str(path) for path in selected], "ignored": ignored,
            "missing_roots": missing_roots, "scan_errors": scan_errors}


def fingerprint_files(paths: Iterable[os.PathLike[str] | str]) -> str:
    digest = hashlib.sha256()
    normalized_paths = {os.path.normcase(os.path.abspath(str(path))) for path in paths}
    for raw in sorted(normalized_paths):
        path = Path(raw)
        try:
            stat = path.stat()
            token = f"{raw}|{stat.st_size}|{stat.st_mtime_ns}"
        except OSError:
            token = f"{raw}|missing"
        digest.update(token.encode("utf-8", errors="surrogatepass"))
    return digest.hexdigest()


def _scan_signature(scan: dict[str, Any]) -> str:
    ignored = [item["file"] for item in scan.get("ignored", [])]
    return fingerprint_files([*scan.get("selected", []), *ignored])


def _sales_book_path(template: str | None, year: int) -> str:
    if not template:
        return ""
    try:
        return template.format(year=year)
    except (KeyError, ValueError):
        return template


def load_employee_master(sales_book: os.PathLike[str] | str | None,
                         override_csv: os.PathLike[str] | str | None = None) -> dict[str, dict[str, str]]:
    """年度営業実績と任意の補正CSVから氏名・社員番号対応を作る。"""
    master: dict[str, dict[str, str]] = {}
    if sales_book and Path(sales_book).exists():
        from openpyxl import load_workbook

        workbook = load_workbook(str(sales_book), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            if sheet.title in {"まとめ", "まとめ (2)", "グラフ"}:
                continue
            for row in sheet.iter_rows(values_only=True):
                values = list(row) + [None] * 11
                if _nfkc(values[10]) != "総受注金額":
                    continue
                employee_id = _nfkc(values[0])
                employee_name = _clean_display_name(values[1])
                if not employee_id or not employee_name:
                    continue
                contract_type = _nfkc(values[3])
                department = {"派遣": "他社向け派遣", "委任": "他社向け委任契約"}.get(contract_type, "")
                master[normalize_name(employee_name)] = {
                    "employee_no": employee_id, "employee_name": employee_name,
                    "partner": re.sub(r"\s*/\s*", "", _nfkc(values[8])).replace("\n", ""),
                    "department": department, "source": f"{Path(sales_book).name}:{sheet.title}",
                    "partner_override": False,
                }
    if override_csv and Path(override_csv).exists():
        for row in _read_csv_rows(Path(override_csv)):
            employee_name = row.get("氏名") or row.get("従業員") or ""
            key = normalize_name(employee_name)
            if not key:
                continue
            current = dict(master.get(key, {}))
            current.update({
                "employee_no": _nfkc(row.get("社員番号") or row.get("管理番号") or current.get("employee_no", "")),
                "employee_name": _clean_display_name(employee_name) or current.get("employee_name", ""),
                "partner": _nfkc(row.get("freee取引先") or row.get("取引先") or current.get("partner", "")),
                "department": _nfkc(row.get("部門") or current.get("department", "")),
                "source": str(override_csv),
                "partner_override": bool(_nfkc(row.get("freee取引先") or row.get("取引先"))),
            })
            master[key] = current
            aliases = re.split(r"[|｜;；]", _nfkc(row.get("氏名別名") or row.get("別名")))
            for alias in aliases:
                if normalize_name(alias):
                    master[normalize_name(alias)] = current
    return master


def _match_employee(document: dict[str, Any], master: dict[str, dict[str, str]]) -> dict[str, str]:
    direct = normalize_name(document.get("employee_name"))
    if direct in master:
        return master[direct]
    haystack = normalize_name(f"{document.get('employee_name', '')} {document.get('text', '')}")
    matches = [(key, value) for key, value in master.items() if len(key) >= 3 and key in haystack]
    if len(matches) == 1:
        return matches[0][1]

    # PDFに担当者名が無い形式は、契約先と対象フォルダから一意に決まる場合だけ補う。
    # BBSやIXのように同じ契約先に複数人いる場合は推測せず、画面補完へ回す。
    company_haystack = _company_key(
        f"{document.get('partner', '')} {document.get('source_file', '')}")
    company_matches = []
    for value in master.values():
        partner_key = _company_key(value.get("partner", ""))
        if len(partner_key) >= 3 and partner_key in company_haystack:
            company_matches.append(value)
    if len(company_matches) == 1:
        return company_matches[0]
    return {"employee_no": "", "employee_name": _clean_display_name(document.get("employee_name", "")),
            "partner": "", "department": "", "source": ""}
def load_departments(employee_ids: Iterable[str], on_date: str, *,
                     cache_path: os.PathLike[str] | str | None = None,
                     client: Any = None) -> dict[str, str]:
    """jinjer のカスタム項目「給与計算関連」から、on_date 時点の部門を引く。

    元データも時点解決も経理モードと同じものを使う（services/keiri_api）。
    部門は履歴を持つので、対象月の末日時点の値を採る（月中異動は異動後になる）。

    キャッシュ（経理モードが作る raw/custom_items.json）があればそれを読む。
    無ければ jinjer を叩くが、取れなくても請求書CSVは作れるべきなので、
    失敗しても例外にはせず空の辞書を返す。

    Returns:
        {社員番号: 部門}。引けなかった人はキーごと入らない。
    """
    ids = [str(e).strip() for e in employee_ids if str(e or "").strip()]
    if not ids:
        return {}
    try:
        from services.keiri_api import (get_client, parse_payroll_custom_history,
                                        resolve_custom_value)
    except Exception:                                          # noqa: BLE001
        return {}

    raw: dict[str, Any] = {}
    if cache_path and Path(cache_path).exists():
        try:
            raw = json.loads(Path(cache_path).read_text(encoding="utf-8")).get("data") or {}
        except Exception:                                      # noqa: BLE001
            raw = {}
    missing = [e for e in ids if e not in raw]
    if missing:
        try:
            raw.update((client or get_client()).get_custom_items(missing))
        except Exception:                                      # noqa: BLE001
            pass   # キャッシュ分だけで続ける（部門が空なら画面で赤くなり人が気づく）

    out: dict[str, str] = {}
    for emp in ids:
        person = raw.get(emp)
        if not person:
            continue
        value = resolve_custom_value(
            parse_payroll_custom_history(person), "部門", on_date)
        if value:
            out[emp] = value
    return out


def _clamp_issue_date(issue_date: str, month_end: date) -> tuple[str, bool]:
    """発生日を対象月の末日に寄せる。

    freee には対象月の末日（2026/7/31）で登録する運用なのに、請求書によっては
    「請求確定日時」「出力日時」など翌月初の日付しか載っていないものがある
    （IXナレッジ様の 2026/08/05 など）。対象月から外れた日付はそのまま使わない。

    Returns:
        (発生日, 元の日付を月末へ置き換えたか)
    """
    fallback = month_end.isoformat()
    if not issue_date:
        return fallback, False
    try:
        parsed = date.fromisoformat(issue_date)
    except ValueError:
        return fallback, True
    if (parsed.year, parsed.month) != (month_end.year, month_end.month):
        return fallback, True
    return issue_date, False


def _common_value(documents: Sequence[dict[str, Any]], key: str) -> tuple[str, bool]:
    values = {_nfkc(doc.get(key)) for doc in documents if _nfkc(doc.get(key))}
    if len(values) == 1:
        return next(iter(values)), False
    return "", len(values) > 1


def _validation_messages(row: dict[str, Any]) -> list[str]:
    row_type = row.get("_row_type", "main")
    required = ["勘定科目", "税区分", "金額", "税計算区分", "税額", "部門", "従業員"]
    if row_type == "main":
        required = ["収支区分", "管理番号", "発生日", "支払期日", "取引先", *required]
    errors = [f"{column}が未入力" for column in required if str(row.get(column, "")).strip() == ""]
    if row_type == "main" and row.get("管理番号") and not re.fullmatch(r"\d+", _nfkc(row["管理番号"])):
        errors.append("管理番号は数字で入力")
    for column in (("発生日", "支払期日") if row_type == "main" else ()):
        if row.get(column):
            try:
                date.fromisoformat(_nfkc(row[column]).replace("/", "-"))
            except ValueError:
                errors.append(f"{column}は正しい日付で入力")
    amount = _parse_money(row.get("金額"))
    tax = _parse_money(row.get("税額"))
    if row.get("金額") not in (None, "") and (amount is None or amount <= 0):
        errors.append("金額は1円以上の整数で入力")
    if row.get("税額") not in (None, "") and (tax is None or tax < 0):
        errors.append("税額は0円以上の整数で入力")
    if amount is not None and tax is not None and tax > amount:
        errors.append("税額が金額を超えています")
    return errors


def build_preview(month: str, *, roots: Sequence[os.PathLike[str] | str] | None = None,
                  target_roots_csv: os.PathLike[str] | str | None = None,
                  sales_book_template: str | None = None,
                  master_csv: os.PathLike[str] | str | None = None,
                  default_department: str = "",
                  custom_items_cache: os.PathLike[str] | str | None = None) -> dict[str, Any]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", _nfkc(month))
    if not match:
        raise InvoiceModeError("対象月は YYYY-MM 形式で指定してください")
    year = int(match.group(1))
    month_end = date(year, int(match.group(2)),
                     calendar.monthrange(year, int(match.group(2)))[1])
    target_roots = list(roots) if roots is not None else load_target_roots(target_roots_csv)
    scan = find_invoice_files(month, target_roots)
    sales_book = _sales_book_path(sales_book_template, year)
    master = load_employee_master(sales_book, master_csv)
    documents: list[dict[str, Any]] = []
    parse_errors: list[str] = []
    for source_file in scan["selected"]:
        try:
            document = parse_invoice_pdf(source_file)
        except Exception as exc:
            parse_errors.append(f"{source_file}: {exc}")
            continue
        employee = _match_employee(document, master)
        if employee.get("employee_name"):
            document["employee_name"] = employee["employee_name"]
        document["employee_no"] = employee.get("employee_no", "")
        document["master_partner"] = employee.get("partner", "")
        document["department"] = employee.get("department", "") or _nfkc(default_department)
        document["master_source"] = employee.get("source", "")
        if employee.get("partner_override") and employee.get("partner"):
            document["partner"] = employee["partner"]
        documents.append(document)

    # 部門は jinjer のカスタム項目「給与計算関連」を正とする（経理モードと同じ元データ）。
    # 引けなかった人だけ、売上簿の契約形態から作った値・既定値を使う。
    jinjer_departments = load_departments(
        {doc.get("employee_no") for doc in documents},
        month_end.isoformat(), cache_path=custom_items_cache)
    for document in documents:
        from_jinjer = jinjer_departments.get(document.get("employee_no") or "")
        if from_jinjer:
            document["department"] = from_jinjer

    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for document in documents:
        employee_key = normalize_name(document.get("employee_name")) or f"file:{document['source_file']}"
        partner = document.get("partner") or document.get("master_partner") or ""
        grouped.setdefault(f"{employee_key}|{normalize_name(partner)}", []).append(document)

    rows: list[dict[str, Any]] = []
    public_documents: list[dict[str, Any]] = []
    for group_index, group_documents in enumerate(grouped.values(), 1):
        main_documents = [doc for doc in group_documents if doc["kind"] == "main"]
        commute_documents = [doc for doc in group_documents if doc["kind"] == "commute"]
        basis = main_documents or commute_documents
        employee_name = _nfkc(basis[0].get("employee_name")) if basis else ""
        employee_no = _nfkc(basis[0].get("employee_no")) if basis else ""
        partner = _nfkc(basis[0].get("partner") or basis[0].get("master_partner")) if basis else ""
        department = _nfkc(basis[0].get("department")) if basis else _nfkc(default_department)
        issue_date, issue_conflict = _common_value(basis, "issue_date")
        issue_date, issue_moved = _clamp_issue_date(issue_date, month_end)
        due_date, due_conflict = _common_value(basis, "due_date")
        main_amount: int | str = ""
        main_tax: int | str = ""
        if main_documents and all(doc.get("main_amount") is not None for doc in main_documents):
            main_amount = sum(int(doc["main_amount"]) for doc in main_documents)
        if main_documents and all(doc.get("main_tax") is not None for doc in main_documents):
            main_tax = sum(int(doc["main_tax"]) for doc in main_documents)
        group_id = f"invoice-{group_index}"
        # 備考・品目は空欄で出す（2026-08 谷津さん指示）。
        # Fieldglass(UAL)分の取込では備考に「総合計請求書：氏名」を入れていたが、
        # 請求書モードで作る分では不要とのこと。戻すならここに文言を組み立てる。
        remarks = ""
        main_row: dict[str, Any] = {
            "収支区分": "収入", "管理番号": employee_no, "発生日": issue_date,
            "支払期日": due_date, "取引先": partner, "勘定科目": "売上高",
            "税区分": "課税売上10%", "金額": main_amount, "税計算区分": "内税",
            "税額": main_tax, "備考": remarks, "品目": "", "部門": department,
            "メモタグ（複数指定可、カンマ区切り）": "", "従業員": employee_name,
            "_row_type": "main", "_group_id": group_id,
            "_sources": [doc["source_file"] for doc in main_documents], "_warnings": [],
        }
        if not main_documents:
            main_row["_warnings"].append("本体請求書が見つかりません")
        if issue_moved:
            main_row["_warnings"].append(
                f"PDFの日付が対象月の外だったので発生日を {month_end.isoformat()} にしました")
        if issue_conflict:
            main_row["_warnings"].append("請求日が複数あります")
        if due_conflict:
            main_row["_warnings"].append("入金期日が複数あります")
        main_row["_errors"] = _validation_messages(main_row)
        rows.append(main_row)

        embedded_commute = sum(int(doc.get("commute_amount") or 0) for doc in main_documents)
        separate_commute = sum(int(doc.get("commute_amount") or 0) for doc in commute_documents)
        commute_amount = embedded_commute + separate_commute
        if commute_amount > 0 or commute_documents:
            embedded_tax = sum(int(doc.get("commute_tax") or 0) for doc in main_documents)
            missing_tax = any(doc.get("commute_tax") is None for doc in commute_documents)
            separate_tax = sum(int(doc.get("commute_tax") or 0) for doc in commute_documents)
            commute_tax: int | str = "" if missing_tax else embedded_tax + separate_tax
            commute_row: dict[str, Any] = {
                "収支区分": "", "管理番号": "", "発生日": "", "支払期日": "",
                "取引先": "", "勘定科目": "売上高（交通費）", "税区分": "課税売上10%",
                "金額": commute_amount or "", "税計算区分": "内税", "税額": commute_tax,
                "備考": remarks, "品目": "", "部門": department,
                "メモタグ（複数指定可、カンマ区切り）": "", "従業員": employee_name,
                "_row_type": "commute", "_group_id": group_id,
                "_sources": [doc["source_file"] for doc in group_documents
                             if int(doc.get("commute_amount") or 0) > 0], "_warnings": [],
            }
            commute_row["_errors"] = _validation_messages(commute_row)
            rows.append(commute_row)
        for document in group_documents:
            public_documents.append({key: value for key, value in document.items() if key != "text"})

    return {
        "month": month, "rows": rows, "documents": public_documents,
        "selected_files": scan["selected"], "ignored": scan["ignored"],
        "missing_roots": scan["missing_roots"], "scan_errors": scan["scan_errors"],
        "parse_errors": parse_errors, "signature": _scan_signature(scan),
        "target_count": len(target_roots), "sales_book": sales_book,
    }


def validate_rows(rows: Sequence[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    cleaned: list[dict[str, Any]] = []
    errors: list[str] = []
    for index, source in enumerate(rows, 1):
        row = {column: source.get(column, "") for column in CSV_COLUMNS}
        row_type = source.get("_row_type", "main")
        row["_row_type"] = row_type
        messages = _validation_messages(row)
        if messages:
            errors.extend(f"{index}行目: {message}" for message in messages)
            continue
        row["金額"] = _parse_money(row["金額"])
        row["税額"] = _parse_money(row["税額"])
        if row_type == "main":
            row["発生日"] = _nfkc(row["発生日"]).replace("-", "/")
            row["支払期日"] = _nfkc(row["支払期日"]).replace("-", "/")
        else:
            for column in ("収支区分", "管理番号", "発生日", "支払期日", "取引先"):
                row[column] = ""
        cleaned.append(row)
    return cleaned, errors


def export_csv(month: str, rows: Sequence[dict[str, Any]], output_root: os.PathLike[str] | str,
               *, log_context: dict[str, Any] | None = None) -> dict[str, Any]:
    """確認済み行を UTF-8 BOM・CRLF のfreee CSVと実行ログへ原子的に保存する。"""
    if not re.fullmatch(r"\d{4}-\d{2}", _nfkc(month)):
        raise InvoiceModeError("対象月は YYYY-MM 形式で指定してください")
    cleaned, errors = validate_rows(rows)
    if errors:
        raise InvoiceModeError("\n".join(errors))
    if not cleaned:
        raise InvoiceModeError("出力する行がありません")
    ym = month.replace("-", "")
    output_dir = Path(output_root) / ym
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_name = f"売上（請求書分）{ym}.csv"
    csv_path = output_dir / csv_name
    fd, temp_name = tempfile.mkstemp(prefix="invoice_", suffix=".csv.tmp", dir=output_dir)
    try:
        with os.fdopen(fd, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, extrasaction="ignore",
                                    lineterminator="\r\n")
            writer.writeheader()
            writer.writerows(cleaned)
        os.replace(temp_name, csv_path)
    except Exception:
        try:
            os.unlink(temp_name)
        except OSError:
            pass
        raise
    log_name = f"売上（請求書分）{ym}_実行ログ.json"
    log_path = output_dir / log_name
    log_payload = {
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "month": month, "csv_file": str(csv_path), "row_count": len(cleaned),
        "rows": cleaned, **(log_context or {}),
    }
    fd, temp_log = tempfile.mkstemp(prefix="invoice_", suffix=".json.tmp", dir=output_dir)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(log_payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        os.replace(temp_log, log_path)
    except Exception:
        try:
            os.unlink(temp_log)
        except OSError:
            pass
        raise
    return {"output_dir": str(output_dir), "csv_path": str(csv_path),
            "csv_name": csv_name, "log_path": str(log_path), "log_name": log_name,
            "row_count": len(cleaned)}


def current_scan_signature(month: str, roots: Sequence[str]) -> tuple[str, dict[str, Any]]:
    scan = find_invoice_files(month, roots)
    return _scan_signature(scan), scan
