import anthropic
import base64
import csv
import io
import json
import time
import logging
import os
import re
import tempfile
from datetime import datetime, date, time as dt_time
import pandas as pd
from config import Config
from services.shift_legend_parser import (
    parse_with_legend_extraction,
    to_legend_dict_for_ui,
)
from services.shift_resolver import resolve_shifts

logger = logging.getLogger(__name__)

SYSTEM_PROMPT = """あなたは勤務表データを解析する専門家です。
与えられた勤務表から、以下の情報を抽出してJSON形式で返してください。

出力形式:
{
  "employee_name": "氏名（漢字）",
  "records": [
    {
      "date": "YYYY-MM-DD",
      "start_time": "HH:MM",
      "end_time": "HH:MM",
      "total_work_time": "HH:MM",
      "comment": "備考があれば記載、なければnull"
    }
  ]
}

ルール:
- 日付は必ず YYYY-MM-DD 形式にする
- 時刻は必ず HH:MM（24時間制）にする
- total_work_time は、その日の休憩控除後の総労働時間を HH:MM 形式で返す
- 日別の「実働」「実労働」「稼働時間」「作業時間」「業務時間数」「合計時間」を優先する
- 日別の正味時間が無く、同じ行に出勤・退勤・休憩があれば退勤－出勤－休憩で算出する
- 時間内・時間外が別列なら両方を合算する
- 「時間内 8時間以内」のような説明や上限表示を総労働時間として扱わない
- 根拠を持って算出できない場合は total_work_time を null にする
- 休日・休暇の行は含めない（出退勤時刻がある行のみ抽出）
- 深夜勤務で日付をまたぐ場合、退勤時刻は翌日の時刻として "25:00" のように表記する
- 氏名が見つからない場合は "employee_name": "不明" とする
- 複数人分のデータがある場合は、配列で返す:
  [{"employee_name": "...", "records": [...]}, ...]
- JSONのみを返し、それ以外のテキストは含めないこと
"""


def _excel_to_text(filepath):
    """ExcelファイルをCSV風テキストに変換"""
    import openpyxl
    workbook_path = filepath
    converted_path = None
    if os.path.splitext(filepath)[1].lower() == ".xlsb":
        converted_path = _convert_xlsb_to_xlsx(filepath)
        if converted_path is None:
            raise ValueError("xlsbをxlsxに変換できませんでした")
        workbook_path = converted_path

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            lines.append(f"=== シート: {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(v) if v is not None else "" for v in row]
                if any(v.strip() for v in row_vals):
                    lines.append(",".join(row_vals))
        return "\n".join(lines)
    finally:
        if converted_path:
            try:
                os.remove(converted_path)
            except OSError:
                pass


def _pdf_to_text_or_bytes(filepath):
    """PDFからテキスト抽出。テキストが少ない場合はバイナリを返す"""
    import pdfplumber
    text_pages = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_pages.append(text)
    full_text = "\n".join(text_pages)
    if len(full_text.strip()) < 100:
        # テキストが少ない → PDFバイナリとして返す
        with open(filepath, "rb") as f:
            return None, f.read()
    return full_text, None


# 画像化したPDFページの長辺の上限(px)。
# Anthropic APIは長辺1568pxを超える画像を内部で縮小するため、これ以上大きくしても
# 精度は上がらず、base64後のサイズだけ膨らんでリクエスト上限(32MB)を超えて413になる。
_PDF_RENDER_MAX_LONG_EDGE = 2000


def _pdf_first_page_to_png_bytes(filepath):
    """文字を持たないPDFを画像解析に回すため、1ページ目をPNG化する。

    レンダリング倍率は固定せず、ページ実寸から長辺が _PDF_RENDER_MAX_LONG_EDGE 以内に
    収まるよう動的に決める。高解像度スキャンPDFでも画像が肥大化せず、Claude APIの
    リクエストサイズ上限(32MB)・画像最大辺(8000px)を超えない。
    """
    try:
        import pypdfium2 as pdfium

        pdf = pdfium.PdfDocument(filepath)
        if len(pdf) == 0:
            return None
        page = pdf[0]
        width_pt, height_pt = page.get_size()  # scale=1.0 相当のピクセル寸法
        native_long_edge = max(width_pt, height_pt)
        if native_long_edge <= 0:
            scale = 4.0
        else:
            # 長辺を _PDF_RENDER_MAX_LONG_EDGE に合わせる倍率。小さなPDFは拡大、
            # 巨大スキャンは縮小する。拡大しすぎを防ぐため上限は4.0倍。
            scale = min(4.0, _PDF_RENDER_MAX_LONG_EDGE / native_long_edge)
        bitmap = page.render(scale=scale)
        image = bitmap.to_pil().convert("RGB")
        buf = io.BytesIO()
        image.save(buf, format="PNG")
        return buf.getvalue()
    except Exception as e:
        logger.warning("PDFの画像化に失敗しました: %s", e)
        return None


def _parse_with_claude(file_content, file_type, media_type=None):
    """Claude APIで勤務表を構造化する（リトライ付き）"""
    client = anthropic.Anthropic()

    if file_type == "text":
        messages = [{
            "role": "user",
            "content": f"以下の勤務表データを解析してください:\n\n{file_content}"
        }]
    elif file_type == "image":
        b64_data = base64.standard_b64encode(file_content).decode("utf-8")
        messages = [{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type or "image/png",
                        "data": b64_data
                    }
                },
                {"type": "text", "text": "この勤務表を解析してください。"}
            ]
        }]
    elif file_type == "pdf":
        b64_data = base64.standard_b64encode(file_content).decode("utf-8")
        messages = [{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": b64_data
                    }
                },
                {"type": "text", "text": "この勤務表を解析してください。"}
            ]
        }]
    else:
        raise ValueError(f"未対応のfile_type: {file_type}")

    for attempt in range(3):
        try:
            response = client.messages.create(
                model=Config.ANTHROPIC_MODEL,
                max_tokens=Config.ANTHROPIC_MAX_TOKENS,
                system=SYSTEM_PROMPT,
                messages=messages
            )
            result_text = response.content[0].text.strip()
            # JSONブロック抽出
            json_match = re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", result_text)
            if json_match:
                result_text = json_match.group(1)
            return json.loads(result_text)
        except json.JSONDecodeError as e:
            logger.warning(f"JSON解析失敗 (試行{attempt+1}/3): {e}")
            if attempt < 2:
                # プロンプトを補強して再送
                messages[-1]["content"] = messages[-1]["content"] if isinstance(messages[-1]["content"], str) else messages[-1]["content"]
                if isinstance(messages[-1]["content"], list):
                    messages[-1]["content"][-1]["text"] = "この勤務表を解析してください。必ずJSONのみを返し、説明文は不要です。"
                time.sleep(2)
        except anthropic.RateLimitError:
            logger.warning(f"レート制限 (試行{attempt+1}/3)、3秒後にリトライ...")
            time.sleep(3)
        except anthropic.APIError as e:
            raise RuntimeError(f"Claude APIエラー: {e}")

    raise RuntimeError("Claude APIからの応答をJSONとして解析できませんでした")


def _normalize_records(claude_result, source_label="勤務表"):
    """Claude APIの結果を統一DataFrameに変換"""
    records = []

    def process_entry(entry):
        name = entry.get("employee_name", "不明")
        for rec in entry.get("records", []):
            try:
                d = datetime.strptime(rec["date"], "%Y-%m-%d").date()
            except (ValueError, KeyError):
                continue

            start = _parse_time_str(rec.get("start_time"))
            end = _parse_time_str(rec.get("end_time"))
            comment = rec.get("comment")
            if comment == "null":
                comment = None
            total_min = _duration_to_minutes(rec.get("total_work_time"))

            if start is None and end is None:
                continue

            records.append({
                "氏名": name,
                "日付": d,
                "出勤時刻": start,
                "退勤時刻": end,
                "コメント": comment,
                "データソース": source_label,
                "総労働時間(分)": total_min,
            })

    if isinstance(claude_result, list):
        for entry in claude_result:
            process_entry(entry)
    else:
        process_entry(claude_result)

    return pd.DataFrame(records, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])


def _parse_time_str(value):
    """HH:MM文字列をdatetime.timeに変換（25:00等の深夜跨ぎ対応）"""
    if not value or value == "null":
        return None
    match = re.match(r"^(\d{1,2}):(\d{2})$", str(value).strip())
    if match:
        hour, minute = int(match.group(1)), int(match.group(2))
        if hour >= 24:
            hour = hour % 24
        try:
            return dt_time(hour, minute)
        except ValueError:
            return None
    return None


def _parse_excel_time(value):
    """Excelセルの時刻値を datetime.time に変換"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, dt_time):
        return value
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime().time().replace(second=0, microsecond=0)
        except Exception:
            pass
    if hasattr(value, "total_seconds"):
        total_minutes = int(value.total_seconds() // 60)
        return dt_time((total_minutes // 60) % 24, total_minutes % 60)

    value_str = str(value).strip()
    if not value_str or value_str in ("nan", "None", "NaT"):
        return None

    match = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", value_str)
    if match:
        hour, minute = int(match.group(1)), int(match.group(2))
        try:
            return dt_time(hour % 24, minute)
        except ValueError:
            return None

    return None


def _parse_excel_date(value):
    """Excelセルの日付値を datetime.date に変換"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime().date()
        except Exception:
            pass

    value_str = str(value).strip()
    if not value_str or value_str in ("nan", "None", "NaT"):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue
    return None


def _parse_hour_minute_cells(hour_value, minute_value):
    """時・分が別セルの値を datetime.time に変換する。"""
    if hour_value is None or minute_value is None:
        return None
    try:
        if pd.isna(hour_value) or pd.isna(minute_value):
            return None
    except (TypeError, ValueError):
        pass

    try:
        hour = int(float(str(hour_value).strip()))
        minute = int(float(str(minute_value).strip()))
    except (TypeError, ValueError):
        return None

    try:
        return dt_time(hour % 24, minute)
    except ValueError:
        return None


def _duration_to_minutes(value):
    """Convert a net work duration to minutes; return None for empty/invalid values."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, dt_time):
        minutes = value.hour * 60 + value.minute + round(value.second / 60)
        return minutes if minutes > 0 else None

    if hasattr(value, "total_seconds"):
        try:
            minutes = round(value.total_seconds() / 60)
            return minutes if minutes > 0 else None
        except (TypeError, ValueError, OverflowError):
            pass

    s = str(value).strip()
    if not s or s in ("nan", "None", "NaT", "-"):
        return None

    m = re.fullmatch(r"(\d{1,3}):(\d{2})(?::(\d{2}))?", s)
    if m:
        seconds = int(m.group(3) or 0)
        minutes = int(m.group(1)) * 60 + int(m.group(2)) + round(seconds / 60)
        return minutes if minutes > 0 else None

    m = re.fullmatch(r"(\d+)\s*時間(?:\s*(\d+)\s*分)?", s)
    if m:
        minutes = int(m.group(1)) * 60 + int(m.group(2) or 0)
        return minutes if minutes > 0 else None

    m = re.fullmatch(r"(\d+)\s*h(?:\s*(\d+)\s*m)?", s, flags=re.IGNORECASE)
    if m:
        minutes = int(m.group(1)) * 60 + int(m.group(2) or 0)
        return minutes if minutes > 0 else None

    try:
        hours = float(s)
    except (TypeError, ValueError):
        return None
    minutes = round(hours * 60)
    return minutes if minutes > 0 else None


def _decimal_hours_to_minutes(value):
    """小数時間（"7.5" / 7.5 / "7:30"）を分に変換する。空・0以下・不正は None。

    請求勤怠ファイルに日別の正味労働時間（NMHTのP列「実働時間」、
    Fieldglassの「エントリ日の労働時間」等）が小数時間で記載されているのを拾う。
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    s = str(value).strip()
    if not s or s in ("nan", "None", "NaT", "-"):
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)  # "7:30" 表記も許容
    if m:
        minutes = int(m.group(1)) * 60 + int(m.group(2))
        return minutes if minutes > 0 else None
    try:
        hours = float(s)
    except (TypeError, ValueError):
        return None
    minutes = round(hours * 60)
    return minutes if minutes > 0 else None


def _hhmm_to_minutes(value):
    """HH:MM 文字列を分に変換する（休憩時間など）。空・不正は 0。"""
    if value is None:
        return 0
    try:
        if pd.isna(value):
            return 0
    except (TypeError, ValueError):
        pass
    m = re.match(r"^(\d{1,2}):(\d{2})$", str(value).strip())
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return 0


def _net_minutes_from_times(start, end, break_minutes=0):
    """datetime.time の出退勤と休憩(分)から正味労働分を返す。

    日跨ぎは翌日退勤として扱う。算出不能・0以下は None（呼び出し側で
    フォールバックさせる）。
    """
    if start is None or end is None:
        return None
    start_min = start.hour * 60 + start.minute
    end_min = end.hour * 60 + end.minute
    if end_min < start_min:
        end_min += 24 * 60
    net = end_min - start_min - max(0, break_minutes or 0)
    return net if net > 0 else None


def _fieldglass_duration_to_minutes(value):
    """Fieldglass PDF の "8h 30m" を分に変換する。"-"/空/0h0m は None。"""
    if not value or value == "-":
        return None
    m = re.fullmatch(r"(\d+)h\s+(\d+)m", str(value).strip())
    if not m:
        return None
    minutes = int(m.group(1)) * 60 + int(m.group(2))
    return minutes if minutes > 0 else None


def _read_text_auto_encoding(filepath):
    """テキストファイルを文字コード自動判定で読み込む"""
    last_error = None
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            with open(filepath, "r", encoding=encoding, newline="") as f:
                return f.read()
        except UnicodeDecodeError as e:
            last_error = e
    raise ValueError(f"テキストの文字コードを判別できませんでした: {last_error}")


def _read_csv_auto_encoding(filepath):
    """CSVを文字コード自動判定で読み込む"""
    last_error = None
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            return pd.read_csv(filepath, encoding=encoding, dtype=object)
        except UnicodeDecodeError as e:
            last_error = e
    raise ValueError(f"CSVの文字コードを判別できませんでした: {last_error}")


def _convert_xlsb_to_xlsx(filepath):
    """Excel COMでxlsbを一時xlsxへ変換する。失敗時はNoneを返す。"""
    if os.path.splitext(filepath)[1].lower() != ".xlsb":
        return None

    fd, converted_path = tempfile.mkstemp(prefix="kintai_xlsb_", suffix=".xlsx")
    os.close(fd)

    excel = None
    workbook = None
    try:
        import pythoncom
        import win32com.client as win32

        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(os.path.abspath(filepath), ReadOnly=True)
        workbook.SaveAs(os.path.abspath(converted_path), FileFormat=51)
        return converted_path
    except Exception:
        logger.exception("xlsbからxlsxへの一時変換に失敗しました: %s", filepath)
        try:
            os.remove(converted_path)
        except OSError:
            pass
        return None
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _extract_itone_name_from_filename(filepath):
    stem = os.path.splitext(os.path.basename(filepath))[0]
    match = re.search(r"(.+?)さん", stem)
    if match:
        return match.group(1).strip()
    match = re.search(r"・([^・()]+)\)_\d{6}$", stem)
    if match:
        return match.group(1).strip()
    return None


def _parse_itone_dispatch_timesheet_file(filepath):
    """IToneの派遣労働者勤務報告書を直接DataFrame化する。"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xlsb"):
        return None

    workbook_path = filepath
    converted_path = None
    if ext == ".xlsb":
        converted_path = _convert_xlsb_to_xlsx(filepath)
        if converted_path is None:
            return None
        workbook_path = converted_path

    try:
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        if "派遣労働者勤務報告書" not in wb.sheetnames:
            return None

        ws = wb["派遣労働者勤務報告書"]
        if str(ws["B12"].value or "").strip() != "日付":
            return None

        employee_name = ws["U6"].value or _extract_itone_name_from_filename(filepath)
        if not employee_name:
            return None
        employee_name = str(employee_name).strip()

        rows = []
        for row_idx in range(13, ws.max_row + 1):
            work_date = _parse_excel_date(ws.cell(row_idx, 2).value)  # B
            if work_date is None:
                continue

            start = _parse_excel_time(ws.cell(row_idx, 24).value)  # X
            end = _parse_excel_time(ws.cell(row_idx, 29).value)  # AC
            if start is None and end is None:
                continue

            comment = ws.cell(row_idx, 12).value or ws.cell(row_idx, 8).value  # L / H
            if comment is not None:
                comment = str(comment).strip() or None
            total_min = _duration_to_minutes(ws[f"CO{row_idx}"].value)

            # CO列は休憩控除後の日別総労働時間
            rows.append({
                "氏名": employee_name,
                "日付": work_date,
                "出勤時刻": start,
                "退勤時刻": end,
                "コメント": comment,
                "データソース": "勤務表",
                "総労働時間(分)": total_min,
            })

        if not rows:
            return None

        return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])
    except Exception:
        logger.exception("ITone派遣労働者勤務報告書の解析に失敗しました: %s", filepath)
        return None
    finally:
        if converted_path:
            try:
                os.remove(converted_path)
            except OSError:
                pass


def _parse_employment_record_file(filepath):
    """Parse the stable 就業記録表 layout without using AI."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xlsb"):
        return None

    workbook_path = filepath
    converted_path = None
    if ext == ".xlsb":
        converted_path = _convert_xlsb_to_xlsx(filepath)
        if converted_path is None:
            return None
        workbook_path = converted_path

    try:
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        if "就業記録表" not in wb.sheetnames:
            return None
        ws = wb["就業記録表"]
        title = re.sub(r"\s+", "", str(ws["A1"].value or ""))
        if title != "就業記録表":
            return None
        if str(ws["A12"].value or "").strip() != "日付":
            return None
        if "開始時刻" not in str(ws["D12"].value or ""):
            return None
        if "終了時刻" not in str(ws["E12"].value or ""):
            return None

        employee_name = str(ws["M3"].value or "").strip()
        if not employee_name:
            return None

        rows = []
        for row_idx in range(13, ws.max_row + 1):
            work_date = _parse_excel_date(ws.cell(row_idx, 1).value)  # A
            if work_date is None:
                continue
            start = _parse_excel_time(ws.cell(row_idx, 4).value)  # D
            end = _parse_excel_time(ws.cell(row_idx, 5).value)  # E
            if start is None and end is None:
                continue

            total_parts = [
                _duration_to_minutes(ws.cell(row_idx, 7).value),  # G: 所定内
                _duration_to_minutes(ws.cell(row_idx, 8).value),  # H: 所定外
            ]
            total_values = [v for v in total_parts if v is not None]
            total_min = sum(total_values) if total_values else None
            comment = ws.cell(row_idx, 10).value  # J
            if comment is not None:
                comment = str(comment).strip() or None

            rows.append({
                "氏名": employee_name,
                "日付": work_date,
                "出勤時刻": start,
                "退勤時刻": end,
                "コメント": comment,
                "データソース": "勤務表",
                "総労働時間(分)": total_min,
            })

        if not rows:
            return None
        return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])
    except Exception:
        logger.exception("就業記録表の解析に失敗しました: %s", filepath)
        return None
    finally:
        if converted_path:
            try:
                os.remove(converted_path)
            except OSError:
                pass


def _find_header_column(ws, rows, names, max_col=40, before_col=None):
    """見出し行から指定ラベルの列番号を探す（見つからなければ None）

    列を決め打ちすると、テンプレートに列が1本挿入されただけで解析できなくなる
    （2026-07 実例: 作業実績報告書 26年度版 v202606 で「勤務内容」列が N に入り、
    実労働ブロックが P→Q へずれて専用パーサーが不発 → AI解析頼みになった）。

    Args:
        rows: 走査する見出し行番号のタプル（左が優先）
        names: 一致させるラベル（前後の空白は無視）
        before_col: この列より左にあるものだけを対象にする
                    （「勤怠区分」は明細用と集計用が2か所にあるため）
    """
    limit = ws.max_column or max_col
    limit = min(limit, max_col)
    if before_col:
        limit = min(limit, before_col - 1)
    for row_idx in rows:
        for col_idx in range(1, limit + 1):
            text = str(ws.cell(row_idx, col_idx).value or "").strip()
            if text and text in names:
                return col_idx
    return None


def _parse_work_result_report_file(filepath):
    """Parse the stable 作業実績報告書 layout without using AI."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xlsb"):
        return None

    workbook_path = filepath
    converted_path = None
    if ext == ".xlsb":
        converted_path = _convert_xlsb_to_xlsx(filepath)
        if converted_path is None:
            return None
        workbook_path = converted_path

    try:
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        if "作業実績報告書" not in wb.sheetnames:
            return None
        ws = wb["作業実績報告書"]
        if str(ws["A1"].value or "").strip() != "作業実績報告書":
            return None
        if str(ws["A14"].value or "").strip() != "日付":
            return None
        # 「実労働」の位置はテンプレート改版で動く（旧版=P / 26年度版 v202606=Q）ため
        # 決め打ちせず見出しから探す。この列があること自体がテンプレートの目印。
        total_col = _find_header_column(ws, (15,), {"実労働"})
        if total_col is None:
            return None

        employee_name = str(ws["D7"].value or "").strip()
        if not employee_name:
            return None

        start_col = _find_header_column(ws, (15,), {"開始時刻"}, before_col=total_col) or 4  # D
        end_col = _find_header_column(ws, (15,), {"終了時刻"}, before_col=total_col) or 5    # E
        # コメントは「勤怠区分」＋「備考」だけを拾う。26年度版で入った「勤務内容」は
        # 毎日同じ値が並ぶだけなので差異一覧のノイズになる＝拾わない。
        # 「勤怠区分」は明細用と集計用の2か所にあるので、実労働より左のものを採る。
        comment_cols = [
            col
            for col in (
                _find_header_column(ws, (14, 15), {"勤怠区分"}, before_col=total_col),
                _find_header_column(ws, (14, 15), {"備考"}, before_col=total_col),
            )
            if col
        ]

        rows = []
        for row_idx in range(16, ws.max_row + 1):
            work_date = _parse_excel_date(ws.cell(row_idx, 1).value)  # A
            if work_date is None:
                continue
            start = _parse_excel_time(ws.cell(row_idx, start_col).value)
            end = _parse_excel_time(ws.cell(row_idx, end_col).value)
            if start is None and end is None:
                continue

            comment_parts = []
            for col_idx in comment_cols:
                value = ws.cell(row_idx, col_idx).value
                if value is None:
                    continue
                text = str(value).strip()
                if text and text not in comment_parts:
                    comment_parts.append(text)

            rows.append({
                "氏名": employee_name,
                "日付": work_date,
                "出勤時刻": start,
                "退勤時刻": end,
                "コメント": " / ".join(comment_parts) or None,
                "データソース": "勤務表",
                "総労働時間(分)": _duration_to_minutes(ws.cell(row_idx, total_col).value),
            })

        if not rows:
            return None
        return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])
    except Exception:
        logger.exception("作業実績報告書の解析に失敗しました: %s", filepath)
        return None
    finally:
        if converted_path:
            try:
                os.remove(converted_path)
            except OSError:
                pass


def _extract_nmht_name_from_filename(filepath):
    stem = os.path.splitext(os.path.basename(filepath))[0]
    match = re.search(r"勤務表[（(]([^）)]+)[）)]", stem)
    if match:
        return match.group(1).strip()
    return _extract_itone_name_from_filename(filepath)


def _parse_nmht_work_time_report_file(filepath):
    """NMHTの勤務時間報告書Ver3系を直接DataFrame化する。"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xlsb"):
        return None

    workbook_path = filepath
    converted_path = None
    if ext == ".xlsb":
        converted_path = _convert_xlsb_to_xlsx(filepath)
        if converted_path is None:
            return None
        workbook_path = converted_path

    try:
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        if "勤務時間報告書" not in wb.sheetnames:
            return None

        ws = wb["勤務時間報告書"]
        version_label = str(ws["F2"].value or "").strip()
        if "勤務時間報告書Ver" not in version_label:
            return None
        if str(ws["C11"].value or "").strip() != "日" or str(ws["E11"].value or "").strip() != "勤務":
            return None

        employee_name = ws["E6"].value or _extract_nmht_name_from_filename(filepath)
        if not employee_name:
            return None
        employee_name = str(employee_name).strip()

        try:
            year = int(float(str(ws["C3"].value).strip()))
            month = int(float(str(ws["E3"].value).strip()))
        except (TypeError, ValueError):
            return None

        rows = []
        for row_idx in range(13, ws.max_row + 1):
            day_value = ws.cell(row_idx, 3).value  # C
            try:
                day = int(float(str(day_value).strip()))
                work_date = date(year, month, day)
            except (TypeError, ValueError):
                continue

            start = _parse_hour_minute_cells(
                ws.cell(row_idx, 6).value,  # F
                ws.cell(row_idx, 7).value,  # G
            )
            end = _parse_hour_minute_cells(
                ws.cell(row_idx, 8).value,  # H
                ws.cell(row_idx, 9).value,  # I
            )
            if start is None and end is None:
                continue

            comment = ws.cell(row_idx, 20).value  # T 備考
            if comment is not None:
                comment = str(comment).strip() or None

            # P列(16)=「実働時間」（休憩控除後の正味、小数時間表記 例:"7.00"）
            total_min = _decimal_hours_to_minutes(ws.cell(row_idx, 16).value)

            rows.append({
                "氏名": employee_name,
                "日付": work_date,
                "出勤時刻": start,
                "退勤時刻": end,
                "コメント": comment,
                "データソース": "勤務表",
                "総労働時間(分)": total_min,
            })

        if not rows:
            return None

        return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])
    except Exception:
        logger.exception("NMHT勤務時間報告書の解析に失敗しました: %s", filepath)
        return None
    finally:
        if converted_path:
            try:
                os.remove(converted_path)
            except OSError:
                pass


def _sap_timesheet_df_from_table(df):
    """SAP/Fieldglass の勤怠取得フォーマット表を統一 DataFrame 化する"""
    required = {"スタッフ", "出勤時刻", "終了時刻", "時間エントリ日"}
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if not required.issubset(set(df.columns)):
        return None

    # 「エントリ日の労働時間 (ブレークダウンなし)」= 日別の正味労働時間（小数時間）
    total_col = next((c for c in df.columns if "エントリ日の労働時間" in c), None)

    rows = []
    for _, row in df.iterrows():
        name = row.get("スタッフ")
        if name is None or str(name).strip() in ("", "nan", "None"):
            continue

        work_date = _parse_excel_date(row.get("時間エントリ日"))
        start = _parse_excel_time(row.get("出勤時刻"))
        end = _parse_excel_time(row.get("終了時刻"))
        if work_date is None or (start is None and end is None):
            continue

        total_min = _decimal_hours_to_minutes(row.get(total_col)) if total_col else None

        # Fieldglassレポートは時刻未確定の日を「00:00〜00:00・24.000h」のプレースホルダで
        # 出力することがある（月末日に多発）。実打刻ではないため時刻・実働は取り込まず、
        # 「特記」付きの行として残す（突合側で突合不可の要確認行になる）。
        if (
            start == dt_time(0, 0)
            and end == dt_time(0, 0)
            and (total_min is None or total_min >= 24 * 60)
        ):
            rows.append({
                "氏名": str(name).strip(),
                "日付": work_date,
                "出勤時刻": None,
                "退勤時刻": None,
                "コメント": None,
                "データソース": "勤務表",
                "総労働時間(分)": None,
                "特記": "Fieldglass時刻なし(00:00-00:00プレースホルダ行)",
            })
            continue

        rows.append({
            "氏名": str(name).strip(),
            "日付": work_date,
            "出勤時刻": start,
            "退勤時刻": end,
            "コメント": None,
            "データソース": "勤務表",
            "総労働時間(分)": total_min,
            "特記": "",
        })

    if not rows:
        return None

    return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)", "特記"])


def _parse_sap_timesheet_file(filepath):
    """SAP/Fieldglass の勤怠取得フォーマットを直接 DataFrame 化する"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        try:
            return _sap_timesheet_df_from_table(_read_csv_auto_encoding(filepath))
        except Exception:
            return None

    if ext not in (".xlsx", ".xls"):
        return None

    try:
        xls = pd.ExcelFile(filepath)
    except Exception:
        return None

    frames = []
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=object)
        except Exception:
            continue

        parsed = _sap_timesheet_df_from_table(df)
        if parsed is not None:
            frames.append(parsed)

    if not frames:
        return None

    result = pd.concat(frames, ignore_index=True)
    return result[["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)", "特記"]]


def _parse_estaffing_timesheet_text(filepath):
    """e-staffing の勤怠処理ツール用テキストを直接 DataFrame 化する"""
    if os.path.splitext(filepath)[1].lower() != ".txt":
        return None

    try:
        text = _read_text_auto_encoding(filepath)
    except Exception:
        return None

    rows = []
    current_name = None
    for fields in csv.reader(text.splitlines(), delimiter="\t"):
        if not fields:
            continue

        row_type = (fields[0] or "").strip()
        if row_type == "H":
            current_name = fields[5].strip() if len(fields) > 5 else None
            continue

        if row_type != "D" or not current_name:
            continue

        work_date = _parse_excel_date(fields[1] if len(fields) > 1 else None)
        start = _parse_excel_time(fields[3] if len(fields) > 3 else None)
        end = _parse_excel_time(fields[4] if len(fields) > 4 else None)
        if work_date is None or (start is None and end is None):
            continue

        comment = None
        if len(fields) > 7 and str(fields[7]).strip():
            comment = str(fields[7]).strip()

        # テキスト版は休憩列が無いため正味は不明（matcher で拘束時間にフォールバック）
        rows.append({
            "氏名": current_name,
            "日付": work_date,
            "出勤時刻": start,
            "退勤時刻": end,
            "コメント": comment,
            "データソース": "勤務表",
            "総労働時間(分)": None,
        })

    if not rows:
        return None

    return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])


def _estaffing_csv_df_from_table(df):
    """e-staffing の請求勤怠 CSV を統一 DataFrame 化する"""
    required = {"スタッフ氏名", "就業年月日", "開始時刻", "終了時刻"}
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if not required.issubset(set(df.columns)):
        return None

    rows = []
    for _, row in df.iterrows():
        name = row.get("スタッフ氏名")
        if name is None or str(name).strip() in ("", "nan", "None"):
            continue

        work_date = _parse_excel_date(row.get("就業年月日"))
        start = _parse_excel_time(row.get("開始時刻"))
        end = _parse_excel_time(row.get("終了時刻"))
        if work_date is None or (start is None and end is None):
            continue

        comment = None
        raw_comment = row.get("備考コメント")
        if raw_comment is not None and str(raw_comment).strip() not in ("", "nan", "None"):
            comment = str(raw_comment).strip()

        # 正味労働 = (終了 − 開始) − 実休憩。e-staffing は総労働列が無いため計算。
        # 実休憩 = 休憩時間(AD列) + 深夜休憩時間(AE列)。夜勤の深夜休憩は別列のため合算必須
        # （AD列のみだと夜勤者の正味が過大になり、偽の総労働時間差異が出る）。
        break_min = _hhmm_to_minutes(row.get("休憩時間")) + _hhmm_to_minutes(row.get("深夜休憩時間"))
        total_min = _net_minutes_from_times(start, end, break_min)

        rows.append({
            "氏名": str(name).strip(),
            "日付": work_date,
            "出勤時刻": start,
            "退勤時刻": end,
            "コメント": comment,
            "データソース": "勤務表",
            "総労働時間(分)": total_min,
        })

    if not rows:
        return None

    return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])


def _parse_estaffing_timesheet_csv(filepath):
    """e-staffing の請求勤怠 CSV を直接 DataFrame 化する"""
    if os.path.splitext(filepath)[1].lower() != ".csv":
        return None

    try:
        return _estaffing_csv_df_from_table(_read_csv_auto_encoding(filepath))
    except Exception:
        return None


def _extract_fieldglass_name_from_filename(filepath):
    stem = os.path.splitext(os.path.basename(filepath))[0]
    match = re.search(r"timesheet_[A-Za-z]+[0-9]+(.+)$", stem, flags=re.IGNORECASE)
    if not match:
        return None

    tail = match.group(1)
    tail = tail.strip(" _,.-")
    if not tail:
        return None

    if re.search(r"[\u3040-\u30ff\u3400-\u9fff]", tail):
        return tail.replace("さん", "").strip()

    tail = tail.replace("_", " ").replace(",", " ").strip()
    tail = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", tail)
    tail = re.sub(r"\s+", " ", tail).strip()
    return tail.upper() if tail else None


def _is_katakana_only_name(name):
    """氏名が（スペースを除き）カタカナのみで構成されるか。

    ERCSTS等の外国人ワーカーは、ファイル名に「ラミタさん」のようなカタカナの
    通称が付くことがある。jinjer側はローマ字（例: MAHARJAN RAMITA）で登録される
    ため、カタカナ通称のままでは突合できない。この判定で通称を検出し、PDF本文の
    ローマ字氏名を優先する。漢字氏名（例: 奈良）はjinjerの漢字氏名と部分一致する
    ため対象外。
    """
    if not name:
        return False
    core = re.sub(r"\s", "", str(name))
    if not core:
        return False
    return all(
        ("゠" <= ch <= "ヿ") or ("ｦ" <= ch <= "ﾟ") or ch == "ー"
        for ch in core
    )


def _choose_fieldglass_name(filename_name, worker_name):
    """Fieldglass勤務表の突合用氏名を決める。

    通常はファイル名由来の氏名を使うが、それがカタカナの通称（jinjerと一致しない）で
    PDF本文にローマ字氏名がある場合は、ローマ字氏名を優先する。
    """
    if filename_name and _is_katakana_only_name(filename_name) and worker_name:
        return worker_name
    return filename_name or worker_name


def _parse_fieldglass_worker_name(text):
    match = re.search(r"\bWorker\s+([^(\n]+)\(", text)
    if not match:
        return None

    raw = match.group(1).strip()
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    if len(parts) == 2:
        raw = f"{parts[1]} {parts[0]}"
    if re.fullmatch(r"[A-Za-z][A-Za-z\s.'-]*", raw):
        return raw.upper()
    return raw


def _fieldglass_date_for_month_day(month, day, period_start, period_end):
    for year in range(period_start.year - 1, period_end.year + 2):
        try:
            candidate = date(year, month, day)
        except ValueError:
            continue
        if period_start <= candidate <= period_end:
            return candidate
    return None


def _fieldglass_dates_from_day_line(line, period_start, period_end):
    dates = []
    for month, day, _weekday in re.findall(r"(\d{1,2})-(\d{2})\s+([A-Za-z]{3})", line):
        dates.append(_fieldglass_date_for_month_day(int(month), int(day), period_start, period_end))
    return dates


def _fieldglass_pad_left(values, dates):
    if len(values) >= len(dates):
        return values[:len(dates)]
    leading_outside_period = 0
    for d in dates:
        if d is None:
            leading_outside_period += 1
        else:
            break
    pad_count = min(len(dates) - len(values), leading_outside_period)
    return (["-"] * pad_count + values)[:len(dates)]


def _fieldglass_12h_to_24h(hour, minute, meridiem):
    """Fieldglass PDF の 12時間制(AM/PM) を "HH:MM"(24時間制) に変換する。

    AM/PM が無い場合はそのまま（既に24時間制とみなす）。
      12:00 AM → 00:00（深夜0時）, 12:00 PM → 12:00（正午）,
      6:00 PM → 18:00, 9:00 AM → 09:00
    """
    if meridiem:
        m = meridiem.upper()
        if m == "AM":
            if hour == 12:
                hour = 0
        elif m == "PM":
            if hour != 12:
                hour += 12
    return f"{hour:02d}:{minute:02d}"


def _fieldglass_time_values(line, prefix, dates):
    # Fieldglass PDF の Time In/Out は "6:00 PM" のような 12時間制(AM/PM)。
    # AM/PM を落とすと 18:00 が 6:00 と誤読され、夜勤跨ぎ判定で +24h されて
    # 30:00 のように化けるため、ここで必ず 24時間制へ変換する。
    body = line[len(prefix):].strip()
    values = []
    for m in re.finditer(r"(\d{1,2}):(\d{2})\s*(AM|PM)?|(-)", body, re.IGNORECASE):
        if m.group(4):  # "-"（打刻なし）
            values.append("-")
        else:
            values.append(
                _fieldglass_12h_to_24h(int(m.group(1)), int(m.group(2)), m.group(3))
            )
    return _fieldglass_pad_left(values, dates)


def _fieldglass_duration_values(line, dates):
    body = line[len("Total"):].strip()
    values = [
        "-" if value == "-" else value
        for value in re.findall(r"-|\d+h\s+\d+m", body)
    ]
    in_period_count = sum(1 for d in dates if d is not None)
    if len(values) in (len(dates) + 1, in_period_count + 1):
        values = values[:-1]
    return _fieldglass_pad_left(values, dates)


def _is_zero_fieldglass_duration(value):
    if not value or value == "-":
        return True
    match = re.fullmatch(r"(\d+)h\s+(\d+)m", value.strip())
    return bool(match and int(match.group(1)) == 0 and int(match.group(2)) == 0)


def _parse_fieldglass_pdf_timesheet(filepath):
    """SAP Fieldglass の Time Sheet PDF を直接 DataFrame 化する"""
    if os.path.splitext(filepath)[1].lower() != ".pdf":
        return None

    try:
        text, pdf_bytes = _pdf_to_text_or_bytes(filepath)
    except Exception:
        return None
    if not text or "Time Sheet" not in text or "Time in/time out" not in text:
        return None

    period_match = re.search(r"\bPeriod\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", text)
    if not period_match:
        return None
    period_start = datetime.strptime(period_match.group(1), "%Y-%m-%d").date()
    period_end = datetime.strptime(period_match.group(2), "%Y-%m-%d").date()

    name = _choose_fieldglass_name(
        _extract_fieldglass_name_from_filename(filepath),
        _parse_fieldglass_worker_name(text),
    )
    if not name:
        return None

    rows = []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for idx, line in enumerate(lines):
        if line != "Time in/time out":
            continue

        day_line = None
        time_in_line = None
        time_out_line = None
        total_line = None
        for candidate in lines[idx + 1: idx + 12]:
            if candidate.startswith("Day "):
                day_line = candidate
            elif candidate.startswith("Time In "):
                time_in_line = candidate
            elif candidate.startswith("Time Out "):
                time_out_line = candidate
            elif candidate.startswith("Total "):
                total_line = candidate
                break

        if not day_line or not time_in_line or not time_out_line:
            continue

        dates = _fieldglass_dates_from_day_line(day_line, period_start, period_end)
        start_values = _fieldglass_time_values(time_in_line, "Time In", dates)
        end_values = _fieldglass_time_values(time_out_line, "Time Out", dates)
        total_values = _fieldglass_duration_values(total_line, dates) if total_line else [""] * len(dates)

        for work_date, start_value, end_value, total_value in zip(dates, start_values, end_values, total_values):
            if work_date is None:
                continue

            start = _parse_time_str(start_value)
            end = _parse_time_str(end_value)
            if start is None and end is None:
                continue
            if _is_zero_fieldglass_duration(total_value):
                continue

            # Total 列（"8h 30m"）= 日別の正味労働時間
            rows.append({
                "氏名": name,
                "日付": work_date,
                "出勤時刻": start,
                "退勤時刻": end,
                "コメント": None,
                "データソース": "勤務表",
                "総労働時間(分)": _fieldglass_duration_to_minutes(total_value),
            })

    if not rows:
        return None

    return pd.DataFrame(rows, columns=["氏名", "日付", "出勤時刻", "退勤時刻", "コメント", "データソース", "総労働時間(分)"])


def _parse_known_timesheet_file(filepath):
    """既知フォーマットをAI解析せずに直接読む"""
    for parser in (
        _parse_nmht_work_time_report_file,
        _parse_itone_dispatch_timesheet_file,
        _parse_employment_record_file,
        _parse_work_result_report_file,
        _parse_sap_timesheet_file,
        _parse_estaffing_timesheet_csv,
        _parse_estaffing_timesheet_text,
        _parse_fieldglass_pdf_timesheet,
    ):
        df = parser(filepath)
        if df is not None:
            return df
    return None


def parse_timesheet(filepath, progress_callback=None):
    """
    勤務表ファイル（Excel/PDF/画像）を解析して統一DataFrameを返す

    progress_callback: 進捗通知用コールバック (任意)

    Note: 後方互換のため "direct" モード相当の挙動のみ。
          記号式シフトの判別が必要な場合は parse_timesheet_smart() を使う。
    """
    ext = os.path.splitext(filepath)[1].lower()
    direct_df = _parse_known_timesheet_file(filepath)
    if direct_df is not None:
        if progress_callback:
            progress_callback(f"勤怠データを直接解析しました: {len(direct_df)}件")
        return direct_df

    if ext in (".xlsx", ".xls", ".xlsb"):
        if progress_callback:
            progress_callback("Excelファイルをテキスト変換中...")
        text = _excel_to_text(filepath)
        result = _parse_with_claude(text, "text")

    elif ext in (".csv", ".txt"):
        if progress_callback:
            progress_callback("テキストファイルを解析中...")
        text = _read_text_auto_encoding(filepath)
        result = _parse_with_claude(text, "text")

    elif ext == ".pdf":
        if progress_callback:
            progress_callback("PDFを解析中...")
        text, pdf_bytes = _pdf_to_text_or_bytes(filepath)
        if text:
            result = _parse_with_claude(text, "text")
        elif image_bytes := _pdf_first_page_to_png_bytes(filepath):
            result = _parse_with_claude(image_bytes, "image", media_type="image/png")
        else:
            result = _parse_with_claude(pdf_bytes, "pdf")

    elif ext in (".png", ".jpg", ".jpeg"):
        if progress_callback:
            progress_callback("画像ファイルを解析中...")
        media_type_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
        with open(filepath, "rb") as f:
            img_bytes = f.read()
        result = _parse_with_claude(img_bytes, "image", media_type=media_type_map[ext])

    else:
        raise ValueError(f"未対応のファイル形式: {ext}")

    return _normalize_records(result)


def _load_file_for_claude(filepath):
    """ファイルから (content, file_type, media_type) を返す"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls", ".xlsb"):
        return _excel_to_text(filepath), "text", None

    if ext in (".csv", ".txt"):
        return _read_text_auto_encoding(filepath), "text", None

    if ext == ".pdf":
        text, pdf_bytes = _pdf_to_text_or_bytes(filepath)
        if text:
            return text, "text", None
        if image_bytes := _pdf_first_page_to_png_bytes(filepath):
            return image_bytes, "image", "image/png"
        return pdf_bytes, "pdf", None

    if ext in (".png", ".jpg", ".jpeg"):
        media_type_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
        with open(filepath, "rb") as f:
            return f.read(), "image", media_type_map[ext]

    raise ValueError(f"未対応のファイル形式: {ext}")


def parse_timesheet_smart(filepath, progress_callback=None):
    """
    勤務表ファイルを自動でモード判定して解析する

    - direct モード: そのまま DataFrame を返す（既存 parse_timesheet と同じ）
    - code   モード: 凡例＋シフト表（生）を返す → ユーザー確認後に shift_resolver で解決

    Returns:
        {
          "mode": "direct" | "code",
          "df": pd.DataFrame,        # direct モードのときのみ
          "legend": [...],           # code モードのときのみ
          "employees": [...],        # code モードのときのみ
          "off_markers": [...],      # code モードのときのみ
          "year": int|None,
          "month": int|None,
          "filename": str,
        }
    """
    if progress_callback:
        progress_callback(f"ファイルを解析中... ({os.path.basename(filepath)})")

    direct_df = _parse_known_timesheet_file(filepath)
    if direct_df is not None:
        return {
            "mode": "direct",
            "df": direct_df,
            "filename": os.path.basename(filepath),
        }

    file_content, file_type, media_type = _load_file_for_claude(filepath)
    fallback_df = None
    try:
        result = parse_with_legend_extraction(file_content, file_type, media_type)
    except Exception:
        logger.exception("凡例対応の勤務表解析に失敗しました。時刻直書き専用解析へフォールバックします。")
        result = {"mode": "direct", "data": []}
        try:
            fallback_df = _normalize_records(_parse_with_claude(file_content, file_type, media_type))
        except Exception:
            logger.exception("時刻直書き専用解析へのフォールバックにも失敗しました。")
            raise

    mode = result.get("mode")
    filename = os.path.basename(filepath)

    if mode == "direct":
        # 既存の direct モード形式に変換して DataFrame 化
        df = _normalize_records(result.get("data") or [])
        if df.empty and fallback_df is None:
            try:
                fallback_df = _normalize_records(_parse_with_claude(file_content, file_type, media_type))
            except Exception:
                logger.exception("空データのため時刻直書き専用解析へフォールバックしましたが失敗しました。")
        if df.empty and fallback_df is not None and not fallback_df.empty:
            df = fallback_df
        if df.empty:
            raise ValueError("AI解析は完了しましたが、出退勤時刻のある勤務行を抽出できませんでした")
        return {
            "mode": "direct",
            "df": df,
            "filename": filename,
        }

    if mode == "code":
        ui_data = to_legend_dict_for_ui(result)
        return {
            "mode": "code",
            "legend": ui_data.get("legend", []),
            "employees": ui_data.get("employees", []),
            "off_markers": ui_data.get("off_markers", []),
            "year": ui_data.get("year"),
            "month": ui_data.get("month"),
            "filename": filename,
        }

    raise RuntimeError(f"未対応の解析モード: {mode}")


def resolve_code_mode_to_df(legend, employees, off_markers=None, source_label="勤務表"):
    """凡例レビュー後の resolve（app.py から呼ばれる）

    Args:
        legend: list of dict（UI で編集された後の凡例）
        employees: list of dict（{name, shifts: [{date, code}]}）
        off_markers: 追加で休扱いとみなす記号セット
        source_label: DataFrame の "データソース" カラム値

    Returns:
        pd.DataFrame（matcher.py に流せる形式）
    """
    return resolve_shifts(legend, employees, off_markers=off_markers, source_label=source_label)
