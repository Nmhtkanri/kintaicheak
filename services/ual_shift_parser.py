"""UAL勤務管理表（KDDI小山）xlsx の構造化パーサ

`Z:\\jinjer移行\\カレンダー\\KDDI小山\\{月}\\UAL勤務管理表*.xlsx` を Claude を経由せず
確定的に解析する。

このブックは **1シート＝1か月** で、過去分のシートを消さずに追加していく運用
（202604, 202605, … 202610 ＋ 年休希望シート・年休まとめ・シフト計算・テンプレ）。
AI読み取りに回すと全シートを一度に渡すことになり、回答が返りきらず時間切れになる
（2026-07-31 に25分以上ハングした実例）。対象月のシートだけを読めばよいので、
構造化パースにする。

シートの構造（YYYYMM シート）:
    B2            = 月初日（datetime）
    C2..AG2       = 日（1〜31）        ※ 日 d は列 d+2
    B4..B14 付近  = 氏名（姓のみ）      ※ 空欄行の次から A/B/明/小計 の集計行
    C4..AG14      = 日別の記号
    D21..D25      = 【ルール】A勤・B勤の時刻

記号:
    A  … A勤（09:00〜17:30 休憩1H）
    B  … B勤（16:45〜33:30 休憩1:45）24時超表記＝翌9:30まで
    明 … 夜勤明け → 「休み」
    ×  … 公休
    年 … 年次有給休暇（全日）
    D  … 特別休暇（集計列「特」に計上される）

⚠️ この表には**他社の方も載っている**。氏名は姓のみで、他社の「小島」が当社の
小島さん(2024044)に名前一致してしまうため、**対象者リストで絞り込むまで
jinjer へ投入してはいけない**（services/employee_alias.py の load_roster_for_source）。
"""

from __future__ import annotations

import calendar
import logging
import re
from datetime import date, datetime

logger = logging.getLogger(__name__)

# シフト表の系統識別子（氏名エイリアス表・対象者リストの適用範囲）
UAL_SOURCE = "kddi_oyama"

# 勤務時間の既定値。ブックの【ルール】欄から読めればそちらを優先する
UAL_A_START, UAL_A_END, UAL_A_BREAK = "9:00", "17:30", 60
UAL_B_START, UAL_B_END, UAL_B_BREAK = "16:45", "33:30", 105

# レイアウト（1始まり）
_COL_NAME = 2          # B列＝氏名
_ROW_DAY_HEADER = 2    # 日付ヘッダー行
_COL_DAY1 = 3          # C列＝1日
_ROW_EMP_START = 4     # 氏名の開始行

# 氏名欄に出てくるが従業員ではない行（集計行）
_SUMMARY_LABELS = {"A", "B", "明", "小計", "合計", "計"}

_SHEET_NAME_RE = re.compile(r"^(20\d{2})(0[1-9]|1[0-2])$")
_RULE_RE = re.compile(
    r"[・･]?\s*([AB])勤\s*[:：]\s*(\d{1,2}:\d{2})\s*[～~〜]\s*(\d{1,2}:\d{2})"
)
_BREAK_RE = re.compile(r"休憩\s*(\d{1,2})\s*(?::(\d{2}))?\s*[Hh時間]?")

# 記号 → (ラベル, 勤務か否か)
_OFF_CODES = {
    "×": "公休",
    "✕": "公休",
    "x": "公休",
    "X": "公休",
}
_AKE_CODE = "明"
_NENKYU_CODE = "年"
_TOKKYU_CODE = "D"


def sheet_name_for(year: int, month: int) -> str:
    """対象年月 → シート名（"202608"）"""
    return f"{year:04d}{month:02d}"


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def _parse_rules(ws) -> dict[str, tuple[str, str, int]]:
    """【ルール】欄から {"A": (開始, 終了, 休憩分)} を読む（読めない項目は既定値）"""
    result = {
        "A": (UAL_A_START, UAL_A_END, UAL_A_BREAK),
        "B": (UAL_B_START, UAL_B_END, UAL_B_BREAK),
    }
    for row in ws.iter_rows(values_only=True):
        for value in row:
            text = _norm(value)
            if not text:
                continue
            m = _RULE_RE.search(text)
            if not m:
                continue
            code, start, end = m.group(1), m.group(2), m.group(3)
            break_minutes = result[code][2] if code in result else 0
            bm = _BREAK_RE.search(text)
            if bm:
                hours = int(bm.group(1))
                minutes = int(bm.group(2) or 0)
                break_minutes = hours * 60 + minutes
            result[code] = (start, end, break_minutes)
            logger.debug("UALルール: %s勤 %s-%s 休憩%d分", code, start, end, break_minutes)
    return result


def build_ual_legend(rules: dict[str, tuple[str, str, int]], seen_codes: set[str]) -> list[dict]:
    """記号 → 凡例エントリ。表に出てきた記号だけを載せる（＋A/Bは常に載せる）"""
    legend: list[dict] = []
    for code in ("A", "B"):
        start, end, break_minutes = rules[code]
        legend.append({
            "code": code,
            "label": f"UAL{code}勤({start}～{end})",
            "start_time": start,
            "end_time": end,
            "break_minutes": break_minutes,
            "is_off": False,
        })
    if _AKE_CODE in seen_codes:
        # label に「明」が入ることで exporter の明け判定（→"休み"）が効く
        legend.append({"code": _AKE_CODE, "label": "夜勤明け", "start_time": None,
                       "end_time": None, "break_minutes": 0, "is_off": True})
    for code in sorted(c for c in seen_codes if c in _OFF_CODES):
        legend.append({"code": code, "label": _OFF_CODES[code], "start_time": None,
                       "end_time": None, "break_minutes": 0, "is_off": True})
    if _NENKYU_CODE in seen_codes:
        # 全日有休は exporter 側で「一般」雛形になる（label の完全一致で判定される）
        legend.append({"code": _NENKYU_CODE, "label": "年次有給休暇", "start_time": None,
                       "end_time": None, "break_minutes": 0, "is_off": True})
    if _TOKKYU_CODE in seen_codes:
        legend.append({"code": _TOKKYU_CODE, "label": "特別休暇", "start_time": None,
                       "end_time": None, "break_minutes": 0, "is_off": True})
    return legend


def _read_day_columns(ws, days_in_month: int) -> dict[int, int]:
    """日付ヘッダー行から {日: 列} を作る。読めない場合は既定レイアウトで補う。"""
    day_col: dict[int, int] = {}
    for col in range(_COL_DAY1, _COL_DAY1 + 40):
        v = _cell(ws, _ROW_DAY_HEADER, col)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            d = int(v)
            if 1 <= d <= 31 and d not in day_col:
                day_col[d] = col
    if len(day_col) < days_in_month:
        # ヘッダーが数式などで読めないケース。C列=1日 の固定レイアウトで補完する
        for d in range(1, days_in_month + 1):
            day_col.setdefault(d, _COL_DAY1 + d - 1)
    return day_col


def parse_ual_worksheet(ws, *, filename: str, year: int, month: int) -> dict:
    """1シート（YYYYMM）→ code_sheet 形式へ解析する純関数

    Raises:
        ValueError: 月初日がシートの内容と一致しない / 従業員行が1件も無い
    """
    days_in_month = calendar.monthrange(year, month)[1]

    # B2 の月初日で「本当にその月のシートか」を検証する（誤月投入の安全弁）
    b2 = _cell(ws, _ROW_DAY_HEADER, _COL_NAME)
    if isinstance(b2, datetime):
        b2 = b2.date()
    if isinstance(b2, date) and (b2.year != year or b2.month != month):
        raise ValueError(
            f"{filename}: シート'{ws.title}'の月初日 {b2} が対象 {year}年{month}月 と一致しません")

    day_col = _read_day_columns(ws, days_in_month)
    rules = _parse_rules(ws)

    employees: list[dict] = []
    seen_codes: set[str] = set()
    blank_run = 0
    for row in range(_ROW_EMP_START, ws.max_row + 1):
        name = _norm(_cell(ws, row, _COL_NAME))
        if not name:
            blank_run += 1
            if blank_run >= 2:
                break          # 氏名欄が続けて空 → 従業員ブロックの終わり
            continue
        if name in _SUMMARY_LABELS:
            break              # A / B / 明 / 小計 の集計行に到達
        blank_run = 0

        shifts = []
        for d in range(1, days_in_month + 1):
            code = _norm(_cell(ws, row, day_col[d]))
            if code:
                seen_codes.add(code)
            shifts.append({"date": date(year, month, d).isoformat(), "code": code})
        employees.append({"name": name, "shifts": shifts})

    if not employees:
        raise ValueError(f"{filename}: シート'{ws.title}'から従業員行を抽出できませんでした")

    legend = build_ual_legend(rules, seen_codes)
    unknown = sorted(c for c in seen_codes
                     if c not in {e["code"] for e in legend})
    if unknown:
        logger.warning("%s: 凡例に無い記号: %s", filename, " / ".join(unknown))

    return {
        "filename": filename,
        "year": year,
        "month": month,
        "legend": legend,
        "employees": employees,
        "off_markers": [c for c in _OFF_CODES if c in seen_codes] or ["×"],
        "source": UAL_SOURCE,
        "unknown_codes": unknown,
        "section_info": {"section_index": None, "sheet": ws.title},
    }


def is_ual_shift_xlsx(filepath: str) -> bool:
    """UAL勤務管理表 xlsx かを軽量判定する（YYYYMM シートと A勤/B勤ルールの有無）"""
    if not str(filepath).lower().endswith((".xlsx", ".xlsm")):
        return False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            month_sheets = [s for s in wb.sheetnames if _SHEET_NAME_RE.match(s)]
            if not month_sheets:
                return False
            ws = wb[month_sheets[0]]
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                for value in row:
                    if value is not None and _RULE_RE.search(str(value)):
                        return True
            return False
        finally:
            wb.close()
    except Exception as e:
        logger.warning("UAL sniff 失敗 %s: %s", filepath, e)
        return False


def parse_ual_shift_xlsx(
    filepath: str,
    target_year: int,
    target_month: int,
) -> dict:
    """UAL勤務管理表 xlsx → code_sheet 形式（対象月のシートのみ読む）

    Raises:
        ValueError: 対象年月のシートが無い / 内容が不整合
    """
    import os
    import openpyxl

    if target_year is None or target_month is None:
        raise ValueError(
            "UAL勤務管理表は対象年月の指定が必要です（シート名 YYYYMM で月を選ぶため）")

    filename = os.path.basename(filepath)
    want = sheet_name_for(target_year, target_month)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        if want not in wb.sheetnames:
            available = " / ".join(s for s in wb.sheetnames if _SHEET_NAME_RE.match(s))
            raise ValueError(
                f"{filename}: 対象月のシート'{want}'がありません（あるのは {available}）")
        result = parse_ual_worksheet(
            wb[want], filename=filename, year=target_year, month=target_month)
    finally:
        wb.close()

    logger.info(
        "UAL勤務管理表を構造化解析: %s シート%s → %d名 凡例%d個",
        filename, want, len(result["employees"]), len(result["legend"]),
    )
    return result
