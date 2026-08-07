"""jinjer給与へのインポート行生成とAPI投入（経費マクロ移植 P2）

集計結果（keihi_classify の bucket）から jinjer給与インポートCSV（11列・谷津さん確定ヘッダー）を作り、
`POST /v1/jinji-imports`（メニュー種別=給与計算、テンプレート「経費APIインポート用」id=44450）で投入する。
※旧テンプレ「経費インポート用」id=37047 は「立替金」裸列が無く計上漏れになるため廃止（2026-07-17 谷津確認）。

列マッピング（live 集計シート実測 2026-07-16。マクロは dRY列振り分け→Module2 の2段変換）:
  - 夜間当番手当       ← F列相当 = 夜間当番(D) + RINK(E)   ※dRYがF値をR/Sペアに入れるため合算値になる
  - 定常外業務対応手当  ← 画面からの手入力（マクロでは「仕訳データ」K/L列 → R/S・T/Uペア）
  - 支給過不足調整     ← **ファイル取込**（1〜30名規模になるため。CSV/xlsxの「社員番号,金額」。負数可）
  - 非課税通勤費       ← 交通費(H)
  - 立替金（顧客請求分）← 顧客請求分(G)
  - 立替金            ← 非課税精算(立替金)(I)
  - その他            ← その他(J)
  - その他手当         ← 画面からの手入力（マクロでは集計シート R/T へ直接手入力）
  - 現物支給          ← 画面からの手入力
  ※テレワーク手当(K) は現行マクロでもインポートCSVに乗らない（dRYでK廃止）。金額があれば警告表示する。

定常外業務対応手当・その他手当・現物支給は経費4ソースから導けない「その月に手で決める」
金額なので画面の手入力欄から受け取る（`parse_manual_allowances`）。支給過不足調整は
対象者が1〜30名規模になり得るためファイル取込にする（`load_allowance_file`）。マクロ側は R/S・T/U の2枠しか無く、
夜間当番手当＋通信手当＋定常外業務対応手当 の3つが揃うと定常外が
「スキップ(両スロット使用済み)」で消える不具合があるが、本実装は項目ごとに独立の
フィールドを持つため取りこぼさない（通信手当はCSV対象外なので枠を食わない）。

投入は人間の最終チェック（プレビューHTML＋CSVダウンロード）を経てから行う（確認後実行）。
"""

from __future__ import annotations

import csv as _csv
import re as _re
import time as _time
from dataclasses import dataclass, field
from pathlib import Path

from services.keihi_classify import B_YAKAN, B_RINK, B_TRANS, B_ETC, B_TW, B_BILL, B_NONTAX

# 谷津さん確定の正式ヘッダー（2026-07-16。jinjerテンプレートと一致させること）
# 社保調整は控除項目（2026-07-21 追加）。テンプレ44450にも列を追加してもらう。
IMPORT_HEADERS = [
    "社員番号", "氏名", "夜間当番手当", "定常外業務対応手当", "支給過不足調整",
    "非課税通勤費", "立替金（顧客請求分）", "立替金", "その他", "その他手当", "現物支給",
    "社保調整",
]
# インポート行が持つ項目（社員番号・氏名以外）。テンプレ列名との照合対象。
ITEM_KEYS = [
    "夜間当番手当", "定常外業務対応手当", "支給過不足調整", "非課税通勤費",
    "立替金（顧客請求分）", "立替金", "その他", "その他手当", "現物支給", "社保調整",
    "テレワーク手当",
]
# 経費4ソースから導けず、画面の「イレギュラー経費」から手で入れる項目。
# 金額は負数可（支給過不足調整の戻し等）。社保調整は控除項目。
MANUAL_ITEM_KEYS = [
    "定常外業務対応手当", "その他手当", "現物支給", "支給過不足調整", "社保調整",
]
# jinjer側テンプレート（GET /v1/master/jinji-import-templates で実測）
DEFAULT_TEMPLATE_ID = "44450"   # 「経費APIインポート用」menu=payroll_calculation（11列・立替金裸列あり・2026-07-17検証済）


def _norm_col(s) -> str:
    """列名の照合キー。空白・全角スペース・先頭の必須マーク「*」を除去。"""
    return _re.sub(r"[\s　*＊]", "", "" if s is None else str(s)).strip()


def _match_column(col: str) -> "str | None":
    """テンプレの列名 → インポート行のどのキーに対応するか。空/不明は None。"""
    n = _norm_col(col)
    if n == "":
        return None
    if n in ("社員番号", "従業員番号"):
        return "社員番号"
    if n in ("氏名", "名前"):
        return "氏名"
    for k in ITEM_KEYS:
        if _norm_col(k) == n:
            return k
    return None


def read_template_header(path: "str | Path") -> list[str]:
    """jinjerからダウンロードした空テンプレCSVの先頭行（列名の並び）を返す。

    列位置対応のインポートに合わせるため、空セルもそのまま（"" として）保持する。
    """
    path = Path(path)
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            with open(path, encoding=enc, newline="") as f:
                header = next(_csv.reader(f))
            return [("" if c is None else str(c)).strip() for c in header]
        except (UnicodeDecodeError, StopIteration):
            continue
    raise ValueError(f"テンプレートCSVを読めませんでした: {path}")


def _parse_amount(s) -> "float | None":
    """金額文字列を数値化する（VBA ValJP 相当）。全角・¥・円・カンマ・(負数) に対応。"""
    import unicodedata as _ud
    t = _ud.normalize("NFKC", "" if s is None else str(s)).strip()
    for ch in ("¥", "￥", "円", ",", " ", "　"):
        t = t.replace(ch, "")
    if t == "":
        return None
    if len(t) >= 2 and t.startswith("(") and t.endswith(")"):
        t = "-" + t[1:-1]
    try:
        return float(t)
    except ValueError:
        return None


def parse_manual_allowances(text: "str | None") -> "tuple[dict[str, int], list[str]]":
    """「社員番号,金額」形式の手入力テキストを {社員番号: 金額} にする。

    定常外業務対応手当・その他手当は経費から導けない手決めの金額なので画面入力を受ける。
    区切りは カンマ / タブ / 全角カンマ（Excelの2列をそのままコピペできる）。
    同一社員の複数行は加算（マクロ「仕訳データ振り分け」と同じ挙動）。
    空行・「#」開始行・ヘッダー行は無視。読めなかった行は errors に理由を返す。
    """
    from services.keihi_summary import excel_coerce, in_company_scope
    out: dict[str, int] = {}
    errors: list[str] = []
    for lineno, raw in enumerate((text or "").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        # タブがあればタブ区切り（Excelの2列コピペ）。無ければカンマ区切り。
        # 金額の桁区切りカンマ（8,000）で割れるため、2列目以降は結合してから数値化する。
        parts = line.split("\t") if "\t" in line else _re.split(r"[,、，]", line)
        if len(parts) < 2:
            errors.append(f"{lineno}行目: 「社員番号,金額」の形式で入力してください → {line!r}")
            continue
        emp = excel_coerce(parts[0].strip())    # "02026012" → "2026012"
        amt = _parse_amount("".join(p.strip() for p in parts[1:]))
        is_id = bool(_re.fullmatch(r"\d+", emp))
        if amt is None:
            if not is_id:
                continue    # 「社員番号,金額」等のヘッダー行 → 無視
            errors.append(f"{lineno}行目: 金額を数値として読めません → {parts[1].strip()!r}")
            continue
        if not is_id:
            errors.append(f"{lineno}行目: 社員番号が数字ではありません → {parts[0].strip()!r}")
            continue
        if not in_company_scope(emp):
            errors.append(f"{lineno}行目: {emp} は給与計算対象外（5/6/9始まり）のため除外しました")
            continue
        if amt == 0:
            continue        # 0 は入れても結果が変わらないので無視
        out[emp] = out.get(emp, 0) + int(round(amt))
    return out, errors


# 一括取込ファイルの列名候補（ヘッダーから自動判別する）
_ID_HEADERS = ("社員番号", "従業員番号", "社員no", "社員ｎｏ", "empno", "employee_id")
_AMT_HEADERS = ("金額", "支給額", "調整額", "amount")
_TYPE_HEADERS = ("項目", "種類", "経費種類", "費目", "手当種類", "項目名")
_ITEM_BY_NORM = {_norm_col(k): k for k in MANUAL_ITEM_KEYS}


def merge_manual(dst: dict, item: str, add: dict) -> None:
    """{項目: {社員番号: 金額}} に加算マージする（同一社員は合算）。"""
    cur = dst.setdefault(item, {})
    for emp, amt in add.items():
        cur[emp] = cur.get(emp, 0) + amt


def _read_table(path: "str | Path", label: str) -> list[tuple]:
    """CSV(cp932/utf-8) または xlsx を2次元のタプル列として読む（空行は除去）。"""
    path = Path(path)
    rows: list[tuple] = []
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = [r for r in wb.worksheets[0].iter_rows(values_only=True) if r is not None]
        finally:
            wb.close()
    else:
        for enc in ("cp932", "utf-8-sig", "utf-8"):
            try:
                with open(path, encoding=enc, newline="") as f:
                    rows = [tuple(r) for r in _csv.reader(f)]
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"{label}ファイルの文字コードを判別できませんでした: {path}")
    return [r for r in rows if any(c is not None and str(c).strip() for c in r)]


def load_irregular_file(path: "str | Path", label: str = "イレギュラー経費",
                        ) -> "tuple[dict[str, dict[str, int]], list[str]]":
    """イレギュラー経費を一括取込する。{項目: {社員番号: 金額}} を返す。

    ヘッダーから2つの形式を自動判別する（どちらもExcelでそのまま作れる）:

    ワイド形式（推奨・出力CSVと同じ並び）
        社員番号, 定常外業務対応手当, その他手当, 現物支給, 支給過不足調整, 社保調整
        2026013,  15000,             5000,       3000,     -5000,          2000

    ロング形式（1行1明細）
        社員番号, 項目,       金額
        2026013,  現物支給,   3000

    金額は負数可。同一社員・同一項目の複数行は加算。
    """
    rows = _read_table(path, label)
    if not rows:
        return {}, [f"{label}ファイルにデータがありません: {path}"]

    head_raw = [_norm_col(c) for c in rows[0]]
    head_low = [h.lower() for h in head_raw]
    id_col = next((i for i, h in enumerate(head_low) if h in _ID_HEADERS), None)
    item_cols = {i: _ITEM_BY_NORM[h] for i, h in enumerate(head_raw) if h in _ITEM_BY_NORM}
    type_col = next((i for i, h in enumerate(head_raw) if h in _TYPE_HEADERS), None)
    amt_col = next((i for i, h in enumerate(head_low) if h in _AMT_HEADERS), None)
    if id_col is None:
        id_col = 0
    body = rows[1:]

    out: dict[str, dict[str, int]] = {}
    errors: list[str] = []

    if item_cols:
        # ワイド形式: 項目名の列ごとに読む
        for ci, item in sorted(item_cols.items()):
            text = "\n".join(f"{r[id_col]}\t{r[ci]}" for r in body
                             if len(r) > max(id_col, ci))
            got, errs = parse_manual_allowances(text)
            if got:
                merge_manual(out, item, got)
            errors += [f"{label}ファイル（{item}）{e}" for e in errs]
    elif type_col is not None and amt_col is not None:
        # ロング形式: 項目列でグルーピングしてから読む
        groups: dict[str, list[str]] = {}
        for n, r in enumerate(body, start=2):
            if len(r) <= max(id_col, type_col, amt_col):
                continue
            item = _ITEM_BY_NORM.get(_norm_col(r[type_col]))
            if item is None:
                name = str(r[type_col] or "").strip()
                if name:
                    errors.append(
                        f"{label}ファイル {n}行目: 項目「{name}」は選択できる項目ではありません"
                        f"（{' / '.join(MANUAL_ITEM_KEYS)}）")
                continue
            groups.setdefault(item, []).append(f"{r[id_col]}\t{r[amt_col]}")
        for item, lines in groups.items():
            got, errs = parse_manual_allowances("\n".join(lines))
            if got:
                merge_manual(out, item, got)
            errors += [f"{label}ファイル（{item}）{e}" for e in errs]
    else:
        errors.append(
            f"{label}ファイルの見出しから項目を判別できませんでした: {path}\n"
            f"　　ワイド形式（社員番号＋項目名の列）またはロング形式（社員番号・項目・金額）"
            f"にしてください。項目名は {' / '.join(MANUAL_ITEM_KEYS)} のいずれかです。")
    return out, errors


@dataclass
class RouteMovePlan:
    """経路突合レビューの選択を、インポート行に反映するための差分。

    集計シート（マクロとの全行一致検証に使う）は動かさず、インポート行の
    非課税通勤費／立替金だけを付け替える（2026-08-06 の上限カットと同じ方針）。
    """
    to_nontax: dict = field(default_factory=dict)      # 社員番号 → 非課税通勤費→立替金 へ移す額
    to_trans: dict = field(default_factory=dict)       # 社員番号 → 立替金→非課税通勤費 へ移す額
    commute_delta: dict = field(default_factory=dict)  # commute_by_id への加算（±・上限判定の材料）
    other_delta: dict = field(default_factory=dict)    # commute_other_by_id への加算（±）
    months_add: dict = field(default_factory=dict)     # 社員番号 → 追加する利用月の集合
    reviewed_emps: set = field(default_factory=set)    # 選択が確定した社員（自動付替を抑止）
    details: list = field(default_factory=list)        # 明細（console・画面表示用）


def plan_route_moves(matched: list, cls_log: list,
                     travel_members: "dict | None" = None) -> tuple[RouteMovePlan, list]:
    """経路突合レビューの [(行, 選択)] から付け替えの差分を組み立てる。

    人間が「通勤費」と判定した金額は、上限3万円の判定にも入れる
    （利用月も複数月警告の材料に加える。2026-08-07 谷津さん決定）。

    整合が取れない選択（付け替えできない行への指定・対象者の行の混入）は
    エラーにして呼び出し側で停止させる。プレビューが古いまま確定された印なので、
    黙って無視すると人が見ていない状態で計上先が決まってしまう。
    """
    from services.keihi_classify import (
        COMMUTE_LIMIT_KEYWORDS, MOVABLE_SIDES, SIDE_TRANS, SIDE_TRAVEL,
        _use_month, entry_side, main_entry_by_row, side_by_row,
    )
    travel_members = travel_members or {}
    plan = RouteMovePlan()
    errors: list[str] = []
    main = main_entry_by_row(cls_log)
    sides = side_by_row(cls_log)

    for r, choice in matched:
        emp = str(r.get("社員番号") or "")
        name = str(r.get("氏名") or "")
        label = f"{emp} {name} {r.get('利用日', '')} {r.get('交通機関', '')}".strip()
        if emp in travel_members:
            errors.append(f"{label}: 移動交通費対象者の行は自動で立替金に計上します"
                          f"（レビュー対象外）。もう一度実行してやり直してください。")
            continue
        row_no = r.get("行番号")
        entry = main.get(row_no)
        side = entry_side(entry)
        why = sides.get(row_no, "対象外（計上なし）")
        if side not in MOVABLE_SIDES:
            # 顧客請求分・夜間当番手当など。画面では固定表示（対象外）にしてあるので、
            # 通勤費/移動交通費 が選ばれていたらプレビューが古い＝突合が壊れている印
            if choice != "対象外":
                errors.append(f"{label}: この行は{why}のため計上先を変えられません。"
                              f"もう一度実行してやり直してください。")
            else:
                r["人間判定"] = why
                r["計上先変更"] = "変更なし"
            continue
        if choice == "対象外":
            errors.append(f"{label}: 計上先（通勤費 / 移動交通費）を選んでください。")
            continue

        plan.reviewed_emps.add(emp)
        amt = int(round(float(entry.amount or 0)))
        if choice == side or amt == 0:
            plan.details.append({
                "社員番号": emp, "氏名": name, "利用日": r.get("利用日", ""),
                "交通機関": r.get("交通機関", ""), "金額": amt,
                "変更": "変更なし", "理由": "人間判定",
            })
            r["人間判定"] = choice
            r["計上先変更"] = "変更なし"
            continue

        if side == SIDE_TRANS:      # 通勤費 → 移動交通費（非課税通勤費 → 立替金）
            plan.to_nontax[emp] = plan.to_nontax.get(emp, 0) + amt
            tgt = (plan.commute_delta if str(entry.matched_kw or "") in COMMUTE_LIMIT_KEYWORDS
                   else plan.other_delta)
            tgt[emp] = tgt.get(emp, 0) - amt
            change = "非課税通勤費 → 立替金"
        else:                       # 移動交通費 → 通勤費（立替金 → 非課税通勤費）
            plan.to_trans[emp] = plan.to_trans.get(emp, 0) + amt
            plan.commute_delta[emp] = plan.commute_delta.get(emp, 0) + amt
            ym = _use_month(r.get("利用日"))
            if ym:
                plan.months_add.setdefault(emp, set()).add(ym)
            change = "立替金 → 非課税通勤費"

        plan.details.append({
            "社員番号": emp, "氏名": name, "利用日": r.get("利用日", ""),
            "交通機関": r.get("交通機関", ""), "金額": amt,
            "変更": change, "理由": "人間判定",
        })
        r["人間判定"] = choice
        r["計上先変更"] = f"{change} {amt:,}円"

    if errors:
        return RouteMovePlan(), errors
    return plan, []


def build_import_rows(by_id: dict, emp_names: dict, roster_names: "dict | None" = None,
                      manual: "dict | None" = None,
                      commute_by_id: "dict | None" = None,
                      commute_other_by_id: "dict | None" = None,
                      commute_limit: "int | None" = None,
                      limit_exempt: "dict | None" = None,
                      travel_members: "dict | None" = None,
                      commute_months: "dict | None" = None,
                      route_plan: "RouteMovePlan | None" = None,
                      moves: "list | None" = None,
                      ) -> tuple[list[dict], list[str], list[dict]]:
    """集計（数字ID→7バケット）＋イレギュラー経費からインポート行・警告・上限カット明細を返す。

    行は 20YY 始まりの社員のみ・社員番号昇順（5/6/9始まりは給与計算対象外）。
    manual は {項目名: {社員番号: 金額}}（項目名は MANUAL_ITEM_KEYS）。
    **経費が無くイレギュラー経費だけの社員も行を出す**（マクロの「A列に無い社員は
    無言スキップ」で起きた計上漏れ事故と同じ轍を踏まないため）。

    commute_limit を渡すと**非課税通勤費に月額上限を適用する**（2026-08-06 谷津さん指定）。
    切るのは交通費(H)のうち通勤系と確信できる分（commute_by_id）だけで、同じHに入る
    駐車場代など（commute_other_by_id）は上限の対象外として素通しする。
    上限が掛からないのは limit_exempt（個別許可）。
    カットは黙って行わず、必ず warnings と cuts の両方に出して投入前に人が見られるようにする。

    travel_members（移動交通費＝立替精算の対象者）の通勤系申請分は、非課税通勤費ではなく
    立替金に計上する（2026-08-07 谷津さん指定）。付け替え後は通勤費0になるので上限も掛からない。
    route_plan（経路突合レビューで人が選んだ結果）がある社員はそちらを正とし、二重に動かさない。
    付け替えも黙って行わず warnings と moves に出す。
    """
    from services.keihi_summary import in_company_scope
    roster_names = roster_names or {}
    manual = manual or {}
    limit_exempt = limit_exempt or {}
    travel_members = travel_members or {}
    # 付け替えで上限判定の材料（通勤費分・利用月）も動く。呼び出し元の dict は壊さずコピーで扱う
    commute_by_id = dict(commute_by_id or {})
    commute_other_by_id = dict(commute_other_by_id or {})
    commute_months = {k: set(v) for k, v in (commute_months or {}).items()}
    if route_plan:
        for _emp, _d in route_plan.commute_delta.items():
            commute_by_id[_emp] = commute_by_id.get(_emp, 0.0) + _d
        for _emp, _d in route_plan.other_delta.items():
            commute_other_by_id[_emp] = commute_other_by_id.get(_emp, 0.0) + _d
        for _emp, _ms in route_plan.months_add.items():
            commute_months.setdefault(_emp, set()).update(_ms)
        if moves is not None:
            moves.extend(route_plan.details)
    manual_ids: set = set()
    for _d in manual.values():
        manual_ids |= set(_d or {})

    rows: list[dict] = []
    warnings: list[str] = []
    cuts: list[dict] = []
    for emp_id in sorted(set(by_id) | manual_ids):
        if not in_company_scope(emp_id):
            continue
        v = by_id.get(emp_id) or [0.0] * 7
        name = emp_names.get(emp_id) or roster_names.get(emp_id, "")
        trans = int(v[B_TRANS])
        nontax = int(v[B_NONTAX])
        if route_plan:
            # 人間が選んだ分を付け替える（合計は不変＝集計シートの判定列に影響しない）
            _out = route_plan.to_nontax.get(emp_id, 0)
            _in = route_plan.to_trans.get(emp_id, 0)
            trans += _in - _out
            nontax += _out - _in
        if emp_id in travel_members and not (route_plan and emp_id in route_plan.reviewed_emps):
            auto = int(round(commute_by_id.get(emp_id, 0)))
            if auto:
                trans -= auto
                nontax += auto
                commute_by_id[emp_id] = 0.0     # 付け替え済み＝上限判定の対象から外す
                if moves is not None:
                    moves.append({
                        "社員番号": emp_id, "氏名": name, "利用日": "", "交通機関": "",
                        "金額": auto, "変更": "非課税通勤費 → 立替金",
                        "理由": "移動交通費対象者（自動）",
                    })
                warnings.append(
                    f"ℹ️ {emp_id} {name}: 通勤費 {auto:,}円 を立替金（移動交通費）へ"
                    f"付け替えました（移動交通費（立替精算）対象者）。")
        if commute_limit:
            trans = _apply_commute_limit(
                emp_id, name, trans, commute_by_id, commute_other_by_id,
                commute_limit, limit_exempt, travel_members,
                commute_months.get(emp_id) or set(), cuts, warnings)
        row = {
            "社員番号": emp_id,
            "氏名": name,
            "夜間当番手当": int(v[B_YAKAN] + v[B_RINK]),   # F列相当（夜間＋RINK）
            "非課税通勤費": trans,
            "立替金（顧客請求分）": int(v[B_BILL]),
            "立替金": nontax,
            "その他": int(v[B_ETC]),
            "テレワーク手当": int(v[B_TW]),
        }
        for key in MANUAL_ITEM_KEYS:
            row[key] = int((manual.get(key) or {}).get(emp_id, 0))
        rows.append(row)

    # 手入力の社員番号が在籍者一覧に無ければ入力ミスの可能性 → 警告（行は出す）
    for emp_id in sorted(manual_ids):
        if in_company_scope(emp_id) and roster_names and emp_id not in roster_names:
            warnings.append(
                f"⚠️ イレギュラー経費: 社員番号 {emp_id} が在籍者一覧にありません（入力ミスの可能性）。")
    if cuts:
        total = sum(c["カット額"] for c in cuts)
        warnings.insert(0, f"✂️ 通勤費の上限カット: {len(cuts)}名 / 計 {total:,}円 を減額しました。"
                           f"投入前に下の明細を必ず確認してください。")
    moved = [m for m in (moves or []) if m.get("変更") != "変更なし"]
    if moved:
        total = sum(m["金額"] for m in moved)
        warnings.insert(0, f"🔁 通勤費↔立替金の付け替え: {len(moved)}件 / 計 {total:,}円。"
                           f"人件費区分が本社の方は経理仕訳で立替金が計上されない仕様なので、"
                           f"該当者がいれば経理モードの警告も確認してください。")
    return rows, warnings, cuts


def _apply_commute_limit(emp_id: str, name: str, trans: int,
                         commute_by_id: dict, commute_other_by_id: dict,
                         limit: int, limit_exempt: dict, travel_members: dict,
                         months: set, cuts: list, warnings: list) -> int:
    """非課税通勤費に月額上限を適用した金額を返す（対象外ならそのまま返す）。"""
    commute = int(round(commute_by_id.get(emp_id, 0)))
    other = int(round(commute_other_by_id.get(emp_id, 0)))
    if commute <= limit:
        return trans
    if len(months) > 1:
        # 定期代を複数月分まとめて申請している。合計は上限超過でも月ごとなら
        # 上限内のことがあり、切ると過少支給になる。自動では切らず人が判断する。
        warnings.append(
            f"⚠️ {emp_id} {name}: 通勤費 {commute:,}円 が上限超過ですが、利用日が "
            f"{'・'.join(sorted(months))} の {len(months)}か月分に分かれています。"
            f"月ごとなら上限内の可能性があるため**自動ではカットしていません**。"
            f"1か月分だけを当月に計上すべきかを確認してください。")
        return trans
    if emp_id in travel_members:
        warnings.append(f"ℹ️ {emp_id} {name}: 通勤費 {commute:,}円 は上限超過ですが、"
                        f"移動交通費（立替精算）対象者のためカットしていません。")
        return trans
    if emp_id in limit_exempt:
        warnings.append(f"ℹ️ {emp_id} {name}: 通勤費 {commute:,}円 は上限超過ですが、"
                        f"上限免除者のためカットしていません。")
        return trans
    # 交通費(H) と通勤費分がズレる＝Hに通勤系と判定できない金額が混ざっている。
    # その分は上限の対象外として残す（切りすぎて過少支給にしない）。
    cut = commute - limit
    after = trans - cut
    cuts.append({
        "社員番号": emp_id, "氏名": name,
        "交通費(H)": trans, "うち通勤費": commute, "うち上限対象外": other,
        "上限": limit, "カット額": cut, "カット後": after,
    })
    if other:
        warnings.append(f"ℹ️ {emp_id} {name}: 交通費(H) のうち {other:,}円 は通勤系と判定できないため"
                        f"（駐車場代など）上限の対象外として残しました。")
    return after


def _escape_csv(s: str) -> str:
    """Module2.EscapeCSV: 文字列は必ずダブルクォート囲み・内部の " は "" に。"""
    return '"' + str(s).replace('"', '""') + '"'


def check_template_coverage(rows: list[dict], template_header: list[str]) -> list[str]:
    """テンプレに列が無いのに金額が発生している支給項目を警告として返す。

    位置対応インポートでは、テンプレに列が無い項目はサイレントに捨てられて計上漏れになる
    （今回の 川口さん立替金12,177円・夜間当番手当30,000円の事故がこれ）。事前に検知する。
    """
    covered = {_match_column(c) for c in template_header}
    warnings: list[str] = []
    for key in ITEM_KEYS:
        if key in covered:
            continue
        total = sum(int(r.get(key, 0) or 0) for r in rows)
        n = sum(1 for r in rows if int(r.get(key, 0) or 0))
        if total:
            warnings.append(
                f"⚠️ テンプレに「{key}」の列がありません → {n}名・計{total:,}円が"
                f"取り込まれず計上漏れになります。jinjer側テンプレートに「{key}」列を追加してください。")
    return warnings


def render_import_csv(rows: list[dict], template_header: "list[str] | None" = None) -> bytes:
    """インポートCSVを Shift-JIS バイト列で返す（Module2 と同形式: 文字列のみクォート）。

    template_header を渡すと、その列名の並び（空セルも保持）に合わせて列を組み立てる。
    jinjer のテンプレインポートは列名でなく**列位置**で対応付けるため、テンプレの並びに
    追従させることが必須（並びがズレると値が別項目に入る）。未指定時は IMPORT_HEADERS。
    """
    header = template_header if template_header is not None else IMPORT_HEADERS
    lines = [",".join(header)]   # ヘッダーは列名をそのまま（テンプレの見出しを忠実に）
    for r in rows:
        cells = []
        for col in header:
            key = _match_column(col)
            if key in ("社員番号", "氏名"):
                cells.append(_escape_csv(r.get(key, "")))
            elif key in ITEM_KEYS:
                cells.append(str(int(r.get(key, 0) or 0)))
            else:
                cells.append("")   # 空列・不明列はそのまま空（テンプレのスキップ列）
        lines.append(",".join(cells))
    return ("\r\n".join(lines) + "\r\n").encode("cp932")


def write_import_csv(rows: list[dict], out_path: "str | Path",
                     template_header: "list[str] | None" = None) -> Path:
    """インポートCSVをファイルに書き出す（人間チェック用ダウンロード）。"""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(render_import_csv(rows, template_header))
    return out_path


# ----------------------------------------------------------------------
# jinjer API 投入（確認後実行）
# ----------------------------------------------------------------------

@dataclass
class PayrollImportResult:
    ok: bool
    month: str = ""
    file_name: str = ""
    status: str = ""      # "1"=全行成功 / "2"=一部失敗 / "0"=処理待ち・不明
    total_rows: int = 0
    failed_rows: int = 0
    import_id: str = ""
    error: str = ""
    logs: list = field(default_factory=list)


def post_payroll_import(
    client,
    month: str,
    csv_bytes: bytes,
    file_name: str,
    template_id: str = DEFAULT_TEMPLATE_ID,
    apply_formulas_off: bool = True,
    poll_seconds: int = 180,
    log_func=print,
) -> PayrollImportResult:
    """給与計算インポートを `POST /v1/jinji-imports` で投入し、完了をポーリングする。

    Args:
        client: JinjerClient（認証済み）
        month: 処理月 "YYYY-MM"（salary_setting.executed_on）
        csv_bytes: Shift-JIS のインポートCSV
        file_name: インポートファイル名
        template_id: 入力テンプレートID（既定 44450「経費APIインポート用」）
        apply_formulas_off: 「編集した項目は計算式を適用しない」を有効にするか
        poll_seconds: 完了待ちの最大秒数（バッチ処理のため時間がかかる）
    """
    result = PayrollImportResult(ok=False, month=month, file_name=file_name)

    def log(msg):
        result.logs.append(msg)
        log_func(msg)

    try:
        data = client.post_jinji_import(
            csv_bytes=csv_bytes, file_name=file_name, template_id=template_id,
            executed_on=month, apply_formulas_off=apply_formulas_off,
        )
        result.import_id = str(data.get("id") or "")
        log(f"[info] jinji-imports 投入予約 OK: {file_name} "
            f"(処理月 {month} / テンプレ {template_id} / id={result.import_id or '不明'})")
    except Exception as e:  # noqa: BLE001
        result.error = f"投入予約に失敗しました: {e}"
        log(f"[error] {result.error}")
        return result

    # ポーリング。jinji-imports の GET には status フィールドが無く、
    # number_of_failed_rows / number_of_total_rows で成否を判定する（2026-07-16 実測。
    # updated_at が入ればバッチ処理完了とみなす）。
    if not result.import_id:
        result.ok = True
        log("[warn] 投入予約は成功しましたが id が取得できず完了確認をスキップします。"
            "jinjer画面のインポート履歴をご確認ください。")
        return result

    waited = 0
    interval = 10
    while waited < poll_seconds:
        _time.sleep(interval)
        waited += interval
        try:
            item = client.find_jinji_import(result.import_id)
        except Exception as e:  # noqa: BLE001
            log(f"[warn] 状況確認に失敗（リトライ）: {e}")
            continue
        if item and str(item.get("updated_at") or "").strip():
            result.total_rows = int(item.get("number_of_total_rows") or 0)
            result.failed_rows = int(item.get("number_of_failed_rows") or 0)
            if result.failed_rows == 0:
                result.status = "1"
                result.ok = True
                log(f"[done] インポート成功: {result.total_rows}行すべて取込（{waited}秒）。"
                    "jinjer給与画面で金額をご確認ください。")
            else:
                result.status = "2"
                result.ok = False
                result.error = (
                    f"{result.total_rows}行中 {result.failed_rows}行が失敗しました。"
                    "通知メールとjinjer画面のインポート履歴でエラー行をご確認ください。")
                log(f"[error] {result.error}")
            return result
        log(f"[info] 処理待ち... ({waited}秒)")

    result.status = "0"
    result.ok = True  # 予約自体は成功（完了待ちタイムアウトは失敗扱いにしない）
    log("[warn] 完了確認がタイムアウトしました（投入予約は成功）。"
        "後ほどjinjer画面のインポート履歴か通知メールをご確認ください。")
    return result
