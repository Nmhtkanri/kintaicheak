# -*- coding: utf-8 -*-
"""kintai_import_runner — アップロード用CSVを jinjer API (kintai-imports) で投入し、反映を検証する

手順③（quick_export が生成したアップロード用CSV）の後工程:

    1. ガード     : 休暇登録がある日の行を除外（jinjerはサイレントに上書きしないため）
    2. 投入       : POST /v1/kintai-imports（種別5、5000行以内・超過時は分割、同時予約1件）
    3. 追跡       : GET /v1/kintai-imports を20秒間隔でポーリング（0=予約/1=成功/2=失敗）
    4. 検証       : work-schedules / attendances API で意図値と現物を1行ずつ突合
    5. レポート   : Excelブック出力
                    - 手動対応リスト（従業員番号/氏名/日付/区分/備考）… 谷津さん指定の必須成果物
                    - 検証結果（全行）
                    - 取込ログ

設計の背景は docs/PLAN_手順3_API直接投入.md を参照。
実証元: Z:\\API連携\\scripts\\fix_schedule_via_api.py / month_schedule_via_api.py（2026-07-09）
"""
from __future__ import annotations

import csv
import io
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from services.jinjer_api_client import JinjerAPIError, JinjerClient

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
COL_NAME = "名前"
COL_EMP = "*従業員ID"
COL_DATE = "*年月日"
COL_SCHED_IN = "出勤予定時刻"
COL_SCHED_OUT = "退勤予定時刻"
COL_PUNCH_IN = "出勤1"
COL_PUNCH_OUT = "退勤1"
BREAK_PAIRS = [(f"休憩予定時刻{i}", f"復帰予定時刻{i}") for i in range(1, 6)]
KYUKA_COLS = ["休日休暇名1", "休日休暇名1：種別", "AM有休", "PM有休"]

MAX_ROWS_PER_IMPORT = 4900  # jinjer上限5000行に対する安全マージン
POLL_INTERVAL_SEC = 20
POLL_MAX_SEC = 900

KUBUN_KYUKA = "休暇日スキップ"
KUBUN_NG = "検証NG"
KUBUN_IMPORT_FAIL = "インポート失敗"
KUBUN_GUARD = "送信前チェックNG"


# ---------------------------------------------------------------------------
# 純粋関数（テスト対象）
# ---------------------------------------------------------------------------
def t2m(v) -> int | None:
    """'8:45' / '08:45:00' / '33:30' → 分。空・不正は None。24時超もそのまま扱う"""
    s = str(v or "").strip()
    m = re.match(r"^(\d{1,3}):(\d{2})(?::\d{2})?$", s)
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def norm_date_iso(v) -> str | None:
    """'2026/6/1' / '2026-06-01' / '6/1/2026' → 'YYYY-MM-DD'"""
    s = str(v or "").strip()
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def _truthy_flag(v: str) -> bool:
    return (v or "").strip() not in ("", "0", "FALSE", "false", "False")


_DATE_JINJER_RE = re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")
_SECONDS_CELL_RE = re.compile(r"^\d{1,3}:\d{2}:\d{2}$")


def validate_upload_csv(
    dict_rows: list[dict[str, str]],
    target_month: str = "",
    month_explicit: bool = False,
    today: datetime | None = None,
) -> list[str]:
    """送信前の門番チェック。エラーメッセージのリストを返す（空なら合格）。

    2026-07-08 の誤インポート事故の再発防止:
    Excel上書き保存で米国式(6/7/2026)に化けた日付を jinjer が日/月/年と
    誤解釈し、各月6日へ誤書込した（7/6法休化事故）。1件でも該当したら送信しない。

    検査項目:
      1. *年月日 が jinjer形式 (YYYY/M/D) か（ISO・米国式は化けの徴候）
      2. 秒付き時刻セルが無いか（Excel保存の徴候。jinjerは秒付きを弾く）
      3. 複数月の混在が無いか / target_month 指定時はその月だけか
      4. 未来月が無いか（無条件NG）・当月と前月より古い月は明示指定時のみ許可
    """
    errors: list[str] = []
    bad_dates: list[str] = []
    sec_cells: list[str] = []
    months: set[str] = set()
    for i, d in enumerate(dict_rows, 2):
        ds = (d.get(COL_DATE) or "").strip()
        if _DATE_JINJER_RE.match(ds):
            y, m, _day = ds.split("/")
            months.add(f"{int(y):04d}-{int(m):02d}")
        elif len(bad_dates) < 3:
            bad_dates.append(f"{i}行目「{ds}」")
        for col, v in d.items():
            if v and _SECONDS_CELL_RE.match(v):
                if len(sec_cells) < 3:
                    sec_cells.append(f"{i}行目 {col}「{v}」")
                break
    if bad_dates:
        errors.append(
            "年月日が jinjer形式(YYYY/M/D) でない行があります: " + " / ".join(bad_dates)
            + " …ISO形式(2026-06-01)や米国式(6/7/2026)はExcelで開いて保存した徴候です。"
              "jinjerは米国式を日/月/年と誤解釈して別の月に書き込むため送信を中止しました。"
              "quick_export でCSVを再生成してください")
    if sec_cells:
        errors.append(
            "秒付き時刻セルがあります: " + " / ".join(sec_cells)
            + " …Excelで開いて上書き保存した徴候です。quick_export でCSVを再生成してください")
    if len(months) > 1:
        errors.append(f"複数の月の日付が混在しています: {sorted(months)} → 月ごとに分けて投入してください")
    if target_month and months and months != {target_month}:
        errors.append(f"対象月 {target_month} 以外の日付が含まれています: {sorted(months)}")
    now = today or datetime.now()
    cur = f"{now.year:04d}-{now.month:02d}"
    prev_y, prev_m = (now.year, now.month - 1) if now.month > 1 else (now.year - 1, 12)
    prev = f"{prev_y:04d}-{prev_m:02d}"
    for mo in sorted(months):
        if mo > cur:
            errors.append(f"未来の月({mo})の日付が含まれています。勤怠の書き戻しで未来月は"
                          "あり得ないため送信を中止しました（日付化けの徴候）")
        elif mo < prev and not month_explicit:
            errors.append(f"当月・前月より古い月({mo})の日付が含まれています。遡及修正で"
                          "意図的な場合は、対象月を明示指定して再実行してください")
    return errors


def is_kyuka_row(row: dict[str, str]) -> tuple[bool, str]:
    """休暇登録がある行か判定し、(該当するか, 備考文) を返す。

    休暇がある日は jinjer がスケジュール書き込みをサイレントに無視する
    （守屋さん6/23の半休で実測）ため、行ごと自動投入から除外する。
    """
    name1 = (row.get("休日休暇名1") or "").strip()
    shubetsu = (row.get("休日休暇名1：種別") or "").strip()
    am = _truthy_flag(row.get("AM有休") or "")
    pm = _truthy_flag(row.get("PM有休") or "")
    if not (name1 or shubetsu or am or pm):
        return False, ""
    parts = []
    if name1:
        parts.append(f"休日休暇名1={name1}")
    if shubetsu:
        parts.append(f"種別={shubetsu}")
    if am:
        parts.append("AM有休")
    if pm:
        parts.append("PM有休")
    return True, "休暇登録あり（スケジュールは上書きされないため除外）: " + " / ".join(parts)


def breaks_of_row(row: dict[str, str]) -> list[tuple[int, int]]:
    """行の休憩予定ペアを分単位で返す（両側そろったペアのみ）"""
    result = []
    for bs_col, be_col in BREAK_PAIRS:
        bs, be = t2m(row.get(bs_col)), t2m(row.get(be_col))
        if bs is not None and be is not None:
            result.append((bs, be))
    return sorted(result)


def compare_row(
    row: dict[str, str],
    sched: dict | None,
    att: dict | None,
) -> list[tuple[str, str, str, str]]:
    """CSV行の意図値と現物（API取得値）を比較し、不一致項目のリストを返す。

    Returns:
        [(項目名, 意図値, 現在値, 追記備考), ...]  一致なら空リスト。
        CSVで空欄の項目は「無処理」なので比較しない。
    """
    ngs: list[tuple[str, str, str, str]] = []
    sched = sched or {}
    att = att or {}

    def _cmp(label, want_s, got_s, note=""):
        want, got = t2m(want_s), t2m(got_s)
        if want is None:
            return
        if got != want:
            ngs.append((label, str(want_s), str(got_s or "(なし)"), note))

    _cmp("出勤予定", row.get(COL_SCHED_IN), sched.get("start"))
    _cmp("退勤予定", row.get(COL_SCHED_OUT), sched.get("end"))
    _cmp("出勤打刻", row.get(COL_PUNCH_IN), att.get("in"),
         "実績はまるめ前表示の可能性あり→画面確認")
    _cmp("退勤打刻", row.get(COL_PUNCH_OUT), att.get("out"),
         "実績はまるめ前表示の可能性あり→画面確認")

    want_breaks = breaks_of_row(row)
    if want_breaks:
        got_breaks = sorted(
            (t2m(s), t2m(e)) for s, e in (sched.get("breaks") or []) if t2m(s) is not None
        )
        if got_breaks != want_breaks:
            def _fmt(pairs):
                return " ".join(f"{m1 // 60}:{m1 % 60:02d}-{m2 // 60}:{m2 % 60:02d}"
                                for m1, m2 in pairs) or "(なし)"
            ngs.append(("休憩予定", _fmt(want_breaks), _fmt(got_breaks), ""))
    return ngs


# ---------------------------------------------------------------------------
# 実行結果
# ---------------------------------------------------------------------------
@dataclass
class ImportRunResult:
    ok: bool = False
    dry_run: bool = True
    total_rows: int = 0
    submitted_rows: int = 0
    excluded: list[dict] = field(default_factory=list)   # 手動対応リストの行
    verified_ok: int = 0
    verified_ng: int = 0
    import_statuses: list[str] = field(default_factory=list)
    report_path: str = ""
    log: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 本体
# ---------------------------------------------------------------------------
def run_api_import(
    upload_csv: Path,
    output_dir: Path,
    executor_id: str = "",
    dry_run: bool = True,
    month: str = "",
    log_func: Callable[[str], None] | None = None,
    client: JinjerClient | None = None,
) -> ImportRunResult:
    """アップロード用CSVを API で jinjer へ投入し、反映を検証して Excel レポートを出す。

    Args:
        upload_csv: quick_export が生成した汎用データ形式CSV（CP932）
        output_dir: レポート出力先フォルダ
        executor_id: kintai-imports の実行者社員番号（完了通知メール宛先）
        dry_run: True ならガードとプラン表示のみ（jinjerへ送信しない）
        month: 検証用の対象月 "YYYY-MM"。空なら CSV の年月日から自動判定
        log_func: 進捗ログの出力先（未指定なら print）
        client: テスト時に差し替え可能な JinjerClient
    """
    result = ImportRunResult(dry_run=dry_run)

    def log(msg: str) -> None:
        result.log.append(msg)
        (log_func or print)(msg)

    # ---- 1. CSV読み込み ----
    with open(upload_csv, encoding="cp932", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        raw_rows = list(reader)
    idx = {h: i for i, h in enumerate(header)}
    for col in (COL_EMP, COL_DATE):
        if col not in idx:
            raise ValueError(f"CSVに必須列がありません: {col}")

    def row_dict(r: list[str]) -> dict[str, str]:
        return {h: (r[i].strip() if i < len(r) else "") for h, i in idx.items()}

    rows = [(r, row_dict(r)) for r in raw_rows]
    result.total_rows = len(rows)
    log(f"入力CSV: {upload_csv} （{len(rows)}行）")
    log("※jinjer画面からの手動インポートは併用しないでください"
        "（2026-07-08のファイル取り違え事故の再発防止。取り込みはこのAPI投入に一本化）")

    month_explicit = bool(month)
    months = {norm_date_iso(d[COL_DATE])[:7] for _r, d in rows if norm_date_iso(d[COL_DATE])}
    if not month:
        month = sorted(months)[-1] if months else datetime.now().strftime("%Y-%m")

    # ---- 1.5 送信前の門番チェック（2026-07-08事故の再発防止） ----
    guard_errors = validate_upload_csv(
        [d for _r, d in rows], target_month=month, month_explicit=month_explicit)
    if guard_errors:
        for e in guard_errors:
            log(f"[中止] {e}")
        log("送信前チェックNGのため、jinjerへは送信していません。")
        for e in guard_errors:
            result.excluded.append({
                "従業員番号": "", "氏名": "", "日付": "",
                "区分": KUBUN_GUARD, "備考": e,
            })
        result.ok = False
        result.report_path = _write_report(output_dir, result, [], month, dry_run=dry_run)
        return result

    # ---- 2. ガード（休暇日除外） ----
    submit_rows: list[list[str]] = []
    intended: dict[tuple[str, str], dict[str, str]] = {}
    for r, d in rows:
        emp = d.get(COL_EMP, "")
        date_iso = norm_date_iso(d.get(COL_DATE)) or ""
        kyuka, why = is_kyuka_row(d)
        if kyuka:
            result.excluded.append({
                "従業員番号": emp, "氏名": d.get(COL_NAME, ""),
                "日付": date_iso, "区分": KUBUN_KYUKA, "備考": why,
            })
            continue
        submit_rows.append(r)
        if emp and date_iso:
            intended[(emp, date_iso)] = d
    result.submitted_rows = len(submit_rows)
    log(f"ガード: 休暇日スキップ {len(result.excluded)} 行 / 投入対象 {len(submit_rows)} 行 "
        f"（{len({k[0] for k in intended})} 名）")

    if dry_run:
        result.ok = True
        result.report_path = _write_report(
            output_dir, result, [], month, dry_run=True)
        log(f"dry-run 完了（jinjerへは送信していません）。手動対応リスト暫定版: {result.report_path}")
        return result

    if not submit_rows:
        result.ok = True
        log("投入対象行がありません")
        result.report_path = _write_report(output_dir, result, [], month, dry_run=False)
        return result

    # ---- 3. 投入（5000行超は分割、同時予約1件のため直列） ----
    cli = client or JinjerClient()
    ts = datetime.now().strftime("%H%M%S")
    chunks = [submit_rows[i:i + MAX_ROWS_PER_IMPORT]
              for i in range(0, len(submit_rows), MAX_ROWS_PER_IMPORT)]
    for n, chunk in enumerate(chunks, 1):
        buf = io.StringIO()
        w = csv.writer(buf, lineterminator="\r\n")
        w.writerow(header)
        w.writerows(chunk)
        csv_bytes = buf.getvalue().encode("cp932")
        file_name = f"API投入_{month.replace('-', '')}_{ts}_{n}.csv"
        log(f"投入 {n}/{len(chunks)}: {file_name} （{len(chunk)}行, {len(csv_bytes):,} bytes）"
            f" executor={executor_id or '(未指定=マスタ)'}")
        resp = cli.post_kintai_import(csv_bytes, file_name, executor_id=executor_id or None)
        log(f"  POST応答: executor={resp.get('executor')} type={resp.get('type')}")

        status = poll_import_status(cli, file_name, log)
        result.import_statuses.append(status)
        if status != "1":
            log(f"[ERROR] インポートが成功しませんでした (status={status})")
            for r in chunk:
                d = row_dict(r)
                result.excluded.append({
                    "従業員番号": d.get(COL_EMP, ""), "氏名": d.get(COL_NAME, ""),
                    "日付": norm_date_iso(d.get(COL_DATE)) or "",
                    "区分": KUBUN_IMPORT_FAIL,
                    "備考": f"インポートstatus={status}。通知メール（executor宛）を確認",
                })
            result.report_path = _write_report(output_dir, result, [], month, dry_run=False)
            return result

    # ---- 4. 検証 ----
    log("反映検証中（work-schedules / attendances API）…")
    verify_rows: list[dict] = []
    emps = sorted({k[0] for k in intended})
    for i, emp in enumerate(emps, 1):
        try:
            sched_month = cli.get_work_schedules(emp, month)
            time.sleep(0.15)
            att_month = cli.get_attendance_times(emp, month)
        except JinjerAPIError as e:
            log(f"[WARN] 検証取得失敗 emp={emp}: {e}")
            sched_month, att_month = {}, {}
        for (e_, d_), rowd in intended.items():
            if e_ != emp:
                continue
            ngs = compare_row(rowd, sched_month.get(d_), att_month.get(d_))
            status = "OK" if not ngs else "NG"
            if ngs:
                result.verified_ng += 1
                for label, want, got, note in ngs:
                    biko = f"{label}: 期待{want} → 実際{got}"
                    if note:
                        biko += f"（{note}）"
                    result.excluded.append({
                        "従業員番号": e_, "氏名": rowd.get(COL_NAME, ""),
                        "日付": d_, "区分": KUBUN_NG, "備考": biko,
                    })
            else:
                result.verified_ok += 1
            verify_rows.append({
                "従業員番号": e_, "氏名": rowd.get(COL_NAME, ""), "日付": d_,
                "判定": status,
                "詳細": " / ".join(f"{l}:期待{w}→実際{g}" for l, w, g, _n in ngs),
            })
        time.sleep(0.15)
        if i % 10 == 0 or i == len(emps):
            log(f"  … {i}/{len(emps)} 名検証済み")

    result.ok = result.verified_ng == 0
    log(f"検証結果: 反映OK {result.verified_ok} / NG {result.verified_ng}")

    # ---- 5. レポート ----
    result.report_path = _write_report(output_dir, result, verify_rows, month, dry_run=False)
    log(f"レポート出力: {result.report_path}")
    return result


def poll_import_status(cli: JinjerClient, file_name: str, log) -> str:
    """インポートステータスを確定までポーリングする。'1'=成功 '2'=失敗 ''=タイムアウト

    POSTが200でも、executor に勤怠管理者権限が無い場合などは予約がキューに
    現れないまま破棄されることがある（2026-07-09 executor=9999999 で実測）。
    キューに一度も現れないまま NOT_SEEN_LIMIT 回経過したら早期に打ち切る。
    schedule_import_runner からも共有される（実証済みロジックを一本化）。
    """
    label = {"0": "予約", "1": "成功", "2": "失敗"}
    NOT_SEEN_LIMIT = 6  # 20秒×6=約2分キューに現れなければ異常とみなす
    not_seen = 0
    seen_once = False
    deadline = time.time() + POLL_MAX_SEC
    while time.time() < deadline:
        time.sleep(POLL_INTERVAL_SEC)
        item = cli.find_kintai_import(file_name)
        if not item:
            not_seen += 1
            log("  （キューに未反映…再確認）")
            if not seen_once and not_seen >= NOT_SEEN_LIMIT:
                log("  [ERROR] 投入予約がキューに現れません。executor の社員番号に"
                    "勤怠管理者権限が無い可能性があります（--executor で管理者の"
                    "社員番号を指定して再実行してください）")
                return ""
            continue
        seen_once = True
        status = str(item.get("status") or "")
        log(f"  status={status}({label.get(status, '?')}) updated={item.get('updated_at')}")
        if status in ("1", "2"):
            return status
    return ""


_poll_status = poll_import_status  # 後方互換エイリアス（既存呼び出し・テスト用）


# ---------------------------------------------------------------------------
# Excel レポート
# ---------------------------------------------------------------------------
def _write_report(
    output_dir: Path,
    result: ImportRunResult,
    verify_rows: list[dict],
    month: str,
    dry_run: bool,
) -> str:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = "dryrun_" if dry_run else ""
    path = output_dir / f"API投入結果_{suffix}{month.replace('-', '')}_{datetime.now():%H%M%S}.xlsx"

    wb = Workbook()
    header_fill = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
    ng_fill = PatternFill("solid", start_color="FCE4EC", end_color="FCE4EC")

    def _sheet(ws, columns, rows_, widths, highlight=None):
        for j, h in enumerate(columns, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.fill = header_fill
            c.font = Font(bold=True)
        for i, row in enumerate(rows_, 2):
            for j, h in enumerate(columns, 1):
                ws.cell(row=i, column=j, value=row.get(h, ""))
            if highlight and highlight(row):
                for j in range(1, len(columns) + 1):
                    ws.cell(row=i, column=j).fill = ng_fill
        for j, w_ in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w_
        ws.freeze_panes = "A2"
        if rows_:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows_) + 1}"

    ws1 = wb.active
    ws1.title = "手動対応リスト"
    _sheet(ws1, ["従業員番号", "氏名", "日付", "区分", "備考"],
           sorted(result.excluded, key=lambda r: (r["区分"], r["従業員番号"], r["日付"])),
           [12, 14, 12, 16, 80],
           highlight=lambda r: r.get("区分") in (KUBUN_NG, KUBUN_IMPORT_FAIL, KUBUN_GUARD))

    ws2 = wb.create_sheet("検証結果")
    _sheet(ws2, ["従業員番号", "氏名", "日付", "判定", "詳細"],
           verify_rows, [12, 14, 12, 8, 80],
           highlight=lambda r: r.get("判定") == "NG")

    ws3 = wb.create_sheet("取込ログ")
    ws3.column_dimensions["A"].width = 120
    for i, line in enumerate(result.log, 1):
        ws3.cell(row=i, column=1, value=line)

    wb.save(path)
    return str(path)
