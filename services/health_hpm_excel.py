"""健診結果の整形済みExcel（スキーマv2）を読む

health-check-pdf-to-excel スキルが健診PDFから起こした Excel を読み、
HPM取込用CSVを組み立てるための構造化データにする。

**このモジュールは値を一切加工しない。** 特に血圧は、1回目と2回目を原票の
まま別々に持ち、平均を作らない・1回分を2回目へ複製しない。CSVを作る側が
「合流させようがない」形（項目＋測定回のキー）で渡すのがここの役目。

スキーマv1（測定回も定性値も持たない旧形式）は読めるだけ読んで
`SCHEMA_TOO_OLD` を付けて返す。画面にプレビューは出すが、CSV生成は
呼び出し側が止める。

Excelの実データはセルの型がまちまちなので、次を素通しで受ける:
  - 受診日   : datetime / date / "2026-07-01 00:00:00" / "2026-07-01" / "2026/7/1"
  - 受診No.  : 132（int）/ "132"（str）→ どちらも "000132" にする
  - 性別     : "男性"/"女性"（受診者一覧の実表記）/ "男"/"女"
  - 数値     : 89.3 / "89.3" / 98.0 → "89.3" / "98"（丸めない・指数表記にしない）
"""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl

logger = logging.getLogger(__name__)

# 対応するスキーマ版。これ未満は CSV 生成を認めない。
REQUIRED_SCHEMA_VERSION = "2.0"

# 変換案内シートのキー名
KEY_SCHEMA_VERSION = "健康診断整形スキーマ版"
KEY_GENPYO_CONFIRMED = "原票確認済み"
KEY_BP_OCCURRENCE = "血圧測定回保持"
KEY_QUALITATIVE = "定性値保持"

# データシートの共通レイアウト（1行目タイトル / 2行目説明 / 3行目ヘッダー / 4行目〜データ）
HEADER_ROW = 3
FIRST_DATA_ROW = 4

SHEET_INFO = "変換案内"
SHEET_RECIPIENTS = "受診者一覧"
SHEET_ITEMS = "項目別データ"

# 値種別（項目別データのv2追加列）
VT_NUMERIC = "数値"
VT_QUALITATIVE = "定性"
VT_TEXT = "文字"
VT_NEEDS_CHECK = "要原票確認"

# 血圧の項目名。測定回が必須なのはこの2つだけ。
BP_ITEMS = ("収縮期血圧", "拡張期血圧")

# HPMへ出力する血圧の測定回。3回目以降は保持するが出力しない。
BP_OUTPUT_OCCURRENCES = (1, 2)

# 陰性を表す括弧つきの表記。括弧の中身がマイナス様1文字のときだけ陰性と見なす。
# 括弧なしの単独ハイフンは、未実施・空欄と区別できないので**絶対に陰性にしない**。
_MINUS_CHARS = "-－−‐‑‒–—―ー゛"
_OPEN_PARENS = "(（"
_CLOSE_PARENS = ")）"
_NEGATIVE_RE = re.compile(
    rf"^[{_OPEN_PARENS}]\s*[{_MINUS_CHARS}]\s*[{_CLOSE_PARENS}]$"
)
_BARE_HYPHEN_RE = re.compile(rf"^[{_MINUS_CHARS}]+$")
# 明記された「陰性」も標準値へ寄せる（原票にそう印字されている場合のみ）
_NEGATIVE_WORDS = ("陰性", "(陰性)", "（陰性）")

# 標準の陰性表記。CSVにはこの文字列をそのまま出す。
NEGATIVE = "(-)"


@dataclass(frozen=True)
class Issue:
    """画面にそのまま出す指摘。level が error なら CSV を作らない。"""

    level: str  # "error" | "warning" | "info"
    code: str
    message: str


@dataclass(frozen=True)
class HealthMetric:
    """検査値1件。値は文字列のまま持ち、数値演算は一切しない。"""

    category: str
    item: str
    occurrence: int  # 測定回。血圧以外は1
    value: str
    value_type: str
    unit: str = ""
    source_judgement: str = ""  # 原票判定A〜G。表示専用。CSVビルダーは参照しない
    source_note: str = ""
    original_display: str = ""
    source_page: str = ""
    source_sheet: str = ""


@dataclass
class PersonRecord:
    key: str
    name: str
    age: int | None
    gender: str
    exam_date: date | None
    exam_no: str
    sheet: str
    metrics: list[HealthMetric] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)

    def blood_pressure(self) -> dict[int, dict[str, str]]:
        """{測定回: {"sys": 収縮期, "dia": 拡張期}}。平均は作らない。"""
        out: dict[int, dict[str, str]] = {}
        for m in self.metrics:
            if m.item == "収縮期血圧":
                out.setdefault(m.occurrence, {})["sys"] = m.value
            elif m.item == "拡張期血圧":
                out.setdefault(m.occurrence, {})["dia"] = m.value
        return out

    def qualitative(self) -> list[HealthMetric]:
        return [m for m in self.metrics if m.value_type == VT_QUALITATIVE]

    def numeric(self) -> list[HealthMetric]:
        return [m for m in self.metrics if m.value_type == VT_NUMERIC]


@dataclass
class WorkbookParseResult:
    schema_version: str = ""
    genpyo_confirmed: bool = False
    bp_occurrence_kept: bool = False
    qualitative_kept: bool = False
    persons: list[PersonRecord] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)
    source_filename: str = ""

    def all_issues(self) -> list[Issue]:
        out = list(self.issues)
        for p in self.persons:
            out.extend(p.issues)
        return out

    def errors(self) -> list[Issue]:
        return [i for i in self.all_issues() if i.level == "error"]

    def warnings(self) -> list[Issue]:
        return [i for i in self.all_issues() if i.level == "warning"]


# ---------------------------------------------------------------------------
# セル値の読み取り
# ---------------------------------------------------------------------------

def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def format_cell_number(value) -> str:
    """数値セルを文字列にする。丸めない・指数表記にしない・末尾の .0 を消す。

    openpyxl は 98 を 98.0 で返すことがあり、そのまま str() すると "98.0" に
    なってHPMの想定と変わる。Decimal 経由で正規化する。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return _text(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        try:
            d = Decimal(str(value)).normalize()
        except InvalidOperation:
            return _text(value)
        # normalize() は 100 を 1E+2 にするので、指数表記になったら戻す
        return format(d, "f")
    text = _text(value)
    if not text:
        return ""
    try:
        d = Decimal(text).normalize()
    except InvalidOperation:
        return text  # 数値でないならそのまま（"1.10" のような原票表記も壊さない）
    if text.lstrip("-+").startswith("0") and "." not in text:
        return text  # "0301619" のような前ゼロは数値化せず原票どおり
    return format(d, "f")


def _to_bool(value) -> bool:
    text = _text(value).upper()
    return text in ("TRUE", "1", "YES", "はい", "○", "有")


def parse_exam_date(value) -> date | None:
    """受診日。datetime/date/文字列（"2026-07-01 00:00:00" 等）を受ける。"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    text = text.split("T")[0].split(" ")[0]
    text = unicodedata.normalize("NFKC", text)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def format_exam_no(value) -> tuple[str, Issue | None]:
    """受診番号を6桁ゼロ埋めの文字列にする。

    Excel側で数値になって 132 で来ることがある。CSVでは "000132" が正。
    （共有CSVがExcel再保存でゼロ落ちした事故があるため、ここで必ず埋める）
    """
    if value is None or value == "":
        return "", Issue("error", "EXAM_NO_MISSING", "受診No.が空欄です")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = _text(value)
    if not text.isdigit():
        return text, Issue(
            "error", "EXAM_NO_INVALID", f"受診No.が数字ではありません: {text!r}"
        )
    if len(text) > 6:
        return text, Issue(
            "error", "EXAM_NO_TOO_LONG", f"受診No.が6桁を超えています: {text!r}"
        )
    return text.zfill(6), None


def normalize_qualitative(raw) -> tuple[str, Issue | None]:
    """定性値を標準表記に寄せる。

    - 括弧つきのマイナス様1文字、明記された「陰性」 → "(-)"
    - (+) (±) (2+) 1+ などは印字どおり保持（全角だけ半角に寄せる）
    - **括弧なしの単独ハイフンは変換しない**（陰性か未実施か原票で確認が要る）
    - 空欄は空欄のまま（勝手に "(-)" にしない）
    """
    text = _text(raw)
    if not text:
        return "", None

    compact = unicodedata.normalize("NFKC", text).replace(" ", "").replace("　", "")

    if _NEGATIVE_RE.match(text.replace(" ", "").replace("　", "")):
        return NEGATIVE, None
    if compact in _NEGATIVE_WORDS or compact == "陰性":
        return NEGATIVE, None
    if _BARE_HYPHEN_RE.match(compact):
        return text, Issue(
            "error",
            "QUAL_BARE_HYPHEN",
            f"括弧のないハイフン {text!r} は陰性か未実施か判断できません。"
            "原票を確認してスキーマ2.0で整形し直してください",
        )
    # NFKC で (２＋) → (2+) のように半角へ寄る。医学的な読み替えはしない。
    return compact, None


# ---------------------------------------------------------------------------
# シート読み取り
# ---------------------------------------------------------------------------

def _read_info_sheet(ws) -> dict[str, str]:
    """変換案内シートを {キー: 値} にする（A列=キー / B列=値）。"""
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        key = _text(row[0] if row else None)
        if not key:
            continue
        value = _text(row[1] if len(row) > 1 else None)
        # タイトル行は同じ文字がA/Bに並ぶので落とす
        if key == value:
            continue
        out.setdefault(key, value)
    return out


def _header_map(ws) -> dict[str, int]:
    """3行目のヘッダーを {列名: 0始まりindex} にする。列順に依存しないため。"""
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for idx, cell in enumerate(row):
            name = _text(cell)
            if name and name not in out:
                out[name] = idx
        break
    return out


def _iter_data_rows(ws, key_indexes: list[int]):
    """4行目以降を回す。主要列がすべて空の行で打ち切る（read_only の max_row 対策）。"""
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if not row:
            break
        if all(not _text(row[i]) for i in key_indexes if i < len(row)):
            break
        yield row


def _get(row, header: dict[str, int], name: str):
    idx = header.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ---------------------------------------------------------------------------
# 本体
# ---------------------------------------------------------------------------

def parse_health_workbook(path: str, original_filename: str = "") -> WorkbookParseResult:
    """整形済みExcelを読み、構造化データと指摘を返す。

    内容の問題は例外にせず Issue に集める（画面で全部見せて人が直せるように）。
    ファイル自体が開けないときだけ ValueError。
    """
    result = WorkbookParseResult(source_filename=original_filename or path)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - openpyxl は多種の例外を投げる
        raise ValueError(f"Excelとして開けません: {e}") from e

    try:
        _parse_into(wb, result)
    finally:
        wb.close()

    return result


def _parse_into(wb, result: WorkbookParseResult) -> None:
    names = set(wb.sheetnames)

    # --- 変換案内（スキーマ版と原票確認フラグ） ---
    if SHEET_INFO not in names:
        result.issues.append(
            Issue("error", "SCHEMA_TOO_OLD",
                  f"「{SHEET_INFO}」シートがありません。"
                  "原票PDFからスキーマ2.0で再整形してください")
        )
        info = {}
    else:
        info = _read_info_sheet(wb[SHEET_INFO])

    result.schema_version = info.get(KEY_SCHEMA_VERSION, "")
    result.genpyo_confirmed = _to_bool(info.get(KEY_GENPYO_CONFIRMED))
    result.bp_occurrence_kept = _to_bool(info.get(KEY_BP_OCCURRENCE))
    result.qualitative_kept = _to_bool(info.get(KEY_QUALITATIVE))

    if not _schema_is_supported(result.schema_version):
        shown = result.schema_version or "（記載なし）"
        result.issues.append(
            Issue("error", "SCHEMA_TOO_OLD",
                  f"整形スキーマ版が {shown} です。"
                  f"CSVを作るには {REQUIRED_SCHEMA_VERSION} が要ります。"
                  "原票PDFからスキーマ2.0で再整形してください")
        )
    else:
        if not result.genpyo_confirmed:
            result.issues.append(
                Issue("error", "GENPYO_FLAG_MISSING",
                      f"変換案内の「{KEY_GENPYO_CONFIRMED}」が TRUE ではありません。"
                      "原票画像との照合が済んだExcelだけCSVにできます")
            )
        if not result.bp_occurrence_kept:
            result.issues.append(
                Issue("error", "BP_OCCURRENCE_FLAG_MISSING",
                      f"変換案内の「{KEY_BP_OCCURRENCE}」が TRUE ではありません。"
                      "血圧の1回目・2回目が保持されているExcelが必要です")
            )
        if not result.qualitative_kept:
            result.issues.append(
                Issue("warning", "QUALITATIVE_FLAG_MISSING",
                      f"変換案内の「{KEY_QUALITATIVE}」が TRUE ではありません。"
                      "定性検査が取り込まれていない可能性があります")
            )

    # --- 受診者一覧 ---
    if SHEET_RECIPIENTS not in names:
        result.issues.append(
            Issue("error", "RECIPIENTS_SHEET_MISSING",
                  f"「{SHEET_RECIPIENTS}」シートがありません")
        )
        return
    persons = _read_recipients(wb[SHEET_RECIPIENTS], result)
    result.persons = persons
    if not persons:
        result.issues.append(
            Issue("error", "NO_RECIPIENTS", "受診者一覧に1名も載っていません")
        )
        return

    # --- 項目別データ ---
    if SHEET_ITEMS not in names:
        result.issues.append(
            Issue("error", "ITEMS_SHEET_MISSING",
                  f"「{SHEET_ITEMS}」シートがありません")
        )
        return
    _read_items(wb[SHEET_ITEMS], persons, result)

    for person in persons:
        _check_blood_pressure(person)


def _schema_is_supported(version: str) -> bool:
    if not version:
        return False
    try:
        parts = [int(p) for p in str(version).strip().split(".")[:2]]
    except ValueError:
        return False
    while len(parts) < 2:
        parts.append(0)
    required = [int(p) for p in REQUIRED_SCHEMA_VERSION.split(".")]
    return parts >= required


def _read_recipients(ws, result: WorkbookParseResult) -> list[PersonRecord]:
    header = _header_map(ws)
    required = ("氏名", "受診日", "受診No.", "個人票シート")
    missing = [c for c in required if c not in header]
    if missing:
        result.issues.append(
            Issue("error", "RECIPIENTS_COLUMNS_MISSING",
                  f"受診者一覧に必要な列がありません: {', '.join(missing)}")
        )
        return []

    key_indexes = [header["氏名"], header["個人票シート"]]
    persons: list[PersonRecord] = []
    seen_sheets: set[str] = set()

    for i, row in enumerate(_iter_data_rows(ws, key_indexes), start=1):
        name = _text(_get(row, header, "氏名"))
        if not name:
            continue
        sheet = _text(_get(row, header, "個人票シート"))
        key = f"p{i:02d}"

        issues: list[Issue] = []
        exam_date = parse_exam_date(_get(row, header, "受診日"))
        if exam_date is None:
            issues.append(
                Issue("error", "EXAM_DATE_INVALID",
                      f"{name}: 受診日を読み取れません "
                      f"({_text(_get(row, header, '受診日'))!r})")
            )
        exam_no, exam_no_issue = format_exam_no(_get(row, header, "受診No."))
        if exam_no_issue:
            issues.append(
                Issue(exam_no_issue.level, exam_no_issue.code,
                      f"{name}: {exam_no_issue.message}")
            )

        age_text = _text(_get(row, header, "年齢"))
        try:
            age = int(float(age_text)) if age_text else None
        except ValueError:
            age = None

        if sheet and sheet in seen_sheets:
            issues.append(
                Issue("error", "DUPLICATE_SHEET",
                      f"{name}: 個人票シート {sheet} が複数の受診者に割り当てられています")
            )
        seen_sheets.add(sheet)

        persons.append(PersonRecord(
            key=key,
            name=name,
            age=age,
            gender=_text(_get(row, header, "性別")),
            exam_date=exam_date,
            exam_no=exam_no,
            sheet=sheet,
            issues=issues,
        ))

    return persons


def _read_items(ws, persons: list[PersonRecord], result: WorkbookParseResult) -> None:
    header = _header_map(ws)
    required = ("氏名", "分類", "項目", "値")
    missing = [c for c in required if c not in header]
    if missing:
        result.issues.append(
            Issue("error", "ITEMS_COLUMNS_MISSING",
                  f"項目別データに必要な列がありません: {', '.join(missing)}")
        )
        return

    # v2 で足した3列。無ければ旧スキーマなので、この時点で明示的に指摘する。
    v2_columns = ("測定回", "値種別", "原票表記")
    missing_v2 = [c for c in v2_columns if c not in header]
    if missing_v2:
        result.issues.append(
            Issue("error", "SCHEMA_TOO_OLD",
                  f"項目別データに {', '.join(missing_v2)} の列がありません。"
                  "血圧の測定回と定性値を保持したスキーマ2.0で再整形してください")
        )

    by_sheet = {p.sheet: p for p in persons if p.sheet}
    by_name: dict[str, list[PersonRecord]] = {}
    for p in persons:
        by_name.setdefault(p.name, []).append(p)

    # (person.key, 分類, 項目, 測定回) -> HealthMetric
    seen: dict[tuple, HealthMetric] = {}
    unmatched_rows = 0

    for row in _iter_data_rows(ws, [header["氏名"], header["項目"]]):
        name = _text(_get(row, header, "氏名"))
        sheet = _text(_get(row, header, "個人票シート")) if "個人票シート" in header else ""

        person = by_sheet.get(sheet)
        if person is None:
            candidates = by_name.get(name) or []
            person = candidates[0] if len(candidates) == 1 else None
        if person is None:
            unmatched_rows += 1
            continue

        category = _text(_get(row, header, "分類"))
        item = _text(_get(row, header, "項目"))
        raw_value = _get(row, header, "値")
        value_type = _text(_get(row, header, "値種別")) or _guess_value_type(raw_value)
        original = _text(_get(row, header, "原票表記"))

        if not item:
            if _text(raw_value):
                person.issues.append(
                    Issue("error", "QUAL_ATTRIBUTION_UNKNOWN",
                          f"{person.name}: 値 {_text(raw_value)!r} がどの検査項目のものか"
                          "分かりません（項目名が空欄です）")
                )
            continue

        occurrence, occ_issue = _read_occurrence(row, header, item, person)
        if occ_issue:
            person.issues.append(occ_issue)
            continue

        if value_type == VT_NEEDS_CHECK:
            person.issues.append(
                Issue("error", "NEEDS_SOURCE_CHECK",
                      f"{person.name}: {category}/{item} が「{VT_NEEDS_CHECK}」です。"
                      "原票画像で確認してから整形し直してください")
            )
            continue

        if value_type == VT_QUALITATIVE:
            value, qual_issue = normalize_qualitative(raw_value)
            if qual_issue:
                person.issues.append(
                    Issue(qual_issue.level, qual_issue.code,
                          f"{person.name}: {category}/{item} — {qual_issue.message}")
                )
                continue
        elif value_type == VT_NUMERIC:
            value = format_cell_number(raw_value)
        else:
            value = _text(raw_value)

        if not value:
            continue  # 空欄は空欄のまま。勝手に埋めない

        metric = HealthMetric(
            category=category,
            item=item,
            occurrence=occurrence,
            value=value,
            value_type=value_type,
            unit=_text(_get(row, header, "単位")),
            source_judgement=_text(_get(row, header, "原票判定")),
            source_note=_text(_get(row, header, "原票注記")),
            original_display=original or _text(raw_value),
            source_page=_text(_get(row, header, "PDFページ")),
            source_sheet=sheet or person.sheet,
        )

        dup_key = (person.key, category, item, occurrence)
        previous = seen.get(dup_key)
        if previous is not None:
            if previous.value != metric.value:
                person.issues.append(
                    Issue("error", "DUP_VALUE_CONFLICT",
                          f"{person.name}: {category}/{item}"
                          f"{'（' + str(occurrence) + '回目）' if item in BP_ITEMS else ''}"
                          f" に異なる値があります: {previous.value!r} と {metric.value!r}")
                )
            continue
        seen[dup_key] = metric
        person.metrics.append(metric)

    if unmatched_rows:
        result.issues.append(
            Issue("warning", "ITEM_ROWS_UNMATCHED",
                  f"項目別データの {unmatched_rows} 行が受診者と結び付きませんでした")
        )


def _guess_value_type(raw) -> str:
    """値種別列が無い（旧スキーマ）ときの当て推量。CSV生成はどのみち止まる。"""
    text = _text(raw)
    if not text:
        return VT_TEXT
    try:
        Decimal(text)
        return VT_NUMERIC
    except InvalidOperation:
        return VT_TEXT


def _read_occurrence(row, header, item: str, person: PersonRecord) -> tuple[int, Issue | None]:
    """測定回を読む。血圧は必須、それ以外は既定1。"""
    raw = _get(row, header, "測定回") if "測定回" in header else None
    text = _text(raw)

    if not text:
        if item in BP_ITEMS:
            return 0, Issue(
                "error", "BP_OCCURRENCE_MISSING",
                f"{person.name}: {item} の測定回が空欄です。"
                "1回目か2回目かを原票で確認してください（推測はしません）",
            )
        return 1, None

    try:
        occurrence = int(float(text))
    except ValueError:
        return 0, Issue(
            "error", "BP_OCCURRENCE_MISSING",
            f"{person.name}: {item} の測定回 {text!r} を読み取れません",
        )
    if occurrence < 1:
        return 0, Issue(
            "error", "BP_OCCURRENCE_MISSING",
            f"{person.name}: {item} の測定回が {occurrence} です",
        )
    return occurrence, None


def _check_blood_pressure(person: PersonRecord) -> None:
    """血圧の欠落・3回目を指摘する。**値には一切触らない。**"""
    bp = person.blood_pressure()
    if not bp:
        return

    extra = sorted(o for o in bp if o not in BP_OUTPUT_OCCURRENCES)
    if extra:
        person.issues.append(
            Issue("warning", "BP_THIRD_IGNORED",
                  f"{person.name}: 血圧の{', '.join(str(o) + '回目' for o in extra)}が"
                  "ありますが、HPMには1回目・2回目だけを出力します")
        )

    if 1 in bp and 2 not in bp:
        person.issues.append(
            Issue("warning", "BP_SINGLE",
                  f"{person.name}: 血圧が1回分だけです。"
                  "2回目は空欄のまま出力します（1回目は複製しません）")
        )

    for occurrence in BP_OUTPUT_OCCURRENCES:
        pair = bp.get(occurrence)
        if not pair:
            continue
        if "sys" not in pair or "dia" not in pair:
            missing = "拡張期血圧" if "sys" in pair else "収縮期血圧"
            person.issues.append(
                Issue("warning", "BP_PAIR_INCOMPLETE",
                      f"{person.name}: 血圧{occurrence}回目の{missing}がありません")
            )
