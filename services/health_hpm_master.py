"""健康診断HPM変換マスタ（Excel）を読む

健診機関コード・健診コース・302列ヘッダー・項目マッピングを Excel の外部
マスタで持つ。HPM側で列が増えたり機関が増えたりしても、マスタを直すだけで
exe の再ビルドなしに反映できるようにするため（経理モードのマッピング表と同じ方式）。

**マスタが壊れている限り1バイトも出力しない。** 列がズレたまま出力すると
別の検査値が別の欄に入ったCSVができてしまい、しかも人には気付きにくい。
そのため読み込み時に総当たりで検証し、少しでも違えば MasterError を投げる。

特に血圧は、マスタが書き換えられても 50〜53 から動かないことをコード側の
定数と二重に突き合わせる（平均を作らない設計の最後の砦）。
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field

import openpyxl

logger = logging.getLogger(__name__)

SHEET_INSTITUTIONS = "健診機関"
SHEET_ALIASES = "機関別名"
SHEET_COURSES = "健診種別"
SHEET_SETTINGS = "設定"
SHEET_ITEM_MAP = "項目マッピング"
SHEET_HEADER = "HPMヘッダー"

REQUIRED_SHEETS = (SHEET_INSTITUTIONS, SHEET_ALIASES, SHEET_COURSES,
                   SHEET_SETTINGS, SHEET_ITEM_MAP, SHEET_HEADER)

HEADER_ROW = 3
FIRST_DATA_ROW = 4

TOTAL_COLS = 302

# --- コード側の固定値（マスタと二重に突き合わせる） ---

# 血圧は測定回ごとに固定の列。1回目と2回目が同じ列に合流する経路を作らない。
BP_EXPECTED_COLS = {
    ("収縮期血圧", 1): 50,
    ("拡張期血圧", 1): 51,
    ("収縮期血圧", 2): 52,
    ("拡張期血圧", 2): 53,
}
BP_EXPECTED_NAMES = {
    50: "血圧（一回目）最高",
    51: "血圧（一回目）最低",
    52: "血圧（二回目）最高",
    53: "血圧（二回目）最低",
}

# 原票判定A〜Gを書いてはいけない列。HPMのマスタコードとは体系が違うため。
JUDGEMENT_COLS = range(183, 198)

# 氏名・生年月日・受診日など、行を組み立てる側が埋める列。
# 項目マッピングがここを指していたら、検査値が識別情報を上書きしてしまう。
IDENTITY_COLS = range(0, 24)

SETTING_VENUE_CODE = "会場コード"


class MasterError(ValueError):
    """マスタが読めない・壊れている。CSV生成は必ず止める。"""


@dataclass(frozen=True)
class Institution:
    name: str
    location_code: str  # "0301619" の前ゼロや "13X5035440" の英字を守るため必ず str
    public_code_ref: str = ""
    hpm_confirmed: bool = False
    note: str = ""


@dataclass(frozen=True)
class Course:
    institution: str
    display_name: str
    hpm_value: str  # 同友会はコース名テキスト、他機関は "10" 等の数値コード


@dataclass(frozen=True)
class ItemMapRule:
    category: str
    item: str
    occurrence: int
    hpm_col: int  # 0始まり
    value_type: str
    method: str = ""  # HBs/HCV等の検査方式。空なら方式を問わない
    note: str = ""

    @property
    def key(self) -> tuple:
        return (self.category, self.item, self.occurrence, self.method)


@dataclass
class HpmMaster:
    path: str
    header: list[str] = field(default_factory=list)
    institutions: dict[str, Institution] = field(default_factory=dict)
    aliases: dict[str, str] = field(default_factory=dict)
    courses: dict[str, list[Course]] = field(default_factory=dict)
    settings: dict[str, str] = field(default_factory=dict)
    item_map: list[ItemMapRule] = field(default_factory=list)

    @property
    def venue_code(self) -> str:
        return self.settings.get(SETTING_VENUE_CODE, "")

    def rules_for(self, category: str, item: str, occurrence: int) -> list[ItemMapRule]:
        """(分類, 項目, 測定回) に一致する規則。方式違いが複数返ることがある。"""
        return [r for r in self.item_map
                if r.category == category and r.item == item
                and r.occurrence == occurrence]


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_bool(value) -> bool:
    return _text(value).upper() in ("TRUE", "1", "YES", "はい", "○", "有")


def _header_map(ws, sheet_name: str) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for idx, cell in enumerate(row):
            name = _text(cell)
            if name and name not in out:
                out[name] = idx
        break
    if not out:
        raise MasterError(f"{sheet_name}シートの{HEADER_ROW}行目にヘッダーがありません")
    return out


def _rows(ws, header: dict[str, int], key_columns: list[str]):
    key_indexes = [header[c] for c in key_columns if c in header]
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if not row:
            break
        if all(not _text(row[i]) for i in key_indexes if i < len(row)):
            break
        yield row


def _get(row, header: dict[str, int], name: str) -> str:
    idx = header.get(name)
    if idx is None or idx >= len(row):
        return ""
    return _text(row[idx])


def _require_columns(header: dict[str, int], columns, sheet_name: str) -> None:
    missing = [c for c in columns if c not in header]
    if missing:
        raise MasterError(
            f"{sheet_name}シートに必要な列がありません: {', '.join(missing)}"
        )


def load_master(path: str) -> HpmMaster:
    """変換マスタを読む。少しでも想定と違えば MasterError。"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError as e:
        raise MasterError(f"変換マスタが見つかりません: {path}") from e
    except Exception as e:  # noqa: BLE001
        raise MasterError(f"変換マスタを開けません: {path} ({e})") from e

    try:
        missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            raise MasterError(
                f"変換マスタに必要なシートがありません: {', '.join(missing)}"
            )
        master = HpmMaster(path=path)
        master.header = _load_header(wb[SHEET_HEADER])
        master.institutions = _load_institutions(wb[SHEET_INSTITUTIONS])
        master.aliases = _load_aliases(wb[SHEET_ALIASES], master.institutions)
        master.courses = _load_courses(wb[SHEET_COURSES], master.institutions)
        master.settings = _load_settings(wb[SHEET_SETTINGS])
        master.item_map = _load_item_map(wb[SHEET_ITEM_MAP], master.header)
    finally:
        wb.close()

    _validate(master)
    return master


def _load_header(ws) -> list[str]:
    header = _header_map(ws, SHEET_HEADER)
    _require_columns(header, ("列番号", "列名"), SHEET_HEADER)

    by_index: dict[int, str] = {}
    for row in _rows(ws, header, ["列番号"]):
        raw = _get(row, header, "列番号")
        try:
            idx = int(float(raw))
        except ValueError as e:
            raise MasterError(
                f"{SHEET_HEADER}シートの列番号が数字ではありません: {raw!r}"
            ) from e
        if idx in by_index:
            raise MasterError(f"{SHEET_HEADER}シートで列番号 {idx} が重複しています")
        by_index[idx] = _get(row, header, "列名")

    if len(by_index) != TOTAL_COLS:
        raise MasterError(
            f"{SHEET_HEADER}シートは{TOTAL_COLS}行必要ですが {len(by_index)} 行です"
        )
    if set(by_index) != set(range(TOTAL_COLS)):
        raise MasterError(
            f"{SHEET_HEADER}シートの列番号が 0〜{TOTAL_COLS - 1} の連番になっていません"
        )
    return [by_index[i] for i in range(TOTAL_COLS)]


def _load_institutions(ws) -> dict[str, Institution]:
    header = _header_map(ws, SHEET_INSTITUTIONS)
    _require_columns(header, ("機関名", "HPM場所コード", "HPM確認済み"), SHEET_INSTITUTIONS)

    out: dict[str, Institution] = {}
    for row in _rows(ws, header, ["機関名"]):
        name = _get(row, header, "機関名")
        if not name:
            continue
        if name in out:
            raise MasterError(f"{SHEET_INSTITUTIONS}シートで機関名 {name} が重複しています")
        out[name] = Institution(
            name=name,
            location_code=_get(row, header, "HPM場所コード"),
            public_code_ref=_get(row, header, "公開医療機関コード参考"),
            hpm_confirmed=_to_bool(_get(row, header, "HPM確認済み")),
            note=_get(row, header, "備考"),
        )
    if not out:
        raise MasterError(f"{SHEET_INSTITUTIONS}シートに1件も登録がありません")
    return out


def _load_aliases(ws, institutions: dict[str, Institution]) -> dict[str, str]:
    header = _header_map(ws, SHEET_ALIASES)
    _require_columns(header, ("別名", "正式機関名"), SHEET_ALIASES)

    out: dict[str, str] = {}
    for row in _rows(ws, header, ["別名"]):
        alias = _get(row, header, "別名")
        official = _get(row, header, "正式機関名")
        if not alias:
            continue
        if official not in institutions:
            raise MasterError(
                f"{SHEET_ALIASES}シートの別名 {alias} が指す "
                f"{official!r} は{SHEET_INSTITUTIONS}シートにありません"
            )
        out[alias] = official
    return out


def _load_courses(ws, institutions: dict[str, Institution]) -> dict[str, list[Course]]:
    header = _header_map(ws, SHEET_COURSES)
    _require_columns(header, ("機関名", "種別表示名", "HPM出力値"), SHEET_COURSES)

    out: dict[str, list[Course]] = {}
    for row in _rows(ws, header, ["機関名", "種別表示名"]):
        institution = _get(row, header, "機関名")
        if institution not in institutions:
            raise MasterError(
                f"{SHEET_COURSES}シートの機関名 {institution!r} は"
                f"{SHEET_INSTITUTIONS}シートにありません"
            )
        course = Course(
            institution=institution,
            display_name=_get(row, header, "種別表示名"),
            hpm_value=_get(row, header, "HPM出力値"),
        )
        if not course.hpm_value:
            raise MasterError(
                f"{SHEET_COURSES}シートの {institution}/{course.display_name} に"
                "HPM出力値がありません"
            )
        out.setdefault(institution, []).append(course)
    if not out:
        raise MasterError(f"{SHEET_COURSES}シートに1件も登録がありません")
    return out


def _load_settings(ws) -> dict[str, str]:
    header = _header_map(ws, SHEET_SETTINGS)
    _require_columns(header, ("キー", "値"), SHEET_SETTINGS)
    out: dict[str, str] = {}
    for row in _rows(ws, header, ["キー"]):
        key = _get(row, header, "キー")
        if key:
            out[key] = _get(row, header, "値")
    return out


def _load_item_map(ws, header_names: list[str]) -> list[ItemMapRule]:
    header = _header_map(ws, SHEET_ITEM_MAP)
    _require_columns(header, ("分類", "項目名", "測定回", "HPM列番号", "値種別"),
                     SHEET_ITEM_MAP)

    rules: list[ItemMapRule] = []
    for row in _rows(ws, header, ["項目名"]):
        item = _get(row, header, "項目名")
        if not item:
            continue
        raw_col = _get(row, header, "HPM列番号")
        try:
            hpm_col = int(float(raw_col))
        except ValueError as e:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートの {item} のHPM列番号が数字ではありません: {raw_col!r}"
            ) from e
        raw_occ = _get(row, header, "測定回") or "1"
        try:
            occurrence = int(float(raw_occ))
        except ValueError as e:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートの {item} の測定回が数字ではありません: {raw_occ!r}"
            ) from e

        if not 0 <= hpm_col < TOTAL_COLS:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートの {item} のHPM列番号 {hpm_col} が範囲外です"
                f"（0〜{TOTAL_COLS - 1}）"
            )
        if hpm_col in JUDGEMENT_COLS:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートの {item} が判定列 {hpm_col}"
                f"（{header_names[hpm_col]}）を指しています。"
                "原票判定A〜GはHPMへ転記しない決まりです"
            )
        if hpm_col in IDENTITY_COLS:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートの {item} が識別列 {hpm_col}"
                f"（{header_names[hpm_col]}）を指しています。"
                "氏名・生年月日・受診日などの列に検査値は書けません"
            )

        rules.append(ItemMapRule(
            category=_get(row, header, "分類"),
            item=item,
            occurrence=occurrence,
            hpm_col=hpm_col,
            value_type=_get(row, header, "値種別"),
            method=_get(row, header, "検査方式"),
            note=_get(row, header, "備考"),
        ))

    if not rules:
        raise MasterError(f"{SHEET_ITEM_MAP}シートに1件も登録がありません")
    return rules


def _validate(master: HpmMaster) -> None:
    """血圧列・重複・設定の総当たり検証。"""
    # --- 血圧列がコード側の固定値と一致するか（最重要） ---
    bp_rules = {(r.item, r.occurrence): r for r in master.item_map
                if (r.item, r.occurrence) in BP_EXPECTED_COLS}
    for key, expected_col in BP_EXPECTED_COLS.items():
        rule = bp_rules.get(key)
        if rule is None:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートに {key[0]}（{key[1]}回目）の行がありません。"
                f"HPM列 {expected_col} へ出す設定が要ります"
            )
        if rule.hpm_col != expected_col:
            raise MasterError(
                f"{key[0]}（{key[1]}回目）のHPM列が {rule.hpm_col} になっています。"
                f"{expected_col} でなければなりません"
                "（血圧の1回目・2回目を取り違えると平均のような値になります）"
            )

    for col, expected_name in BP_EXPECTED_NAMES.items():
        actual = master.header[col]
        if actual != expected_name:
            raise MasterError(
                f"{SHEET_HEADER}シートの列 {col} が {actual!r} です。"
                f"{expected_name!r} でなければなりません"
            )

    # --- 同じ (分類, 項目, 測定回, 方式) が複数の列を指していないか ---
    seen_keys: dict[tuple, ItemMapRule] = {}
    for rule in master.item_map:
        previous = seen_keys.get(rule.key)
        if previous is not None:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートで {rule.category}/{rule.item}"
                f"（{rule.occurrence}回目・方式{rule.method or '指定なし'}）が"
                f"列 {previous.hpm_col} と {rule.hpm_col} の両方を指しています"
            )
        seen_keys[rule.key] = rule

    # --- 同じ列を複数の項目が指していないか ---
    seen_cols: dict[int, ItemMapRule] = {}
    for rule in master.item_map:
        previous = seen_cols.get(rule.hpm_col)
        if previous is not None:
            raise MasterError(
                f"{SHEET_ITEM_MAP}シートで列 {rule.hpm_col}"
                f"（{master.header[rule.hpm_col]}）を "
                f"{previous.item} と {rule.item} の両方が指しています"
            )
        seen_cols[rule.hpm_col] = rule

    # --- 設定 ---
    if not master.venue_code:
        raise MasterError(f"{SHEET_SETTINGS}シートに「{SETTING_VENUE_CODE}」がありません")


def resolve_institution(master: HpmMaster, name: str) -> Institution | None:
    """機関名または別名から機関を引く。"""
    name = (name or "").strip()
    if not name:
        return None
    if name in master.institutions:
        return master.institutions[name]
    official = master.aliases.get(name)
    if official:
        return master.institutions.get(official)
    return None


def courses_of(master: HpmMaster, institution_name: str) -> list[Course]:
    institution = resolve_institution(master, institution_name)
    if institution is None:
        return []
    return list(master.courses.get(institution.name, []))


def find_course(master: HpmMaster, institution_name: str, hpm_value: str) -> Course | None:
    """HPM出力値（画面から返ってくる値）でコースを引く。"""
    for course in courses_of(master, institution_name):
        if course.hpm_value == hpm_value:
            return course
    return None
