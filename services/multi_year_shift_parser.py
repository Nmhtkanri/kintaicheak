"""多年度横並びレイアウトのシフト表 xlsx と、シフトルール xlsx の構造化パーサ

以下の2種類の xlsx を Claude API を使わず確定的に解析する。

1. シフトルール (legend) xlsx
   行ごとに | 記号 | 開始時間 | 終了時間 | 休憩開始 | 休憩終了 |
   末尾に「休(m/d) | 振休」「休 | 休暇」といった休扱い行が続く。

2. シフト表 xlsx (多年度横並び)
   - 1〜2 行目: 休マーカー / 日付行（Excel シリアル値が並ぶ）
   - 3 行目: 曜日（土,日,月,...）
   - 4 行目以降: 氏名 + シフトパターン列 + 各日のシフト記号
   - 同じ「4/1 〜 3/31」または「1/1 〜 12/31」のパターンが複数年度分
     横方向に積まれており、年ラベルは無い。日付セルだけは全年度とも
     2026 シリアルが入っている（コピーで上書きされた状態）が、
     **3 行目の曜日は元の年度のまま**なので、target_year の曜日と一致
     するセクションが「本物の」当該年度データである。

target_year/target_month に対応するセクションを曜日マッチで自動選定し、
{name, shifts:[{date, code}]} 形式の従業員データを返す。"""

from __future__ import annotations

import calendar
import logging
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Iterable

import openpyxl

# 他形式のシフト表パーサ。以前は sniff/parse の関数内で import していたが、
# 1 本の ImportError が同じ try ブロックにいる無関係な形式判定
# （多年度シフト・月次シフト・凡例）まで巻き添えでスキップさせ、
# except Exception が warning ログに落とすだけで静かに壊れる作りだった。
# モジュール先頭に上げて、モジュールが欠けているなら起動時に大きく落ちるようにする。
# （どちらも stdlib と openpyxl しか import せず循環参照は無い。
#   kdx 側の pdfplumber は関数内 import なので起動コストも増えない）
from services.higashi_shift_parser import (
    is_higashi_shift_xlsx,
    parse_higashi_shift_xlsx,
)
from services.kdx_shift_parser import is_kdx_shift_pdf, parse_kdx_shift_pdf
from services.ual_shift_parser import is_ual_shift_xlsx, parse_ual_shift_xlsx

logger = logging.getLogger(__name__)


# 曜日 → 漢字 1 文字（Mon=0..Sun=6）
_WD_KANJI = ["月", "火", "水", "木", "金", "土", "日"]

# Excel epoch (1900 leap-year bug 込み)
_EXCEL_EPOCH = date(1899, 12, 30)

_FULLWIDTH_DIGIT_TRANS = str.maketrans("０１２３４５６７８９", "0123456789")


# =============================================================================
# 共通ユーティリティ
# =============================================================================

def _to_date(v) -> date | None:
    """セル値を date に変換できれば返す（数値シリアル / datetime / date のみ対応）"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            n = int(v)
        except Exception:
            return None
        if n < 1 or n > 60_000:
            return None
        try:
            return _EXCEL_EPOCH + timedelta(days=n)
        except Exception:
            return None
    return None


def _to_hhmm(v) -> str | None:
    """セル値を "HH:MM" 文字列に変換"""
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # "07:00:00" / "7:00" 等を切り詰める
        parts = s.split(":")
        if len(parts) >= 2:
            try:
                h = int(parts[0])
                m = int(parts[1])
                return f"{h:02d}:{m:02d}"
            except ValueError:
                return None
        return None
    return None


def _clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_digits(s: str) -> str:
    return str(s).translate(_FULLWIDTH_DIGIT_TRANS)


def _time_to_minutes(hhmm: str) -> int | None:
    try:
        h, m = map(int, hhmm.split(":"))
    except (AttributeError, ValueError):
        return None
    return h * 60 + m


def _format_time_match(hour: str, minute: str) -> str:
    return f"{int(_normalize_digits(hour)):02d}:{int(_normalize_digits(minute)):02d}"


def _default_break_minutes(start: str, end: str) -> int:
    """勤務時間から休憩を推定する。月次表は休憩列を持たないため、6時間超を60分扱いにする。"""
    s = _time_to_minutes(start)
    e = _time_to_minutes(end)
    if s is None or e is None:
        return 0
    if e <= s:
        e += 24 * 60
    return 60 if (e - s) > 6 * 60 else 0


# =============================================================================
# 1. シフトルール xlsx パーサ
# =============================================================================

# 「休扱い」とみなすラベル（凡例に開始/終了時刻が無い行用）
_OFF_LABEL_PATTERNS = ("振休", "休暇", "公休", "休日", "代休", "明け")


def parse_legend_xlsx(filepath: str) -> list[dict]:
    """シフトルール xlsx を解析して凡例リストを返す

    Returns:
        [{code, label, start_time, end_time, break_minutes, is_off}, ...]
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    legend: list[dict] = []
    seen_codes: set[str] = set()

    for row in rows:
        if not row:
            continue
        code = _clean_str(row[0])
        if not code:
            continue
        # ヘッダ行は除外
        if code in ("記号", "コード", "シフト記号"):
            continue

        col1 = _clean_str(row[1]) if len(row) > 1 else ""

        # 開始/終了時刻が文字列ラベルの場合（休系）
        start_time = _to_hhmm(row[1]) if len(row) > 1 else None
        end_time = _to_hhmm(row[2]) if len(row) > 2 else None

        if start_time is None and end_time is None:
            # 「休」「休(m/d)」などの休扱い行
            if any(p in col1 for p in _OFF_LABEL_PATTERNS) or code.startswith("休"):
                if code in seen_codes:
                    continue
                seen_codes.add(code)
                legend.append({
                    "code": code,
                    "label": col1 or code,
                    "start_time": "",
                    "end_time": "",
                    "break_minutes": 0,
                    "is_off": True,
                })
            continue

        if not start_time or not end_time:
            continue

        # 休憩は「開始」「終了」が直接入っている前提で分換算
        break_minutes = 0
        b_start = _to_hhmm(row[3]) if len(row) > 3 else None
        b_end = _to_hhmm(row[4]) if len(row) > 4 else None
        if b_start and b_end:
            try:
                sh, sm = map(int, b_start.split(":"))
                eh, em = map(int, b_end.split(":"))
                diff = (eh * 60 + em) - (sh * 60 + sm)
                if diff > 0:
                    break_minutes = diff
            except ValueError:
                pass

        # 重複コードは最初の登場分のみ採用
        if code in seen_codes:
            logger.warning("シフトルール内で重複コード '%s' を検出（後続行をスキップ）", code)
            continue
        seen_codes.add(code)

        legend.append({
            "code": code,
            "label": code,  # シフトルールには label 列が無いのでコード自身を使う
            "start_time": start_time,
            "end_time": end_time,
            "break_minutes": break_minutes,
            "is_off": False,
        })

    return legend


def is_legend_xlsx(filepath: str) -> bool:
    """シフトルール xlsx かどうかをファイル内容から判定する

    判定条件:
      - 単一シート
      - 行数が比較的少ない（<= 30）
      - 1 列目に A/B/C... のような 1-2 文字コードが並ぶ
      - 2-3 列目に時刻型セルがある
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return False

    if len(rows) > 30:
        return False

    code_count = 0
    time_count = 0
    for row in rows:
        if not row or row[0] is None:
            continue
        code = _clean_str(row[0])
        if 1 <= len(code) <= 3 and any(c.isalpha() or c == "休" for c in code):
            code_count += 1
        if len(row) > 1 and isinstance(row[1], time):
            time_count += 1
        elif len(row) > 1 and isinstance(row[1], datetime):
            time_count += 1
    return code_count >= 3 and time_count >= 3


# =============================================================================
# 2. 多年度横並び シフト表 xlsx パーサ
# =============================================================================

@dataclass
class _Section:
    """日付行を連続セクション（次の日に増分する範囲）に分けた1ブロック"""
    index: int           # 1-based
    start_col: int       # 1-based
    end_col: int         # 1-based, inclusive
    dates: dict[int, date]  # col_idx (1-based) -> date


def _find_date_row(rows: list[tuple]) -> int:
    """日付行の 0-indexed 行番号を返す。ヒューリスティック:
    日付シリアルに変換できるセルが最も多い行を選ぶ。
    """
    best_idx = -1
    best_count = 0
    for i, row in enumerate(rows[:10]):  # 上位 10 行のみ
        if not row:
            continue
        cnt = sum(1 for v in row if _to_date(v) is not None)
        if cnt > best_count:
            best_count = cnt
            best_idx = i
    if best_count < 30:  # 最低 30 セルくらいは欲しい
        raise ValueError("日付行が見つかりません（30セル以上の日付シリアル列が無い）")
    return best_idx


def _split_date_sections(date_row: tuple) -> list[_Section]:
    """日付行を連続日付のセクションに分割する。日付が前のセルの翌日でないところで切る。"""
    sections: list[_Section] = []
    sec_start_col: int | None = None
    sec_dates: dict[int, date] = {}
    last_date: date | None = None

    for i, v in enumerate(date_row):
        col_1based = i + 1
        d = _to_date(v)
        if d is None:
            # 文字列「2/29」など → 連続性を切らずにスキップ
            continue
        if last_date is None or (d - last_date).days != 1:
            # 新セクション開始
            if sec_dates:
                sections.append(_Section(
                    index=len(sections) + 1,
                    start_col=sec_start_col,
                    end_col=last_col,
                    dates=sec_dates,
                ))
            sec_start_col = col_1based
            sec_dates = {}
        sec_dates[col_1based] = d
        last_date = d
        last_col = col_1based

    if sec_dates:
        sections.append(_Section(
            index=len(sections) + 1,
            start_col=sec_start_col,
            end_col=last_col,
            dates=sec_dates,
        ))

    return sections


def _score_section_weekdays(
    section: _Section,
    weekday_row: tuple,
    target_year: int,
    target_month: int,
) -> tuple[int, int]:
    """セクション内 target_year/month の曜日が weekday_row の表記と一致する数を返す

    Returns:
        (matched, total)
    """
    matched = 0
    total = 0
    for col_1based, d in section.dates.items():
        if d.year != target_year or d.month != target_month:
            continue
        total += 1
        wd_actual = _WD_KANJI[d.weekday()]
        cell = weekday_row[col_1based - 1] if col_1based - 1 < len(weekday_row) else None
        wd_excel = _clean_str(cell)
        if wd_excel == wd_actual:
            matched += 1
    return matched, total


def _pick_target_section(
    sections: list[_Section],
    weekday_row: tuple,
    target_year: int,
    target_month: int,
) -> _Section | None:
    """target_year/month の曜日が最も一致するセクションを返す

    タイブレイク: rightmost (start_col が大きい) を優先（直近年度の想定）
    """
    candidates: list[tuple[float, int, _Section]] = []
    for sec in sections:
        # target year/month の日が 1 件でも含まれていなければスキップ
        contains = any(
            d.year == target_year and d.month == target_month
            for d in sec.dates.values()
        )
        if not contains:
            continue
        matched, total = _score_section_weekdays(sec, weekday_row, target_year, target_month)
        ratio = (matched / total) if total else 0.0
        candidates.append((ratio, sec.start_col, sec))

    if not candidates:
        return None

    # ratio 降順 → start_col 降順 (rightmost) でソート
    candidates.sort(key=lambda x: (-x[0], -x[1]))
    return candidates[0][2]


def _find_employee_rows(rows: list[tuple], data_start_row: int) -> list[int]:
    """従業員行の 0-indexed リストを返す

    判定: data_start_row 以降で col 0 が空でなく、コード値（D など）が
    複数個出現している行。
    """
    result: list[int] = []
    for i in range(data_start_row, len(rows)):
        row = rows[i]
        if not row:
            continue
        name = _clean_str(row[0])
        if not name:
            continue
        # コード値が 5 個以上ある行を従業員行とみなす（パターン列だけ埋まっている行を除外）
        code_count = 0
        for v in row[2:]:  # col 0 = name, col 1 = "シフトパターン" 想定
            s = _clean_str(v)
            if s and s != "シフトパターン":
                code_count += 1
                if code_count >= 5:
                    break
        if code_count >= 5:
            result.append(i)
    return result


def parse_multi_year_shift_xlsx(
    filepath: str,
    target_year: int,
    target_month: int,
) -> dict:
    """多年度横並びレイアウトのシフト表 xlsx を解析

    Args:
        filepath: シフト表 xlsx のパス
        target_year, target_month: 抽出対象（必須）

    Returns:
        {
          "year": int,
          "month": int,
          "filename": str,
          "employees": [{"name": str, "shifts": [{"date": "YYYY-MM-DD", "code": str}]}],
          "off_markers": list[str],
          "section_info": {...},  # デバッグ用: 採用したセクション
        }

    Raises:
        ValueError: レイアウトが該当しない / 該当セクションが見つからない場合
    """
    if not target_year or not target_month:
        raise ValueError("target_year / target_month は必須です")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("空のシートです")

    # 日付行と曜日行を特定
    date_row_idx = _find_date_row(rows)
    weekday_row_idx = date_row_idx + 1
    if weekday_row_idx >= len(rows):
        raise ValueError("曜日行が見つかりません")

    date_row = rows[date_row_idx]
    weekday_row = rows[weekday_row_idx]

    # 日付セクション分割
    sections = _split_date_sections(date_row)
    if not sections:
        raise ValueError("日付セクションが分割できません")

    target_section = _pick_target_section(sections, weekday_row, target_year, target_month)
    if target_section is None:
        raise ValueError(
            f"{target_year}年{target_month}月 を含むセクションが見つかりません "
            f"(検出セクション数: {len(sections)})"
        )

    # 採用セクションが本当に当該年度データかチェック
    matched, total = _score_section_weekdays(
        target_section, weekday_row, target_year, target_month
    )
    if total > 0 and matched / total < 0.5:
        logger.warning(
            "採用セクション section%d の曜日マッチ率が低いです (%d/%d)",
            target_section.index, matched, total,
        )

    # 従業員行を抽出
    employee_row_indices = _find_employee_rows(rows, weekday_row_idx + 1)
    if not employee_row_indices:
        raise ValueError("従業員行が見つかりません")

    days_in_month = calendar.monthrange(target_year, target_month)[1]
    target_dates = {
        col: d
        for col, d in target_section.dates.items()
        if d.year == target_year and d.month == target_month
    }

    employees: list[dict] = []
    for ridx in employee_row_indices:
        row = rows[ridx]
        name = _clean_str(row[0])
        if not name:
            continue
        # 「シフトパターン」のラベル列は無視
        shifts = []
        for col_1based, d in sorted(target_dates.items(), key=lambda x: x[1]):
            cell = row[col_1based - 1] if col_1based - 1 < len(row) else None
            code = _clean_str(cell)
            shifts.append({"date": d.isoformat(), "code": code})
        employees.append({"name": name, "shifts": shifts})

    return {
        "year": target_year,
        "month": target_month,
        "filename": os.path.basename(filepath),
        "employees": employees,
        "off_markers": ["休"],  # 「休」を含むセルはデフォルトで休扱い
        "section_info": {
            "section_index": target_section.index,
            "start_col": target_section.start_col,
            "end_col": target_section.end_col,
            "weekday_match_ratio": (matched / total) if total else 0.0,
            "weekday_matched": matched,
            "weekday_total": total,
            "total_sections": len(sections),
        },
    }


def is_multi_year_shift_xlsx(filepath: str) -> bool:
    """多年度横並びレイアウトに該当するか判定する

    判定条件:
      - 1 シート
      - 上位 5 行内に「日付シリアルが 100 セル以上ある行」が 1 つある
      - その下に曜日（土,日,月,...）が並ぶ行がある
      - 列方向の幅が 200 以上ある（短い1ヶ月シートではない）
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    except Exception:
        return False

    if not rows:
        return False
    max_width = max((len(r) for r in rows if r), default=0)
    if max_width < 200:
        return False

    has_long_date_row = False
    for row in rows:
        if not row:
            continue
        cnt = sum(1 for v in row if _to_date(v) is not None)
        if cnt >= 100:
            has_long_date_row = True
            break
    return has_long_date_row


# =============================================================================
# 2b. 月次横並び シフト表 xlsx パーサ
# =============================================================================

def _find_monthly_date_row(rows: list[tuple]) -> int:
    """月次表の「氏名 ... 日付列」行を探す。"""
    best_idx = -1
    best_count = 0
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        has_name_header = any(_clean_str(v) == "氏名" for v in row[:5])
        date_count = sum(1 for v in row if _to_date(v) is not None)
        if has_name_header and date_count > best_count:
            best_idx = i
            best_count = date_count
    if best_idx < 0 or best_count < 15:
        raise ValueError("月次シフト表の日付行が見つかりません")
    return best_idx


def _monthly_date_columns(date_row: tuple) -> dict[int, date]:
    cols: dict[int, date] = {}
    for col_1based, value in enumerate(date_row, start=1):
        d = _to_date(value)
        if d is not None:
            cols[col_1based] = d
    if not cols:
        raise ValueError("月次シフト表の日付列が見つかりません")
    return cols


def _pick_month_from_date_columns(
    date_cols: dict[int, date],
    target_year: int | None,
    target_month: int | None,
) -> tuple[int, int]:
    """対象月を決める。指定年月が表に無い場合は、表内で最も多い年月を採用する。"""
    if target_year and target_month:
        if any(d.year == target_year and d.month == target_month for d in date_cols.values()):
            return target_year, target_month
        logger.warning(
            "指定年月 %s年%s月 が月次シフト表の日付列に無いため、表内の年月を採用します",
            target_year,
            target_month,
        )

    counts: dict[tuple[int, int], int] = {}
    for d in date_cols.values():
        counts[(d.year, d.month)] = counts.get((d.year, d.month), 0) + 1
    if not counts:
        raise ValueError("月次シフト表の対象年月を判定できません")
    return max(counts.items(), key=lambda kv: kv[1])[0]


def _looks_like_monthly_employee_row(row: tuple, date_cols: dict[int, date]) -> bool:
    name = _clean_str(row[0] if row else "")
    if not name or "勤" in name or "シフト" in name or "ｼﾌﾄ" in name or name in ("氏名", "日付"):
        return False
    serial_like_count = 0
    code_count = 0
    for col_1based in date_cols:
        value = row[col_1based - 1] if col_1based - 1 < len(row) else None
        if isinstance(value, (int, float)) and value > 1000:
            serial_like_count += 1
            if serial_like_count >= 5:
                return False
        if _clean_str(value):
            code_count += 1
            if code_count >= 5:
                return True
    return False


_SHIFT_TIME_RE = re.compile(
    r"([0-9０-９]{1,2})[:：]([0-9０-９]{2})\s*[～~-]\s*(?:翌)?\s*([0-9０-９]{1,2})[:：]([0-9０-９]{2})"
)


def _parse_monthly_shift_legend_row(text: str) -> dict | None:
    normalized = _normalize_digits(_clean_str(text))
    if not normalized:
        return None

    code_match = re.match(r"^([0-9]+)\s*勤", normalized)
    if not code_match:
        return None
    code = code_match.group(1)

    time_match = _SHIFT_TIME_RE.search(text)
    if not time_match:
        return None
    start = _format_time_match(time_match.group(1), time_match.group(2))
    end = _format_time_match(time_match.group(3), time_match.group(4))

    label = normalized
    time_start = time_match.start()
    if time_start > 0:
        label = _normalize_digits(text[:time_start]).strip(" 　")
    return {
        "code": code,
        "label": label or code,
        "start_time": start,
        "end_time": end,
        "break_minutes": _default_break_minutes(start, end),
        "is_off": False,
    }


def _parse_monthly_embedded_legend(
    rows: list[tuple],
    *,
    summary_start_col: int | None = None,
) -> tuple[list[dict], list[str]]:
    """月次表内の勤務コード説明と休暇系コードを凡例化する。"""
    legend: list[dict] = []
    seen: set[str] = set()

    for row in rows:
        if not row:
            continue
        parsed = _parse_monthly_shift_legend_row(_clean_str(row[0]))
        if parsed and parsed["code"] not in seen:
            seen.add(parsed["code"])
            legend.append(parsed)

    off_entries = {
        "休": "休",
        "明": "明け休",
    }
    start_idx = max((summary_start_col or 1) - 1, 0)
    for row in rows:
        values = [_clean_str(v) for v in row[start_idx:]]
        for i in range(0, max(len(values) - 1, 0)):
            code = values[i]
            label = values[i + 1]
            if not code:
                continue
            if code in seen or code in ("氏名", "勤務時間", "出"):
                continue
            if label and any(p in label for p in ("休", "有給", "在宅")):
                off_entries[code] = label

    for code, label in off_entries.items():
        if code in seen:
            continue
        seen.add(code)
        legend.append({
            "code": code,
            "label": label,
            "start_time": "",
            "end_time": "",
            "break_minutes": 0,
            "is_off": True,
        })

    return legend, list(off_entries.keys())


def parse_monthly_shift_xlsx(
    filepath: str,
    target_year: int | None = None,
    target_month: int | None = None,
) -> dict:
    """1か月分の横並びシフト表 xlsx を解析する。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("空のシートです")

    date_row_idx = _find_monthly_date_row(rows)
    date_cols_all = _monthly_date_columns(rows[date_row_idx])
    year, month = _pick_month_from_date_columns(date_cols_all, target_year, target_month)
    target_cols = {
        col: d
        for col, d in date_cols_all.items()
        if d.year == year and d.month == month
    }
    if not target_cols:
        raise ValueError(f"{year}年{month}月の日付列が見つかりません")

    employee_rows = [
        row for row in rows[date_row_idx + 1:]
        if _looks_like_monthly_employee_row(row, target_cols)
    ]
    if not employee_rows:
        raise ValueError("月次シフト表の従業員行が見つかりません")

    employees: list[dict] = []
    for row in employee_rows:
        name = _clean_str(row[0])
        shifts = []
        for col_1based, d in sorted(target_cols.items(), key=lambda x: x[1]):
            cell = row[col_1based - 1] if col_1based - 1 < len(row) else None
            code = _normalize_digits(_clean_str(cell))
            shifts.append({"date": d.isoformat(), "code": code})
        employees.append({"name": name, "shifts": shifts})

    legend, off_markers = _parse_monthly_embedded_legend(
        rows,
        summary_start_col=max(date_cols_all) + 1,
    )
    return {
        "year": year,
        "month": month,
        "filename": os.path.basename(filepath),
        "legend": legend,
        "employees": employees,
        "off_markers": off_markers,
        "section_info": {
            "section_index": 1,
            "start_col": min(target_cols),
            "end_col": max(target_cols),
            "weekday_match_ratio": 1.0,
            "weekday_matched": len(target_cols),
            "weekday_total": len(target_cols),
            "total_sections": 1,
        },
    }


def is_monthly_shift_xlsx(filepath: str) -> bool:
    """1か月分の横並びシフト表に該当するか判定する。"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    except Exception:
        return False

    try:
        date_row_idx = _find_monthly_date_row(rows)
        date_cols = _monthly_date_columns(rows[date_row_idx])
    except ValueError:
        return False
    return len(date_cols) >= 15


# =============================================================================
# 3. 高レベルエントリ: 複数ファイルから「凡例 + 従業員シフト」を組み立てる
# =============================================================================

def parse_structured_files(
    filepaths: Iterable[str],
    target_year: "int | None",
    target_month: "int | None",
) -> tuple[list[dict], list[str], list[str]] | None:
    """複数のアップロードパスを sniff し、構造化解析できるファイルだけを処理する

    対象: 多年度シフト表 .xlsx / 月次シフト表 .xlsx / KDX勤務シフト表 .pdf。
    1 つも見つからなければ何もせず None を返す。見つかった場合は構造化パースし、
    consumed パス（後続の Claude フォールバックから除外すべきパス）と、
    ユーザーに見せるべき warning（構造化解析に失敗して AI 読み取りへ
    フォールバックするファイル等）を合わせて返す。

    Args:
        filepaths: アップロードされた timesheet 系ファイルのパス
        target_year, target_month: 抽出対象。None（画面で未入力）の場合、
            年月をファイル自身から特定できる KDX PDF だけを処理し、
            対象月の指定が必要な xlsx 系はスキップして warning を返す

    Returns:
        (sheets, consumed_paths, warnings) — sheets は code_sheets と同形式
        構造化解析の対象ファイルが無い場合は None
    """
    paths = list(filepaths)
    if not paths:
        return None
    warnings: list[str] = []

    legend_paths: list[str] = []
    shift_paths: list[str] = []
    monthly_shift_paths: list[str] = []
    higashi_shift_paths: list[str] = []
    kdx_pdf_paths: list[str] = []
    ual_shift_paths: list[str] = []
    for p in paths:
        low = p.lower()
        if low.endswith(".pdf"):
            try:
                if is_kdx_shift_pdf(p):
                    kdx_pdf_paths.append(p)
            except Exception as e:
                logger.warning("pdf sniff 失敗 %s: %s", p, e)
            continue
        if not low.endswith((".xlsx", ".xls")):
            continue
        try:
            if low.endswith(".xlsx"):
                if is_higashi_shift_xlsx(p):
                    higashi_shift_paths.append(p)
                    continue
                if is_ual_shift_xlsx(p):
                    ual_shift_paths.append(p)
                    continue
            if is_multi_year_shift_xlsx(p):
                shift_paths.append(p)
            elif is_monthly_shift_xlsx(p):
                monthly_shift_paths.append(p)
            elif is_legend_xlsx(p):
                legend_paths.append(p)
        except Exception as e:
            logger.warning("xlsx sniff 失敗 %s: %s", p, e)

    if (not shift_paths and not monthly_shift_paths
            and not higashi_shift_paths and not kdx_pdf_paths
            and not ual_shift_paths):
        return None

    # 凡例をマージ（複数渡された場合は重複コードを除外しつつ全件統合）
    legend: list[dict] = []
    seen: set[str] = set()
    for lp in legend_paths:
        try:
            for entry in parse_legend_xlsx(lp):
                if entry["code"] in seen:
                    continue
                seen.add(entry["code"])
                legend.append(entry)
        except Exception as e:
            logger.warning("シフトルール %s の解析に失敗: %s", lp, e)

    # シフト表ごとにシートを作る
    sheets: list[dict] = []
    consumed: list[str] = []

    for kp in kdx_pdf_paths:
        try:
            result = parse_kdx_shift_pdf(kp, target_year, target_month)
        except Exception as e:
            logger.warning("KDXシフト表 %s の構造化解析に失敗: %s", kp, e)
            warnings.append(
                f"KDXシフト表の構造化解析に失敗したため AI 読み取りへフォールバックします: "
                f"{os.path.basename(kp)} — {e}")
            continue
        sheets.append({
            "mode": "code",
            "filename": result["filename"],
            "legend": result["legend"],
            "employees": result["employees"],
            "off_markers": result["off_markers"],
            "year": result["year"],
            "month": result["month"],
            "source": result.get("source", ""),
            "section_info": result.get("section_info"),
        })
        consumed.append(kp)

    for up in ual_shift_paths:
        try:
            result = parse_ual_shift_xlsx(up, target_year, target_month)
        except Exception as e:
            logger.warning("UAL勤務管理表 %s の構造化解析に失敗: %s", up, e)
            warnings.append(
                f"UAL勤務管理表の構造化解析に失敗しました: {os.path.basename(up)} — {e}")
            # AI 読み取りへは流さない。このブックは全シート（過去月）を渡すことになり
            # 応答が返らず時間切れになるため（2026-07-31 実例）。
            consumed.append(up)
            continue
        sheets.append({
            "mode": "code",
            "filename": result["filename"],
            "legend": result["legend"],
            "employees": result["employees"],
            "off_markers": result["off_markers"],
            "year": result["year"],
            "month": result["month"],
            "source": result.get("source", ""),
            "section_info": result.get("section_info"),
        })
        if result.get("unknown_codes"):
            warnings.append(
                f"{result['filename']}: 凡例に無い記号がありました "
                f"({' / '.join(result['unknown_codes'])})。凡例確認画面で内容を指定してください。")
        consumed.append(up)

    for hp in higashi_shift_paths:
        try:
            result = parse_higashi_shift_xlsx(hp, target_year, target_month)
        except Exception as e:
            logger.warning("東さん形式シフト表 %s の構造化解析に失敗: %s", hp, e)
            warnings.append(
                f"東さん形式シフト表の構造化解析に失敗したため "
                f"AI 読み取りへフォールバックします: {os.path.basename(hp)} — {e}"
            )
            continue
        sheets.append({
            "mode": "code",
            "filename": result["filename"],
            "legend": result["legend"],
            "employees": result["employees"],
            "off_markers": result["off_markers"],
            "year": result["year"],
            "month": result["month"],
            "section_info": result.get("section_info"),
        })
        consumed.append(hp)

    # xlsx 系の構造化パースは対象年月の指定が必須（多年度表から対象月を切り出すため）
    if (monthly_shift_paths or shift_paths) and not (target_year and target_month):
        skipped = [os.path.basename(p) for p in monthly_shift_paths + shift_paths]
        warnings.append(
            "対象年月が未入力のため、シフト表 xlsx の構造化解析をスキップして "
            f"AI 読み取りへフォールバックします: {', '.join(skipped)}"
            "（画面の「対象年月」を入れると確定的に解析できます）")
        monthly_shift_paths = []
        shift_paths = []

    for mp in monthly_shift_paths:
        try:
            result = parse_monthly_shift_xlsx(mp, target_year, target_month)
        except Exception as e:
            logger.warning("月次シフト表 %s の構造化解析に失敗: %s", mp, e)
            continue
        sheets.append({
            "mode": "code",
            "filename": result["filename"],
            "legend": result["legend"],
            "employees": result["employees"],
            "off_markers": result["off_markers"],
            "year": result["year"],
            "month": result["month"],
            "section_info": result.get("section_info"),
        })
        consumed.append(mp)

    for sp in shift_paths:
        try:
            result = parse_multi_year_shift_xlsx(sp, target_year, target_month)
        except Exception as e:
            logger.warning("シフト表 %s の構造化解析に失敗: %s", sp, e)
            continue
        sheets.append({
            "mode": "code",
            "filename": result["filename"],
            "legend": legend,
            "employees": result["employees"],
            "off_markers": result["off_markers"],
            "year": result["year"],
            "month": result["month"],
            "section_info": result.get("section_info"),
        })
        consumed.append(sp)

    if not sheets and not warnings:
        return None

    # 構造化解析が成功したシフト表があれば、シフトルール .xlsx も消費扱いにする
    if sheets:
        consumed.extend(legend_paths)
    return sheets, consumed, warnings
