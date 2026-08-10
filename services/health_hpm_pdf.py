"""健診PDFの原票画像を読んで、整形済みExcelと同じ構造に起こす

チェッカーの健康診断HPMモードに、整形済みExcelを介さずPDFを直接入れるための層。
出来上がりは `health_hpm_excel.parse_health_workbook` と同じ `WorkbookParseResult`
なので、以降の照合・CSV生成はExcel経路とまったく同じコードが動く。

**読み取り方（実測で決めたので変えないこと）**

ページ全体を1枚だけ送ると、A4が長辺1568pxに縮小されて細かい数字が潰れる。
2026-08-10 の検証では、それで **%肺活量(103.4) と 1秒率(82.5) が入れ替わった**。
どちらも実在する値なのでBMI検算のような整合チェックでは気付けない。
そのため「ページ全体 ＋ 上半分の拡大 ＋ 下半分の拡大」の3枚を送る。
この形で同友会6名・228項目が全一致し、同じページを3回読ませてもブレなかった。

**このモジュールが守ること**

- 血圧は測定回ごとに別々。平均を作らない。読み手が測定回を判断できなければ
  値ごと捨ててエラーにする（推測して埋めない）。
- `(-)` は陰性という値。空欄を `(-)` にしない。判断できない表記も値にしない。
- 原票の判定A〜Gは `source_judgement` にだけ入れ、検査値の側へ混ぜない。
- 読めなかったものは AI 自身に申告させ、`READ_NEEDS_CHECK`（情報レベル）として
  画面と監査用Excelに残す。生成は止めないが、人が見る材料にはする。
"""

from __future__ import annotations

import base64
import io
import json
import logging
import re
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

from config import Config
from services.health_hpm_excel import (
    HealthMetric,
    Issue,
    PersonRecord,
    WorkbookParseResult,
    _check_blood_pressure,
    format_exam_no,
    normalize_qualitative,
    parse_exam_date,
)

logger = logging.getLogger(__name__)

# --- 読み取りの設定（検証済みの値。むやみに変えない） ---
PDF_RENDER_DPI = 200
BAND_TOP = (0.10, 0.55)      # 上バンド。身体計測〜血圧〜血液一般あたり
BAND_BOTTOM = (0.45, 0.90)   # 下バンド。肝機能〜大腸〜肺機能あたり。上と重ねて境目の行を落とさない
PAGE_MAX_TOKENS = 8000       # 実測の出力は約2,400トークン
RATE_LIMIT_RETRIES = 3
BMI_TOLERANCE = 0.15         # 身長体重から計算したBMIとの許容差

BP_ITEMS = {"収縮期血圧": "systolic", "拡張期血圧": "diastolic"}
BP_OUTPUT_OCCURRENCES = (1, 2)

# 読み取り結果の項目名を、変換マスタが期待する名前と測定回へ寄せる。
# 検証で実際に出てきた表記だけを入れてある（推測で増やさない）。
ITEM_ALIASES = {
    "便潜血①": ("便潜血", 1),
    "便潜血②": ("便潜血", 2),
    "便潜血1": ("便潜血", 1),
    "便潜血2": ("便潜血", 2),
    "ウロビリノーゲン": ("尿ウロビリノーゲン", 1),
    "尿ウロビリノーゲン": ("尿ウロビリノーゲン", 1),
}

# 同じ値が別欄にも印字される項目。正規の項目が別にあるので落とす。
DROP_ITEMS = {"尿糖（糖代謝）", "尿糖(糖代謝)"}

# 値の頭に付く注記記号。値からは外して原票注記へ回す。
_NOTE_MARKS = "＊*"


SYSTEM_PROMPT = """あなたは健康診断結果の原票（スキャン画像）から検査値を読み取る担当です。
読めないものを推測してはいけません。読めない値は null にして、needs_check に理由を書いてください。

守ること:
- 血圧は「何回目の測定か」を必ず判断する。原票の印字位置やラベル（1回目/2回目/再測定など）で決める。
  平均値を計算してはいけない。原票に平均が印字されていても blood_pressure には入れない。
  1回しか測っていないなら blood_pressure は1件だけにする（2件目を複製しない）。
  どちらが1回目か判断できないときは occurrence を null にして needs_check に書く。
- 定性検査（尿蛋白・尿糖など）の (-) は「陰性」という有効な値。そのまま "(-)" と書く。
  空欄・未実施は null にする。空欄を (-) にしてはいけない。
  括弧のない単独のハイフンは、陰性かどうか判断できないので null + needs_check にする。
- 判定のA〜Gは judgements に入れる。検査値の欄には入れない。
- 数値は原票の桁・小数点をそのまま。単位は付けない。
- 同じ検査値が複数の画像に写っている場合は1件だけ出す（重複させない）。"""

USER_PROMPT = """この健康診断結果の原票画像から、次のJSONだけを出力してください（前置き・説明は不要）。
1枚目はページ全体、2枚目以降は同じページの一部を拡大したものです。数字は拡大側で確認してください。

{
  "identity": {"氏名": "", "年齢": 0, "性別": "", "受診日": "YYYY-MM-DD", "受診No": ""},
  "blood_pressure": [
    {"occurrence": 1, "systolic": 0, "diastolic": 0}
  ],
  "metrics": [
    {"category": "身体計測", "item": "身長", "value": "181.1", "note": null}
  ],
  "qualitative": [
    {"category": "尿検査", "item": "尿蛋白", "value": "(-)", "method": null}
  ],
  "judgements": {"身体計測": "B"},
  "needs_check": ["読めなかった項目とその理由"]
}

値の横に ＊ などの注記記号が印字されている場合は "note" に入れてください（"value" には含めない）。

metrics の項目名は原票の印字どおりで構いませんが、次の標準名があるものはそれに合わせてください:
身長, 体重, BMI, 腹囲, 視力（右）（裸眼）, 視力（左）（裸眼）, 視力（右）（矯正）, 視力（左）（矯正）,
眼圧（右）, 眼圧（左）, 努力性肺活量, %肺活量, 1秒率, 1秒量, 白血球数, 赤血球数, 血色素量,
ヘマトクリット, MCV, MCH, MCHC, 血小板数, 総蛋白, 総ビリルビン, アルブミン, AST, ALT, γ-GT,
LD, ALP, ChE, 総コレステロール, 中性脂肪, HDL-C, LDL-C, 尿素窒素, クレアチニン, eGFR, 尿酸,
空腹時血糖, HbA1c, CRP

血圧は metrics ではなく blood_pressure に入れてください。"""


class PdfReadError(RuntimeError):
    """PDFそのものが扱えない（開けない・0ページ・ページ数が多すぎる）。"""


class PageReadError(RuntimeError):
    """1ページ分の読み取りが確定的に失敗した（他ページは続行できる）。"""


@dataclass
class PdfAnalysisResult:
    parse: WorkbookParseResult
    pages: dict[str, int] = field(default_factory=dict)       # person.key → 1始まりのページ番号
    page_pngs: dict[int, bytes] = field(default_factory=dict)  # ページ番号 → PNG


# ---------------------------------------------------------------------------
# 画像化
# ---------------------------------------------------------------------------

def render_pdf_pages(pdf_path: str, dpi: int = PDF_RENDER_DPI) -> list[bytes]:
    """PDFの全ページをPNGにする。"""
    import pypdfium2 as pdfium

    try:
        pdf = pdfium.PdfDocument(str(pdf_path))
    except Exception as e:  # noqa: BLE001 - pypdfium2 は多様な例外を投げる
        raise PdfReadError(f"PDFを開けません: {e}") from e

    try:
        total = len(pdf)
        if total == 0:
            raise PdfReadError("PDFにページがありません")
        limit = Config.HEALTH_HPM_PDF_MAX_PAGES
        if total > limit:
            raise PdfReadError(
                f"{total}ページあります（1度に読めるのは{limit}ページまで）。"
                "PDFを分けてから読み込んでください"
            )
        scale = dpi / 72.0
        images = []
        for i in range(total):
            pil = pdf[i].render(scale=scale).to_pil().convert("RGB")
            buf = io.BytesIO()
            pil.save(buf, format="PNG")
            images.append(buf.getvalue())
        return images
    finally:
        try:
            pdf.close()
        except Exception:  # noqa: BLE001
            pass


def crop_band(png: bytes, top: float, bottom: float) -> bytes:
    """ページの一部を切り出す。細かい数字を潰さずAIへ渡すため。"""
    from PIL import Image

    im = Image.open(io.BytesIO(png))
    width, height = im.size
    band = im.crop((0, int(height * top), width, int(height * bottom)))
    buf = io.BytesIO()
    band.save(buf, format="PNG")
    return buf.getvalue()


def page_images_for_reading(png: bytes) -> list[bytes]:
    """1ページを読ませるときに送る画像一式（全体＋上下の拡大）。"""
    return [png, crop_band(png, *BAND_TOP), crop_band(png, *BAND_BOTTOM)]


# ---------------------------------------------------------------------------
# Claude 呼び出し（APIに触るのはこの関数だけ。テストではここを差し替える）
# ---------------------------------------------------------------------------

def build_client():
    """timeout を必ず入れる。未指定だと応答が返らないとき画面が固まったままになる。

    SDK 側の自動リトライは 0 にして、下のリトライループと掛け算にならないようにする
    （shift_legend_parser と同じ方針）。
    """
    import anthropic

    return anthropic.Anthropic(
        timeout=Config.ANTHROPIC_TIMEOUT_SECONDS,
        max_retries=0,
    )


def _extract_json(text: str) -> dict:
    fenced = re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", text)
    if fenced:
        text = fenced.group(1)
    match = re.search(r"\{[\s\S]*\}", text)
    if not match:
        raise PageReadError(f"読み取り結果がJSONではありませんでした: {text[:200]}")
    try:
        return json.loads(match.group())
    except json.JSONDecodeError as e:
        raise PageReadError(f"読み取り結果のJSONを解釈できません: {e}") from e


def _call_claude_for_page(client, images: list[bytes]) -> dict:
    """原票画像1ページ分を読み取ってJSONを返す。"""
    import anthropic

    content: list[dict] = []
    for png in images:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": base64.standard_b64encode(png).decode("ascii"),
            },
        })
    content.append({"type": "text", "text": USER_PROMPT})

    last_error: Exception | None = None
    for attempt in range(RATE_LIMIT_RETRIES):
        try:
            res = client.messages.create(
                model=Config.ANTHROPIC_MODEL,
                max_tokens=PAGE_MAX_TOKENS,
                temperature=0,  # 同じ原票からは同じ値が出てほしい
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content}],
            )
        except anthropic.APITimeoutError as e:
            # ここで粘ると画面が何分も固まる。1ページ諦めて次へ行く方がまし。
            raise PageReadError(
                f"読み取りが{Config.ANTHROPIC_TIMEOUT_SECONDS:.0f}秒以内に終わりませんでした"
            ) from e
        except anthropic.RateLimitError as e:
            last_error = e
            wait = 20 * (attempt + 1)
            logger.warning("Claude レート制限。%d秒待って再試行します", wait)
            time.sleep(wait)
            continue
        except anthropic.APIError as e:
            raise PageReadError(f"Claude APIエラー: {e}") from e

        text = "".join(block.text for block in res.content if block.type == "text")
        return _extract_json(text)

    raise PageReadError(f"Claude のレート制限が続いたため読み取れませんでした: {last_error}")


# ---------------------------------------------------------------------------
# 読み取り結果 → PersonRecord
# ---------------------------------------------------------------------------

def build_item_lookup(master) -> dict[str, tuple[str, str]]:
    """変換マスタから {項目名: (分類, 値種別)} を作る。分類の正はマスタ側。"""
    lookup: dict[str, tuple[str, str]] = {}
    for rule in master.item_map:
        lookup.setdefault(rule.item, (rule.category, rule.value_type))
    return lookup


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_note(value: str) -> tuple[str, str]:
    """値の先頭・末尾に付いた注記記号（＊）を外し、(値, 注記) にする。"""
    note = ""
    text = value.strip()
    while text and text[0] in _NOTE_MARKS:
        note, text = "＊", text[1:].strip()
    while text and text[-1] in _NOTE_MARKS:
        note, text = "＊", text[:-1].strip()
    return text, note


def _to_decimal(value: str):
    try:
        return Decimal(unicodedata.normalize("NFKC", value))
    except (InvalidOperation, ValueError):
        return None


def _alias(item: str) -> tuple[str, int]:
    return ITEM_ALIASES.get(item, (item, 1))


def readings_to_person(raw: dict, page_no: int, item_lookup: dict) -> PersonRecord:
    """1ページ分の読み取り結果を PersonRecord にする（API・I/Oなしの純ロジック）。"""
    identity = raw.get("identity") or {}
    name = _text(identity.get("氏名"))
    key = f"p{page_no:02d}"
    sheet = f"P{page_no:02d}_{name.replace(' ', '').replace('　', '')}"[:31]
    issues: list[Issue] = []

    exam_date = parse_exam_date(identity.get("受診日"))
    if exam_date is None:
        issues.append(Issue(
            "error", "EXAM_DATE_INVALID",
            f"{name}: 受診日を読み取れません（{_text(identity.get('受診日'))!r}）"))

    exam_no, exam_no_issue = format_exam_no(identity.get("受診No"))
    if exam_no_issue:
        issues.append(Issue(exam_no_issue.level, exam_no_issue.code,
                            f"{name}: {exam_no_issue.message}"))

    age_text = _text(identity.get("年齢"))
    try:
        age = int(float(age_text)) if age_text else None
    except ValueError:
        age = None

    judgements = raw.get("judgements") or {}
    person = PersonRecord(
        key=key, name=name, age=age, gender=_text(identity.get("性別")),
        exam_date=exam_date, exam_no=exam_no, sheet=sheet, issues=issues,
    )

    _read_blood_pressure(raw, person, judgements, page_no)
    _read_metrics(raw, person, judgements, item_lookup, page_no)
    _read_qualitative(raw, person, item_lookup, page_no)
    _check_bmi(person)

    for note in raw.get("needs_check") or []:
        text = _text(note)
        if text:
            person.issues.append(Issue("info", "READ_NEEDS_CHECK", f"{name}: {text}"))

    return person


def _read_blood_pressure(raw, person: PersonRecord, judgements: dict, page_no: int) -> None:
    """血圧。測定回が分からないものは値ごと捨てる（推測して埋めない）。"""
    seen: set[tuple[str, int]] = set()
    for entry in raw.get("blood_pressure") or []:
        occurrence = entry.get("occurrence")
        if occurrence is None:
            person.issues.append(Issue(
                "error", "BP_OCCURRENCE_UNKNOWN",
                f"{person.name}: 血圧が何回目の測定か原票から判断できませんでした。"
                "原票を確認し、整形済みExcel（スキーマ2.0）を作ってから読み込んでください"))
            continue
        try:
            occurrence = int(occurrence)
        except (TypeError, ValueError):
            person.issues.append(Issue(
                "error", "BP_OCCURRENCE_UNKNOWN",
                f"{person.name}: 血圧の測定回 {occurrence!r} を読み取れません"))
            continue

        for item, source_key in BP_ITEMS.items():
            value, note = _split_note(_text(entry.get(source_key)))
            if not value:
                continue
            if (item, occurrence) in seen:
                continue
            seen.add((item, occurrence))
            person.metrics.append(HealthMetric(
                category="血圧", item=item, occurrence=occurrence, value=value,
                value_type="数値", unit="mmHg",
                source_judgement=_text(judgements.get("血圧")),
                source_note=note, original_display=value,
                source_page=str(page_no), source_sheet=person.sheet,
            ))


def _read_metrics(raw, person: PersonRecord, judgements: dict,
                  item_lookup: dict, page_no: int) -> None:
    for entry in raw.get("metrics") or []:
        item = _text(entry.get("item"))
        if not item or item in DROP_ITEMS:
            continue
        value, note = _split_note(_text(entry.get("value")))
        if not value:
            continue  # 空欄は空欄のまま。勝手に埋めない
        alias, occurrence = _alias(item)
        category, value_type = item_lookup.get(
            alias, (_text(entry.get("category")), "数値"))
        person.metrics.append(HealthMetric(
            category=category, item=alias, occurrence=occurrence, value=value,
            value_type=value_type or "数値",
            source_judgement=_text(judgements.get(category)),
            source_note=note or _text(entry.get("note")),
            original_display=value,
            source_page=str(page_no), source_sheet=person.sheet,
        ))


def _read_qualitative(raw, person: PersonRecord, item_lookup: dict, page_no: int) -> None:
    for entry in raw.get("qualitative") or []:
        item = _text(entry.get("item"))
        if not item or item in DROP_ITEMS:
            continue
        raw_value = _text(entry.get("value"))
        if not raw_value:
            continue  # 未実施・空欄。(-) にはしない
        value, issue = normalize_qualitative(raw_value)
        alias, occurrence = _alias(item)
        category, _vt = item_lookup.get(alias, (_text(entry.get("category")), "定性"))
        if issue is not None:
            person.issues.append(Issue(
                issue.level, issue.code,
                f"{person.name}: {category}/{alias} — {issue.message}"))
            continue
        if not value:
            continue
        method = _text(entry.get("method"))
        display = f"{item} {method} {raw_value}".strip() if method else raw_value
        person.metrics.append(HealthMetric(
            category=category, item=alias, occurrence=occurrence, value=value,
            value_type="定性", original_display=display,
            source_page=str(page_no), source_sheet=person.sheet,
        ))


def _check_bmi(person: PersonRecord) -> None:
    """身長・体重から出したBMIと読み取ったBMIがずれていたら警告する。

    読み違いに気付くための網。値は直さない（原票が正）。
    """
    values = {m.item: m.value for m in person.metrics if m.category != "血圧"}
    height = _to_decimal(values.get("身長", ""))
    weight = _to_decimal(values.get("体重", ""))
    bmi = _to_decimal(values.get("BMI", ""))
    if height is None or weight is None or bmi is None or height <= 0:
        return
    meters = height / Decimal(100)
    calculated = weight / (meters * meters)
    if abs(calculated - bmi) > Decimal(str(BMI_TOLERANCE)):
        person.issues.append(Issue(
            "warning", "BMI_MISMATCH",
            f"{person.name}: BMI {bmi} が身長{height}cm・体重{weight}kgから計算した値"
            f"（{calculated:.1f}）と合いません。原票を確認してください"))


# ---------------------------------------------------------------------------
# 全体の制御
# ---------------------------------------------------------------------------

def analyze_health_pdf(pdf_path: str, master, *, source_filename: str = "",
                       client=None, dpi: int = PDF_RENDER_DPI):
    """PDFを読み、進捗を yield しながら最後に PdfAnalysisResult を返すジェネレータ。

    ("progress", {...}) を都度、最後に ("result", PdfAnalysisResult)。
    SSE は yield でしか送れないので、コールバックではなくジェネレータにしている。
    """
    parse = WorkbookParseResult(
        schema_version="2.0",
        genpyo_confirmed=False,   # 画面で人がチェックするまで False
        bp_occurrence_kept=True,
        qualitative_kept=True,
        source_filename=source_filename or pdf_path,
    )
    result = PdfAnalysisResult(parse=parse)
    item_lookup = build_item_lookup(master)

    yield ("progress", {"message": f"PDFを画像にしています（{dpi}dpi）…"})
    pages = render_pdf_pages(pdf_path, dpi)
    total = len(pages)

    if client is None:
        client = build_client()

    for page_no, png in enumerate(pages, start=1):
        result.page_pngs[page_no] = png
        yield ("progress", {"message": f"{page_no}/{total}ページ目を読み取っています…",
                            "page": page_no, "total": total})
        try:
            raw = _call_claude_for_page(client, page_images_for_reading(png))
        except PageReadError as e:
            parse.issues.append(Issue(
                "error", "PDF_PAGE_READ_FAILED",
                f"{page_no}ページ目を読み取れませんでした: {e}"))
            yield ("progress", {"message": f"{page_no}/{total}ページ目は読み取れませんでした",
                                "page": page_no, "total": total})
            continue

        name = _text((raw.get("identity") or {}).get("氏名"))
        if not name:
            yield ("progress", {
                "message": f"{page_no}/{total}ページ目は個人票ではないので飛ばします",
                "page": page_no, "total": total})
            continue

        person = readings_to_person(raw, page_no, item_lookup)
        parse.persons.append(person)
        result.pages[person.key] = page_no
        numeric = len(person.numeric())
        qualitative = len(person.qualitative())
        yield ("progress", {
            "message": (f"{page_no}/{total}ページ目 {person.name} を読み取りました"
                        f"（数値{numeric}件・定性{qualitative}件）"),
            "page": page_no, "total": total, "name": person.name})

    if not parse.persons:
        parse.issues.append(Issue(
            "error", "PDF_NO_PERSONS",
            "個人票のページが1枚も見つかりませんでした。健診結果のPDFか確認してください"))

    # 血圧の欠落・3回目の指摘はExcel経路と同じ関数で付ける（挙動を1本にするため）
    for person in parse.persons:
        _check_blood_pressure(person)

    yield ("result", result)


# ---------------------------------------------------------------------------
# 監査用の整形済みExcel
# ---------------------------------------------------------------------------
# CSVを作ったあと、後から「何を根拠にこの値を出したか」を辿れるように、
# 原票画像を貼った整形済みExcel（スキーマ2.0）をCSVと同じフォルダへ残す。
# 中身は PersonRecord から書くので、これを読み直せばエラー0件になる。

RECIPIENT_HEADERS = ["PDFページ", "氏名", "年齢", "性別", "受診日", "受診No.", "個人票シート"]
ITEM_HEADERS = ["PDFページ", "氏名", "受診日", "分類", "項目", "値", "単位",
                "原票判定", "原票注記", "個人票シート", "測定回", "値種別", "原票表記"]
SUMMARY_HEADERS = ["PDFページ", "氏名", "年齢", "性別", "受診日", "受診No.",
                   "身長(cm)", "体重(kg)", "BMI", "腹囲(cm)",
                   "収縮期血圧1回目(mmHg)", "拡張期血圧1回目(mmHg)",
                   "収縮期血圧2回目(mmHg)", "拡張期血圧2回目(mmHg)",
                   "A以外の原票判定", "個人票シート", "原票ページ確認"]

_AUDIT_MAX_SUFFIX = 9


def _audit_sheet(wb, title, headers, rows, description="", widths=None):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title)
    cell = ws.cell(1, 1)
    cell.value = title
    cell.font = Font(name="Meiryo UI", size=12, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor="FF1F4E79")
    if description:
        ws.cell(2, 1).value = description
        ws.cell(2, 1).font = Font(name="Meiryo UI", size=9, color="FF555555")
    for i, name in enumerate(headers, start=1):
        head = ws.cell(3, i)
        head.value = name
        head.font = Font(name="Meiryo UI", size=9.5, bold=True)
        head.fill = PatternFill("solid", fgColor="FFD9E2F3")
        head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=4):
        for c, value in enumerate(row, start=1):
            target = ws.cell(r, c)
            target.value = value
            target.font = Font(name="Meiryo UI", size=9.5)
    if widths:
        from openpyxl.utils import get_column_letter
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"
    return ws


def _unique_output_path(out_path: str) -> str:
    """同名があれば _2, _3 … を付ける。既存ファイルは絶対に壊さない。"""
    import os

    if not os.path.exists(out_path):
        return out_path
    stem, ext = os.path.splitext(out_path)
    for n in range(2, _AUDIT_MAX_SUFFIX + 1):
        candidate = f"{stem}_{n}{ext}"
        if not os.path.exists(candidate):
            return candidate
    raise FileExistsError(f"同名のファイルが多すぎます: {out_path}")


def write_audit_workbook(out_path: str, parse: WorkbookParseResult, pages: dict[str, int],
                         page_png_bytes: dict[int, bytes], *, pdf_name: str,
                         confirmed_at: str) -> str:
    """原票画像つきの整形済みExcel（スキーマ2.0）を書き、実際のパスを返す。"""
    import os

    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill

    out_path = _unique_output_path(out_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- 変換案内（スキーマの宣言。ここが揃っていないと読み直しでエラーになる） ---
    info = wb.create_sheet("変換案内")
    info.cell(1, 1).value = "健康診断PDF - Excel変換版（スキーマ2.0）"
    info.cell(1, 1).font = Font(name="Meiryo UI", size=12, bold=True, color="FFFFFFFF")
    info.cell(1, 1).fill = PatternFill("solid", fgColor="FF1F4E79")
    info.cell(3, 1).value = "項目"
    info.cell(3, 2).value = "内容"
    for col in (1, 2):
        info.cell(3, col).font = Font(name="Meiryo UI", size=9.5, bold=True)
        info.cell(3, col).fill = PatternFill("solid", fgColor="FFD9E2F3")
    pairs = [
        ("原本", pdf_name),
        ("対象", f"健康診断結果{len(parse.persons)}名"),
        ("変換方法", "勤怠チェッカーが原票画像をClaudeで読み取り（ページ全体＋上下2分割の拡大）"),
        ("健康診断整形スキーマ版", "2.0"),
        ("原票確認済み", "TRUE"),
        ("血圧測定回保持", "TRUE"),
        ("定性値保持", "TRUE"),
        ("確認方法", f"チェッカー画面で原票画像と照合（{confirmed_at}）"),
        ("注意", "血圧は1回目・2回目を原票どおり別々に持ちます（平均は作りません）。"
                 "(-) は陰性を表す有効な値です。"),
        ("原本保護", "元PDFは変更していません。"),
    ]
    for r, (key, value) in enumerate(pairs, start=4):
        info.cell(r, 1).value = key
        info.cell(r, 2).value = value
        info.cell(r, 1).font = Font(name="Meiryo UI", size=9.5, bold=key in (
            "健康診断整形スキーマ版", "原票確認済み", "確認方法"))
        info.cell(r, 2).font = Font(name="Meiryo UI", size=9.5)
        info.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 96

    # --- 各シートの行を作る ---
    rec_rows, item_rows, sum_rows = [], [], []
    for person in parse.persons:
        page = pages.get(person.key, 0)
        page_text = str(page) if page else ""
        exam_date = person.exam_date.isoformat() if person.exam_date else ""
        rec_rows.append([page_text, person.name, person.age, person.gender,
                         exam_date, person.exam_no, person.sheet])

        for metric in person.metrics:
            item_rows.append([
                page_text, person.name, exam_date, metric.category, metric.item,
                metric.value, metric.unit, metric.source_judgement, metric.source_note,
                person.sheet, metric.occurrence, metric.value_type,
                metric.original_display or metric.value,
            ])

        bp = person.blood_pressure()
        values = {m.item: m.value for m in person.metrics if m.category != "血圧"}
        judged = sorted({(m.category, m.source_judgement) for m in person.metrics
                         if m.source_judgement and m.source_judgement != "A"})
        sum_rows.append([
            page_text, person.name, person.age, person.gender, exam_date, person.exam_no,
            values.get("身長", ""), values.get("体重", ""), values.get("BMI", ""),
            values.get("腹囲", ""),
            bp.get(1, {}).get("sys", ""), bp.get(1, {}).get("dia", ""),
            bp.get(2, {}).get("sys", ""), bp.get(2, {}).get("dia", ""),
            " / ".join(f"{c}:{j}" for c, j in judged),
            person.sheet, f"PDF {page}ページ" if page else "",
        ])

    _audit_sheet(wb, "受診者一覧", RECIPIENT_HEADERS, rec_rows,
                 widths=[10, 18, 8, 8, 14, 12, 22])

    # --- 原票画像のページシート ---
    for person in parse.persons:
        page = pages.get(person.key)
        png = page_png_bytes.get(page) if page else None
        ws = wb.create_sheet(person.sheet)
        ws.cell(1, 1).value = f"{person.name}／原票 {page}ページ"
        ws.cell(1, 1).font = Font(name="Meiryo UI", size=12, bold=True)
        ws.cell(2, 1).value = ("この画像が正本です。血圧の1回目・2回目と定性検査の (-) は"
                               "この画像で確認したものです。")
        ws.cell(2, 1).font = Font(name="Meiryo UI", size=9, color="FFA32D2D")
        ws.cell(3, 1).value = pdf_name
        if png:
            image = XLImage(io.BytesIO(png))
            scale = 1000 / image.height if image.height else 1
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
            ws.add_image(image, "A5")

    _audit_sheet(wb, "整形サマリー", SUMMARY_HEADERS, sum_rows,
                 description="血圧は1回目・2回目を別の列に出しています（平均は作りません）。",
                 widths=[9, 16, 6, 6, 12, 10] + [12] * 8 + [30, 20, 14])
    _audit_sheet(wb, "項目別データ", ITEM_HEADERS, item_rows,
                 description="1項目1行。測定回・値種別・原票表記はスキーマ2.0の追加列です。",
                 widths=[9, 16, 12, 12, 20, 10, 8, 9, 9, 20, 7, 9, 22])

    notes = [[f"[{i.level}] {i.message}"] for i in parse.all_issues()
             if i.level in ("info", "warning")]
    _audit_sheet(wb, "要確認メモ", ["内容"], notes,
                 description="読み取り時に「読めなかった／確認が要る」と申告された項目です。",
                 widths=[130])

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
