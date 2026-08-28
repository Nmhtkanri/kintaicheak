# -*- coding: utf-8 -*-
"""台帳テンプレート（.xlsx）を旧フォーム（派遣元管理台帳202504.xlsm の『派遣元管理台帳』シート）から作る。

- 見た目（Meiryo UI 10pt・罫線・B:C ラベル結合・D:N 値結合・A4縦）は旧フォームをそのまま複製する
- 値（前任者が最後に出力した個人のデータ）は一切コピーしない。ラベルだけ写す
- 旧マクロで行がズレていた 就業曜日(24)/休日(26)/就業時間外(28) は正しい行に割り当てる
- 旧フォームに無かった法定記載事項（抵触日・資格取得届の理由・教育訓練・キャリコン・雇用安定措置・
  苦情処理状況・派遣先通知）を 57 行目以降に同じ様式で追加する
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .config import SOURCE_FORM_SHEET, TEMPLATE_SHEET

ORIGINAL_ROWS = 56        # 旧フォームの最終行（備考）
FIRST_COL, LAST_COL = 1, 15   # A..O（印刷範囲）
LABEL_STYLE_CELL = "B56"  # 1行ラベルの見本
VALUE_STYLE_CELL = "D56"  # 1行値欄の見本

# 旧フォーム上の小見出し（D17/D18/D20/D22 は旧フォームでは「契約期間/就業時間1/2/3」）
SUB_LABELS = {
    "D17": "契約期間",
    "D18": "就業時間",        # D18:D19 結合
    "D20": "休憩\n（続き）",  # D20:D21 結合。D列は狭いので折り返し位置を明示する
    "D22": "契約書\n確定日",  # D22:D23 結合
    "G3": "加入有無",
    "I2": "健康保険",
    "K2": "厚生年金",
    "M2": "雇用保険",
}

# 57行目以降に追加する行: (ラベル, 項目キー)
EXTRA_ROWS: list[tuple[str, str]] = [
    ("事業所単位の抵触日", "事業所単位の抵触日"),
    ("個人単位の抵触日", "個人単位の抵触日"),
    ("資格取得届の提出\n（健康保険）", "資格取得届_健康保険"),
    ("資格取得届の提出\n（厚生年金）", "資格取得届_厚生年金"),
    ("資格取得届の提出\n（雇用保険）", "資格取得届_雇用保険"),
    ("便宜供与", "便宜供与"),
    ("安全及び衛生", "安全及び衛生"),
    ("教育訓練の\n日時及び内容", "教育訓練の日時及び内容"),
    ("キャリア\nコンサルティングの\n日時及び内容", "キャリアコンサルティングの日時及び内容"),
    ("雇用安定措置の内容", "雇用安定措置の内容"),
    ("苦情の処理状況", "苦情の処理状況"),
    ("派遣先へ\n通知した事項", "派遣先へ通知した事項"),
]
LAST_ROW = ORIGINAL_ROWS + len(EXTRA_ROWS)

# 項目キー → 書き込みセル（結合範囲の左上）
CELL_MAP: dict[str, str] = {
    "社員番号": "D2",
    "氏名": "E2",
    "生年月日": "D3",
    "氏名・年齢・性別": "E3",
    "健康保険_加入": "I3",
    "厚生年金_加入": "K3",
    "雇用保険_加入": "M3",
    "派遣先名称": "D5",
    "派遣先の事業所所在地": "D6",
    "派遣先TEL": "D7",
    "就業場所": "D8",
    "就業場所住所": "D9",
    "就業場所TEL": "D10",
    "組織単位": "D11",
    "業務内容": "D12",
    "責任の程度": "D13",
    "協定対象派遣労働者であるか否かの別": "D14",
    "有期か無期かの別": "D15",
    "60歳以上の者であるか否かの別": "D16",
    "契約期間": "E17",
    "就業時間": "E18",
    "休憩時間1": "E19",
    "休憩時間2": "E20",
    "職種": "E21",
    "契約書確定日": "E22",
    "契約No": "E23",
    "就業曜日": "D24",
    "就業曜日備考": "D25",
    "休日": "D26",
    "休日備考": "D27",
    "就業時間外の労働": "D28",
    "休日勤務": "D29",
    "就業状況": "D30",
    "派遣先責任者_部署": "D31",
    "派遣先責任者_役職氏名": "D32",
    "派遣先責任者_TEL": "D33",
    "苦情申出先_部署": "D34",
    "苦情申出先_役職氏名": "D35",
    "苦情申出先_TEL": "D36",
    "製造業務専門派遣先責任者": "D37",
    "指揮命令者_部署": "D40",
    "指揮命令者_役職氏名": "D41",
    "指揮命令者_TEL": "D42",
    "派遣元責任者_部署": "D43",
    "派遣元責任者_役職氏名": "D44",
    "派遣元責任者_TEL": "D45",
    "派遣元苦情処理担当者_部署": "D46",
    "派遣元苦情処理担当者_役職氏名": "D47",
    "派遣元苦情処理担当者_TEL": "D48",
    "製造業務専門派遣元責任者": "D49",
    "日数限定業務・産休代替等": "D52",
    "期間制限の対象外理由": "D53",
    "有期プロジェクト業務": "D55",
    "備考": "D56",
}
for _i, (_label, _key) in enumerate(EXTRA_ROWS):
    CELL_MAP[_key] = f"D{ORIGINAL_ROWS + 1 + _i}"

# 長文になりうる項目（行の高さを内容に合わせる）
LONG_TEXT_KEYS = {
    "業務内容", "責任の程度", "組織単位", "就業場所", "就業時間外の労働", "休日", "休日備考", "備考",
    "安全及び衛生", "派遣先へ通知した事項", "個人単位の抵触日", "資格取得届_健康保険", "資格取得届_厚生年金",
    "資格取得届_雇用保険", "就業曜日", "派遣先の事業所所在地", "就業場所住所",
    "就業時間", "休憩時間1", "休憩時間2", "職種",   # E列ブロック（シフト制の説明文が長くなる）
}


def _clone_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def build_template(source_xlsm: Path | str, out_xlsx: Path | str, *, keep_values: bool = False) -> Path:
    """旧フォームから値を持たないテンプレートを作る。戻り値は出力パス。"""
    src_wb = openpyxl.load_workbook(source_xlsm, data_only=True)
    if SOURCE_FORM_SHEET not in src_wb.sheetnames:
        raise ValueError(f"旧フォームに『{SOURCE_FORM_SHEET}』シートがありません: {source_xlsm}")
    src = src_wb[SOURCE_FORM_SHEET]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET

    # 1) セルの様式を複製（値はラベル列 B と G2 だけ）
    for row in src.iter_rows(min_row=1, max_row=ORIGINAL_ROWS, min_col=FIRST_COL, max_col=LAST_COL):
        for cell in row:
            dst = ws.cell(row=cell.row, column=cell.column)
            if cell.has_style:
                _clone_style(cell, dst)
            if cell.value is not None and (cell.column == 2 or cell.coordinate == "G2" or keep_values):
                dst.value = cell.value
    for coord, text in SUB_LABELS.items():
        c = ws[coord]
        c.value = text
        # 折り返しON（D列は全角4文字強しかなく、OFFだと「契約書確定日」等の右側が印刷で切れる）
        a = c.alignment
        c.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical or "center", wrap_text=True)
    # 旧フォーム由来のラベルのうち、3行結合で縦が詰まる「日数限定業務・産休代替等」は2行に組み直す
    ws["B52"].value = "日数限定業務・\n産休代替等"

    # 2) 結合セル・列幅・行高
    for rng in src.merged_cells.ranges:
        if rng.max_row <= ORIGINAL_ROWS and rng.max_col <= LAST_COL:
            ws.merge_cells(str(rng))
    for col in range(FIRST_COL, LAST_COL + 1):
        letter = get_column_letter(col)
        w = src.column_dimensions[letter].width if letter in src.column_dimensions else None
        if w:
            ws.column_dimensions[letter].width = w
    for r in range(1, ORIGINAL_ROWS + 1):
        h = src.row_dimensions[r].height if r in src.row_dimensions else None
        if h:
            ws.row_dimensions[r].height = h

    # 先頭ブロックの手当て: E列は全角4文字強しかなく「契約開始時 47歳　男性」等が切れるため
    # E:F を結合して幅を確保し、年齢・性別の行は2行分の高さにする
    ws.merge_cells("E2:F2")
    ws.merge_cells("E3:F3")
    ws.row_dimensions[3].height = 32
    for r in (20, 21, 22, 23):
        ws.row_dimensions[r].height = 15   # D列小見出し（休憩(続き)/契約書確定日）が2行に折り返す分

    # 3) 追加行（ラベル B:C 結合、値 D:N 結合。様式は備考行を見本にする）
    label_src, value_src = ws[LABEL_STYLE_CELL], ws[VALUE_STYLE_CELL]
    for i, (label, _key) in enumerate(EXTRA_ROWS):
        r = ORIGINAL_ROWS + 1 + i
        for col in range(FIRST_COL, LAST_COL + 1):
            dst = ws.cell(row=r, column=col)
            model = label_src if col in (2, 3) else value_src if 4 <= col <= 14 else None
            if model is not None:
                _clone_style(model, dst)
        lbl = ws.cell(row=r, column=2)
        lbl.value = label
        # 折り返しON必須。OFFだと改行が効かず1行扱いになり、中央揃えのまま左右が印刷で切れる
        lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=14)
        n_lines = label.count("\n") + 1
        ws.row_dimensions[r].height = {1: 18, 2: 32}.get(n_lines, 46)

    # 4) 印刷設定（A4縦・1契約＝1枚に収める。旧フォームも1枚だった）
    ws.print_area = f"A1:{get_column_letter(LAST_COL)}{LAST_ROW}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    out = Path(out_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
