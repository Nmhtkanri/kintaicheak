# -*- coding: utf-8 -*-
"""LedgerRecord をテンプレートに流し込み、四半期1冊の .xlsx（1契約=1シート）＋一覧CSV＋警告CSVを書く。"""
from __future__ import annotations

import csv
import datetime as dt
import math
import re
import unicodedata
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import TEMPLATE_SHEET, quarter_label
from .records import CSV_COLUMNS, LedgerRecord
from .template import CELL_MAP, LAST_COL, LAST_ROW, LONG_TEXT_KEYS

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
VALUE_WIDTH_UNITS = 92.0   # D:N 結合セルのおおよその幅（半角文字数相当）
LINE_HEIGHT_PT = 14.5

# 先頭ブロックの狭いセルは折り返すと行高が足りず印刷で切れる。
# 氏名は E2:F2 に1行で左詰め、生年月日は右隣（年齢・性別）が埋まっていてはみ出せないので縮小で収める
NO_WRAP_KEYS = {"氏名"}
SHRINK_KEYS = {"生年月日"}


def _display_width(s: str) -> float:
    return sum(2.0 if unicodedata.east_asian_width(ch) in ("W", "F", "A") else 1.0 for ch in s)


def estimate_row_height(text: str, width_units: float = VALUE_WIDTH_UNITS) -> float:
    lines = 0
    for seg in str(text).split("\n"):
        lines += max(1, math.ceil(_display_width(seg) / width_units))
    if lines <= 1:
        return 15.0          # 1行なら旧フォームの既定の高さのまま
    return min(409.0, lines * LINE_HEIGHT_PT + 2)


def sheet_title(rec: LedgerRecord, used: set[str]) -> str:
    base = f"{rec.emp_id or 'X'}_{re.sub(r'[\s　]+', '', rec.name)}"
    base = _INVALID_SHEET_CHARS.sub("", base)[:28] or "台帳"
    title, n = base, 1
    while title in used:
        n += 1
        title = f"{base[:28 - len(str(n)) - 1]}_{n}"
    used.add(title)
    return title


def _sort_key(rec: LedgerRecord):
    c = rec.contract
    kana = (c.t("スタッフ姓（カナ）") + c.t("スタッフ名（カナ）")) if c else ""
    return (kana or "ﾝ", rec.name, c.start if c and c.start else dt.date.min, rec.fields.get("契約No", ""))


def fill_sheet(ws, rec: LedgerRecord) -> None:
    for key, coord in CELL_MAP.items():
        value = rec.fields.get(key, "")
        cell = ws[coord]
        cell.value = value if value != "" else None
        al = copy(cell.alignment) if cell.alignment else Alignment()
        if key in NO_WRAP_KEYS:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        elif key in SHRINK_KEYS:
            cell.alignment = Alignment(horizontal=al.horizontal, vertical="center",
                                       wrap_text=False, shrink_to_fit=True)
        else:
            cell.alignment = Alignment(horizontal=al.horizontal, vertical="center", wrap_text=True,
                                       indent=al.indent, shrink_to_fit=False)
        if key in LONG_TEXT_KEYS and value:
            row = cell.row
            # E列始まりの結合（E:N）はD:Nより1列ぶん狭い
            needed = estimate_row_height(value, 84.0 if coord.startswith("E") else VALUE_WIDTH_UNITS)
            current = ws.row_dimensions[row].height or 15.0
            if needed > current:
                ws.row_dimensions[row].height = needed
    ws.print_area = f"A1:{get_column_letter(LAST_COL)}{LAST_ROW}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1          # 1契約＝1枚
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.showGridLines = False


def _write_index(ws, rows: list[dict], quarter: str) -> None:
    ws["A1"] = f"派遣元管理台帳　{quarter_label(quarter)}（{quarter}）　目次"
    ws["A1"].font = Font(name="Meiryo UI", size=12, bold=True)
    ws["A2"] = f"作成: {dt.datetime.now():%Y/%m/%d %H:%M}　件数: {len(rows)}　※シート名をクリックで該当台帳へ"
    ws["A2"].font = Font(name="Meiryo UI", size=9)
    headers = ["No", "シート", "社員番号", "氏名", "派遣先", "就業場所", "契約期間", "契約No", "警告数"]
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(name="Meiryo UI", size=10, bold=True)
        c.fill = head_fill
    for i, r in enumerate(rows, start=1):
        vals = [i, r["sheet"], r["emp_id"], r["name"], r["client"], r["place"], r["period"], r["contract_no"], r["warn"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=4 + i, column=j, value=v)
            c.font = Font(name="Meiryo UI", size=10, color="0563C1" if j == 2 else None, underline="single" if j == 2 else None)
        ws.cell(row=4 + i, column=2).hyperlink = f"#'{r['sheet']}'!A1"
    for col, w in zip("ABCDEFGHI", (5, 24, 10, 16, 24, 36, 24, 18, 8)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def _write_warnings(ws, items: list[tuple[str, str, str, str]]) -> None:
    ws["A1"] = "警告・要確認事項"
    ws["A1"].font = Font(name="Meiryo UI", size=12, bold=True)
    ws["A2"] = "台帳は作成済み。内容を確認のうえ、必要なら元データ（e-staffing / jinjer）を直して再生成する。"
    ws["A2"].font = Font(name="Meiryo UI", size=9)
    for j, h in enumerate(["区分", "契約No", "氏名", "内容"], start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(name="Meiryo UI", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="FCE4D6")
    for i, (kind, no, name, text) in enumerate(items, start=5):
        for j, v in enumerate((kind, no, name, text), start=1):
            ws.cell(row=i, column=j, value=v).font = Font(name="Meiryo UI", size=10)
    for col, w in zip("ABCD", (12, 18, 16, 100)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def write_quarter(records: list[LedgerRecord], quarter: str, template_path: Path | str, out_dir: Path | str,
                  global_warnings: list[str] | None = None) -> dict[str, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    records = sorted(records, key=_sort_key)
    wb = openpyxl.load_workbook(template_path)
    tpl = wb[TEMPLATE_SHEET]
    index_ws = wb.create_sheet("目次", 0)
    warn_ws = wb.create_sheet("警告", 1)

    used: set[str] = set()
    index_rows: list[dict] = []
    warn_items: list[tuple[str, str, str, str]] = [("全体", "", "", w) for w in (global_warnings or [])]
    for rec in records:
        ws = wb.copy_worksheet(tpl)
        ws.title = sheet_title(rec, used)
        fill_sheet(ws, rec)
        index_rows.append({
            "sheet": ws.title, "emp_id": rec.emp_id, "name": rec.name,
            "client": rec.fields.get("派遣先名称", ""), "place": rec.fields.get("就業場所", ""),
            "period": rec.fields.get("契約期間", ""), "contract_no": rec.fields.get("契約No", ""),
            "warn": len(rec.warnings),
        })
        for w in rec.warnings:
            warn_items.append(("台帳", rec.fields.get("契約No", ""), rec.name, w))
    wb.remove(tpl)
    _write_index(index_ws, index_rows, quarter)
    _write_warnings(warn_ws, warn_items)
    wb.active = 0

    xlsx = out_dir / f"派遣元管理台帳_{quarter}.xlsx"
    wb.save(xlsx)

    csv_path = out_dir / f"派遣元管理台帳_{quarter}_一覧.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fp:
        wr = csv.writer(fp)
        wr.writerow(CSV_COLUMNS)
        for rec in records:
            wr.writerow([rec.fields.get(k, "") for k in CSV_COLUMNS])

    warn_path = out_dir / f"派遣元管理台帳_{quarter}_警告.csv"
    with warn_path.open("w", encoding="utf-8-sig", newline="") as fp:
        wr = csv.writer(fp)
        wr.writerow(["区分", "契約No", "氏名", "内容"])
        wr.writerows(warn_items)
    return {"xlsx": xlsx, "csv": csv_path, "warnings": warn_path}
