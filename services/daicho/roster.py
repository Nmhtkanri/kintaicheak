# -*- coding: utf-8 -*-
"""jinjer の従業員一覧（インポート形式 .xlsx）を人マスタとして読む。

e-staffing は社員番号を持たないので、氏名（姓＋名）で橋渡しして
社員番号・生年月日・性別・雇用区分・所属グループ・勤務先を取る。

シート1（従業員一覧）: ＊従業員ID / ＊氏 / ＊名 / 所属グループ / 退職日 / 勤務先 …
シート「jinjer_勤務先_*」: 社員番号 / 生年月日 / 性別 / 雇用区分 / 入社年月日 / 在籍区分 …
シート「所属G」: 社員番号 / 氏名 / 所属グループ（ヘッダー行なし）
"""
from __future__ import annotations

import datetime as dt
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# 旧字体・異体字 → 新字体（氏名突合用。表示には使わない）
_VARIANTS = str.maketrans({
    "髙": "高", "﨑": "崎", "嵜": "崎", "齋": "斎", "齊": "斉", "濵": "浜", "濱": "浜", "邊": "辺", "邉": "辺",
    "國": "国", "澤": "沢", "櫻": "桜", "廣": "広", "嶋": "島", "嶌": "島", "德": "徳", "淺": "浅", "眞": "真",
    "瀨": "瀬", "冨": "富", "峯": "峰", "橫": "横", "黑": "黒", "靜": "静", "圓": "円", "會": "会", "學": "学",
    "龍": "竜", "榮": "栄", "曾": "曽", "寶": "宝", "惠": "恵", "賴": "頼", "鄕": "郷", "藏": "蔵", "壽": "寿",
    "彌": "弥", "礒": "磯", "舘": "館", "槇": "槙", "萠": "萌", "ヶ": "ケ", "ヱ": "エ", "ゐ": "い", "ゑ": "え",
})


def normalize_name(*parts: str) -> str:
    """突合キー: 全角/半角を揃え、空白を除き、旧字体を新字体に寄せる。"""
    s = "".join(p or "" for p in parts)
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\s　]+", "", s)
    return s.translate(_VARIANTS)


def _to_date(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@dataclass
class Person:
    emp_id: str
    sei: str
    mei: str
    birth: dt.date | None = None
    sex: str = ""
    employment_type: str = ""   # 雇用区分（正社員/契約社員/役員）
    joined: dt.date | None = None
    status: str = ""            # 在籍区分（在籍/休職/退職）
    group: str = ""             # 所属グループ（OT：UAL（首都圏）など）
    workplace: str = ""         # 勤務先
    retired: dt.date | None = None
    notes: list[str] = field(default_factory=list)

    @property
    def name(self) -> str:
        return f"{self.sei} {self.mei}".strip()

    @property
    def key(self) -> str:
        return normalize_name(self.sei, self.mei)

    def age_at(self, on: dt.date | None) -> int | None:
        if self.birth is None or on is None:
            return None
        years = on.year - self.birth.year
        if (on.month, on.day) < (self.birth.month, self.birth.day):
            years -= 1
        return years


@dataclass
class Roster:
    people: dict[str, Person]                     # 社員番号 → Person
    by_name: dict[str, list[Person]]              # 正規化氏名 → Person（同姓同名は複数）
    warnings: list[str] = field(default_factory=list)

    def find(self, sei: str, mei: str) -> tuple[Person | None, str]:
        """氏名で探す。戻り値 (Person|None, 状態)。状態: ok / none / ambiguous。"""
        return self.find_by_key(normalize_name(sei, mei))

    def find_by_key(self, key: str) -> tuple[Person | None, str]:
        hits = self.by_name.get(key, [])
        if len(hits) == 1:
            return hits[0], "ok"
        if not hits:
            return None, "none"
        # 同姓同名: 在籍者を優先、それでも複数なら曖昧
        active = [p for p in hits if p.status != "退職" and p.retired is None]
        if len(active) == 1:
            return active[0], "ok"
        return None, "ambiguous"


def _header_index(row: list) -> dict[str, int]:
    return {str(v).strip(): i for i, v in enumerate(row) if v not in (None, "")}


def _pick(row: list, idx: dict[str, int], *names: str):
    for n in names:
        if n in idx and idx[n] < len(row):
            v = row[idx[n]]
            if v not in (None, ""):
                return v
    return None


def load_roster(path: Path | str) -> Roster:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    warnings: list[str] = []
    people: dict[str, Person] = {}

    # --- シート1: 従業員一覧（インポート形式） ---
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise ValueError(f"従業員一覧が空です: {path}")
    idx = _header_index(rows[0])
    id_col = next((i for n, i in idx.items() if n.replace("＊", "") in ("従業員ID", "社員番号")), None)
    sei_col = next((i for n, i in idx.items() if n.replace("＊", "") in ("氏", "姓")), None)
    mei_col = next((i for n, i in idx.items() if n.replace("＊", "") == "名"), None)
    if id_col is None or sei_col is None or mei_col is None:
        raise ValueError("従業員一覧に 従業員ID／氏／名 の列が見つかりません")
    # 勤務先・所属グループは同名列が2組ある（左=現在値、右=更新値）。右を優先する
    work_cols = [i for i, v in enumerate(rows[0]) if str(v).strip() == "勤務先"]
    group_cols = [i for i, v in enumerate(rows[0]) if str(v).strip() == "所属グループ"]
    retire_col = idx.get("退職日")
    for r in rows[1:]:
        if id_col >= len(r) or r[id_col] in (None, ""):
            continue
        emp_id = str(r[id_col]).strip()
        if emp_id in people:
            warnings.append(f"従業員一覧に社員番号 {emp_id} が重複")
            continue
        workplace = next((str(r[i]).strip() for i in reversed(work_cols)
                          if i < len(r) and r[i] not in (None, "", 0, "0")), "")
        group = next((str(r[i]).strip() for i in reversed(group_cols)
                      if i < len(r) and r[i] not in (None, "")), "")
        people[emp_id] = Person(
            emp_id=emp_id, sei=str(r[sei_col] or "").strip(), mei=str(r[mei_col] or "").strip(),
            workplace=workplace, group=group,
            retired=_to_date(r[retire_col]) if retire_col is not None and retire_col < len(r) else None,
        )

    # --- シート jinjer_勤務先_*: 生年月日・性別・雇用区分・入社日・在籍区分 ---
    detail_ws = next((wb[n] for n in wb.sheetnames if n.startswith("jinjer_勤務先")), None)
    if detail_ws is None:
        for n in wb.sheetnames[1:]:
            first = next(wb[n].iter_rows(min_row=1, max_row=1, values_only=True), ())
            names = [str(v).strip() for v in first if v is not None]
            if "社員番号" in names and "生年月日" in names:
                detail_ws = wb[n]
                break
    if detail_ws is not None:
        drows = [list(r) for r in detail_ws.iter_rows(values_only=True)]
        didx = _header_index(drows[0])
        for r in drows[1:]:
            emp_id = _pick(r, didx, "社員番号")
            if emp_id is None:
                continue
            emp_id = str(emp_id).strip()
            p = people.get(emp_id)
            if p is None:
                # 一覧に無い人（退職者など）も人マスタとしては持っておく
                p = Person(emp_id=emp_id,
                           sei=str(_pick(r, didx, "職場氏名(氏)", "氏") or ""),
                           mei=str(_pick(r, didx, "職場氏名(名)", "名") or ""))
                people[emp_id] = p
            p.birth = _to_date(_pick(r, didx, "生年月日"))
            p.sex = str(_pick(r, didx, "性別") or "")
            p.employment_type = str(_pick(r, didx, "雇用区分") or "")
            p.joined = _to_date(_pick(r, didx, "入社年月日"))
            p.status = str(_pick(r, didx, "在籍区分") or "")
            if not p.workplace:
                p.workplace = str(_pick(r, didx, "就業先企業名(就業先情報)") or "")
    else:
        warnings.append("jinjer_勤務先 シートが無いため 生年月日・雇用区分 が取れません（60歳判定は契約データの宣言のみ）")

    # --- シート 所属G（ヘッダー無し: 社員番号, 氏名, 所属グループ） ---
    if "所属G" in wb.sheetnames:
        for r in wb["所属G"].iter_rows(values_only=True):
            r = list(r)
            if len(r) >= 3 and r[0] not in (None, "") and r[2] not in (None, ""):
                key = str(r[0]).strip()
                if key in people and not people[key].group:
                    people[key].group = str(r[2]).strip()

    by_name: dict[str, list[Person]] = {}
    for p in people.values():
        by_name.setdefault(p.key, []).append(p)
    dup = [k for k, v in by_name.items() if len(v) > 1]
    if dup:
        warnings.append(f"同姓同名が {len(dup)} 組あります（在籍者を優先して突合）")
    return Roster(people=people, by_name=by_name, warnings=warnings)
