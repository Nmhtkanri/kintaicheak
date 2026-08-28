# -*- coding: utf-8 -*-
"""直接契約の契約書（Excel/テキストPDF）から「直接契約マスタ_自動.csv」を再生成する。

対応様式:
- nmht  … 自社様式（シート[個別契約書/管理台帳/通知書]）。オリゾン・BCS・日本ディクス・
          ルミナス・マンパワーが同一様式（.xls/.xlsx 両対応）
- itone … アイ・ティー・ワンの月次テキストPDF（Meeepa出力）

スキャンPDF等は対象外＝「直接契約マスタ_手入力.csv」に手で行を起こす（load側で手入力が勝つ）。
"""
from __future__ import annotations

import csv
import datetime as dt
import re
import unicodedata
from pathlib import Path

from .direct import AUTO_CSV, COLUMNS
from .estaffing import parse_date

CONTRACT_BASE = Path(r"Z:\NetMarks以外(常駐）\1,契約書")
KEEP_FROM = dt.date(2025, 7, 1)      # これより前に終わる契約はマスタに載せない

FAMILIES = [
    {"name": "オリゾン", "glob": "オリゾンシステムズ/労働者派遣個別契約書_*.xls*", "format": "nmht"},
    {"name": "BCS", "glob": "BCS/FY*/労働者派遣個別契約書_*.xls*", "format": "nmht"},
    {"name": "日本ディクス", "glob": "日本ディクス/**/労働者派遣個別契約書_*.xls*", "format": "nmht"},
    {"name": "ルミナス", "glob": "ルミナス・ビー・ジャパン/**/*派遣個別契約書_*.xls*", "format": "nmht"},
    {"name": "マンパワー", "glob": "マンパワーグループ/20*年度/**/【マンパワー様】労働者派遣個別契約書_*.xls*", "format": "nmht"},
    {"name": "アイ・ティー・ワン", "glob": "アイ・ティー・ワン/労働者派遣個別契約書_*.pdf", "format": "itone"},
    {"name": "ペンギン", "glob": "ペンギンソリューションズ（旧ストラタス）/**/派遣個別契約書_菅原孝_*.xlsx",
     "format": "penguin", "worker": "菅原 孝"},
]

_norm = lambda s: re.sub(r"[\s　]+", "", unicodedata.normalize("NFKC", str(s or "")))


# ---------------------------------------------------------------- グリッド化
def _cell_to_str(v, datemode=None) -> str:
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, dt.date):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, float):
        if datemode is not None and 20000 < v < 80000:   # Excelシリアル日付とみなす
            import xlrd
            try:
                return xlrd.xldate_as_datetime(v, datemode).strftime("%Y/%m/%d")
            except Exception:
                pass
        if v == int(v):
            return str(int(v))
    return str(v).strip()


def load_grids(path: Path) -> dict[str, list[list[str]]]:
    """{シート名: 2次元文字列} を返す。.xls は xlrd、.xlsx は openpyxl。"""
    grids: dict[str, list[list[str]]] = {}
    if path.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        for sn in wb.sheet_names():
            sh = wb.sheet_by_name(sn)
            grids[sn] = [[_cell_to_str(sh.cell_value(r, c), wb.datemode)
                          for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            grids[sn] = [[_cell_to_str(v) for v in row]
                         for row in wb[sn].iter_rows(values_only=True)]
    return grids


# ---------------------------------------------------------------- 自社様式(nmht)
_BLOCKS = {"派遣先責任者", "指揮命令者", "苦情の申出を受ける者", "派遣元責任者"}
_SUBLABELS = {"就業曜日", "就業時間", "休憩時間", "実労働時間", "休日", "時間外労働", "名称", "住所",
              "部門名", "役職", "氏名", "TEL", "ＴＥＬ", "許可番号または届出番号"}


def _right_value(row: list[str], idx: int, skip: set[str] = frozenset()) -> str:
    for c in range(idx + 1, len(row)):
        v = row[c].strip()
        if v and _norm(v) not in skip:
            return v
    return ""


def _dates_in_row(row: list[str], idx: int) -> list[str]:
    out = []
    for c in range(idx + 1, len(row)):
        v = row[c].strip()
        if v in ("から", "まで") or not v:
            continue
        if parse_date(v):
            out.append(parse_date(v).strftime("%Y/%m/%d"))
    return out


def parse_nmht(grids: dict[str, list[list[str]]]) -> dict:
    f: dict[str, str] = {}

    def scan(grid: list[list[str]], person_sheet: bool):
        section = ""
        ins_values: list[str] = []
        for ri, row in enumerate(grid):
            cells = [(ci, v) for ci, v in enumerate(row) if str(v).strip()]
            if not cells:
                continue
            head = _norm(cells[0][1])
            if head in ("派遣先", "派遣元", "派遣条件"):
                section = head
            for ci, raw in cells:
                lab = _norm(raw)
                if lab == "契約番号":
                    f.setdefault("契約番号", _right_value(row, ci))
                elif lab == "事業所名" and section == "派遣先":
                    f.setdefault("派遣先名称", _right_value(row, ci))
                elif lab == "就業場所":
                    v = _right_value(row, ci, skip={"名称"})
                    f.setdefault("就業場所名称", v)
                    nxt = grid[ri + 1] if ri + 1 < len(grid) else []
                    addr = next((x.strip() for x in nxt if str(x).strip().startswith("〒")), "")
                    if not addr:
                        addr = _right_value(nxt, 0, skip={"住所"}) if any(_norm(x) == "住所" for x in nxt) else ""
                    if addr.startswith("〒") or "都" in addr or "県" in addr:
                        f.setdefault("就業場所住所", addr)
                elif lab == "組織単位":
                    f.setdefault("組織単位", _right_value(row, ci))
                elif lab == "派遣業務の内容":
                    f.setdefault("業務内容", _right_value(row, ci))
                elif lab == "責任の程度":
                    f.setdefault("責任の程度", _right_value(row, ci))
                elif lab == "派遣期間":
                    ds = _dates_in_row(row, ci)
                    if len(ds) >= 2:
                        f.setdefault("派遣期間開始", ds[0])
                        f.setdefault("派遣期間終了", ds[1])
                elif lab in ("就業曜日", "休日", "時間外労働", "実労働時間"):
                    key = {"就業曜日": "就業曜日", "休日": "休日", "時間外労働": "時間外労働", "実労働時間": "実労働時間"}[lab]
                    f.setdefault(key, _right_value(row, ci, skip=_SUBLABELS - {lab}))
                elif lab in ("就業時間", "休憩時間"):
                    v = _right_value(row, ci, skip=_SUBLABELS - {lab})
                    if re.search(r"\d", v):
                        f.setdefault(lab, v)
                elif lab == "安全及び衛生":
                    f.setdefault("安全及び衛生", _right_value(row, ci))
                elif lab == "福利厚生施設の利用等":
                    f.setdefault("便宜供与", _right_value(row, ci))
                elif lab == "備考" and ci == 0:
                    v = _right_value(row, ci)
                    if "限定" in v:
                        f.setdefault("限定の別", v.splitlines()[0])
                elif lab in {_norm(b) for b in _BLOCKS}:
                    block = next(b for b in _BLOCKS if _norm(b) == lab)
                    if block == "苦情の申出を受ける者":
                        block = "苦情申出先" if section == "派遣先" else "派遣元苦情"
                    elif block == "派遣元責任者":
                        block = "派遣元責任者"
                    pair1 = dict(_pairs(row, ci))
                    pair2 = dict(_pairs(grid[ri + 1], -1)) if ri + 1 < len(grid) else {}
                    f.setdefault(f"{block}_部署", pair1.get("部門名", ""))
                    f.setdefault(f"{block}_役職", pair1.get("役職", ""))
                    f.setdefault(f"{block}_氏名", pair2.get("氏名", ""))
                    f.setdefault(f"{block}_TEL", pair2.get("TEL", pair2.get("ＴＥＬ", "")))
                elif lab == "派遣労働者氏名" and person_sheet:
                    f.setdefault("氏名カナ", _right_value(row, ci))
                    nxt = grid[ri + 1] if ri + 1 < len(grid) else []
                    f.setdefault("氏名", _right_value(nxt, -1))
                    for rr in (row, nxt):
                        for cj in range(len(rr) - 1):
                            if str(rr[cj]).strip() == "■" and _norm(rr[cj + 1]) in ("男性", "女性"):
                                f.setdefault("性別", _norm(rr[cj + 1])[0])
                elif lab in ("社保･雇保の適用状況", "社保・雇保の適用状況") and person_sheet:
                    ins_values = []
                    for rr in grid[ri:ri + 3]:
                        for cj in range(len(rr) - 1):
                            if _norm(rr[cj]) in ("健康保険", "厚生年金", "雇用保険") or (
                                    ins_values and _norm(rr[cj]) == "健康保険"):
                                v = _right_value(rr, cj)
                                if v in ("有", "無"):
                                    ins_values.append(v)
                                break
                    for kind, v in zip(("健康保険", "厚生年金", "雇用保険"), ins_values):
                        f.setdefault(kind, v)
                elif lab == "雇用形態" and person_sheet:
                    for cj in range(ci + 1, len(row)):
                        t = _norm(row[cj])
                        if t.startswith("■") and ("無期" in t or "有期" in t):
                            f.setdefault("雇用形態", "無期雇用" if "無期" in t else "有期雇用")

    for sn, grid in grids.items():
        if "個別契約書" in sn:
            scan(grid, person_sheet=False)
    for sn, grid in grids.items():
        if "通知書" in sn or "台帳" in sn:
            scan(grid, person_sheet=True)
    f["様式"] = "自社様式"
    return f


def _pairs(row: list[str], start: int) -> list[tuple[str, str]]:
    """行の中の『ラベル, 値』並びを拾う（部門名/役職/氏名/TEL 用）。"""
    out = []
    cells = [(ci, str(v).strip()) for ci, v in enumerate(row) if str(v).strip()]
    for i, (ci, v) in enumerate(cells):
        if _norm(v) in ("部門名", "役職", "氏名", "TEL", "ＴＥＬ") and i + 1 < len(cells):
            out.append((_norm(v).replace("ＴＥＬ", "TEL"), cells[i + 1][1]))
    return out


# ---------------------------------------------------------------- ペンギン様式(xlsx)
def parse_penguin(grids: dict[str, list[list[str]]]) -> dict:
    """ペンギンソリューションズの個別契約書（1シート・ラベルは0列目）。"""
    grid = next(g for sn, g in grids.items() if "個別契約書" in sn)
    f: dict[str, str] = {"様式": "ペンギン様式"}
    sub_labels = {"部署名", "氏名", "役職名", "TEL", "ＴＥＬ", "事業所名", "所在地", "住所"}

    def pairs(row: list[str]) -> dict[str, str]:
        out = {}
        cells = [(i, str(v).strip()) for i, v in enumerate(row) if str(v).strip()]
        for i, (_, v) in enumerate(cells):
            key = _norm(v).rstrip("：:")
            if key in sub_labels and i + 1 < len(cells):
                nxt = cells[i + 1][1]
                if not _norm(nxt).endswith("：") and not _norm(nxt).endswith(":"):
                    out[key.replace("ＴＥＬ", "TEL")] = nxt
        return out

    def rv(row: list[str]) -> str:
        for c in range(1, len(row)):
            v = str(row[c]).strip()
            if v and not _norm(v).rstrip("：:") in sub_labels and v not in ("（派遣先）", "（派遣元）"):
                return v
        return ""

    def nums(row: list[str], width: int) -> list[str]:
        return [str(v).strip() for v in row[1:] if re.fullmatch(rf"\d{{1,{width}}}", str(v).strip())]

    for ri, row in enumerate(grid):
        head = _norm(row[0]) if row else ""
        nxt = grid[ri + 1] if ri + 1 < len(grid) else []
        if head == "派遣先名称":
            f["派遣先名称"] = rv(row)
        elif head == "就業場所":
            p = pairs(row)
            f["就業場所名称"] = p.get("事業所名", rv(row))
            for rr in grid[ri + 1:ri + 4]:
                pp = pairs(rr)
                if "部署名" in pp:
                    f["就業場所名称"] = f.get("就業場所名称", "") + "　" + pp["部署名"]
                if "所在地" in pp:
                    f["就業場所住所"] = pp["所在地"]
        elif head == "組織単位":
            f["組織単位"] = rv(row)
        elif head in ("指揮命令者", "派遣先責任者", "派遣元責任者"):
            p1, p2 = pairs(row), pairs(nxt)
            f[f"{head}_部署"] = p1.get("部署名", "")
            f[f"{head}_役職"] = p1.get("役職名", "")
            f[f"{head}_氏名"] = p2.get("氏名", "")
            f[f"{head}_TEL"] = p2.get("TEL", "")
        elif head.startswith("苦情処理"):
            side = ""
            for rr in grid[ri:ri + 4]:
                joined = "".join(str(v) for v in rr)
                if "（派遣先）" in joined:
                    side = "苦情申出先"
                elif "（派遣元）" in joined:
                    side = "派遣元苦情"
                pp = pairs(rr)
                if side and "部署名" in pp:
                    f[f"{side}_部署"] = pp["部署名"]
                    f[f"{side}_役職"] = pp.get("役職名", "")
                if side and "氏名" in pp:
                    f[f"{side}_氏名"] = pp["氏名"]
                    f[f"{side}_TEL"] = pp.get("TEL", "")
        elif head == "業務内容":
            f["業務内容"] = rv(row)
        elif "責任の程度" in head:
            m = re.search(r"■\s*([^□■]+)", rv(row))
            f["責任の程度"] = m.group(1).strip().rstrip("：:") if m else rv(row)
        elif head == "派遣期間":
            n = nums(row, 4)
            if len(n) >= 6:
                f["派遣期間開始"] = f"{n[0]}/{int(n[1]):02d}/{int(n[2]):02d}"
                f["派遣期間終了"] = f"{n[3]}/{int(n[4]):02d}/{int(n[5]):02d}"
        elif head == "勤務日":
            f["就業曜日"] = " ".join(re.findall(r"■\s*([月火水木金土日祝])", rv(row)))
        elif head in ("就業時間", "休憩"):
            n = nums(row, 2)
            if len(n) >= 4:
                key = "就業時間" if head == "就業時間" else "休憩時間"
                f[key] = f"{int(n[0])}:{int(n[1]):02d}～{int(n[2])}:{int(n[3]):02d}"
        elif head == "休日":
            f["休日"] = rv(row)
        elif head.startswith("時間外労働"):
            f["時間外労働"] = rv(row)
        elif head == "安全及び衛生":
            f["安全及び衛生"] = rv(row)
        elif head == "便宜供与":
            f["便宜供与"] = rv(row)
    return f


# ---------------------------------------------------------------- アイ・ティー・ワン(PDF)
def parse_itone(path: Path) -> dict:
    import pymupdf
    doc = pymupdf.open(str(path))
    text = "\n".join(pg.get_text() for pg in doc)
    t = re.sub(r"[ \t]+", " ", text)
    flat = re.sub(r"\s+", " ", text)

    def find(pattern, src=None, group=1):
        m = re.search(pattern, src or flat)
        return m.group(group).strip() if m else ""

    f: dict[str, str] = {"様式": "ITOne-PDF"}
    f["契約番号"] = find(r"発注番号\s*(\S+)")
    f["派遣先名称"] = "株式会社アイ・ティー・ワン"
    f["就業場所住所"] = find(r"就業場所\s*所在地\s*(.+?)(?:\s名称|\s部署|\s組織)")
    f["就業場所名称"] = find(r"所在地\s.+?名称\s*(株式会社[^\s]+|[^\s]+株式会社)")
    f["組織単位"] = find(r"指揮命令者\s*部署名\s*(.+?)\s*氏名")
    f["業務内容"] = find(r"業務内容及び\s*付随的な業務\s*(.+?)\s*指揮命令者")
    f["責任の程度"] = find(r"(メンバー（役職なし）|リーダー（[^）]*）|[^\s]*役職[^\s]*)\s*責任の程度")
    m = re.search(r"派遣期間.*?(\d{4}[/年]\d{1,2}[/月]\d{1,2}日?)\s*[～〜]?\s*(\d{4}[/年]\d{1,2}[/月]\d{1,2}日?)", flat)
    if not m:
        m = re.search(r"(\d{4}/\d{2}/\d{2})\s*[～〜]\s*(\d{4}/\d{2}/\d{2})", flat)
    if m:
        d1, d2 = parse_date(re.sub(r"[年月]", "/", m.group(1)).rstrip("日")), parse_date(re.sub(r"[年月]", "/", m.group(2)).rstrip("日"))
        if d1 and d2:
            f["派遣期間開始"], f["派遣期間終了"] = d1.strftime("%Y/%m/%d"), d2.strftime("%Y/%m/%d")
    f["就業曜日"] = find(r"就業日\s*([月火水木金土日祝・\s]+?)\s*休日")
    f["休日"] = find(r"休日\s*([^\n]+?)\s*就業時間")
    f["就業時間"] = find(r"就業時間\s*(\d{1,2}:\d{2}\s*[～〜]\s*\d{1,2}:\d{2})").replace(" ", "")
    f["休憩時間"] = find(r"休憩時間\s*(\d{1,2}:\d{2}\s*[～〜]\s*\d{1,2}:\d{2})").replace(" ", "")
    f["時間外労働"] = find(r"時間外・休日労働\s*[（(]36協定[）)]\s*(.+?)(?:特別条項\s*[:：][^。]*。?|派遣料金|明細)")
    f["指揮命令者_部署"] = find(r"指揮命令者\s*部署名\s*(.+?)\s*氏名")
    f["指揮命令者_氏名"] = find(r"指揮命令者\s*部署名.+?氏名\s*(.+?)\s*(?:派遣元|電話)")
    f["派遣元責任者_部署"] = find(r"派遣元\s*責任者\s*部署名\s*(.+?)\s*氏名")
    f["派遣元責任者_氏名"] = find(r"派遣元\s*責任者\s*部署名.+?氏名\s*(.+?)\s*電話番号")
    f["派遣元責任者_TEL"] = find(r"派遣元\s*責任者.+?電話番号\s*([\d\-]+)")
    f["派遣先責任者_部署"] = find(r"派遣先\s*責任者\s*部署名\s*(.+?)\s*(?:役職|氏名)")
    f["派遣先責任者_氏名"] = find(r"派遣先\s*責任者.+?氏名\s*(.+?)\s*(?:電話|派遣先\s*苦情)")
    f["派遣先責任者_TEL"] = find(r"派遣先\s*責任者.+?電話番号\s*([\d\-]+)")
    f["苦情申出先_部署"] = find(r"苦情申し?出先\s*部署名\s*氏名\s*電話番号\s*(.+?)\s+[^\s]+\s+[\d\-]+") or find(r"苦情申し?出先\s*部署名\s*(.+?)\s*氏名")
    m = re.search(r"苦情申し?出先\s*部署名\s*氏名\s*電話番号\s*(.+?)\s+([^\s]+\s?[^\s]*)\s+([\d\-]{10,})", flat)
    if m:
        f["苦情申出先_部署"], f["苦情申出先_氏名"], f["苦情申出先_TEL"] = m.group(1), m.group(2), m.group(3)
    f["事業所抵触日"] = find(r"事業所抵触日\s*(?:事業所名\s*)?(\d{4}年\d{1,2}月\d{1,2}日|\d{4}/\d{1,2}/\d{1,2})")
    d = parse_date(re.sub(r"[年月]", "/", f["事業所抵触日"]).rstrip("日")) if f["事業所抵触日"] else None
    if d:
        f["事業所抵触日"] = d.strftime("%Y/%m/%d")
    f["限定の別"] = find(r"(労使協定方式限定[^\s]*|無期[^\s]*に?限定[^\s]*)")
    return f


# ---------------------------------------------------------------- 実行
def extract_all(base: Path | str = CONTRACT_BASE, out_csv: Path | str = AUTO_CSV,
                debug: bool = False) -> tuple[int, list[str]]:
    base = Path(base)
    warnings: list[str] = []
    rows: dict[tuple, tuple[float, dict]] = {}
    for fam in FAMILIES:
        files = sorted(base.glob(fam["glob"]))
        n = 0
        for p in files:
            if any(k in p.name for k in ("コピー", "org")):
                continue
            try:
                if fam["format"] == "nmht":
                    f = parse_nmht(load_grids(p))
                elif fam["format"] == "penguin":
                    f = parse_penguin(load_grids(p))
                else:
                    f = parse_itone(p)
            except Exception as e:
                warnings.append(f"{fam['name']}: {p.name} 解析失敗 {e}")
                continue
            f["出所ファイル"] = str(p)
            if fam.get("worker"):
                f.setdefault("氏名", fam["worker"])   # 契約書に氏名が無い様式（人数のみ）向けの固定値
            if fam["format"] == "itone":
                f.setdefault("氏名", "中澤 寿代")   # ITOneの個別契約書は氏名非記載（人数のみ）。対象者は1名で確定
            end = parse_date(f.get("派遣期間終了", ""))
            start = parse_date(f.get("派遣期間開始", ""))
            if not (start and end):
                warnings.append(f"{fam['name']}: {p.name} 期間が読めない → スキップ")
                continue
            if end < KEEP_FROM:
                continue
            key = (_norm(f.get("氏名", "")), f["派遣期間開始"], f["派遣期間終了"], f.get("派遣先名称", ""))
            mtime = p.stat().st_mtime
            if key not in rows or mtime >= rows[key][0]:   # 修正版・_改 は後勝ち
                rows[key] = (mtime, f)
            n += 1
            if debug:
                print(f"--- {p.name}")
                for k in COLUMNS:
                    if f.get(k):
                        print(f"    {k}: {str(f[k])[:80]}")
        if n == 0:
            warnings.append(f"{fam['name']}: 対象ファイルなし（glob={fam['glob']}）")
    out = Path(out_csv)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8-sig", newline="") as fp:
        w = csv.DictWriter(fp, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        for _, f in sorted(rows.values(), key=lambda t: (t[1].get("派遣先名称", ""), t[1].get("氏名", ""), t[1].get("派遣期間開始", ""))):
            w.writerow({k: f.get(k, "") for k in COLUMNS})
    return len(rows), warnings
