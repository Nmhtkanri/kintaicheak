"""BBS（ブロードバンドセキュリティ）勤務表 xlsx の構造化パーサ

`Z:\\jinjer移行\\カレンダー\\BBSカレンダー\\{月}\\*NMshift.xlsx` を Claude を経由せず
確定的に解析する。

なぜ構造化パースが要るか:
    この表は「1人＝3行（計画／リーダー／時間）」で、何日のセルかは**列位置だけ**が
    手がかりになる。AI読み取り（_excel_to_text → Claude）に回すと CSV 風テキストの
    連続カンマを数え違えて、**行頭が空欄の人だけ1日ズレる**。
    2026-09 の加藤 英人さん（9/1が空欄で、他の4名は9/1に記号あり）で3回中2回ズレた
    （2026-08-31 実測）。列位置で読めば空欄もそのまま「その日は休み」として扱える。

シートの構造:
    1行目          「令和8年(2026)年09月勤務表」          ← 対象年月
    日付ヘッダー行  A列「枠」 / D列以降に 1,2,3… の日番号
    その次の行      曜日（火,水,木…）                     ← 年月の検算に使う
    その次の行      祝日マーカー（「祝」）
    以降 3行で1人   C列が 計画 / リーダー / 時間
                    B列＝氏名（当社社員は「(N)」始まり）
                    計画     … その日のシフト記号。**空欄＝休み**
                    リーダー … リーダー当番（BL/AL）。まれに「有給休暇」等のメモも入る
                    時間     … その日の実働時間（10.75 / 8 など。取り込まない）
    最下段          凡例「B ＝20:00～32:15」「A ＝8:00～20:15」…（無い月もある）

記号（凡例がシートに書かれていない月は BBS_DEFAULT_TIMES で補う）:
    A … 8:00～20:15    B … 20:00～32:15（24時超表記＝翌8:15）
    E … 9:00～18:00    L … 11:00～20:00    調 … 10:00～17:00
    有 … 有給休暇（全日 → exporter 側で「一般」雛形になる）

休憩は凡例欄の「24シフト休憩時間＝1.5時間」「通常日勤休憩時間＝1時間」を読み、
拘束12時間以上を長い方に割り当てる（A/B=1.5h、E/L/調=1h）。

⚠️ この表には BBS 側の要員も載りうる（「枠」番号が歯抜けなのはその名残）。
**当社社員は氏名が「(N)」で始まる**ため、(N) 付きの行だけを取り込む。
1人も見つからない場合は解析失敗にして AI 読み取りへ回す
（他社の方のスケジュールを jinjer へ投入しないため）。
"""

from __future__ import annotations

import calendar
import logging
import os
import re
from datetime import date

logger = logging.getLogger(__name__)

# シフト表の系統識別子（氏名エイリアス表・対象者リストの適用範囲）
BBS_SOURCE = "bbs"

# レイアウト（1始まりの列番号）
_COL_NAME = 2        # B列＝氏名
_COL_ROW_LABEL = 3   # C列＝計画 / リーダー / 時間

_LABEL_PLAN = "計画"
_LABEL_LEADER = "リーダー"

# 日付ヘッダーを探す範囲と、必要な日数（月末が28日の2月でも通るように28）
_HEADER_SEARCH_ROWS = 12
_MIN_DAY_COLUMNS = 28

# 曜日行が対象年月と一致しているとみなす下限
_WEEKDAY_MATCH_MIN_RATIO = 0.9

_WEEKDAY_KANJI = ["月", "火", "水", "木", "金", "土", "日"]

# 「令和8年(2026)年09月勤務表」の (2026)年09月 を拾う
_TITLE_RE = re.compile(r"[(（](\d{4})[)）]\s*年\s*(\d{1,2})\s*月")
# シート名「2026-09」
_SHEET_NAME_RE = re.compile(r"^(20\d{2})[-_/](0[1-9]|1[0-2])$")
# 当社社員の目印「(N)」
_OUR_STAFF_RE = re.compile(r"^[(（]\s*[NnＮｎ]\s*[)）]\s*")

# 凡例セル「＝20:00～32:15」／「＝有給休暇」
_LEGEND_TIME_RE = re.compile(
    r"^[＝=]\s*(\d{1,2})\s*[:：]\s*(\d{2})\s*[～~〜]\s*(\d{1,2})\s*[:：]\s*(\d{2})\s*$"
)
_LEGEND_DESC_RE = re.compile(r"^[＝=]\s*(\S.*)$")
# 「24シフト休憩時間＝1.5時間」「通常日勤休憩時間＝1時間」
_BREAK_RE = re.compile(r"(24シフト|通常日勤)休憩時間\s*[＝=]\s*(\d+(?:\.\d+)?)\s*時間")

# 凡例が書かれていない月に補う既定値（2026-04〜2026-08 の同ブックの凡例より）
BBS_DEFAULT_TIMES = {
    "A": ("8:00", "20:15"),
    "B": ("20:00", "32:15"),   # 24時超表記＝翌8:15
    "E": ("9:00", "18:00"),
    "L": ("11:00", "20:00"),
    "調": ("10:00", "17:00"),
}

# 全日有休。exporter 側は code / label の完全一致で「一般」雛形へ振る
PAID_LEAVE_CODES = {"有", "有休", "有給", "有給休暇"}
_PAID_LEAVE_LABEL = "有給休暇"

# 休憩の既定値（凡例欄が読めない月用）と、長い休憩を当てる拘束時間のしきい値
LONG_SHIFT_MINUTES = 12 * 60
DEFAULT_LONG_BREAK_MIN = 90
DEFAULT_NORMAL_BREAK_MIN = 60

# リーダー行のメモを「その日の記号」として拾う語（有給休暇など）。
# 「新宿」のような勤務地メモは拾わず、警告に回す。
_LEAVE_KEYWORDS = ("休", "有給", "有休")


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _cell(row, col_1based: int) -> str:
    """行タプルから1始まりの列を取り出す（範囲外は空文字）"""
    if not row or col_1based is None:
        return ""
    idx = col_1based - 1
    if idx < 0 or idx >= len(row):
        return ""
    return _clean(row[idx])


def is_our_staff(raw_name) -> bool:
    """氏名が当社社員の目印「(N)」で始まるか"""
    return bool(_OUR_STAFF_RE.match(_clean(raw_name)))


def normalize_name(raw_name) -> str:
    """「(N)瀧澤　智也」→「瀧澤 智也」（(N) を外し空白を半角1つに揃える）"""
    name = _OUR_STAFF_RE.sub("", _clean(raw_name))
    return re.sub(r"[\s\u3000]+", " ", name).strip()


def _span_minutes(start: str, end: str) -> int:
    """拘束時間（分）。終了が開始以下なら翌日扱い（24時超表記もそのまま扱える）"""
    def to_minutes(text: str) -> int:
        hour, minute = str(text).split(":")
        return int(hour) * 60 + int(minute)

    start_min, end_min = to_minutes(start), to_minutes(end)
    if end_min <= start_min:
        end_min += 24 * 60
    return end_min - start_min


def _break_minutes_for(start: str, end: str, long_break: int, normal_break: int) -> int:
    return long_break if _span_minutes(start, end) >= LONG_SHIFT_MINUTES else normal_break


def find_year_month(sheet_title, rows) -> "tuple[int, int] | None":
    """タイトル行 →（無ければ）シート名 の順で対象年月を決める"""
    for row in rows[:4]:
        for value in row or ():
            match = _TITLE_RE.search(_clean(value))
            if match:
                year, month = int(match.group(1)), int(match.group(2))
                if 1 <= month <= 12:
                    return year, month
    match = _SHEET_NAME_RE.match(_clean(sheet_title))
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def find_day_header(rows) -> "tuple[int, dict[int, int]]":
    """日付ヘッダー行を探して (行index, {日: 列番号}) を返す

    表の下にも同じ日番号の行が残っている月があるため、**上から最初に見つかった行**
    を採用する。日番号は 1 から連番で並んでいることを条件にして、
    「労働日数」「勤務年数」等の数値セルを拾わないようにする。

    Raises:
        ValueError: 日付ヘッダー行が無い
    """
    for idx, row in enumerate(rows[:_HEADER_SEARCH_ROWS]):
        day_col: "dict[int, int]" = {}
        for col_0based, value in enumerate(row or ()):
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                continue
            day = int(value)
            if day != len(day_col) + 1 or not 1 <= day <= 31:
                continue
            day_col[day] = col_0based + 1
        if len(day_col) >= _MIN_DAY_COLUMNS:
            return idx, day_col
    raise ValueError("日付ヘッダー行（1〜31の連番）が見つかりません")


def weekday_match(rows, day_row_idx: int, day_col: "dict[int, int]",
                  year: int, month: int) -> "tuple[int, int]":
    """日付ヘッダーの次行（曜日）が対象年月と合う日数を数える → (一致, 判定対象)"""
    weekday_row = rows[day_row_idx + 1] if day_row_idx + 1 < len(rows) else ()
    days_in_month = calendar.monthrange(year, month)[1]
    matched = total = 0
    for day in range(1, days_in_month + 1):
        text = _cell(weekday_row, day_col.get(day))
        if not text:
            continue
        total += 1
        if text[0] == _WEEKDAY_KANJI[date(year, month, day).weekday()]:
            matched += 1
    return matched, total


def parse_legend_block(rows) -> "tuple[dict[str, tuple[str, str]], dict[str, str]]":
    """最下段の凡例（「A」「＝8:00～20:15」の2セル並び）を読む

    Returns:
        (times, descriptions)
        times        … {記号: (出勤, 退勤)}
        descriptions … {記号: 説明}（「有」→「有給休暇」など時刻を持たないもの）
    """
    times: "dict[str, tuple[str, str]]" = {}
    descriptions: "dict[str, str]" = {}
    for row in rows:
        cells = [_clean(v) for v in (row or ()) if _clean(v)]
        for code, desc in zip(cells, cells[1:]):
            if len(code) > 3 or "＝" in code or "=" in code:
                continue
            time_match = _LEGEND_TIME_RE.match(desc)
            if time_match:
                start = f"{int(time_match.group(1))}:{time_match.group(2)}"
                end = f"{int(time_match.group(3))}:{time_match.group(4)}"
                times.setdefault(code, (start, end))
                continue
            desc_match = _LEGEND_DESC_RE.match(desc)
            if desc_match:
                descriptions.setdefault(code, desc_match.group(1).strip())
    return times, descriptions


def parse_break_minutes(rows) -> "tuple[int, int]":
    """凡例欄の休憩時間表記 → (長い勤務の休憩分, 通常日勤の休憩分)"""
    long_break, normal_break = DEFAULT_LONG_BREAK_MIN, DEFAULT_NORMAL_BREAK_MIN
    for row in rows:
        for value in row or ():
            for kind, hours in _BREAK_RE.findall(_clean(value)):
                minutes = int(round(float(hours) * 60))
                if kind == "24シフト":
                    long_break = minutes
                else:
                    normal_break = minutes
    return long_break, normal_break


def build_bbs_legend(
    sheet_times: "dict[str, tuple[str, str]]",
    sheet_descriptions: "dict[str, str]",
    seen_codes: "set[str]",
    long_break: int,
    normal_break: int,
) -> "tuple[list[dict], list[str]]":
    """凡例エントリを作る

    シートに書かれている凡例を優先し、表で使われているのに凡例が無い記号だけ
    BBS_DEFAULT_TIMES で補う（凡例欄が省かれている月がある）。

    Returns:
        (legend, filled_from_default) — filled_from_default は既定値で補った記号
    """
    legend: "list[dict]" = []
    added: "set[str]" = set()
    filled: "list[str]" = []

    def add_work(code: str, start: str, end: str, from_default: bool) -> None:
        legend.append({
            "code": code,
            "label": f"BBS{code}勤({start}～{end})",
            "start_time": start,
            "end_time": end,
            "break_minutes": _break_minutes_for(start, end, long_break, normal_break),
            "is_off": False,
        })
        added.add(code)
        if from_default:
            filled.append(code)

    for code, (start, end) in sheet_times.items():
        add_work(code, start, end, from_default=False)
    for code in sorted(seen_codes):
        if code in added:
            continue
        default = BBS_DEFAULT_TIMES.get(code)
        if default:
            add_work(code, default[0], default[1], from_default=True)

    # 全日有休。シートの凡例（有＝有給休暇）と、表に直接書かれた「有給休暇」の両方を拾う
    leave_codes = {c for c in seen_codes if c in PAID_LEAVE_CODES}
    leave_codes |= {c for c in sheet_descriptions if c in PAID_LEAVE_CODES}
    for code in sorted(leave_codes):
        if code in added:
            continue
        legend.append({
            "code": code,
            "label": sheet_descriptions.get(code) or _PAID_LEAVE_LABEL,
            "start_time": None,
            "end_time": None,
            "break_minutes": 0,
            "is_off": True,
        })
        added.add(code)

    return legend, filled


def _looks_like_leave(text: str) -> bool:
    return any(keyword in text for keyword in _LEAVE_KEYWORDS)


def parse_bbs_worksheet(
    ws,
    *,
    filename: str,
    target_year: "int | None" = None,
    target_month: "int | None" = None,
) -> dict:
    """1シート → code_sheet 形式へ解析する

    Raises:
        ValueError: 年月が読めない / 対象年月と不一致 / 日付ヘッダーが無い /
                    曜日行が合わない / 当社社員の行が無い
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{filename}: 空のシートです")

    found = find_year_month(ws.title, rows)
    if not found:
        raise ValueError(f"{filename}: 対象年月（タイトルの「(YYYY)年MM月」）が読めません")
    year, month = found
    if (target_year is not None and target_month is not None
            and (year != target_year or month != target_month)):
        raise ValueError(
            f"{filename}: 勤務表の年月 {year}年{month}月 が対象 "
            f"{target_year}年{target_month}月 と一致しません")

    day_row_idx, day_col = find_day_header(rows)
    matched, total = weekday_match(rows, day_row_idx, day_col, year, month)
    if total and matched / total < _WEEKDAY_MATCH_MIN_RATIO:
        raise ValueError(
            f"{filename}: 曜日行が {year}年{month}月 と一致しません（{matched}/{total}日）")

    days_in_month = calendar.monthrange(year, month)[1]
    employees: "list[dict]" = []
    seen_codes: "set[str]" = set()
    skipped_names: "list[str]" = []
    leader_notes: "list[str]" = []

    for idx in range(day_row_idx + 1, len(rows)):
        row = rows[idx]
        if _cell(row, _COL_ROW_LABEL) != _LABEL_PLAN:
            continue
        raw_name = _cell(row, _COL_NAME)
        if not raw_name:
            continue
        if not is_our_staff(raw_name):
            skipped_names.append(raw_name)
            continue

        name = normalize_name(raw_name)
        next_row = rows[idx + 1] if idx + 1 < len(rows) else ()
        leader_row = next_row if _cell(next_row, _COL_ROW_LABEL) == _LABEL_LEADER else ()

        shifts = []
        for day in range(1, days_in_month + 1):
            col = day_col.get(day)
            code = _cell(row, col)
            if not code:
                # 計画が空の日にリーダー行へ書かれるメモ。有給休暇だけ記号として拾い、
                # 「新宿」のような勤務地メモは取り込まず警告に回す。
                memo = _cell(leader_row, col)
                if memo and _looks_like_leave(memo):
                    code = memo
                elif memo:
                    leader_notes.append(f"{name} {month}/{day}「{memo}」")
            if code:
                seen_codes.add(code)
            shifts.append({"date": date(year, month, day).isoformat(), "code": code})
        employees.append({"name": name, "shifts": shifts})

    if not employees:
        raise ValueError(
            f"{filename}: 当社社員の行（氏名が「(N)」で始まる行）が見つかりませんでした")

    sheet_times, sheet_descriptions = parse_legend_block(rows)
    long_break, normal_break = parse_break_minutes(rows)
    legend, filled_from_default = build_bbs_legend(
        sheet_times, sheet_descriptions, seen_codes, long_break, normal_break)
    unknown_codes = sorted(c for c in seen_codes if c not in {e["code"] for e in legend})

    if skipped_names:
        logger.info("%s: 当社社員でない行を除外: %s", filename, " / ".join(skipped_names))
    if filled_from_default:
        logger.info("%s: 凡例がシートに無いため既定値で補完: %s",
                    filename, " / ".join(filled_from_default))
    if unknown_codes:
        logger.warning("%s: 凡例に無い記号: %s", filename, " / ".join(unknown_codes))

    return {
        "filename": filename,
        "year": year,
        "month": month,
        "legend": legend,
        "employees": employees,
        # 空欄＝休み。exporter は空文字を休扱いにするが、AI 経路と同じ形にそろえる
        "off_markers": [""],
        "source": BBS_SOURCE,
        "unknown_codes": unknown_codes,
        "legend_filled_from_default": filled_from_default,
        "skipped_names": skipped_names,
        "leader_notes": leader_notes,
        "section_info": {
            "section_index": None,
            "sheet": ws.title,
            "weekday_matched": matched,
            "weekday_total": total,
        },
    }


def _pick_worksheet(wb, target_year: "int | None", target_month: "int | None"):
    """対象年月のシート名（2026-09）があればそれを、無ければ先頭シートを使う"""
    if target_year and target_month:
        want = f"{target_year:04d}-{target_month:02d}"
        for ws in wb.worksheets:
            if _clean(ws.title).replace("_", "-").replace("/", "-") == want:
                return ws
    return wb.active


def is_bbs_shift_xlsx(filepath: str) -> bool:
    """BBS勤務表 xlsx かを軽量判定する（日付ヘッダー＋計画/リーダー行＋年月）"""
    if not str(filepath).lower().endswith((".xlsx", ".xlsm")):
        return False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            title = ws.title
            rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
        finally:
            wb.close()
    except Exception as e:
        logger.warning("BBS sniff 失敗 %s: %s", filepath, e)
        return False

    if find_year_month(title, rows) is None:
        return False
    try:
        day_row_idx, _ = find_day_header(rows)
    except ValueError:
        return False
    labels = {_cell(row, _COL_ROW_LABEL) for row in rows[day_row_idx + 1:]}
    return {_LABEL_PLAN, _LABEL_LEADER} <= labels


def parse_bbs_shift_xlsx(
    filepath: str,
    target_year: "int | None" = None,
    target_month: "int | None" = None,
) -> dict:
    """BBS勤務表 xlsx → code_sheet 形式（parse_structured_files と同形）

    target_year/month 未指定（None）の場合はタイトルの年月を採用する。
    """
    import openpyxl

    filename = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        result = parse_bbs_worksheet(
            _pick_worksheet(wb, target_year, target_month),
            filename=filename,
            target_year=target_year,
            target_month=target_month,
        )
    finally:
        wb.close()

    logger.info(
        "BBS勤務表を構造化解析: %s → %d年%d月 従業員%d人 凡例%d個（曜日一致 %d/%d）",
        filename, result["year"], result["month"],
        len(result["employees"]), len(result["legend"]),
        result["section_info"]["weekday_matched"], result["section_info"]["weekday_total"],
    )
    return result
